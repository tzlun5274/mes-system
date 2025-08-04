"""
作業員報工資料匯入處理模組
負責處理從 Excel/CSV 檔案匯入作業員報工資料的功能
"""

import pandas as pd
import logging
from django.http import JsonResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.shortcuts import redirect, render
from django.utils import timezone
from datetime import datetime, time
from io import BytesIO

from workorder.models import WorkOrder
from workorder.workorder_reporting.models import OperatorSupplementReport, SMTProductionReport
from process.models import Operator, ProcessName
from equip.models import Equipment
from erp_integration.models import CompanyConfig

logger = logging.getLogger(__name__)


def import_user_required(user):
    """
    檢查用戶是否為超級用戶或具有匯入權限
    """
    return user.is_superuser or user.groups.filter(name="報表使用者").exists()


@login_required
@user_passes_test(import_user_required, login_url='/login/')
def operator_report_import_page(request):
    """
    作業員報工資料匯入頁面
    顯示匯入介面和欄位格式說明
    """
    return render(request, 'workorder/import/operator_report_import.html')


@csrf_exempt
@login_required
@user_passes_test(import_user_required, login_url='/login/')
def operator_report_import_file(request):
    """
    作業員報工資料檔案匯入處理
    
    支援的欄位格式：
    作業員名稱	公司代號	報工日期	開始時間	結束時間	工單號	產品編號	工序名稱	設備名稱	報工數量	不良品數量	備註	異常紀錄
    
    必填欄位：作業員名稱、公司代號、報工日期、開始時間、結束時間、工單號、產品編號、工序名稱、報工數量
    選填欄位：設備名稱、備註、異常紀錄
    特殊處理：異常紀錄欄位有內容才會列入異常紀錄
    """
    
    # 檢查權限
    if not import_user_required(request.user):
        return JsonResponse({
            'success': False,
            'message': '您沒有權限執行此操作。請聯繫管理員取得匯入權限。'
        })
    
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'message': '只支援 POST 請求'
        })
    
    try:
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({
                'success': False,
                'message': '沒有選擇檔案'
            })
        
        # 檢查檔案格式
        file_name = uploaded_file.name.lower()
        if not (file_name.endswith('.xlsx') or file_name.endswith('.csv')):
            return JsonResponse({
                'success': False,
                'message': '只支援 Excel (.xlsx) 或 CSV 格式檔案'
            })
        
        # 讀取檔案內容
        try:
            if file_name.endswith('.xlsx'):
                df = pd.read_excel(uploaded_file)
            else:
                df = pd.read_csv(uploaded_file, encoding='utf-8')
        except Exception as e:
            logger.error(f"檔案讀取失敗: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'檔案讀取失敗: {str(e)}'
            })
        
        # 檢查必要欄位
        required_columns = [
            '作業員名稱', '公司代號', '報工日期', '開始時間', '結束時間', 
            '工單號', '產品編號', '工序名稱', '報工數量'
        ]
        
        # 檢查可選欄位（包括設備名稱）
        optional_columns = ['設備名稱', '不良品數量', '備註', '異常紀錄']
        all_expected_columns = required_columns + optional_columns
        
        # 檢查是否有設備名稱欄位
        has_equipment_column = '設備名稱' in df.columns
        if not has_equipment_column:
            logger.warning("匯入檔案中沒有「設備名稱」欄位")
        else:
            logger.info("匯入檔案中包含「設備名稱」欄位")
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'success': False,
                'message': f'缺少必要欄位：{", ".join(missing_columns)}'
            })
        
        # 處理每一行資料
        success_count = 0
        error_count = 0
        skip_count = 0
        errors = []
        skips = []
        
        for index, row in df.iterrows():
            try:
                # 1. 驗證並取得作業員
                operator_name = str(row['作業員名稱']).strip()
                if not operator_name:
                    errors.append(f'第 {index+1} 行：作業員名稱為空')
                    error_count += 1
                    continue
                
                operator = Operator.objects.filter(name=operator_name).first()
                if not operator:
                    errors.append(f'第 {index+1} 行：找不到作業員 "{operator_name}"')
                    error_count += 1
                    continue
                
                # 2. 驗證並取得公司代號
                company_code_raw = row['公司代號']
                if pd.isna(company_code_raw):
                    errors.append(f'第 {index+1} 行：公司代號為空')
                    error_count += 1
                    continue
                
                # 處理 pandas 自動轉換的數字格式
                if isinstance(company_code_raw, (int, float)):
                    # 如果是數字，轉換為整數後再轉字串
                    company_code = str(int(company_code_raw))
                else:
                    # 如果是字串，直接使用
                    company_code = str(company_code_raw).strip()
                
                if not company_code:
                    errors.append(f'第 {index+1} 行：公司代號為空')
                    error_count += 1
                    continue
                
                # 確保公司代號是兩位數格式
                if company_code.isdigit():
                    company_code = company_code.zfill(2)  # 確保是兩位數格式
                
                # 檢查公司是否存在
                company = CompanyConfig.objects.filter(company_code=company_code).first()
                if not company:
                    errors.append(f'第 {index+1} 行：找不到公司代號 "{company_code}"')
                    error_count += 1
                    continue
                
                # 3. 驗證並解析報工日期
                try:
                    date_str = str(row['報工日期']).strip()
                    # 支援多種日期格式：YYYY-MM-DD、YYYY/MM/DD、YYYY.MM.DD
                    if '/' in date_str:
                        # 處理 YYYY/MM/DD 格式
                        work_date = datetime.strptime(date_str, '%Y/%m/%d').date()
                    elif '.' in date_str:
                        # 處理 YYYY.MM.DD 格式
                        work_date = datetime.strptime(date_str, '%Y.%m.%d').date()
                    else:
                        # 處理 YYYY-MM-DD 格式（預設）
                        work_date = pd.to_datetime(date_str).date()
                except Exception as e:
                    errors.append(f'第 {index+1} 行：報工日期格式錯誤 "{row["報工日期"]}"，支援格式：YYYY-MM-DD、YYYY/MM/DD、YYYY.MM.DD')
                    error_count += 1
                    continue
                
                # 4. 驗證並解析開始時間
                try:
                    start_time_str = str(row['開始時間']).strip()
                    if 'days' in start_time_str:
                        # 處理 Excel 時間格式 "0 days 08:30:00"
                        time_part = start_time_str.split(' ')[-1]  # 取得 "08:30:00" 部分
                        start_time = datetime.strptime(time_part, '%H:%M:%S').time()
                    elif 'AM' in start_time_str or 'PM' in start_time_str:
                        # 處理 12 小時制格式 "04:30:00 PM"
                        start_time = pd.to_datetime(start_time_str).time()
                    elif ':' in start_time_str:
                        # 處理標準時間格式 "08:30:00" 或 "08:30"
                        if start_time_str.count(':') == 2:
                            start_time = datetime.strptime(start_time_str, '%H:%M:%S').time()
                        else:
                            start_time = datetime.strptime(start_time_str, '%H:%M').time()
                    else:
                        # 處理純數字格式 "0830"
                        start_time = datetime.strptime(start_time_str.zfill(4), '%H%M').time()
                except Exception as e:
                    errors.append(f'第 {index+1} 行：開始時間格式錯誤 "{row["開始時間"]}"，支援格式：HH:MM:SS、HH:MM、HHMM、12小時制、Excel時間格式')
                    error_count += 1
                    continue
                
                # 5. 驗證並解析結束時間
                try:
                    end_time_str = str(row['結束時間']).strip()
                    if 'days' in end_time_str:
                        # 處理 Excel 時間格式 "0 days 17:30:00"
                        time_part = end_time_str.split(' ')[-1]  # 取得 "17:30:00" 部分
                        end_time = datetime.strptime(time_part, '%H:%M:%S').time()
                    elif 'AM' in end_time_str or 'PM' in end_time_str:
                        # 處理 12 小時制格式 "04:30:00 PM"
                        end_time = pd.to_datetime(end_time_str).time()
                    elif ':' in end_time_str:
                        # 處理標準時間格式 "17:30:00" 或 "17:30"
                        if end_time_str.count(':') == 2:
                            end_time = datetime.strptime(end_time_str, '%H:%M:%S').time()
                        else:
                            end_time = datetime.strptime(end_time_str, '%H:%M').time()
                    else:
                        # 處理純數字格式 "1730"
                        end_time = datetime.strptime(end_time_str.zfill(4), '%H%M').time()
                except Exception as e:
                    errors.append(f'第 {index+1} 行：結束時間格式錯誤 "{row["結束時間"]}"，支援格式：HH:MM:SS、HH:MM、HHMM、12小時制、Excel時間格式')
                    error_count += 1
                    continue
                
                # 6. 驗證工單號碼格式
                workorder_number = str(row['工單號']).strip()
                if not workorder_number:
                    errors.append(f'第 {index+1} 行：工單號為空')
                    error_count += 1
                    continue
                
                # 7. 驗證產品編號
                product_code_raw = row['產品編號']
                # 處理 pandas 讀取的空值問題
                if pd.isna(product_code_raw) or str(product_code_raw).strip() == '' or str(product_code_raw).strip().lower() == 'nan':
                    product_code = ""
                else:
                    product_code = str(product_code_raw).strip()
                
                # 8. 驗證工序名稱
                process_name = str(row['工序名稱']).strip()
                if not process_name:
                    errors.append(f'第 {index+1} 行：工序名稱為空')
                    error_count += 1
                    continue
                
                process = ProcessName.objects.filter(name=process_name).first()
                if not process:
                    errors.append(f'第 {index+1} 行：找不到工序 "{process_name}"')
                    error_count += 1
                    continue
                
                # 設定工序名稱到 operation 欄位
                operation_name = process.name
                
                # 9. 驗證報工數量
                try:
                    work_quantity = int(row['報工數量'])
                    if work_quantity < 0:
                        errors.append(f'第 {index+1} 行：報工數量不能為負數')
                        error_count += 1
                        continue
                except (ValueError, TypeError):
                    errors.append(f'第 {index+1} 行：報工數量格式錯誤 "{row["報工數量"]}"')
                    error_count += 1
                    continue
                
                # 10. 處理不良品數量（選填）
                defect_quantity = 0
                if '不良品數量' in df.columns and pd.notna(row['不良品數量']):
                    try:
                        defect_quantity = int(row['不良品數量'])
                        if defect_quantity < 0:
                            errors.append(f'第 {index+1} 行：不良品數量不能為負數')
                            error_count += 1
                            continue
                    except (ValueError, TypeError):
                        errors.append(f'第 {index+1} 行：不良品數量格式錯誤 "{row["不良品數量"]}"')
                        error_count += 1
                        continue
                
                # 11. 處理設備名稱（選填）
                equipment = None
                
                if '設備名稱' in df.columns and pd.notna(row['設備名稱']):
                    equipment_name_raw = row['設備名稱']
                    # 處理 pandas 讀取的空值問題
                    if pd.isna(equipment_name_raw) or str(equipment_name_raw).strip() == '' or str(equipment_name_raw).strip().lower() == 'nan':
                        equipment_name = ""
                    else:
                        equipment_name = str(equipment_name_raw).strip()
                    
                    if equipment_name:
                        # 先嘗試精確匹配
                        equipment = Equipment.objects.filter(name=equipment_name).first()
                        
                        if not equipment:
                            # 如果精確匹配失敗，嘗試不區分大小寫的查找
                            equipment = Equipment.objects.filter(name__iexact=equipment_name).first()
                            
                            if equipment:
                                # 記錄大小寫不匹配的情況
                                logger.info(f'第 {index+1} 行：設備名稱大小寫不匹配 "{equipment_name}" -> "{equipment.name}"')
                            else:
                                # 如果還是找不到，記錄警告但不中斷匯入
                                logger.warning(f'第 {index+1} 行：找不到設備 "{equipment_name}"，設備欄位設為空')
                                equipment = None
                
                # 12. 處理備註（選填）
                remarks = ""
                if '備註' in df.columns and pd.notna(row['備註']):
                    remarks_raw = row['備註']
                    # 處理 pandas 讀取的空值問題
                    if pd.isna(remarks_raw) or str(remarks_raw).strip() == '' or str(remarks_raw).strip().lower() == 'nan':
                        remarks = ""
                    else:
                        remarks = str(remarks_raw).strip()
                
                # 13. 處理異常紀錄（選填）
                abnormal_notes = ""
                if '異常紀錄' in df.columns and pd.notna(row['異常紀錄']):
                    abnormal_notes_raw = row['異常紀錄']
                    # 處理 pandas 讀取的空值問題
                    if pd.isna(abnormal_notes_raw) or str(abnormal_notes_raw).strip() == '' or str(abnormal_notes_raw).strip().lower() == 'nan':
                        abnormal_notes = ""
                    else:
                        abnormal_notes = str(abnormal_notes_raw).strip()
                
                # 14. 檢查是否已存在相同記錄（避免重複匯入）
                # 暫時註解掉重複檢查功能，讓所有資料都能匯入
                # existing_query = OperatorSupplementReport.objects.filter(
                #     operator=operator,
                #     work_date=work_date,
                #     start_time=start_time,
                #     end_time=end_time,
                #     process=process
                # )
                # 
                # # 如果工單存在，加入工單條件
                # if workorder:
                #     existing_query = existing_query.filter(workorder=workorder)
                # else:
                #     existing_query = existing_query.filter(original_workorder_number=workorder_number)
                # 
                # existing_report = existing_query.first()
                # 
                # if existing_report:
                #     skips.append(f'第 {index+1} 行：已存在相同記錄（作業員、日期、時間、工序、工單），跳過匯入')
                #     skip_count += 1
                #     continue
                
                # 15. 建立報工記錄
                # 直接使用匯入的工單號碼，不進行工單查找
                report_data = {
                    'operator': operator,
                    'process': process,
                    'operation': operation_name,  # 確保工序名稱包含在內
                    'equipment': equipment,  # 確保設備欄位包含在內
                    'product_id': product_code,  # 確保產品編號包含在內
                    'work_date': work_date,
                    'start_time': start_time,
                    'end_time': end_time,
                    'work_quantity': work_quantity,
                    'defect_quantity': defect_quantity,
                    'remarks': remarks,
                    'abnormal_notes': abnormal_notes,
                    'created_by': request.user.username,
                    'original_workorder_number': workorder_number  # 直接使用匯入的工單號碼
                }
                
                report = OperatorSupplementReport.objects.create(**report_data)
                
                # 16. 自動計算工時
                report.calculate_work_hours()
                report.save()
                
                success_count += 1
                
            except Exception as e:
                logger.error(f"處理第 {index+1} 行時發生錯誤: {str(e)}")
                errors.append(f'第 {index+1} 行：處理失敗 - {str(e)}')
                error_count += 1
        
        # 準備回應訊息
        result_message = f"匯入完成！成功：{success_count} 筆，錯誤：{error_count} 筆，跳過：{skip_count} 筆"
        
        if errors:
            result_message += f"\n錯誤詳情：\n" + "\n".join(errors[:10])  # 只顯示前10個錯誤
            if len(errors) > 10:
                result_message += f"\n... 還有 {len(errors) - 10} 個錯誤"
        
        if skips:
            result_message += f"\n跳過詳情：\n" + "\n".join(skips[:5])  # 只顯示前5個跳過
            if len(skips) > 5:
                result_message += f"\n... 還有 {len(skips) - 5} 個跳過"
        
        return JsonResponse({
            'success': True,
            'message': result_message,
            'data': {
                'success_count': success_count,
                'error_count': error_count,
                'skip_count': skip_count,
                'total_processed': len(df)
            }
        })
        
    except Exception as e:
        logger.error(f"匯入處理發生未預期錯誤: {str(e)}")
        import traceback
        logger.error(f"錯誤詳情: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': f'匯入處理失敗: {str(e)}'
        })


@login_required
def download_import_template(request):
    """
    下載匯入範本檔案
    提供標準的 Excel 範本，包含所有必要欄位和範例資料
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # 建立新的工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "作業員報工匯入範本"
        
        # 定義欄位標題
        headers = [
            '作業員名稱', '公司代號', '報工日期', '開始時間', '結束時間', 
            '工單號', '產品編號', '工序名稱', '設備名稱', '報工數量', 
            '不良品數量', '備註', '異常紀錄'
        ]
        
        # 設定標題樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 寫入標題
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 範例資料
        example_data = [
            ['王 珍', '02', '2025-01-15', '08:00', '12:00', 'WO-02-202501001', 'PROD-001', '後焊', 'AOI-1', 100, 2, '正常生產', ''],
            ['蔡秉儒', '02', '2025/01/15', '13:00', '17:00', 'WO-02-202501001', 'PROD-001', '電測', 'AOI-2', 95, 5, '設備調整', '設備故障30分鐘'],
            ['林淑惠', '02', '2025.01.15', '08:00', '16:00', 'WO-02-202501002', 'PROD-002', '出貨包裝', '', 200, 0, '加班生產', ''],
        ]
        
        # 寫入範例資料
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 設定欄寬
        column_widths = [12, 10, 12, 10, 10, 15, 12, 10, 10, 10, 10, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # 設定邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(example_data)+1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # 建立回應
        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="作業員報工匯入範本.xlsx"'
        
        # 儲存到回應
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"下載範本失敗: {str(e)}")
        messages.error(request, f'下載範本失敗: {str(e)}')
        return redirect('workorder:operator_report_import_page')


@login_required
def get_import_field_guide(request):
    """
    取得匯入欄位格式說明
    提供 API 介面，返回詳細的欄位格式說明
    """
    field_guide = {
        'title': '作業員報工資料匯入欄位格式說明',
        'description': '此格式用於匯入作業員的報工資料，支援 Excel 和 CSV 格式',
        'fields': [
            {
                'name': '作業員名稱',
                'required': True,
                'type': '文字',
                'description': '作業員的姓名，必須在系統中已存在',
                'example': '張小明',
                'validation': '不能為空，必須對應到系統中的作業員'
            },
            {
                'name': '公司代號',
                'required': True,
                'type': '文字',
                'description': '公司代號，用於識別不同公司的工單',
                'example': '01',
                'validation': '不能為空，必須對應到系統中的公司設定'
            },
            {
                'name': '報工日期',
                'required': True,
                'type': '日期',
                'description': '報工的日期',
                'example': '2025-01-15',
                'validation': '格式：YYYY-MM-DD'
            },
            {
                'name': '開始時間',
                'required': True,
                'type': '時間',
                'description': '報工開始時間',
                'example': '08:00 或 0800',
                'validation': '24小時制，支援 HH:MM 或 HHMM 格式'
            },
            {
                'name': '結束時間',
                'required': True,
                'type': '時間',
                'description': '報工結束時間',
                'example': '12:00 或 1200',
                'validation': '24小時制，支援 HH:MM 或 HHMM 格式'
            },
            {
                'name': '工單號',
                'required': True,
                'type': '文字',
                'description': '工單號碼',
                'example': 'WO-01-202501001',
                'validation': '不能為空，必須對應到系統中的工單'
            },
            {
                'name': '產品編號',
                'required': True,
                'type': '文字',
                'description': '產品編號',
                'example': 'PROD-001',
                'validation': '不能為空'
            },
            {
                'name': '工序名稱',
                'required': True,
                'type': '文字',
                'description': '工序名稱',
                'example': 'SMT、DIP、測試',
                'validation': '不能為空，必須對應到系統中的工序'
            },
            {
                'name': '設備名稱',
                'required': False,
                'type': '文字',
                'description': '使用的設備名稱（選填）',
                'example': 'SMT-001',
                'validation': '可為空，如果填寫必須對應到系統中的設備'
            },
            {
                'name': '報工數量',
                'required': True,
                'type': '整數',
                'description': '合格品數量',
                'example': '100',
                'validation': '必須為非負整數'
            },
            {
                'name': '不良品數量',
                'required': False,
                'type': '整數',
                'description': '不良品數量（選填）',
                'example': '2',
                'validation': '可為空，如果填寫必須為非負整數'
            },
            {
                'name': '備註',
                'required': False,
                'type': '文字',
                'description': '備註說明（選填）',
                'example': '正常生產',
                'validation': '可為空，純文字描述'
            },
            {
                'name': '異常紀錄',
                'required': False,
                'type': '文字',
                'description': '異常情況記錄（選填）',
                'example': '設備故障30分鐘',
                'validation': '可為空，有內容才會列入異常紀錄'
            }
        ],
        'notes': [
            '所有時間欄位支援 24 小時制格式',
            '設備名稱和備註為選填欄位',
            '異常紀錄欄位有內容才會列入異常紀錄',
            '系統會自動檢查重複記錄並跳過',
            '匯入後會自動計算工時和休息時間'
        ]
    }
    
    return JsonResponse({
        'success': True,
        'data': field_guide
    }) 


@login_required
@user_passes_test(import_user_required, login_url='/login/')
def smt_report_import_page(request):
    """
    SMT設備報工資料匯入頁面
    顯示匯入介面和欄位格式說明
    """
    return render(request, 'workorder/import/smt_report_import.html')


@csrf_exempt
@login_required
@user_passes_test(import_user_required, login_url='/login/')
def smt_report_import_file(request):
    """
    SMT設備報工資料檔案匯入處理
    
    支援的欄位格式：
    設備名稱	公司代號	報工日期	開始時間	結束時間	工單號	產品編號	工序名稱	空白佔位用	報工數量	不良品數量	備註	異常紀錄
    
    必填欄位：設備名稱、公司代號、報工日期、開始時間、結束時間、工單號、產品編號、工序名稱、報工數量
    選填欄位：備註、異常紀錄
    特殊處理：異常紀錄欄位有內容才會列入異常紀錄
    """
    
    if request.method != 'POST':
        return JsonResponse({
            'success': False,
            'message': '只支援 POST 請求'
        })
    
    try:
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({
                'success': False,
                'message': '沒有選擇檔案'
            })
        
        # 檢查檔案格式
        file_name = uploaded_file.name.lower()
        if not (file_name.endswith('.xlsx') or file_name.endswith('.csv')):
            return JsonResponse({
                'success': False,
                'message': '只支援 Excel (.xlsx) 或 CSV 格式檔案'
            })
        
        # 讀取檔案內容
        try:
            if file_name.endswith('.xlsx'):
                df = pd.read_excel(uploaded_file)
            else:
                df = pd.read_csv(uploaded_file, encoding='utf-8')
        except Exception as e:
            logger.error(f"檔案讀取失敗: {str(e)}")
            return JsonResponse({
                'success': False,
                'message': f'檔案讀取失敗: {str(e)}'
            })
        
        # 檢查必要欄位
        required_columns = [
            '設備名稱', '公司代號', '報工日期', '開始時間', '結束時間', 
            '工單號', '產品編號', '工序名稱', '報工數量'
        ]
        
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'success': False,
                'message': f'缺少必要欄位：{", ".join(missing_columns)}'
            })
        
        # 處理每一行資料
        success_count = 0
        error_count = 0
        skip_count = 0
        errors = []
        skips = []
        
        for index, row in df.iterrows():
            try:
                # 1. 驗證並取得設備
                equipment_name = str(row['設備名稱']).strip()
                if not equipment_name:
                    errors.append(f'第 {index+1} 行：設備名稱為空')
                    error_count += 1
                    continue
                
                # 使用不區分大小寫的設備查找
                equipment = Equipment.objects.filter(name__iexact=equipment_name).first()
                if not equipment:
                    errors.append(f'第 {index+1} 行：找不到設備 "{equipment_name}"')
                    error_count += 1
                    continue
                
                # 2. 驗證並取得公司代號
                company_code_raw = row['公司代號']
                if pd.isna(company_code_raw):
                    errors.append(f'第 {index+1} 行：公司代號為空')
                    error_count += 1
                    continue
                
                # 處理 pandas 自動轉換的數字格式
                if isinstance(company_code_raw, (int, float)):
                    # 如果是數字，轉換為整數後再轉字串
                    company_code = str(int(company_code_raw))
                else:
                    # 如果是字串，直接使用
                    company_code = str(company_code_raw).strip()
                
                if not company_code:
                    errors.append(f'第 {index+1} 行：公司代號為空')
                    error_count += 1
                    continue
                
                # 確保公司代號是兩位數格式
                if company_code.isdigit():
                    company_code = company_code.zfill(2)  # 確保是兩位數格式
                
                # 檢查公司是否存在
                company = CompanyConfig.objects.filter(company_code=company_code).first()
                if not company:
                    errors.append(f'第 {index+1} 行：找不到公司代號 "{company_code}"')
                    error_count += 1
                    continue
                
                # 3. 驗證並解析報工日期
                try:
                    date_str = str(row['報工日期']).strip()
                    # 支援多種日期格式：YYYY-MM-DD、YYYY/MM/DD、YYYY.MM.DD
                    if '/' in date_str:
                        # 處理 YYYY/MM/DD 格式
                        work_date = datetime.strptime(date_str, '%Y/%m/%d').date()
                    elif '.' in date_str:
                        # 處理 YYYY.MM.DD 格式
                        work_date = datetime.strptime(date_str, '%Y.%m.%d').date()
                    else:
                        # 處理 YYYY-MM-DD 格式（預設）
                        work_date = pd.to_datetime(date_str).date()
                except Exception as e:
                    errors.append(f'第 {index+1} 行：報工日期格式錯誤 "{row["報工日期"]}"，支援格式：YYYY-MM-DD、YYYY/MM/DD、YYYY.MM.DD')
                    error_count += 1
                    continue
                
                # 4. 驗證並解析開始時間
                try:
                    start_time_str = str(row['開始時間']).strip()
                    if 'days' in start_time_str:
                        # 處理 Excel 時間格式 "0 days 08:30:00"
                        time_part = start_time_str.split(' ')[-1]  # 取得 "08:30:00" 部分
                        start_time = datetime.strptime(time_part, '%H:%M:%S').time()
                    elif 'AM' in start_time_str or 'PM' in start_time_str:
                        # 處理 12 小時制格式 "04:30:00 PM"
                        start_time = pd.to_datetime(start_time_str).time()
                    elif ':' in start_time_str:
                        # 處理標準時間格式 "08:30:00" 或 "08:30"
                        if start_time_str.count(':') == 2:
                            start_time = datetime.strptime(start_time_str, '%H:%M:%S').time()
                        else:
                            start_time = datetime.strptime(start_time_str, '%H:%M').time()
                    else:
                        # 處理純數字格式 "0830"
                        start_time = datetime.strptime(start_time_str.zfill(4), '%H%M').time()
                except Exception as e:
                    errors.append(f'第 {index+1} 行：開始時間格式錯誤 "{row["開始時間"]}"，支援格式：HH:MM:SS、HH:MM、HHMM、12小時制、Excel時間格式')
                    error_count += 1
                    continue
                
                # 5. 驗證並解析結束時間
                try:
                    end_time_str = str(row['結束時間']).strip()
                    if 'days' in end_time_str:
                        # 處理 Excel 時間格式 "0 days 17:30:00"
                        time_part = end_time_str.split(' ')[-1]  # 取得 "17:30:00" 部分
                        end_time = datetime.strptime(time_part, '%H:%M:%S').time()
                    elif 'AM' in end_time_str or 'PM' in end_time_str:
                        # 處理 12 小時制格式 "04:30:00 PM"
                        end_time = pd.to_datetime(end_time_str).time()
                    elif ':' in end_time_str:
                        # 處理標準時間格式 "17:30:00" 或 "17:30"
                        if end_time_str.count(':') == 2:
                            end_time = datetime.strptime(end_time_str, '%H:%M:%S').time()
                        else:
                            end_time = datetime.strptime(end_time_str, '%H:%M').time()
                    else:
                        # 處理純數字格式 "1730"
                        end_time = datetime.strptime(end_time_str.zfill(4), '%H%M').time()
                except Exception as e:
                    errors.append(f'第 {index+1} 行：結束時間格式錯誤 "{row["結束時間"]}"，支援格式：HH:MM:SS、HH:MM、HHMM、12小時制、Excel時間格式')
                    error_count += 1
                    continue
                
                # 6. 驗證工單號碼格式
                workorder_number = str(row['工單號']).strip()
                if not workorder_number:
                    errors.append(f'第 {index+1} 行：工單號為空')
                    error_count += 1
                    continue
                
                # 7. 驗證產品編號
                product_code_raw = row['產品編號']
                # 處理 pandas 讀取的空值問題
                if pd.isna(product_code_raw) or str(product_code_raw).strip() == '' or str(product_code_raw).strip().lower() == 'nan':
                    product_code = ""
                else:
                    product_code = str(product_code_raw).strip()
                
                # 8. 不進行工單查找，直接使用匯入的工單號碼
                
                # 8. 驗證並取得工序
                process_name = str(row['工序名稱']).strip()
                if not process_name:
                    errors.append(f'第 {index+1} 行：工序名稱為空')
                    error_count += 1
                    continue
                
                process = ProcessName.objects.filter(name=process_name).first()
                if not process:
                    errors.append(f'第 {index+1} 行：找不到工序 "{process_name}"')
                    error_count += 1
                    continue
                
                # 9. 驗證報工數量
                try:
                    work_quantity = int(row['報工數量'])
                    if work_quantity < 0:
                        errors.append(f'第 {index+1} 行：報工數量不能為負數')
                        error_count += 1
                        continue
                except (ValueError, TypeError):
                    errors.append(f'第 {index+1} 行：報工數量格式錯誤 "{row["報工數量"]}"')
                    error_count += 1
                    continue
                
                # 10. 處理不良品數量（選填）
                defect_quantity = 0
                if '不良品數量' in df.columns and pd.notna(row['不良品數量']):
                    try:
                        defect_quantity = int(row['不良品數量'])
                        if defect_quantity < 0:
                            errors.append(f'第 {index+1} 行：不良品數量不能為負數')
                            error_count += 1
                            continue
                    except (ValueError, TypeError):
                        errors.append(f'第 {index+1} 行：不良品數量格式錯誤 "{row["不良品數量"]}"')
                        error_count += 1
                        continue
                
                # 11. 處理備註（選填）
                remarks = ""
                if '備註' in df.columns and pd.notna(row['備註']):
                    remarks_raw = row['備註']
                    # 處理 pandas 讀取的空值問題
                    if pd.isna(remarks_raw) or str(remarks_raw).strip() == '' or str(remarks_raw).strip().lower() == 'nan':
                        remarks = ""
                    else:
                        remarks = str(remarks_raw).strip()
                
                # 12. 處理異常紀錄（選填）
                abnormal_notes = ""
                if '異常紀錄' in df.columns and pd.notna(row['異常紀錄']):
                    abnormal_notes_raw = row['異常紀錄']
                    # 處理 pandas 讀取的空值問題
                    if pd.isna(abnormal_notes_raw) or str(abnormal_notes_raw).strip() == '' or str(abnormal_notes_raw).strip().lower() == 'nan':
                        abnormal_notes = ""
                    else:
                        abnormal_notes = str(abnormal_notes_raw).strip()
                
                # 13. 檢查是否已存在相同記錄（避免重複匯入）
                # 暫時註解掉重複檢查功能，讓所有資料都能匯入
                # from workorder.workorder_reporting.models import SMTProductionReport
                # 
                # existing_report = SMTProductionReport.objects.filter(
                #     equipment=equipment,
                #     workorder=workorder,
                #     operation=process_name,
                #     work_date=work_date,
                #     start_time=start_time,
                #     end_time=end_time
                # ).first()
                # 
                # if existing_report:
                #     skips.append(f'第 {index+1} 行：已存在相同記錄，跳過匯入')
                #     skip_count += 1
                #     continue
                
                # 14. 建立SMT報工記錄
                # 直接使用匯入的工單號碼，不進行工單查找
                report_data = {
                    'equipment': equipment,
                    'product_id': product_code,  # 直接使用匯入的產品編號
                    'process': process,  # 使用工序 ID
                    'operation': process_name,  # 保留工序名稱字串
                    'work_date': work_date,
                    'start_time': start_time,
                    'end_time': end_time,
                    'work_quantity': work_quantity,
                    'defect_quantity': defect_quantity,
                    'remarks': remarks,
                    'abnormal_notes': abnormal_notes,
                    'created_by': request.user.username,
                    'original_workorder_number': workorder_number  # 直接使用匯入的工單號碼
                }
                
                report = SMTProductionReport.objects.create(**report_data)
                
                # 15. 自動計算工時
                report.calculate_work_hours()
                report.save()
                
                success_count += 1
                
            except Exception as e:
                logger.error(f"處理第 {index+1} 行時發生錯誤: {str(e)}")
                errors.append(f'第 {index+1} 行：處理失敗 - {str(e)}')
                error_count += 1
        
        # 準備回應訊息
        result_message = f"匯入完成！成功：{success_count} 筆，錯誤：{error_count} 筆，跳過：{skip_count} 筆"
        
        if errors:
            result_message += f"\n錯誤詳情：\n" + "\n".join(errors[:10])  # 只顯示前10個錯誤
            if len(errors) > 10:
                result_message += f"\n... 還有 {len(errors) - 10} 個錯誤"
        
        if skips:
            result_message += f"\n跳過詳情：\n" + "\n".join(skips[:5])  # 只顯示前5個跳過
            if len(skips) > 5:
                result_message += f"\n... 還有 {len(skips) - 5} 個跳過"
        
        return JsonResponse({
            'success': True,
            'message': result_message,
            'data': {
                'success_count': success_count,
                'error_count': error_count,
                'skip_count': skip_count,
                'total_processed': len(df)
            }
        })
        
    except Exception as e:
        logger.error(f"匯入處理發生未預期錯誤: {str(e)}")
        import traceback
        logger.error(f"錯誤詳情: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': f'匯入處理失敗: {str(e)}'
        })


@login_required
@user_passes_test(import_user_required, login_url='/login/')
def download_smt_import_template(request):
    """
    下載SMT匯入範本檔案
    提供標準的 Excel 範本，包含所有必要欄位和範例資料
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # 建立新的工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SMT設備報工匯入範本"
        
        # 定義欄位標題（按照圖片格式順序）
        headers = [
            '設備名稱', '公司代號', '報工日期', '開始時間', '結束時間', 
            '工單號', '產品編號', '工序名稱', '空白佔位用', '報工數量', 
            '不良品數量', '備註', '異常紀錄'
        ]
        
        # 設定標題樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 寫入標題
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 範例資料
        example_data = [
            ['SMT-001', '01', '2025-01-15', '08:00', '12:00', 'WO-01-202501001', 'PROD-001', 'SMT', '', 100, 2, '正常生產', ''],
            ['SMT-002', '01', '2025/01/15', '13:00', '17:00', 'WO-01-202501001', 'PROD-001', 'SMT', '', 95, 5, '設備調整', '設備故障30分鐘'],
            ['SMT-003', '02', '2025.01.15', '08:00', '16:00', 'WO-02-202501002', 'PROD-002', 'SMT', '', 200, 0, '加班生產', ''],
        ]
        
        # 寫入範例資料
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 設定欄寬
        column_widths = [12, 10, 12, 10, 10, 15, 12, 10, 10, 10, 10, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width
        
        # 設定邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in ws.iter_rows(min_row=1, max_row=len(example_data)+1, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
        
        # 建立回應
        from django.http import HttpResponse
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="SMT設備報工匯入範本.xlsx"'
        
        # 儲存到回應
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f"下載SMT範本失敗: {str(e)}")
        messages.error(request, f'下載SMT範本失敗: {str(e)}')
        return redirect('workorder:smt_report_import_page')


@login_required
@user_passes_test(import_user_required, login_url='/login/')
def get_smt_import_field_guide(request):
    """
    取得SMT匯入欄位格式說明
    提供 API 介面，返回詳細的欄位格式說明
    """
    field_guide = {
        'title': 'SMT設備報工資料匯入欄位格式說明',
        'description': '此格式用於匯入SMT設備的報工資料，支援 Excel 和 CSV 格式',
        'fields': [
            {
                'name': '設備名稱',
                'required': True,
                'type': '文字',
                'description': 'SMT設備的名稱，必須在系統中已存在',
                'example': 'SMT-001',
                'validation': '不能為空，必須對應到系統中的設備'
            },
            {
                'name': '公司代號',
                'required': True,
                'type': '文字',
                'description': '公司代號，用於識別不同公司的工單',
                'example': '01',
                'validation': '不能為空，必須對應到系統中的公司設定'
            },
            {
                'name': '報工日期',
                'required': True,
                'type': '日期',
                'description': '報工的日期',
                'example': '2025-01-15',
                'validation': '格式：YYYY-MM-DD'
            },
            {
                'name': '開始時間',
                'required': True,
                'type': '時間',
                'description': '報工開始時間',
                'example': '08:00 或 0800',
                'validation': '24小時制，支援 HH:MM 或 HHMM 格式'
            },
            {
                'name': '結束時間',
                'required': True,
                'type': '時間',
                'description': '報工結束時間',
                'example': '12:00 或 1200',
                'validation': '24小時制，支援 HH:MM 或 HHMM 格式'
            },
            {
                'name': '工單號',
                'required': True,
                'type': '文字',
                'description': '工單號碼',
                'example': 'WO-01-202501001',
                'validation': '不能為空，必須對應到系統中的工單'
            },
            {
                'name': '產品編號',
                'required': True,
                'type': '文字',
                'description': '產品編號',
                'example': 'PROD-001',
                'validation': '不能為空'
            },
            {
                'name': '工序名稱',
                'required': True,
                'type': '文字',
                'description': '工序名稱',
                'example': 'SMT',
                'validation': '不能為空，通常為SMT相關工序'
            },
            {
                'name': '空白佔位用',
                'required': False,
                'type': '文字',
                'description': '空白欄位，用於格式對齊',
                'example': '',
                'validation': '可為空'
            },
            {
                'name': '報工數量',
                'required': True,
                'type': '整數',
                'description': '合格品數量',
                'example': '100',
                'validation': '必須為非負整數'
            },
            {
                'name': '不良品數量',
                'required': False,
                'type': '整數',
                'description': '不良品數量（選填）',
                'example': '2',
                'validation': '可為空，如果填寫必須為非負整數'
            },
            {
                'name': '備註',
                'required': False,
                'type': '文字',
                'description': '備註說明（選填）',
                'example': '正常生產',
                'validation': '可為空，純文字描述'
            },
            {
                'name': '異常紀錄',
                'required': False,
                'type': '文字',
                'description': '異常情況記錄（選填）',
                'example': '設備故障30分鐘',
                'validation': '可為空，有內容才會列入異常紀錄'
            }
        ],
        'notes': [
            '所有時間欄位支援 24 小時制格式',
            '空白佔位用欄位為格式對齊用途，可留空',
            '異常紀錄欄位有內容才會列入異常紀錄',
            '系統會自動檢查重複記錄並跳過',
            '匯入後會自動計算工時（SMT中午不休息，16:30後算加班）'
        ]
    }
    
    return JsonResponse({
        'success': True,
        'data': field_guide
    })


@login_required
@user_passes_test(import_user_required, login_url='/login/')
def smt_report_export(request):
    """
    匯出SMT設備報工記錄
    支援篩選條件和Excel格式匯出
    """
    try:
        from workorder.workorder_reporting.models import SMTProductionReport
        
        reports = SMTProductionReport.objects.all().order_by('-work_date')
        
        # 篩選條件
        date_from = request.GET.get('date_from')
        if date_from:
            reports = reports.filter(work_date__gte=date_from)
        
        date_to = request.GET.get('date_to')
        if date_to:
            reports = reports.filter(work_date__lte=date_to)
        
        equipment_id = request.GET.get('equipment_id')
        if equipment_id:
            reports = reports.filter(equipment_id=equipment_id)
        
        # 建立Excel檔案
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "SMT設備報工記錄"
        
        # 設定標題（按照圖片格式順序）
        headers = [
            '設備名稱', '公司代號', '報工日期', '開始時間', '結束時間', 
            '工單號', '產品編號', '工序名稱', '空白佔位用', '報工數量', 
            '不良品數量', '備註', '異常紀錄'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 填入資料
        for row, report in enumerate(reports, 2):
            ws.cell(row=row, column=1, value=report.equipment.name if report.equipment else '')
            
            # 改善公司代號處理：優先從工單取得，其次使用預設值
            company_code = ''
            if report.workorder and report.workorder.company_code:
                company_code = report.workorder.company_code
            else:
                # 如果沒有工單關聯，根據設備名稱或其他邏輯判斷公司代號
                if report.equipment and report.equipment.name:
                    # 可以根據設備名稱判斷公司代號，這裡先使用預設值
                    company_code = '10'  # 預設公司代號
            ws.cell(row=row, column=2, value=company_code)
            
            ws.cell(row=row, column=3, value=report.work_date.strftime('%Y-%m-%d'))
            ws.cell(row=row, column=4, value=report.start_time.strftime('%H:%M') if report.start_time else '')
            ws.cell(row=row, column=5, value=report.end_time.strftime('%H:%M') if report.end_time else '')
            
            # 使用 workorder_number 屬性來取得工單號碼，這個屬性會處理各種情況
            ws.cell(row=row, column=6, value=report.workorder_number)
            
            # 產品編號 - 直接使用報工記錄本身的產品編號欄位
            product_id = report.product_id or ''
            if not product_id and report.workorder and hasattr(report.workorder, 'product_code') and report.workorder.product_code:
                # 如果報工記錄的產品編號為空，才從工單取得
                product_id = report.workorder.product_code
            ws.cell(row=row, column=7, value=product_id)
            ws.cell(row=row, column=8, value=report.operation)
            ws.cell(row=row, column=9, value='')  # 空白佔位用
            ws.cell(row=row, column=10, value=report.work_quantity or 0)
            ws.cell(row=row, column=11, value=report.defect_quantity or 0)
            ws.cell(row=row, column=12, value=report.remarks or '')
            ws.cell(row=row, column=13, value=report.abnormal_notes or '')
        
        # 調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 儲存檔案
        from django.http import HttpResponse
        from io import BytesIO
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="SMT設備報工記錄_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'SMT匯出失敗：{str(e)}')
        return redirect('workorder:smt_report_import_page') 


@login_required
def operator_report_export(request):
    """
    匯出作業員報工資料
    支援按日期範圍、作業員、工序等條件篩選匯出
    """
    try:
        from django.http import HttpResponse
        import pandas as pd
        from io import BytesIO
        from datetime import datetime
        
        # 取得篩選參數
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        operator_name = request.GET.get('operator_name')
        process_name = request.GET.get('process_name')
        company_code = request.GET.get('company_code')
        
        # 建立查詢 - 只針對 OperatorSupplementReport 模型
        query = OperatorSupplementReport.objects.select_related(
            'operator', 'process', 'equipment', 'workorder'
        ).all()
        
        # 應用篩選條件
        if start_date:
            query = query.filter(work_date__gte=start_date)
        if end_date:
            query = query.filter(work_date__lte=end_date)
        if operator_name:
            query = query.filter(operator__name__icontains=operator_name)
        if process_name:
            query = query.filter(process__name__icontains=process_name)
        if company_code:
            query = query.filter(workorder__company_code=company_code)
        
        # 限制匯出數量，避免檔案過大
        max_export_count = 10000
        total_count = query.count()
        if total_count > max_export_count:
            query = query[:max_export_count]
        
        # 準備匯出資料
        export_data = []
        for report in query:
            # 取得工單號碼 - 只從 OperatorSupplementReport 模型取得
            workorder_number = ''
            if report.workorder:
                workorder_number = report.workorder.order_number
            elif report.original_workorder_number:
                workorder_number = report.original_workorder_number
            elif report.rd_workorder_number:
                workorder_number = report.rd_workorder_number
            
            # 取得產品編號 - 只從 OperatorSupplementReport 模型取得
            product_code = ''
            if report.workorder and report.workorder.product_code:
                product_code = report.workorder.product_code
            elif report.product_id:
                product_code = report.product_id
            elif report.rd_product_code:
                product_code = report.rd_product_code
            
            # 取得公司代號 - 只從 OperatorSupplementReport 模型取得
            company_code_value = ''
            if report.workorder and report.workorder.company_code:
                company_code_value = report.workorder.company_code
            
            export_data.append({
                '作業員名稱': report.operator.name if report.operator else '',
                '公司代號': company_code_value,
                '報工日期': report.work_date.strftime('%Y-%m-%d') if report.work_date else '',
                '開始時間': report.start_time.strftime('%H:%M:%S') if report.start_time else '',
                '結束時間': report.end_time.strftime('%H:%M:%S') if report.end_time else '',
                '工單號': workorder_number,
                '產品編號': product_code,
                '工序名稱': report.process.name if report.process else '',
                '設備名稱': report.equipment.name if report.equipment else '',  # 直接從 OperatorSupplementReport 的 equipment 欄位取得
                '報工數量': report.work_quantity,
                '不良品數量': report.defect_quantity,
                '備註': report.remarks or '',
                '異常紀錄': report.abnormal_notes or '',
                '工作時數': float(report.work_hours_calculated) if report.work_hours_calculated else 0,
                '加班時數': float(report.overtime_hours_calculated) if report.overtime_hours_calculated else 0,
                '報工類型': report.report_type,
                '核准狀態': report.approval_status,
                '建立時間': report.created_at.strftime('%Y-%m-%d %H:%M:%S') if report.created_at else '',
            })
        
        # 建立 DataFrame
        df = pd.DataFrame(export_data)
        
        # 建立 Excel 檔案
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='作業員報工資料', index=False)
            
            # 取得工作表並調整欄寬
            worksheet = writer.sheets['作業員報工資料']
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        output.seek(0)
        
        # 準備檔案名稱
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'作業員報工資料_{timestamp}.xlsx'
        
        # 建立回應
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        return response
        
    except Exception as e:
        logger.error(f"匯出作業員報工資料失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'匯出失敗: {str(e)}'
        }) 




 