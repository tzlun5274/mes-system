"""
工單模組視圖檔案
注意：此檔案中的函數式視圖已被標記為棄用，建議使用新的類別視圖
新的類別視圖位於 workorder/views/ 目錄下
"""

import logging
import warnings
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from django.core.paginator import Paginator
from .models import (
    WorkOrder,
    WorkOrderProcess,
    WorkOrderProcessLog,
    WorkOrderAssignment,
    DispatchLog,
)

# 導入子模組的模型
from .workorder_erp.models import PrdMKOrdMain, SystemConfig, CompanyOrder
# from .tasks import get_standard_processes
from .forms import WorkOrderForm
# from .forms import from django.contrib import messages

from datetime import datetime, timedelta, date
from django.db import models
import psycopg2
from django.core.management.base import BaseCommand
from django.core.management import call_command
from django.core.serializers.json import DjangoJSONEncoder
from django.http import JsonResponse
from django.utils import timezone

from process.models import (
    ProductProcessRoute,
    ProcessEquipment,
    Operator,
    OperatorSkill,
    ProcessName,
    ProductProcessStandardCapacity,
)
from django.views.decorators.http import require_GET

from equip.models import Equipment
from django.contrib.auth.decorators import login_required, user_passes_test
from collections import defaultdict
from django.views.decorators.http import require_POST
from django import forms
from django.conf import settings
from django.db.models import Q
from django.utils.decorators import method_decorator
from django.views import View
import json
from django.views.decorators.csrf import ensure_csrf_cookie
# 完工判斷服務已移除，避免資料汙染
# from .services.completion_service import WorkOrderCompletionService

# 棄用警告函數
def deprecated_warning(func_name):
    """顯示棄用警告"""
    warnings.warn(
        f"函數 {func_name} 已被標記為棄用，請使用新的類別視圖替代。"
        f"新的類別視圖位於 workorder/views/ 目錄下。",
        DeprecationWarning,
        stacklevel=2
    )

# 設定工單管理模組的日誌記錄器
workorder_logger = logging.getLogger("workorder")
workorder_handler = logging.FileHandler("/var/log/mes/workorder.log")
workorder_handler.setFormatter(
    logging.Formatter("%(levelname)s %(asctime)s %(module)s %(message)s")
)
workorder_logger.addHandler(workorder_handler)
workorder_logger.setLevel(logging.INFO)

# Create your views here.

# 工單列表 (已棄用 - 請使用 WorkOrderListView)
def index(request):
    deprecated_warning('index')
    workorders = WorkOrder.objects.all().order_by("created_at")
    
    # 修正後的生產中工單數量計算邏輯
    # 生產中工單 = 所有工單（不管狀態）
    # 這是一個流動式的統計，反映整個工單資料表的狀態
    
    # 計算生產中工單數量（所有工單）
    active_workorders_count = WorkOrder.objects.count()
    
    context = {
        "workorders": workorders,
        "is_debug": settings.DEBUG,  # 傳遞是否為測試環境
        "active_workorders_count": active_workorders_count,
    }
    return render(request, "workorder/workorder/workorder_list.html", context)

# 工單詳情 (已棄用 - 請使用 WorkOrderDetailView)
def detail(request, pk):
    deprecated_warning('detail')
    workorder = get_object_or_404(WorkOrder, pk=pk)
    
    # 計算統計數據
    completed_processes_count = workorder.processes.filter(status='completed').count()
    in_progress_processes_count = workorder.processes.filter(status='in_progress').count()
    
    context = {
        "workorder": workorder,
        "completed_processes_count": completed_processes_count,
        "in_progress_processes_count": in_progress_processes_count,
    }
    return render(request, "workorder/workorder/workorder_detail.html", context)

# 工單列表視圖 (已棄用 - 請使用 WorkOrderListView)
def list_view(request):
    deprecated_warning('list_view')
    workorders = WorkOrder.objects.all().order_by("created_at")
    
    # 搜尋功能
    search = request.GET.get('search', '')
    if search:
        workorders = workorders.filter(
            Q(order_number__icontains=search) |
            Q(product_code__icontains=search) |
            Q(company_code__icontains=search)
        )
    
    # 分頁
    paginator = Paginator(workorders, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        "page_obj": page_obj,
        "search": search,
        "total_count": workorders.count(),
    }
    return render(request, "workorder/workorder/workorder_list.html", context)

# 新增工單
def create(request):
    error = None
    if request.method == "POST":
        form = WorkOrderForm(request.POST)
        if form.is_valid():
            company_code = form.cleaned_data["company_code"]
            order_number = form.cleaned_data["order_number"]
            # 檢查公司代號+工單編號是否重複
            if WorkOrder.objects.filter(
                company_code=company_code, order_number=order_number
            ).exists():
                error = "公司代號＋工單編號已存在，不能重複新增！"
            else:
                workorder = form.save()
                # 新增成功後自動導向工序明細頁面
                return redirect(
                    reverse("workorder:workorder_process_detail", args=[workorder.id])
                )
    else:
        form = WorkOrderForm()
    return render(
        request, "workorder/workorder/workorder_form.html", {"form": form, "error": error}
    )

@require_GET
@login_required
def get_company_order_info(request):
    """
    AJAX 視圖：根據產品編號取得公司製令單資訊
    回傳工單號與數量
    """
    product_id = request.GET.get("product_id")
    company_code = request.GET.get("company_code")

    if not product_id:
        return JsonResponse({"status": "error", "message": "請提供產品編號"})

    try:
        # 查詢公司製令單
        query = CompanyOrder.objects.filter(product_id=product_id, is_converted=False)
        if company_code:
            query = query.filter(company_code=company_code)

        company_order = query.first()

        if company_order:
            return JsonResponse(
                {
                    "status": "success",
                    "data": {
                        "mkordno": company_order.mkordno,
                        "prodt_qty": company_order.prodt_qty,
                        "company_code": company_order.company_code,
                    },
                }
            )
        else:
            return JsonResponse(
                {
                    "status": "error",
                    "message": f"找不到產品編號 {product_id} 的未轉換製令單",
                }
            )

    except Exception as e:
        return JsonResponse({"status": "error", "message": f"查詢失敗：{str(e)}"})

# 編輯工單
def edit(request, pk):
    workorder = get_object_or_404(WorkOrder, pk=pk)
    if request.method == "POST":
        form = WorkOrderForm(request.POST, instance=workorder)
        if form.is_valid():
            form.save()
            return redirect(reverse("workorder:index"))
    else:
        form = WorkOrderForm(instance=workorder)
    return render(
        request, "workorder/workorder/workorder_form.html", {"form": form, "edit_mode": True}
    )

# 刪除工單
def delete(request, pk):
    workorder = get_object_or_404(WorkOrder, pk=pk)
    if request.method == "POST":
        try:
            # 直接刪除工單（會自動刪除相關的工序）
            workorder.delete()
            messages.success(request, "工單刪除成功！")
            return redirect(reverse("workorder:index"))
            
        except Exception as e:
            messages.error(request, f"刪除工單時發生錯誤：{str(e)}")
            return redirect(reverse("workorder:index"))
            
    return render(
        request, "workorder/workorder/workorder_confirm_delete.html", {"workorder": workorder}
    )

# 刪除所有未派工工單（支援 all=1 參數時同時刪除已派工工單）
def delete_pending_workorders(request):
    """
    刪除所有未派工工單（狀態為 pending）
    若帶有 ?all=1 參數，則同時刪除已派工工單（狀態為 in_progress）
    """
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "只有管理員可以執行此操作")
        return redirect("workorder:index")

    delete_all = request.GET.get("all") == "1"

    if request.method == "POST":
        try:
            if delete_all:
                # 同時刪除 pending 和 in_progress
                pending_count = WorkOrder.objects.filter(status="pending").count()
                in_progress_count = WorkOrder.objects.filter(status="in_progress").count()
                
                # 刪除工單
                WorkOrder.objects.filter(status__in=["pending", "in_progress"]).delete()
                
                messages.success(
                    request,
                    f"成功刪除所有工單！共刪除 {pending_count + in_progress_count} 個工單（待生產：{pending_count}，生產中：{in_progress_count}）"
                )
            else:
                # 只刪除 pending
                pending_count = WorkOrder.objects.filter(status="pending").count()
                
                # 刪除工單
                WorkOrder.objects.filter(status="pending").delete()
                
                messages.success(
                    request, 
                    f"成功刪除所有未派工工單！共刪除 {pending_count} 個工單"
                )
            return redirect("workorder_dispatch:dispatch_list")
        except Exception as e:
            messages.error(request, f"刪除工單時發生錯誤：{str(e)}")
            return redirect("workorder_dispatch:dispatch_list")

    # GET 請求顯示確認頁面
    pending_count = WorkOrder.objects.filter(status="pending").count()
    in_progress_count = (
        WorkOrder.objects.filter(status="in_progress").count() if delete_all else 0
    )
    return render(
        request,
        "workorder/workorder/delete_pending_confirm.html",
        {
            "pending_count": pending_count,
            "in_progress_count": in_progress_count,
            "delete_all": delete_all,
        },
    )

# 刪除所有已派工工單
def delete_in_progress_workorders(request):
    """
    刪除所有已派工工單（狀態為 in_progress），支援 AJAX 回傳 JSON
    """
    is_ajax = (
        request.headers.get("x-requested-with") == "XMLHttpRequest"
        or request.content_type == "application/json"
        or request.headers.get("Accept") == "application/json"
    )
    if not (request.user.is_staff or request.user.is_superuser):
        msg = "只有管理員可以執行此操作"
        if is_ajax:
            return JsonResponse({"status": "danger", "message": msg})
        messages.error(request, msg)
        return redirect("workorder:index")

    if request.method == "POST":
        try:
            in_progress_count = WorkOrder.objects.filter(status="in_progress").count()
            WorkOrder.objects.filter(status="in_progress").update(status="pending")
            msg = f"成功停止所有生產中工單！共處理 {in_progress_count} 個工單"
            if is_ajax:
                return JsonResponse({"status": "success", "message": msg})
            messages.success(request, msg)
            return redirect("workorder_dispatch:dispatch_list")
        except Exception as e:
            msg = f"停止所有生產中工單時發生錯誤：{str(e)}"
            if is_ajax:
                return JsonResponse({"status": "danger", "message": msg})
            messages.error(request, msg)
            return redirect("workorder_dispatch:dispatch_list")

    # GET 請求顯示確認頁面
    in_progress_count = WorkOrder.objects.filter(status="in_progress").count()
    return render(
        request,
        "workorder/workorder/delete_in_progress_confirm.html",
        {"in_progress_count": in_progress_count},
    )

# 派工單管理頁（可重用原本工單列表）
def dispatch_list(request):
    """
    派工單管理頁：分成待生產（pending）和已派工（in_progress）兩個區塊
    並提供每個工單的補登記錄數量與待核准數量
    包含所有類型的工單：正式生產、PP試產(330-)、重工(339-)等
    """
    from process.models import ProductProcessRoute

    # 只顯示待生產和生產中的工單，不顯示已完成工單
    all_orders = WorkOrder.objects.filter(status__in=["pending", "in_progress"]).order_by("created_at")
    
    # 按狀態分組
    pending_orders = [order for order in all_orders if order.status == "pending"]
    in_progress_orders = [order for order in all_orders if order.status == "in_progress"]
    completed_orders = []  # 派工單管理頁面不顯示已完成工單
    other_orders = []  # 派工單管理頁面不顯示其他狀態工單

    # 計算統計數據（只計算待生產和生產中的工單）
    total_workorders = WorkOrder.objects.filter(status__in=["pending", "in_progress"]).count()
    pending_count = len(pending_orders)
    in_progress_count = len(in_progress_orders)
    completed_count = 0  # 派工單管理頁面不顯示已完成工單
    other_count = 0  # 派工單管理頁面不顯示其他狀態工單
    
    # 計算工藝路線已設定總數
    process_route_set_count = 0
    # 計算分配資訊已分配總數
    assignment_set_count = 0
    
    # 檢查待生產和生產中工單的工藝路線和分配狀態
    active_workorders = WorkOrder.objects.filter(status__in=["pending", "in_progress"])
    for workorder in active_workorders:
        # 檢查工藝路線設定
        has_process_route = ProductProcessRoute.objects.filter(
            product_id=workorder.product_code
        ).exists()
        if has_process_route:
            process_route_set_count += 1
            
        # 檢查分配資訊
        has_assignment = (
            WorkOrderProcess.objects.filter(workorder=workorder)
            .filter(
                models.Q(assigned_operator__isnull=False, assigned_operator__gt="")
                | models.Q(assigned_equipment__isnull=False, assigned_equipment__gt="")
            )
            .exists()
        )
        if has_assignment:
            assignment_set_count += 1

    # 檢查每個工單的工序設定狀態和分配資訊
    for order in list(pending_orders) + list(in_progress_orders):
        order.has_process_route = ProductProcessRoute.objects.filter(
            product_id=order.product_code
        ).exists()
        # 新增：檢查工序明細裡有沒有分配作業員或設備
        order.has_assignment = (
            WorkOrderProcess.objects.filter(workorder=order)
            .filter(
                models.Q(assigned_operator__isnull=False, assigned_operator__gt="")
                | models.Q(assigned_equipment__isnull=False, assigned_equipment__gt="")
            )
            .exists()
        )
        # 移除 has_process_detail 與 process_configured 判斷，回復只看 has_process_route
        # 只有同時有工藝路線且有工序明細才算「已設定」
        # order.process_configured = order.has_process_route and order.has_process_detail
        # 預載入分配資訊
        order.workorderassignment_set.all()

        # 計算作業員工作負載
        for assignment in order.workorderassignment_set.all():
            if assignment.operator_id:
                assignment.operator_current_load = WorkOrderAssignment.objects.filter(
                    operator_id=assignment.operator_id, workorder__status="in_progress"
                ).count()

            # 添加設備名稱屬性
            if assignment.equipment_id:
                try:
                    equipment = Equipment.objects.get(id=assignment.equipment_id)
                    assignment.equipment_name = equipment.name
                except (Equipment.DoesNotExist, ValueError):
                    assignment.equipment_name = f"未知設備({assignment.equipment_id})"
            else:
                assignment.equipment_name = "未分配"

    return render(
        request,
        "workorder/dispatch/dispatch_list.html",
        {
            "pending_orders": pending_orders,
            "in_progress_orders": in_progress_orders,
            "completed_orders": completed_orders,
            "other_orders": other_orders,
            "total_workorders": total_workorders,
            "pending_count": pending_count,
            "in_progress_count": in_progress_count,
            "completed_count": completed_count,
            "other_count": other_count,
            "process_route_set_count": process_route_set_count,
            "assignment_set_count": assignment_set_count,
        },
    )

# 公司製令單列表（依公司分群）
# 公司製令單管理功能已移至 CompanyOrderListView 類別視圖
# 此函數已廢棄，保留註解以說明功能遷移

# 手動同步製令單
def manual_sync_orders(request):
    """
    手動同步各公司製令單到 CompanyOrder 表
    """
    if not (request.user.is_staff or request.user.is_superuser):
        workorder_logger.warning(
            f"非管理員 {request.user} 嘗試手動同步製令單，已拒絕。IP: {request.META.get('REMOTE_ADDR')}"
        )
        messages.error(request, "只有管理員可以執行此操作")
        return redirect("workorder:company_orders")

    if request.method != "POST":
        return redirect("workorder:company_orders")

    try:
        # 執行同步命令
        call_command("sync_pending_workorders")
        workorder_logger.info(
            f"管理員 {request.user} 手動同步公司製令單成功。IP: {request.META.get('REMOTE_ADDR')}"
        )
        messages.success(request, "手動同步製令單完成！")
    except Exception as e:
        workorder_logger.error(
            f"管理員 {request.user} 手動同步公司製令單失敗：{e}。IP: {request.META.get('REMOTE_ADDR')}"
        )
        messages.error(request, f"同步失敗：{e}")

    return redirect("workorder:company_orders")

# 手動轉換工單
def manual_convert_orders(request):
    """
    手動轉換製令單為 MES 工單
    支援 AJAX 請求，回傳 JSON。
    新增：在轉換時自動分配作業員和設備
    """
    from process.models import ProductProcessRoute, ProductProcessStandardCapacity, Operator, OperatorSkill, ProcessEquipment
    from equip.models import Equipment

    is_ajax = (
        request.headers.get("x-requested-with") == "XMLHttpRequest"
        or request.content_type == "application/json"
        or request.headers.get("Accept") == "application/json"
    )
    if request.method != "POST":
        msg = "請用 POST 方式呼叫手動轉換"
        if is_ajax:
            return JsonResponse({"status": "danger", "message": msg})
        messages.error(request, msg)
        return redirect("workorder:company_orders")

    try:
        # 取得未轉換的製令單
        pending_orders = CompanyOrder.objects.filter(is_converted=False)
        if not pending_orders.exists():
            msg = "沒有需要轉換的製令單！"
            if is_ajax:
                return JsonResponse({"status": "warning", "message": msg})
            messages.warning(request, msg)
            return redirect("workorder:company_orders")

        count_converted = 0
        count_processes_created = 0
        count_auto_assigned = 0

        for company_order in pending_orders:
            # 檢查工單是否已存在（使用公司代號和製令單號的組合）
            existing_workorder = WorkOrder.objects.filter(
                company_code=company_order.company_code,
                order_number=company_order.mkordno
            ).first()
            
            if existing_workorder:
                # 如果工單已存在，跳過並標記為已轉換
                company_order.is_converted = True
                company_order.save()
                print(f"⚠️ 工單已存在，跳過轉換：{company_order.mkordno}")
                continue
            
            # 建立工單（使用製令單號作為工單號碼）
            workorder = WorkOrder.objects.create(
                order_number=company_order.mkordno,  # 直接使用製令單號
                product_code=company_order.product_id,  # 使用 product_id
                quantity=company_order.prodt_qty,  # 使用 prodt_qty
                status="pending",
                company_code=company_order.company_code,
            )
            count_converted += 1

            # 建立工序明細
            try:
                # 從產品工藝路線取得工序資料
                routes = ProductProcessRoute.objects.filter(
                    product_id=workorder.product_code
                ).order_by("step_order")

                if routes.exists():
                    # 使用產品工藝路線建立工序明細
                    for route in routes:
                        # 查詢標準產能資料
                        capacity_data = ProductProcessStandardCapacity.objects.filter(
                            product_code=workorder.product_code,
                            process_name=route.process_name.name,
                            is_active=True
                        ).order_by('-version').first()
                        
                        # 使用標準產能或預設值
                        target_hourly_output = (
                            capacity_data.standard_capacity_per_hour 
                            if capacity_data else 1000
                        )
                        
                        # 建立工序明細
                        process = WorkOrderProcess.objects.create(
                            workorder=workorder,
                            process_name=route.process_name.name,
                            step_order=route.step_order,
                            planned_quantity=workorder.quantity,
                            target_hourly_output=target_hourly_output,
                        )
                        count_processes_created += 1

                        # ====== 自動分配作業員和設備 ======
                        assigned_operator = None
                        assigned_equipment = None

                        # 1. 自動分配作業員：根據工序名稱找到有對應技能的作業員
                        try:
                            # 找到有該工序技能的作業員
                            skilled_operators = Operator.objects.filter(
                                skills__process_name__name=route.process_name.name
                            ).distinct()
                            
                            if skilled_operators.exists():
                                # 選擇工作負載最輕的作業員
                                operator_loads = {}
                                for op in skilled_operators:
                                    # 計算該作業員目前的工作負載（進行中的工序數量）
                                    current_load = WorkOrderProcess.objects.filter(
                                        assigned_operator=op.name,
                                        status='in_progress'
                                    ).count()
                                    operator_loads[op] = current_load
                                
                                # 選擇負載最輕的作業員
                                assigned_operator = min(operator_loads.items(), key=lambda x: x[1])[0]
                                process.assigned_operator = assigned_operator.name
                                print(f"✅ 自動分配作業員：{assigned_operator.name} (工序：{route.process_name.name})")
                            else:
                                print(f"⚠️ 找不到有 {route.process_name.name} 技能的作業員")
                        except Exception as e:
                            print(f"⚠️ 分配作業員失敗：{str(e)}")

                        # 2. 自動分配設備：根據工序名稱找到可用的設備
                        try:
                            # 檢查工序是否有指定可用設備
                            if route.usable_equipment_ids:
                                # 解析可用設備ID列表
                                equipment_ids = [
                                    int(eid.strip()) 
                                    for eid in route.usable_equipment_ids.split(',') 
                                    if eid.strip().isdigit()
                                ]
                                
                                if equipment_ids:
                                    # 找到可用且狀態為閒置的設備
                                    available_equipments = Equipment.objects.filter(
                                        id__in=equipment_ids,
                                        status='idle'
                                    )
                                    
                                    if available_equipments.exists():
                                        # 選擇第一個可用設備
                                        assigned_equipment = available_equipments.first()
                                        process.assigned_equipment = assigned_equipment.name
                                        print(f"✅ 自動分配設備：{assigned_equipment.name} (工序：{route.process_name.name})")
                                    else:
                                        print(f"⚠️ 工序 {route.process_name.name} 的指定設備都不可用")
                                else:
                                    print(f"⚠️ 工序 {route.process_name.name} 的可用設備ID格式錯誤")
                            else:
                                # 如果沒有指定設備，嘗試從工序設備對應表找
                                process_equipments = ProcessEquipment.objects.filter(
                                    process_name__name=route.process_name.name
                                )
                                
                                if process_equipments.exists():
                                    equipment_ids = [pe.equipment_id for pe in process_equipments]
                                    available_equipments = Equipment.objects.filter(
                                        id__in=equipment_ids,
                                        status='idle'
                                    )
                                    
                                    if available_equipments.exists():
                                        assigned_equipment = available_equipments.first()
                                        process.assigned_equipment = assigned_equipment.name
                                        print(f"✅ 自動分配設備：{assigned_equipment.name} (工序：{route.process_name.name})")
                                    else:
                                        print(f"⚠️ 工序 {route.process_name.name} 的對應設備都不可用")
                                else:
                                    print(f"⚠️ 工序 {route.process_name.name} 沒有對應的設備設定")
                        except Exception as e:
                            print(f"⚠️ 分配設備失敗：{str(e)}")

                        # 3. 如果有分配到作業員或設備，更新狀態並記錄
                        if assigned_operator or assigned_equipment:
                            process.save()
                            count_auto_assigned += 1
                            
                            # 記錄分配日誌
                            try:
                                WorkOrderProcessLog.objects.create(
                                    workorder_process=process,
                                    action='auto_assignment',
                                    operator=assigned_operator.name if assigned_operator else None,
                                    equipment=assigned_equipment.name if assigned_equipment else None,
                                    notes=f"製令轉工單時自動分配 - 工序：{route.process_name.name}",
                                    quantity_before=0,
                                    quantity_after=0,
                                )
                            except Exception as e:
                                print(f"⚠️ 記錄分配日誌失敗：{str(e)}")

                else:
                    # 使用標準工序建立工序明細
                    standard_processes = [
                        {"name": "SMT 貼片", "target_hourly_output": 1000},
                        {"name": "測試", "target_hourly_output": 500},
                        {"name": "包裝", "target_hourly_output": 200},
                    ]

                    for step_order, process_info in enumerate(
                        standard_processes, 1
                    ):
                        WorkOrderProcess.objects.create(
                            workorder=workorder,
                            process_name=process_info["name"],
                            step_order=step_order,
                            planned_quantity=workorder.quantity,
                            target_hourly_output=process_info[
                                "target_hourly_output"
                            ],
                        )
                        count_processes_created += 1
            except Exception as e:
                # 如果建立工序明細失敗，記錄錯誤但不影響工單轉換
                print(f"建立工序明細失敗 (工單: {workorder.order_number}): {e}")

        # 更新 CompanyOrder 的轉換狀態
        if count_converted > 0:
            CompanyOrder.objects.filter(is_converted=False).update(is_converted=True)
            
            # 更新成功訊息，包含自動分配資訊
            auto_assignment_msg = f"，並自動分配 {count_auto_assigned} 個工序的作業員和設備" if count_auto_assigned > 0 else ""
            
            messages.success(
                request,
                f"手動轉換完成！共轉換 {count_converted} 筆製令單為 MES 工單，並自動建立 {count_processes_created} 個工序明細{auto_assignment_msg}",
            )
        else:
            messages.info(request, "沒有需要轉換的製令單")

        workorder_logger.info(
            f"管理員 {request.user} 手動轉換工單成功，轉換 {count_converted} 筆，建立工序 {count_processes_created} 筆，自動分配 {count_auto_assigned} 筆。IP: {request.META.get('REMOTE_ADDR')}"
        )
    except Exception as e:
        workorder_logger.error(
            f"管理員 {request.user} 手動轉換工單失敗：{e}。IP: {request.META.get('REMOTE_ADDR')}"
        )
        messages.error(request, f"轉換失敗：{e}")

    return redirect("workorder:company_orders")

# 派工單工序明細管理
def workorder_process_detail(request, workorder_id):
    """
    派工單工序明細頁面：顯示工單的所有工序及執行狀況
    完工工單：從實際報工資料生成工序明細
    未完工工單：顯示預先建立的工序明細
    """
    workorder = get_object_or_404(WorkOrder, pk=workorder_id)
    
    # 檢查是否為完工工單
    if workorder.status == "completed":
        # 完工工單：從實際報工資料生成工序明細
        processes = generate_completed_workorder_processes(workorder)
        has_process_route = True  # 完工工單一定有實際工序
    else:
        # 未完工工單：顯示預先建立的工序明細
        processes = WorkOrderProcess.objects.filter(workorder=workorder).order_by("step_order")
        
        # 檢查產品是否有工藝路線設定
        from process.models import ProductProcessRoute
        has_process_route = ProductProcessRoute.objects.filter(
            product_id=workorder.product_code
        ).exists()

    # 計算整體進度
    total_planned = sum(p.planned_quantity for p in processes)
    total_completed = sum(p.completed_quantity for p in processes)
    overall_progress = round(
        (total_completed / total_planned * 100) if total_planned > 0 else 0, 2
    )

    # 取得所有工序名稱供下拉選單使用
    from process.models import ProcessName
    process_names = ProcessName.objects.all().order_by('name')

    return render(
        request,
        "workorder/process/workorder_process_detail.html",
        {
            "workorder": workorder,
            "processes": processes,
            "total_planned": total_planned,
            "total_completed": total_completed,
            "overall_progress": overall_progress,
            "has_process_route": has_process_route,
            "process_names": process_names,
            "is_completed_workorder": workorder.status == "completed",
        },
    )

def generate_completed_workorder_processes(workorder):
    """
    為完工工單從實際報工資料生成工序明細
    整合作業員報工和SMT報工資料
    """
    from collections import defaultdict
    from datetime import datetime, time
    
    # 收集所有報工資料
    process_data = defaultdict(lambda: {
        'process_name': '',
        'planned_quantity': workorder.quantity,
        'completed_quantity': 0,
        'assigned_operators': set(),
        'assigned_equipments': set(),
        'actual_start_time': None,
        'actual_end_time': None,
        'total_work_hours': 0,
        'total_overtime_hours': 0,
        'defect_quantity': 0,
        'abnormal_notes': [],
        'step_order': 0
    })
    
    # 1. 收集作業員報工資料
    # operator_reports = OperatorSupplementReport.objects.filter(
    #     workorder=workorder,
    #     approval_status='approved'
    # ).select_related('operator', 'process', 'equipment')
    operator_reports = []
    
    for report in operator_reports:
        process_name = report.process.name if report.process else report.operation
        if process_name:
            process_data[process_name]['process_name'] = process_name
            process_data[process_name]['completed_quantity'] += report.work_quantity
            process_data[process_name]['defect_quantity'] += report.defect_quantity
            
            if report.operator:
                process_data[process_name]['assigned_operators'].add(report.operator.name)
            if report.equipment:
                process_data[process_name]['assigned_equipments'].add(report.equipment.name)
            
            # 計算時間
            report_start_time = datetime.combine(report.work_date, report.start_time)
            report_end_time = datetime.combine(report.work_date, report.end_time)
            
            if not process_data[process_name]['actual_start_time'] or report_start_time < process_data[process_name]['actual_start_time']:
                process_data[process_name]['actual_start_time'] = report_start_time
            if not process_data[process_name]['actual_end_time'] or report_end_time > process_data[process_name]['actual_end_time']:
                process_data[process_name]['actual_end_time'] = report_end_time
            
            process_data[process_name]['total_work_hours'] += float(report.work_hours_calculated or 0)
            process_data[process_name]['total_overtime_hours'] += float(report.overtime_hours_calculated or 0)
            
            if report.abnormal_notes:
                process_data[process_name]['abnormal_notes'].append(report.abnormal_notes)
    
    # 2. 收集SMT報工資料
    # smt_reports = SMTSupplementReport.objects.filter(
    #     workorder=workorder,
    #     approval_status='approved'
    # ).select_related('equipment')
    smt_reports = []
    
    for report in smt_reports:
        process_name = report.operation
        if process_name:
            process_data[process_name]['process_name'] = process_name
            process_data[process_name]['completed_quantity'] += report.work_quantity
            process_data[process_name]['defect_quantity'] += report.defect_quantity
            
            if report.equipment:
                process_data[process_name]['assigned_equipments'].add(report.equipment.name)
            
            # 計算時間
            report_start_time = datetime.combine(report.work_date, report.start_time)
            report_end_time = datetime.combine(report.work_date, report.end_time)
            
            if not process_data[process_name]['actual_start_time'] or report_start_time < process_data[process_name]['actual_start_time']:
                process_data[process_name]['actual_start_time'] = report_start_time
            if not process_data[process_name]['actual_end_time'] or report_end_time > process_data[process_name]['actual_end_time']:
                process_data[process_name]['actual_end_time'] = report_end_time
            
            process_data[process_name]['total_work_hours'] += float(report.work_hours_calculated or 0)
            process_data[process_name]['total_overtime_hours'] += float(report.overtime_hours_calculated or 0)
            
            if report.abnormal_notes:
                process_data[process_name]['abnormal_notes'].append(report.abnormal_notes)
    
    # 3. 生成工序明細物件
    processes = []
    step_order = 1
    
    # 按開始時間排序工序
    sorted_processes = sorted(
        process_data.items(),
        key=lambda x: x[1]['actual_start_time'] or datetime.max
    )
    
    for process_name, data in sorted_processes:
        if data['completed_quantity'] > 0:  # 只顯示有實際產出的工序
            # 創建一個類似 WorkOrderProcess 的物件
            process_obj = type('CompletedProcess', (), {
                'step_order': step_order,
                'process_name': data['process_name'],
                'planned_quantity': data['planned_quantity'],
                'completed_quantity': data['completed_quantity'],
                'status': 'completed',
                'assigned_operator': ', '.join(data['assigned_operators']) if data['assigned_operators'] else '-',
                'assigned_equipment': ', '.join(data['assigned_equipments']) if data['assigned_equipments'] else '-',
                'actual_start_time': data['actual_start_time'],
                'actual_end_time': data['actual_end_time'],
                'capacity_multiplier': len(data['assigned_operators']) or 1,
                'target_hourly_output': 0,  # 完工工單不需要目標產能
                'estimated_hours': data['total_work_hours'] + data['total_overtime_hours'],
                'defect_quantity': data['defect_quantity'],
                'abnormal_notes': '; '.join(data['abnormal_notes']) if data['abnormal_notes'] else '',
                'completion_rate': f"{(data['completed_quantity'] / data['planned_quantity'] * 100):.1f}%" if data['planned_quantity'] > 0 else "100%",
                'get_status_display': lambda: "已完成",
                'equipment_display_name': ', '.join(data['assigned_equipments']) if data['assigned_equipments'] else '-',
            })()
            
            processes.append(process_obj)
            step_order += 1
    
    return processes

def create_workorder_processes(request, workorder_id):
    """
    為指定工單建立工序明細
    支援 AJAX 請求，回傳 JSON。
    新增：在建立工序時自動分配作業員和設備
    """
    from process.models import ProductProcessRoute, ProductProcessStandardCapacity, Operator, OperatorSkill, ProcessEquipment
    from equip.models import Equipment

    is_ajax = (
        request.headers.get("x-requested-with") == "XMLHttpRequest"
        or request.content_type == "application/json"
        or request.headers.get("Accept") == "application/json"
    )

    try:
        workorder = WorkOrder.objects.get(id=workorder_id)
    except WorkOrder.DoesNotExist:
        msg = "工單不存在！"
        if is_ajax:
            return JsonResponse({"status": "danger", "message": msg})
        messages.error(request, msg)
        return redirect("workorder:list")

    try:
        # 檢查是否已有工序明細
        existing_processes = WorkOrderProcess.objects.filter(workorder=workorder)
        if existing_processes.exists():
            msg = "此工單已有工序明細，無法重複建立！"
            if is_ajax:
                return JsonResponse({"status": "warning", "message": msg})
            messages.warning(request, msg)
            return redirect("workorder:workorder_process_detail", workorder_id=workorder_id)

        created_count = 0
        auto_assigned_count = 0

        # 從產品工藝路線取得工序資料
        routes = ProductProcessRoute.objects.filter(
            product_id=workorder.product_code
        ).order_by("step_order")

        if routes.exists():
            # 使用產品工藝路線建立工序明細
            for route in routes:
                # 查詢標準產能資料
                capacity_data = ProductProcessStandardCapacity.objects.filter(
                    product_code=workorder.product_code,
                    process_name=route.process_name.name,
                    is_active=True
                ).order_by('-version').first()
                
                # 使用標準產能或預設值
                target_hourly_output = (
                    capacity_data.standard_capacity_per_hour 
                    if capacity_data else 1000
                )
                
                # 建立工序明細
                process = WorkOrderProcess.objects.create(
                    workorder=workorder,
                    process_name=route.process_name.name,
                    step_order=route.step_order,
                    planned_quantity=workorder.quantity,
                    target_hourly_output=target_hourly_output,
                )
                created_count += 1

                # ====== 自動分配作業員和設備 ======
                assigned_operator = None
                assigned_equipment = None

                # 1. 自動分配作業員：根據工序名稱找到有對應技能的作業員
                try:
                    # 找到有該工序技能的作業員
                    skilled_operators = Operator.objects.filter(
                        skills__process_name__name=route.process_name.name
                    ).distinct()
                    
                    if skilled_operators.exists():
                        # 選擇工作負載最輕的作業員
                        operator_loads = {}
                        for op in skilled_operators:
                            # 計算該作業員目前的工作負載（進行中的工序數量）
                            current_load = WorkOrderProcess.objects.filter(
                                assigned_operator=op.name,
                                status='in_progress'
                            ).count()
                            operator_loads[op] = current_load
                        
                        # 選擇負載最輕的作業員
                        assigned_operator = min(operator_loads.items(), key=lambda x: x[1])[0]
                        process.assigned_operator = assigned_operator.name
                        print(f"✅ 自動分配作業員：{assigned_operator.name} (工序：{route.process_name.name})")
                    else:
                        print(f"⚠️ 找不到有 {route.process_name.name} 技能的作業員")
                except Exception as e:
                    print(f"⚠️ 分配作業員失敗：{str(e)}")

                # 2. 自動分配設備：根據工序名稱找到可用的設備
                try:
                    # 檢查工序是否有指定可用設備
                    if route.usable_equipment_ids:
                        # 解析可用設備ID列表
                        equipment_ids = [
                            int(eid.strip()) 
                            for eid in route.usable_equipment_ids.split(',') 
                            if eid.strip().isdigit()
                        ]
                        
                        if equipment_ids:
                            # 找到可用且狀態為閒置的設備
                            available_equipments = Equipment.objects.filter(
                                id__in=equipment_ids,
                                status='idle'
                            )
                            
                            if available_equipments.exists():
                                # 選擇第一個可用設備
                                assigned_equipment = available_equipments.first()
                                process.assigned_equipment = assigned_equipment.name
                                print(f"✅ 自動分配設備：{assigned_equipment.name} (工序：{route.process_name.name})")
                            else:
                                print(f"⚠️ 工序 {route.process_name.name} 的指定設備都不可用")
                        else:
                            print(f"⚠️ 工序 {route.process_name.name} 的可用設備ID格式錯誤")
                    else:
                        # 如果沒有指定設備，嘗試從工序設備對應表找
                        process_equipments = ProcessEquipment.objects.filter(
                            process_name__name=route.process_name.name
                        )
                        
                        if process_equipments.exists():
                            equipment_ids = [pe.equipment_id for pe in process_equipments]
                            available_equipments = Equipment.objects.filter(
                                id__in=equipment_ids,
                                status='idle'
                            )
                            
                            if available_equipments.exists():
                                assigned_equipment = available_equipments.first()
                                process.assigned_equipment = assigned_equipment.name
                                print(f"✅ 自動分配設備：{assigned_equipment.name} (工序：{route.process_name.name})")
                            else:
                                print(f"⚠️ 工序 {route.process_name.name} 的對應設備都不可用")
                        else:
                            print(f"⚠️ 工序 {route.process_name.name} 沒有對應的設備設定")
                except Exception as e:
                    print(f"⚠️ 分配設備失敗：{str(e)}")

                # 3. 如果有分配到作業員或設備，更新狀態並記錄
                if assigned_operator or assigned_equipment:
                    process.save()
                    auto_assigned_count += 1
                    
                    # 記錄分配日誌
                    try:
                        WorkOrderProcessLog.objects.create(
                            workorder_process=process,
                            action='auto_assignment',
                            operator=assigned_operator.name if assigned_operator else None,
                            equipment=assigned_equipment.name if assigned_equipment else None,
                            notes=f"建立工序時自動分配 - 工序：{route.process_name.name}",
                            quantity_before=0,
                            quantity_after=0,
                        )
                    except Exception as e:
                        print(f"⚠️ 記錄分配日誌失敗：{str(e)}")

            # 更新成功訊息，包含自動分配資訊
            auto_assignment_msg = f"，並自動分配 {auto_assigned_count} 個工序的作業員和設備" if auto_assigned_count > 0 else ""
            messages.success(
                request, f"成功建立 {created_count} 個工序明細（使用產品工藝路線）{auto_assignment_msg}"
            )
        else:
            # 使用標準工序建立工序明細
            standard_processes = get_standard_processes(workorder.product_code)

            for step_order, process_info in enumerate(standard_processes, 1):
                WorkOrderProcess.objects.create(
                    workorder=workorder,
                    process_name=process_info["name"],
                    step_order=step_order,
                    planned_quantity=workorder.quantity,
                    target_hourly_output=process_info["target_hourly_output"],
                )
                created_count += 1

            messages.success(
                request, f"成功建立 {created_count} 個工序明細（使用標準工序）"
            )

    except Exception as e:
        messages.error(request, f"建立工序明細失敗：{e}")

    return redirect("workorder:workorder_process_detail", workorder_id=workorder_id)

@require_GET
@login_required
def get_processes_only(request):
    """
    API：根據工單ID或工單編號回傳該工單所有工序（JSON格式），給前端 AJAX 用
    支援傳入數字型工單 id 或字串型工單編號（order_number）
    """
    workorder_id = request.GET.get("workorder_id")
    if not workorder_id:
        return JsonResponse({"success": False, "message": "缺少工單ID"})

    # 判斷傳入的是數字 id 還是工單編號
    workorder = None
    if workorder_id.isdigit():
        # 如果是數字，直接用 id 查
        try:
            workorder = WorkOrder.objects.get(id=int(workorder_id))
        except WorkOrder.DoesNotExist:
            return JsonResponse({"success": False, "message": "查無此工單"})
    else:
        # 如果是字串，改用工單編號查
        try:
            workorder = WorkOrder.objects.get(order_number=workorder_id)
        except WorkOrder.DoesNotExist:
            return JsonResponse({"success": False, "message": "查無此工單編號"})

    # 查詢該工單的所有工序
    processes = WorkOrderProcess.objects.filter(workorder=workorder).order_by(
        "step_order"
    )
    process_list = [
        {
            "id": p.id,
            "process_name": p.process_name,
            "step_order": p.step_order,
        }
        for p in processes
    ]
    return JsonResponse({"success": True, "processes": process_list})

# 完工工單相關函數已移除

def clear_data(request):
    """
    清除數據頁面：清除派工單、公司製令單等數據
    只有管理員可以執行此操作，需要確認頁面
    """
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "只有管理員可以執行此操作")
        return redirect("workorder:index")

    if request.method == "POST":
        try:
            # 取得要清除的數據統計
            workorder_count = WorkOrder.objects.count()
            company_order_count = CompanyOrder.objects.count()
            dispatch_log_count = DispatchLog.objects.count()

            # 清除所有數據
            WorkOrder.objects.all().delete()
            CompanyOrder.objects.all().delete()
            DispatchLog.objects.all().delete()

            messages.success(
                request,
                f"成功清除所有數據！共清除：工單 {workorder_count} 筆、公司製令單 {company_order_count} 筆、"
                f"派工記錄 {dispatch_log_count} 筆",
            )
            return redirect("workorder:index")

        except Exception as e:
            messages.error(request, f"清除數據失敗：{str(e)}")
            return redirect("workorder:clear_data")

    # GET 請求顯示確認頁面
    workorder_count = WorkOrder.objects.count()
    company_order_count = CompanyOrder.objects.count()
    dispatch_log_count = DispatchLog.objects.count()
    
    context = {
        "workorder_count": workorder_count,
        "company_order_count": company_order_count,
        "dispatch_log_count": dispatch_log_count,
        "total_count": workorder_count + company_order_count + dispatch_log_count,
    }

    return render(request, "workorder/clear_data_confirm.html", context)

def start_production(request, pk):
    """
    開始生產：將工單狀態從 pending 轉換為 in_progress
    只有管理員可以執行此操作
    """
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "只有管理員可以執行此操作")
        return redirect("workorder_dispatch:dispatch_list")

    if request.method != "POST":
        messages.error(request, "無效的請求方法")
        return redirect("workorder_dispatch:dispatch_list")

    try:
        # 取得工單
        workorder = WorkOrder.objects.get(pk=pk)

        # 檢查工單狀態
        if workorder.status != "pending":
            messages.error(
                request, f"工單 {workorder.order_number} 狀態不是待生產，無法開始生產"
            )
            return redirect("workorder_dispatch:dispatch_list")

        # 檢查是否有工序路線
        from process.models import ProductProcessRoute

        has_process_route = ProductProcessRoute.objects.filter(
            product_id=workorder.product_code
        ).exists()

        if not has_process_route:
            messages.error(
                request, f"工單 {workorder.order_number} 尚未設定工序路線，無法開始生產"
            )
            return redirect("workorder_dispatch:dispatch_list")

        # 更新工單狀態
        workorder.status = "in_progress"
        workorder.updated_at = timezone.now()
        workorder.save()

        # 記錄派工日誌
        # 注意：DispatchLog 模型沒有 action 欄位，所以我們只記錄基本資訊
        # 如果需要記錄動作類型，可以考慮使用 WorkOrderProcessLog 或添加 action 欄位到 DispatchLog
        DispatchLog.objects.create(
            workorder=workorder,
            process=workorder.processes.first(),  # 使用第一個工序作為記錄
            operator=None,  # 暫時設為 None，因為需要 Operator 對象
            quantity=workorder.quantity,
            created_by=request.user.username,
        )

        messages.success(request, f"工單 {workorder.order_number} 已成功開始生產")

    except WorkOrder.DoesNotExist:
        messages.error(request, "工單不存在")
    except Exception as e:
        messages.error(request, f"開始生產失敗：{str(e)}")

    return redirect("workorder_dispatch:dispatch_list")

def stop_production(request, pk):
    """
    停止生產：將工單狀態從 in_progress 轉換為 pending
    只有管理員可以執行此操作
    """
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "只有管理員可以執行此操作")
        return redirect("workorder_dispatch:dispatch_list")

    if request.method != "POST":
        messages.error(request, "無效的請求方法")
        return redirect("workorder_dispatch:dispatch_list")

    try:
        # 取得工單
        workorder = WorkOrder.objects.get(pk=pk)

        # 檢查工單狀態
        if workorder.status != "in_progress":
            messages.error(
                request, f"工單 {workorder.order_number} 狀態不是生產中，無法停止生產"
            )
            return redirect("workorder_dispatch:dispatch_list")

        # 更新工單狀態
        workorder.status = "pending"
        workorder.updated_at = timezone.now()
        workorder.save()

        # 記錄派工日誌
        # 注意：DispatchLog 模型沒有 action 欄位，所以我們只記錄基本資訊
        DispatchLog.objects.create(
            workorder=workorder,
            process=workorder.processes.first(),  # 使用第一個工序作為記錄
            operator=None,  # 暫時設為 None，因為需要 Operator 對象
            quantity=workorder.quantity,
            created_by=request.user.username,
        )

        messages.success(request, f"工單 {workorder.order_number} 已成功停止生產")

    except WorkOrder.DoesNotExist:
        messages.error(request, "工單不存在")
    except Exception as e:
        messages.error(request, f"停止生產失敗：{str(e)}")

    return redirect("workorder_dispatch:dispatch_list")

def selective_revert_orders(request):
    """
    選擇性將已轉換的公司製令單退回（is_converted 設回 False）
    只有管理員可操作。
    """
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "只有管理員可以執行此操作")
        return redirect("workorder:company_orders")

    from .models import CompanyOrder

    if request.method == "POST":
        ids = request.POST.getlist("order_ids")
        if ids:
            updated = CompanyOrder.objects.filter(id__in=ids, is_converted=True).update(
                is_converted=False
            )
            workorder_logger.info(
                f"管理員 {request.user} 選擇性退回 {updated} 筆製令單。IP: {request.META.get('REMOTE_ADDR')}"
            )
            messages.success(request, f"已成功退回 {updated} 筆製令單！")
        else:
            messages.warning(request, "請至少選擇一筆要退回的製令單！")
        return redirect("workorder:company_orders")

    # GET 顯示所有已轉換的製令單
    orders = CompanyOrder.objects.filter(is_converted=True).order_by("-sync_time")
    return render(request, "workorder/selective_revert_orders.html", {"orders": orders})

def user_supplement_form(request, workorder_id):
    """
    使用者補登表單 - 只能補登，不能核准
    """
    from django.shortcuts import get_object_or_404, redirect
    from datetime import datetime

    workorder = get_object_or_404(WorkOrder, id=workorder_id)
    processes = workorder.processes.all()
    # 取得所有作業員與設備
    operators = list(Operator.objects.all())
    equipments = list(Equipment.objects.all())
    error_msgs = []
    
    if request.method == "POST":
        for process in processes:
            completed_quantity = request.POST.get(f"completed_quantity_{process.id}")
            work_date = request.POST.get(f"work_date_{process.id}")
            start_time = request.POST.get(f"start_time_{process.id}")
            end_time = request.POST.get(f"end_time_{process.id}")
            operator_id = request.POST.get(f"operator_{process.id}")
            equipment_id = request.POST.get(f"equipment_{process.id}")
            notes = request.POST.get(f"notes_{process.id}")
            completed_flag = request.POST.get(f"completed_flag_{process.id}")

            # 判斷工序類型
            is_smt = (
                "SMT" in process.process_name
                or "smt" in process.process_name
                or "貼片" in process.process_name
            )
            # 驗證必填欄位
            if not is_smt and not operator_id:
                error_msgs.append(f"工序「{process.process_name}」作業員必填！")
            if is_smt and not equipment_id:
                error_msgs.append(f"工序「{process.process_name}」設備必填！")

            # 組合日期和時間為 datetime
            supplement_datetime = None
            end_datetime = None

            if work_date and start_time:
                try:
                    supplement_datetime = datetime.strptime(
                        f"{work_date} {start_time}", "%Y-%m-%d %H:%M"
                    )
                    supplement_datetime = timezone.make_aware(supplement_datetime)
                except ValueError:
                    error_msgs.append(
                        f"工序「{process.process_name}」開始時間格式錯誤！"
                    )

            if work_date and end_time:
                try:
                    end_datetime = datetime.strptime(
                        f"{work_date} {end_time}", "%Y-%m-%d %H:%M"
                    )
                    end_datetime = timezone.make_aware(end_datetime)
                except ValueError:
                    error_msgs.append(
                        f"工序「{process.process_name}」結束時間格式錯誤！"
                    )

            # 驗證時間邏輯
            if supplement_datetime and end_datetime:
                if supplement_datetime >= end_datetime:
                    error_msgs.append(
                        f"工序「{process.process_name}」結束時間必須晚於開始時間！"
                    )
                if (
                    supplement_datetime > timezone.now()
                    or end_datetime > timezone.now()
                ):
                    error_msgs.append(
                        f"工序「{process.process_name}」補登時間不能超過現在時間！"
                    )

            # 只要有填寫完成數量才儲存
            if completed_quantity:
                # 取得作業員/設備物件
                operator_obj = (
                    Operator.objects.filter(id=operator_id).first()
                    if operator_id
                    else None
                )
                equipment_name = None
                if equipment_id:
                    eq_obj = Equipment.objects.filter(id=equipment_id).first()
                    equipment_name = eq_obj.name if eq_obj else None
                # 儲存補登紀錄
                QuickSupplementLog.objects.create(
                    product_code=workorder.product_code,
                    workorder=workorder,
                    process=process,
                    action_type="complete",
                    supplement_time=supplement_datetime or timezone.now(),
                    end_time=end_datetime,
                    completed_quantity=completed_quantity or 0,
                    defect_qty=0,  # 預設不良品數量為 0
                    notes=notes,
                    created_by=request.user.username,
                    operator=operator_obj.name if operator_obj else None,
                    equipment=equipment_name,
                )
                # 若勾選已完工，直接設工序狀態為已完成
                if completed_flag:
                    process.status = "completed"
                    process.save()
        if error_msgs:
            messages.error(request, "；".join(error_msgs))
        else:
            messages.success(request, "補登資料已提交，等待管理員核准")
            return redirect("workorder:user_supplement_index")
    
    return render(
        request,
        "workorder/user_supplement_form.html",
        {
            "workorder": workorder,
            "processes": processes,
            "operators": operators,
            "equipments": equipments,
            "today": datetime.now().date(),
        },
    )

def edit_my_supplement(request, supplement_id):
    """
    編輯自己的補登記錄 - 只能編輯未核准的記錄
    """
    from django.shortcuts import get_object_or_404, redirect
    from datetime import datetime
    
    supplement = get_object_or_404(QuickSupplementLog, id=supplement_id)
    
    # 檢查權限 - 只能編輯自己的未核准記錄
    if supplement.created_by != request.user.username:
        messages.error(request, "您只能編輯自己的補登記錄")
        return redirect('workorder:user_supplement_index')
    
    if supplement.approval_status != "pending":
        messages.error(request, "已核准的記錄無法修改")
        return redirect('workorder:user_supplement_index')
    
    if request.method == "POST":
        # 更新補登記錄
        supplement.completed_quantity = request.POST.get('completed_quantity', 0)
        supplement.defect_qty = request.POST.get('defect_qty', 0)
        supplement.notes = request.POST.get('notes', '')
        
        # 更新時間
        work_date = request.POST.get('work_date')
        start_time = request.POST.get('start_time')
        if work_date and start_time:
            try:
                supplement_datetime = datetime.strptime(
                    f"{work_date} {start_time}", "%Y-%m-%d %H:%M"
                )
                supplement.supplement_time = timezone.make_aware(supplement_datetime)
            except ValueError:
                messages.error(request, "時間格式錯誤")
                return redirect('workorder:edit_my_supplement', supplement_id=supplement_id)
        
        supplement.save()
        messages.success(request, "補登記錄已更新")
        return redirect('workorder:user_supplement_index')
    
    return render(
        request,
        "workorder/edit_my_supplement.html",
        {
            "supplement": supplement,
            "today": datetime.now().date(),
        },
    )

def delete_my_supplement(request, supplement_id):
    """
    刪除自己的補登記錄 - 只能刪除未核准的記錄
    """
    from django.shortcuts import get_object_or_404, redirect
    
    supplement = get_object_or_404(QuickSupplementLog, id=supplement_id)
    
    # 檢查權限 - 只能刪除自己的未核准記錄
    if supplement.created_by != request.user.username:
        messages.error(request, "您只能刪除自己的補登記錄")
        return redirect('workorder:user_supplement_index')
    
    if supplement.approval_status != "pending":
        messages.error(request, "已核准的記錄無法刪除")
        return redirect('workorder:user_supplement_index')
    
    supplement.delete()
    messages.success(request, "補登記錄已刪除")
    return redirect('workorder:user_supplement_index')

def process_logs(request, process_id):
    """
    顯示工序日誌頁面
    """
    from django.shortcuts import get_object_or_404

    process = get_object_or_404(WorkOrderProcess, pk=process_id)
    logs = DispatchLog.objects.filter(workorder_process=process).order_by("-created_at")

    return render(
        request,
        "workorder/process/process_logs.html",
        {
            "process": process,
            "logs": logs,
        },
    )

def edit_supplement_mobile(request):
    """
    行動裝置補登編輯頁面
    """
    log_id = request.GET.get("log_id")
    if log_id:
        log = get_object_or_404(QuickSupplementLog, pk=log_id)
        return render(request, "workorder/edit_supplement_mobile.html", {"log": log})
    else:
        messages.error(request, "缺少日誌ID")
        return redirect("workorder:index")

def pending_approval_list(request):
    """待審核報工列表 - 已移至主管功能模組"""
    from django.shortcuts import redirect
    return redirect('workorder:supervisor:pending_approval_list')

def approved_reports_list(request):
    """已核准報工列表 - 已移至 reporting 模組"""
    from django.shortcuts import redirect
    return redirect('reporting:placeholder', function_name='approved_reports_list')

@login_required
def get_operators_and_equipments(request):
    """
    API：取得所有作業員和設備資料，供前端 AJAX 使用
    """
    try:
        from process.models import Operator
        from equip.models import Equipment
        # 取得作業員資料，包含技能
        operators = []
        for op in Operator.objects.all():
            skills = list(op.skills.all().values_list('process_name__name', flat=True))
            operators.append({
                "id": op.id,
                "name": op.name,
                "skills": skills,
            })
        # 取得設備資料，包含型別和狀態
        equipments = []
        for eq in Equipment.objects.all():
            equipments.append({
                "id": eq.id,
                "name": eq.name,
                "model": getattr(eq, 'model', ''),
                "type": getattr(eq, 'type', ''),
                "status": getattr(eq, 'status', ''),
            })
        return JsonResponse({
            "success": True,
            "operators": operators,
            "equipments": equipments,
        })
    except Exception as e:
        return JsonResponse({"success": False, "message": f"取得資料失敗：{str(e)}"})

@require_GET
@login_required
def get_operators_only(request):
    """
    API：取得所有作業員資料，包含技能清單，供前端 AJAX 使用
    作業員沒有指定技能就不顯示
    """
    try:
        from process.models import Operator, OperatorSkill, ProcessName
        
        # 取得工序名稱參數（可選）
        process_name = request.GET.get('process_name', '').strip()
        
        operators = []
        
        if process_name:
            # 查詢有該工序技能的作業員
            skilled_operators = Operator.objects.filter(
                skills__process_name__name=process_name
            ).distinct()
            
            for op in skilled_operators:
                # 取得技能清單
                skills = list(op.skills.all().values_list('process_name__name', flat=True))
                operators.append({
                    "id": op.id,
                    "name": op.name,
                    "skills": skills,
                })
            # 如果沒有作業員有該技能，operators 保持空列表，不顯示任何作業員
        else:
            # 如果沒有指定工序名稱，顯示所有作業員（用於一般查詢）
            for op in Operator.objects.all():
                skills = list(op.skills.all().values_list('process_name__name', flat=True))
                operators.append({
                    "id": op.id,
                    "name": op.name,
                    "skills": skills,
                })
        
        return JsonResponse({
            "success": True,
            "operators": operators,
        })
    except Exception as e:
        return JsonResponse({"success": False, "message": f"取得作業員資料失敗：{str(e)}"})

@require_GET
@login_required
def get_equipments_only(request):
    """
    API：取得所有設備資料，供前端 AJAX 使用
    工序沒有指定設備就不顯示
    """
    try:
        from equip.models import Equipment
        from process.models import ProcessEquipment, ProcessName
        
        # 取得工序名稱參數（可選）
        process_name = request.GET.get('process_name', '').strip()
        
        equipments = []
        
        if process_name:
            # 查詢工序是否有指定設備
            suitable_equipment_ids = ProcessEquipment.objects.filter(
                process_name__name=process_name
            ).values_list('equipment_id', flat=True)
            
            # 如果工序有指定設備，才顯示這些設備
            if suitable_equipment_ids:
                suitable_equipments = Equipment.objects.filter(id__in=suitable_equipment_ids)
                for eq in suitable_equipments:
                    equipments.append({
                        "id": eq.id,
                        "name": eq.name,
                        "model": getattr(eq, 'model', ''),
                        "status": getattr(eq, 'status', ''),
                    })
            # 如果工序沒有指定設備，equipments 保持空列表，不顯示任何設備
        
        return JsonResponse({
            "success": True,
            "equipments": equipments,
        })
    except Exception as e:
        return JsonResponse({"success": False, "message": f"取得設備資料失敗：{str(e)}"})

# ==================== 行動裝置補登管理作業 ====================

def mobile_quick_supplement_index(request):
    """
    行動裝置補登管理作業主頁
    針對行動裝置設計的簡潔介面，讓使用者依產品編號查詢工單進行補登
    """
    workorders = None
    product_code = request.GET.get("product_code")
    if product_code:
        workorders = WorkOrder.objects.filter(product_code__icontains=product_code)
        # 轉成列表並加上產品名稱欄位
        workorders = [
            {
                "id": wo.id,
                "order_number": wo.order_number,
                "product_name": wo.product_code,
                "quantity": wo.quantity,
                "status": wo.status,
                "get_status_display": wo.get_status_display(),
            }
            for wo in workorders
        ]
    return render(
        request,
        "workorder/mobile_quick_supplement_index.html",
        {"workorders": workorders, "product_code": product_code},
    )

def mobile_quick_supplement_form(request, workorder_id):
    """
    行動裝置補登管理作業表單
    針對行動裝置設計的表單，簡化操作流程，適合觸控操作
    """
    from django.shortcuts import get_object_or_404, redirect
    from datetime import datetime

    workorder = get_object_or_404(WorkOrder, id=workorder_id)
    processes = workorder.processes.all()
    # 取得所有作業員與設備
    operators = list(Operator.objects.all())
    equipments = list(Equipment.objects.all())
    error_msgs = []

    if request.method == "POST":
        for process in processes:
            completed_quantity = request.POST.get(f"completed_quantity_{process.id}")
            work_date = request.POST.get(f"work_date_{process.id}")
            start_time = request.POST.get(f"start_time_{process.id}")
            end_time = request.POST.get(f"end_time_{process.id}")
            operator_id = request.POST.get(f"operator_{process.id}")
            equipment_id = request.POST.get(f"equipment_{process.id}")
            notes = request.POST.get(f"notes_{process.id}")
            completed_flag = request.POST.get(f"completed_flag_{process.id}")

            # 判斷工序類型
            is_smt = (
                "SMT" in process.process_name
                or "smt" in process.process_name
                or "貼片" in process.process_name
            )

            # 驗證必填欄位
            if not is_smt and not operator_id:
                error_msgs.append(f"工序「{process.process_name}」作業員必填！")
            if is_smt and not equipment_id:
                error_msgs.append(f"工序「{process.process_name}」設備必填！")

            # 組合日期和時間為 datetime
            supplement_datetime = None
            end_datetime = None

            if work_date and start_time:
                try:
                    supplement_datetime = datetime.strptime(
                        f"{work_date} {start_time}", "%Y-%m-%d %H:%M"
                    )
                    supplement_datetime = timezone.make_aware(supplement_datetime)
                except ValueError:
                    error_msgs.append(
                        f"工序「{process.process_name}」開始時間格式錯誤！"
                    )

            if work_date and end_time:
                try:
                    end_datetime = datetime.strptime(
                        f"{work_date} {end_time}", "%Y-%m-%d %H:%M"
                    )
                    end_datetime = timezone.make_aware(end_datetime)
                except ValueError:
                    error_msgs.append(
                        f"工序「{process.process_name}」結束時間格式錯誤！"
                    )

            # 驗證時間邏輯
            if supplement_datetime and end_datetime:
                if supplement_datetime >= end_datetime:
                    error_msgs.append(
                        f"工序「{process.process_name}」結束時間必須晚於開始時間！"
                    )
                if (
                    supplement_datetime > timezone.now()
                    or end_datetime > timezone.now()
                ):
                    error_msgs.append(
                        f"工序「{process.process_name}」補登時間不能超過現在時間！"
                    )

            # 只要有填寫完成數量才儲存
            if completed_quantity:
                # 取得作業員/設備物件
                operator_obj = (
                    Operator.objects.filter(id=operator_id).first()
                    if operator_id
                    else None
                )
                equipment_name = None
                if equipment_id:
                    eq_obj = Equipment.objects.filter(id=equipment_id).first()
                    equipment_name = eq_obj.name if eq_obj else None

                # 儲存補登紀錄
                QuickSupplementLog.objects.create(
                    product_code=workorder.product_code,
                    workorder=workorder,
                    process=process,
                    action_type="complete",
                    supplement_time=supplement_datetime or timezone.now(),
                    end_time=end_datetime,
                    completed_quantity=completed_quantity or 0,
                    notes=notes,
                    created_by=request.user.username,
                    operator=operator_obj.name if operator_obj else None,
                    equipment=equipment_name,
                )

                # 若勾選已完工，直接設工序狀態為已完成
                if completed_flag:
                    process.status = "completed"
                    process.save()

        if error_msgs:
            messages.error(request, "；".join(error_msgs))
        else:
            messages.success(request, "補登資料已成功儲存！")
            return redirect("workorder:mobile_quick_supplement_index")

    return render(
        request,
        "workorder/mobile_quick_supplement_form.html",
        {
            "workorder": workorder,
            "processes": processes,
            "operators": operators,
            "equipments": equipments,
        },
    )

def clear_all_production_reports(request):
    """
    清除所有報工紀錄的視圖函數
    包括：作業員補登報工、SMT補登報工、SMT現場報工
    只有管理員可以執行此操作
    """
    if not (request.user.is_staff or request.user.is_superuser):
        messages.error(request, "只有管理員可以執行此操作")
        return redirect("workorder:index")

    if request.method == "POST":
        try:
            # 統計要清除的資料數量
            from workorder.models import CompletedProductionReport
            operator_reports_count = CompletedProductionReport.objects.filter(report_type='operator').count()
            smt_supplement_count = CompletedProductionReport.objects.filter(report_type='smt').count()
            smt_on_site_count = 0  # 已移除 report_type 欄位
            
            total_count = operator_reports_count + smt_supplement_count + smt_on_site_count
            
            if total_count == 0:
                messages.info(request, "目前沒有任何報工紀錄需要清除")
                return redirect("workorder:index")
            
            # 清除所有報工紀錄
            CompletedProductionReport.objects.all().delete()
            
            # 記錄操作日誌
            from system.models import OperationLog
            OperationLog.objects.create(
                user=request.user.username,
                module="workorder",
                action=f"清除所有報工紀錄（作業員：{operator_reports_count}，SMT補登：{smt_supplement_count}，SMT現場：{smt_on_site_count}）",
                timestamp=timezone.now(),
                ip_address=request.META.get('REMOTE_ADDR', ''),
            )
            
            messages.success(
                request, 
                f"成功清除所有報工紀錄！共清除 {total_count} 筆記錄：\n"
                f"• 作業員補登報工：{operator_reports_count} 筆\n"
                f"• SMT補登報工：{smt_supplement_count} 筆\n"
                f"• SMT現場報工：{smt_on_site_count} 筆"
            )
            return redirect("workorder:index")
            
        except Exception as e:
            messages.error(request, f"清除報工紀錄失敗：{str(e)}")
            return redirect("workorder:clear_all_production_reports")

    # GET 請求顯示確認頁面
    from workorder.models import CompletedProductionReport
    operator_reports_count = CompletedProductionReport.objects.filter(report_type='operator').count()
    smt_supplement_count = CompletedProductionReport.objects.filter(report_type='smt').count()
    smt_on_site_count = 0  # 已移除 report_type 欄位
    
    total_count = operator_reports_count + smt_supplement_count + smt_on_site_count
    
    context = {
        "operator_reports_count": operator_reports_count,
        "smt_supplement_count": smt_supplement_count,
        "smt_on_site_count": smt_on_site_count,
        "total_count": total_count,
    }
    
    return render(request, "workorder/clear_production_reports_confirm.html", context)

@require_GET
@login_required
def mobile_get_workorder_info(request):
    """
    行動裝置 API：根據產品編號取得工單資訊
    """
    product_code = request.GET.get("product_code")
    if not product_code:
        return JsonResponse({"success": False, "message": "請提供產品編號"})

    try:
        workorders = WorkOrder.objects.filter(product_code__icontains=product_code)
        workorder_list = []

        for wo in workorders:
            workorder_list.append(
                {
                    "id": wo.id,
                    "order_number": wo.order_number,
                    "company_code": wo.company_code,
                    "product_code": wo.product_code,
                    "quantity": wo.quantity,
                    "status": wo.status,
                    "status_display": wo.get_status_display(),
                    "created_at": (
                        wo.created_at.strftime("%Y-%m-%d %H:%M")
                        if wo.created_at
                        else ""
                    ),
                }
            )

        return JsonResponse(
            {
                "success": True,
                "workorders": workorder_list,
                "count": len(workorder_list),
            }
        )

    except Exception as e:
        return JsonResponse({"success": False, "message": f"查詢失敗：{str(e)}"})

@require_GET
@login_required
def mobile_get_process_info(request):
    """
    行動裝置 API：取得指定工單的工序資訊
    """
    workorder_id = request.GET.get("workorder_id")
    if not workorder_id:
        return JsonResponse({"success": False, "message": "請提供工單ID"})

    try:
        workorder = WorkOrder.objects.get(id=workorder_id)
        processes = workorder.processes.all()
        process_list = []

        for process in processes:
            process_list.append(
                {
                    "id": process.id,
                    "process_name": process.process_name,
                    "status": process.status,
                    "status_display": process.get_status_display(),
                    "is_smt": "SMT" in process.process_name
                    or "smt" in process.process_name
                    or "貼片" in process.process_name,
                }
            )

        return JsonResponse(
            {
                "success": True,
                "workorder": {
                    "id": workorder.id,
                    "order_number": workorder.order_number,
                    "company_code": workorder.company_code,
                    "product_code": workorder.product_code,
                    "quantity": workorder.quantity,
                    "status": workorder.status,
                    "status_display": workorder.get_status_display(),
                },
                "processes": process_list,
            }
        )

    except WorkOrder.DoesNotExist:
        return JsonResponse({"success": False, "message": "找不到指定的工單"})
    except Exception as e:
        return JsonResponse({"success": False, "message": f"查詢失敗：{str(e)}"})

@require_POST
@login_required
def mobile_submit_supplement(request):
    """
    行動裝置 API：提交補登資料
    """
    try:
        from datetime import datetime

        workorder_id = request.POST.get("workorder_id")
        process_id = request.POST.get("process_id")
        completed_quantity = request.POST.get("completed_quantity", 0)
        work_date = request.POST.get("work_date")
        start_time = request.POST.get("start_time")
        end_time = request.POST.get("end_time")
        operator_id = request.POST.get("operator_id")
        equipment_id = request.POST.get("equipment_id")
        notes = request.POST.get("notes", "")
        completed_flag = request.POST.get("completed_flag") == "true"

        # 驗證必填欄位
        if not workorder_id or not process_id:
            return JsonResponse({"success": False, "message": "缺少必要參數"})

        workorder = WorkOrder.objects.get(id=workorder_id)
        process = WorkOrderProcess.objects.get(id=process_id)

        # 判斷工序類型並驗證
        is_smt = (
            "SMT" in process.process_name
            or "smt" in process.process_name
            or "貼片" in process.process_name
        )

        if not is_smt and not operator_id:
            return JsonResponse(
                {
                    "success": False,
                    "message": f"工序「{process.process_name}」作業員必填！",
                }
            )

        if is_smt and not equipment_id:
            return JsonResponse(
                {
                    "success": False,
                    "message": f"工序「{process.process_name}」設備必填！",
                }
            )

        # 組合日期和時間為 datetime
        supplement_datetime = None
        end_datetime = None

        if work_date and start_time:
            try:
                supplement_datetime = datetime.strptime(
                    f"{work_date} {start_time}", "%Y-%m-%d %H:%M"
                )
                supplement_datetime = timezone.make_aware(supplement_datetime)
            except ValueError:
                return JsonResponse({"success": False, "message": "開始時間格式錯誤！"})

        if work_date and end_time:
            try:
                end_datetime = datetime.strptime(
                    f"{work_date} {end_time}", "%Y-%m-%d %H:%M"
                )
                end_datetime = timezone.make_aware(end_datetime)
            except ValueError:
                return JsonResponse({"success": False, "message": "結束時間格式錯誤！"})

        # 驗證時間邏輯
        if supplement_datetime and end_datetime:
            if supplement_datetime >= end_datetime:
                return JsonResponse(
                    {"success": False, "message": "結束時間必須晚於開始時間！"}
                )
            if supplement_datetime > timezone.now() or end_datetime > timezone.now():
                return JsonResponse(
                    {"success": False, "message": "補登時間不能超過現在時間！"}
                )

        # 取得作業員/設備物件
        operator_obj = (
            Operator.objects.filter(id=operator_id).first() if operator_id else None
        )
        equipment_name = None
        if equipment_id:
            eq_obj = Equipment.objects.filter(id=equipment_id).first()
            equipment_name = eq_obj.name if eq_obj else None

        # 儲存補登紀錄
        QuickSupplementLog.objects.create(
            product_code=workorder.product_code,
            workorder=workorder,
            process=process,
            action_type="complete",
            supplement_time=supplement_datetime or timezone.now(),
            end_time=end_datetime,
            completed_quantity=completed_quantity or 0,
            notes=notes,
            created_by=request.user.username,
            operator=operator_obj.name if operator_obj else None,
            equipment=equipment_name,
        )

        # 若勾選已完工，直接設工序狀態為已完成
        if completed_flag:
            process.status = "completed"
            process.save()

        return JsonResponse({"success": True, "message": "補登資料已成功儲存！"})

    except WorkOrder.DoesNotExist:
        return JsonResponse({"success": False, "message": "找不到指定的工單"})
    except WorkOrderProcess.DoesNotExist:
        return JsonResponse({"success": False, "message": "找不到指定的工序"})
    except Exception as e:
        return JsonResponse({"success": False, "message": f"儲存失敗：{str(e)}"})

def mobile_api_test(request):
    """
    行動裝置 API 測試頁面
    用於測試行動裝置快速補登的 API 功能，方便開發和除錯
    """
    return render(request, "workorder/mobile_api_test.html")

@require_GET
@login_required
def quick_get_product_codes(request):
    """
    一般快速補登 API：取得所有有工單的產品編號清單
    """
    try:
        codes = WorkOrder.objects.values_list("product_code", flat=True).distinct()
        codes = sorted(set(codes))
        return JsonResponse({"success": True, "product_codes": list(codes)})
    except Exception as e:
        return JsonResponse({"success": False, "message": f"查詢失敗：{str(e)}"})

@require_GET
@login_required
def get_product_codes(request):
    """
    行動裝置 API：取得所有有工單的產品編號清單（去重、排序）
    """
    try:
        product_codes = (
            WorkOrder.objects.values_list("product_code", flat=True)
            .distinct()
            .order_by("product_code")
        )
        return JsonResponse(
            {
                "success": True,
                "product_codes": list(product_codes),
                "count": len(product_codes),
            }
        )
    except Exception as e:
        return JsonResponse({"success": False, "message": f"查詢失敗：{str(e)}"})

@require_POST
@login_required
def move_process(request):
    """
    移動工序順序（上下移動）
    修正：移動時使用臨時值避免唯一性約束衝突。
    """
    try:
        process_id = request.POST.get("process_id")
        direction = request.POST.get("direction")  # 'up' 或 'down'
        
        if not process_id or not direction:
            return JsonResponse({"success": False, "message": "缺少必要參數"})
        
        # 取得當前工序
        current_process = WorkOrderProcess.objects.get(id=process_id)
        current_order = current_process.step_order
        
        # 取得同工單的所有工序，按順序排序
        all_processes = WorkOrderProcess.objects.filter(
            workorder=current_process.workorder
        ).order_by("step_order")
        
        if direction == "up":
            # 向上移動：與前一個工序交換順序
            if current_order > 1:
                prev_process = all_processes.filter(step_order=current_order - 1).first()
                if prev_process:
                    # 使用臨時值避免唯一性約束衝突
                    temp_order = 999999
                    # 先將當前工序移到臨時位置
                    current_process.step_order = temp_order
                    current_process.save()
                    # 將前一個工序移到當前位置
                    prev_process.step_order = current_order
                    prev_process.save()
                    # 將當前工序移到前一個位置
                    current_process.step_order = current_order - 1
                    current_process.save()
                    workorder_logger.info(
                        f"使用者 {request.user} 將工序「{current_process.process_name}」向上移動。IP: {request.META.get('REMOTE_ADDR')}"
                    )
                    return JsonResponse({"success": True, "message": "工序順序已更新"})
        elif direction == "down":
            # 向下移動：與後一個工序交換順序
            max_order = all_processes.count()
            if current_order < max_order:
                next_process = all_processes.filter(step_order=current_order + 1).first()
                if next_process:
                    # 使用臨時值避免唯一性約束衝突
                    temp_order = 999999
                    # 先將當前工序移到臨時位置
                    current_process.step_order = temp_order
                    current_process.save()
                    # 將後一個工序移到當前位置
                    next_process.step_order = current_order
                    next_process.save()
                    # 將當前工序移到後一個位置
                    current_process.step_order = current_order + 1
                    current_process.save()
                    workorder_logger.info(
                        f"使用者 {request.user} 將工序「{current_process.process_name}」向下移動。IP: {request.META.get('REMOTE_ADDR')}"
                    )
                    return JsonResponse({"success": True, "message": "工序順序已更新"})
        
        return JsonResponse({"success": False, "message": "無法移動工序順序"})
        
    except WorkOrderProcess.DoesNotExist:
        return JsonResponse({"success": False, "message": "找不到指定的工序"})
    except Exception as e:
        workorder_logger.error(f"移動工序失敗：{str(e)}")
        return JsonResponse({"success": False, "message": f"移動失敗：{str(e)}"})

@require_POST
@login_required
def add_process(request, workorder_id):
    """
    新增工序到工單
    """
    try:
        # 取得工單
        workorder = WorkOrder.objects.get(id=workorder_id)
        
        # 取得表單資料
        process_name = request.POST.get("process_name")
        step_order = request.POST.get("step_order")
        planned_quantity = request.POST.get("planned_quantity")
        target_hourly_output = request.POST.get("target_hourly_output")
        
        # 驗證必填欄位
        if not process_name or not step_order:
            return JsonResponse({"status": "error", "message": "工序名稱和順序為必填欄位"})
        
        try:
            step_order = int(step_order)
            planned_quantity = int(planned_quantity) if planned_quantity else workorder.quantity
            target_hourly_output = int(target_hourly_output) if target_hourly_output else 100
        except ValueError:
            return JsonResponse({"status": "error", "message": "數值欄位格式錯誤"})
        
        # 檢查工序順序是否重複
        if WorkOrderProcess.objects.filter(workorder=workorder, step_order=step_order).exists():
            return JsonResponse({"status": "error", "message": f"工序順序 {step_order} 已存在"})
        
        # 調整後續工序的順序（使用臨時值避免唯一性約束衝突）
        existing_processes = WorkOrderProcess.objects.filter(
            workorder=workorder, 
            step_order__gte=step_order
        ).order_by("-step_order")
        
        # 先將所有需要調整的工序移到臨時位置
        temp_start = 999999
        for i, process in enumerate(existing_processes):
            process.step_order = temp_start + i
            process.save()
        
        # 再將工序移到正確的位置
        for i, process in enumerate(existing_processes):
            process.step_order = step_order + 1 + i
            process.save()
        
        # 建立新工序
        new_process = WorkOrderProcess.objects.create(
            workorder=workorder,
            process_name=process_name,
            step_order=step_order,
            planned_quantity=planned_quantity,
            target_hourly_output=target_hourly_output,
            status="pending"
        )
        
        # 計算預計工時
        if target_hourly_output > 0:
            estimated_hours = planned_quantity / target_hourly_output
            new_process.estimated_hours = round(estimated_hours, 2)
            new_process.save()
        
        workorder_logger.info(
            f"使用者 {request.user} 為工單 {workorder.order_number} 新增工序「{process_name}」。IP: {request.META.get('REMOTE_ADDR')}"
        )
        
        return JsonResponse({
            "status": "success", 
            "message": f"工序「{process_name}」新增成功"
        })
        
    except WorkOrder.DoesNotExist:
        return JsonResponse({"status": "error", "message": "找不到指定的工單"})
    except Exception as e:
        workorder_logger.error(f"新增工序失敗：{str(e)}")
        return JsonResponse({"status": "error", "message": f"新增失敗：{str(e)}"})

@require_POST
@login_required
def edit_process(request, process_id):
    """
    編輯工序
    """
    try:
        # 取得工序
        process = WorkOrderProcess.objects.get(id=process_id)
        
        # 取得表單資料
        process_name = request.POST.get("process_name")
        step_order = request.POST.get("step_order")
        planned_quantity = request.POST.get("planned_quantity")
        target_hourly_output = request.POST.get("target_hourly_output")
        
        # 驗證必填欄位
        if not process_name or not step_order:
            return JsonResponse({"status": "error", "message": "工序名稱和順序為必填欄位"})
        
        try:
            step_order = int(step_order)
            planned_quantity = int(planned_quantity) if planned_quantity else process.planned_quantity
            target_hourly_output = int(target_hourly_output) if target_hourly_output else process.target_hourly_output
        except ValueError:
            return JsonResponse({"status": "error", "message": "數值欄位格式錯誤"})
        
        # 檢查工序順序是否重複（排除自己）
        if WorkOrderProcess.objects.filter(
            workorder=process.workorder, 
            step_order=step_order
        ).exclude(id=process_id).exists():
            return JsonResponse({"status": "error", "message": f"工序順序 {step_order} 已存在"})
        
        # 如果工序順序有變更，需要重新排序
        old_step_order = process.step_order
        if step_order != old_step_order:
            # 先將當前工序移到臨時位置
            temp_order = 999999
            process.step_order = temp_order
            process.save()
            
            # 調整其他工序的順序
            if step_order > old_step_order:
                # 向下移動：將中間的工序向上移動
                processes_to_move = WorkOrderProcess.objects.filter(
                    workorder=process.workorder,
                    step_order__gt=old_step_order,
                    step_order__lte=step_order
                ).exclude(id=process_id).order_by('step_order')
                
                for proc in processes_to_move:
                    proc.step_order -= 1
                    proc.save()
            else:
                # 向上移動：將中間的工序向下移動
                processes_to_move = WorkOrderProcess.objects.filter(
                    workorder=process.workorder,
                    step_order__gte=step_order,
                    step_order__lt=old_step_order
                ).exclude(id=process_id).order_by('-step_order')
                
                for proc in processes_to_move:
                    proc.step_order += 1
                    proc.save()
        
        # 更新工序資料
        process.process_name = process_name
        process.step_order = step_order
        process.planned_quantity = planned_quantity
        process.target_hourly_output = target_hourly_output
        
        # 計算預計工時
        if target_hourly_output > 0:
            estimated_hours = planned_quantity / target_hourly_output
            process.estimated_hours = round(estimated_hours, 2)
        
        process.save()
        
        workorder_logger.info(
            f"使用者 {request.user} 編輯工序「{process_name}」。IP: {request.META.get('REMOTE_ADDR')}"
        )
        
        return JsonResponse({
            "status": "success", 
            "message": f"工序「{process_name}」更新成功"
        })
        
    except WorkOrderProcess.DoesNotExist:
        return JsonResponse({"status": "error", "message": "找不到指定的工序"})
    except Exception as e:
        workorder_logger.error(f"編輯工序失敗：{str(e)}")
        return JsonResponse({"status": "error", "message": f"編輯失敗：{str(e)}"})

@require_POST
@login_required
def delete_process(request, process_id):
    """
    刪除工序
    """
    try:
        # 取得工序
        process = WorkOrderProcess.objects.get(id=process_id)
        
        # 檢查工序狀態
        if process.status != "pending":
            return JsonResponse({"status": "error", "message": "只有待生產狀態的工序才能刪除"})
        
        process_name = process.process_name
        workorder = process.workorder
        
        # 刪除工序
        process.delete()
        
        # 重新排序剩餘工序
        remaining_processes = WorkOrderProcess.objects.filter(
            workorder=workorder
        ).order_by("step_order")
        
        for i, proc in enumerate(remaining_processes, 1):
            proc.step_order = i
            proc.save()
        
        workorder_logger.info(
            f"使用者 {request.user} 刪除工序「{process_name}」。IP: {request.META.get('REMOTE_ADDR')}"
        )
        
        return JsonResponse({
            "status": "success", 
            "message": f"工序「{process_name}」刪除成功"
        })
        
    except WorkOrderProcess.DoesNotExist:
        return JsonResponse({"status": "error", "message": "找不到指定的工序"})
    except Exception as e:
        workorder_logger.error(f"刪除工序失敗：{str(e)}")
        return JsonResponse({"status": "error", "message": f"刪除失敗：{str(e)}"})

@require_POST
@login_required
def batch_approve_supplements(request):
    """
    批量核准補登記錄 API
    """
    try:
        workorder_ids = request.POST.getlist('workorder_ids[]')
        if not workorder_ids:
            return JsonResponse({
                'success': False,
                'message': '請選擇要核准的工單'
            })
        
        approved_count = 0
        for workorder_id in workorder_ids:
            # 核准該工單的所有待核准補登記錄
            supplements = QuickSupplementLog.objects.filter(
                workorder_id=workorder_id,
                approval_status='pending'
            )
            supplements.update(
                approval_status='approved',
                approved_at=timezone.now(),
                approved_by=request.user.username
            )
            approved_count += supplements.count()
        
        return JsonResponse({
            'success': True,
            'message': f'成功核准 {approved_count} 筆補登記錄',
            'approved_count': approved_count
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'核准失敗：{str(e)}'
        })

@require_POST
@login_required
def batch_approve_pending(request):
    """
    批次審核所有待審核報工記錄 API
    只有超級管理員可以使用此功能
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    # 檢查是否為超級管理員
    if not request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'message': '只有超級管理員可以使用批次審核功能'
        })
    
    return JsonResponse({
        'success': False,
        'message': '此功能已棄用，請使用新的報工系統進行審核'
    })

@require_POST
@login_required
def quick_approve_workorder(request, workorder_id):
    """
    快速核准指定工單的所有待核准補登記錄
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    return JsonResponse({
        'success': False,
        'message': '此功能已棄用，請使用新的報工系統進行審核'
    })

@require_GET
@login_required
def supplement_statistics(request):
    """
    補登統計資料 API
    """
    from datetime import datetime, timedelta
    from django.db.models import Q, Count, Sum
    
    try:
        # 取得查詢參數
        period = request.GET.get('period', 'today')  # today, week, month
        product_code = request.GET.get('product_code', '')
        
        # 設定時間範圍
        today = datetime.now().date()
        if period == 'today':
            start_date = today
            end_date = today
        elif period == 'week':
            start_date = today - timedelta(days=today.weekday())
            end_date = today
        elif period == 'month':
            start_date = today.replace(day=1)
            end_date = today
        else:
            start_date = today
            end_date = today
        
        # 建立查詢條件
        query = Q(supplement_time__date__gte=start_date, supplement_time__date__lte=end_date)
        if product_code:
            query &= Q(product_code__icontains=product_code)
        
        # 統計資料
        supplements = QuickSupplementLog.objects.filter(query)
        
        stats = supplements.aggregate(
            total_records=Count('id'),
            pending_records=Count('id', filter=Q(approval_status='pending')),
            approved_records=Count('id', filter=Q(approval_status='approved')),
            total_completed=Sum('completed_quantity'),
            total_defect=Sum('defect_qty'),
            total_workorders=Count('workorder', distinct=True)
        )
        
        # 按日期分組統計
        daily_stats = supplements.values('supplement_time__date').annotate(
            daily_records=Count('id'),
            daily_completed=Sum('completed_quantity'),
            daily_defect=Sum('defect_qty')
        ).order_by('supplement_time__date')
        
        return JsonResponse({
            'success': True,
            'data': {
                'period': period,
                'start_date': start_date.isoformat(),
                'end_date': end_date.isoformat(),
                'summary': stats,
                'daily_stats': list(daily_stats)
            }
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'統計失敗：{str(e)}'
        })

@require_GET
@login_required
def get_process_capacity_info(request, process_id):
    """
    取得工序產能資訊 API
    回傳工序的標準產能、當前產能倍數、額外作業員和設備清單
    """
    from .models import WorkOrderProcess
    try:
        process = WorkOrderProcess.objects.get(id=process_id)
        # 取得標準產能（從產品工序標準產能表）
        from process.models import ProductProcessStandardCapacity
        capacity_data = ProductProcessStandardCapacity.objects.filter(
            product_code=process.workorder.product_code,
            process_name=process.process_name,
            is_active=True
        ).order_by('-version').first()
        standard_capacity = capacity_data.standard_capacity_per_hour if capacity_data else 1000
        # 取得額外作業員和設備清單
        additional_operators = process.get_additional_operators_list()
        additional_equipments = process.get_additional_equipments_list()
        # 取得所有可用作業員和設備供選擇
        from process.models import Operator
        from equip.models import Equipment
        all_operators = list(Operator.objects.values('id', 'name'))
        all_equipments = list(Equipment.objects.values('id', 'name'))
        return JsonResponse({
            'success': True,
            'capacity_info': {
                'standard_capacity': standard_capacity,
                'current_multiplier': process.capacity_multiplier,
                'current_target_hourly_output': process.target_hourly_output,
                'additional_operators': additional_operators,
                'additional_equipments': additional_equipments,
                'all_operators': all_operators,
                'all_equipments': all_equipments,
            }
        })
    except WorkOrderProcess.DoesNotExist:
        return JsonResponse({'success': False, 'message': '工序不存在'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'取得產能資訊失敗：{str(e)}'}, status=500)

@require_POST
@login_required
def update_process_capacity(request, process_id):
    """
    更新工序產能設定 API
    更新產能倍數、目標每小時產出、額外作業員和設備
    """
    from .models import WorkOrderProcess
    try:
        process = WorkOrderProcess.objects.get(id=process_id)
        # 取得表單資料
        capacity_multiplier = int(request.POST.get('capacity_multiplier', 1))
        target_hourly_output = int(request.POST.get('target_hourly_output', 0))
        additional_operators = request.POST.get('additional_operators', '')
        additional_equipments = request.POST.get('additional_equipments', '')
        # 驗證產能倍數
        if capacity_multiplier < 1 or capacity_multiplier > 10:
            return JsonResponse({'success': False, 'message': '產能倍數必須在 1-10 之間'}, status=400)
        # 更新工序資料
        process.capacity_multiplier = capacity_multiplier
        process.target_hourly_output = target_hourly_output
        process.additional_operators = additional_operators
        process.additional_equipments = additional_equipments
        # 重新計算預計工時
        process.estimated_hours = process.calculate_estimated_hours()
        process.save()
        # 記錄操作日誌（如有 WorkOrderProcessLog 模型）
        try:
            from .models import WorkOrderProcessLog
            WorkOrderProcessLog.objects.create(
                workorder_process=process,
                action='capacity_update',
                description=f'更新產能設定：倍數={capacity_multiplier}x，目標產能={target_hourly_output}pcs/hr',
                created_by=request.user.username if request.user.is_authenticated else 'system'
            )
        except Exception:
            pass
        return JsonResponse({
            'success': True,
            'message': '產能設定更新成功',
            'updated_data': {
                'capacity_multiplier': capacity_multiplier,
                'target_hourly_output': target_hourly_output,
                'estimated_hours': float(process.estimated_hours),
                'completion_rate': process.completion_rate,
            }
        })
    except WorkOrderProcess.DoesNotExist:
        return JsonResponse({'success': False, 'message': '工序不存在'}, status=404)
    except ValueError as e:
        return JsonResponse({'success': False, 'message': f'資料格式錯誤：{str(e)}'}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'更新產能設定失敗：{str(e)}'}, status=500)

@require_GET
@login_required
def get_capacity_calculation_info(request, process_id):
    """
    取得產能計算資訊 API
    用於產能計算器的詳細資訊
    """
    from .models import WorkOrderProcess
    try:
        process = WorkOrderProcess.objects.get(id=process_id)
        # 取得標準產能資料
        from process.models import ProductProcessStandardCapacity
        capacity_data = ProductProcessStandardCapacity.objects.filter(
            product_code=process.workorder.product_code,
            process_name=process.process_name,
            is_active=True
        ).order_by('-version').first()
        if capacity_data:
            calculation_info = {
                'standard_capacity': capacity_data.standard_capacity_per_hour,
                'efficiency_factor': float(capacity_data.efficiency_factor),
                'learning_curve_factor': float(capacity_data.learning_curve_factor),
                'setup_time_minutes': capacity_data.setup_time_minutes,
                'teardown_time_minutes': capacity_data.teardown_time_minutes,
                'cycle_time_seconds': float(capacity_data.cycle_time_seconds),
                'optimal_batch_size': capacity_data.optimal_batch_size,
                'expected_defect_rate': float(capacity_data.expected_defect_rate),
                'rework_time_factor': float(capacity_data.rework_time_factor),
            }
        else:
            # 使用預設值
            calculation_info = {
                'standard_capacity': 1000,
                'efficiency_factor': 1.0,
                'learning_curve_factor': 1.0,
                'setup_time_minutes': 30,
                'teardown_time_minutes': 15,
                'cycle_time_seconds': 3.6,
                'optimal_batch_size': 100,
                'expected_defect_rate': 0.02,
                'rework_time_factor': 1.05,
            }
        # 加入當前工序資訊
        calculation_info.update({
            'current_multiplier': process.capacity_multiplier,
            'planned_quantity': process.planned_quantity,
            'completed_quantity': process.completed_quantity,
            'remaining_quantity': process.remaining_quantity,
        })
        return JsonResponse({'success': True, 'calculation_info': calculation_info})
    except WorkOrderProcess.DoesNotExist:
        return JsonResponse({'success': False, 'message': '工序不存在'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'取得計算資訊失敗：{str(e)}'}, status=500)

@require_POST
@login_required
def update_process_status(request, process_id):
    """
    API：更新工序分配（作業員、設備、備註），並變更狀態為生產中
    純手動分配，不進行自動規範檢查
    """
    from workorder.models import WorkOrderProcess
    from equip.models import Equipment
    
    try:
        process = WorkOrderProcess.objects.get(id=process_id)
        operator_name = request.POST.get('operator', '').strip()
        equipment_name = request.POST.get('equipment', '').strip()
        notes = request.POST.get('notes', '').strip()
        
        # 除錯資訊
        print(f"=== 手動分配 ===")
        print(f"工序ID: {process_id}")
        print(f"工序名稱: {process.process_name}")
        print(f"產品編號: {process.workorder.product_code}")
        print(f"收到的作業員: '{operator_name}'")
        print(f"收到的設備: '{equipment_name}'")
        
        # 基本驗證
        if not operator_name and not equipment_name:
            return JsonResponse({
                "status": "error", 
                "message": "請至少選擇作業員或設備其中一項"
            })
        
        # ====== 純手動分配，不進行自動檢查 ======
        print(f"✅ 執行手動分配")
        
        # 更新欄位
        process.assigned_operator = operator_name
        process.assigned_equipment = equipment_name
        process.status = 'in_progress'
        process.actual_start_time = timezone.now()
        process.save()
        
        # 如果有分配設備，更新設備狀態
        if equipment_name:
            try:
                equipment = Equipment.objects.get(name=equipment_name)
                equipment.status = 'running'
                equipment.save()
                print(f"✅ 設備狀態更新為運轉中：{equipment_name}")
            except Exception as e:
                print(f"⚠️ 設備狀態更新失敗：{str(e)}")
        
        # 重新讀取確認儲存
        process.refresh_from_db()
        print(f"✅ 分配完成")
        print(f"更新後的作業員: {process.assigned_operator}")
        print(f"更新後的設備: {process.assigned_equipment}")
        print(f"更新後的狀態: {process.status}")
        print(f"=== 手動分配結束 ===")
        
        # 記錄操作日誌
        try:
            from workorder.models import WorkOrderProcessLog
            WorkOrderProcessLog.objects.create(
                workorder_process=process,
                action='manual_assignment',
                operator=operator_name,
                equipment=equipment_name,
                notes=f"手動分配 - {notes}" if notes else "手動分配",
                quantity_before=process.completed_quantity,
                quantity_after=process.completed_quantity,
            )
            print(f"✅ 操作日誌記錄完成")
        except Exception as e:
            print(f"⚠️ 操作日誌記錄失敗：{str(e)}")
        
        return JsonResponse({
            "status": "success", 
            "message": "手動分配成功",
            "assigned_operator": operator_name,
            "assigned_equipment": equipment_name,
            "process_status": process.status
        })
        
    except WorkOrderProcess.DoesNotExist:
        return JsonResponse({"status": "error", "message": "找不到工序資料"})
    except Exception as e:
        print(f"❌ 分配過程發生錯誤：{str(e)}")
        return JsonResponse({"status": "error", "message": f"更新失敗：{str(e)}"})



@login_required
def supervisor_report_index(request):
    """
    主管審核首頁 - 統計儀表板
    顯示所有報工類型的統計資訊和快速操作
    """
    from .services.statistics_service import StatisticsService
    
    # 使用統一的統計服務
    stats = StatisticsService.get_supervisor_dashboard_statistics()
    
    # 取得最近審核記錄
    recent_reviews = []
    
    # 從填報記錄中獲取最近審核記錄
    from workorder.fill_work.models import FillWork
    
    fill_work_reviews = FillWork.objects.select_related(
        'process'
    ).filter(
        approval_status='approved'
    ).order_by('-approved_at')[:5]
    
    for review in fill_work_reviews:
        # 判斷報工類型
        if review.process and 'SMT' in review.process.name:
            report_type = 'SMT報工'
            operator_name = 'SMT設備'
        else:
            report_type = '作業員報工'
            operator_name = review.operator
        
        recent_reviews.append({
            'type': report_type,
            'time': review.approved_at,
            'operator': operator_name,
            'workorder': review.workorder,
            'process': review.process.name if review.process else '-',
            'quantity': review.work_quantity,
            'reviewer': review.approved_by,
        })
    
    context = {
        'stats': stats,
        'recent_reviews': recent_reviews,
    }
    
    return render(request, 'supervisor/index.html', context)

@login_required
def supervisor_functions(request):
    """
    主管功能首頁 - 主管專用功能
    包含報工統計、異常處理、系統設定等功能
    """
    from .services.statistics_service import StatisticsService
    
    # 使用統一的統計服務
    stats_data = StatisticsService.get_report_statistics()
    
    # 格式化為主管功能頁面需要的格式
    stats = {
        # 報工統計
        'total_reports_today': stats_data['total']['total_today'],
        'total_reports_month': stats_data['total']['total_month'],
        'pending_reports': stats_data['total']['total_pending'],
        'abnormal_reports': stats_data['total']['total_abnormal'],
        
        # 各類型統計
        'operator_reports_today': stats_data['operator']['today'],
        'smt_reports_today': stats_data['smt']['today'],
        'supervisor_reports_today': 0,  # 主管不報工
        
        'operator_reports_month': stats_data['operator']['month'],
        'smt_reports_month': stats_data['smt']['month'],
        'supervisor_reports_month': 0,  # 主管不報工
        
        # 待審核統計
        'pending_operator': stats_data['operator']['pending'],
        'pending_smt': stats_data['smt']['pending'],
        'pending_supervisor': 0,  # 主管不報工
        
        # 異常統計
        'abnormal_operator': stats_data['operator']['abnormal'],
        'abnormal_smt': stats_data['smt']['abnormal'],
        'abnormal_supervisor': 0,  # 主管不報工
    }
    
    # 取得最近異常記錄
    recent_abnormal = []
    
    # 從填報記錄中獲取異常記錄
    from workorder.fill_work.models import FillWork
    
    abnormal_reports = FillWork.objects.select_related(
        'process'
    ).filter(
        Q(abnormal_notes__isnull=False) & ~Q(abnormal_notes='')
    ).order_by('-work_date', '-created_at')[:5]
    
    for report in abnormal_reports:
        # 判斷報工類型
        if report.process and 'SMT' in report.process.name:
            report_type = 'SMT報工'
            operator_name = 'SMT設備'
        else:
            report_type = '作業員報工'
            operator_name = report.operator
        
        recent_abnormal.append({
            'type': report_type,
            'time': report.work_date,
            'operator': operator_name,
            'workorder': report.workorder,
            'process': report.process.name if report.process else '-',
            'remarks': report.abnormal_notes,
            'status': report.approval_status,
        })
    
    # SMT異常
    # 異常記錄已經在上面處理了，這裡不需要重複處理
    
    # 主管異常
    # 主管不報工，移除主管報工相關邏輯
    
    # 按時間排序
    recent_abnormal.sort(key=lambda x: x['time'], reverse=True)
    recent_abnormal = recent_abnormal[:10]  # 只取前10筆
    
    context = {
        'stats': stats,
        'recent_abnormal': recent_abnormal,
    }
    
    return render(request, 'supervisor/functions.html', context)

@login_required
def report_statistics(request):
    """
    報工統計分析頁面
    提供詳細的報工統計資訊、產能分析、效率評估等
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    context = {
        'stats': {
            'total_reports_today': 0,
            'total_reports_week': 0,
            'total_reports_month': 0,
            'total_reports_year': 0,
            'avg_daily_reports': 0,
            'avg_weekly_reports': 0,
            'avg_monthly_reports': 0,
        },
        'operator_stats': {
            'today': 0,
            'week': 0,
            'month': 0,
            'year': 0,
            'pending': 0,
            'approved': 0,
            'rejected': 0,
        },
        'smt_stats': {
            'today': 0,
            'week': 0,
            'month': 0,
            'year': 0,
            'pending': 0,
            'approved': 0,
            'rejected': 0,
        },
        'capacity_analysis': {
            'operator_efficiency': 0,
            'smt_efficiency': 0,
            'total_efficiency': 0,
        },
        'message': '此功能已棄用，請使用新的報工系統查看統計資訊'
    }
    
    return render(request, 'workorder/report_statistics.html', context)

@login_required
def abnormal_management(request):
    """
    異常處理頁面
    處理報工異常情況，包含異常原因分析、處理方案制定等
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    context = {
        'operator_abnormal': [],
        'smt_abnormal': [],
        'abnormal_stats': {
            'total_abnormal': 0,
            'operator_abnormal': 0,
            'smt_abnormal': 0,
            'resolved': 0,
            'pending': 0,
            'critical': 0,
        },
        'message': '此功能已棄用，請使用新的報工系統處理異常'
    }
    
    return render(request, 'supervisor/abnormal.html', context)

@require_POST
@login_required
def batch_resolve_abnormal(request):
    """
    批次解決異常 - 只有超級管理員可以使用
    將所有待處理的異常標記為已解決
    """
    import json
    from django.http import JsonResponse
    
    # 檢查權限
    if not request.user.is_superuser:
        return JsonResponse({
            'success': False,
            'error': '只有超級管理員可以使用此功能'
        }, status=403)
    
    try:
        # 解析請求資料
        data = json.loads(request.body)
        action = data.get('action')
        
        if action != 'batch_resolve':
            return JsonResponse({
                'success': False,
                'error': '無效的操作'
            })
        
        # 取得所有待處理的異常
        from datetime import date, timedelta
        from django.db.models import Q
        from django.utils import timezone
        
        today = date.today()
        week_start = today - timedelta(days=7)
        
        # 作業員異常 - 標記為已解決（包含所有狀態的異常）
        operator_abnormal = OperatorSupplementReport.objects.filter(
            Q(abnormal_notes__isnull=False) & ~Q(abnormal_notes='') & 
            Q(work_date__gte=week_start)
        )
        
        # 批次更新作業員異常（只處理尚未解決的）
        current_time = timezone.now().strftime('%Y-%m-%d %H:%M:%S')
        operator_resolved_count = 0
        for report in operator_abnormal:
            if report.approval_status != 'approved':
                report.approval_status = 'approved'
                report.approved_at = timezone.now()
                report.approved_by = request.user.username
                report.abnormal_notes = f"{report.abnormal_notes} [批次解決 - {request.user.username} - {current_time}]"
                report.save()
                operator_resolved_count += 1
        
        # SMT異常 - 標記為已解決（包含所有狀態的異常）
        smt_abnormal = SMTSupplementReport.objects.filter(
            Q(abnormal_notes__isnull=False) & ~Q(abnormal_notes='') & 
            Q(work_date__gte=week_start)
        )
        
        # 批次更新SMT異常（只處理尚未解決的）
        smt_resolved_count = 0
        for report in smt_abnormal:
            if report.approval_status != 'approved':
                report.approval_status = 'approved'
                report.approved_at = timezone.now()
                report.approved_by = request.user.username
                report.abnormal_notes = f"{report.abnormal_notes} [批次解決 - {request.user.username} - {current_time}]"
                report.save()
                smt_resolved_count += 1
        
        total_resolved = operator_resolved_count + smt_resolved_count
        
        # 記錄操作日誌
        from django.contrib import messages
        messages.success(request, f'批次解決完成！已解決 {total_resolved} 個異常')
        
        return JsonResponse({
            'success': True,
            'resolved_count': total_resolved,
            'operator_resolved': operator_resolved_count,
            'smt_resolved': smt_resolved_count,
            'message': f'批次解決完成！已解決 {total_resolved} 個異常'
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': '請求格式錯誤'
        })
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'處理過程中發生錯誤：{str(e)}'
        })

def abnormal_detail(request, abnormal_type, abnormal_id):
    """
    異常詳情頁面
    顯示特定異常的詳細資訊
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    from django.contrib import messages
    from django.shortcuts import redirect
    
    messages.warning(request, '此功能已棄用，請使用新的報工系統查看異常詳情')
    return redirect('workorder:abnormal_management')

# 系統設定功能已移至系統管理模組，請使用 system:workorder_settings

@login_required
def data_maintenance(request):
    """
    資料維護頁面
    維護報工相關資料，包含資料清理、備份還原等
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    context = {
        'data_stats': {
            'total_operator_reports': 0,
            'total_smt_reports': 0,
            'old_reports_30d': 0,
            'old_reports_90d': 0,
            'duplicate_reports': 0,
            'orphaned_reports': 0,
        },
        'maintenance_options': [
            {'id': 'cleanup_old', 'name': '清理舊資料', 'description': '清理30天前的報工記錄'},
            {'id': 'cleanup_duplicates', 'name': '清理重複資料', 'description': '清理重複的報工記錄'},
            {'id': 'optimize_database', 'name': '資料庫優化', 'description': '優化資料庫效能'},
            {'id': 'export_archive', 'name': '匯出歸檔', 'description': '匯出舊資料進行歸檔'},
        ],
        'db_size': '0 MB',
        'message': '此功能已棄用，請使用新的報工系統進行資料維護'
    }
    
    return render(request, 'supervisor/data_maintenance.html', context)

@require_POST
@login_required
def execute_maintenance(request):
    """
    執行維護操作的API端點
    處理各種資料維護功能
    """
    from datetime import date, timedelta
    from django.db.models import Q, Count
    from django.db import connection
    from django.http import JsonResponse
    import json
    import os
    from django.conf import settings
    
    action = request.POST.get('action')
    
    if not action:
        return JsonResponse({'success': False, 'message': '缺少操作類型'})
    
    try:
        if action == 'cleanup_old':
            # 清理30天前的舊資料
            cutoff_date = date.today() - timedelta(days=30)
            
            # 清理作業員報工記錄
            old_operator_reports = OperatorSupplementReport.objects.filter(
                work_date__lt=cutoff_date
            )
            operator_count = old_operator_reports.count()
            old_operator_reports.delete()
            
            # 清理SMT報工記錄
            old_smt_reports = SMTSupplementReport.objects.filter(
                work_date__lt=cutoff_date
            )
            smt_count = old_smt_reports.count()
            old_smt_reports.delete()
            
            total_deleted = operator_count + smt_count
            
            return JsonResponse({
                'success': True,
                'message': f'成功清理 {total_deleted} 筆舊資料（作業員報工：{operator_count}，SMT報工：{smt_count}）'
            })
            
        elif action == 'cleanup_duplicates':
            # 清理重複資料
            deleted_count = 0
            
            # 清理作業員報工重複記錄
            # 定義重複：相同的作業員、工單、工序、日期、時間範圍
            operator_duplicates = OperatorSupplementReport.objects.values(
                'operator', 'workorder', 'process', 'work_date', 'start_time', 'end_time'
            ).annotate(
                count=Count('id')
            ).filter(count__gt=1)
            
            for duplicate in operator_duplicates:
                # 保留最新的記錄，刪除其他重複記錄
                reports = OperatorSupplementReport.objects.filter(
                    operator=duplicate['operator'],
                    workorder=duplicate['workorder'],
                    process=duplicate['process'],
                    work_date=duplicate['work_date'],
                    start_time=duplicate['start_time'],
                    end_time=duplicate['end_time']
                ).order_by('-work_date', '-start_time')
                
                # 刪除除最新記錄外的所有重複記錄
                reports_to_delete = reports[1:]
                deleted_count += reports_to_delete.count()
                reports_to_delete.delete()
            
            # 清理SMT報工重複記錄
            # 定義重複：相同的設備、工單、產品、報工時間
            smt_duplicates = SMTSupplementReport.objects.values(
                'equipment', 'workorder', 'rd_product_code', 'work_date', 'start_time', 'end_time'
            ).annotate(
                count=Count('id')
            ).filter(count__gt=1)
            
            for duplicate in smt_duplicates:
                # 保留最新的記錄，刪除其他重複記錄
                reports = SMTSupplementReport.objects.filter(
                    equipment=duplicate['equipment'],
                    workorder=duplicate['workorder'],
                    rd_product_code=duplicate['rd_product_code'],
                    work_date=duplicate['work_date'],
                    start_time=duplicate['start_time'],
                    end_time=duplicate['end_time']
                ).order_by('-work_date', '-start_time')
                
                # 刪除除最新記錄外的所有重複記錄
                reports_to_delete = reports[1:]
                deleted_count += reports_to_delete.count()
                reports_to_delete.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'成功清理 {deleted_count} 筆重複資料'
            })
            
        elif action == 'optimize_database':
            # 資料庫優化
            with connection.cursor() as cursor:
                # 分析資料表
                cursor.execute("ANALYZE workorder_operatorsupplementreport;")
                cursor.execute("ANALYZE workorder_smt_supplement_report;")
                cursor.execute("ANALYZE workorder_workorder;")
                cursor.execute("ANALYZE workorder_workorderprocess;")
                
                # 重新整理資料表
                cursor.execute("REINDEX TABLE workorder_operatorsupplementreport;")
                cursor.execute("REINDEX TABLE workorder_smt_supplement_report;")
                cursor.execute("REINDEX TABLE workorder_workorder;")
                cursor.execute("REINDEX TABLE workorder_workorderprocess;")
                
                # 清理過期的統計資訊
                cursor.execute("VACUUM ANALYZE workorder_operatorsupplementreport;")
                cursor.execute("VACUUM ANALYZE workorder_smt_supplement_report;")
                cursor.execute("VACUUM ANALYZE workorder_workorder;")
                cursor.execute("VACUUM ANALYZE workorder_workorderprocess;")
            
            return JsonResponse({
                'success': True,
                'message': '資料庫優化完成，已重新整理索引並更新統計資訊'
            })
            
        elif action == 'export_archive':
            # 匯出歸檔資料
            from django.core import serializers
            
            # 取得90天前的資料進行歸檔
            cutoff_date = date.today() - timedelta(days=90)
            
            # 收集要歸檔的資料
            archive_data = {
                'operator_reports': OperatorSupplementReport.objects.filter(
                    created_at__lt=cutoff_date
                ),
                'smt_reports': SMTSupplementReport.objects.filter(
                    created_at__lt=cutoff_date
                ),
                'completed_workorders': WorkOrder.objects.filter(
                    status='completed',
                    created_at__lt=cutoff_date
                )
            }
            
            # 建立歸檔目錄
            archive_dir = os.path.join(settings.MEDIA_ROOT, 'archives')
            os.makedirs(archive_dir, exist_ok=True)
            
            # 建立歸檔檔案
            archive_filename = f'workorder_archive_{date.today().strftime("%Y%m%d")}.json'
            archive_path = os.path.join(archive_dir, archive_filename)
            
            # 序列化資料
            all_data = []
            for model_name, queryset in archive_data.items():
                for obj in queryset:
                    all_data.append({
                        'model': f'workorder.{obj._meta.model_name}',
                        'pk': obj.pk,
                        'fields': serializers.serialize('json', [obj])[1:-1]  # 移除外層的方括號
                    })
            
            # 寫入檔案
            with open(archive_path, 'w', encoding='utf-8') as f:
                json.dump(all_data, f, ensure_ascii=False, indent=2)
            
            total_records = (
                archive_data['operator_reports'].count() +
                archive_data['smt_reports'].count() +
                archive_data['completed_workorders'].count()
            )
            
            return JsonResponse({
                'success': True,
                'message': f'成功匯出 {total_records} 筆歸檔資料到 {archive_filename}',
                'download_url': f'/media/archives/{archive_filename}'
            })
            
        elif action == 'check_duplicates':
            # 檢查重複資料但不刪除，只回報統計
            duplicate_info = {
                'operator_duplicates': [],
                'smt_duplicates': [],
                'total_operator_duplicates': 0,
                'total_smt_duplicates': 0
            }
            
            # 檢查作業員報工重複記錄
            # 定義重複：所有重要欄位都完全相同
            operator_duplicates = OperatorSupplementReport.objects.values(
                'operator', 'workorder', 'process', 'work_date', 'start_time', 'end_time',
                'work_quantity', 'defect_quantity', 'report_type', 'product_id',
                'equipment', 'remarks', 'abnormal_notes'
            ).annotate(
                count=Count('id')
            ).filter(count__gt=1)
            
            for duplicate in operator_duplicates:
                reports = OperatorSupplementReport.objects.filter(
                    operator=duplicate['operator'],
                    workorder=duplicate['workorder'],
                    process=duplicate['process'],
                    work_date=duplicate['work_date'],
                    start_time=duplicate['start_time'],
                    end_time=duplicate['end_time'],
                    work_quantity=duplicate['work_quantity'],
                    defect_quantity=duplicate['defect_quantity'],
                    report_type=duplicate['report_type'],
                    product_id=duplicate['product_id'],
                    equipment=duplicate['equipment'],
                    remarks=duplicate['remarks'],
                    abnormal_notes=duplicate['abnormal_notes']
                ).select_related('operator', 'workorder', 'process').order_by('-created_at')
                
                duplicate_info['operator_duplicates'].append({
                    'operator_name': reports[0].operator.name if reports[0].operator else '未知',
                    'workorder_number': reports[0].workorder.order_number if reports[0].workorder else '未知',
                    'process_name': reports[0].process.name if reports[0].process else '未知',
                    'work_date': duplicate['work_date'].isoformat(),
                    'start_time': duplicate['start_time'].isoformat(),
                    'end_time': duplicate['end_time'].isoformat(),
                    'duplicate_count': duplicate['count'],
                    'report_ids': [str(r.id) for r in reports],
                    'created_times': [r.created_at.isoformat() for r in reports]
                })
                duplicate_info['total_operator_duplicates'] += duplicate['count'] - 1
            
            # 檢查SMT報工重複記錄
            # 定義重複：所有重要欄位都完全相同
            smt_duplicates = SMTSupplementReport.objects.values(
                'equipment', 'workorder', 'rd_product_code', 'work_date', 'start_time', 'end_time',
                'work_quantity', 'defect_quantity', 'report_type', 'product_id',
                'operation', 'remarks', 'abnormal_notes'
            ).annotate(
                count=Count('id')
            ).filter(count__gt=1)
            
            for duplicate in smt_duplicates:
                reports = SMTSupplementReport.objects.filter(
                    equipment=duplicate['equipment'],
                    workorder=duplicate['workorder'],
                    rd_product_code=duplicate['rd_product_code'],
                    work_date=duplicate['work_date'],
                    start_time=duplicate['start_time'],
                    end_time=duplicate['end_time'],
                    work_quantity=duplicate['work_quantity'],
                    defect_quantity=duplicate['defect_quantity'],
                    report_type=duplicate['report_type'],
                    product_id=duplicate['product_id'],
                    operation=duplicate['operation'],
                    remarks=duplicate['remarks'],
                    abnormal_notes=duplicate['abnormal_notes']
                ).select_related('equipment', 'workorder').order_by('-created_at')
                
                duplicate_info['smt_duplicates'].append({
                    'equipment_name': reports[0].equipment.name if reports[0].equipment else '未知',
                    'workorder_number': reports[0].workorder.order_number if reports[0].workorder else '未知',
                    'product_code': duplicate['rd_product_code'] or '未知',
                    'work_date': duplicate['work_date'].isoformat(),
                    'start_time': duplicate['start_time'].isoformat(),
                    'end_time': duplicate['end_time'].isoformat(),
                    'duplicate_count': duplicate['count'],
                    'report_ids': [str(r.id) for r in reports],
                    'created_times': [r.created_at.isoformat() for r in reports]
                })
                duplicate_info['total_smt_duplicates'] += duplicate['count'] - 1
            
            return JsonResponse({
                'success': True,
                'message': f'檢查完成：作業員報工重複 {duplicate_info["total_operator_duplicates"]} 筆，SMT報工重複 {duplicate_info["total_smt_duplicates"]} 筆',
                'data': duplicate_info
            })
            
        elif action == 'cleanup_duplicates':
            # 清理重複資料
            deleted_count = 0
            
            # 清理作業員報工重複記錄
            # 定義重複：所有重要欄位都完全相同
            operator_duplicates = OperatorSupplementReport.objects.values(
                'operator', 'workorder', 'process', 'work_date', 'start_time', 'end_time',
                'work_quantity', 'defect_quantity', 'report_type', 'product_id',
                'equipment', 'remarks', 'abnormal_notes'
            ).annotate(
                count=Count('id')
            ).filter(count__gt=1)
            
            for duplicate in operator_duplicates:
                # 保留最新的記錄，刪除其他重複記錄
                reports = OperatorSupplementReport.objects.filter(
                    operator=duplicate['operator'],
                    workorder=duplicate['workorder'],
                    process=duplicate['process'],
                    work_date=duplicate['work_date'],
                    start_time=duplicate['start_time'],
                    end_time=duplicate['end_time'],
                    work_quantity=duplicate['work_quantity'],
                    defect_quantity=duplicate['defect_quantity'],
                    report_type=duplicate['report_type'],
                    product_id=duplicate['product_id'],
                    equipment=duplicate['equipment'],
                    remarks=duplicate['remarks'],
                    abnormal_notes=duplicate['abnormal_notes']
                ).order_by('-work_date', '-start_time')
                
                # 刪除除最新記錄外的所有重複記錄
                reports_to_delete = reports[1:]
                deleted_count += reports_to_delete.count()
                reports_to_delete.delete()
            
            # 清理SMT報工重複記錄
            # 定義重複：相同的設備、工單、產品、報工時間
            smt_duplicates = SMTSupplementReport.objects.values(
                'equipment', 'workorder', 'rd_product_code', 'work_date', 'start_time', 'end_time'
            ).annotate(
                count=Count('id')
            ).filter(count__gt=1)
            
            for duplicate in smt_duplicates:
                # 保留最新的記錄，刪除其他重複記錄
                reports = SMTSupplementReport.objects.filter(
                    equipment=duplicate['equipment'],
                    workorder=duplicate['workorder'],
                    rd_product_code=duplicate['rd_product_code'],
                    work_date=duplicate['work_date'],
                    start_time=duplicate['start_time'],
                    end_time=duplicate['end_time']
                ).order_by('-work_date', '-start_time')
                
                # 刪除除最新記錄外的所有重複記錄
                reports_to_delete = reports[1:]
                deleted_count += reports_to_delete.count()
                reports_to_delete.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'成功清理 {deleted_count} 筆重複資料'
            })
            
        else:
            return JsonResponse({'success': False, 'message': '不支援的操作類型'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'執行失敗：{str(e)}'})







@require_POST
@login_required
def submit_smt_report(request):
    """
    API：提交 SMT 報工記錄
    SMT 設備為自動化運作，不需要作業員
    """
    try:
        equipment_id = request.POST.get('equipment_id')
        workorder_id = request.POST.get('workorder_id')
        process_id = request.POST.get('process_id')
        quantity = request.POST.get('quantity')
        notes = request.POST.get('notes', '')
        
        # 基本驗證
        if not all([equipment_id, workorder_id, process_id]):
            return JsonResponse({
                'status': 'error',
                'message': '請填寫所有必要欄位'
            })
        
        # 處理報工數量，允許為 0
        if quantity is None or quantity == '':
            quantity = 0
        else:
            try:
                quantity = int(quantity)
                if quantity < 0:
                    return JsonResponse({
                        'status': 'error',
                        'message': '報工數量不能為負數'
                    })
            except ValueError:
                return JsonResponse({
                    'status': 'error',
                    'message': '報工數量格式錯誤'
                })
        
        # 取得設備和工單
        from equip.models import Equipment
        equipment = Equipment.objects.get(id=equipment_id)
        workorder = WorkOrder.objects.get(id=workorder_id)
        
        # 建立報工記錄
        from process.models import ProcessName
        from django.utils import timezone
        
        # 取得選擇的工序（從 WorkOrderProcess 讀取）
        try:
            selected_process = WorkOrderProcess.objects.get(id=process_id)
            process_name = selected_process.process_name
        except WorkOrderProcess.DoesNotExist:
            return JsonResponse({
                'status': 'error',
                'message': '找不到指定的工序'
            })
        
        # 使用SMT作業員服務建立報工記錄
        from .services.smt_operator_service import SMTOperatorService
        
        report = SMTOperatorService.create_smt_report_with_operator_name(
            equipment=equipment,
            product_id=workorder.product_code,
            workorder=workorder,
            planned_quantity=workorder.quantity,
            operation=process_name,
            work_date=timezone.now().date(),
            start_time=timezone.now().time(),
            end_time=timezone.now().time(),
            work_quantity=quantity,
            defect_quantity=0,
            is_completed=False,
            remarks=notes,
            created_by=request.user.username if request.user.is_authenticated else 'system'
        )
        
        return JsonResponse({
            'status': 'success',
            'message': 'SMT 報工記錄提交成功',
            'report_id': report.id
        })
        
    except Equipment.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': '找不到指定的設備'
        })
    except WorkOrder.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': '找不到指定的工單'
        })
    except WorkOrderProcess.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': '找不到指定的工序'
        })
    except ValueError as e:
        return JsonResponse({
            'status': 'error',
            'message': f'資料格式錯誤：{str(e)}'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'提交失敗：{str(e)}'
        })

# ============================================================================
# 作業員補登報工功能視圖函數
# ============================================================================

# 此函數已被移除，請使用 OperatorSupplementReportListView 類別視圖
# 位置：workorder/views/report_views.py

# 此函數已被移除，請使用 OperatorSupplementReportCreateView 類別視圖
# 位置：workorder/views/report_views.py

# 此函數已被移除，請使用 OperatorSupplementReportUpdateView 類別視圖
# 位置：workorder/views/report_views.py

# 此函數已被移除，請使用 OperatorSupplementReportDeleteView 類別視圖
# 位置：workorder/views/report_views.py

# 此函數已被移除，請使用 OperatorSupplementReportDetailView 類別視圖
# 位置：workorder/views/report_views.py



# ============================================================================
# SMT補登報工功能視圖函數
# ============================================================================

# 此函數已被移除，請使用 SMTSupplementReportListView 類別視圖
# 位置：workorder/views/report_views.py

# 此函數已被移除，請使用 SMTSupplementReportCreateView 類別視圖
# 位置：workorder/views/report_views.py

# 此函數已被移除，請使用 SMTSupplementReportUpdateView 類別視圖
# 位置：workorder/views/report_views.py

# 此函數已被移除，請使用 SMTSupplementReportDeleteView 類別視圖
# 位置：workorder/views/report_views.py

# 此函數已被移除，請使用 SMTSupplementReportDetailView 類別視圖
# 位置：workorder/views/report_views.py

# ============================================================================
# 主管審核功能視圖函數
# ============================================================================

@login_required
def supervisor_approve_reports(request):
    """主管審核報工記錄
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    context = {
        'page_obj': [],
        'search': '',
        'total_count': 0,
        'message': '此功能已棄用，請使用新的報工系統進行審核'
    }
    return render(request, 'supervisor/pending_approval_list.html', context)

@login_required
def approve_report(request, report_id):
    """審核單筆報工記錄
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    from django.contrib import messages
    messages.warning(request, '此功能已棄用，請使用新的報工系統進行審核')
    return redirect('workorder:supervisor_approve_reports')

@login_required
def reject_report(request, report_id):
    """駁回單筆報工記錄
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    from django.contrib import messages
    messages.warning(request, '此功能已棄用，請使用新的報工系統進行審核')
    return redirect('workorder:supervisor_approve_reports')

# ============================================================================
# API功能視圖函數
# ============================================================================

@require_GET
@login_required
def api_get_operator_reports(request):
    """API：取得作業員報工記錄
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    return JsonResponse({
        'status': 'error',
        'message': '此功能已棄用，請使用新的報工系統',
        'data': [],
        'total_count': 0
    })

@require_GET
@login_required
def api_get_smt_reports(request):
    """API：取得SMT報工記錄
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    return JsonResponse({
        'status': 'error',
        'message': '此功能已棄用，請使用新的報工系統',
        'data': [],
        'total_count': 0
    })

# ============================================================================
# 報表匯出功能視圖函數
# ============================================================================

@login_required
def export_operator_reports(request):
    """匯出作業員報工記錄
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    from django.contrib import messages
    messages.warning(request, '此功能已棄用，請使用新的報工系統進行匯出')
    return redirect('workorder:index')

# ============================================================================
# 審核功能視圖函數
# ============================================================================

@login_required
def operator_supplement_report_approve(request, report_id):
    """審核作業員補登報工
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    from django.contrib import messages
    messages.warning(request, '此功能已棄用，請使用新的報工系統進行審核')
    return redirect('workorder:index')

@login_required
def operator_supplement_report_reject(request, report_id):
    """駁回作業員補登報工
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    from django.contrib import messages
    messages.warning(request, '此功能已棄用，請使用新的報工系統進行審核')
    return redirect('workorder:index')

@login_required
def smt_supplement_report_approve(request, report_id):
    """審核SMT補登報工"""
    try:
        report = SMTSupplementReport.objects.get(id=report_id)
        report.approval_status = 'approved'
        report.approved_by = request.user.username
        report.approved_at = timezone.now()
        report.save()
        
        # 核准成功後，同步到生產中工單詳情資料表並更新工序完成數量
        try:
            from workorder.services import ProductionReportSyncService
            from workorder.services.process_update_service import ProcessUpdateService
            if hasattr(report, 'workorder') and report.workorder:
                # 同步到生產詳情資料表
                ProductionReportSyncService.sync_specific_workorder(report.workorder.id)
                # 更新工序完成數量
                ProcessUpdateService.update_workorder_processes(report.workorder.id)
        except Exception as sync_error:
            # 同步失敗不影響核准流程，只記錄錯誤
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"同步SMT報工記錄到生產詳情失敗: {str(sync_error)}")
        
        messages.success(request, 'SMT報工記錄審核通過！')
    except SMTSupplementReport.DoesNotExist:
        messages.error(request, '找不到指定的SMT報工記錄！')
    
    return redirect('workorder:smt_supplement_report_index')

@login_required
def smt_supplement_report_reject(request, report_id):
    """駁回SMT補登報工"""
    try:
        report = SMTSupplementReport.objects.get(id=report_id)
        report.approval_status = 'rejected'
        report.approved_by = request.user.username
        report.approved_at = timezone.now()
        report.save()
        messages.success(request, 'SMT報工記錄已駁回！')
    except SMTSupplementReport.DoesNotExist:
        messages.error(request, '找不到指定的SMT報工記錄！')
    
    return redirect('workorder:smt_supplement_report_index')

# ============================================================================
# 批次處理功能視圖函數
# ============================================================================



def operator_supplement_import_file(request):
    """作業員補登報工檔案匯入處理
    注意：此功能已棄用，相關模型已移至新的報工系統
    """
    return JsonResponse({
        'success': False,
        'message': '此功能已棄用，請使用新的報工系統進行匯入'
    })
    
    try:
        uploaded_file = request.FILES.get('file')
        if not uploaded_file:
            return JsonResponse({
                'success': False,
                'message': '沒有選擇檔案'
            })
        
        # 檢查檔案格式
        file_name = uploaded_file.name.lower()
        if not (file_name.endswith('.xlsx') or file_name.endswith('.csv')):
            return JsonResponse({
                'success': False,
                'message': '只支援 Excel (.xlsx) 或 CSV 格式檔案'
            })
        
        # 讀取檔案內容
        import pandas as pd
        from io import BytesIO
        
        if file_name.endswith('.xlsx'):
            df = pd.read_excel(uploaded_file)
        else:
            df = pd.read_csv(uploaded_file)
        
        # 檢查必要欄位 - 只支援新格式
        required_columns = ['作業員名稱', '公司代號', '報工日期', '開始時間', '結束時間', '工單號', '產品編號', '工序名稱', '報工數量']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return JsonResponse({
                'success': False,
                'message': f'缺少必要欄位：{", ".join(missing_columns)}'
            })
        
        # 處理每一行資料
        success_count = 0
        error_count = 0
        errors = []
        
        for index, row in df.iterrows():
            try:
                # 新格式處理
                operator_name = str(row['作業員名稱']).strip()
                
                # 處理公司代號 - 修正 pandas 自動轉換的數字格式
                company_code_raw = row['公司代號']
                if pd.isna(company_code_raw):
                    raise ValueError('公司代號為空')
                
                # 處理 pandas 自動轉換的數字格式
                if isinstance(company_code_raw, (int, float)):
                    # 如果是數字，轉換為整數後再轉字串
                    company_code = str(int(company_code_raw))
                else:
                    # 如果是字串，直接使用
                    company_code = str(company_code_raw).strip()
                
                # 確保公司代號是兩位數格式
                if company_code.isdigit():
                    company_code = company_code.zfill(2)  # 確保是兩位數格式
                
                workorder_number = str(row['工單號']).strip()
                process_name = str(row['工序名稱']).strip()
                work_quantity = int(row['報工數量'])
                
                # 處理產品編號
                product_code_raw = row['產品編號']
                # 處理 pandas 讀取的空值問題
                if pd.isna(product_code_raw) or str(product_code_raw).strip() == '' or str(product_code_raw).strip().lower() == 'nan':
                    product_code = ""
                else:
                    product_code = str(product_code_raw).strip()
                
                # 解析日期 - 支援多種格式
                date_str = str(row['報工日期']).strip()
                if '/' in date_str:
                    work_date = datetime.strptime(date_str, '%Y/%m/%d').date()
                elif '.' in date_str:
                    work_date = datetime.strptime(date_str, '%Y.%m.%d').date()
                else:
                    work_date = pd.to_datetime(date_str).date()
                
                # 解析時間 - 支援所有 Excel 時間格式
                start_time_str = str(row['開始時間']).strip()
                try:
                    if 'days' in start_time_str:
                        # 處理 Excel 時間格式 "0 days 08:30:00"
                        time_part = start_time_str.split(' ')[-1]  # 取得 "08:30:00" 部分
                        start_time = datetime.strptime(time_part, '%H:%M:%S').time()
                    elif 'AM' in start_time_str or 'PM' in start_time_str:
                        # 處理 12 小時制格式 "04:30:00 PM"
                        start_time = pd.to_datetime(start_time_str).time()
                    elif ':' in start_time_str:
                        # 處理標準時間格式 "08:30:00" 或 "08:30"
                        if start_time_str.count(':') == 2:
                            start_time = datetime.strptime(start_time_str, '%H:%M:%S').time()
                        else:
                            start_time = datetime.strptime(start_time_str, '%H:%M').time()
                    else:
                        # 處理純數字格式 "0830"
                        start_time = datetime.strptime(start_time_str.zfill(4), '%H%M').time()
                except Exception as e:
                    errors.append(f'第 {index+1} 行：開始時間格式錯誤 "{start_time_str}"')
                    error_count += 1
                    continue
                
                end_time_str = str(row['結束時間']).strip()
                try:
                    if 'days' in end_time_str:
                        # 處理 Excel 時間格式 "0 days 17:30:00"
                        time_part = end_time_str.split(' ')[-1]  # 取得 "17:30:00" 部分
                        end_time = datetime.strptime(time_part, '%H:%M:%S').time()
                    elif 'AM' in end_time_str or 'PM' in end_time_str:
                        # 處理 12 小時制格式 "04:30:00 PM"
                        end_time = pd.to_datetime(end_time_str).time()
                    elif ':' in end_time_str:
                        # 處理標準時間格式 "17:30:00" 或 "17:30"
                        if end_time_str.count(':') == 2:
                            end_time = datetime.strptime(end_time_str, '%H:%M:%S').time()
                        else:
                            end_time = datetime.strptime(end_time_str, '%H:%M').time()
                    else:
                        # 處理純數字格式 "1730"
                        end_time = datetime.strptime(end_time_str.zfill(4), '%H%M').time()
                except Exception as e:
                    errors.append(f'第 {index+1} 行：結束時間格式錯誤 "{end_time_str}"')
                    error_count += 1
                    continue
                
                # 處理選填欄位
                defect_quantity = int(row.get('不良品數量', 0))
                
                # 處理備註（選填）
                remarks_raw = row.get('備註', '')
                if pd.isna(remarks_raw) or str(remarks_raw).strip() == '' or str(remarks_raw).strip().lower() == 'nan':
                    remarks = ""
                else:
                    remarks = str(remarks_raw).strip()
                
                # 處理異常紀錄（選填）
                abnormal_notes_raw = row.get('異常紀錄', '')
                if pd.isna(abnormal_notes_raw) or str(abnormal_notes_raw).strip() == '' or str(abnormal_notes_raw).strip().lower() == 'nan':
                    abnormal_notes = ""
                else:
                    abnormal_notes = str(abnormal_notes_raw).strip()
                
                # 不進行工單查找，直接使用匯入的工單號碼
                
                # 驗證作業員
                operator = Operator.objects.filter(name=operator_name).first()
                if not operator:
                    errors.append(f'第 {index+1} 行：找不到作業員 "{operator_name}"')
                    error_count += 1
                    continue
                
                # 對於舊資料匯入，不檢查工單是否存在，直接使用工單號碼
                # 如果工單不存在，將使用工單號碼作為字串儲存
                
                # 驗證工序
                process = ProcessName.objects.filter(name=process_name).first()
                if not process:
                    errors.append(f'第 {index+1} 行：找不到工序 "{process_name}"')
                    error_count += 1
                    continue
                
                # 建立報工記錄
                # 直接使用匯入的工單號碼，不進行工單查找
                report_data = {
                    'operator': operator,
                    'process': process,
                    'product_id': product_code,  # 加入產品編號
                    'work_date': work_date,
                    'start_time': start_time,
                    'end_time': end_time,
                    'work_quantity': work_quantity,
                    'defect_quantity': defect_quantity,
                    'remarks': remarks,
                    'abnormal_notes': abnormal_notes,
                    'created_by': request.user.username,
                    'original_workorder_number': workorder_number  # 直接使用匯入的工單號碼
                }
                
                # 注意：相關模型已棄用
                pass
                
                success_count += 1
                
            except Exception as e:
                errors.append(f'第 {index+1} 行：{str(e)}')
                error_count += 1
        
        return JsonResponse({
            'success': True,
            'success_count': success_count,
            'error_count': error_count,
            'total_count': len(df),
            'errors': errors[:10]  # 只回傳前10個錯誤
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'匯入失敗：{str(e)}'
        })



# ============================================================================
# 匯出功能視圖函數
# ============================================================================

@login_required
def operator_supplement_export(request):
    """匯出作業員補登報工記錄"""
    try:
        # 注意：相關模型已棄用
        reports = []
        # 篩選條件
        date_from = request.GET.get('date_from')
        if date_from:
            reports = reports.filter(work_date__gte=date_from)
        
        date_to = request.GET.get('date_to')
        if date_to:
            reports = reports.filter(work_date__lte=date_to)
        
        operator_id = request.GET.get('operator_id')
        if operator_id:
            reports = reports.filter(operator_id=operator_id)
        
        # 建立Excel檔案
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "作業員補登報工記錄"
        
        # 設定標題 - 與匯入格式保持一致
        headers = ['作業員名稱', '公司代號', '報工日期', '開始時間', '結束時間', '工單號', '產品編號', '工序名稱', '設備名稱', '報工數量', '不良品數量', '備註', '異常紀錄', '審核狀態', '審核者', '審核時間']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 填入資料
        for row, report in enumerate(reports, 2):
            # 作業員名稱
            ws.cell(row=row, column=1, value=report.operator.name if report.operator else '')
            
            # 公司代號 - 改善處理邏輯
            company_code = ''
            if report.workorder and report.workorder.company_code:
                company_code = report.workorder.company_code
            else:
                # 如果沒有工單關聯，使用預設值
                company_code = '10'  # 預設公司代號
            ws.cell(row=row, column=2, value=company_code)
            
            # 報工日期
            ws.cell(row=row, column=3, value=report.work_date.strftime('%Y-%m-%d'))
            
            # 開始時間
            ws.cell(row=row, column=4, value=report.start_time.strftime('%H:%M') if report.start_time else '')
            
            # 結束時間
            ws.cell(row=row, column=5, value=report.end_time.strftime('%H:%M') if report.end_time else '')
            
            # 工單號 - 使用 workorder_number 屬性
            ws.cell(row=row, column=6, value=report.workorder_number)
            
            # 產品編號 - 直接使用報工記錄本身的產品編號欄位
            product_id = report.product_id or ''
            if not product_id and report.workorder and hasattr(report.workorder, 'product_code') and report.workorder.product_code:
                # 如果報工記錄的產品編號為空，才從工單取得
                product_id = report.workorder.product_code
            ws.cell(row=row, column=7, value=product_id)
            
            # 工序名稱
            ws.cell(row=row, column=8, value=report.process.name if report.process else '')
            
            # 設備名稱
            ws.cell(row=row, column=9, value=report.equipment.name if report.equipment else '')
            
            # 報工數量
            ws.cell(row=row, column=10, value=report.work_quantity or 0)
            
            # 不良品數量
            ws.cell(row=row, column=11, value=report.defect_quantity or 0)
            
            # 備註
            ws.cell(row=row, column=12, value=report.remarks or '')
            
            # 異常紀錄
            ws.cell(row=row, column=13, value=report.abnormal_notes or '')
            
            # 審核狀態
            ws.cell(row=row, column=14, value=report.get_approval_status_display())
            
            # 審核者
            ws.cell(row=row, column=15, value=report.approved_by or '')
            
            # 審核時間
            ws.cell(row=row, column=16, value=report.approved_at.strftime('%Y-%m-%d %H:%M:%S') if report.approved_at else '')
        
        # 調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 儲存檔案
        from django.http import HttpResponse
        from io import BytesIO
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="作業員補登報工記錄_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'匯出失敗：{str(e)}')
        return redirect('workorder:operator_supplement_report_index')

# ============================================================================
# 模板下載功能視圖函數
# ============================================================================

@login_required
def operator_supplement_template(request):
    """下載作業員補登報工模板"""
    try:
        # 建立Excel檔案
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        
        wb = Workbook()
        ws = wb.active
        ws.title = "作業員補登報工模板"
        
        # 設定標題 - 使用新格式
        headers = ['作業員名稱', '公司代號', '報工日期', '開始時間', '結束時間', '工單號', '產品編號', '工序名稱', '設備名稱', '報工數量', '不良品數量', '備註', '異常紀錄']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 添加範例資料 - 使用新格式
        example_data = [
            ['張小明', '01', '2025-01-15', '08:00', '12:00', 'WO-01-202501001', 'PROD-001', 'SMT', 'SMT-001', 100, 2, '正常生產', ''],
            ['李小華', '01', '2025/01/15', '13:00', '17:00', 'WO-01-202501001', 'PROD-001', 'DIP', 'DIP-001', 95, 5, '設備調整', '設備故障30分鐘'],
            ['王小美', '02', '2025.01.15', '08:00', '16:00', 'WO-02-202501002', 'PROD-002', '測試', '', 200, 0, '加班生產', ''],
        ]
        
        for row, data in enumerate(example_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # 調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 儲存檔案
        from django.http import HttpResponse
        from io import BytesIO
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="作業員補登報工模板_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        return response
        
    except Exception as e:
        messages.error(request, f'模板下載失敗：{str(e)}')
        return redirect('workorder:operator_supplement_report_index')

@require_POST
@login_required
def operator_supplement_batch_create(request):
    """API：批次建立作業員補登報工"""
    try:
        data = json.loads(request.body)
        reports_data = data.get('reports', [])
        
        created_count = 0
        errors = []
        
        for report_data in reports_data:
            try:
                # 驗證必要欄位
                required_fields = ['operator_id', 'workorder_id', 'process_id', 'work_quantity', 'work_date']
                for field in required_fields:
                    if field not in report_data:
                        errors.append(f'缺少必要欄位：{field}')
                        continue
                
                # 注意：OperatorSupplementReport 模型已棄用
                pass
                created_count += 1
                
            except Exception as e:
                errors.append(f'建立記錄失敗：{str(e)}')
        
        return JsonResponse({
            'status': 'success',
            'created_count': created_count,
            'errors': errors
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@require_POST
@login_required
def smt_supplement_batch_create(request):
    """API：批次建立SMT補登報工"""
    try:
        data = json.loads(request.body)
        reports_data = data.get('reports', [])
        
        created_count = 0
        errors = []
        
        for report_data in reports_data:
            try:
                # 驗證必要欄位
                required_fields = ['equipment_id', 'workorder_id', 'product_id', 'work_quantity', 'work_date']
                for field in required_fields:
                    if field not in report_data:
                        errors.append(f'缺少必要欄位：{field}')
                        continue
                
                # 使用SMT作業員服務自動設定作業員名稱
                from workorder.services.smt_operator_service import SMTOperatorService
                from equip.models import Equipment
                
                # 取得設備並設定作業員名稱
                equipment = Equipment.objects.get(id=report_data['equipment_id'])
                equipment_operator_name = SMTOperatorService.get_smt_equipment_operator_name(equipment.name)
                
                # 注意：SMTSupplementReport 模型已棄用
                pass
                created_count += 1
                
            except Exception as e:
                errors.append(f'建立記錄失敗：{str(e)}')
        
        return JsonResponse({
            'status': 'success',
            'created_count': created_count,
            'errors': errors
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

# ============================================================================
# 主管功能視圖函數
# ============================================================================

@login_required
def supervisor_index(request):
    """主管功能首頁"""
    # 統計資料
    # 注意：相關模型已棄用
    total_operator_reports = 0
    pending_operator_reports = 0
    approved_operator_reports = 0
    total_smt_reports = 0
    pending_smt_reports = 0
    approved_smt_reports = 0
    
    context = {
        'total_operator_reports': total_operator_reports,
        'pending_operator_reports': pending_operator_reports,
        'approved_operator_reports': approved_operator_reports,
        'total_smt_reports': total_smt_reports,
        'pending_smt_reports': pending_smt_reports,
        'approved_smt_reports': approved_smt_reports,
    }
    return render(request, 'workorder/report/supervisor/index.html', context)

# ============================================================================
# 其他API功能視圖函數
# ============================================================================

@require_GET
@login_required
def get_workorders_by_operator(request):
    """API：根據作業員取得工單"""
    try:
        operator_id = request.GET.get('operator_id')
        if not operator_id:
            return JsonResponse({
                'status': 'error',
                'message': '請提供作業員ID'
            })
        
        # 移除所有過濾：顯示所有工單，完全自由選擇
        workorders = WorkOrder.objects.all().distinct()
        
        workorders_data = []
        for workorder in workorders:
            # 檢查工單是否有進行中的工序
            has_in_progress = workorder.processes.filter(status='in_progress').exists()
            process_status = 'in_progress' if has_in_progress else 'pending'
            
            workorders_data.append({
                'id': workorder.id,
                'order_number': workorder.order_number,
                'company_code': workorder.company_code,
                'product_code': workorder.product_code,
                'quantity': workorder.quantity,
                'status': workorder.status,
                'process_status': process_status
            })
        
        return JsonResponse({
            'status': 'success',
            'workorders': workorders_data
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@require_GET
@login_required
def get_processes_by_workorder_for_operator(request):
    """API：根據工單取得工序（作業員用），自動排除SMT工序"""
    try:
        workorder_id = request.GET.get('workorder_id')
        if not workorder_id:
            return JsonResponse({
                'status': 'error',
                'message': '請提供工單ID'
            })
        from workorder.models import WorkOrderProcess
        processes = WorkOrderProcess.objects.filter(
            workorder_id=workorder_id
        ).exclude(process_name__icontains='SMT').order_by('step_order')
        data = []
        for process in processes:
            data.append({
                'id': process.id,
                'process_name': process.process_name,
                'sequence': process.step_order,
                'status': process.status,
            })
        return JsonResponse({
            'status': 'success',
            'processes': data
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'載入失敗:{str(e)}'
        })

@require_POST
@login_required
def submit_operator_report(request):
    """API：提交作業員報工"""
    try:
        data = json.loads(request.body)
        
        # 驗證必要欄位
        required_fields = ['operator_id', 'workorder_id', 'process_id', 'quantity']
        for field in required_fields:
            if field not in data:
                return JsonResponse({
                    'status': 'error',
                    'message': f'缺少必要欄位：{field}'
                })
        
        # 注意：相關模型已棄用
        pass
        
        return JsonResponse({
            'status': 'success',
            'message': '報工記錄提交成功',
            'report_id': report.id
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@require_GET
@login_required
def get_smt_workorders_by_equipment(request):
    """API：根據設備取得工單（SMT用）"""
    try:
        equipment_id = request.GET.get('equipment_id')
        if not equipment_id:
            return JsonResponse({
                'status': 'error',
                'message': '請提供設備ID'
            })
        
        # 修正查詢條件：使用 processes 作為 related_name
        workorders = WorkOrder.objects.filter(
            processes__assigned_equipment=equipment_id
        ).distinct()
        
        data = []
        for workorder in workorders:
            # 取得該工單的工序狀態
            processes = WorkOrderProcess.objects.filter(
                workorder=workorder,
                assigned_equipment=equipment_id
            ).order_by('step_order')
            
            # 檢查是否有正在進行的工序
            process_status = 'pending'
            for process in processes:
                if process.status == 'in_progress':
                    process_status = 'in_progress'
                    break
                elif process.status == 'completed':
                    process_status = 'completed'
            
            data.append({
                'id': workorder.id,
                'order_number': workorder.order_number,
                'company_code': workorder.company_code,
                'product_code': workorder.product_code,
                'quantity': workorder.quantity,
                'process_status': process_status,
            })
        
        return JsonResponse({
            'status': 'success',
            'workorders': data
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': f'載入失敗:{str(e)}'
        })

@require_GET
@login_required
def get_workorders_by_product(request):
    """API：根據產品編號取得相關工單（直接從工單表取得）"""
    try:
        product_code = request.GET.get('product_id') or request.GET.get('product_code')
        if not product_code:
            return JsonResponse({
                'status': 'error',
                'message': '請提供產品編號'
            })
        
        # 直接從工單表中查詢指定產品編號的工單
        workorders = WorkOrder.objects.filter(
            product_code__icontains=product_code
        ).exclude(
            order_number__icontains="RD樣品"
        ).exclude(
            order_number__icontains="RD-樣品"
        ).exclude(
            order_number__icontains="RD樣本"
        ).exclude(
            status="completed"
        ).order_by('-created_at')  # 按建立時間倒序排列
        
        data = []
        for workorder in workorders:
            data.append({
                'id': workorder.id,
                'order_number': workorder.order_number,
                'company_code': workorder.company_code,
                'product_code': workorder.product_code,
                'quantity': workorder.quantity,
                'status': workorder.status,
            })
        
        return JsonResponse({
            'status': 'success',
            'workorders': data
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@require_GET
@login_required
def get_workorder_details(request):
    """API：取得工單詳細資料"""
    try:
        workorder_id = request.GET.get('workorder_id')
        if not workorder_id:
            return JsonResponse({
                'status': 'error',
                'message': '請提供工單ID'
            })
        
        workorder = WorkOrder.objects.get(id=workorder_id)
        processes = WorkOrderProcess.objects.filter(workorder=workorder).order_by('step_order')
        
        process_data = []
        for process in processes:
            process_data.append({
                'id': process.id,
                'process_name': process.process_name,
                'sequence': process.step_order,  # 修正欄位名稱
                'assigned_equipment': process.assigned_equipment,  # 修正欄位名稱
                'assigned_operator': process.assigned_operator,  # 修正欄位名稱
            })
        
        data = {
            'id': workorder.id,
            'order_number': workorder.order_number,
            'company_code': workorder.company_code,
            'product_code': workorder.product_code,
            'quantity': workorder.quantity,
            'status': workorder.status,
            'processes': process_data,
        }
        
        return JsonResponse({
            'status': 'success',
            'workorder': data
        })
        
    except WorkOrder.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': '工單不存在'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@require_GET
@login_required
def get_product_by_workorder(request):
    """API：根據工單號碼取得產品編號（相向自動帶出）"""
    try:
        workorder_id = request.GET.get('workorder_id')
        if not workorder_id:
            return JsonResponse({
                'status': 'error',
                'message': '請提供工單ID'
            })
        
        workorder = WorkOrder.objects.get(id=workorder_id)
        
        data = {
            'product_code': workorder.product_code,
            'quantity': workorder.quantity,
            'company_code': workorder.company_code,
            'order_number': workorder.order_number,
        }
        
        return JsonResponse({
            'status': 'success',
            'product': data
        })
        
    except WorkOrder.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': '工單不存在'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@require_GET
@login_required
def get_product_codes_for_autocomplete(request):
    """API：獲取產品編號列表用於自動完成"""
    try:
        # 從工單中獲取所有產品編號
        product_codes = WorkOrder.objects.exclude(
            status="completed"
        ).exclude(
            product_code__isnull=True
        ).exclude(
            product_code=""
        ).values_list('product_code', flat=True).distinct().order_by('product_code')
        
        # 從派工單中獲取產品編號
        from .models import WorkOrderDispatch
        dispatch_product_codes = WorkOrderDispatch.objects.exclude(
            product_code__isnull=True
        ).exclude(
            product_code=""
        ).values_list('product_code', flat=True).distinct().order_by('product_code')
        
        # 合併並去重
        all_product_codes = list(set(list(product_codes) + list(dispatch_product_codes)))
        all_product_codes.sort()
        
        return JsonResponse({
            'status': 'success',
            'product_codes': all_product_codes
        })
        
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

@require_GET
@login_required
def get_workorder_info(request):
    """API：取得工單基本資訊"""
    try:
        workorder_id = request.GET.get('workorder_id')
        if not workorder_id:
            return JsonResponse({
                'status': 'error',
                'message': '請提供工單ID'
            })
        
        workorder = WorkOrder.objects.get(id=workorder_id)
        
        data = {
            'id': workorder.id,
            'order_number': workorder.order_number,
            'company_code': workorder.company_code,
            'product_code': workorder.product_code,
            'quantity': workorder.quantity,
            'status': workorder.status,
        }
        
        return JsonResponse({
            'status': 'success',
            'data': data
        })
        
    except WorkOrder.DoesNotExist:
        return JsonResponse({
            'status': 'error',
            'message': '找不到指定的工單'
        })
    except Exception as e:
        return JsonResponse({
            'status': 'error',
            'message': str(e)
        })

# 派工單管理視圖函數
@login_required
def dispatch_dashboard(request):
    """派工單儀表板"""
    from .workorder_dispatch.models import WorkOrderDispatch
    
    # 統計資料
    total_dispatches = WorkOrderDispatch.objects.count()
    pending_dispatches = WorkOrderDispatch.objects.filter(status='pending').count()
    in_progress_dispatches = WorkOrderDispatch.objects.filter(status='in_progress').count()
    completed_dispatches = WorkOrderDispatch.objects.filter(status='completed').count()
    
    context = {
        'total_dispatches': total_dispatches,
        'pending_dispatches': pending_dispatches,
        'in_progress_dispatches': in_progress_dispatches,
        'completed_dispatches': completed_dispatches,
    }
    
    return render(request, 'workorder_dispatch/dispatch_dashboard.html', context)

@login_required
def dispatch_add(request):
    """新增派工單"""
    from .workorder_dispatch.models import WorkOrderDispatch
    from .workorder_dispatch.forms import WorkOrderDispatchForm
    
    if request.method == 'POST':
        form = WorkOrderDispatchForm(request.POST)
        if form.is_valid():
            form.instance.created_by = request.user.username
            form.save()
            messages.success(request, '派工單建立成功！')
            return redirect('workorder_dispatch:dispatch_list')
    else:
        form = WorkOrderDispatchForm()
    
    context = {
        'form': form,
        'title': '新增派工單',
        'submit_text': '建立派工單'
    }
    return render(request, 'workorder_dispatch/dispatch_form.html', context)

@login_required
def dispatch_edit(request, pk):
    """編輯派工單"""
    from .workorder_dispatch.models import WorkOrderDispatch
    from .workorder_dispatch.forms import WorkOrderDispatchForm
    
    dispatch = get_object_or_404(WorkOrderDispatch, pk=pk)
    
    if request.method == 'POST':
        form = WorkOrderDispatchForm(request.POST, instance=dispatch)
        if form.is_valid():
            form.save()
            messages.success(request, '派工單更新成功！')
            return redirect('workorder_dispatch:dispatch_list')
    else:
        form = WorkOrderDispatchForm(instance=dispatch)
    
    context = {
        'form': form,
        'dispatch': dispatch,
        'title': '編輯派工單',
        'submit_text': '更新派工單'
    }
    return render(request, 'workorder_dispatch/dispatch_form.html', context)

@login_required
def dispatch_detail(request, pk):
    """派工單詳情"""
    from .workorder_dispatch.models import WorkOrderDispatch
    
    dispatch = get_object_or_404(WorkOrderDispatch, pk=pk)
    
    context = {
        'dispatch': dispatch
    }
    return render(request, 'workorder_dispatch/dispatch_detail.html', context)

@login_required
def dispatch_delete(request, pk):
    """刪除派工單"""
    from .workorder_dispatch.models import WorkOrderDispatch
    
    dispatch = get_object_or_404(WorkOrderDispatch, pk=pk)
    
    if request.method == 'POST':
        dispatch.delete()
        messages.success(request, '派工單刪除成功！')
        return redirect('workorder_dispatch:dispatch_list')
    
    context = {
        'dispatch': dispatch
    }
    return render(request, 'workorder_dispatch/dispatch_confirm_delete.html', context)

@login_required
def active_workorders(request):
    """
    生產執行監控視圖
    顯示主管審核後的真正填報紀錄統計
    功能：基於已核准填報記錄監控生產執行狀況
    """
    from django.db.models import Q, Count, Sum
    from workorder.models import WorkOrder
    from workorder.fill_work.models import FillWork
    from erp_integration.models import CompanyConfig
    from datetime import date, timedelta
    
    # 獲取今天的日期
    today = date.today()
    
    # 新的生產中工單判斷邏輯：
    # 生產中工單 = 所有有已核准填報記錄的工單（只計算主管審核後的記錄）
    # 這樣可以監控所有真正在生產的工單
    
    # 獲取所有有已核准填報記錄的工單（正確區分公司）
    # 先獲取已核准填報記錄的公司和工單號碼對應關係
    approved_fill_works = FillWork.objects.filter(
        approval_status='approved'
    ).select_related('process')
    
    # 建立公司代號到公司名稱的對應
    company_name_to_code = {config.company_name: config.company_code for config in CompanyConfig.objects.all()}
    
    # 獲取有已核准填報記錄的工單（正確區分公司）
    workorders_with_approved_reports = []
    for fill_work in approved_fill_works:
        # 從填報記錄的公司名稱找到對應的公司代號
        company_code = company_name_to_code.get(fill_work.company_name)
        if company_code:
            # 根據公司代號和工單號碼查找對應的工單
            workorder = WorkOrder.objects.filter(
                company_code=company_code,
                order_number=fill_work.workorder
            ).first()
            if workorder:
                workorders_with_approved_reports.append(workorder)
    
    # 去重並排序
    workorders_with_approved_reports = list(set(workorders_with_approved_reports))
    workorders_with_approved_reports.sort(key=lambda x: x.created_at, reverse=True)
    
    # 獲取公司配置資訊，用於顯示公司名稱
    company_configs = {config.company_code: config.company_name for config in CompanyConfig.objects.all()}
    
    # 為每個工單添加公司名稱
    for workorder in workorders_with_approved_reports:
        workorder.company_name = company_configs.get(workorder.company_code, workorder.company_code or '-')
    
    # 獲取統計數據
    total_active = len(workorders_with_approved_reports)  # 有已核准填報記錄的工單數量
    total_pending = WorkOrder.objects.filter(status='pending').count()
    total_completed = WorkOrder.objects.filter(status='completed').count()
    
    # 已核准填報記錄統計（只計算主管審核後的記錄）
    total_approved_reports = FillWork.objects.filter(approval_status='approved').count()
    
    # 計算有對應工單的已核准填報記錄數量（正確區分公司）
    total_approved_reports_with_workorder = 0
    for fill_work in approved_fill_works:
        company_code = company_name_to_code.get(fill_work.company_name)
        if company_code:
            workorder_exists = WorkOrder.objects.filter(
                company_code=company_code,
                order_number=fill_work.workorder
            ).exists()
            if workorder_exists:
                total_approved_reports_with_workorder += 1
    
    # 計算總合格品數量和總工作時數（只計算已核准的記錄）
    total_good_quantity = FillWork.objects.filter(approval_status='approved').aggregate(
        total=Sum('work_quantity')
    )['total'] or 0
    
    total_work_hours = FillWork.objects.filter(approval_status='approved').aggregate(
        total=Sum('work_hours_calculated')
    )['total'] or 0.0
    
    total_overtime_hours = FillWork.objects.filter(approval_status='approved').aggregate(
        total=Sum('overtime_hours_calculated')
    )['total'] or 0.0
    
    # 確保數值類型一致，轉換為 float
    total_work_hours = float(total_work_hours) if total_work_hours else 0.0
    total_overtime_hours = float(total_overtime_hours) if total_overtime_hours else 0.0
    
    context = {
        'total_active': total_active,
        'total_pending': total_pending,
        'total_completed': total_completed,
        'total_approved_reports': total_approved_reports,
        'total_approved_reports_with_workorder': total_approved_reports_with_workorder,
        'total_good_quantity': total_good_quantity,
        'total_work_hours': total_work_hours + total_overtime_hours,
        'workorders_with_approved_reports': workorders_with_approved_reports,
        'today': today,
    }
    
    return render(request, 'workorder/active_workorders.html', context)

@login_required
def check_workorder_completion(request):
    """
    工單完工檢查視圖
    允許使用者手動檢查和處理工單完工
    """
    if request.method == 'POST':
        workorder_id = request.POST.get('workorder_id')
        action = request.POST.get('action')
        
        if action == 'check_specific':
            # 檢查特定工單
            if workorder_id:
                try:
                    workorder = WorkOrder.objects.get(id=workorder_id)
                    # 完工判斷功能已移除，避免資料汙染
                    # packaging_quantity = WorkOrderCompletionService._get_packaging_quantity(workorder)
                    packaging_quantity = 0
                    
                    if packaging_quantity >= workorder.quantity:
                        # 完工判斷功能已移除，避免資料汙染
                        # success = WorkOrderCompletionService.check_and_complete_workorder(workorder.id)
                        success = False
                        if success:
                            messages.success(request, f'工單 {workorder.order_number} 已成功完工並轉移')
                        else:
                            messages.error(request, f'工單 {workorder.order_number} 完工處理失敗')
                    else:
                        messages.warning(request, f'工單 {workorder.order_number} 尚未達到完工條件 (出貨包裝數量: {packaging_quantity}/{workorder.quantity})')
                        
                except WorkOrder.DoesNotExist:
                    messages.error(request, f'工單 ID {workorder_id} 不存在')
                except Exception as e:
                    messages.error(request, f'檢查工單時發生錯誤: {str(e)}')
            else:
                messages.error(request, '請提供工單ID')
                
        elif action == 'check_all':
            # 檢查所有工單
            try:
                all_workorders = WorkOrder.objects.all()
                completed_count = 0
                error_count = 0
                
                for workorder in all_workorders:
                    try:
                        # 完工判斷功能已移除，避免資料汙染
                        # packaging_quantity = WorkOrderCompletionService._get_packaging_quantity(workorder)
                        packaging_quantity = 0
                        
                        if packaging_quantity >= workorder.quantity:
                            # 完工判斷功能已移除，避免資料汙染
                            # success = WorkOrderCompletionService.check_and_complete_workorder(workorder.id)
                            success = False
                            if success:
                                completed_count += 1
                            else:
                                error_count += 1
                                
                    except Exception as e:
                        error_count += 1
                        logger.error(f"檢查工單 {workorder.order_number} 時發生錯誤: {str(e)}")
                
                if completed_count > 0:
                    messages.success(request, f'成功完工 {completed_count} 個工單')
                if error_count > 0:
                    messages.warning(request, f'處理失敗 {error_count} 個工單')
                if completed_count == 0 and error_count == 0:
                    messages.info(request, '沒有工單達到完工條件')
                    
            except Exception as e:
                messages.error(request, f'批量檢查時發生錯誤: {str(e)}')
    
    # 獲取生產中工單列表（只顯示真正在生產中的工單）
    from django.core.paginator import Paginator
    
    production_workorders = WorkOrder.objects.filter(
        status='in_progress'
    ).select_related(
        'production_record'
    ).prefetch_related(
        'processes'
    ).order_by('-created_at')
    
    # 為每個工單計算出貨包裝數量（功能已移除）
    for workorder in production_workorders:
        # workorder.packaging_quantity = WorkOrderCompletionService._get_packaging_quantity(workorder)
        workorder.packaging_quantity = 0
        workorder.completion_rate = (workorder.packaging_quantity / workorder.quantity * 100) if workorder.quantity > 0 else 0
    
    # 分頁處理
    paginator = Paginator(production_workorders, 20)  # 每頁顯示20筆
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'production_workorders': page_obj,
        'page_obj': page_obj,
        'page_title': '工單完工檢查',
        'total_workorders': production_workorders.count(),
    }
    
    return render(request, 'workorder/completion_check.html', context)

@login_required
@require_POST
def force_complete_workorder(request, pk):
    """
    強制完工工單
    將工單標記為完工並轉移到已完工工單模組
    只有管理員可以使用此功能
    """
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({
            "success": False,
            "message": "只有管理員可以使用強制完工功能"
        }, status=403)
    
    try:
        workorder = WorkOrder.objects.get(pk=pk)
        
        # 檢查工單是否已經完工
        if workorder.status == 'completed':
            return JsonResponse({
                "success": False,
                "message": "此工單已經完工"
            }, status=400)
        
        # 記錄操作前的狀態
        original_status = workorder.status
        original_completed_quantity = workorder.completed_quantity
        
        # 更新工單狀態為完工
        workorder.status = 'completed'
        workorder.completed_at = timezone.now()
        
        # 如果完成數量為0，設定為計劃數量
        if workorder.completed_quantity == 0:
            workorder.completed_quantity = workorder.quantity
        
        workorder.save()
        
        # 轉移到已完工工單模組
        try:
            from workorder.services import WorkOrderCompletionService
            completed_workorder = WorkOrderCompletionService.transfer_workorder_to_completed(workorder.id)
            
            # 記錄操作日誌
            from system.models import OperationLog
            OperationLog.objects.create(
                user=request.user.username,
                module="workorder",
                action=f"強制完工工單 {workorder.order_number}（原狀態：{original_status}，原完成數量：{original_completed_quantity}）",
                timestamp=timezone.now(),
                ip_address=request.META.get('REMOTE_ADDR', ''),
            )
            
            return JsonResponse({
                "success": True,
                "message": f"工單 {workorder.order_number} 強制完工成功，已轉移到已完工工單模組",
                "workorder_id": workorder.id,
                "completed_workorder_id": completed_workorder.id if completed_workorder else None
            })
            
        except Exception as transfer_error:
            # 如果轉移失敗，回滾工單狀態
            workorder.status = original_status
            workorder.completed_quantity = original_completed_quantity
            workorder.completed_at = None
            workorder.save()
            
            logger.error(f"強制完工工單 {workorder.order_number} 轉移失敗：{str(transfer_error)}")
            
            return JsonResponse({
                "success": False,
                "message": f"工單狀態已更新為完工，但轉移到已完工模組失敗：{str(transfer_error)}"
            }, status=500)
            
    except WorkOrder.DoesNotExist:
        return JsonResponse({
            "success": False,
            "message": "找不到指定的工單"
        }, status=404)
        
    except Exception as e:
        logger.error(f"強制完工工單失敗：{str(e)}")
        return JsonResponse({
            "success": False,
            "message": f"強制完工失敗：{str(e)}"
        }, status=500)

@login_required
@require_POST
def auto_complete_workorder(request, pk):
    """
    自動完工工單
    檢查完工條件並自動完工（如果達到條件）
    只有管理員可以使用此功能
    """
    if not (request.user.is_staff or request.user.is_superuser):
        return JsonResponse({
            "success": False,
            "message": "只有管理員可以使用自動完工功能"
        }, status=403)
    
    try:
        from workorder.services.completion_service import WorkOrderCompletionService
        
        # 執行自動完工
        result = WorkOrderCompletionService.auto_complete_workorder(pk)
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"自動完工工單 {pk} 失敗: {str(e)}")
        return JsonResponse({
            "success": False,
            "message": f"自動完工失敗: {str(e)}"
        }, status=500)
