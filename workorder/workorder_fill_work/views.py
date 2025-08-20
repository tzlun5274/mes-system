"""
填報作業管理子模組 - 視圖定義
負責填報作業的網頁視圖和表單處理
"""

from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, DetailView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.http import JsonResponse
from django.db.models import Q
from django.urls import reverse_lazy
from django.forms import ModelForm
from django import forms
from datetime import datetime, time
from decimal import Decimal
from django.conf import settings
from mes_config.date_utils import get_today_string, convert_date_for_html_input, normalize_date_string
from mes_config.custom_fields import smart_time_field
from django.views.decorators.http import require_POST
from django.utils.decorators import method_decorator
from django.contrib.auth.decorators import login_required
import csv
from io import StringIO
from django.http import HttpResponse
from io import StringIO, BytesIO
from django.utils import timezone
try:
    import openpyxl
except Exception:
    openpyxl = None

from .models import FillWork
from workorder.workorder_dispatch.models import WorkOrderDispatch
from erp_integration.models import CompanyConfig
from process.models import ProcessName, Operator
from equip.models import Equipment
from workorder.models import WorkOrder
from workorder.services.production_sync_service import ProductionReportSyncService
from .services import MultiFilterService


# ==================== 表單類別定義 ====================

class OperatorBackfillForm(ModelForm):
    """作業員補登填報表單"""
    
    # 產品編號 - 下拉式選單（選項由前端載入；後端不做 choices 驗證）
    product_id = forms.CharField(
        label="產品編號",
        widget=forms.Select(attrs={'class': 'form-select', 'placeholder': '請選擇產品編號'}),
        required=True,
        help_text="選擇產品編號會自動更新相關的工單號碼選項",
    )
    
    # 工單號碼 - 下拉式選單（選項由前端載入）
    workorder = forms.ChoiceField(
        choices=[],
        label="工單號碼",
        widget=forms.Select(attrs={'class': 'form-select', 'placeholder': '請選擇工單號碼'}),
        required=True,
        help_text="選擇工單號碼會自動填入公司名稱和預設生產數量",
    )
    
    # 下拉式日期（最近 30 天）
    work_date = forms.ChoiceField(
        choices=[],
        label="填報日期",
        widget=forms.Select(attrs={'class': 'form-select'}),
        required=True,
        help_text="請選擇日期（預設今天）"
    )
    
    # 公司名稱欄位 - 下拉式選單
    company_name = forms.ChoiceField(
        choices=[],
        label="公司名稱",
        widget=forms.Select(attrs={'class': 'form-select', 'placeholder': '請選擇公司名稱'}),
        required=True,
        help_text="選擇公司名稱會更新相關的產品編號選項",
    )
    
    # 作業員欄位 - 下拉式選單
    operator = forms.ChoiceField(
        choices=[],
        label="作業員",
        widget=forms.Select(attrs={'class': 'form-select', 'placeholder': '請選擇作業員'}),
        required=True,
        help_text="請選擇作業員",
    )
    
    # 使用的設備欄位 - 下拉式選單
    equipment = forms.ChoiceField(
        choices=[],
        label="使用的設備",
        widget=forms.Select(attrs={'class': 'form-select', 'placeholder': '請選擇使用的設備'}),
        required=False,
        help_text="請選擇此次填報使用的設備,選擇後會自動填入作業員欄位",
    )
    
    class Meta:
        model = FillWork
        fields = [
            'product_id', 'workorder', 'company_name', 'operator', 'process', 'equipment',
            'planned_quantity', 'work_quantity', 'defect_quantity', 'is_completed',
            'work_date', 'start_time', 'end_time', 'remarks', 'abnormal_notes'
        ]
        widgets = {
            # product_id 和 workorder 由上方 ChoiceField 控制
            'planned_quantity': forms.NumberInput(attrs={'class': 'form-control', 'readonly': 'readonly'}),
            'work_quantity': forms.NumberInput(attrs={'class': 'form-control', 'min': '0'}),
            'defect_quantity': forms.NumberInput(attrs={'class': 'form-control', 'min': '0'}),
            'is_completed': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'process': forms.Select(attrs={'class': 'form-select'}),
            'start_time': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'HH:MM',
                'pattern': r'\d{2}:\d{2}',
                'maxlength': '5',
                'title': '請輸入24小時制時間格式，例如：08:30、17:30'
            }),
            'end_time': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'HH:MM',
                'pattern': r'\d{2}:\d{2}',
                'maxlength': '5',
                'title': '請輸入24小時制時間格式，例如：08:30、17:30'
            }),
            'remarks': forms.Textarea(attrs={'class': 'form-control', 'rows': '3'}),
            'abnormal_notes': forms.Textarea(attrs={'class': 'form-control', 'rows': '3'}),
        }
    
    def __init__(self, *args, **kwargs):
        self.user = kwargs.pop('user', None)
        super().__init__(*args, **kwargs)
        
        # 使用多重過濾服務
        from .services import MultiFilterService
        
        # 獲取多重過濾後的選項
        filtered_choices = MultiFilterService.get_multi_filtered_choices(
            user=self.user,  # 傳入使用者參數
            form_type='operator',
            permission_type='both'
        )
        
        # 設定過濾後的選項
        self.fields['process'].queryset = MultiFilterService.get_multi_filtered_process_queryset(
            user=self.user, form_type='operator', permission_type='both'
        )
        self.fields['operator'].choices = filtered_choices['operators']
        self.fields['equipment'].choices = filtered_choices['equipments']
        
        # 設定 planned_quantity 可留空
        self.fields['planned_quantity'].required = False
        
        # 載入選項
        self.fields['company_name'].choices = self.get_company_choices()
        self.fields['workorder'].choices = self.get_workorder_choices()
        
        # 設定預設值與日期選項
        if not self.instance.pk:
            from datetime import timedelta, date
            today = date.today()
            options = []
            for i in range(0, 30):
                d = today - timedelta(days=i)
                s = d.strftime('%Y-%m-%d')
                options.append((s, s))
            self.fields['work_date'].choices = options
            self.fields['work_date'].initial = today.strftime('%Y-%m-%d')
    
    def get_company_choices(self):
        choices = [("", "請選擇公司名稱")]
        try:
            from erp_integration.models import CompanyConfig
            companies = CompanyConfig.objects.all().order_by('company_name')
            for company in companies:
                choices.append((company.company_name, company.company_name))
        except ImportError:
            pass
        return choices
    
    def get_operator_choices(self):
        choices = [("", "請選擇作業員")]
        try:
            from process.models import Operator
            operators = Operator.objects.all().order_by('name')
            for operator in operators:
                choices.append((operator.name, operator.name))
        except ImportError:
            pass
        return choices
    
    def get_equipment_choices(self):
        choices = [("", "請選擇使用的設備")]
        try:
            from equip.models import Equipment
            equipments = Equipment.objects.exclude(name__icontains='SMT').order_by('name')
            for equipment in equipments:
                choices.append((equipment.name, equipment.name))
        except ImportError:
            pass
        return choices
    
    def get_workorder_choices(self):
        """獲取工單號碼選項"""
        choices = [("", "請選擇工單號碼")]
        try:
            from workorder.workorder_dispatch.models import WorkOrderDispatch
            workorders = WorkOrderDispatch.objects.all().order_by('order_number')
            for workorder in workorders:
                choices.append((workorder.order_number, workorder.order_number))
        except ImportError:
            pass
        return choices
    
    def clean_start_time(self):
        """驗證開始時間格式"""
        start_time = self.cleaned_data.get('start_time')
        if start_time:
            try:
                # 驗證 HH:MM 格式
                from datetime import datetime, time as time_cls
                if isinstance(start_time, time_cls):
                    return start_time
                time_obj = datetime.strptime(str(start_time), "%H:%M").time()
                return time_obj
            except (ValueError, TypeError):
                raise forms.ValidationError("請輸入正確的時間格式：HH:MM（24小時制）")
        return start_time
    
    def clean_end_time(self):
        """驗證結束時間格式"""
        end_time = self.cleaned_data.get('end_time')
        if not end_time:
            return end_time
        from datetime import datetime, time as time_cls
        if isinstance(end_time, time_cls):
            return end_time
        try:
            return datetime.strptime(str(end_time), "%H:%M").time()
        except (ValueError, TypeError):
            raise forms.ValidationError("請輸入正確的時間格式：HH:MM（24小時制）")


class OperatorRDBackfillForm(ModelForm):
    """作業員RD樣品補登填報表單 - 使用智慧日期欄位，支援任何格式輸入"""
    
    # 下拉式日期（最近 30 天）
    work_date = forms.ChoiceField(
        choices=[],
        label="填報日期",
        widget=forms.Select(attrs={'class': 'form-select'}),
        required=True,
        help_text="請選擇日期（預設今天）"
    )
    
    # 公司名稱欄位 - 下拉式選單
    company_name = forms.ChoiceField(
        choices=[],
        label="公司名稱",
        widget=forms.Select(attrs={
            'class': 'form-select',
            'placeholder': '請選擇公司名稱'
        }),
        required=True,
        help_text="請選擇公司名稱",
    )
    
    # 作業員欄位 - 下拉式選單
    operator = forms.ChoiceField(
        choices=[],
        label="作業員",
        widget=forms.Select(attrs={
            'class': 'form-select',
            'placeholder': '請選擇作業員'
        }),
        required=True,
        help_text="請選擇作業員",
    )
    
    # 使用的設備欄位 - 下拉式選單
    equipment = forms.ChoiceField(
        choices=[],
        label="使用的設備",
        widget=forms.Select(attrs={
            'class': 'form-select',
            'placeholder': '請選擇使用的設備'
        }),
        required=False,
        help_text="請選擇使用的設備",
    )
    
    class Meta:
        model = FillWork
        fields = [
            'product_id', 'workorder', 'company_name', 'operator', 'process', 'equipment',
            'planned_quantity', 'work_quantity', 'defect_quantity', 'is_completed',
            'work_date', 'start_time', 'end_time', 'remarks', 'abnormal_notes'
        ]
        widgets = {
            'product_id': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': '請輸入產品編號',
                'title': '請輸入RD樣品的產品編號'
            }),
            'workorder': forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'}),
            'planned_quantity': forms.NumberInput(attrs={'class': 'form-control', 'readonly': 'readonly'}),
            'work_quantity': forms.NumberInput(attrs={'class': 'form-control', 'min': '0'}),
            'defect_quantity': forms.NumberInput(attrs={'class': 'form-control', 'min': '0'}),
            'is_completed': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'process': forms.Select(attrs={'class': 'form-select'}),
            'start_time': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'HH:MM',
                'pattern': r'\d{2}:\d{2}',
                'maxlength': '5',
                'title': '請輸入24小時制時間格式，例如：08:30、17:30'
            }),
            'end_time': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'HH:MM',
                'pattern': r'\d{2}:\d{2}',
                'maxlength': '5',
                'title': '請輸入24小時制時間格式，例如：08:30、17:30'
            }),
            'remarks': forms.Textarea(attrs={'class': 'form-control', 'rows': '3'}),
            'abnormal_notes': forms.Textarea(attrs={'class': 'form-control', 'rows': '3'}),
        }
    
    def __init__(self, *args, **kwargs):
        self.user = kwargs.pop('user', None)
        super().__init__(*args, **kwargs)
        
        # 使用多重過濾服務
        from .services import MultiFilterService
        
        # 獲取多重過濾後的選項
        filtered_choices = MultiFilterService.get_multi_filtered_choices(
            user=self.user,  # 傳入使用者參數
            form_type='operator',
            permission_type='both'
        )
        
        # 設定過濾後的選項
        self.fields['process'].queryset = MultiFilterService.get_multi_filtered_process_queryset(
            user=self.user, form_type='operator', permission_type='both'
        )
        self.fields['operator'].choices = filtered_choices['operators']
        self.fields['equipment'].choices = filtered_choices['equipments']
        
        # 設定planned_quantity為非必填
        self.fields['planned_quantity'].required = False
        
        # 載入選項
        self.fields['company_name'].choices = self.get_company_choices()
        
        # 設定預設值
        if not self.instance.pk:  # 新增時
            # 構建最近 30 天的日期選項（含今天，往前 29 天）
            from datetime import timedelta, date
            today = date.today()
            options = []
            for i in range(0, 30):
                d = today - timedelta(days=i)
                s = d.strftime('%Y-%m-%d')
                options.append((s, s))
            self.fields['work_date'].choices = options
            self.fields['work_date'].initial = today.strftime('%Y-%m-%d')
            self.fields['product_id'].initial = 'PFP-CCT'  # 預設值，但可以手動修改
            self.fields['workorder'].initial = 'RD樣品'
            self.fields['planned_quantity'].initial = 0
            # 不設定時間初始值，讓前端JavaScript處理
    
    def get_company_choices(self):
        """獲取公司選項"""
        choices = [("", "請選擇公司名稱")]
        
        try:
            from erp_integration.models import CompanyConfig
            companies = CompanyConfig.objects.all().order_by('company_name')
            for company in companies:
                choices.append((company.company_name, company.company_name))
        except ImportError:
            pass
        
        return choices
    
    def get_operator_choices(self):
        choices = [("", "請選擇作業員")]
        try:
            # 排除SMT相關的作業員，因為SMT是設備/產線，不是人員
            operators = Operator.objects.exclude(name__icontains='SMT').order_by('name')
            for o in operators:
                choices.append((o.name, o.name))
        except Exception:
            pass
        return choices
    
    def get_equipment_choices(self):
        """獲取設備選項（排除SMT相關設備）"""
        choices = [("", "請選擇使用的設備")]
        
        try:
            from equip.models import Equipment
            equipments = Equipment.objects.exclude(name__icontains='SMT').order_by('name')
            for equipment in equipments:
                choices.append((equipment.name, equipment.name))
        except ImportError:
            pass
        
        return choices
    
    def clean_start_time(self):
        """驗證開始時間格式"""
        start_time = self.cleaned_data.get('start_time')
        if not start_time:
            return start_time
        from datetime import datetime, time as time_cls
        if isinstance(start_time, time_cls):
            return start_time
        try:
            return datetime.strptime(str(start_time), "%H:%M").time()
        except (ValueError, TypeError):
            raise forms.ValidationError("請輸入正確的時間格式：HH:MM（24小時制）")
    
    def clean_end_time(self):
        """驗證結束時間格式"""
        end_time = self.cleaned_data.get('end_time')
        if end_time:
            try:
                # 驗證 HH:MM 格式
                from datetime import datetime, time as time_cls
                if isinstance(end_time, time_cls):
                    return end_time
                time_obj = datetime.strptime(str(end_time), "%H:%M").time()
                return time_obj
            except (ValueError, TypeError):
                raise forms.ValidationError("請輸入正確的時間格式：HH:MM（24小時制）")
        return end_time
     



class SMTBackfillForm(ModelForm):
    """SMT補登填報表單"""

    # 產品編號 - 下拉式選單（選項由前端載入）
    product_id = forms.CharField(
        label="產品編號",
        widget=forms.Select(attrs={'class': 'form-select', 'placeholder': '請選擇產品編號'}),
        required=True,
        help_text="選擇產品編號會自動更新相關的工單號碼選項",
    )
    
    # 工單號碼 - 下拉式選單（選項由前端載入）
    workorder = forms.ChoiceField(
        choices=[],
        label="工單號碼",
        widget=forms.Select(attrs={'class': 'form-select', 'placeholder': '請選擇工單號碼'}),
        required=True,
        help_text="選擇工單號碼會自動填入公司名稱和預設生產數量",
    )
    
    # 下拉式日期（最近 30 天）
    work_date = forms.ChoiceField(
        choices=[],
        label="填報日期",
        widget=forms.Select(attrs={'class': 'form-select'}),
        required=True,
        help_text="請選擇日期（預設今天）"
    )

    # 公司名稱、作業員、設備（僅顯示SMT相關設備）
    company_name = forms.ChoiceField(
        choices=[], label="公司名稱",
        widget=forms.Select(attrs={'class': 'form-select', 'placeholder': '請選擇公司名稱'}), required=True,
        help_text="選擇公司名稱會更新相關的產品編號選項"
    )
    operator = forms.ChoiceField(
        choices=[], label="作業員",
        widget=forms.Select(attrs={'class': 'form-select', 'placeholder': '請選擇作業員'}), required=True,
        help_text="請選擇作業員"
    )
    equipment = forms.ChoiceField(
        choices=[], label="使用的設備",
        widget=forms.Select(attrs={'class': 'form-select', 'placeholder': '請選擇使用的設備'}), required=True,
        help_text="請選擇此次填報使用的設備,選擇後會自動填入作業員欄位"
    )
    
    class Meta:
        model = FillWork
        fields = [
            'product_id', 'workorder', 'company_name', 'operator', 'process', 'equipment',
            'planned_quantity', 'work_quantity', 'defect_quantity', 'is_completed',
            'work_date', 'start_time', 'end_time', 'remarks', 'abnormal_notes'
        ]
        widgets = {
            'product_id': forms.TextInput(attrs={'class': 'form-control'}),
            'workorder': forms.TextInput(attrs={'class': 'form-control'}),
            'planned_quantity': forms.NumberInput(attrs={'class': 'form-control', 'readonly': 'readonly'}),
            'work_quantity': forms.NumberInput(attrs={'class': 'form-control', 'min': '0'}),
            'defect_quantity': forms.NumberInput(attrs={'class': 'form-control', 'min': '0'}),
            'is_completed': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'process': forms.Select(attrs={'class': 'form-select'}),
            'start_time': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'HH:MM',
                'pattern': r'\d{2}:\d{2}',
                'maxlength': '5',
                'title': '請輸入24小時制時間格式，例如：08:30、17:30'
            }),
            'end_time': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'HH:MM',
                'pattern': r'\d{2}:\d{2}',
                'maxlength': '5',
                'title': '請輸入24小時制時間格式，例如：08:30、17:30'
            }),
            'remarks': forms.Textarea(attrs={'class': 'form-control', 'rows': '3'}),
            'abnormal_notes': forms.Textarea(attrs={'class': 'form-control', 'rows': '3'}),
        }
    
    def __init__(self, *args, **kwargs):
        self.user = kwargs.pop('user', None)
        super().__init__(*args, **kwargs)
        
        # 使用多重過濾服務
        from .services import MultiFilterService
        
        # 獲取多重過濾後的選項
        filtered_choices = MultiFilterService.get_multi_filtered_choices(
            user=self.user,  # 傳入使用者參數
            form_type='smt',
            permission_type='both'
        )
        
        # 設定過濾後的選項
        self.fields['process'].queryset = MultiFilterService.get_multi_filtered_process_queryset(
            user=self.user, form_type='smt', permission_type='both'
        )
        self.fields['operator'].choices = filtered_choices['operators']
        self.fields['equipment'].choices = filtered_choices['equipments']
        
        # planned_quantity 非必填
        self.fields['planned_quantity'].required = False
        
        # 載入選項
        self.fields['company_name'].choices = self.get_company_choices()
        
        # 最近30天日期選項
        from datetime import timedelta, date
        today = date.today()
        options = []
        for i in range(0, 30):
            d = today - timedelta(days=i)
            s = d.strftime('%Y-%m-%d')
            options.append((s, s))
        self.fields['work_date'].choices = options
        if not self.instance.pk:
            self.fields['work_date'].initial = today.strftime('%Y-%m-%d')

    def get_company_choices(self):
        choices = [("", "請選擇公司名稱")]
        try:
            companies = CompanyConfig.objects.all().order_by('company_name')
            for c in companies:
                choices.append((c.company_name, c.company_name))
        except Exception:
            pass
        return choices

    def get_operator_choices(self):
        choices = [("", "請選擇作業員")]
        try:
            # 排除SMT相關的作業員，因為SMT是設備/產線，不是人員
            operators = Operator.objects.exclude(name__icontains='SMT').order_by('name')
            for o in operators:
                choices.append((o.name, o.name))
        except Exception:
            pass
        return choices

    def get_equipment_choices(self):
        choices = [("", "請選擇使用的設備")]
        try:
            equipments = Equipment.objects.filter(name__icontains='SMT').order_by('name')
            for e in equipments:
                choices.append((e.name, e.name))
        except Exception:
            pass
        return choices
    
    def clean_start_time(self):
        """驗證開始時間格式"""
        start_time = self.cleaned_data.get('start_time')
        if not start_time:
            return start_time
        from datetime import datetime, time as time_cls
        if isinstance(start_time, time_cls):
            return start_time
        try:
            return datetime.strptime(str(start_time), "%H:%M").time()
        except (ValueError, TypeError):
            raise forms.ValidationError("請輸入正確的時間格式：HH:MM（24小時制）")
    
    def clean_end_time(self):
        """驗證結束時間格式"""
        end_time = self.cleaned_data.get('end_time')
        if not end_time:
            return end_time
        from datetime import datetime, time as time_cls
        if isinstance(end_time, time_cls):
            return end_time
        try:
            return datetime.strptime(str(end_time), "%H:%M").time()
        except (ValueError, TypeError):
            raise forms.ValidationError("請輸入正確的時間格式：HH:MM（24小時制）")


class SMTRDBackfillForm(ModelForm):
    """SMT_RD樣品補登填報表單"""

    # 下拉式日期（最近 30 天）
    work_date = forms.ChoiceField(
        choices=[],
        label="填報日期",
        widget=forms.Select(attrs={'class': 'form-select'}),
        required=True,
        help_text="請選擇日期（預設今天）"
    )

    # 公司名稱、作業員、設備（僅顯示SMT相關設備）
    company_name = forms.ChoiceField(
        choices=[], label="公司名稱",
        widget=forms.Select(attrs={'class': 'form-select'}), required=True
    )
    operator = forms.CharField(
        label="作業員",
        widget=forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'}),
        required=False
    )
    equipment = forms.ChoiceField(
        choices=[], label="使用的設備",
        widget=forms.Select(attrs={'class': 'form-select'}), required=True
    )
    
    class Meta:
        model = FillWork
        fields = [
            'product_id', 'workorder', 'company_name', 'operator', 'process', 'equipment',
            'planned_quantity', 'work_quantity', 'defect_quantity', 'is_completed',
            'work_date', 'start_time', 'end_time', 'remarks', 'abnormal_notes'
        ]
        widgets = {
            'product_id': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': '請輸入產品編號',
                'title': '請輸入SMT RD樣品的產品編號'
            }),
            'workorder': forms.TextInput(attrs={'class': 'form-control', 'readonly': 'readonly'}),
            'planned_quantity': forms.NumberInput(attrs={'class': 'form-control', 'readonly': 'readonly'}),
            'work_quantity': forms.NumberInput(attrs={'class': 'form-control', 'min': '0'}),
            'defect_quantity': forms.NumberInput(attrs={'class': 'form-control', 'min': '0'}),
            'is_completed': forms.CheckboxInput(attrs={'class': 'form-check-input'}),
            'process': forms.Select(attrs={'class': 'form-select'}),
            'start_time': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'HH:MM',
                'pattern': r'\d{2}:\d{2}',
                'maxlength': '5',
                'title': '請輸入24小時制時間格式，例如：08:30、17:30'
            }),
            'end_time': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'HH:MM',
                'pattern': r'\d{2}:\d{2}',
                'maxlength': '5',
                'title': '請輸入24小時制時間格式，例如：08:30、17:30'
            }),
            'remarks': forms.Textarea(attrs={'class': 'form-control', 'rows': '3'}),
            'abnormal_notes': forms.Textarea(attrs={'class': 'form-control', 'rows': '3'}),
        }
    
    def __init__(self, *args, **kwargs):
        self.user = kwargs.pop('user', None)
        super().__init__(*args, **kwargs)
        
        # 使用多重過濾服務
        from .services import MultiFilterService
        
        # 獲取多重過濾後的選項
        filtered_choices = MultiFilterService.get_multi_filtered_choices(
            user=self.user,  # 傳入使用者參數
            form_type='smt',
            permission_type='both'
        )
        
        # 設定過濾後的選項
        self.fields['process'].queryset = MultiFilterService.get_multi_filtered_process_queryset(
            user=self.user, form_type='smt', permission_type='both'
        )
        self.fields['operator'].choices = filtered_choices['operators']
        self.fields['equipment'].choices = filtered_choices['equipments']
        
        # planned_quantity 非必填
        self.fields['planned_quantity'].required = False
        
        # 載入選項
        self.fields['company_name'].choices = self.get_company_choices()

        # 設定預設值
        if not self.instance.pk:  # 新增時
            from datetime import date, timedelta
            today = date.today()
            # 最近30天日期選項
            options = []
            for i in range(0, 30):
                d = today - timedelta(days=i)
                s = d.strftime('%Y-%m-%d')
                options.append((s, s))
            self.fields['work_date'].choices = options
            self.fields['work_date'].initial = today.strftime('%Y-%m-%d')
            self.fields['product_id'].initial = 'PFP-CCT'
            self.fields['workorder'].initial = 'RD樣品'
            self.fields['planned_quantity'].initial = 0
            # 不設定時間初始值，讓前端JavaScript處理

    def get_company_choices(self):
        choices = [("", "請選擇公司名稱")]
        try:
            companies = CompanyConfig.objects.all().order_by('company_name')
            for c in companies:
                choices.append((c.company_name, c.company_name))
        except Exception:
            pass
        return choices

    def get_operator_choices(self):
        choices = [("", "請選擇作業員")]
        try:
            operators = Operator.objects.all().order_by('name')
            for o in operators:
                choices.append((o.name, o.name))
        except Exception:
            pass
        return choices

    def get_equipment_choices(self):
        choices = [("", "請選擇使用的設備")]
        try:
            equipments = Equipment.objects.filter(name__icontains='SMT').order_by('name')
            for e in equipments:
                choices.append((e.name, e.name))
        except Exception:
            pass
        return choices
    
    def clean_start_time(self):
        """驗證開始時間格式"""
        start_time = self.cleaned_data.get('start_time')
        if not start_time:
            return start_time
        from datetime import datetime, time as time_cls
        if isinstance(start_time, time_cls):
            return start_time
        try:
            return datetime.strptime(str(start_time), "%H:%M").time()
        except (ValueError, TypeError):
            raise forms.ValidationError("請輸入正確的時間格式：HH:MM（24小時制）")
    
    def clean_end_time(self):
        """驗證結束時間格式"""
        end_time = self.cleaned_data.get('end_time')
        if not end_time:
            return end_time
        from datetime import datetime, time as time_cls
        if isinstance(end_time, time_cls):
            return end_time
        try:
            return datetime.strptime(str(end_time), "%H:%M").time()
        except (ValueError, TypeError):
            raise forms.ValidationError("請輸入正確的時間格式：HH:MM（24小時制）")


# ==================== 視圖類別定義 ====================

class FillWorkIndexView(LoginRequiredMixin, ListView):
    """填報管理首頁"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/index.html'
    context_object_name = 'fill_works'
    paginate_by = 20
    
    def get_queryset(self):
        return FillWork.objects.all().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        """提供首頁統計卡用的統計數據。"""
        context = super().get_context_data(**kwargs)
        base_qs = FillWork.objects.all()
        context['pending_count'] = base_qs.filter(approval_status='pending').count()
        context['approved_count'] = base_qs.filter(approval_status='approved').count()
        context['rejected_count'] = base_qs.filter(approval_status='rejected').count()
        context['total_count'] = base_qs.count()
        return context


class OperatorIndexView(LoginRequiredMixin, ListView):
    """作業員填報首頁"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/operator_index.html'
    context_object_name = 'fill_works'
    paginate_by = 20
    
    def get_queryset(self):
        return FillWork.objects.exclude(
            Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
        ).order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        """提供統計數據與最近記錄，供首頁儀表板顯示"""
        context = super().get_context_data(**kwargs)
        base_qs = FillWork.objects.exclude(
            Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
        )
        context['operator_pending_count'] = base_qs.filter(approval_status='pending').count()
        context['operator_approved_count'] = base_qs.filter(approval_status='approved').count()
        context['operator_rejected_count'] = base_qs.filter(approval_status='rejected').count()
        context['operator_total_count'] = base_qs.count()
        context['recent_operator_records'] = base_qs.order_by('-work_date', '-start_time', '-created_at')[:10]
        return context


class OperatorBackfillCreateView(LoginRequiredMixin, CreateView):
    """新增作業員補登填報"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/operator_backfill_form.html'
    form_class = OperatorBackfillForm
    success_url = reverse_lazy('workorder:fill_work:operator_index')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['hours_range'] = [f"{h:02d}" for h in range(0, 24)]
        context['minutes_range'] = [f"{m:02d}" for m in range(0, 60)]
        return context
    
    def get_form_kwargs(self):
        """傳遞用戶參數給表單"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        try:
            # 設定休息時間（作業員填報：12:00-13:00）
            form.instance.has_break = True
            form.instance.break_start_time = time(12, 0)
            form.instance.break_end_time = time(13, 0)
            
            # 設定建立者
            form.instance.created_by = self.request.user.username
            
            response = super().form_valid(form)
            messages.success(self.request, '作業員補登填報新增成功！')
            return response
        except Exception as e:
            messages.error(self.request, f'儲存失敗：{str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        # 彙整詳細欄位錯誤
        if form.errors:
            details = []
            for field, errs in form.errors.items():
                for err in errs:
                    details.append(f"{field}: {err}")
            if details:
                messages.error(self.request, "資料有誤：" + "；".join(details))
        else:
            messages.error(self.request, '請檢查表單資料，有必填欄位未填寫或格式不正確')
        return self.render_to_response(self.get_context_data(form=form))


class OperatorRDBackfillCreateView(LoginRequiredMixin, CreateView):
    """新增作業員RD樣品補登填報"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/operator_rd_backfill_form.html'
    form_class = OperatorRDBackfillForm
    success_url = reverse_lazy('workorder:fill_work:operator_index')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # 後端提供正確的時間選單：00-23 與 00-59
        context['hours_range'] = [f"{h:02d}" for h in range(0, 24)]
        context['minutes_range'] = [f"{m:02d}" for m in range(0, 60)]
        return context
    
    def get_form_kwargs(self):
        """傳遞用戶參數給表單"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        try:
            from workorder.models import WorkOrder
            from workorder.workorder_dispatch.models import WorkOrderDispatch
            from erp_integration.models import CompanyConfig
            from django.utils import timezone
            
            # 取得表單資料
            company_name = form.cleaned_data.get('company_name')
            product_code = form.cleaned_data.get('product_id')
            
            # 取得公司代號
            company_config = CompanyConfig.objects.filter(company_name=company_name).first()
            if not company_config:
                messages.error(self.request, '找不到對應的公司設定')
                return self.form_invalid(form)
            
            company_code_value = company_config.company_code
            
            # 查找現有工單（只根據公司代號和工單號碼，因為唯一性約束是 (company_code, order_number)）
            existing_workorder = WorkOrder.objects.filter(
                company_code=company_code_value,
                order_number='RD樣品'
            ).first()
            
            workorder = None
            workorder_created = False
            
            if existing_workorder:
                workorder = existing_workorder
                messages.info(self.request, f'找到現有RD樣品工單: {workorder.order_number}')
            else:
                # 建立新工單
                workorder = WorkOrder.objects.create(
                    company_code=company_code_value,
                    order_number='RD樣品',
                    product_code=product_code,
                    quantity=0,
                    status='in_progress',
                    order_source='MES手動建立'
                )
                workorder_created = True
                messages.info(self.request, f'建立新RD樣品工單: {workorder.order_number}')
            
            # 檢查並建立派工單（比對公司代號+工單號碼+產品編號+工序）
            existing_dispatch = WorkOrderDispatch.objects.filter(
                company_code=company_code_value,
                order_number='RD樣品',
                product_code=product_code,
                process_name=process_name
            ).first()
            
            dispatch_created = False
            
            if existing_dispatch:
                messages.info(self.request, f'找到現有RD樣品派工單: {existing_dispatch.order_number}')
            else:
                # 建立新派工單
                new_dispatch = WorkOrderDispatch.objects.create(
                    company_code=company_code_value,
                    order_number='RD樣品',
                    product_code=product_code,
                    product_name=f'RD樣品-{product_code}',
                    planned_quantity=0,
                    status='in_production',
                    dispatch_date=timezone.now().date(),
                    assigned_operator=form.cleaned_data.get('operator'),
                    assigned_equipment=form.cleaned_data.get('equipment') or '',
                    process_name=process_name,
                    notes=f"作業員RD樣品補登填報自動建立 - 建立時間: {timezone.now()} - 注意：RD樣品無預定工序流程",
                    created_by=self.request.user.username
                )
                dispatch_created = True
                messages.info(self.request, f'建立新RD樣品派工單: {new_dispatch.order_number}')
            
            # 設定休息時間（作業員RD填報：12:00-13:00）
            form.instance.has_break = True
            form.instance.break_start_time = time(12, 0)
            form.instance.break_end_time = time(13, 0)
            
            # 設定建立者
            form.instance.created_by = self.request.user.username
            
            # 儲存表單
            response = super().form_valid(form)
            
            # 顯示建立結果
            success_message = '作業員RD樣品補登填報新增成功！'
            if workorder_created:
                success_message += f' - 已建立新工單: {workorder.order_number}'
            if dispatch_created:
                success_message += f' - 已建立新派工單'
            
            messages.success(self.request, success_message)
            
            return response
            
        except Exception as e:
            # 錯誤訊息
            messages.error(self.request, f'儲存失敗：{str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        """表單驗證失敗時的處理"""
        # 顯示錯誤訊息
        messages.error(self.request, '請檢查表單資料，有必填欄位未填寫或格式不正確')
        
        # 返回表單頁面，讓使用者繼續編輯
        return self.render_to_response(self.get_context_data(form=form))


class SMTIndexView(LoginRequiredMixin, ListView):
    """SMT填報首頁
    提供 SMT 相關的統計數字與最近填報記錄，供首頁儀表板顯示。
    """
    model = FillWork
    template_name = 'workorder/workorder_fill_work/smt_index.html'
    context_object_name = 'fill_works'
    paginate_by = 20
    
    def get_queryset(self):
        return FillWork.objects.filter(
            Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
        ).order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        """組裝 SMT 首頁需要的統計資訊與最近記錄清單。
        - smt_pending_count / smt_approved_count / smt_rejected_count / smt_total_count
        - recent_smt_records：最近 10 筆 SMT 填報
        """
        context = super().get_context_data(**kwargs)
        base_qs = FillWork.objects.filter(
            Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
        )
        context['smt_pending_count'] = base_qs.filter(approval_status='pending').count()
        context['smt_approved_count'] = base_qs.filter(approval_status='approved').count()
        context['smt_rejected_count'] = base_qs.filter(approval_status='rejected').count()
        context['smt_total_count'] = base_qs.count()
        context['recent_smt_records'] = base_qs.order_by('-created_at')[:10]
        return context


class SMTBackfillCreateView(LoginRequiredMixin, CreateView):
    """新增SMT補登填報"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/smt_backfill_form.html'
    form_class = SMTBackfillForm
    success_url = reverse_lazy('workorder:fill_work:smt_index')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['hours_range'] = [f"{h:02d}" for h in range(0, 24)]
        context['minutes_range'] = [f"{m:02d}" for m in range(0, 60)]
        return context

    def get_form_kwargs(self):
        """傳遞用戶參數給表單"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs

    def form_valid(self, form):
        try:
            # 設定無休息時間（SMT填報：12:00-12:00）
            form.instance.has_break = True
            form.instance.break_start_time = time(12, 0)
            form.instance.break_end_time = time(12, 0)
            
            # 設定建立者
            form.instance.created_by = self.request.user.username
            
            response = super().form_valid(form)
            messages.success(self.request, 'SMT補登填報新增成功！')
            return response
        except Exception as e:
            messages.error(self.request, f'儲存失敗：{str(e)}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        # 彙整詳細欄位錯誤
        if form.errors:
            details = []
            for field, errs in form.errors.items():
                for err in errs:
                    details.append(f"{field}: {err}")
            if details:
                messages.error(self.request, "資料有誤：" + "；".join(details))
        else:
            messages.error(self.request, '請檢查表單資料，有必填欄位未填寫或格式不正確')
        return self.render_to_response(self.get_context_data(form=form))


class SMTRDBackfillCreateView(LoginRequiredMixin, CreateView):
    """新增SMT_RD樣品補登填報"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/smt_rd_backfill_form.html'
    form_class = SMTRDBackfillForm
    success_url = reverse_lazy('workorder:fill_work:smt_index')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['hours_range'] = [f"{h:02d}" for h in range(0, 24)]
        context['minutes_range'] = [f"{m:02d}" for m in range(0, 60)]
        return context
    
    def get_form_kwargs(self):
        """傳遞用戶參數給表單"""
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        try:
            from workorder.models import WorkOrder
            from workorder.workorder_dispatch.models import WorkOrderDispatch
            from erp_integration.models import CompanyConfig
            from django.utils import timezone
            
            # 取得表單資料
            company_name = form.cleaned_data.get('company_name')
            product_code = form.cleaned_data.get('product_id')
            
            # 取得公司代號
            company_config = CompanyConfig.objects.filter(company_name=company_name).first()
            if not company_config:
                messages.error(self.request, '找不到對應的公司設定')
                return self.form_invalid(form)
            
            company_code_value = company_config.company_code
            
            # 查找現有工單（只根據公司代號和工單號碼，因為唯一性約束是 (company_code, order_number)）
            existing_workorder = WorkOrder.objects.filter(
                company_code=company_code_value,
                order_number='RD樣品'
            ).first()
            
            workorder = None
            workorder_created = False
            
            if existing_workorder:
                workorder = existing_workorder
                messages.info(self.request, f'找到現有RD樣品工單: {workorder.order_number}')
            else:
                # 建立新工單
                workorder = WorkOrder.objects.create(
                    company_code=company_code_value,
                    order_number='RD樣品',
                    product_code=product_code,
                    quantity=0,
                    status='in_progress',
                    order_source='MES手動建立'
                )
                workorder_created = True
                messages.info(self.request, f'建立新RD樣品工單: {workorder.order_number}')
            
            # 檢查並建立派工單（比對公司代號+工單號碼+產品編號+工序）
            process_name = form.cleaned_data.get('process').name if form.cleaned_data.get('process') else ''
            existing_dispatch = WorkOrderDispatch.objects.filter(
                company_code=company_code_value,
                order_number='RD樣品',
                product_code=product_code,
                process_name=process_name
            ).first()
            
            dispatch_created = False
            
            if existing_dispatch:
                messages.info(self.request, f'找到現有RD樣品派工單: {existing_dispatch.order_number}')
            else:
                # 建立新派工單
                new_dispatch = WorkOrderDispatch.objects.create(
                    company_code=company_code_value,
                    order_number='RD樣品',
                    product_code=product_code,
                    product_name=f'RD樣品-{product_code}',
                    planned_quantity=0,
                    status='in_production',
                    dispatch_date=timezone.now().date(),
                    assigned_operator=form.cleaned_data.get('operator') or '',
                    assigned_equipment=form.cleaned_data.get('equipment') or '',
                    process_name=process_name,
                    notes=f"SMT_RD樣品補登填報自動建立 - 建立時間: {timezone.now()} - 注意：RD樣品無預定工序流程",
                    created_by=self.request.user.username
                )
                dispatch_created = True
                messages.info(self.request, f'建立新RD樣品派工單: {new_dispatch.order_number}')
            
            # 設定無休息時間（SMT RD填報：12:00-12:00）
            form.instance.has_break = True
            form.instance.break_start_time = time(12, 0)
            form.instance.break_end_time = time(12, 0)
            
            # 設定建立者
            form.instance.created_by = self.request.user.username
            
            response = super().form_valid(form)
            
            # 顯示建立結果
            success_message = 'SMT_RD樣品補登填報新增成功！'
            if workorder_created:
                success_message += f' - 已建立新工單: {workorder.order_number}'
            if dispatch_created:
                success_message += f' - 已建立新派工單'
            
            messages.success(self.request, success_message)
            return response
        except Exception as e:
            messages.error(self.request, f'儲存失敗：{str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        if form.errors:
            details = []
            for field, errs in form.errors.items():
                for err in errs:
                    details.append(f"{field}: {err}")
            if details:
                messages.error(self.request, "資料有誤：" + "；".join(details))
        else:
            messages.error(self.request, '請檢查表單資料，有必填欄位未填寫或格式不正確')
        return self.render_to_response(self.get_context_data(form=form))


class SMTBackfillUpdateView(LoginRequiredMixin, UpdateView):
    """編輯 SMT 補登/SMT_RD樣品 補登填報（僅允許未審核）"""
    model = FillWork
    # 預設使用 SMT 補登模板，RD 會在 get_template_names 動態切換
    template_name = 'workorder/workorder_fill_work/smt_backfill_form.html'
    success_url = reverse_lazy('workorder:fill_work:smt_index')

    def dispatch(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.approval_status != 'pending':
            messages.error(request, '此筆記錄已審核，無法修改')
            return redirect('workorder:fill_work:smt_index')
        return super().dispatch(request, *args, **kwargs)

    def get_form_class(self):
        obj = self.get_object()
        # 依工單號碼是否為 RD樣品 決定使用的表單
        if str(obj.workorder).strip() == 'RD樣品':
            return SMTRDBackfillForm
        return SMTBackfillForm

    def get_template_names(self):
        obj = self.get_object()
        if str(obj.workorder).strip() == 'RD樣品':
            return ['workorder/fill_work/smt_rd_backfill_form.html']
        return ['workorder/fill_work/smt_backfill_form.html']

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['hours_range'] = [f"{h:02d}" for h in range(0, 24)]
        context['minutes_range'] = [f"{m:02d}" for m in range(0, 60)]
        return context

    def form_valid(self, form):
        try:
            form.instance.created_by = self.request.user.username or form.instance.created_by
            response = super().form_valid(form)
            # 依 RD 與否顯示不同成功訊息
            if str(self.object.workorder).strip() == 'RD樣品':
                messages.success(self.request, 'SMT_RD樣品補登填報修改成功！')
            else:
                messages.success(self.request, 'SMT補登填報修改成功！')
            return response
        except Exception as e:
            messages.error(self.request, f'儲存失敗：{str(e)}')
            return self.form_invalid(form)
    
    def form_invalid(self, form):
        if form.errors:
            details = []
            for field, errs in form.errors.items():
                for err in errs:
                    details.append(f"{field}: {err}")
            if details:
                messages.error(self.request, "資料有誤：" + "；".join(details))
        else:
            messages.error(self.request, '請檢查表單資料，有必填欄位未填寫或格式不正確')
        return self.render_to_response(self.get_context_data(form=form))


class OperatorBackfillUpdateView(LoginRequiredMixin, UpdateView):
    """編輯作業員補登填報（僅允許未審核）"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/operator_backfill_form.html'
    form_class = OperatorBackfillForm
    success_url = reverse_lazy('workorder:fill_work:operator_index')

    def dispatch(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.approval_status != 'pending':
            messages.error(request, '此筆記錄已審核，無法修改')
            return redirect('workorder:fill_work:operator_index')
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['hours_range'] = [f"{h:02d}" for h in range(0, 24)]
        context['minutes_range'] = [f"{m:02d}" for m in range(0, 60)]
        return context

    def form_valid(self, form):
        try:
            form.instance.created_by = self.request.user.username or form.instance.created_by
            response = super().form_valid(form)
            messages.success(self.request, '作業員補登填報修改成功！')
            return response
        except Exception as e:
            messages.error(self.request, f'儲存失敗：{str(e)}')
            return self.form_invalid(form)

    def form_invalid(self, form):
        if form.errors:
            details = []
            for field, errs in form.errors.items():
                for err in errs:
                    details.append(f"{field}: {err}")
            if details:
                messages.error(self.request, "資料有誤：" + "；".join(details))
        else:
            messages.error(self.request, '請檢查表單資料，有必填欄位未填寫或格式不正確')
        return self.render_to_response(self.get_context_data(form=form))


class FillWorkListView(LoginRequiredMixin, ListView):
    """填報記錄列表視圖"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/fill_work_list.html'
    context_object_name = 'fill_works'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = FillWork.objects.all()
        
        # 根據類型過濾
        fill_type = self.request.GET.get('type')
        if fill_type == 'operator':
            queryset = queryset.exclude(
                Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
            )
        elif fill_type == 'smt':
            queryset = queryset.filter(
                Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
            )
        
        # 根據狀態過濾
        status = self.request.GET.get('status')
        if status == 'pending':
            queryset = queryset.filter(approval_status='pending')
        elif status == 'approved':
            queryset = queryset.filter(approval_status='approved')
        elif status == 'rejected':
            queryset = queryset.filter(approval_status='rejected')
        
        return queryset.order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['fill_type'] = self.request.GET.get('type', '')
        context['status'] = self.request.GET.get('status', '')
        # 統計區塊：以 type 篩選為主，不受 status 影響
        base_qs = FillWork.objects.all()
        fill_type = self.request.GET.get('type')
        if fill_type == 'operator':
            base_qs = base_qs.exclude(Q(operator__icontains='SMT') | Q(process__name__icontains='SMT'))
        elif fill_type == 'smt':
            base_qs = base_qs.filter(Q(operator__icontains='SMT') | Q(process__name__icontains='SMT'))
        context['pending_count'] = base_qs.filter(approval_status='pending').count()
        context['approved_count'] = base_qs.filter(approval_status='approved').count()
        context['rejected_count'] = base_qs.filter(approval_status='rejected').count()
        return context


class FillWorkDetailView(LoginRequiredMixin, DetailView):
    """填報記錄詳情視圖"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/fill_work_detail.html'
    context_object_name = 'fill_work'


class FillWorkDeleteView(LoginRequiredMixin, DeleteView):
    """刪除填報記錄（僅允許待核准）"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/fillwork_confirm_delete.html'
    success_url = reverse_lazy('workorder:fill_work:fill_work_list')

    def dispatch(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj.approval_status != 'pending':
            messages.error(request, '此筆記錄非待核准狀態，無法刪除')
            return redirect(self.success_url)
        return super().dispatch(request, *args, **kwargs)

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        obj_id = obj.id
        response = super().delete(request, *args, **kwargs)
        messages.success(request, f'已刪除填報記錄（ID: {obj_id}）')
        return response


def approve_fill_work(request, pk: int):
    """核准填報記錄"""
    from django.shortcuts import redirect
    from django.http import JsonResponse
    from .services import FillWorkApprovalService
    
    try:
        record = FillWork.objects.get(pk=pk)
        
        # 若已核准，避免重複核准與重複同步
        if record.approval_status == 'approved':
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': '此筆填報記錄已核准，已略過重複操作'
                })
            messages.info(request, '此筆填報記錄已核准，已略過重複操作')
            return redirect('workorder:fill_work:fill_work_list')
        
        # 使用服務層處理核准流程
        approval_result = FillWorkApprovalService.approve_fill_work_record(
            record, request.user.username
        )
        
        if approval_result['success']:
            # 顯示核准成功訊息
            success_message = approval_result['message']
            
            # 如果有RD樣品工單建立，顯示額外訊息
            if approval_result.get('workorder_created'):
                success_message += f"，已自動建立RD樣品工單"
            if approval_result.get('dispatch_created'):
                success_message += f"，已自動建立RD樣品派工單"
            
            # 如果是 AJAX 請求，返回 JSON 回應
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_message
                })
            
            messages.success(request, success_message)
            
            # 如果有RD樣品處理訊息，顯示詳細資訊
            if approval_result.get('rd_message'):
                messages.info(request, f"RD樣品處理: {approval_result['rd_message']}")
        else:
            # 如果是 AJAX 請求，返回 JSON 回應
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': approval_result['message']
                })
            messages.error(request, approval_result['message'])
        
        # 嘗試同步到生產執行監控
        try:
            # 使用公司名稱查找對應的工單，確保資料分離正確
            from erp_integration.models import CompanyConfig
            
            # 先從公司名稱找到公司代號
            company_config = CompanyConfig.objects.filter(
                company_name=record.company_name
            ).first()
            
            if company_config:
                # 使用公司代號和工單號碼查詢工單
                workorder = WorkOrder.objects.filter(
                    company_code=company_config.company_code,
                    order_number=record.workorder
                ).first()
            else:
                # 如果找不到公司配置，使用舊方式查詢（向後相容）
                workorder = WorkOrder.objects.filter(
                    order_number=record.workorder
                ).first()
            
            if workorder:
                # 判斷類型：SMT 或 作業員
                is_smt = False
                try:
                    if (record.operator and 'SMT' in record.operator.upper()) or (record.process and 'SMT' in record.process.name.upper()):
                        is_smt = True
                except Exception:
                    is_smt = False
                report_type = 'smt' if is_smt else 'operator'

                ProductionReportSyncService._create_or_update_production_detail(
                    workorder=workorder,
                    process_name=(record.process.name if record.process else record.operation or ''),
                    report_date=record.work_date,
                    report_time=timezone.now(),
                    work_quantity=record.work_quantity or 0,
                    defect_quantity=record.defect_quantity or 0,
                    operator=(record.operator or None),
                    equipment=(record.equipment or None),
                    report_source='fill_work',
                    start_time=record.start_time,
                    end_time=record.end_time,
                    remarks=record.remarks,
                    abnormal_notes=record.abnormal_notes,
                    original_report_id=record.id,
                    original_report_type='fill_work',
                    work_hours=float(record.work_hours_calculated or 0),
                    overtime_hours=float(record.overtime_hours_calculated or 0),
                    has_break=bool(record.has_break),
                    break_start_time=record.break_start_time,
                    break_end_time=record.break_end_time,
                    break_hours=float(record.break_hours or 0),
                    report_type=report_type,
                    allocated_quantity=0,
                    quantity_source='original',
                    allocation_notes='',
                    is_completed=bool(record.is_completed),
                    completion_method='manual',
                    auto_completed=False,
                    completion_time=None,
                    cumulative_quantity=0,
                    cumulative_hours=float(record.work_hours_calculated or 0),
                    approval_status='approved',
                    approved_by=request.user.username,
                    approved_at=timezone.now(),
                    approval_remarks=record.approval_remarks or ''
                )
                
                # 更新工單狀態
                try:
                    from workorder.services.workorder_status_service import WorkOrderStatusService
                    WorkOrderStatusService.update_workorder_status(workorder.id)
                except Exception as status_error:
                    # 狀態更新錯誤不阻斷使用者操作
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.error(f"工單狀態更新失敗: {str(status_error)}")
                    
        except Exception:
            # 同步錯誤不阻斷使用者操作
            pass
    except FillWork.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': '記錄不存在'
            })
        messages.error(request, '記錄不存在')
    
    # 如果不是 AJAX 請求，返回重定向
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return redirect('workorder:fill_work:fill_work_list')
    
    # 預設返回成功回應
    return JsonResponse({
        'success': True,
        'message': '核准成功'
    })


def cancel_approve_fill_work(request, pk: int):
    """取消核准：將記錄狀態從已核准改回待審核，並移除對應的生產監控明細（若存在）。
    僅限主管與超級管理員使用。
    """
    from django.shortcuts import redirect
    record = get_object_or_404(FillWork, pk=pk)
    if not (request.user.is_superuser or request.user.groups.filter(name__in=['系統管理員', '主管']).exists()):
        messages.error(request, '無權限取消核准')
        return redirect('workorder:fill_work:fill_work_detail', pk=pk)

    if record.approval_status != 'approved':
        messages.info(request, '此筆記錄目前非已核准狀態，無需取消')
        return redirect('workorder:fill_work:fill_work_detail', pk=pk)

    try:
        # 1) 回復狀態
        record.approval_status = 'pending'
        record.save(update_fields=['approval_status', 'updated_at'])

        # 2) 嘗試移除對應生產監控的明細（以來源鍵對應）
        try:
            # 使用公司名稱查找對應的工單，確保資料分離正確
            from erp_integration.models import CompanyConfig
            
            # 先從公司名稱找到公司代號
            company_config = CompanyConfig.objects.filter(
                company_name=record.company_name
            ).first()
            
            if company_config:
                # 使用公司代號和工單號碼查詢工單
                workorder = WorkOrder.objects.filter(
                    company_code=company_config.company_code,
                    order_number=record.workorder
                ).first()
            else:
                # 如果找不到公司配置，使用舊方式查詢（向後相容）
                workorder = WorkOrder.objects.filter(
                    order_number=record.workorder
                ).first()
            
            if workorder:
                from workorder.models import WorkOrderProduction, WorkOrderProductionDetail
                production_record = WorkOrderProduction.objects.filter(workorder=workorder).first()
                if production_record:
                    WorkOrderProductionDetail.objects.filter(
                        workorder_production=production_record,
                        original_report_id=record.id,
                        original_report_type='fill_work'
                    ).delete()
        except Exception:
            pass

        messages.success(request, '已取消核准，記錄已回到待審核')
    except Exception as e:
        messages.error(request, f'取消核准失敗：{e}')
    return redirect('workorder:fill_work:fill_work_detail', pk=pk)


def reject_fill_work(request, pk: int):
    """駁回填報記錄"""
    from django.shortcuts import redirect
    from django.http import JsonResponse
    
    try:
        record = FillWork.objects.get(pk=pk)
        record.approval_status = 'rejected'
        record.rejected_by = request.user.username
        record.rejected_at = timezone.now()
        record.rejection_reason = request.POST.get('reason', '')
        record.save(update_fields=['approval_status', 'rejected_by', 'rejected_at', 'rejection_reason', 'updated_at'])
        
        # 如果是 AJAX 請求，返回 JSON 回應
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': True,
                'message': '已駁回該筆填報記錄'
            })
        
        messages.info(request, '已駁回該筆填報記錄')
    except FillWork.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': '記錄不存在'
            })
        messages.error(request, '記錄不存在')
    
    # 如果不是 AJAX 請求，返回重定向
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return redirect('workorder:fill_work:fill_work_list')
    
    # 預設返回成功回應
    return JsonResponse({
        'success': True,
        'message': '駁回成功'
    })


def delete_fill_work(request, pk: int):
    """刪除填報記錄（僅允許待核准），不使用確認模板，直接刪除後導回列表。"""
    try:
        obj = get_object_or_404(FillWork, pk=pk)
        if obj.approval_status != 'pending':
            messages.error(request, '此筆記錄非待核准狀態，無法刪除')
            return redirect('workorder:fill_work:fill_work_list')
        obj_id = obj.id
        obj.delete()
        messages.success(request, f'已刪除填報記錄（ID: {obj_id}）')
    except Exception as exc:
        messages.error(request, f'刪除失敗：{exc}')
    return redirect('workorder:fill_work:fill_work_list')


# ==================== API 視圖 ====================

# 填報模組的獨立API函數已移除，統一使用 /workorder/api/ 路徑的統一API


# 取得全部派工單號碼的API已移除，統一使用 /workorder/api/ 路徑的統一API


class FillWorkSettingsView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    template_name = 'workorder/workorder_fill_work/settings.html'

    def test_func(self) -> bool:
        user = self.request.user
        return bool(getattr(user, 'is_superuser', False) or user.groups.filter(name__in=['系統管理員', '主管']).exists())

    def handle_no_permission(self):
        messages.error(self.request, '無權限存取填報功能設定')
        return redirect('workorder:fill_work:fill_work_index')


class FillWorkDataSettingsView(FillWorkSettingsView):
    template_name = 'workorder/workorder_fill_work/settings_data.html'


class FillWorkDataOperatorView(FillWorkSettingsView):
    template_name = 'workorder/workorder_fill_work/settings_data_operator.html'


class FillWorkDataSMTView(FillWorkSettingsView):
    template_name = 'workorder/workorder_fill_work/settings_data_smt.html'


@require_POST
@login_required
def import_fill_work_records(request):
    """示意：匯入填報資料時的工時計算策略。
    - 若匯入資料已提供 work_hours_calculated / overtime_hours_calculated（其中一個或兩個），則以匯入值為準，並跳過自動計算。
    - 若未提供，則根據 start_time/end_time 等欄位呼叫 calculate_work_hours() 自動計算。
    注意：此為示意接口，實務上需加入檔案解析與欄位驗證。
    """
    # 這裡省略檔案解析，假設已得到一筆或多筆資料 dicts
    records = []  # TODO: 從上傳檔案解析而來
    created = 0
    for data in records:
        obj = FillWork(
            operator=data.get('operator',''),
            company_name=data.get('company_name',''),
            workorder=data.get('workorder',''),
            product_id=data.get('product_id',''),
            planned_quantity=data.get('planned_quantity') or 0,
            process_id=data.get('process_id'),
            operation=data.get('operation',''),
            equipment=data.get('equipment',''),
            work_date=data.get('work_date'),
            start_time=data.get('start_time'),
            end_time=data.get('end_time'),
            has_break=data.get('has_break') or False,
            break_start_time=data.get('break_start_time'),
            break_end_time=data.get('break_end_time'),
            work_quantity=data.get('work_quantity') or 0,
            defect_quantity=data.get('defect_quantity') or 0,
            approval_status=data.get('approval_status') or 'pending',
            created_by=request.user.username,
        )
        # 判斷是否已有工時計算數值
        has_hours = ('work_hours_calculated' in data and data['work_hours_calculated'] is not None)
        has_overtime = ('overtime_hours_calculated' in data and data['overtime_hours_calculated'] is not None)
        if has_hours or has_overtime:
            if has_hours:
                obj.work_hours_calculated = data['work_hours_calculated']
            if has_overtime:
                obj.overtime_hours_calculated = data['overtime_hours_calculated']
            # 設定臨時旗標，讓 save() 略過自動計算
            obj._skip_auto_hours_calculation = True
            obj.save()
        else:
            # 未提供則自動計算
            obj.calculate_work_hours()
            obj.save()
        created += 1
    messages.success(request, f'匯入完成，共建立 {created} 筆資料')
    return redirect('workorder:fill_work:fill_work_list')


TEMPLATE_HEADERS = [
    '作業員名稱','公司名稱','報工日期','開始時間','結束時間','工單號','產品編號','工序名稱','設備名稱','報工數量','不良品數量','備註','異常紀錄','工作時數','加班時數'
]


def _build_csv_response(filename: str, rows: list[list[str]]):
    output = StringIO()
    writer = csv.writer(output)
    for row in rows:
        writer.writerow(row)
    resp = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@login_required
def download_fill_work_template_operator(request):
    rows = [TEMPLATE_HEADERS]
    return _build_csv_response('operator_fill_work_template.csv', rows)


@login_required
def download_fill_work_template_smt(request):
    rows = [TEMPLATE_HEADERS]
    return _build_csv_response('smt_fill_work_template.csv', rows)


# 匯入（CSV：UTF-8，含標頭）
@login_required
def import_fill_work_records_operator(request):
    if request.method != 'POST':
        messages.error(request, '請使用表單上傳檔案')
        return redirect('workorder:fill_work:fill_work_settings_data_operator')
    file = request.FILES.get('file')
    if not file:
        messages.error(request, '未選擇檔案')
        return redirect('workorder:fill_work:fill_work_settings_data_operator')
    created, errors = _import_any(file, is_smt=False, username=request.user.username)
    if errors:
        if created > 0:
            preview = '；'.join(errors[:5])
            more = '' if len(errors) <= 5 else f'（另有 {len(errors) - 5} 筆錯誤省略）'
            messages.warning(request, f'部分匯入成功：成功 {created} 筆，失敗 {len(errors)} 筆。錯誤摘要：{preview}{more}')
        else:
            messages.error(request, f'匯入失敗：{errors[0]}')
    else:
        messages.success(request, f'匯入完成，共建立 {created} 筆')
    return redirect('workorder:fill_work:fill_work_settings_data_operator')


@login_required
def import_fill_work_records_smt(request):
    if request.method != 'POST':
        messages.error(request, '請使用表單上傳檔案')
        return redirect('workorder:fill_work:fill_work_settings_data_smt')
    file = request.FILES.get('file')
    if not file:
        messages.error(request, '未選擇檔案')
        return redirect('workorder:fill_work:fill_work_settings_data_smt')
    created, errors = _import_any(file, is_smt=True, username=request.user.username)
    if errors:
        if created > 0:
            preview = '；'.join(errors[:5])
            more = '' if len(errors) <= 5 else f'（另有 {len(errors) - 5} 筆錯誤省略）'
            messages.warning(request, f'部分匯入成功：成功 {created} 筆，失敗 {len(errors)} 筆。錯誤摘要：{preview}{more}')
        else:
            messages.error(request, f'匯入失敗：{errors[0]}')
    else:
        messages.success(request, f'匯入完成，共建立 {created} 筆')
    return redirect('workorder:fill_work:fill_work_settings_data_smt')


# 匯出（目前依照 type 參數）
@login_required
def export_fill_work_records_operator(request):
    qs = FillWork.objects.exclude(Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')).order_by('-created_at')
    return _export_fill_work_qs(qs, 'operator_fill_work_export.csv')


@login_required
def export_fill_work_records_smt(request):
    qs = FillWork.objects.filter(Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')).order_by('-created_at')
    return _export_fill_work_qs(qs, 'smt_fill_work_export.csv')


def _export_fill_work_qs(qs, filename: str):
    rows = [TEMPLATE_HEADERS]
    for r in qs:
        rows.append([
            r.operator,
            r.company_name,
            r.work_date.strftime('%Y-%m-%d') if r.work_date else '',
            r.start_time.strftime('%H:%M') if r.start_time else '',
            r.end_time.strftime('%H:%M') if r.end_time else '',
            r.workorder,
            r.product_id,
            r.operation,
            r.equipment,
            r.work_quantity,
            r.defect_quantity,
            r.remarks or '',
            r.abnormal_notes or '',
            str(r.work_hours_calculated),
            str(r.overtime_hours_calculated),
        ])
    return _build_csv_response(filename, rows)


def _parse_time(value: str):
    from datetime import datetime, time as time_cls
    if value is None or value == "":
        return None
    # 允許直接給 datetime.time / datetime.datetime
    if isinstance(value, time_cls):
        return value
    if isinstance(value, datetime):
        return value.time()
    s = str(value).strip()
    for fmt in ('%H:%M', '%H%M', '%H.%M', '%H:%M:%S'):
        try:
            return datetime.strptime(s, fmt).time()
        except Exception:
            pass
    raise ValueError(f'時間格式錯誤：{value}（需 HH:MM）')


def _parse_date(value):
    from datetime import datetime, date as date_cls
    if value is None or value == "":
        return None
    if isinstance(value, date_cls):
        return value
    if isinstance(value, datetime):
        return value.date()
    s = str(value).strip()
    for fmt in ('%Y-%m-%d', '%Y/%m/%d', '%Y.%m.%d', '%Y%m%d'):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            pass
    raise ValueError(f'日期格式錯誤：{value}（需 YYYY-MM-DD）')


def _to_int(val, default=0):
    if val in (None, ""):
        return default
    try:
        return int(val)
    except Exception:
        try:
            return int(float(val))
        except Exception:
            raise ValueError(f'整數欄位格式錯誤：{val}')


def _to_decimal(val):
    if val in (None, ""):
        return None
    try:
        return Decimal(str(val))
    except Exception:
        raise ValueError(f'數值欄位格式錯誤：{val}')


def _create_fillwork_from_row(row: dict, is_smt: bool, username: str) -> FillWork:
    from datetime import datetime
    # 讀取欄位（依指定格式）
    operator_raw = row.get('作業員名稱')
    company_name_raw = row.get('公司名稱')
    work_date_raw = row.get('報工日期')
    start_time_raw = row.get('開始時間')
    end_time_raw = row.get('結束時間')
    workorder_raw = row.get('工單號')
    product_id_raw = row.get('產品編號')
    operation_raw = row.get('工序名稱')
    equipment_raw = row.get('設備名稱')
    work_qty_raw = row.get('報工數量')
    defect_qty_raw = row.get('不良品數量')
    remarks_raw = row.get('備註')
    abnormal_raw = row.get('異常紀錄')
    work_hours_raw = row.get('工作時數')
    overtime_hours_raw = row.get('加班時數')

    # 必要欄位檢查
    required = {
        '公司名稱': company_name_raw,
        '報工日期': work_date_raw,
        '開始時間': start_time_raw,
        '結束時間': end_time_raw,
        '工單號': workorder_raw,
        '產品編號': product_id_raw,
        '工序名稱': operation_raw,
        '報工數量': work_qty_raw,
    }
    missing = [k for k, v in required.items() if v in (None, "")]
    if missing:
        raise ValueError(f"必要欄位缺漏：{', '.join(missing)}")

    # 正規化日期/時間/數值
    work_date = _parse_date(work_date_raw)
    start_time = _parse_time(start_time_raw)
    end_time = _parse_time(end_time_raw)
    work_quantity = _to_int(work_qty_raw, 0)
    defect_quantity = _to_int(defect_qty_raw, 0)

    operator = (str(operator_raw).strip() if operator_raw not in (None, "") else "")
    company_name = str(company_name_raw).strip()
    workorder = str(workorder_raw).strip()
    product_id = str(product_id_raw).strip()
    operation = str(operation_raw).strip()
    equipment = str(equipment_raw).strip() if equipment_raw not in (None, "") else ""
    remarks = "" if remarks_raw in (None, "") else str(remarks_raw)
    abnormal_notes = "" if abnormal_raw in (None, "") else str(abnormal_raw)

    # 公司名稱轉公司代號（統一使用 CompanyConfig）
    company_code = None
    from erp_integration.models import CompanyConfig
    cfg = CompanyConfig.objects.filter(company_name=company_name).first()
    if cfg:
        company_code = cfg.company_code
    else:
        # 若名稱找不到，嘗試用代號本身（使用者可能直接填代號）
        cfg2 = CompanyConfig.objects.filter(company_code=company_name).first()
        if cfg2:
            company_code = cfg2.company_code
        else:
            raise ValueError(f"找不到公司設定：{company_name}")

    # 智能判斷是否為 SMT 類型
    is_smt = is_smt or ('SMT' in (operation.upper() if operation else '')) or ('SMT' in (equipment.upper() if equipment else '')) or ('SMT' in (operator.upper() if operator else ''))

    # 建立 FillWork
    # 對應工序名稱到 ProcessName.id（若不存在則自動建立）
    process_obj = ProcessName.objects.filter(name=operation).first()
    if not process_obj:
        process_obj = ProcessName.objects.create(name=operation)
    obj = FillWork(
        operator=operator or (equipment if is_smt else operator),
        company_name=company_name,
        workorder=workorder,
        product_id=product_id,
        planned_quantity=0,
        operation=operation,
        equipment=equipment,
        work_date=work_date,
        start_time=start_time,
        end_time=end_time,
        has_break=(not is_smt),
        break_start_time=(None if is_smt else _parse_time('12:00')),
        break_end_time=(None if is_smt else _parse_time('13:00')),
        work_quantity=work_quantity,
        defect_quantity=defect_quantity,
        approval_status='pending',
        created_by=username,
        process_id=process_obj.id,
        remarks=remarks,
        abnormal_notes=abnormal_notes,
    )

    # 若匯入有提供工時數值 → 直接使用並跳過自動計算
    provided_hours = False
    if work_hours_raw not in (None, ""):
        obj.work_hours_calculated = _to_decimal(work_hours_raw) or Decimal('0')
        provided_hours = True
    if overtime_hours_raw not in (None, ""):
        obj.overtime_hours_calculated = _to_decimal(overtime_hours_raw) or Decimal('0')
        provided_hours = True

    if provided_hours:
        obj._skip_auto_hours_calculation = True
        obj.save()
    else:
        # 未提供則自動計算（依 has_break 決定規則）
        obj.calculate_work_hours()
        obj.save()
    return obj


@login_required
def download_fill_work_template_operator_xlsx(request):
    if openpyxl is None:
        messages.error(request, '系統未安裝 openpyxl，無法產生 Excel 範本')
        return redirect('workorder:fill_work:fill_work_settings_data_operator')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'OperatorTemplate'
    ws.append(TEMPLATE_HEADERS)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="operator_fill_work_template.xlsx"'
    return resp


@login_required
def download_fill_work_template_smt_xlsx(request):
    if openpyxl is None:
        messages.error(request, '系統未安裝 openpyxl，無法產生 Excel 範本')
        return redirect('workorder:fill_work:fill_work_settings_data_smt')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'SMTTemplate'
    ws.append(TEMPLATE_HEADERS)
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = 'attachment; filename="smt_fill_work_template.xlsx"'
    return resp


# 匯入：支援 CSV 與 XLSX
@login_required
def import_fill_work_records_operator(request):
    if request.method != 'POST':
        messages.error(request, '請使用表單上傳檔案')
        return redirect('workorder:fill_work:fill_work_settings_data_operator')
    file = request.FILES.get('file')
    if not file:
        messages.error(request, '未選擇檔案')
        return redirect('workorder:fill_work:fill_work_settings_data_operator')
    created, errors = _import_any(file, is_smt=False, username=request.user.username)
    if errors:
        if created > 0:
            preview = '；'.join(errors[:5])
            more = '' if len(errors) <= 5 else f'（另有 {len(errors) - 5} 筆錯誤省略）'
            messages.warning(request, f'部分匯入成功：成功 {created} 筆，失敗 {len(errors)} 筆。錯誤摘要：{preview}{more}')
        else:
            messages.error(request, f'匯入失敗：{errors[0]}')
    else:
        messages.success(request, f'匯入完成，共建立 {created} 筆')
    return redirect('workorder:fill_work:fill_work_settings_data_operator')


@login_required
def import_fill_work_records_smt(request):
    if request.method != 'POST':
        messages.error(request, '請使用表單上傳檔案')
        return redirect('workorder:fill_work:fill_work_settings_data_smt')
    file = request.FILES.get('file')
    if not file:
        messages.error(request, '未選擇檔案')
        return redirect('workorder:fill_work:fill_work_settings_data_smt')
    created, errors = _import_any(file, is_smt=True, username=request.user.username)
    if errors:
        if created > 0:
            preview = '；'.join(errors[:5])
            more = '' if len(errors) <= 5 else f'（另有 {len(errors) - 5} 筆錯誤省略）'
            messages.warning(request, f'部分匯入成功：成功 {created} 筆，失敗 {len(errors)} 筆。錯誤摘要：{preview}{more}')
        else:
            messages.error(request, f'匯入失敗：{errors[0]}')
    else:
        messages.success(request, f'匯入完成，共建立 {created} 筆')
    return redirect('workorder:fill_work:fill_work_settings_data_smt')


def _import_any(file, is_smt: bool, username: str):
    name = (file.name or '').lower()
    created = 0
    errors: list[str] = []
    try:
        if file.size == 0:
            return 0, ['上傳檔案為空']
        if name.endswith('.xlsx'):
            if openpyxl is None:
                return 0, ['系統未安裝 openpyxl，無法解析 Excel']
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            headers = [str(c.value).strip() if c.value is not None else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]
            header_map = {h: i for i, h in enumerate(headers)}
            missing = [h for h in TEMPLATE_HEADERS if h not in header_map]
            if missing:
                return 0, [f"必要欄位缺漏：{', '.join(missing)}"]
            if ws.max_row <= 1:
                return 0, ['Excel 無資料內容']
            for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                # 跳過整列空白
                if row is None or all((cell is None or str(cell).strip() == '') for cell in row):
                    continue
                try:
                    data = {h: (row[header_map[h]] if header_map[h] < len(row) else '') for h in TEMPLATE_HEADERS}
                    # 再次檢查必要欄位是否皆空，若皆空則忽略
                    if all((data.get(h) in (None, '') or str(data.get(h)).strip() == '') for h in TEMPLATE_HEADERS):
                        continue
                    _create_fillwork_from_row(data, is_smt=is_smt, username=username)
                    created += 1
                except Exception as e:
                    errors.append(f'第 {idx} 列：{e}')
            return created, errors
        else:
            # CSV
            try:
                decoded = file.read().decode('utf-8')
            except Exception:
                return 0, ['CSV 編碼非 UTF-8，請轉存為 UTF-8 後再上傳']
            reader = csv.DictReader(StringIO(decoded))
            headers = [h.strip() for h in (reader.fieldnames or [])]
            header_map = {h: i for i, h in enumerate(headers)}
            missing = [h for h in TEMPLATE_HEADERS if h not in header_map]
            if missing:
                return 0, [f"必要欄位缺漏：{', '.join(missing)}"]
            any_row = False
            for idx, row in enumerate(reader, start=2):
                # 跳過整列空白
                if all((row.get(h) in (None, '') or str(row.get(h)).strip() == '') for h in row.keys()):
                    continue
                any_row = True
                try:
                    # 若必要欄位皆空白，忽略
                    if all((str(row.get(h, '')).strip() == '') for h in TEMPLATE_HEADERS):
                        continue
                    _create_fillwork_from_row(row, is_smt=is_smt, username=username)
                    created += 1
                except Exception as e:
                    errors.append(f'第 {idx} 列：{e}')
            if not any_row:
                return 0, ['CSV 無資料內容']
            return created, errors
    except Exception as e:
        errors.append(f'匯入過程發生未預期錯誤：{e}')
        return created, errors


# 匯出：新增 XLSX
@login_required
def export_fill_work_records_operator_xlsx(request):
    qs = FillWork.objects.exclude(Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')).order_by('-created_at')
    return _export_fill_work_qs_xlsx(qs, 'operator_fill_work_export.xlsx')


@login_required
def export_fill_work_records_smt_xlsx(request):
    qs = FillWork.objects.filter(Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')).order_by('-created_at')
    return _export_fill_work_qs_xlsx(qs, 'smt_fill_work_export.xlsx')


def _export_fill_work_qs_xlsx(qs, filename: str):
    if openpyxl is None:
        messages.error(request, '系統未安裝 openpyxl，無法匯出 Excel')
        return redirect('workorder:fill_work:fill_work_settings')
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'FillWork'
    ws.append(TEMPLATE_HEADERS)
    for r in qs:
        ws.append([
            r.operator,
            r.company_name,
            r.work_date.strftime('%Y-%m-%d') if r.work_date else '',
            r.start_time.strftime('%H:%M') if r.start_time else '',
            r.end_time.strftime('%H:%M') if r.end_time else '',
            r.workorder,
            r.product_id,
            r.operation,
            r.equipment,
            r.work_quantity,
            r.defect_quantity,
            r.remarks or '',
            r.abnormal_notes or '',
            float(r.work_hours_calculated or 0),
            float(r.overtime_hours_calculated or 0),
        ])
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


def delete_all_fill_work_records(request):
    """刪除所有填報記錄（僅限超級管理員）
    注意：為了與前端一致，允許在 GET 觸發刪除（前端已彈出確認）。
    """
    if not request.user.is_authenticated:
        return redirect('workorder:fill_work:fill_work_list')
    if not request.user.is_superuser:
        messages.error(request, '只有超級管理員可以執行全部刪除')
        return redirect('workorder:fill_work:fill_work_list')
    try:
        total = FillWork.objects.count()
        FillWork.objects.all().delete()
        messages.success(request, f'已刪除全部填報記錄，共 {total} 筆。')
    except Exception as exc:
        messages.error(request, f'刪除失敗：{exc}')
    return redirect('workorder:fill_work:fill_work_list')


class SupervisorApprovalOperatorView(LoginRequiredMixin, ListView):
    """主管審核（作業員專用）列表視圖
    僅顯示非SMT的填報紀錄，提供統計與最近待審核清單。
    """
    model = FillWork
    template_name = 'workorder/workorder_fill_work/supervisor_approval_operator.html'
    context_object_name = 'fill_works'
    paginate_by = 20

    def get_queryset(self):
        return FillWork.objects.exclude(
            Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
        ).order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        base_qs = FillWork.objects.exclude(
            Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
        )
        context['pending_count'] = base_qs.filter(approval_status='pending').count()
        context['approved_count'] = base_qs.filter(approval_status='approved').count()
        context['rejected_count'] = base_qs.filter(approval_status='rejected').count()
        context['total_count'] = base_qs.count()
        context['recent_pending_records'] = base_qs.filter(approval_status='pending').order_by('-created_at')[:10]
        return context


class SupervisorApprovalSMTView(LoginRequiredMixin, ListView):
    """主管審核（SMT專用）列表視圖
    僅顯示SMT相關的填報紀錄，提供統計與最近待審核清單。
    """
    model = FillWork
    template_name = 'workorder/workorder_fill_work/supervisor_approval_smt.html'
    context_object_name = 'fill_works'
    paginate_by = 20

    def get_queryset(self):
        return FillWork.objects.filter(
            Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
        ).order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        base_qs = FillWork.objects.filter(
            Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
        )
        context['pending_count'] = base_qs.filter(approval_status='pending').count()
        context['approved_count'] = base_qs.filter(approval_status='approved').count()
        context['rejected_count'] = base_qs.filter(approval_status='rejected').count()
        context['total_count'] = base_qs.count()
        context['recent_pending_records'] = base_qs.filter(approval_status='pending').order_by('-created_at')[:10]
        return context


class SupervisorApprovalIndexView(LoginRequiredMixin, ListView):
    """主管審核首頁"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/supervisor_approval_index.html'
    context_object_name = 'fill_works'
    paginate_by = 20
    
    def get_queryset(self):
        return FillWork.objects.all().order_by('-created_at')
    
    def get_context_data(self, **kwargs):
        """提供主管首頁統計資訊與最近待審核清單。"""
        from django.utils import timezone
        from datetime import timedelta
        context = super().get_context_data(**kwargs)
        base_qs = FillWork.objects.all()
        # 統計卡
        context['pending_count'] = base_qs.filter(approval_status='pending').count()
        context['approved_count'] = base_qs.filter(approval_status='approved').count()
        context['rejected_count'] = base_qs.filter(approval_status='rejected').count()
        context['total_count'] = base_qs.count()
        # 最近待審核清單
        context['recent_pending_records'] = base_qs.filter(approval_status='pending').order_by('-created_at')[:10]
        # 今日核准/駁回
        today = timezone.localdate()
        start_today = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.min.time()))
        end_today = timezone.make_aware(timezone.datetime.combine(today, timezone.datetime.max.time()))
        context['today_approved'] = base_qs.filter(approval_status='approved', approved_at__range=(start_today, end_today)).count()
        context['today_rejected'] = base_qs.filter(approval_status='rejected', rejected_at__range=(start_today, end_today)).count()
        # 近7天平均審核時間（分鐘）
        seven_days_ago = timezone.now() - timedelta(days=7)
        samples = list(base_qs.filter(
            Q(approval_status='approved', approved_at__gte=seven_days_ago) |
            Q(approval_status='rejected', rejected_at__gte=seven_days_ago)
        ).values('created_at', 'approved_at', 'rejected_at'))
        total_minutes = 0
        count = 0
        for s in samples:
            created_at = s.get('created_at')
            approved_at = s.get('approved_at')
            rejected_at = s.get('rejected_at')
            decision_at = approved_at or rejected_at
            if created_at and decision_at:
                diff = decision_at - created_at
                total_minutes += max(int(diff.total_seconds() // 60), 0)
                count += 1
        context['avg_approval_time'] = round(total_minutes / count) if count else 0
        return context


class SupervisorPendingListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """主管審核－待審核列表（獨立模板）"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/supervisor_pending_list.html'
    context_object_name = 'fill_works'
    paginate_by = 20

    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.groups.filter(name__in=['系統管理員', '主管']).exists()

    def get_queryset(self):
        from datetime import datetime
        qs = FillWork.objects.filter(approval_status='pending')
        operator = self.request.GET.get('operator', '').strip()
        workorder = self.request.GET.get('workorder', '').strip()
        product_id = self.request.GET.get('product_id', '').strip()
        operation = self.request.GET.get('operation', '').strip()
        start_date = self.request.GET.get('start_date', '').strip()
        end_date = self.request.GET.get('end_date', '').strip()
        if operator:
            qs = qs.filter(operator__icontains=operator)
        if workorder:
            qs = qs.filter(workorder__icontains=workorder)
        if product_id:
            qs = qs.filter(product_id__icontains=product_id)
        if operation:
            qs = qs.filter(Q(operation__icontains=operation) | Q(process__name__icontains=operation))
        if start_date:
            try:
                d1 = datetime.strptime(start_date, '%Y-%m-%d').date()
                qs = qs.filter(work_date__gte=d1)
            except ValueError:
                pass
        if end_date:
            try:
                d2 = datetime.strptime(end_date, '%Y-%m-%d').date()
                qs = qs.filter(work_date__lte=d2)
            except ValueError:
                pass
        return qs.order_by('-work_date', '-start_time', '-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 使用與 get_queryset 相同的過濾邏輯來計算統計資料
        from datetime import datetime
        base_qs = FillWork.objects.filter(approval_status='pending')
        
        # 應用相同的過濾條件
        operator = self.request.GET.get('operator', '').strip()
        workorder = self.request.GET.get('workorder', '').strip()
        product_id = self.request.GET.get('product_id', '').strip()
        operation = self.request.GET.get('operation', '').strip()
        start_date = self.request.GET.get('start_date', '').strip()
        end_date = self.request.GET.get('end_date', '').strip()
        
        if operator:
            base_qs = base_qs.filter(operator__icontains=operator)
        if workorder:
            base_qs = base_qs.filter(workorder__icontains=workorder)
        if product_id:
            base_qs = base_qs.filter(product_id__icontains=product_id)
        if operation:
            base_qs = base_qs.filter(Q(operation__icontains=operation) | Q(process__name__icontains=operation))
        if start_date:
            try:
                d1 = datetime.strptime(start_date, '%Y-%m-%d').date()
                base_qs = base_qs.filter(work_date__gte=d1)
            except ValueError:
                pass
        if end_date:
            try:
                d2 = datetime.strptime(end_date, '%Y-%m-%d').date()
                base_qs = base_qs.filter(work_date__lte=d2)
            except ValueError:
                pass
        
        # 計算過濾後的統計資料
        context['total_count'] = base_qs.count()
        context['operator_count'] = base_qs.exclude(Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')).count()
        context['smt_count'] = base_qs.filter(Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')).count()
        
        # 回填查詢值
        request_get = self.request.GET
        context['filter_operator'] = request_get.get('operator', '').strip()
        context['filter_workorder'] = request_get.get('workorder', '').strip()
        context['filter_product_id'] = request_get.get('product_id', '').strip()
        context['filter_operation'] = request_get.get('operation', '').strip()
        context['filter_start_date'] = request_get.get('start_date', '').strip()
        context['filter_end_date'] = request_get.get('end_date', '').strip()
        
        # 加入除錯資訊（僅在開發環境）
        if settings.DEBUG:
            context['debug_info'] = {
                'total_records': context['total_count'],
                'current_page': context.get('page_obj', {}).number if context.get('page_obj') else 1,
                'total_pages': context.get('page_obj', {}).paginator.num_pages if context.get('page_obj') else 1,
                'filters_applied': {
                    'operator': operator,
                    'workorder': workorder,
                    'product_id': product_id,
                    'operation': operation,
                    'start_date': start_date,
                    'end_date': end_date,
                }
            }
        
        return context


class SupervisorReviewedListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """主管審核－已審核列表（獨立模板，含核准與駁回）"""
    model = FillWork
    template_name = 'workorder/workorder_fill_work/supervisor_reviewed_list.html'
    context_object_name = 'fill_works'
    paginate_by = 20

    def test_func(self):
        user = self.request.user
        return user.is_superuser or user.groups.filter(name__in=['系統管理員', '主管']).exists()

    def get_queryset(self):
        from datetime import datetime
        qs = FillWork.objects.exclude(approval_status='pending')
        
        # 篩選參數
        operator = self.request.GET.get('operator', '').strip()
        workorder = self.request.GET.get('workorder', '').strip()
        product_id = self.request.GET.get('product_id', '').strip()
        operation = self.request.GET.get('operation', '').strip()
        approval_status = self.request.GET.get('approval_status', '').strip()
        start_date = self.request.GET.get('start_date', '').strip()
        end_date = self.request.GET.get('end_date', '').strip()
        report_type = self.request.GET.get('report_type', '').strip()
        
        # 應用篩選條件
        if operator:
            qs = qs.filter(operator__icontains=operator)
        if workorder:
            qs = qs.filter(workorder__icontains=workorder)
        if product_id:
            qs = qs.filter(product_id__icontains=product_id)
        if operation:
            qs = qs.filter(Q(operation__icontains=operation) | Q(process__name__icontains=operation))
        if approval_status:
            qs = qs.filter(approval_status=approval_status)
        if start_date:
            try:
                d1 = datetime.strptime(start_date, '%Y-%m-%d').date()
                qs = qs.filter(work_date__gte=d1)
            except ValueError:
                pass
        if end_date:
            try:
                d2 = datetime.strptime(end_date, '%Y-%m-%d').date()
                qs = qs.filter(work_date__lte=d2)
            except ValueError:
                pass
        if report_type:
            if report_type == 'smt':
                qs = qs.filter(Q(operator__icontains='SMT') | Q(process__name__icontains='SMT'))
            elif report_type == 'operator':
                qs = qs.exclude(Q(operator__icontains='SMT') | Q(process__name__icontains='SMT'))
        
        return qs.order_by('-work_date', '-start_time', '-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        base_qs = FillWork.objects.exclude(approval_status='pending')
        
        # 統計數據
        context['total_count'] = base_qs.count()
        context['approved_count'] = base_qs.filter(approval_status='approved').count()
        context['rejected_count'] = base_qs.filter(approval_status='rejected').count()
        
        # 篩選參數（用於保持表單狀態）
        context['filter_operator'] = self.request.GET.get('operator', '')
        context['filter_workorder'] = self.request.GET.get('workorder', '')
        context['filter_product_id'] = self.request.GET.get('product_id', '')
        context['filter_operation'] = self.request.GET.get('operation', '')
        context['filter_approval_status'] = self.request.GET.get('approval_status', '')
        context['filter_start_date'] = self.request.GET.get('start_date', '')
        context['filter_end_date'] = self.request.GET.get('end_date', '')
        context['filter_report_type'] = self.request.GET.get('report_type', '')
        
        # 處理匯出功能
        if self.request.GET.get('export') == 'excel':
            return self.export_to_excel()
        
        return context
    
    def export_to_excel(self):
        """匯出篩選結果到 Excel"""
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from django.http import HttpResponse
        
        # 獲取篩選後的數據
        queryset = self.get_queryset()
        
        # 創建 Excel 工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "已審核填報作業"
        
        # 設定標題行
        headers = [
            '狀態', '類型', '作業員/設備', '工單', '產品', '日期', 
            '開始時間', '結束時間', '數量', '不良品數量', '工作時數', 
            '加班時數', '工序', '備註', '異常記錄', '審核人', '審核時間'
        ]
        
        # 設定標題樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 填充數據
        for row, record in enumerate(queryset, 2):
            # 判斷類型
            report_type = 'SMT' if ('SMT' in (record.operator or '').upper() or 
                                   'SMT' in (record.process.name if record.process else '').upper()) else '作業員'
            
            # 狀態
            status = '已核准' if record.approval_status == 'approved' else '已駁回'
            
            # 審核時間
            review_time = record.approved_at if record.approved_at else record.rejected_at
            
            ws.cell(row=row, column=1, value=status)
            ws.cell(row=row, column=2, value=report_type)
            ws.cell(row=row, column=3, value=record.operator or record.equipment or '')
            ws.cell(row=row, column=4, value=record.workorder or '')
            ws.cell(row=row, column=5, value=record.product_id or '')
            ws.cell(row=row, column=6, value=record.work_date.strftime('%Y-%m-%d') if record.work_date else '')
            ws.cell(row=row, column=7, value=record.start_time.strftime('%H:%M') if record.start_time else '')
            ws.cell(row=row, column=8, value=record.end_time.strftime('%H:%M') if record.end_time else '')
            ws.cell(row=row, column=9, value=record.work_quantity or 0)
            ws.cell(row=row, column=10, value=record.defect_quantity or 0)
            ws.cell(row=row, column=11, value=record.work_hours_calculated or 0)
            ws.cell(row=row, column=12, value=record.overtime_hours_calculated or 0)
            ws.cell(row=row, column=13, value=record.operation or (record.process.name if record.process else '') or '')
            ws.cell(row=row, column=14, value=record.remarks or '')
            ws.cell(row=row, column=15, value=record.abnormal_notes or '')
            ws.cell(row=row, column=16, value=record.approved_by or record.rejected_by or '')
            ws.cell(row=row, column=17, value=review_time.strftime('%Y-%m-%d %H:%M:%S') if review_time else '')
        
        # 調整欄寬
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 保存到記憶體
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 生成檔案名稱
        from datetime import datetime
        filename = f"已審核填報作業_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # 返回 Excel 檔案
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


@login_required
@require_POST
def batch_approve_fill_work(request):
    """
    批次核准填報記錄
    處理多筆填報記錄的批量核准操作
    """
    from django.http import JsonResponse
    
    if request.method != 'POST':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': '無效的請求方法'
            })
        return redirect('workorder:fill_work:supervisor_pending_list')
    
    # 檢查權限
    if not (request.user.is_superuser or request.user.groups.filter(name__in=['系統管理員', '主管']).exists()):
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': '您沒有權限執行批次核准操作'
            })
        messages.error(request, '您沒有權限執行批次核准操作')
        return redirect('workorder:fill_work:supervisor_pending_list')
    
    try:
        # 取得要核准的記錄ID列表
        record_ids = request.POST.getlist('record_ids')
        
        if not record_ids:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': '請選擇要核准的記錄'
                })
            messages.warning(request, '請選擇要核准的記錄')
            return redirect('workorder:fill_work:supervisor_pending_list')
        
        # 使用服務層處理批量核准
        from .services import FillWorkApprovalService
        batch_result = FillWorkApprovalService.batch_approve_fill_work_records(
            record_ids, request.user.username
        )
        
        if batch_result['success']:
            # 顯示核准成功訊息
            success_message = batch_result['message']
            
            # 如果有RD樣品工單建立，顯示額外訊息
            if batch_result.get('rd_workorders_created', 0) > 0:
                success_message += f"，已自動建立 {batch_result['rd_workorders_created']} 個RD樣品工單"
            if batch_result.get('rd_dispatches_created', 0) > 0:
                success_message += f"，已自動建立 {batch_result['rd_dispatches_created']} 個RD樣品派工單"
            
            # 如果有警告訊息，添加到成功訊息中
            if batch_result.get('warnings'):
                warning_text = '；'.join(batch_result['warnings'])
                success_message += f"（注意：{warning_text}）"
            
            # 如果是 AJAX 請求，返回 JSON 回應
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': success_message,
                    'approved_count': batch_result.get('approved_count', 0),
                    'warnings': batch_result.get('warnings', [])
                })
            
            messages.success(request, success_message)
        else:
            # 如果是 AJAX 請求，返回 JSON 回應
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': batch_result['message']
                })
            messages.error(request, batch_result['message'])
        
        # 批量同步到生產執行監控
        try:
            # 取得所有已核准的記錄
            approved_records = FillWork.objects.filter(
                id__in=record_ids,
                approval_status='approved'
            )
            
            synced_workorders = set()  # 記錄需要同步的工單
            
            for record in approved_records:
                if record.workorder:
                    synced_workorders.add(record.workorder)
            
            if synced_workorders:
                for workorder_number in synced_workorders:
                    # 使用公司名稱和工單號碼組合查詢，確保唯一性
                    # 注意：這裡需要從填報記錄中獲取公司名稱
                    fill_work_record = FillWork.objects.filter(workorder=workorder_number).first()
                    if fill_work_record and fill_work_record.company_name:
                        # 先從公司名稱找到公司代號，再查詢工單
                        from erp_integration.models import CompanyConfig
                        company_config = CompanyConfig.objects.filter(
                            company_name=fill_work_record.company_name
                        ).first()
                        
                        if company_config:
                            workorder = WorkOrder.objects.filter(
                                company_code=company_config.company_code,
                                order_number=workorder_number
                            ).first()
                        else:
                            # 如果找不到公司配置，使用舊方式查詢（向後相容）
                            workorder = WorkOrder.objects.filter(order_number=workorder_number).first()
                    else:
                        # 如果沒有公司名稱，使用舊方式查詢（向後相容）
                        workorder = WorkOrder.objects.filter(order_number=workorder_number).first()
                    if workorder:
                        # 取得該工單的所有已核准填報記錄
                        approved_fill_works = FillWork.objects.filter(
                            workorder=workorder_number,
                            approval_status='approved'
                        )
                        
                        for fill_work in approved_fill_works:
                            # 判斷類型：SMT 或 作業員
                            is_smt = False
                            try:
                                if (fill_work.operator and 'SMT' in fill_work.operator.upper()) or (fill_work.process and 'SMT' in fill_work.process.name.upper()):
                                    is_smt = True
                            except Exception:
                                is_smt = False
                            report_type = 'smt' if is_smt else 'operator'

                            ProductionReportSyncService._create_or_update_production_detail(
                                workorder=workorder,
                                process_name=(fill_work.process.name if fill_work.process else fill_work.operation or ''),
                                report_date=fill_work.work_date,
                                report_time=timezone.now(),
                                work_quantity=fill_work.work_quantity or 0,
                                defect_quantity=fill_work.defect_quantity or 0,
                                operator=(fill_work.operator or None),
                                equipment=(fill_work.equipment or None),
                                report_source='fill_work',
                                start_time=fill_work.start_time,
                                end_time=fill_work.end_time,
                                remarks=fill_work.remarks,
                                abnormal_notes=fill_work.abnormal_notes,
                                original_report_id=fill_work.id,
                                original_report_type='fill_work',
                                work_hours=float(fill_work.work_hours_calculated or 0),
                                overtime_hours=float(fill_work.overtime_hours_calculated or 0),
                                has_break=bool(fill_work.has_break),
                                break_start_time=fill_work.break_start_time,
                                break_end_time=fill_work.break_end_time,
                                break_hours=float(fill_work.break_hours or 0),
                                report_type=report_type,
                                allocated_quantity=0,
                                quantity_source='original',
                                allocation_notes='',
                                is_completed=bool(fill_work.is_completed),
                                completion_method='manual',
                                auto_completed=False,
                                completion_time=None,
                                cumulative_quantity=0,
                                cumulative_hours=float(fill_work.work_hours_calculated or 0),
                                approval_status='approved',
                                approved_by=fill_work.approved_by,
                                approved_at=fill_work.approved_at,
                                approval_remarks=fill_work.approval_remarks or ''
                            )
        except Exception as sync_error:
            # 同步失敗不影響核准流程，只記錄錯誤
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"批次同步填報記錄到生產詳情失敗: {str(sync_error)}")
        
        # 批量更新工單狀態
        try:
            from workorder.services.workorder_status_service import WorkOrderStatusService
            
            # 取得所有相關工單並更新狀態
            synced_workorders = set()
            for record in approved_records:
                if record.workorder:
                    synced_workorders.add(record.workorder)
            
            for workorder_number in synced_workorders:
                # 查找對應的工單
                fill_work_record = FillWork.objects.filter(workorder=workorder_number).first()
                if fill_work_record and fill_work_record.company_name:
                    from erp_integration.models import CompanyConfig
                    company_config = CompanyConfig.objects.filter(
                        company_name=fill_work_record.company_name
                    ).first()
                    
                    if company_config:
                        workorder = WorkOrder.objects.filter(
                            company_code=company_config.company_code,
                            order_number=workorder_number
                        ).first()
                    else:
                        workorder = WorkOrder.objects.filter(order_number=workorder_number).first()
                else:
                    workorder = WorkOrder.objects.filter(order_number=workorder_number).first()
                
                if workorder:
                    WorkOrderStatusService.update_workorder_status(workorder.id)
                    
        except Exception as status_error:
            # 狀態更新錯誤不阻斷使用者操作
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"批量工單狀態更新失敗: {str(status_error)}")
        
    except Exception as e:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'success': False,
                'message': f'批次核准失敗: {str(e)}'
            })
        messages.error(request, f'批次核准失敗: {str(e)}')
    
    # 如果不是 AJAX 請求，返回重定向
    if not request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return redirect('workorder:fill_work:supervisor_pending_list')
    
    # 預設返回成功回應
    return JsonResponse({
        'success': True,
        'message': '批次核准完成'
    })


@login_required
@require_POST
def batch_delete_fill_work(request):
    """
    批次刪除填報記錄
    處理多筆填報記錄的批量刪除操作（僅限待核准狀態）
    """
    if request.method != 'POST':
        return redirect('workorder:fill_work:supervisor_pending_list')
    
    # 檢查權限
    if not request.user.is_superuser:
        messages.error(request, '您沒有權限執行批次刪除操作')
        return redirect('workorder:fill_work:supervisor_pending_list')
    
    try:
        action = request.POST.get('action')
        
        if action == 'delete_all':
            # 刪除所有填報記錄（不管狀態）
            total_count = FillWork.objects.count()
            FillWork.objects.all().delete()
            messages.success(request, f'成功刪除全部 {total_count} 筆填報記錄')
        else:
            # 批次刪除選定的記錄
            record_ids = request.POST.getlist('record_ids')
            
            if not record_ids:
                messages.warning(request, '請選擇要刪除的記錄')
                return redirect('workorder:fill_work:supervisor_pending_list')
            
            # 查詢待核准的記錄（只能刪除待核准狀態的記錄）
            pending_records = FillWork.objects.filter(
                id__in=record_ids,
                approval_status='pending'
            )
            
            if not pending_records.exists():
                messages.info(request, '沒有找到可以刪除的記錄（只能刪除待核准狀態的記錄）')
                return redirect('workorder:fill_work:supervisor_pending_list')
            
            deleted_count = 0
            
            # 批量刪除記錄
            for record in pending_records:
                record_id = record.id
                record.delete()
                deleted_count += 1
            
            messages.success(request, f'成功刪除 {deleted_count} 筆填報記錄')
        
    except Exception as e:
        messages.error(request, f'批次刪除失敗: {str(e)}')
    
    return redirect('workorder:fill_work:supervisor_pending_list')


@login_required
@require_POST
def batch_unapprove_fill_work(request):
    """
    批次取消審核填報記錄
    處理多筆已審核填報記錄的批量取消審核操作
    """
    if request.method != 'POST':
        return redirect('workorder:fill_work:supervisor_reviewed_list')
    
    # 檢查權限
    if not (request.user.is_superuser or request.user.is_staff or request.user.groups.filter(name__in=['系統管理員', '主管', '工單使用者']).exists()):
        messages.error(request, '您沒有權限執行批次取消審核操作')
        return redirect('workorder:fill_work:supervisor_reviewed_list')
    
    try:
        # 取得要取消審核的記錄ID列表
        record_ids = request.POST.getlist('record_ids')
        
        if not record_ids:
            messages.warning(request, '請選擇要取消審核的記錄')
            return redirect('workorder:fill_work:supervisor_reviewed_list')
        
        # 查詢已審核的記錄（只能取消已核准或已駁回狀態的記錄）
        reviewed_records = FillWork.objects.filter(
            id__in=record_ids
        ).exclude(approval_status='pending')
        
        if not reviewed_records.exists():
            messages.info(request, '沒有找到可以取消審核的記錄')
            return redirect('workorder:fill_work:supervisor_reviewed_list')
        
        unapproved_count = 0
        
        # 批量取消審核記錄
        for record in reviewed_records:
            # 重置審核狀態
            record.approval_status = 'pending'
            record.approved_by = ''  # 設為空字串而不是 None
            record.approved_at = None
            record.approval_remarks = ''
            record.rejected_by = ''
            record.rejected_at = None
            record.rejection_reason = ''
            record.save()
            unapproved_count += 1
        
        messages.success(request, f'成功取消審核 {unapproved_count} 筆填報記錄')
        
    except Exception as e:
        messages.error(request, f'批次取消審核失敗: {str(e)}')
    
    return redirect('workorder:fill_work:supervisor_reviewed_list')


# ==================== 測試視圖 ====================

@login_required
def test_multi_filter(request):
    """測試多重過濾功能"""
    try:
        from .services import MultiFilterService
        
        # 獲取當前使用者的多重過濾選項
        filtered_choices = MultiFilterService.get_multi_filtered_choices(
            user=request.user,
            form_type='operator',
            permission_type='both'
        )
        
        # 獲取工序查詢集
        process_queryset = MultiFilterService.get_multi_filtered_process_queryset(
            user=request.user,
            form_type='operator',
            permission_type='both'
        )
        
        # 獲取作業員查詢集
        operator_queryset = MultiFilterService.get_multi_filtered_operator_queryset(
            user=request.user,
            form_type='operator',
            permission_type='both'
        )
        
        # 獲取設備查詢集
        equipment_queryset = MultiFilterService.get_multi_filtered_equipment_queryset(
            user=request.user,
            form_type='operator',
            permission_type='both'
        )
        
        context = {
            'user': request.user,
            'filtered_choices': filtered_choices,
            'process_queryset': process_queryset,
            'operator_queryset': operator_queryset,
            'equipment_queryset': equipment_queryset,
            'process_count': process_queryset.count(),
            'operator_count': operator_queryset.count(),
            'equipment_count': equipment_queryset.count(),
        }
        
        return render(request, 'workorder/fill_work/test_multi_filter.html', context)
        
    except Exception as e:
        messages.error(request, f'測試多重過濾功能失敗: {str(e)}')
        return redirect('workorder:fill_work:fill_work_index')