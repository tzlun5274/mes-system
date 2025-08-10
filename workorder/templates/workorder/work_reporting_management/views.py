"""
統一補登報工視圖
提供統一補登報工的完整管理功能，包含列表、新增、編輯、核准、搜尋、匯入匯出等
"""

from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required, permission_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.template.loader import render_to_string
from django.urls import reverse
import json
import csv
import logging
from datetime import datetime, timedelta
from decimal import Decimal

from .models import UnifiedWorkReport, UnifiedWorkReportLog
from .forms import (
    UnifiedWorkReportForm,
    UnifiedWorkReportSimpleForm,
    UnifiedWorkReportApprovalForm,
    UnifiedWorkReportSearchForm,
    UnifiedWorkReportImportForm
)

logger = logging.getLogger(__name__)


@login_required
@permission_required('unified_work_reporting.can_view_unified_work_report', raise_exception=True)
def unified_work_report_list(request):
    """
    統一補登報工列表頁面
    顯示所有統一補登報工記錄，支援搜尋和分頁
    """
    # 處理搜尋表單
    search_form = UnifiedWorkReportSearchForm(request.GET)
    reports = UnifiedWorkReport.objects.all()
    
    if search_form.is_valid():
        # 基本搜尋條件
        operator = search_form.cleaned_data.get('operator')
        if operator:
            reports = reports.filter(operator__icontains=operator)
        
        company_code = search_form.cleaned_data.get('company_code')
        if company_code:
            reports = reports.filter(company_code__icontains=company_code)
        
        workorder_number = search_form.cleaned_data.get('workorder_number')
        if workorder_number:
            reports = reports.filter(
                Q(original_workorder_number__icontains=workorder_number) |
                Q(workorder__order_number__icontains=workorder_number)
            )
        
        product_id = search_form.cleaned_data.get('product_id')
        if product_id:
            reports = reports.filter(product_id__icontains=product_id)
        
        # 日期範圍搜尋
        date_from = search_form.cleaned_data.get('date_from')
        if date_from:
            reports = reports.filter(work_date__gte=date_from)
        
        date_to = search_form.cleaned_data.get('date_to')
        if date_to:
            reports = reports.filter(work_date__lte=date_to)
        
        # 核准狀態搜尋
        approval_status = search_form.cleaned_data.get('approval_status')
        if approval_status:
            reports = reports.filter(approval_status=approval_status)
        
        # 工序搜尋
        process = search_form.cleaned_data.get('process')
        if process:
            reports = reports.filter(process=process)
        
        # 完工狀態搜尋
        is_completed = search_form.cleaned_data.get('is_completed')
        if is_completed:
            reports = reports.filter(is_completed=is_completed == 'True')
    
    # 排序
    reports = reports.order_by('-created_at')
    
    # 分頁
    paginator = Paginator(reports, 20)  # 每頁顯示20筆
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 統計資訊
    total_reports = reports.count()
    pending_reports = reports.filter(approval_status='pending').count()
    approved_reports = reports.filter(approval_status='approved').count()
    rejected_reports = reports.filter(approval_status='rejected').count()
    
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
        'total_reports': total_reports,
        'pending_reports': pending_reports,
        'approved_reports': approved_reports,
        'rejected_reports': rejected_reports,
    }
    
    return render(request, 'unified_work_reporting/unified_work_report_list.html', context)


@login_required
@permission_required('unified_work_reporting.can_add_unified_work_report', raise_exception=True)
def unified_work_report_create(request):
    """
    新增統一補登報工
    """
    if request.method == 'POST':
        form = UnifiedWorkReportSimpleForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user.username
            report.save()
            
            # 記錄操作日誌
            UnifiedWorkReportLog.objects.create(
                work_report=report,
                action='created',
                operator=request.user.username,
                remarks='新增統一補登報工記錄'
            )
            
            messages.success(request, '統一補登報工記錄已成功建立！')
            return redirect('unified_work_reporting:unified_work_report_detail', pk=report.pk)
    else:
        form = UnifiedWorkReportSimpleForm()
    
    context = {
        'form': form,
        'title': '新增統一補登報工',
        'submit_text': '建立報工記錄',
    }
    
    return render(request, 'unified_work_reporting/unified_work_report_form_simple.html', context)


@login_required
@permission_required('unified_work_reporting.can_edit_unified_work_report', raise_exception=True)
def unified_work_report_update(request, pk):
    """
    編輯統一補登報工
    """
    report = get_object_or_404(UnifiedWorkReport, pk=pk)
    
    # 檢查是否可以編輯（已核准的記錄不能編輯）
    if report.approval_status == 'approved':
        messages.error(request, '已核准的報工記錄不能編輯！')
        return redirect('unified_work_reporting:unified_work_report_detail', pk=pk)
    
    if request.method == 'POST':
        form = UnifiedWorkReportSimpleForm(request.POST, instance=report)
        if form.is_valid():
            report = form.save()
            
            # 記錄操作日誌
            UnifiedWorkReportLog.objects.create(
                work_report=report,
                action='updated',
                operator=request.user.username,
                remarks='編輯統一補登報工記錄'
            )
            
            messages.success(request, '統一補登報工記錄已成功更新！')
            return redirect('unified_work_reporting:unified_work_report_detail', pk=report.pk)
    else:
        form = UnifiedWorkReportSimpleForm(instance=report)
    
    context = {
        'form': form,
        'report': report,
        'title': '編輯統一補登報工',
        'submit_text': '更新報工記錄',
    }
    
    return render(request, 'unified_work_reporting/unified_work_report_form_simple.html', context)


@login_required
@permission_required('unified_work_reporting.can_view_unified_work_report', raise_exception=True)
def unified_work_report_detail(request, pk):
    """
    統一補登報工詳細資訊
    """
    report = get_object_or_404(UnifiedWorkReport, pk=pk)
    
    # 取得操作日誌
    logs = report.logs.all()[:10]  # 最近10筆操作記錄
    
    context = {
        'report': report,
        'logs': logs,
    }
    
    return render(request, 'unified_work_reporting/unified_work_report_detail.html', context)


@login_required
@permission_required('unified_work_reporting.can_delete_unified_work_report', raise_exception=True)
def unified_work_report_delete(request, pk):
    """
    刪除統一補登報工
    """
    report = get_object_or_404(UnifiedWorkReport, pk=pk)
    
    # 檢查是否可以刪除（已核准的記錄不能刪除）
    if report.approval_status == 'approved':
        messages.error(request, '已核准的報工記錄不能刪除！')
        return redirect('unified_work_reporting:unified_work_report_detail', pk=pk)
    
    if request.method == 'POST':
        # 記錄操作日誌
        UnifiedWorkReportLog.objects.create(
            work_report=report,
            action='deleted',
            operator=request.user.username,
            remarks='刪除統一補登報工記錄'
        )
        
        report.delete()
        messages.success(request, '統一補登報工記錄已成功刪除！')
        return redirect('unified_work_reporting:unified_work_report_list')
    
    context = {
        'report': report,
    }
    
    return render(request, 'unified_work_reporting/unified_work_report_confirm_delete.html', context)


@login_required
@permission_required('unified_work_reporting.can_approve_unified_work_report', raise_exception=True)
def unified_work_report_approval(request, pk):
    """
    統一補登報工核准/駁回
    """
    report = get_object_or_404(UnifiedWorkReport, pk=pk)
    
    # 檢查是否可以核准（只有待核准的記錄可以核准）
    if report.approval_status != 'pending':
        messages.error(request, '只有待核准的報工記錄可以進行核准操作！')
        return redirect('unified_work_reporting:unified_work_report_detail', pk=pk)
    
    if request.method == 'POST':
        form = UnifiedWorkReportApprovalForm(request.POST)
        if form.is_valid():
            approval_action = form.cleaned_data['approval_action']
            
            if approval_action == 'approve':
                approval_remarks = form.cleaned_data['approval_remarks']
                report.approve(request.user.username, approval_remarks)
                
                # 記錄操作日誌
                UnifiedWorkReportLog.objects.create(
                    work_report=report,
                    action='approved',
                    operator=request.user.username,
                    remarks=f'核准報工記錄，備註：{approval_remarks}'
                )
                
                messages.success(request, '統一補登報工記錄已成功核准！')
            else:
                rejection_reason = form.cleaned_data['rejection_reason']
                report.reject(request.user.username, rejection_reason)
                
                # 記錄操作日誌
                UnifiedWorkReportLog.objects.create(
                    work_report=report,
                    action='rejected',
                    operator=request.user.username,
                    remarks=f'駁回報工記錄，原因：{rejection_reason}'
                )
                
                messages.success(request, '統一補登報工記錄已成功駁回！')
            
            return redirect('unified_work_reporting:unified_work_report_detail', pk=report.pk)
    else:
        form = UnifiedWorkReportApprovalForm()
    
    context = {
        'form': form,
        'report': report,
    }
    
    return render(request, 'unified_work_reporting/unified_work_report_approval.html', context)


@login_required
@permission_required('unified_work_reporting.can_view_unified_work_report', raise_exception=True)
def unified_work_report_export(request):
    """
    匯出統一補登報工資料
    支援 Excel 和 CSV 格式
    """
    # 取得搜尋條件
    search_form = UnifiedWorkReportSearchForm(request.GET)
    reports = UnifiedWorkReport.objects.all()
    
    if search_form.is_valid():
        # 套用搜尋條件（與列表頁面相同的邏輯）
        operator = search_form.cleaned_data.get('operator')
        if operator:
            reports = reports.filter(operator__icontains=operator)
        
        company_code = search_form.cleaned_data.get('company_code')
        if company_code:
            reports = reports.filter(company_code__icontains=company_code)
        
        workorder_number = search_form.cleaned_data.get('workorder_number')
        if workorder_number:
            reports = reports.filter(
                Q(original_workorder_number__icontains=workorder_number) |
                Q(workorder__order_number__icontains=workorder_number)
            )
        
        product_id = search_form.cleaned_data.get('product_id')
        if product_id:
            reports = reports.filter(product_id__icontains=product_id)
        
        date_from = search_form.cleaned_data.get('date_from')
        if date_from:
            reports = reports.filter(work_date__gte=date_from)
        
        date_to = search_form.cleaned_data.get('date_to')
        if date_to:
            reports = reports.filter(work_date__lte=date_to)
        
        approval_status = search_form.cleaned_data.get('approval_status')
        if approval_status:
            reports = reports.filter(approval_status=approval_status)
        
        process = search_form.cleaned_data.get('process')
        if process:
            reports = reports.filter(process=process)
        
        is_completed = search_form.cleaned_data.get('is_completed')
        if is_completed:
            reports = reports.filter(is_completed=is_completed == 'True')
    
    # 排序
    reports = reports.order_by('-created_at')
    
    # 取得匯出格式
    export_format = request.GET.get('format', 'csv')
    
    if export_format == 'excel':
        return _export_to_excel(reports)
    else:
        return _export_to_csv(reports)


def _export_to_csv(reports):
    """匯出為 CSV 格式"""
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="統一補登報工_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    
    # 寫入 BOM 以支援中文
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # 寫入標題列
    writer.writerow([
        '作業員', '公司代號', '工單號碼', '原始工單號碼', '產品編號', '工單預設生產數量',
        '工序', '工序名稱', '使用的設備', '日期', '開始時間', '結束時間',
        '是否有休息時間', '休息開始時間', '休息結束時間', '休息時數',
        '工作時數', '加班時數', '工作數量', '不良品數量', '總數量',
        '是否已完工', '核准狀態', '核准人員', '核准時間', '核准備註',
        '駁回原因', '駁回人員', '駁回時間', '備註', '異常記錄',
        '建立人員', '建立時間', '更新時間'
    ])
    
    # 寫入資料列
    for report in reports:
        writer.writerow([
            report.operator,
            report.company_code,
            report.workorder.order_number if report.workorder else '',
            report.original_workorder_number,
            report.product_id,
            report.planned_quantity,
            report.process.name if report.process else '',
            report.operation,
            report.equipment,
            report.work_date.strftime('%Y-%m-%d') if report.work_date else '',
            report.start_time.strftime('%H:%M') if report.start_time else '',
            report.end_time.strftime('%H:%M') if report.end_time else '',
            '是' if report.has_break else '否',
            report.break_start_time.strftime('%H:%M') if report.break_start_time else '',
            report.break_end_time.strftime('%H:%M') if report.break_end_time else '',
            str(report.break_hours),
            str(report.work_hours_calculated),
            str(report.overtime_hours_calculated),
            report.work_quantity,
            report.defect_quantity,
            report.total_quantity,
            '是' if report.is_completed else '否',
            report.approval_status_display,
            report.approved_by or '',
            report.approved_at.strftime('%Y-%m-%d %H:%M:%S') if report.approved_at else '',
            report.approval_remarks or '',
            report.rejection_reason or '',
            report.rejected_by or '',
            report.rejected_at.strftime('%Y-%m-%d %H:%M:%S') if report.rejected_at else '',
            report.remarks or '',
            report.abnormal_notes or '',
            report.created_by,
            report.created_at.strftime('%Y-%m-%d %H:%M:%S'),
            report.updated_at.strftime('%Y-%m-%d %H:%M:%S')
        ])
    
    return response


def _export_to_excel(reports):
    """匯出為 Excel 格式"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        messages.error(request, '請安裝 openpyxl 套件以支援 Excel 匯出功能')
        return redirect('unified_work_reporting:unified_work_report_list')
    
    # 建立工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "統一補登報工"
    
    # 設定標題列樣式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 標題列
    headers = [
        '作業員', '公司代號', '工單號碼', '原始工單號碼', '產品編號', '工單預設生產數量',
        '工序', '工序名稱', '使用的設備', '日期', '開始時間', '結束時間',
        '是否有休息時間', '休息開始時間', '休息結束時間', '休息時數',
        '工作時數', '加班時數', '工作數量', '不良品數量', '總數量',
        '是否已完工', '核准狀態', '核准人員', '核准時間', '核准備註',
        '駁回原因', '駁回人員', '駁回時間', '備註', '異常記錄',
        '建立人員', '建立時間', '更新時間'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 資料列
    for row, report in enumerate(reports, 2):
        ws.cell(row=row, column=1, value=report.operator)
        ws.cell(row=row, column=2, value=report.company_code)
        ws.cell(row=row, column=3, value=report.workorder.order_number if report.workorder else '')
        ws.cell(row=row, column=4, value=report.original_workorder_number)
        ws.cell(row=row, column=5, value=report.product_id)
        ws.cell(row=row, column=6, value=report.planned_quantity)
        ws.cell(row=row, column=7, value=report.process.name if report.process else '')
        ws.cell(row=row, column=8, value=report.operation)
        ws.cell(row=row, column=9, value=report.equipment)
        ws.cell(row=row, column=10, value=report.work_date.strftime('%Y-%m-%d') if report.work_date else '')
        ws.cell(row=row, column=11, value=report.start_time.strftime('%H:%M') if report.start_time else '')
        ws.cell(row=row, column=12, value=report.end_time.strftime('%H:%M') if report.end_time else '')
        ws.cell(row=row, column=13, value='是' if report.has_break else '否')
        ws.cell(row=row, column=14, value=report.break_start_time.strftime('%H:%M') if report.break_start_time else '')
        ws.cell(row=row, column=15, value=report.break_end_time.strftime('%H:%M') if report.break_end_time else '')
        ws.cell(row=row, column=16, value=float(report.break_hours))
        ws.cell(row=row, column=17, value=float(report.work_hours_calculated))
        ws.cell(row=row, column=18, value=float(report.overtime_hours_calculated))
        ws.cell(row=row, column=19, value=report.work_quantity)
        ws.cell(row=row, column=20, value=report.defect_quantity)
        ws.cell(row=row, column=21, value=report.total_quantity)
        ws.cell(row=row, column=22, value='是' if report.is_completed else '否')
        ws.cell(row=row, column=23, value=report.approval_status_display)
        ws.cell(row=row, column=24, value=report.approved_by or '')
        ws.cell(row=row, column=25, value=report.approved_at.strftime('%Y-%m-%d %H:%M:%S') if report.approved_at else '')
        ws.cell(row=row, column=26, value=report.approval_remarks or '')
        ws.cell(row=row, column=27, value=report.rejection_reason or '')
        ws.cell(row=row, column=28, value=report.rejected_by or '')
        ws.cell(row=row, column=29, value=report.rejected_at.strftime('%Y-%m-%d %H:%M:%S') if report.rejected_at else '')
        ws.cell(row=row, column=30, value=report.remarks or '')
        ws.cell(row=row, column=31, value=report.abnormal_notes or '')
        ws.cell(row=row, column=32, value=report.created_by)
        ws.cell(row=row, column=33, value=report.created_at.strftime('%Y-%m-%d %H:%M:%S'))
        ws.cell(row=row, column=34, value=report.updated_at.strftime('%Y-%m-%d %H:%M:%S'))
    
    # 調整欄寬
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 建立回應
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="統一補登報工_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


@login_required
@permission_required('unified_work_reporting.can_add_unified_work_report', raise_exception=True)
def unified_work_report_import(request):
    """
    匯入統一補登報工資料
    支援 Excel 和 CSV 格式
    """
    if request.method == 'POST':
        form = UnifiedWorkReportImportForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                import_result = _import_from_file(
                    request.FILES['file'],
                    form.cleaned_data['company_code'],
                    form.cleaned_data['created_by'],
                    form.cleaned_data['skip_duplicates']
                )
                
                messages.success(
                    request, 
                    f'匯入完成！成功匯入 {import_result["success_count"]} 筆記錄，'
                    f'失敗 {import_result["error_count"]} 筆記錄。'
                )
                
                if import_result['errors']:
                    for error in import_result['errors'][:5]:  # 只顯示前5個錯誤
                        messages.warning(request, error)
                
                return redirect('unified_work_reporting:unified_work_report_list')
                
            except Exception as e:
                logger.error(f"匯入統一補登報工時發生錯誤: {e}")
                messages.error(request, f'匯入失敗：{str(e)}')
    else:
        form = UnifiedWorkReportImportForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'unified_work_reporting/unified_work_report_import.html', context)


def _import_from_file(file, company_code, created_by, skip_duplicates):
    """
    從檔案匯入統一補登報工資料
    """
    import_result = {
        'success_count': 0,
        'error_count': 0,
        'errors': []
    }
    
    try:
        if file.name.lower().endswith('.csv'):
            # 處理 CSV 檔案
            import_result = _import_from_csv(file, company_code, created_by, skip_duplicates)
        else:
            # 處理 Excel 檔案
            import_result = _import_from_excel(file, company_code, created_by, skip_duplicates)
    
    except Exception as e:
        logger.error(f"匯入檔案時發生錯誤: {e}")
        import_result['errors'].append(f"檔案處理錯誤：{str(e)}")
        import_result['error_count'] += 1
    
    return import_result


def _import_from_csv(file, company_code, created_by, skip_duplicates):
    """
    從 CSV 檔案匯入資料
    """
    import_result = {
        'success_count': 0,
        'error_count': 0,
        'errors': []
    }
    
    # 讀取 CSV 檔案
    decoded_file = file.read().decode('utf-8-sig')  # 處理 BOM
    csv_data = csv.reader(decoded_file.splitlines())
    
    # 跳過標題列
    next(csv_data, None)
    
    for row_num, row in enumerate(csv_data, 2):
        try:
            if len(row) < 15:  # 最少需要15個欄位
                import_result['errors'].append(f"第 {row_num} 行：欄位數量不足")
                import_result['error_count'] += 1
                continue
            
            # 解析資料
            report_data = {
                'operator': row[0].strip(),
                'company_code': company_code,
                'original_workorder_number': row[3].strip(),
                'product_id': row[4].strip(),
                'planned_quantity': int(row[5]) if row[5].strip().isdigit() else 0,
                'operation': row[7].strip(),
                'equipment': row[8].strip(),
                'work_date': datetime.strptime(row[9], '%Y-%m-%d').date() if row[9].strip() else None,
                'start_time': datetime.strptime(row[10], '%H:%M').time() if row[10].strip() else None,
                'end_time': datetime.strptime(row[11], '%H:%M').time() if row[11].strip() else None,
                'has_break': row[12].strip() == '是',
                'break_start_time': datetime.strptime(row[13], '%H:%M').time() if row[13].strip() else None,
                'break_end_time': datetime.strptime(row[14], '%H:%M').time() if row[14].strip() else None,
                'work_quantity': int(row[18]) if row[18].strip().isdigit() else 0,
                'defect_quantity': int(row[19]) if row[19].strip().isdigit() else 0,
                'is_completed': row[21].strip() == '是',
                'remarks': row[28].strip() if len(row) > 28 else '',
                'abnormal_notes': row[29].strip() if len(row) > 29 else '',
                'created_by': created_by,
            }
            
            # 檢查重複
            if skip_duplicates:
                existing = UnifiedWorkReport.objects.filter(
                    operator=report_data['operator'],
                    original_workorder_number=report_data['original_workorder_number'],
                    work_date=report_data['work_date'],
                    start_time=report_data['start_time']
                ).first()
                
                if existing:
                    import_result['errors'].append(f"第 {row_num} 行：重複記錄，已跳過")
                    continue
            
            # 建立記錄
            report = UnifiedWorkReport(**report_data)
            report.save()
            
            import_result['success_count'] += 1
            
        except Exception as e:
            import_result['errors'].append(f"第 {row_num} 行：{str(e)}")
            import_result['error_count'] += 1
    
    return import_result


def _import_from_excel(file, company_code, created_by, skip_duplicates):
    """
    從 Excel 檔案匯入資料
    """
    try:
        import openpyxl
    except ImportError:
        raise Exception("請安裝 openpyxl 套件以支援 Excel 匯入功能")
    
    import_result = {
        'success_count': 0,
        'error_count': 0,
        'errors': []
    }
    
    # 讀取 Excel 檔案
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    
    # 跳過標題列
    for row_num in range(2, ws.max_row + 1):
        try:
            row = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_num, column=col).value
                row.append(str(cell_value) if cell_value is not None else '')
            
            if len(row) < 15:  # 最少需要15個欄位
                import_result['errors'].append(f"第 {row_num} 行：欄位數量不足")
                import_result['error_count'] += 1
                continue
            
            # 解析資料（與 CSV 相同的邏輯）
            report_data = {
                'operator': row[0].strip(),
                'company_code': company_code,
                'original_workorder_number': row[3].strip(),
                'product_id': row[4].strip(),
                'planned_quantity': int(row[5]) if row[5].strip().isdigit() else 0,
                'operation': row[7].strip(),
                'equipment': row[8].strip(),
                'work_date': datetime.strptime(row[9], '%Y-%m-%d').date() if row[9].strip() else None,
                'start_time': datetime.strptime(row[10], '%H:%M').time() if row[10].strip() else None,
                'end_time': datetime.strptime(row[11], '%H:%M').time() if row[11].strip() else None,
                'has_break': row[12].strip() == '是',
                'break_start_time': datetime.strptime(row[13], '%H:%M').time() if row[13].strip() else None,
                'break_end_time': datetime.strptime(row[14], '%H:%M').time() if row[14].strip() else None,
                'work_quantity': int(row[18]) if row[18].strip().isdigit() else 0,
                'defect_quantity': int(row[19]) if row[19].strip().isdigit() else 0,
                'is_completed': row[21].strip() == '是',
                'remarks': row[28].strip() if len(row) > 28 else '',
                'abnormal_notes': row[29].strip() if len(row) > 29 else '',
                'created_by': created_by,
            }
            
            # 檢查重複
            if skip_duplicates:
                existing = UnifiedWorkReport.objects.filter(
                    operator=report_data['operator'],
                    original_workorder_number=report_data['original_workorder_number'],
                    work_date=report_data['work_date'],
                    start_time=report_data['start_time']
                ).first()
                
                if existing:
                    import_result['errors'].append(f"第 {row_num} 行：重複記錄，已跳過")
                    continue
            
            # 建立記錄
            report = UnifiedWorkReport(**report_data)
            report.save()
            
            import_result['success_count'] += 1
            
        except Exception as e:
            import_result['errors'].append(f"第 {row_num} 行：{str(e)}")
            import_result['error_count'] += 1
    
    return import_result


@login_required
@permission_required('unified_work_reporting.can_view_unified_work_report', raise_exception=True)
def unified_work_report_dashboard(request):
    """
    統一補登報工儀表板
    顯示統計資訊和圖表
    """
    # 取得統計資料
    total_reports = UnifiedWorkReport.objects.count()
    pending_reports = UnifiedWorkReport.objects.filter(approval_status='pending').count()
    approved_reports = UnifiedWorkReport.objects.filter(approval_status='approved').count()
    rejected_reports = UnifiedWorkReport.objects.filter(approval_status='rejected').count()
    
    # 最近30天的統計
    thirty_days_ago = timezone.now().date() - timedelta(days=30)
    recent_reports = UnifiedWorkReport.objects.filter(created_at__date__gte=thirty_days_ago)
    
    # 按核准狀態分組
    status_stats = recent_reports.values('approval_status').annotate(
        count=Count('id')
    ).order_by('approval_status')
    
    # 按日期分組
    daily_stats = recent_reports.extra(
        select={'day': 'date(created_at)'}
    ).values('day').annotate(
        count=Count('id')
    ).order_by('day')
    
    # 按作業員分組
    operator_stats = recent_reports.values('operator').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    context = {
        'total_reports': total_reports,
        'pending_reports': pending_reports,
        'approved_reports': approved_reports,
        'rejected_reports': rejected_reports,
        'status_stats': list(status_stats),
        'daily_stats': list(daily_stats),
        'operator_stats': list(operator_stats),
    }
    
    return render(request, 'unified_work_reporting/unified_work_report_dashboard.html', context)


@csrf_exempt
@require_http_methods(["POST"])
def unified_work_report_api(request):
    """
    統一補登報工 API
    提供 AJAX 介面
    """
    try:
        data = json.loads(request.body)
        action = data.get('action')
        
        if action == 'get_workorder_info':
            # 取得工單資訊
            workorder_number = data.get('workorder_number')
            workorder = WorkOrder.objects.filter(order_number=workorder_number).first()
            
            if workorder:
                return JsonResponse({
                    'success': True,
                    'data': {
                        'product_code': workorder.product_code,
                        'quantity': workorder.quantity,
                        'status': workorder.status,
                    }
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': '找不到指定的工單'
                })
        
        elif action == 'get_process_info':
            # 取得工序資訊
            process_name = data.get('process_name')
            process = ProcessName.objects.filter(name=process_name).first()
            
            if process:
                return JsonResponse({
                    'success': True,
                    'data': {
                        'name': process.name,
                        'description': process.description,
                    }
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': '找不到指定的工序'
                })
        
        else:
            return JsonResponse({
                'success': False,
                'message': '不支援的操作'
            })
    
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'message': '無效的 JSON 格式'
        })
    except Exception as e:
        logger.error(f"API 錯誤: {e}")
        return JsonResponse({
            'success': False,
            'message': f'系統錯誤：{str(e)}'
        })
