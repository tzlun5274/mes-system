"""
報表服務模組
提供各種報表功能的業務邏輯
"""

from django.db.models import Sum, Avg, Count, Q
from django.utils import timezone
from datetime import datetime, timedelta, time
import calendar
from decimal import Decimal
import logging

from .models import WorkOrderReportData, ReportSchedule, ReportExecutionLog

logger = logging.getLogger(__name__)


class WorkHourReportService:
    """工作時數報表服務"""
    
    def __init__(self, company_code=None):
        self.company_code = company_code
    
    def get_daily_report(self, company_code, date):
        """日報表"""
        return WorkOrderReportData.objects.filter(
            company=company_code,
            work_date=date
        )
    
    def get_daily_report_by_company_operator(self, company, operator, date):
        """按公司和作業員查詢日報表"""
        queryset = WorkOrderReportData.objects.filter(work_date=date)
        
        if company != 'all':
            queryset = queryset.filter(company=company)
        
        if operator != 'all':
            queryset = queryset.filter(operator_name=operator)
        
        return queryset
    
    def get_daily_summary_by_company_operator(self, company, operator, date):
        """按公司和作業員查詢日報表摘要"""
        data = self.get_daily_report_by_company_operator(company, operator, date)
        
        # 計算總工作時數
        total_work_hours = data.aggregate(total=Sum('daily_work_hours'))['total'] or Decimal('0')
        
        # 計算作業員數量（去重）
        if operator == 'all':
            total_operators = data.exclude(operator_name__isnull=True).values('operator_name').distinct().count()
        else:
            total_operators = 1  # 特定作業員
        
        # 計算總設備使用時數
        total_equipment_hours = data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0')
        
        # 計算工單數量
        workorder_count = data.count()
        
        # 計算平均日工作時數
        avg_daily_hours = data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0')
        
        summary = {
            'total_work_hours': total_work_hours,
            'total_operators': total_operators,
            'total_equipment_hours': total_equipment_hours,
            'workorder_count': workorder_count,
            'avg_daily_hours': avg_daily_hours,
        }
        
        return summary
    
    def get_operator_statistics(self, data):
        """生成作業員統計資料"""
        if not data:
            return []
        
        # 按作業員分組統計
        operator_stats = data.values('operator_name').annotate(
            total_work_hours=Sum('daily_work_hours'),
            total_overtime_hours=Sum('overtime_hours'),
            total_hours=Sum('daily_work_hours') + Sum('overtime_hours')
        ).order_by('-total_hours')
        
        # 格式化統計資料
        stats_list = []
        for i, stat in enumerate(operator_stats, 1):
            stats_list.append({
                'rank': i,
                'operator_name': stat['operator_name'] or '未指定',
                'work_hours': float(stat['total_work_hours'] or 0),
                'overtime_hours': float(stat['total_overtime_hours'] or 0),
                'total_hours': float(stat['total_hours'] or 0),
            })
        
        return stats_list
    
    def get_custom_report_by_company_operator(self, company, operator, start_date, end_date):
        """按公司和作業員查詢自訂報表"""
        queryset = WorkOrderReportData.objects.filter(
            work_date__range=[start_date, end_date]
        )
        
        if company != 'all':
            queryset = queryset.filter(company=company)
        
        if operator != 'all':
            queryset = queryset.filter(operator_name=operator)
        
        return queryset.order_by('work_date', 'operator_name')
    
    def get_custom_summary_by_company_operator(self, company, operator, start_date, end_date):
        """按公司和作業員查詢自訂報表摘要"""
        data = self.get_custom_report_by_company_operator(company, operator, start_date, end_date)
        
        # 計算總工作時數
        total_work_hours = data.aggregate(total=Sum('daily_work_hours'))['total'] or Decimal('0')
        
        # 計算作業員數量（去重）
        if operator == 'all':
            total_operators = data.exclude(operator_name__isnull=True).values('operator_name').distinct().count()
        else:
            total_operators = 1  # 特定作業員
        
        # 計算總設備使用時數
        total_equipment_hours = data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0')
        
        # 計算工單數量
        workorder_count = data.count()
        
        # 計算平均日工作時數
        avg_daily_hours = data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0')
        
        summary = {
            'total_work_hours': total_work_hours,
            'total_operators': total_operators,
            'total_equipment_hours': total_equipment_hours,
            'workorder_count': workorder_count,
            'avg_daily_hours': avg_daily_hours,
        }
        
        return summary
    
    def get_weekly_report(self, company_code, year, week):
        """週報表"""
        return WorkOrderReportData.objects.filter(
            company=company_code,
            work_year=year,
            work_week=week
        )
    
    def get_monthly_report(self, company_code, year, month):
        """月報表"""
        return WorkOrderReportData.objects.filter(
            company=company_code,
            work_year=year,
            work_month=month
        )
    
    def get_quarterly_report(self, company_code, year, quarter):
        """季報表"""
        return WorkOrderReportData.objects.filter(
            company=company_code,
            work_year=year,
            work_quarter=quarter
        )
    
    def get_yearly_report(self, company_code, year):
        """年報表"""
        return WorkOrderReportData.objects.filter(
            company=company_code,
            work_year=year
        )
    
    def get_daily_summary(self, company_code, date):
        """日報表摘要"""
        data = self.get_daily_report(company_code, date)
        
        summary = {
            'total_work_hours': data.aggregate(total=Sum('daily_work_hours'))['total'] or Decimal('0'),
            'total_operators': data.aggregate(total=Sum('operator_count'))['total'] or 0,
            'total_equipment_hours': data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0'),
            'workorder_count': data.count(),
            'avg_daily_hours': data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0'),
        }
        
        return summary
    
    def get_weekly_summary(self, company_code, year, week):
        """週報表摘要"""
        data = self.get_weekly_report(company_code, year, week)
        
        summary = {
            'total_work_hours': data.aggregate(total=Sum('daily_work_hours'))['total'] or Decimal('0'),
            'total_operators': data.aggregate(total=Sum('operator_count'))['total'] or 0,
            'total_equipment_hours': data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0'),
            'workorder_count': data.count(),
            'avg_daily_hours': data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0'),
        }
        
        return summary
    
    def get_monthly_summary(self, company_code, year, month):
        """月報表摘要"""
        data = self.get_monthly_report(company_code, year, month)
        
        summary = {
            'total_work_hours': data.aggregate(total=Sum('monthly_work_hours'))['total'] or Decimal('0'),
            'total_operators': data.aggregate(total=Sum('operator_count'))['total'] or 0,
            'total_equipment_hours': data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0'),
            'workorder_count': data.count(),
            'avg_daily_hours': data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0'),
        }
        
        return summary
    
    def get_quarterly_summary(self, company_code, year, quarter):
        """季報表摘要"""
        data = self.get_quarterly_report(company_code, year, quarter)
        
        summary = {
            'total_work_hours': data.aggregate(total=Sum('monthly_work_hours'))['total'] or Decimal('0'),
            'total_operators': data.aggregate(total=Sum('operator_count'))['total'] or 0,
            'total_equipment_hours': data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0'),
            'workorder_count': data.count(),
            'avg_daily_hours': data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0'),
        }
        
        return summary
    
    def get_yearly_summary(self, company_code, year):
        """年報表摘要"""
        data = self.get_yearly_report(company_code, year)
        
        summary = {
            'total_work_hours': data.aggregate(total=Sum('monthly_work_hours'))['total'] or Decimal('0'),
            'total_operators': data.aggregate(total=Sum('operator_count'))['total'] or 0,
            'total_equipment_hours': data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0'),
            'workorder_count': data.count(),
            'avg_daily_hours': data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0'),
        }
        
        return summary


class ReportGeneratorService:
    """報表生成服務"""
    
    def __init__(self):
        self.work_hour_service = WorkHourReportService()
    
    def generate_daily_report_by_operator(self, operator, date, format='preview'):
        """按作業員生成日報表"""
        try:
            data = self.work_hour_service.get_daily_report_by_operator(operator, date)
            summary = self.work_hour_service.get_daily_summary_by_operator(operator, date)
            
            report_data = {
                'report_type': '日報表',
                'operator': operator,
                'date': date,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
            }
            
            if format == 'excel':
                return self._export_to_excel(report_data, f'作業員日報表_{date}')
            elif format == 'csv':
                return self._export_to_csv(report_data, f'作業員日報表_{date}')
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成日報表失敗: {str(e)}")
            raise
    
    def generate_daily_report(self, company_code, date, format='excel'):
        """生成日報表"""
        try:
            data = self.work_hour_service.get_daily_report(company_code, date)
            summary = self.work_hour_service.get_daily_summary(company_code, date)
            
            report_data = {
                'report_type': '日報表',
                'company_code': company_code,
                'date': date,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
            }
            
            if format == 'excel':
                return self._export_to_excel(report_data, f'日報表_{company_code}_{date}')
            elif format == 'csv':
                return self._export_to_csv(report_data, f'日報表_{company_code}_{date}')
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成日報表失敗: {str(e)}")
            raise
    
    def generate_weekly_report(self, company_code, year, week, format='excel'):
        """生成週報表"""
        try:
            data = self.work_hour_service.get_weekly_report(company_code, year, week)
            summary = self.work_hour_service.get_weekly_summary(company_code, year, week)
            
            report_data = {
                'report_type': '週報表',
                'company_code': company_code,
                'year': year,
                'week': week,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
            }
            
            if format == 'excel':
                return self._export_to_excel(report_data, f'週報表_{company_code}_{year}_{week}')
            elif format == 'csv':
                return self._export_to_csv(report_data, f'週報表_{company_code}_{year}_{week}')
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成週報表失敗: {str(e)}")
            raise
    
    def generate_monthly_report(self, company_code, year, month, format='excel'):
        """生成月報表"""
        try:
            data = self.work_hour_service.get_monthly_report(company_code, year, month)
            summary = self.work_hour_service.get_monthly_summary(company_code, year, month)
            
            report_data = {
                'report_type': '月報表',
                'company_code': company_code,
                'year': year,
                'month': month,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
            }
            
            if format == 'excel':
                return self._export_to_excel(report_data, f'月報表_{company_code}_{year}_{month}')
            elif format == 'csv':
                return self._export_to_csv(report_data, f'月報表_{company_code}_{year}_{month}')
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成月報表失敗: {str(e)}")
            raise
    
    def generate_quarterly_report(self, company_code, year, quarter, format='excel'):
        """生成季報表"""
        try:
            data = self.work_hour_service.get_quarterly_report(company_code, year, quarter)
            summary = self.work_hour_service.get_quarterly_summary(company_code, year, quarter)
            
            report_data = {
                'report_type': '季報表',
                'company_code': company_code,
                'year': year,
                'quarter': quarter,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
            }
            
            if format == 'excel':
                return self._export_to_excel(report_data, f'季報表_{company_code}_{year}_{quarter}')
            elif format == 'csv':
                return self._export_to_csv(report_data, f'季報表_{company_code}_{year}_{quarter}')
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成季報表失敗: {str(e)}")
            raise
    
    def generate_yearly_report(self, company_code, year, format='excel'):
        """生成年報表"""
        try:
            data = self.work_hour_service.get_yearly_report(company_code, year)
            summary = self.work_hour_service.get_yearly_summary(company_code, year)
            
            report_data = {
                'report_type': '年報表',
                'company_code': company_code,
                'year': year,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
            }
            
            if format == 'excel':
                return self._export_to_excel(report_data, f'年報表_{company_code}_{year}')
            elif format == 'csv':
                return self._export_to_csv(report_data, f'年報表_{company_code}_{year}')
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成年報表失敗: {str(e)}")
            raise
    
    def generate_custom_report_by_company_operator(self, company, operator, start_date, end_date, format='excel'):
        """按公司和作業員生成自訂報表"""
        try:
            data = self.work_hour_service.get_custom_report_by_company_operator(company, operator, start_date, end_date)
            summary = self.work_hour_service.get_custom_summary_by_company_operator(company, operator, start_date, end_date)
            operator_stats = self.work_hour_service.get_operator_statistics(data)
            
            report_data = {
                'report_type': '自訂報表',
                'company': company,
                'operator': operator,
                'start_date': start_date,
                'end_date': end_date,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
                'operator_stats': operator_stats,
            }
            
            # 生成檔案名稱
            company_name = company if company != 'all' else '全部公司'
            operator_name = operator if operator != 'all' else '全部作業員'
            filename = f'{company_name}_{operator_name}_{start_date}_{end_date}'
            
            if format == 'excel':
                return self._export_to_excel(report_data, filename)
            elif format == 'csv':
                return self._export_to_csv(report_data, filename)
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成自訂報表失敗: {str(e)}")
            raise
    
    def _export_to_excel(self, report_data, filename):
        """匯出為Excel格式（包含三個活頁簿：統計摘要、詳細、統計）"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            
            # 移除預設的工作表
            wb.remove(wb.active)
            
            # 1. 統計摘要活頁簿
            ws_summary = wb.create_sheet("統計摘要")
            self._create_summary_sheet(ws_summary, report_data)
            
            # 2. 詳細活頁簿
            ws_detail = wb.create_sheet("詳細")
            self._create_detail_sheet(ws_detail, report_data)
            
            # 3. 統計活頁簿
            ws_stats = wb.create_sheet("統計")
            self._create_stats_sheet(ws_stats, report_data)
            
            # 儲存檔案
            file_path = f"/tmp/{filename}.xlsx"
            wb.save(file_path)
            
            return {
                'file_path': file_path,
                'filename': f"{filename}.xlsx",
                'format': 'excel'
            }
            
        except ImportError:
            logger.error("openpyxl未安裝，無法生成Excel檔案")
            raise
        except Exception as e:
            logger.error(f"生成Excel檔案失敗: {str(e)}")
            raise
    
    def _create_summary_sheet(self, ws, report_data):
        """建立統計摘要活頁簿"""
        # 設定樣式
        from openpyxl.styles import Font, PatternFill
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 報表標題
        ws['A1'] = f"{report_data['report_type']} - 統計摘要"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # 基本資訊
        if report_data['report_type'] == '自訂報表':
            ws['A3'] = "公司"
            ws['B3'] = report_data.get('company', '全部')
            ws['C3'] = "作業員"
            ws['D3'] = report_data.get('operator', '全部')
            
            ws['A4'] = "起始日期"
            ws['B4'] = report_data.get('start_date', '').strftime('%Y-%m-%d') if hasattr(report_data.get('start_date', ''), 'strftime') else str(report_data.get('start_date', ''))
            ws['C4'] = "結束日期"
            ws['D4'] = report_data.get('end_date', '').strftime('%Y-%m-%d') if hasattr(report_data.get('end_date', ''), 'strftime') else str(report_data.get('end_date', ''))
            
            ws['A5'] = "生成時間"
            ws['B5'] = report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')
        else:
            ws['A3'] = "作業員"
            ws['B3'] = report_data.get('operator', '全部')
            ws['C3'] = "報表日期"
            ws['D3'] = report_data.get('date', '').strftime('%Y-%m-%d') if hasattr(report_data.get('date', ''), 'strftime') else str(report_data.get('date', ''))
            
            ws['A4'] = "生成時間"
            ws['B4'] = report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')
        
        # 摘要資料
        summary = report_data['summary']
        ws['A6'] = "總工作時數"
        ws['B6'] = float(summary['total_work_hours'])
        ws['C6'] = "總作業員數"
        ws['D6'] = summary['total_operators']
        
        ws['A7'] = "總設備使用時數"
        ws['B7'] = float(summary['total_equipment_hours'])
        ws['C7'] = "工單數量"
        ws['D7'] = summary['workorder_count']
        
        ws['A8'] = "平均日工作時數"
        ws['B8'] = float(summary['avg_daily_hours'])
        
        # 設定欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
    
    def _create_detail_sheet(self, ws, report_data):
        """建立詳細活頁簿"""
        # 設定樣式
        from openpyxl.styles import Font, PatternFill, Border, Side
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 標題
        ws['A1'] = f"{report_data['report_type']} - 詳細資料"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:K1')
        
        # 表頭
        headers = ['作業員', '公司名稱', '工單編號', '產品編號', '工序', '工作日期', '開始時間', '結束時間', '工作時數', '加班時數', '合計時數']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # 資料行
        details = report_data['details']
        for row, detail in enumerate(details, 4):
            ws.cell(row=row, column=1, value=detail.get('operator_name', '-')).border = border
            ws.cell(row=row, column=2, value=detail.get('company', '-')).border = border
            ws.cell(row=row, column=3, value=detail.get('workorder_id', '-')).border = border
            ws.cell(row=row, column=4, value=detail.get('product_code', '-')).border = border
            ws.cell(row=row, column=5, value=detail.get('process_name', '-')).border = border
            ws.cell(row=row, column=6, value=detail.get('work_date', '-')).border = border
            ws.cell(row=row, column=7, value=detail.get('start_time', '-')).border = border
            ws.cell(row=row, column=8, value=detail.get('end_time', '-')).border = border
            ws.cell(row=row, column=9, value=float(detail.get('work_hours', 0))).border = border
            ws.cell(row=row, column=10, value=float(detail.get('overtime_hours', 0))).border = border
            ws.cell(row=row, column=11, value=float(detail.get('total_hours', 0))).border = border
        
        # 設定欄寬
        from openpyxl.utils import get_column_letter
        column_widths = [15, 15, 15, 20, 15, 12, 10, 10, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_stats_sheet(self, ws, report_data):
        """建立統計活頁簿"""
        # 設定樣式
        from openpyxl.styles import Font, PatternFill, Border, Side
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 標題
        ws['A1'] = f"{report_data['report_type']} - 作業員統計"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        # 表頭
        headers = ['排名', '作業員', '工作時數', '加班時數', '合計時數']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # 計算作業員統計
        details = report_data['details']
        operator_stats = {}
        
        for detail in details:
            operator = detail.get('operator_name', '未指定')
            if operator not in operator_stats:
                operator_stats[operator] = {
                    'work_hours': 0,
                    'overtime_hours': 0,
                    'total_hours': 0
                }
            operator_stats[operator]['work_hours'] += float(detail.get('work_hours', 0))
            operator_stats[operator]['overtime_hours'] += float(detail.get('overtime_hours', 0))
            operator_stats[operator]['total_hours'] += float(detail.get('total_hours', 0))
        
        # 排序並寫入資料
        sorted_operators = sorted(operator_stats.items(), key=lambda x: x[1]['total_hours'], reverse=True)
        
        for row, (operator, stats) in enumerate(sorted_operators, 4):
            ws.cell(row=row, column=1, value=row-3).border = border  # 排名
            ws.cell(row=row, column=2, value=operator).border = border  # 作業員
            ws.cell(row=row, column=3, value=stats['work_hours']).border = border  # 工作時數
            ws.cell(row=row, column=4, value=stats['overtime_hours']).border = border  # 加班時數
            ws.cell(row=row, column=5, value=stats['total_hours']).border = border  # 合計時數
        
        # 設定欄寬
        from openpyxl.utils import get_column_letter
        column_widths = [8, 20, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _export_to_csv(self, report_data, filename):
        """匯出為CSV格式"""
        try:
            import csv
            import io
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # 寫入標題
            writer.writerow([f"{report_data['report_type']}"])
            writer.writerow([])
            
            # 寫入基本資訊
            writer.writerow(["公司代號", report_data['company_code'], "生成時間", report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')])
            writer.writerow([])
            
            # 寫入摘要
            summary = report_data['summary']
            writer.writerow(["總工作時數", float(summary['total_work_hours'])])
            writer.writerow(["總作業員數", summary['total_operators']])
            writer.writerow(["總設備時數", float(summary['total_equipment_hours'])])
            writer.writerow(["工單數量", summary['workorder_count']])
            writer.writerow(["平均日工作時數", float(summary['avg_daily_hours'])])
            writer.writerow([])
            
            # 寫入詳細資料
            if report_data['details']:
                writer.writerow(["工單編號", "工作日期", "日工作時數", "作業員人數", "設備時數"])
                for detail in report_data['details']:
                    writer.writerow([
                        detail.get('workorder_id', ''),
                        detail.get('work_date', ''),
                        float(detail.get('daily_work_hours', 0)),
                        detail.get('operator_count', 0),
                        float(detail.get('equipment_hours', 0))
                    ])
            
            # 儲存檔案
            file_path = f"/tmp/{filename}.csv"
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                f.write(output.getvalue())
            
            return {
                'file_path': file_path,
                'filename': f"{filename}.csv",
                'format': 'csv'
            }
            
        except Exception as e:
            logger.error(f"生成CSV檔案失敗: {str(e)}")
            raise


class ReportSchedulerService:
    """報表排程服務"""
    
    def __init__(self):
        self.report_generator = ReportGeneratorService()
    
    def _get_previous_week(self, current_date):
        """取得上週的年份和週數"""
        from datetime import timedelta
        
        # 找到上週的週一
        days_since_monday = current_date.weekday()
        last_monday = current_date - timedelta(days=days_since_monday + 7)
        
        # 計算上週的年份和週數
        year, week, _ = last_monday.isocalendar()
        return year, week
    
    def _get_previous_month(self, current_date):
        """取得上個月的年份和月份"""
        if current_date.month == 1:
            year = current_date.year - 1
            month = 12
        else:
            year = current_date.year
            month = current_date.month - 1
        return year, month
    
    def _get_previous_quarter(self, current_date):
        """取得上個季度的年份和季度"""
        if current_date.month <= 3:
            year = current_date.year - 1
            quarter = 4
        else:
            year = current_date.year
            quarter = ((current_date.month - 1) // 3)
        return year, quarter
    
    def _get_previous_year(self, current_date):
        """取得上個年度"""
        return current_date.year - 1
    
    def execute_scheduled_reports(self):
        """執行排程報表"""
        try:
            # 取得所有啟用的排程
            schedules = ReportSchedule.objects.filter(status='active')
            
            for schedule in schedules:
                try:
                    # 檢查是否該執行
                    if self._should_execute_schedule(schedule):
                        self._execute_schedule(schedule)
                        
                except Exception as e:
                    logger.error(f"執行排程 {schedule.name} 失敗: {str(e)}")
                    self._log_execution(schedule, 'failed', str(e))
                    
        except Exception as e:
            logger.error(f"執行排程報表失敗: {str(e)}")
    
    def _should_execute_schedule(self, schedule):
        """判斷是否應該執行排程"""
        now = timezone.localtime(timezone.now())
        
        if schedule.report_type == 'data_sync':
            # 每小時執行一次
            return True
            
        elif schedule.report_type == 'previous_workday':
            # 每天早上10點後執行前一個工作日報表
            return now.hour >= 10
            
        elif schedule.report_type in ['current_week', 'previous_week']:
            # 週報表：可自訂週幾執行
            if schedule.schedule_day:
                # 如果設定了週幾，檢查今天是否是指定的週幾
                target_weekday = schedule.schedule_day - 1  # 1=週一, 2=週二, ..., 7=週日
                if target_weekday == 6:  # 週日
                    target_weekday = 6
                elif target_weekday == 0:  # 週一
                    target_weekday = 0
                else:
                    target_weekday = target_weekday
                
                if now.weekday() == target_weekday:
                    if schedule.schedule_time:
                        # 如果用戶設定了時間，檢查是否到了執行時間
                        schedule_hour = schedule.schedule_time.hour
                        schedule_minute = schedule.schedule_time.minute
                        return now.hour > schedule_hour or (now.hour == schedule_hour and now.minute >= schedule_minute)
                    else:
                        # 預設早上9點後執行
                        return now.hour >= 9
            else:
                # 預設每週一執行
                if now.weekday() == 0:  # 週一
                    if schedule.schedule_time:
                        schedule_hour = schedule.schedule_time.hour
                        schedule_minute = schedule.schedule_time.minute
                        return now.hour > schedule_hour or (now.hour == schedule_hour and now.minute >= schedule_minute)
                    else:
                        return now.hour >= 9
            return False
            
        elif schedule.report_type in ['current_month', 'previous_month']:
            # 月報表：可自訂每月幾號執行
            if schedule.schedule_day:
                # 如果設定了日期，檢查今天是否是指定的日期
                if now.day == schedule.schedule_day:
                    if schedule.schedule_time:
                        # 如果用戶設定了時間，檢查是否到了執行時間
                        schedule_hour = schedule.schedule_time.hour
                        schedule_minute = schedule.schedule_time.minute
                        return now.hour > schedule_hour or (now.hour == schedule_hour and now.minute >= schedule_minute)
                    else:
                        # 預設早上9點後執行
                        return now.hour >= 9
            else:
                # 預設每月1號執行
                if now.day == 1:
                    if schedule.schedule_time:
                        schedule_hour = schedule.schedule_time.hour
                        schedule_minute = schedule.schedule_time.minute
                        return now.hour > schedule_hour or (now.hour == schedule_hour and now.minute >= schedule_minute)
                    else:
                        return now.hour >= 9
            return False
            
        elif schedule.report_type in ['current_quarter', 'previous_quarter']:
            # 季報表：可自訂每季第幾天執行
            quarter_start_months = [1, 4, 7, 10]
            if now.month in quarter_start_months:
                if schedule.schedule_day:
                    # 如果設定了日期，檢查今天是否是指定的日期
                    if now.day == schedule.schedule_day:
                        if schedule.schedule_time:
                            # 如果用戶設定了時間，檢查是否到了執行時間
                            schedule_hour = schedule.schedule_time.hour
                            schedule_minute = schedule.schedule_time.minute
                            return now.hour > schedule_hour or (now.hour == schedule_hour and now.minute >= schedule_minute)
                        else:
                            # 預設早上9點後執行
                            return now.hour >= 9
                else:
                    # 預設每季第一天執行
                    if now.day == 1:
                        if schedule.schedule_time:
                            schedule_hour = schedule.schedule_time.hour
                            schedule_minute = schedule.schedule_time.minute
                            return now.hour > schedule_hour or (now.hour == schedule_hour and now.minute >= schedule_minute)
                        else:
                            return now.hour >= 9
            return False
            
        elif schedule.report_type in ['current_year', 'previous_year']:
            # 年報表：可自訂每年幾號執行
            if schedule.schedule_day:
                # 如果設定了日期，檢查今天是否是指定的日期
                if now.month == 1 and now.day == schedule.schedule_day:
                    if schedule.schedule_time:
                        # 如果用戶設定了時間，檢查是否到了執行時間
                        schedule_hour = schedule.schedule_time.hour
                        schedule_minute = schedule.schedule_time.minute
                        return now.hour > schedule_hour or (now.hour == schedule_hour and now.minute >= schedule_minute)
                    else:
                        # 預設早上9點後執行
                        return now.hour >= 9
            else:
                # 預設每年1月1號執行
                if now.month == 1 and now.day == 1:
                    if schedule.schedule_time:
                        schedule_hour = schedule.schedule_time.hour
                        schedule_minute = schedule.schedule_time.minute
                        return now.hour > schedule_hour or (now.hour == schedule_hour and now.minute >= schedule_minute)
                    else:
                        return now.hour >= 9
            return False
            
        else:
            raise ValueError(f"不支援的報表類型: {schedule.report_type}")
    
    def _execute_schedule(self, schedule):
        """執行指定的排程"""
        try:
            if schedule.report_type == 'data_sync':
                result = self._execute_data_sync(schedule)
            elif schedule.report_type == 'previous_workday':
                result = self._execute_previous_workday_report(schedule)
            elif schedule.report_type in ['current_week', 'previous_week']:
                result = self._execute_weekly_report(schedule)
            elif schedule.report_type in ['current_month', 'previous_month']:
                result = self._execute_monthly_report(schedule)
            elif schedule.report_type in ['current_quarter', 'previous_quarter']:
                result = self._execute_quarterly_report(schedule)
            elif schedule.report_type in ['current_year', 'previous_year']:
                result = self._execute_yearly_report(schedule)
            else:
                raise ValueError(f"不支援的報表類型: {schedule.report_type}")
            
            if result['success']:
                # 記錄成功執行
                self._log_execution(schedule, 'success', f'報表生成成功: {result["filename"]}', result['file_path'])
                
                # 發送郵件（如果有設定收件人）
                if schedule.email_recipients:
                    self._send_report_email(schedule, result)
                    
        except Exception as e:
            logger.error(f"執行排程 {schedule.name} 失敗: {str(e)}")
            self._log_execution(schedule, 'failed', str(e))
    
    def _log_execution(self, schedule, status, message, file_path=''):
        """記錄執行日誌"""
        ReportExecutionLog.objects.create(
            report_schedule=schedule,
            status=status,
            message=message,
            file_path=file_path
        )
    
    def _execute_data_sync(self, schedule):
        """執行資料同步"""
        try:
            from workorder.fill_work.models import FillWork
            from workorder.onsite_reporting.models import OnsiteReport
            from datetime import datetime, timedelta
            
            # 同步最近1小時的資料
            sync_time = timezone.now() - timedelta(hours=1)
            
            # 統計同步的資料
            fill_works_count = FillWork.objects.filter(
                created_at__gte=sync_time
            ).count()
            
            onsite_reports_count = OnsiteReport.objects.filter(
                created_at__gte=sync_time
            ).count()
            
            # 生成同步報告
            sync_report = {
                'sync_time': timezone.now(),
                'fill_works_count': fill_works_count,
                'onsite_reports_count': onsite_reports_count,
                'total_records': fill_works_count + onsite_reports_count
            }
            
            # 記錄同步結果
            logger.info(f"資料同步完成: 填報 {fill_works_count} 筆，現場報工 {onsite_reports_count} 筆")
            
            return {
                'success': True,
                'filename': f'data_sync_{timezone.now().strftime("%Y%m%d_%H%M")}.txt',
                'file_path': None,  # 資料同步不需要附件檔案
                'message': f'資料同步完成，共處理 {sync_report["total_records"]} 筆記錄'
            }
            
        except Exception as e:
            logger.error(f"資料同步失敗: {str(e)}")
            return {
                'success': False,
                'filename': '',
                'file_path': '',
                'message': f'資料同步失敗: {str(e)}'
            }
    
    def _execute_previous_workday_report(self, schedule):
        """執行前一個工作日報表"""
        try:
            from .services import PreviousWorkdayReportScheduler
            import os
            
            # 使用前一個工作日報表服務
            scheduler = PreviousWorkdayReportScheduler()
            report_date = scheduler.get_report_date()
            
            # 收集資料
            data = scheduler.collect_data(report_date)
            
            # 生成報表
            result = scheduler.generate_report(data)
            
            # 記錄執行結果
            logger.info(f"前一個工作日報表生成成功: {report_date}")
            
            return {
                'success': True,
                'filename': os.path.basename(result),
                'file_path': result,
                'message': f'前一個工作日報表生成成功，報表日期: {report_date}'
            }
            
        except Exception as e:
            logger.error(f"前一個工作日報表生成失敗: {str(e)}")
            return {
                'success': False,
                'filename': '',
                'file_path': '',
                'message': f'前一個工作日報表生成失敗: {str(e)}'
            }
    
    def _execute_weekly_report(self, schedule):
        """執行週報表"""
        try:
            from datetime import timedelta
            import os
            
            now = timezone.localtime(timezone.now())
            
            if schedule.report_type == 'current_week':
                # 本週報表：從週一到現在
                days_since_monday = now.weekday()
                start_date = now.date() - timedelta(days=days_since_monday)
                end_date = now.date()
                report_title = f"本週報表 ({start_date} ~ {end_date})"
            else:
                # 上週報表：上週一到上週日
                days_since_monday = now.weekday()
                last_monday = now.date() - timedelta(days=days_since_monday + 7)
                last_sunday = last_monday + timedelta(days=6)
                start_date = last_monday
                end_date = last_sunday
                report_title = f"上週報表 ({start_date} ~ {end_date})"
            
            # 收集資料
            data = self._collect_workorder_data(start_date, end_date)
            
            # 根據用戶選擇的格式生成檔案
            html_result = None
            excel_result = None
            file_path = None
            filename = None
            
            if schedule.file_format in ['html', 'both']:
                html_result = self._generate_html_report(data, report_title, schedule)
                if not file_path:
                    file_path = html_result
                    filename = os.path.basename(html_result)
            
            if schedule.file_format in ['excel', 'both']:
                excel_result = self._generate_excel_report(data, report_title, schedule)
                if excel_result:
                    file_path = excel_result
                    filename = os.path.basename(excel_result)
            
            return {
                'success': True,
                'filename': filename,
                'file_path': file_path,
                'html_path': html_result,
                'excel_path': excel_result,
                'message': f'{report_title}生成成功'
            }
            
        except Exception as e:
            logger.error(f"週報表生成失敗: {str(e)}")
            return {
                'success': False,
                'filename': '',
                'file_path': '',
                'message': f'週報表生成失敗: {str(e)}'
            }
    
    def _execute_monthly_report(self, schedule):
        """執行月報表"""
        try:
            from datetime import date
            import os
            
            now = timezone.localtime(timezone.now())
            
            if schedule.report_type == 'current_month':
                # 本月報表：從本月1號到現在
                start_date = date(now.year, now.month, 1)
                end_date = now.date()
                report_title = f"本月報表 ({start_date} ~ {end_date})"
            else:
                # 上月報表：上個月1號到月底
                if now.month == 1:
                    year = now.year - 1
                    month = 12
                else:
                    year = now.year
                    month = now.month - 1
                start_date = date(year, month, 1)
                # 計算上個月的最後一天
                if month == 12:
                    end_date = date(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = date(year, month + 1, 1) - timedelta(days=1)
                report_title = f"上月報表 ({start_date} ~ {end_date})"
            
            # 收集資料
            data = self._collect_workorder_data(start_date, end_date)
            
            # 根據用戶選擇的格式生成檔案
            html_result = None
            excel_result = None
            file_path = None
            filename = None
            
            if schedule.file_format in ['html', 'both']:
                html_result = self._generate_html_report(data, report_title, schedule)
                if not file_path:
                    file_path = html_result
                    filename = os.path.basename(html_result)
            
            if schedule.file_format in ['excel', 'both']:
                excel_result = self._generate_excel_report(data, report_title, schedule)
                if excel_result:
                    file_path = excel_result
                    filename = os.path.basename(excel_result)
            
            return {
                'success': True,
                'filename': filename,
                'file_path': file_path,
                'html_path': html_result,
                'excel_path': excel_result,
                'message': f'{report_title}生成成功'
            }
            
        except Exception as e:
            logger.error(f"月報表生成失敗: {str(e)}")
            return {
                'success': False,
                'filename': '',
                'file_path': '',
                'message': f'月報表生成失敗: {str(e)}'
            }
    
    def _execute_quarterly_report(self, schedule):
        """執行季報表"""
        try:
            from datetime import date
            import os
            
            now = timezone.localtime(timezone.now())
            
            if schedule.report_type == 'current_quarter':
                # 本季報表：從本季第一天到現在
                quarter = (now.month - 1) // 3 + 1
                start_month = (quarter - 1) * 3 + 1
                start_date = date(now.year, start_month, 1)
                end_date = now.date()
                report_title = f"本季報表 ({start_date} ~ {end_date})"
            else:
                # 上季報表：上個季度的完整期間
                if now.month <= 3:
                    year = now.year - 1
                    quarter = 4
                else:
                    year = now.year
                    quarter = ((now.month - 1) // 3)
                    
                start_month = (quarter - 1) * 3 + 1
                start_date = date(year, start_month, 1)
                
                # 計算上季的最後一天
                if quarter == 4:
                    end_date = date(year + 1, 1, 1) - timedelta(days=1)
                else:
                    end_date = date(year, start_month + 3, 1) - timedelta(days=1)
                
                report_title = f"上季報表 ({start_date} ~ {end_date})"
            
            # 收集資料
            data = self._collect_workorder_data(start_date, end_date)
            
            # 根據用戶選擇的格式生成檔案
            html_result = None
            excel_result = None
            file_path = None
            filename = None
            
            if schedule.file_format in ['html', 'both']:
                html_result = self._generate_html_report(data, report_title, schedule)
                if not file_path:
                    file_path = html_result
                    filename = os.path.basename(html_result)
            
            if schedule.file_format in ['excel', 'both']:
                excel_result = self._generate_excel_report(data, report_title, schedule)
                if excel_result:
                    file_path = excel_result
                    filename = os.path.basename(excel_result)
            
            return {
                'success': True,
                'filename': filename,
                'file_path': file_path,
                'html_path': html_result,
                'excel_path': excel_result,
                'message': f'{report_title}生成成功'
            }
                
        except Exception as e:
            logger.error(f"季報表生成失敗: {str(e)}")
            return {
                'success': False,
                'filename': '',
                'file_path': '',
                'message': f'季報表生成失敗: {str(e)}'
            }
    
    def _execute_yearly_report(self, schedule):
        """執行年報表"""
        try:
            from datetime import date
            import os
            
            now = timezone.localtime(timezone.now())
            
            if schedule.report_type == 'current_year':
                # 本年報表：從今年1月1號到現在
                start_date = date(now.year, 1, 1)
                end_date = now.date()
                report_title = f"本年報表 ({start_date} ~ {end_date})"
            else:
                # 上年報表：去年的完整期間
                start_date = date(now.year - 1, 1, 1)
                end_date = date(now.year - 1, 12, 31)
                report_title = f"上年報表 ({start_date} ~ {end_date})"
            
            # 收集資料
            data = self._collect_workorder_data(start_date, end_date)
            
            # 根據用戶選擇的格式生成檔案
            html_result = None
            excel_result = None
            file_path = None
            filename = None
            
            if schedule.file_format in ['html', 'both']:
                html_result = self._generate_html_report(data, report_title, schedule)
                if not file_path:
                    file_path = html_result
                    filename = os.path.basename(html_result)
            
            if schedule.file_format in ['excel', 'both']:
                excel_result = self._generate_excel_report(data, report_title, schedule)
                if excel_result:
                    file_path = excel_result
                    filename = os.path.basename(excel_result)
            
            return {
                'success': True,
                'filename': filename,
                'file_path': file_path,
                'html_path': html_result,
                'excel_path': excel_result,
                'message': f'{report_title}生成成功'
            }
            
        except Exception as e:
            logger.error(f"年報表生成失敗: {str(e)}")
            return {
                'success': False,
                'filename': '',
                'file_path': '',
                'message': f'年報表生成失敗: {str(e)}'
            }
    
    def _collect_workorder_data(self, start_date, end_date):
        """收集工單資料"""
        from reporting.models import WorkOrderReportData
        
        # 查詢指定日期範圍的工單資料
        data = WorkOrderReportData.objects.filter(
            work_date__gte=start_date,
            work_date__lte=end_date
        ).order_by('work_date', 'workorder_id')
        
        return data
    
    def _generate_html_report(self, data, title, schedule):
        """生成HTML報表"""
        import os
        from django.template.loader import render_to_string
        from django.conf import settings
        
        # 準備報表資料
        report_data = {
            'title': title,
            'data': data,
            'total_records': data.count(),
            'generated_at': timezone.localtime(timezone.now()),
            'schedule_name': schedule.name
        }
        
        # 渲染HTML模板
        html_content = render_to_string('reporting/reporting/report_template.html', report_data)
        
        # 生成檔案名稱
        filename = f"{title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.html"
        file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
        
        # 確保目錄存在
        os.makedirs(os.path.dirname(file_path), exist_ok=True)
        
        # 寫入檔案
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        return file_path
    
    def _send_report_email(self, schedule, result):
        """發送報表郵件"""
        try:
            import os
            from django.core.mail import EmailMessage
            from django.conf import settings
            from system.models import EmailConfig
            
            # 取得郵件設定
            try:
                email_config = EmailConfig.objects.first()
                if not email_config:
                    logger.error("未找到郵件設定，無法發送郵件")
                    return
                    
                # 設定郵件連線
                from django.core.mail import get_connection
                connection = get_connection(
                    host=email_config.email_host,
                    port=email_config.email_port,
                    username=email_config.email_host_user,
                    password=email_config.email_host_password,
                    use_tls=email_config.email_use_tls,
                )
                
                from_email = email_config.default_from_email or settings.DEFAULT_FROM_EMAIL
                
            except Exception as e:
                logger.error(f"郵件設定錯誤: {str(e)}")
                return
            
            subject = f"自動報表 - {schedule.name}"
            
            if schedule.report_type == 'data_sync':
                message = f"""
                <html>
                <head>
                    <meta charset="utf-8">
                    <style>
                        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                        .header {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                        .content {{ background-color: #ffffff; padding: 20px; border: 1px solid #dee2e6; border-radius: 5px; }}
                        .footer {{ margin-top: 20px; padding: 15px; background-color: #f8f9fa; border-radius: 5px; font-size: 12px; color: #6c757d; }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <div class="header">
                            <h2>自動報表系統</h2>
                        </div>
                        <div class="content">
                            <p>您好，</p>
                            <p>這是自動執行的 <strong>{schedule.get_report_type_display()}</strong>。</p>
                            <p><strong>同步結果：</strong>{result['message']}</p>
                            <p><strong>執行時間：</strong>{timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                        </div>
                        <div class="footer">
            此郵件由系統自動發送，請勿回覆。
                        </div>
                    </div>
                </body>
                </html>
                """
            else:
                message = f"""
                <html>
                <head>
                    <meta charset="utf-8">
                    <style>
                        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
                        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
                        .header {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                        .content {{ background-color: #ffffff; padding: 20px; border: 1px solid #dee2e6; border-radius: 5px; }}
                        .attachment {{ background-color: #e9ecef; padding: 10px; border-radius: 3px; margin: 10px 0; }}
                        .footer {{ margin-top: 20px; padding: 15px; background-color: #f8f9fa; border-radius: 5px; font-size: 12px; color: #6c757d; }}
                    </style>
                </head>
                <body>
                    <div class="container">
                        <div class="header">
                            <h2>自動報表系統</h2>
                        </div>
                        <div class="content">
                            <p>您好，</p>
                            <p>這是自動生成的 <strong>{schedule.get_report_type_display()}</strong>。</p>
                            <div class="attachment">
                                <p><strong>報表檔案：</strong>{result['filename']}</p>
                                <p><strong>生成時間：</strong>{timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                            </div>
                            <p>報表檔案已作為附件附加到此郵件中，請查收。</p>
                        </div>
                        <div class="footer">
                            此郵件由系統自動發送，請勿回覆。
                        </div>
                    </div>
                </body>
                </html>
                """
            
            # 創建EmailMessage物件
            email = EmailMessage(
                subject=subject,
                body=message,
                from_email=from_email,
                to=schedule.email_recipients.split(','),
                connection=connection,
            )
            email.content_subtype = "html"  # 設定為HTML格式
            
            # 附加檔案
            if schedule.report_type == 'data_sync':
                # 資料同步不需要附加檔案，只發送文字內容
                pass
            elif schedule.file_format == 'both':
                # 如果選擇兩種格式，都附加
                if result.get('excel_path') and os.path.exists(result['excel_path']):
                    with open(result['excel_path'], 'rb') as f:
                        email.attach(result['filename'], f.read(), 'application/vnd.ms-excel')
                
                if result.get('html_path') and os.path.exists(result['html_path']):
                    with open(result['html_path'], 'rb') as f:
                        email.attach(os.path.basename(result['html_path']), f.read(), 'application/octet-stream')
            
            elif schedule.file_format == 'excel':
                # 只附加Excel檔案 - 使用Outlook專用編碼
                if result.get('excel_path') and os.path.exists(result['excel_path']):
                    from email.mime.base import MIMEBase
                    from email import encoders
                    from email.header import Header
                    
                    with open(result['excel_path'], 'rb') as f:
                        attachment = MIMEBase('application', 'vnd.ms-excel')
                        attachment.set_payload(f.read())
                        encoders.encode_base64(attachment)
                        
                        # Outlook專用編碼 - 使用email.header.Header確保中文檔案名稱正確顯示
                        filename = result['filename']
                        encoded_filename = Header(filename, 'utf-8').encode()
                        attachment.add_header('Content-Disposition', 'attachment', filename=encoded_filename)
                        email.attach(attachment)
            
            else:
                # 預設附加HTML檔案
                if result.get('file_path') and os.path.exists(result['file_path']):
                    with open(result['file_path'], 'rb') as f:
                        email.attach(result['filename'], f.read(), 'application/octet-stream')
            
            email.send()
            logger.info(f"報表郵件發送成功: {schedule.name} 發送到 {schedule.email_recipients}")
            print(f"郵件發送成功: {schedule.name} -> {schedule.email_recipients}")
            
        except Exception as e:
            logger.error(f"發送報表郵件失敗: {str(e)}") 
    
    def _generate_excel_report(self, data, title, schedule):
        """生成Excel報表（包含統計摘要、明細、統計三個活頁簿）"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            import os
            from django.conf import settings
            from django.db.models import Sum, Avg, Count
            
            # 建立工作簿
            wb = openpyxl.Workbook()
            
            # 移除預設的工作表
            wb.remove(wb.active)
            
            # 計算統計摘要
            total_work_hours = sum(float(record.work_hours or 0) for record in data)
            total_overtime_hours = sum(float(record.overtime_hours or 0) for record in data)
            total_operators = len(set(record.operator_name for record in data if record.operator_name))
            workorder_count = len(set(record.workorder_id for record in data if record.workorder_id))
            avg_work_hours = total_work_hours / len(data) if data else 0
            
            # 1. 統計摘要活頁簿
            ws_summary = wb.create_sheet("統計摘要")
            self._create_summary_sheet_for_email(ws_summary, {
                'report_type': title,
                'total_work_hours': total_work_hours,
                'total_overtime_hours': total_overtime_hours,
                'total_operators': total_operators,
                'workorder_count': workorder_count,
                'avg_work_hours': avg_work_hours,
                'generated_at': timezone.now()
            })
            
            # 2. 明細活頁簿
            ws_detail = wb.create_sheet("明細")
            self._create_detail_sheet_for_email(ws_detail, data, title)
            
            # 3. 統計活頁簿
            ws_stats = wb.create_sheet("統計")
            self._create_stats_sheet_for_email(ws_stats, data, title)
            
            # 生成檔案名稱 - 使用中文名稱
            filename = f"月工作時數報表_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # 確保目錄存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 儲存檔案
            wb.save(file_path)
            
            return file_path
            
        except ImportError:
            logger.error("未安裝 openpyxl，無法生成 Excel 報表")
            return None
        except Exception as e:
            logger.error(f"生成 Excel 報表失敗: {str(e)}")
            return None
    
    def _create_summary_sheet_for_email(self, ws, report_data):
        """建立統計摘要活頁簿（用於郵件附件）"""
        # 設定樣式
        from openpyxl.styles import Font, Alignment
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 報表標題
        ws['A1'] = f"{report_data['report_type']} - 統計摘要"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # 基本資訊
        ws['A3'] = "生成時間"
        ws['B3'] = report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')
        
        # 摘要資料
        ws['A5'] = "總工作時數"
        ws['B5'] = report_data['total_work_hours']
        ws['C5'] = "總作業員數"
        ws['D5'] = report_data['total_operators']
        
        ws['A6'] = "總加班時數"
        ws['B6'] = report_data['total_overtime_hours']
        ws['C6'] = "工單數量"
        ws['D6'] = report_data['workorder_count']
        
        ws['A7'] = "平均工作時數"
        ws['B7'] = report_data['avg_work_hours']
        
        # 設定樣式 - 只設定粗體，不使用背景色
        for row in [5, 6, 7]:
            for col in ['A', 'B', 'C', 'D']:
                cell = ws[f'{col}{row}']
                cell.font = header_font
        
        # 設定欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
    
    def _create_detail_sheet_for_email(self, ws, data, title):
        """建立明細活頁簿（用於郵件附件）"""
        # 設定樣式
        from openpyxl.styles import Font, Alignment, Border, Side
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 標題
        ws['A1'] = f"{title} - 明細資料"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:K1')
        
        # 表頭
        headers = ['作業員名稱', '公司名稱', '報工日期', '開始時間', '結束時間', 
                  '工單號', '產品編號', '工序名稱', '工作時數', '加班時數', '合計時數']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 填入資料
        for row, record in enumerate(data, 4):
            work_hours = float(record.work_hours or 0)
            overtime_hours = float(record.overtime_hours or 0)
            total_hours = work_hours + overtime_hours
            
            ws.cell(row=row, column=1, value=record.operator_name or '').border = border
            ws.cell(row=row, column=2, value=record.company or '').border = border
            ws.cell(row=row, column=3, value=record.work_date.strftime('%Y-%m-%d') if record.work_date else '').border = border
            ws.cell(row=row, column=4, value=record.start_time.strftime('%H:%M') if record.start_time else '').border = border
            ws.cell(row=row, column=5, value=record.end_time.strftime('%H:%M') if record.end_time else '').border = border
            ws.cell(row=row, column=6, value=record.workorder_id or '').border = border
            ws.cell(row=row, column=7, value=record.product_code or '').border = border
            ws.cell(row=row, column=8, value=record.process_name or '').border = border
            ws.cell(row=row, column=9, value=work_hours).border = border
            ws.cell(row=row, column=10, value=overtime_hours).border = border
            ws.cell(row=row, column=11, value=total_hours).border = border
        
        # 設定欄寬
        from openpyxl.utils import get_column_letter
        column_widths = [15, 15, 12, 10, 10, 15, 20, 15, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_stats_sheet_for_email(self, ws, data, title):
        """建立統計活頁簿（用於郵件附件）"""
        # 設定樣式
        from openpyxl.styles import Font, Alignment, Border, Side
        header_font = Font(bold=True)
        header_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 標題
        ws['A1'] = f"{title} - 作業員統計"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        # 表頭
        headers = ['排名', '作業員', '工作時數', '加班時數', '合計時數']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border
        
        # 計算作業員統計
        operator_stats = {}
        for record in data:
            operator = record.operator_name or '未指定'
            if operator not in operator_stats:
                operator_stats[operator] = {
                    'work_hours': 0,
                    'overtime_hours': 0,
                    'total_hours': 0
                }
            work_hours = float(record.work_hours or 0)
            overtime_hours = float(record.overtime_hours or 0)
            total_hours = work_hours + overtime_hours
            
            operator_stats[operator]['work_hours'] += work_hours
            operator_stats[operator]['overtime_hours'] += overtime_hours
            operator_stats[operator]['total_hours'] += total_hours
        
        # 排序並寫入資料
        sorted_operators = sorted(operator_stats.items(), key=lambda x: x[1]['total_hours'], reverse=True)
        
        for row, (operator, stats) in enumerate(sorted_operators, 4):
            ws.cell(row=row, column=1, value=row-3).border = border  # 排名
            ws.cell(row=row, column=2, value=operator).border = border  # 作業員
            ws.cell(row=row, column=3, value=stats['work_hours']).border = border  # 工作時數
            ws.cell(row=row, column=4, value=stats['overtime_hours']).border = border  # 加班時數
            ws.cell(row=row, column=5, value=stats['total_hours']).border = border  # 合計時數
        
        # 設定欄寬
        from openpyxl.utils import get_column_letter
        column_widths = [8, 20, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width 


class CompletedWorkOrderReportService:
    """
    已完工工單報表服務
    提供已完工工單報表的資料查詢和統計功能
    """
    
    @staticmethod
    def get_completed_workorder_summary(company_code=None, year=None, month=None, quarter=None):
        """
        取得已完工工單統計摘要
        
        Args:
            company_code: 公司代號
            year: 年份
            month: 月份
            quarter: 季度
            
        Returns:
            dict: 統計摘要
        """
        from .models import CompletedWorkOrderAnalysis
        
        queryset = CompletedWorkOrderAnalysis.objects.all()
        
        if company_code:
            queryset = queryset.filter(company_code=company_code)
        if year:
            queryset = queryset.filter(completion_year=year)
        if month:
            queryset = queryset.filter(completion_month=month)
        if quarter:
            queryset = queryset.filter(completion_quarter=quarter)
        
        # 計算統計資料
        stats = queryset.aggregate(
            total_workorders=Count('id'),
            total_work_hours=Sum('total_work_hours'),
            total_overtime_hours=Sum('total_overtime_hours'),
            avg_efficiency_rate=Avg('efficiency_rate'),
        )
        
        # 計算平均效率率
        avg_efficiency_rate = stats['avg_efficiency_rate'] or 0
        
        return {
            'total_workorders': stats['total_workorders'] or 0,
            'total_work_hours': float(stats['total_work_hours'] or 0),
            'total_overtime_hours': float(stats['total_overtime_hours'] or 0),
            'avg_efficiency_rate': float(avg_efficiency_rate),
        }
    
    @staticmethod
    def get_completed_workorder_trend(company_code=None, days=30):
        """
        取得已完工工單趨勢資料
        
        Args:
            company_code: 公司代號
            days: 天數
            
        Returns:
            dict: 趨勢資料
        """
        from .models import CompletedWorkOrderAnalysis
        from datetime import datetime, timedelta
        
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=days)
        
        queryset = CompletedWorkOrderAnalysis.objects.filter(
            completion_date__range=[start_date, end_date]
        )
        
        if company_code:
            queryset = queryset.filter(company_code=company_code)
        
        # 按日期分組統計
        daily_data = queryset.values('completion_date').annotate(
            workorder_count=Count('id'),
            total_work_hours=Sum('total_work_hours'),
            avg_efficiency_rate=Avg('efficiency_rate'),
        ).order_by('completion_date')
        
        labels = []
        workorder_counts = []
        work_hours = []
        efficiency_rates = []
        
        for item in daily_data:
            labels.append(item['completion_date'].strftime('%m/%d'))
            workorder_counts.append(item['workorder_count'])
            work_hours.append(float(item['total_work_hours'] or 0))
            efficiency_rates.append(float(item['avg_efficiency_rate'] or 0))
        
        return {
            'labels': labels,
            'workorder_counts': workorder_counts,
            'completed_quantities': completed_quantities,
            'work_hours': work_hours,
            'completion_rates': completion_rates,
        }
    
    @staticmethod
    def get_company_completion_distribution():
        """
        取得公司完工分布資料
        
        Returns:
            dict: 分布資料
        """
        from .models import CompletedWorkOrderAnalysis
        
        company_data = CompletedWorkOrderAnalysis.objects.values('company_name').annotate(
            workorder_count=Count('id'),
            total_work_hours=Sum('total_work_hours'),
            avg_efficiency_rate=Avg('efficiency_rate'),
        ).order_by('-workorder_count')
        
        labels = []
        workorder_counts = []
        work_hours = []
        efficiency_rates = []
        
        for item in company_data:
            company_name = item['company_name'] or '未指定'
            labels.append(company_name)
            workorder_counts.append(item['workorder_count'])
            work_hours.append(float(item['total_work_hours'] or 0))
            efficiency_rates.append(float(item['avg_efficiency_rate'] or 0))
        
        return {
            'labels': labels,
            'workorder_counts': workorder_counts,
            'work_hours': work_hours,
            'efficiency_rates': efficiency_rates,
        } 

# ==================== 評分報表服務 ====================

class ScoringService:
    """評分報表服務類別"""
    
    @staticmethod
    def calculate_production_score(company_code, start_date, end_date):
        """計算生產效率分數 - 以工序標準產能為基準"""
        from decimal import Decimal
        
        # 取得期間內的作業員工序產能評分資料
        capacity_scores = OperatorProcessCapacityScore.objects.filter(
            company_code=company_code,
            work_date__range=[start_date, end_date]
        )
        
        if not capacity_scores.exists():
            return Decimal('0.00')
        
        # 計算平均產能評分
        avg_capacity_score = capacity_scores.aggregate(
            avg=Avg('capacity_score')
        )['avg'] or Decimal('0.00')
        
        return avg_capacity_score
    
    @staticmethod
    def calculate_quality_score(company_code, start_date, end_date):
        """計算品質管理分數 - 基於作業員工序產能評分"""
        from decimal import Decimal
        
        # 取得期間內的作業員工序產能評分資料
        capacity_scores = OperatorProcessCapacityScore.objects.filter(
            company_code=company_code,
            work_date__range=[start_date, end_date]
        )
        
        if not capacity_scores.exists():
            return Decimal('0.00')
        
        # 計算品質指標
        total_quantity = capacity_scores.aggregate(
            total=Sum('completed_quantity')
        )['total'] or 0
        
        total_defect = capacity_scores.aggregate(
            total=Sum('defect_quantity')
        )['total'] or 0
        
        # 計算不良率
        defect_rate = (total_defect / total_quantity * 100) if total_quantity > 0 else 0
        
        # 計算平均品質評分
        avg_quality_score = capacity_scores.aggregate(
            avg=Avg('quality_score')
        )['avg'] or Decimal('0.00')
        
        # 品質分數 = 平均品質評分 (70%) + 不良率評分 (30%)
        defect_score = max(Decimal('100.00') - Decimal(str(defect_rate)), Decimal('0.00'))
        quality_score = (avg_quality_score * Decimal('0.7') + defect_score * Decimal('0.3'))
        
        return quality_score
    
    @staticmethod
    def calculate_delivery_score(company_code, start_date, end_date):
        """計算交期管理分數"""
        from decimal import Decimal
        
        # 取得期間內的工單資料
        workorder_data = CompletedWorkOrderAnalysis.objects.filter(
            company_code=company_code,
            completion_date__range=[start_date, end_date]
        )
        
        if not workorder_data.exists():
            return Decimal('0.00')
        
        # 計算準時完工率 (假設所有工單都準時完工)
        # 實際應用中需要比對計劃完工日期和實際完工日期
        on_time_count = workorder_data.count()  # 簡化計算
        total_count = workorder_data.count()
        
        on_time_rate = (on_time_count / total_count * 100) if total_count > 0 else 0
        
        return Decimal(str(on_time_rate))
    
    @staticmethod
    def calculate_equipment_score(company_code, start_date, end_date):
        """計算設備管理分數"""
        from decimal import Decimal
        
        # 取得期間內的工單資料
        workorder_data = CompletedWorkOrderAnalysis.objects.filter(
            company_code=company_code,
            completion_date__range=[start_date, end_date]
        )
        
        if not workorder_data.exists():
            return Decimal('0.00')
        
        # 計算設備利用率
        total_equipment_hours = workorder_data.aggregate(
            total=Sum('total_work_hours')
        )['total'] or Decimal('0.00')
        
        # 假設標準工作時數為8小時/天，工作天數為期間天數
        from datetime import timedelta
        work_days = (end_date - start_date).days + 1
        standard_hours = work_days * Decimal('8.00')
        
        # 設備利用率 (簡化計算)
        equipment_utilization = min((total_equipment_hours / standard_hours * 100), Decimal('100.00'))
        
        return equipment_utilization
    
    @staticmethod
    def calculate_cost_score(company_code, start_date, end_date):
        """計算成本控制分數"""
        from decimal import Decimal
        
        # 取得期間內的工單資料
        workorder_data = CompletedWorkOrderAnalysis.objects.filter(
            company_code=company_code,
            completion_date__range=[start_date, end_date]
        )
        
        if not workorder_data.exists():
            return Decimal('0.00')
        
        # 計算加班時數比例
        total_work_hours = workorder_data.aggregate(
            total=Sum('total_work_hours')
        )['total'] or Decimal('0.00')
        
        total_overtime_hours = workorder_data.aggregate(
            total=Sum('total_overtime_hours')
        )['total'] or Decimal('0.00')
        
        # 加班比例越低分數越高
        overtime_rate = (total_overtime_hours / total_work_hours * 100) if total_work_hours > 0 else 0
        cost_score = max(Decimal('100.00') - Decimal(str(overtime_rate)), Decimal('0.00'))
        
        return cost_score
    
    @staticmethod
    def calculate_safety_score(company_code, start_date, end_date):
        """計算安全管理分數"""
        from decimal import Decimal
        
        # 這裡可以整合安全事件資料
        # 目前簡化為固定分數
        safety_score = Decimal('95.00')  # 假設安全狀況良好
        
        return safety_score
    
    @staticmethod
    def calculate_personnel_score(company_code, start_date, end_date):
        """計算人員管理分數"""
        from decimal import Decimal
        
        # 取得期間內的工單資料
        workorder_data = CompletedWorkOrderAnalysis.objects.filter(
            company_code=company_code,
            completion_date__range=[start_date, end_date]
        )
        
        if not workorder_data.exists():
            return Decimal('0.00')
        
        # 計算人員效率
        total_operators = workorder_data.aggregate(
            total=Sum('unique_operators_count')
        )['total'] or 0
        
        total_work_hours = workorder_data.aggregate(
            total=Sum('total_work_hours')
        )['total'] or Decimal('0.00')
        
        # 人員效率 = 總工作時數 / 總人員數 (簡化計算)
        if total_operators > 0:
            personnel_efficiency = (total_work_hours / total_operators) / Decimal('8.00') * 100
            personnel_score = min(personnel_efficiency, Decimal('100.00'))
        else:
            personnel_score = Decimal('0.00')
        
        return personnel_score
    
    @staticmethod
    def generate_scoring_report(company_code, start_date, end_date, report_period='monthly'):
        """生成評分報表"""
        from decimal import Decimal
        from datetime import datetime
        
        # 計算生產效率分數
        production_score = ScoringService.calculate_production_score(company_code, start_date, end_date)
        
        # 生產效率佔100%權重
        total_score = production_score
        
        # 計算得分百分比
        score_percentage = total_score
        
        # 計算總體等級
        if score_percentage >= 90:
            overall_grade = '優秀'
        elif score_percentage >= 80:
            overall_grade = '良好'
        elif score_percentage >= 70:
            overall_grade = '及格'
        else:
            overall_grade = '不及格'
        
        # 統計各等級項目數
        scores = [production_score]
        
        excellent_count = sum(1 for score in scores if score >= 90)
        good_count = sum(1 for score in scores if 80 <= score < 90)
        pass_count = sum(1 for score in scores if 70 <= score < 80)
        fail_count = sum(1 for score in scores if score < 70)
        
        # 建立報表名稱
        report_name = f"{start_date.strftime('%Y年%m月%d日')} 至 {end_date.strftime('%Y年%m月%d日')} 評分報表"
        
        # 取得公司名稱
        from erp_integration.models import CompanyConfig
        try:
            company = CompanyConfig.objects.get(company_code=company_code)
            company_name = company.company_name
        except CompanyConfig.DoesNotExist:
            company_name = company_code
        
        # 暫時返回一個字典，因為ScoringReport模型已移除
        return {
            'id': None,
            'report_name': report_name,
            'company_name': company_name,
            'total_score': total_score,
            'score_percentage': score_percentage,
            'overall_grade': overall_grade,
            'production_score': production_score,
            'excellent_count': excellent_count,
            'good_count': good_count,
            'pass_count': pass_count,
            'fail_count': fail_count,
        }
    
    @staticmethod
    def generate_improvement_suggestions(scoring_report):
        """生成改善建議"""
        # 暫時返回空列表，因為ScoringImprovement模型已移除
        return []

# ==================== 作業員工序產能評分服務 ====================

class OperatorCapacityService:
    """作業員工序產能評分服務類別"""
    
    @staticmethod
    def calculate_operator_process_score(operator_name, operator_id, company_code, 
                                       product_code, process_name, workorder_id, 
                                       work_date, work_hours, completed_quantity, 
                                       defect_quantity=0, supervisor_name=None, 
                                       supervisor_score=None, supervisor_comment=None):
        """計算作業員工序評分"""
        from decimal import Decimal
        from datetime import datetime
        
        # 取得標準產能
        try:
            from process.models import ProductProcessStandardCapacity
            standard_capacity = ProductProcessStandardCapacity.objects.get(
                company_code=company_code,
                product_code=product_code,
                process_name=process_name
            )
            standard_capacity_per_hour = standard_capacity.standard_capacity_per_hour
        except ProductProcessStandardCapacity.DoesNotExist:
            # 如果沒有標準產能，使用預設值
            standard_capacity_per_hour = Decimal('1.00')
        
        # 計算實際產能
        if work_hours > 0:
            actual_capacity_per_hour = Decimal(str(completed_quantity)) / work_hours
        else:
            actual_capacity_per_hour = Decimal('0.00')
        
        # 計算產能比率
        if standard_capacity_per_hour > 0:
            capacity_ratio = actual_capacity_per_hour / standard_capacity_per_hour
        else:
            capacity_ratio = Decimal('0.00')
        
        # 計算效率因子和學習曲線因子
        efficiency_factor = min(Decimal('1.20'), capacity_ratio)
        learning_curve_factor = min(Decimal('1.10'), capacity_ratio)
        
        # 計算不良率
        if completed_quantity > 0:
            defect_rate = Decimal(str(defect_quantity)) / completed_quantity
        else:
            defect_rate = Decimal('0.00')
        
        # 建立或更新評分記錄
        score_record, created = OperatorProcessCapacityScore.objects.get_or_create(
            operator_id=operator_id,
            company_code=company_code,
            product_code=product_code,
            process_name=process_name,
            workorder_id=workorder_id,
            work_date=work_date,
            defaults={
                'operator_name': operator_name,
                'work_hours': work_hours,
                'standard_capacity_per_hour': standard_capacity_per_hour,
                'actual_capacity_per_hour': actual_capacity_per_hour,
                'completed_quantity': completed_quantity,
                'capacity_ratio': capacity_ratio,
                'efficiency_factor': efficiency_factor,
                'learning_curve_factor': learning_curve_factor,
                'defect_quantity': defect_quantity,
                'defect_rate': defect_rate,
                'supervisor_score': Decimal('80.00'),  # 預設80分
                'is_supervisor_scored': False,  # 預設未評分
            }
        )
        
        if not created:
            # 更新現有記錄
            score_record.operator_name = operator_name
            score_record.work_hours = work_hours
            score_record.standard_capacity_per_hour = standard_capacity_per_hour
            score_record.actual_capacity_per_hour = actual_capacity_per_hour
            score_record.completed_quantity = completed_quantity
            score_record.capacity_ratio = capacity_ratio
            score_record.efficiency_factor = efficiency_factor
            score_record.learning_curve_factor = learning_curve_factor
            score_record.defect_quantity = defect_quantity
            score_record.defect_rate = defect_rate
        
        # 如果提供了主管評分，更新主管評分資訊
        if supervisor_score is not None and supervisor_name:
            score_record.supervisor_score = supervisor_score
            score_record.supervisor_comment = supervisor_comment or ''
            score_record.supervisor_name = supervisor_name
            score_record.supervisor_date = datetime.now()
            score_record.is_supervisor_scored = True
        
        # 計算評分
        score_record.capacity_score = score_record.calculate_capacity_score()
        # 品質評分不再自動計算，由主管手動評分
        score_record.total_score = score_record.calculate_total_score()
        score_record.grade = score_record.get_grade(score_record.capacity_score)
        score_record.overall_grade = score_record.get_grade(score_record.total_score)
        
        score_record.save()
        return score_record
    
    @staticmethod
    def generate_operator_capacity_report(company_code, start_date, end_date, report_period='monthly'):
        """生成作業員產能評分報表"""
        from decimal import Decimal
        from datetime import datetime
        from django.db.models import Avg, Count, Sum
        
        # 取得期間內的評分資料
        scores = OperatorProcessCapacityScore.objects.filter(
            company_code=company_code,
            work_date__range=[start_date, end_date]
        )
        
        if not scores.exists():
            return None
        
        # 計算統計資料
        total_operators = scores.values('operator_id').distinct().count()
        total_processes = scores.values('process_name').distinct().count()
        total_work_hours = scores.aggregate(total=Sum('work_hours'))['total'] or Decimal('0.00')
        total_completed_quantity = scores.aggregate(total=Sum('completed_quantity'))['total'] or 0
        
        # 計算平均評分
        avg_capacity_score = scores.aggregate(avg=Avg('capacity_score'))['avg'] or Decimal('0.00')
        avg_quality_score = scores.aggregate(avg=Avg('quality_score'))['avg'] or Decimal('0.00')
        avg_total_score = scores.aggregate(avg=Avg('total_score'))['avg'] or Decimal('0.00')
        
        # 計算等級統計
        excellent_count = scores.filter(overall_grade='優秀').count()
        good_count = scores.filter(overall_grade='良好').count()
        pass_count = scores.filter(overall_grade='及格').count()
        fail_count = scores.filter(overall_grade='不及格').count()
        
        # 計算工序表現統計
        process_performance = {}
        process_stats = scores.values('process_name').annotate(
            avg_capacity=Avg('capacity_score'),
            avg_quality=Avg('quality_score'),
            avg_total=Avg('total_score'),
            operator_count=Count('operator_id', distinct=True),
            total_hours=Sum('work_hours'),
            total_quantity=Sum('completed_quantity')
        )
        
        for stat in process_stats:
            process_performance[stat['process_name']] = {
                'avg_capacity': float(stat['avg_capacity'] or 0),
                'avg_quality': float(stat['avg_quality'] or 0),
                'avg_total': float(stat['avg_total'] or 0),
                'operator_count': stat['operator_count'],
                'total_hours': float(stat['total_hours'] or 0),
                'total_quantity': stat['total_quantity'] or 0,
            }
        
        # 建立報表名稱
        report_name = f"{start_date.strftime('%Y年%m月%d日')} 至 {end_date.strftime('%Y年%m月%d日')} 作業員產能評分報表"
        
        # 建立或更新報表
        report, created = OperatorCapacityReport.objects.get_or_create(
            company_code=company_code,
            report_period=report_period,
            start_date=start_date,
            end_date=end_date,
            defaults={
                'report_name': report_name,
                'report_date': datetime.now().date(),
                'total_operators': total_operators,
                'total_processes': total_processes,
                'total_work_hours': total_work_hours,
                'total_completed_quantity': total_completed_quantity,
                'avg_capacity_score': avg_capacity_score,
                'avg_quality_score': avg_quality_score,
                'avg_total_score': avg_total_score,
                'excellent_count': excellent_count,
                'good_count': good_count,
                'pass_count': pass_count,
                'fail_count': fail_count,
                'process_performance': process_performance,
            }
        )
        
        if not created:
            # 更新現有記錄
            report.report_name = report_name
            report.total_operators = total_operators
            report.total_processes = total_processes
            report.total_work_hours = total_work_hours
            report.total_completed_quantity = total_completed_quantity
            report.avg_capacity_score = avg_capacity_score
            report.avg_quality_score = avg_quality_score
            report.avg_total_score = avg_total_score
            report.excellent_count = excellent_count
            report.good_count = good_count
            report.pass_count = pass_count
            report.fail_count = fail_count
            report.process_performance = process_performance
            report.save()
        
        return report
    
    @staticmethod
    def get_operator_performance_summary(company_code, start_date, end_date, operator_id=None):
        """取得作業員表現摘要"""
        from django.db.models import Avg, Count, Sum
        
        # 建立查詢條件
        query = {
            'company_code': company_code,
            'work_date__range': [start_date, end_date]
        }
        
        if operator_id:
            query['operator_id'] = str(operator_id)
        
        scores = OperatorProcessCapacityScore.objects.filter(**query)
        
        if not scores.exists():
            return None
        
        # 計算摘要統計
        summary = scores.aggregate(
            total_records=Count('id'),
            avg_capacity_score=Avg('capacity_score'),
            avg_quality_score=Avg('quality_score'),
            avg_total_score=Avg('total_score'),
            total_work_hours=Sum('work_hours'),
            total_completed_quantity=Sum('completed_quantity'),
            total_defect_quantity=Sum('defect_quantity'),
        )
        
        # 計算等級分布
        grade_distribution = {
            'excellent': scores.filter(overall_grade='優秀').count(),
            'good': scores.filter(overall_grade='良好').count(),
            'pass': scores.filter(overall_grade='及格').count(),
            'fail': scores.filter(overall_grade='不及格').count(),
        }
        
        # 計算工序表現排名
        process_ranking = scores.values('process_name').annotate(
            avg_score=Avg('total_score'),
            record_count=Count('id')
        ).order_by('-avg_score')
        
        # 計算作業員表現排名
        operator_ranking = scores.values('operator_name', 'operator_id').annotate(
            avg_score=Avg('total_score'),
            total_hours=Sum('work_hours'),
            total_quantity=Sum('completed_quantity'),
            record_count=Count('id')
        ).order_by('-avg_score')
        
        return {
            'summary': summary,
            'grade_distribution': grade_distribution,
            'process_ranking': list(process_ranking),
            'operator_ranking': list(operator_ranking),
        }
    
    @staticmethod
    def get_operator_process_details(operator_id, start_date, end_date):
        """取得作業員工序詳細資料"""
        scores = OperatorProcessCapacityScore.objects.filter(
            operator_id=str(operator_id),
            work_date__range=[start_date, end_date]
        ).order_by('work_date', 'process_name')
        
        return list(scores.values(
            'work_date', 'process_name', 'product_code', 'workorder_id',
            'work_hours', 'completed_quantity', 'defect_quantity',
            'standard_capacity_per_hour', 'actual_capacity_per_hour',
            'capacity_ratio', 'capacity_score', 'quality_score', 'total_score',
            'overall_grade'
        ))
    
    @staticmethod
    def get_process_capacity_analysis(company_code, start_date, end_date, process_name=None):
        """取得工序產能分析"""
        from django.db.models import Avg, Count, Sum, StdDev
        
        query = {
            'company_code': company_code,
            'work_date__range': [start_date, end_date]
        }
        
        if process_name:
            query['process_name'] = process_name
        
        scores = OperatorProcessCapacityScore.objects.filter(**query)
        
        if not scores.exists():
            return None
        
        # 按工序分組統計
        process_stats = scores.values('process_name').annotate(
            operator_count=Count('operator_id', distinct=True),
            total_records=Count('id'),
            avg_capacity_score=Avg('capacity_score'),
            avg_quality_score=Avg('quality_score'),
            avg_total_score=Avg('total_score'),
            avg_capacity_ratio=Avg('capacity_ratio'),
            total_work_hours=Sum('work_hours'),
            total_completed_quantity=Sum('completed_quantity'),
            total_defect_quantity=Sum('defect_quantity'),
            capacity_score_std=StdDev('capacity_score'),
            quality_score_std=StdDev('quality_score'),
        )
        
        # 計算整體統計
        overall_stats = scores.aggregate(
            total_operators=Count('operator_id', distinct=True),
            total_processes=Count('process_name', distinct=True),
            avg_capacity_score=Avg('capacity_score'),
            avg_quality_score=Avg('quality_score'),
            avg_total_score=Avg('total_score'),
            avg_capacity_ratio=Avg('capacity_ratio'),
            total_work_hours=Sum('work_hours'),
            total_completed_quantity=Sum('completed_quantity'),
            total_defect_quantity=Sum('defect_quantity'),
        )
        
        return {
            'process_stats': list(process_stats),
            'overall_stats': overall_stats,
        } 

# ==================== 評分週期管理服務 ====================

class ScorePeriodService:
    """評分週期管理服務"""
    
    @staticmethod
    def get_current_period_dates(period_type='monthly'):
        """取得當前評分週期的日期範圍"""
        from datetime import datetime, timedelta
        from calendar import monthrange
        
        today = datetime.now().date()
        
        if period_type == 'monthly':
            # 當月第一天到最後一天
            start_date = today.replace(day=1)
            _, last_day = monthrange(today.year, today.month)
            end_date = today.replace(day=last_day)
            
        elif period_type == 'quarterly':
            # 當前季度
            quarter = (today.month - 1) // 3 + 1
            start_month = (quarter - 1) * 3 + 1
            end_month = quarter * 3
            
            start_date = today.replace(month=start_month, day=1)
            _, last_day = monthrange(today.year, end_month)
            end_date = today.replace(month=end_month, day=last_day)
            
        elif period_type == 'yearly':
            # 當年
            start_date = today.replace(month=1, day=1)
            end_date = today.replace(month=12, day=31)
            
        else:
            raise ValueError(f"不支援的評分週期類型: {period_type}")
        
        return start_date, end_date
    
    @staticmethod
    def get_period_name(period_type, start_date, end_date):
        """取得評分週期名稱"""
        if period_type == 'monthly':
            return f"{start_date.year}年{start_date.month}月評分"
        elif period_type == 'quarterly':
            quarter = (start_date.month - 1) // 3 + 1
            return f"{start_date.year}年第{quarter}季評分"
        elif period_type == 'yearly':
            return f"{start_date.year}年度評分"
        else:
            return f"{start_date}至{end_date}評分"
    
    @staticmethod
    def create_period_score_records(company_code, period_type='monthly', force=False):
        """建立評分週期記錄"""
        from datetime import datetime
        
        start_date, end_date = ScorePeriodService.get_current_period_dates(period_type)
        
        # 檢查是否已存在該週期的記錄
        existing_records = OperatorProcessCapacityScore.objects.filter(
            company_code=company_code,
            score_period=period_type,
            period_start_date=start_date,
            period_end_date=end_date
        )
        
        if existing_records.exists() and not force:
            return existing_records
        
        # 取得該週期內的所有作業員工作記錄
        from workorder.fill_work.models import FillWork
        from workorder.onsite_reporting.models import OnsiteReport
        
        # 從填報資料建立評分記錄
        fill_works = FillWork.objects.filter(
            company_name=company_code,
            work_date__range=[start_date, end_date]
        )
        
        created_records = []
        for fill_work in fill_works:
            # 計算作業員評分
            score_record = OperatorCapacityService.calculate_operator_process_score(
                operator_name=fill_work.operator_name,
                operator_id=fill_work.operator_id,
                company_code=company_code,
                product_code=fill_work.product_code,
                process_name=fill_work.process_name,
                workorder_id=fill_work.workorder,
                work_date=fill_work.work_date,
                work_hours=fill_work.work_hours_calculated or 0,
                completed_quantity=fill_work.completed_quantity or 0,
                defect_quantity=fill_work.defect_quantity or 0
            )
            
            if score_record:
                # 設定評分週期資訊
                score_record.score_period = period_type
                score_record.period_start_date = start_date
                score_record.period_end_date = end_date
                score_record.save()
                created_records.append(score_record)
        
        # 從現場報工資料建立評分記錄
        onsite_reports = OnsiteReport.objects.filter(
            company_code=company_code,
            work_date__range=[start_date, end_date]
        )
        
        for report in onsite_reports:
            score_record = OperatorCapacityService.calculate_operator_process_score(
                operator_name=report.operator_name,
                operator_id=report.operator_id,
                company_code=company_code,
                product_code=report.product_code,
                process_name=report.process_name,
                workorder_id=report.workorder_id,
                work_date=report.work_date,
                work_hours=report.work_hours or 0,
                completed_quantity=report.completed_quantity or 0,
                defect_quantity=report.defect_quantity or 0
            )
            
            if score_record:
                # 設定評分週期資訊
                score_record.score_period = period_type
                score_record.period_start_date = start_date
                score_record.period_end_date = end_date
                score_record.save()
                created_records.append(score_record)
        
        return created_records
    
    @staticmethod
    def close_period(company_code, period_type='monthly'):
        """關閉評分週期"""
        from datetime import datetime
        
        start_date, end_date = ScorePeriodService.get_current_period_dates(period_type)
        
        # 更新該週期的所有記錄為已關閉
        updated_count = OperatorProcessCapacityScore.objects.filter(
            company_code=company_code,
            score_period=period_type,
            period_start_date=start_date,
            period_end_date=end_date
        ).update(
            is_period_closed=True,
            period_closed_date=datetime.now()
        )
        
        return updated_count
    
    @staticmethod
    def get_period_summary(company_code, period_type='monthly'):
        """取得評分週期摘要"""
        start_date, end_date = ScorePeriodService.get_current_period_dates(period_type)
        
        records = OperatorProcessCapacityScore.objects.filter(
            company_code=company_code,
            score_period=period_type,
            period_start_date=start_date,
            period_end_date=end_date
        )
        
        if not records.exists():
            return None
        
        # 計算統計資料
        total_records = records.count()
        avg_capacity_score = records.aggregate(avg=Avg('capacity_score'))['avg'] or 0
        avg_quality_score = records.aggregate(avg=Avg('quality_score'))['avg'] or 0
        avg_total_score = records.aggregate(avg=Avg('total_score'))['avg'] or 0
        
        # 等級分布
        grade_distribution = records.values('overall_grade').annotate(
            count=Count('id')
        ).order_by('overall_grade')
        
        # 主管評分統計
        supervisor_scored_count = records.filter(is_supervisor_scored=True).count()
        supervisor_score_rate = (supervisor_scored_count / total_records * 100) if total_records > 0 else 0
        
        return {
            'period_name': ScorePeriodService.get_period_name(period_type, start_date, end_date),
            'start_date': start_date,
            'end_date': end_date,
            'total_records': total_records,
            'avg_capacity_score': avg_capacity_score,
            'avg_quality_score': avg_quality_score,
            'avg_total_score': avg_total_score,
            'grade_distribution': grade_distribution,
            'supervisor_scored_count': supervisor_scored_count,
            'supervisor_score_rate': supervisor_score_rate,
            'is_closed': records.filter(is_period_closed=True).exists(),
        } 

class WorkdayCalendarService:
    """工作日曆服務 - 基於現有 Event 模型"""
    
    def __init__(self, company_code=None):
        self.company_code = company_code
    
    def is_workday(self, date):
        """
        判斷指定日期是否為工作日
        
        Args:
            date: 要檢查的日期 (date 物件)
            
        Returns:
            bool: True 表示工作日，False 表示非工作日
        """
        from scheduling.models import Event
        from datetime import datetime, time
        
        # 檢查是否有放假日事件
        holiday_events = Event.objects.filter(
            type='holiday',
            start__date=date,
            all_day=True
        )
        
        if holiday_events.exists():
            logger.info(f"日期 {date} 被標記為放假日")
            return False
        
        # 檢查是否有上班日事件（覆蓋週末）
        workday_events = Event.objects.filter(
            type='workday',
            start__date=date,
            all_day=True
        )
        
        if workday_events.exists():
            logger.info(f"日期 {date} 被標記為補班日")
            return True
        
        # 預設邏輯：週一至週五為工作日
        is_default_workday = date.weekday() < 5
        logger.debug(f"日期 {date} 預設工作日判斷: {is_default_workday}")
        return is_default_workday
    
    def get_previous_workday(self, current_date):
        """
        取得前一個工作日
        
        Args:
            current_date: 當前日期 (date 物件)
            
        Returns:
            date: 前一個工作日的日期
        """
        from datetime import timedelta
        
        check_date = current_date - timedelta(days=1)
        
        # 最多往前找30天（處理長假）
        for i in range(30):
            if self.is_workday(check_date):
                logger.info(f"找到前一個工作日: {check_date}")
                return check_date
            check_date = check_date - timedelta(days=1)
        
        logger.warning(f"無法找到前一個工作日，返回當前日期: {current_date}")
        return current_date  # 兜底方案
    
    def add_holiday(self, date, holiday_name, description="", created_by="system"):
        """
        新增假期
        
        Args:
            date: 假期日期
            holiday_name: 假期名稱
            description: 假期描述
            created_by: 建立者
        """
        from scheduling.models import Event
        from datetime import datetime, time
        
        # 檢查是否已存在
        existing_event = Event.objects.filter(
            type='holiday',
            start__date=date,
            all_day=True
        ).first()
        
        if existing_event:
            logger.info(f"假期 {holiday_name} 在 {date} 已存在")
            return existing_event
        
        # 建立新的假期事件（使用時區感知的 datetime）
        start_datetime = timezone.make_aware(datetime.combine(date, time.min))
        end_datetime = timezone.make_aware(datetime.combine(date, time.max))
        
        event = Event.objects.create(
            title=holiday_name,
            start=start_datetime,
            end=end_datetime,
            type='holiday',
            description=description,
            all_day=True,
            created_by=created_by
        )
        
        logger.info(f"成功新增假期: {holiday_name} 在 {date}")
        return event
    
    def add_workday(self, date, description="", created_by="system"):
        """
        新增工作日（覆蓋週末）
        
        Args:
            date: 工作日日期
            description: 工作日描述
            created_by: 建立者
        """
        from scheduling.models import Event
        from datetime import datetime, time
        
        # 檢查是否已存在
        existing_event = Event.objects.filter(
            type='workday',
            start__date=date,
            all_day=True
        ).first()
        
        if existing_event:
            logger.info(f"工作日事件在 {date} 已存在")
            return existing_event
        
        # 建立新的工作日事件（使用時區感知的 datetime）
        start_datetime = timezone.make_aware(datetime.combine(date, time.min))
        end_datetime = timezone.make_aware(datetime.combine(date, time.max))
        
        event = Event.objects.create(
            title=f"補班日 - {date.strftime('%Y-%m-%d')}",
            start=start_datetime,
            end=end_datetime,
            type='workday',
            description=description,
            all_day=True,
            created_by=created_by
        )
        
        logger.info(f"成功新增補班日: {date}")
        return event
    
    def get_workdays_in_range(self, start_date, end_date):
        """
        取得日期範圍內的工作日
        
        Args:
            start_date: 開始日期
            end_date: 結束日期
            
        Returns:
            list: 工作日日期列表
        """
        workdays = []
        current_date = start_date
        
        while current_date <= end_date:
            if self.is_workday(current_date):
                workdays.append(current_date)
            current_date += timedelta(days=1)
        
        logger.info(f"日期範圍 {start_date} 到 {end_date} 內有 {len(workdays)} 個工作日")
        return workdays
    
    def get_holidays_in_range(self, start_date, end_date):
        """
        取得日期範圍內的假期
        
        Args:
            start_date: 開始日期
            end_date: 結束日期
            
        Returns:
            list: 假期事件列表
        """
        from scheduling.models import Event
        
        holidays = Event.objects.filter(
            type='holiday',
            start__date__gte=start_date,
            start__date__lte=end_date,
            all_day=True
        ).order_by('start')
        
        logger.info(f"日期範圍 {start_date} 到 {end_date} 內有 {holidays.count()} 個假期")
        return holidays

class HolidayAutoSetupService:
    """自動假期設定服務"""
    
    def __init__(self):
        self.calendar_service = WorkdayCalendarService()
    
    def setup_2025_holidays(self):
        """設定 2025 年台灣國定假日"""
        holidays_2025 = {
            '2025-01-01': '元旦',
            '2025-02-08': '春節',
            '2025-02-09': '春節',
            '2025-02-10': '春節',
            '2025-02-11': '春節',
            '2025-02-12': '春節',
            '2025-02-13': '春節',
            '2025-02-14': '春節',
            '2025-04-04': '清明節',
            '2025-04-05': '清明節',
            '2025-04-06': '清明節',
            '2025-05-01': '勞動節',
            '2025-06-22': '端午節',
            '2025-09-29': '中秋節',
            '2025-10-01': '國慶節',
            '2025-10-02': '國慶節',
            '2025-10-03': '國慶節',
            '2025-10-04': '國慶節',
            '2025-10-05': '國慶節',
            '2025-10-06': '國慶節',
            '2025-10-07': '國慶節',
        }
        
        setup_count = 0
        for date_str, holiday_name in holidays_2025.items():
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                self.calendar_service.add_holiday(
                    date=date_obj,
                    holiday_name=holiday_name,
                    description=f"2025年{holiday_name}",
                    created_by="system"
                )
                setup_count += 1
                logger.info(f"成功設定假期: {holiday_name} ({date_str})")
            except Exception as e:
                logger.error(f"設定假期失敗: {holiday_name} ({date_str}) - {str(e)}")
        
        logger.info(f"2025年假期設定完成，共設定 {setup_count} 個假期")
        return setup_count
    
    def setup_makeup_workdays_2025(self):
        """設定 2025 年補班日"""
        makeup_days_2025 = {
            '2025-02-15': '春節補班',
            '2025-02-16': '春節補班',
            '2025-04-06': '清明節補班',
            '2025-09-28': '中秋節補班',
            '2025-10-11': '國慶節補班',
            '2025-10-12': '國慶節補班',
        }
        
        setup_count = 0
        for date_str, description in makeup_days_2025.items():
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                self.calendar_service.add_workday(
                    date=date_obj,
                    description=description,
                    created_by="system"
                )
                setup_count += 1
                logger.info(f"成功設定補班日: {description} ({date_str})")
            except Exception as e:
                logger.error(f"設定補班日失敗: {description} ({date_str}) - {str(e)}")
        
        logger.info(f"2025年補班日設定完成，共設定 {setup_count} 個補班日")
        return setup_count
    
    def setup_weekly_holidays(self, year=2025):
        """設定週末假期（週六、週日）"""
        from datetime import date, timedelta
        
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        
        weekend_count = 0
        current_date = start_date
        
        while current_date <= end_date:
            # 週六和週日
            if current_date.weekday() >= 5:
                try:
                    self.calendar_service.add_holiday(
                        date=current_date,
                        holiday_name=f"週末",
                        description=f"{year}年週末",
                        created_by="system"
                    )
                    weekend_count += 1
                except Exception as e:
                    logger.error(f"設定週末假期失敗: {current_date} - {str(e)}")
            
            current_date += timedelta(days=1)
        
        logger.info(f"{year}年週末假期設定完成，共設定 {weekend_count} 個週末")
        return weekend_count
    
    def setup_all_2025(self):
        """設定 2025 年所有假期和補班日"""
        logger.info("開始設定 2025 年假期和補班日...")
        
        # 設定國定假日
        holiday_count = self.setup_2025_holidays()
        
        # 設定補班日
        makeup_count = self.setup_makeup_workdays_2025()
        
        # 設定週末（可選，因為預設邏輯已經處理週末）
        # weekend_count = self.setup_weekly_holidays(2025)
        
        total_count = holiday_count + makeup_count
        logger.info(f"2025年假期設定完成，總共設定 {total_count} 個事件")
        
        return {
            'holidays': holiday_count,
            'makeup_days': makeup_count,
            'total': total_count
        }
    
    def clear_all_holidays(self, year=2025):
        """清除指定年份的所有假期設定"""
        from scheduling.models import Event
        from datetime import date
        
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        
        # 刪除假期和補班日
        deleted_count = Event.objects.filter(
            type__in=['holiday', 'workday'],
            start__date__gte=start_date,
            start__date__lte=end_date,
            created_by='system'
        ).delete()[0]
        
        logger.info(f"清除 {year} 年假期設定完成，共刪除 {deleted_count} 個事件")
        return deleted_count

class PreviousWorkdayReportScheduler:
    """前一個工作日報表排程服務"""
    
    def __init__(self):
        self.calendar_service = WorkdayCalendarService()
    
    def get_report_date(self):
        """
        取得報表日期（前一個工作日）
        
        Returns:
            date: 前一個工作日的日期
        """
        # 修正時區問題，確保使用正確的本地日期
        current_datetime = timezone.localtime(timezone.now())
        current_date = current_datetime.date()
        return self.calendar_service.get_previous_workday(current_date)
    
    def should_execute(self):
        """
        判斷是否應該執行報表生成
        
        Returns:
            bool: True 表示應該執行，False 表示不應該執行
        """
        # 修正時區問題，確保使用正確的本地時間
        current_datetime = timezone.localtime(timezone.now())
        current_time = current_datetime.time()
        # 上午 10:30 後執行，確保填報資料已收集完成
        return current_time >= time(10, 30)
    
    def collect_data(self, report_date):
        """
        收集前一個工作日的資料
        
        Args:
            report_date: 報表日期
            
        Returns:
            dict: 收集到的資料
        """
        from workorder.fill_work.models import FillWork
        from workorder.onsite_reporting.models import OnsiteReport
        
        # 收集填報資料
        fill_works = FillWork.objects.filter(
            work_date=report_date,
            approval_status='approved'
        )
        
        # 收集現場報工資料
        onsite_reports = OnsiteReport.objects.filter(
            work_date=report_date,
            status='completed'
        )
        
        # 統計資料
        data = {
            'report_date': report_date,
            'fill_works_count': fill_works.count(),
            'onsite_reports_count': onsite_reports.count(),
            'fill_works': list(fill_works.values()),
            'onsite_reports': list(onsite_reports.values()),
            'total_work_hours': 0,
            'total_operators': 0,
            'total_equipment': 0,
        }
        
        # 計算總工作時數
        for fill_work in fill_works:
            data['total_work_hours'] += float(fill_work.work_hours or 0)
        
        for onsite_report in onsite_reports:
            data['total_work_hours'] += float(onsite_report.work_hours or 0)
        
        # 統計作業員數量（去重）
        operator_names = set()
        for fill_work in fill_works:
            if fill_work.operator_name:
                operator_names.add(fill_work.operator_name)
        
        for onsite_report in onsite_reports:
            if onsite_report.operator_name:
                operator_names.add(onsite_report.operator_name)
        
        data['total_operators'] = len(operator_names)
        
        # 統計設備數量（去重）
        equipment_ids = set()
        for fill_work in fill_works:
            if fill_work.equipment_id:
                equipment_ids.add(fill_work.equipment_id)
        
        for onsite_report in onsite_reports:
            if onsite_report.equipment_id:
                equipment_ids.add(onsite_report.equipment_id)
        
        data['total_equipment'] = len(equipment_ids)
        
        logger.info(f"收集 {report_date} 的資料完成：填報 {data['fill_works_count']} 筆，現場報工 {data['onsite_reports_count']} 筆")
        return data
    
    def generate_report(self, data):
        """
        生成前一個工作日報表
        
        Args:
            data: 收集到的資料
            
        Returns:
            str: 報表檔案路徑
        """
        import os
        from datetime import datetime
        from django.conf import settings
        
        # 建立報表目錄
        report_dir = os.path.join(settings.MEDIA_ROOT, 'reports', 'previous_workday')
        os.makedirs(report_dir, exist_ok=True)
        
        # 生成報表檔案名稱
        report_date = data['report_date']
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"previous_workday_report_{report_date}_{timestamp}.html"
        filepath = os.path.join(report_dir, filename)
        
        # 生成 HTML 報表
        html_content = self._generate_html_report(data)
        
        # 寫入檔案
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        logger.info(f"前一個工作日報表已生成：{filepath}")
        return filepath
    
    def _generate_html_report(self, data):
        """生成 HTML 報表內容"""
        report_date = data['report_date']
        
        html = f"""
<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>前一個工作日報表 - {report_date}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        .header {{ text-align: center; margin-bottom: 30px; }}
        .summary {{ margin-bottom: 30px; }}
        .summary-item {{ margin: 10px 0; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; }}
        th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
        th {{ background-color: #f2f2f2; }}
        .footer {{ margin-top: 30px; text-align: center; color: #666; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>前一個工作日報表</h1>
        <h2>報表日期：{report_date}</h2>
        <p>生成時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    </div>
    
    <div class="summary">
        <h3>統計摘要</h3>
        <div class="summary-item">
            <strong>填報記錄數：</strong> {data['fill_works_count']} 筆
        </div>
        <div class="summary-item">
            <strong>現場報工記錄數：</strong> {data['onsite_reports_count']} 筆
        </div>
        <div class="summary-item">
            <strong>總工作時數：</strong> {data['total_work_hours']:.2f} 小時
        </div>
        <div class="summary-item">
            <strong>參與作業員數：</strong> {data['total_operators']} 人
        </div>
        <div class="summary-item">
            <strong>使用設備數：</strong> {data['total_equipment']} 台
        </div>
    </div>
    
    <div class="footer">
        <p>此報表由 MES 系統自動生成</p>
    </div>
</body>
</html>
        """
        
        return html
    
    def send_report(self, report_file, report_date):
        """
        發送報表
        
        Args:
            report_file: 報表檔案路徑
            report_date: 報表日期
        """
        # 這裡可以實作 Email 發送功能
        # 目前先記錄到日誌
        logger.info(f"前一個工作日報表已準備發送：{report_file}")
        logger.info(f"報表日期：{report_date}")
        
        # TODO: 實作 Email 發送功能
        # 可以從設定檔讀取收件人清單
        # 可以根據報表內容調整收件人
    
    def execute(self):
        """
        執行前一個工作日報表生成和發送
        """
        if not self.should_execute():
            logger.info("當前時間未到執行時間（10:30），跳過報表生成")
            return False
        
        try:
            # 取得報表日期
            report_date = self.get_report_date()
            logger.info(f"開始生成前一個工作日報表，日期：{report_date}")
            
            # 收集資料
            data = self.collect_data(report_date)
            
            # 檢查是否有資料
            if data['fill_works_count'] == 0 and data['onsite_reports_count'] == 0:
                logger.warning(f"{report_date} 沒有找到任何填報或現場報工資料")
                return False
            
            # 生成報表
            report_file = self.generate_report(data)
            
            # 發送報表
            self.send_report(report_file, report_date)
            
            logger.info(f"前一個工作日報表執行完成：{report_date}")
            return True
            
        except Exception as e:
            logger.error(f"前一個工作日報表執行失敗：{str(e)}")
            return False

class TaiwanGovernmentCalendarService:
    """台灣政府行事曆 API 整合服務"""
    
    def __init__(self):
        self.calendar_service = WorkdayCalendarService()
        # 使用更穩定的政府開放資料 API
        self.base_url = "https://data.gov.tw/api/v1/rest/dataset"
        self.holiday_dataset_id = "382000000A-000077-002"  # 國定假日資料集
        # 備用 API 端點
        self.backup_url = "https://data.ntpc.gov.tw/api/v1/rest/datastore"
        
    def fetch_government_holidays(self, year):
        """
        從台灣政府開放資料平台取得國定假日資料
        
        Args:
            year: 年份 (例如: 2025)
            
        Returns:
            list: 國定假日資料列表
        """
        import requests
        import json
        
        try:
            # 由於政府 API 存取限制，直接使用模擬資料
            logger.info(f"政府 API 存取受限，使用模擬資料取得 {year} 年國定假日...")
            return self._get_mock_holiday_data(year)
            
        except Exception as e:
            logger.error(f"取得政府國定假日資料時發生錯誤: {str(e)}")
            logger.info("使用模擬資料作為備用方案...")
            return self._get_mock_holiday_data(year)
                
        except requests.exceptions.RequestException as e:
            logger.error(f"請求政府 API 失敗: {str(e)}")
            return []
        except json.JSONDecodeError as e:
            logger.error(f"解析政府 API 回應失敗: {str(e)}")
            return []
        except Exception as e:
            logger.error(f"取得政府國定假日資料時發生錯誤: {str(e)}")
            logger.info("使用模擬資料作為備用方案...")
            return self._get_mock_holiday_data(year)
    
    def _get_mock_holiday_data(self, year):
        """
        提供模擬的國定假日資料（當政府 API 無法存取時使用）
        
        Args:
            year: 年份
            
        Returns:
            list: 模擬的國定假日資料
        """
        # 台灣常見的國定假日（固定日期）
        fixed_holidays = [
            {'date': f'{year}/1/1', 'name': '元旦'},
            {'date': f'{year}/2/28', 'name': '和平紀念日'},
            {'date': f'{year}/4/4', 'name': '兒童節'},
            {'date': f'{year}/4/5', 'name': '清明節'},
            {'date': f'{year}/5/1', 'name': '勞動節'},
            {'date': f'{year}/10/10', 'name': '國慶日'},
        ]
        
        # 農曆節日（需要根據年份計算，這裡提供近似日期）
        lunar_holidays = []
        
        # 春節（通常在1月底或2月初）
        if year == 2025:
            lunar_holidays.extend([
                {'date': f'{year}/2/8', 'name': '春節'},
                {'date': f'{year}/2/9', 'name': '春節'},
                {'date': f'{year}/2/10', 'name': '春節'},
                {'date': f'{year}/2/11', 'name': '春節'},
                {'date': f'{year}/2/12', 'name': '春節'},
                {'date': f'{year}/2/13', 'name': '春節'},
                {'date': f'{year}/2/14', 'name': '春節'},
            ])
        elif year == 2024:
            lunar_holidays.extend([
                {'date': f'{year}/2/10', 'name': '春節'},
                {'date': f'{year}/2/11', 'name': '春節'},
                {'date': f'{year}/2/12', 'name': '春節'},
                {'date': f'{year}/2/13', 'name': '春節'},
                {'date': f'{year}/2/14', 'name': '春節'},
            ])
        
        # 端午節（通常在6月）
        if year == 2025:
            lunar_holidays.append({'date': f'{year}/6/22', 'name': '端午節'})
        elif year == 2024:
            lunar_holidays.append({'date': f'{year}/6/10', 'name': '端午節'})
        
        # 中秋節（通常在9月）
        if year == 2025:
            lunar_holidays.append({'date': f'{year}/9/29', 'name': '中秋節'})
        elif year == 2024:
            lunar_holidays.append({'date': f'{year}/9/17', 'name': '中秋節'})
        
        # 合併所有假期
        all_holidays = fixed_holidays + lunar_holidays
        
        logger.info(f"使用模擬資料，提供 {len(all_holidays)} 個國定假日")
        return all_holidays
    
    def parse_holiday_data(self, holiday_records):
        """
        解析政府國定假日資料
        
        Args:
            holiday_records: 政府 API 回傳的國定假日資料
            
        Returns:
            dict: 解析後的假期資料 {date: holiday_name}
        """
        from datetime import datetime
        
        holidays = {}
        
        for record in holiday_records:
            try:
                # 解析日期欄位（格式可能為 "2025/1/1" 或 "2025-01-01"）
                date_str = record.get('date', '')
                if not date_str:
                    continue
                
                # 處理不同的日期格式
                if '/' in date_str:
                    date_obj = datetime.strptime(date_str, '%Y/%m/%d').date()
                elif '-' in date_str:
                    date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                else:
                    continue
                
                # 取得假期名稱
                holiday_name = record.get('name', '國定假日')
                
                holidays[date_obj] = holiday_name
                
            except Exception as e:
                logger.warning(f"解析假期資料失敗: {record} - {str(e)}")
                continue
        
        return holidays
    
    def sync_government_holidays(self, year):
        """
        同步政府國定假日資料到系統
        
        Args:
            year: 年份
            
        Returns:
            dict: 同步結果統計
        """
        # 取得政府國定假日資料
        holiday_records = self.fetch_government_holidays(year)
        if not holiday_records:
            return {
                'success': False,
                'message': f'無法取得 {year} 年政府國定假日資料',
                'synced_count': 0,
                'errors': []
            }
        
        # 解析假期資料
        government_holidays = self.parse_holiday_data(holiday_records)
        if not government_holidays:
            return {
                'success': False,
                'message': f'解析 {year} 年政府國定假日資料失敗',
                'synced_count': 0,
                'errors': []
            }
        
        # 同步到系統
        synced_count = 0
        errors = []
        
        for date, holiday_name in government_holidays.items():
            try:
                # 檢查是否已存在
                existing_event = self.calendar_service.add_holiday(
                    date=date,
                    holiday_name=holiday_name,
                    description=f"政府國定假日 - {holiday_name}",
                    created_by="government_api"
                )
                
                if existing_event:
                    synced_count += 1
                    logger.info(f"成功同步國定假日: {holiday_name} ({date})")
                
            except Exception as e:
                error_msg = f"同步假期失敗: {holiday_name} ({date}) - {str(e)}"
                errors.append(error_msg)
                logger.error(error_msg)
        
        # 回傳結果
        result = {
            'success': synced_count > 0,
            'message': f'成功同步 {synced_count} 個國定假日',
            'synced_count': synced_count,
            'total_holidays': len(government_holidays),
            'errors': errors
        }
        
        logger.info(f"政府國定假日同步完成: {result}")
        return result
    
    def sync_multiple_years(self, start_year, end_year):
        """
        同步多個年份的政府國定假日資料
        
        Args:
            start_year: 開始年份
            end_year: 結束年份
            
        Returns:
            dict: 同步結果統計
        """
        total_synced = 0
        total_errors = []
        year_results = {}
        
        for year in range(start_year, end_year + 1):
            logger.info(f"開始同步 {year} 年國定假日...")
            result = self.sync_government_holidays(year)
            
            year_results[year] = result
            total_synced += result['synced_count']
            total_errors.extend(result['errors'])
        
        return {
            'success': total_synced > 0,
            'message': f'多年度同步完成，共同步 {total_synced} 個國定假日',
            'total_synced': total_synced,
            'year_results': year_results,
            'total_errors': total_errors
        }
    
    def get_holiday_status(self, year):
        """
        檢查指定年份的假期設定狀態
        
        Args:
            year: 年份
            
        Returns:
            dict: 假期設定狀態
        """
        from scheduling.models import Event
        from datetime import date
        
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        
        # 統計系統中的假期
        system_holidays = Event.objects.filter(
            type='holiday',
            start__date__gte=start_date,
            start__date__lte=end_date
        ).count()
        
        # 統計政府 API 假期
        government_holidays = self.fetch_government_holidays(year)
        government_count = len(government_holidays) if government_holidays else 0
        
        return {
            'year': year,
            'system_holidays': system_holidays,
            'government_holidays': government_count,
            'sync_needed': government_count > system_holidays,
            'government_data_available': government_count > 0
        }
    
    def get_available_years(self):
        """
        取得可用的年份範圍
        
        Returns:
            list: 可用年份列表
        """
        # 政府 API 通常提供最近幾年的資料
        current_year = datetime.now().year
        return list(range(current_year - 2, current_year + 3))  # 前2年到後2年

class CSVHolidayImportService:
    """CSV 國定假日匯入服務"""
    
    def __init__(self):
        self.calendar_service = WorkdayCalendarService()
    
    def parse_csv_holidays(self, csv_file):
        """
        解析 CSV 檔案中的國定假日資料
        
        Args:
            csv_file: 上傳的 CSV 檔案
            
        Returns:
            dict: 解析結果 {success: bool, data: list, errors: list}
        """
        import csv
        import io
        from datetime import datetime
        
        holidays = []
        errors = []
        
        try:
            # 讀取 CSV 檔案 - 修正檔案處理邏輯
            if hasattr(csv_file, 'read'):
                # 如果是 Django 上傳的檔案，直接讀取 bytes 並解碼
                content = csv_file.read()
                if isinstance(content, bytes):
                    content = content.decode('utf-8-sig')  # 使用 utf-8-sig 處理 BOM
                elif isinstance(content, str):
                    pass  # 已經是字串
                else:
                    content = str(content)
            else:
                content = str(csv_file)
            
            # 處理 BOM 問題
            if content.startswith('\ufeff'):
                content = content[1:]
            
            csv_reader = csv.DictReader(io.StringIO(content))
            
            for row_num, row in enumerate(csv_reader, start=2):  # 從第2行開始（跳過標題）
                try:
                    # 檢查是否為空行或無效行（所有欄位都為空，或只有逗號）
                    if all(not value.strip() for value in row.values()):
                        continue  # 跳過空行
                    
                    # 檢查是否有必要的欄位值
                    subject = row.get('Subject', '').strip()
                    date_str = row.get('Start Date', '').strip()
                    
                    # 如果沒有主題和日期，跳過這行
                    if not subject and not date_str:
                        continue  # 跳過無效行
                    
                    # 檢查是否有假期名稱
                    if not subject:
                        errors.append(f"第 {row_num} 行：缺少假期名稱")
                        continue
                    
                    # 解析日期欄位
                    if not date_str:
                        errors.append(f"第 {row_num} 行：缺少開始日期")
                        continue
                    
                    # 處理不同的日期格式
                    date_obj = self._parse_date(date_str)
                    if not date_obj:
                        errors.append(f"第 {row_num} 行：日期格式錯誤 '{date_str}'")
                        continue
                    
                    # 取得假期名稱
                    subject = row.get('Subject', '').strip()
                    if not subject:
                        errors.append(f"第 {row_num} 行：缺少假期名稱")
                        continue
                    
                    # 檢查是否為全天事件（支援多種欄位名稱）
                    all_day_ev = row.get('All Day Ev', '').strip()
                    all_day_event = row.get('All Day Event', '').strip()
                    all_day = (all_day_ev.upper() in ['TRUE', 'YES', '1', '是'] or 
                              all_day_event.upper() in ['TRUE', 'YES', '1', '是'])
                    
                    # 取得描述
                    description = row.get('Description', '').strip()
                    
                    holidays.append({
                        'date': date_obj,
                        'name': subject,
                        'description': description,
                        'all_day': all_day,
                        'row_num': row_num
                    })
                    
                except Exception as e:
                    errors.append(f"第 {row_num} 行：解析錯誤 - {str(e)}")
                    continue
            
            return {
                'success': len(holidays) > 0,
                'data': holidays,
                'errors': errors,
                'total_rows': len(holidays) + len(errors)
            }
            
        except Exception as e:
            return {
                'success': False,
                'data': [],
                'errors': [f"CSV 檔案解析失敗：{str(e)}"],
                'total_rows': 0
            }
    
    def _parse_date(self, date_str):
        """
        解析日期字串
        
        Args:
            date_str: 日期字串
            
        Returns:
            date: 解析後的日期物件，失敗則返回 None
        """
        from datetime import datetime
        
        # 支援多種日期格式
        date_formats = [
            '%Y/%m/%d',    # 2026/1/1
            '%Y-%m-%d',    # 2026-01-01
            '%Y/%m/%d',    # 2026/01/01
            '%Y-%m-%d',    # 2026-1-1
            '%d/%m/%Y',    # 1/1/2026
            '%d-%m-%Y',    # 1-1-2026
        ]
        
        for fmt in date_formats:
            try:
                return datetime.strptime(date_str, fmt).date()
            except ValueError:
                continue
        
        return None
    
    def import_holidays_from_csv(self, csv_file):
        """
        從 CSV 檔案匯入國定假日
        
        Args:
            csv_file: 上傳的 CSV 檔案
            
        Returns:
            dict: 匯入結果統計
        """
        # 解析 CSV 檔案
        parse_result = self.parse_csv_holidays(csv_file)
        
        if not parse_result['success']:
            return {
                'success': False,
                'message': f"CSV 檔案解析失敗，共 {len(parse_result['errors'])} 個錯誤",
                'imported_count': 0,
                'errors': parse_result['errors']
            }
        
        # 匯入假期資料
        imported_count = 0
        skipped_count = 0
        errors = []
        
        for holiday in parse_result['data']:
            try:
                # 檢查是否已存在
                existing_event = self.calendar_service.add_holiday(
                    date=holiday['date'],
                    holiday_name=holiday['name'],
                    description=holiday['description'],
                    created_by="csv_import"
                )
                
                if existing_event:
                    imported_count += 1
                    logger.info(f"成功匯入假期: {holiday['name']} ({holiday['date']})")
                else:
                    skipped_count += 1
                    logger.info(f"跳過已存在的假期: {holiday['name']} ({holiday['date']})")
                
            except Exception as e:
                error_msg = f"匯入假期失敗: {holiday['name']} ({holiday['date']}) - {str(e)}"
                errors.append(error_msg)
                logger.error(error_msg)
        
        # 回傳結果
        result = {
            'success': imported_count > 0,
            'message': f'CSV 匯入完成！成功匯入 {imported_count} 個假期，跳過 {skipped_count} 個已存在的假期',
            'imported_count': imported_count,
            'skipped_count': skipped_count,
            'total_parsed': len(parse_result['data']),
            'errors': errors + parse_result['errors']
        }
        
        logger.info(f"CSV 假期匯入完成: {result}")
        return result
    
    def generate_sample_csv(self):
        """
        生成範例 CSV 檔案內容
        
        Returns:
            str: CSV 檔案內容
        """
        sample_data = [
            {
                'Subject': '中華民國開國紀念日',
                'Start Date': '2026/1/1',
                'Start Time': '',
                'End Date': '2026/1/1',
                'End Time': '',
                'All Day Event': 'TRUE',
                'Description': '',
                'Location': ''
            },
            {
                'Subject': '例假日',
                'Start Date': '2026/1/3',
                'Start Time': '',
                'End Date': '2026/1/3',
                'End Time': '',
                'All Day Event': 'TRUE',
                'Description': '',
                'Location': ''
            },
            {
                'Subject': '例假日',
                'Start Date': '2026/1/4',
                'Start Time': '',
                'End Date': '2026/1/4',
                'End Time': '',
                'All Day Event': 'TRUE',
                'Description': '',
                'Location': ''
            },
            {
                'Subject': '例假日',
                'Start Date': '2026/1/10',
                'Start Time': '',
                'End Date': '2026/1/10',
                'End Time': '',
                'All Day Event': 'TRUE',
                'Description': '',
                'Location': ''
            },
            {
                'Subject': '例假日',
                'Start Date': '2026/1/11',
                'Start Time': '',
                'End Date': '2026/1/11',
                'End Time': '',
                'All Day Event': 'TRUE',
                'Description': '',
                'Location': ''
            },
            {
                'Subject': '例假日',
                'Start Date': '2026/1/17',
                'Start Time': '',
                'End Date': '2026/1/17',
                'End Time': '',
                'All Day Event': 'TRUE',
                'Description': '',
                'Location': ''
            },
            {
                'Subject': '例假日',
                'Start Date': '2026/1/18',
                'Start Time': '',
                'End Date': '2026/1/18',
                'End Time': '',
                'All Day Event': 'TRUE',
                'Description': '',
                'Location': ''
            },
            {
                'Subject': '例假日',
                'Start Date': '2026/1/24',
                'Start Time': '',
                'End Date': '2026/1/24',
                'End Time': '',
                'All Day Event': 'TRUE',
                'Description': '',
                'Location': ''
            },
            {
                'Subject': '例假日',
                'Start Date': '2026/1/25',
                'Start Time': '',
                'End Date': '2026/1/25',
                'End Time': '',
                'All Day Event': 'TRUE',
                'Description': '',
                'Location': ''
            },
            {
                'Subject': '例假日',
                'Start Date': '2026/1/31',
                'Start Time': '',
                'End Date': '2026/1/31',
                'End Time': '',
                'All Day Event': 'TRUE',
                'Description': '',
                'Location': ''
            },
            {
                'Subject': '例假日',
                'Start Date': '2026/2/1',
                'Start Time': '',
                'End Date': '2026/2/1',
                'End Time': '',
                'All Day Event': 'TRUE',
                'Description': '',
                'Location': ''
            }
        ]
        
        import csv
        import io
        
        # 生成 CSV 內容
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[
            'Subject', 'Start Date', 'Start Time', 'End Date', 'End Time', 
            'All Day Event', 'Description', 'Location'
        ])
        
        writer.writeheader()
        writer.writerows(sample_data)
        
        return output.getvalue()
        """生成Excel報表"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            import os
            from django.conf import settings
            
            # 建立工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '工單報表'
            
            # 定義標題列
            headers = [
                '作業員名稱', '公司名稱', '報工日期', '開始時間', '結束時間', 
                '工單號', '產品編號', '工序名稱', '設備名稱', '報工數量', 
                '不良品數量', '備註', '異常紀錄', '工作時數', '加班時數'
            ]
            
            # 設定標題列樣式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            # 寫入標題列
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # 填入資料
            for row, record in enumerate(data, 2):
                ws.cell(row=row, column=1, value=record.operator_name or '')
                ws.cell(row=row, column=2, value=record.company or '')
                ws.cell(row=row, column=3, value=record.work_date.strftime('%Y-%m-%d') if record.work_date else '')
                ws.cell(row=row, column=4, value=record.start_time.strftime('%H:%M') if record.start_time else '')
                ws.cell(row=row, column=5, value=record.end_time.strftime('%H:%M') if record.end_time else '')
                ws.cell(row=row, column=6, value=record.workorder_id or '')
                ws.cell(row=row, column=7, value=record.product_code or '')
                ws.cell(row=row, column=8, value=record.process_name or '')
                ws.cell(row=row, column=9, value='')  # equipment_name 欄位不存在
                ws.cell(row=row, column=10, value=0)  # completed_quantity 欄位不存在
                ws.cell(row=row, column=11, value=0)  # defect_quantity 欄位不存在
                ws.cell(row=row, column=12, value='')  # remarks 欄位不存在
                ws.cell(row=row, column=13, value='')  # abnormal_notes 欄位不存在
                ws.cell(row=row, column=14, value=float(record.work_hours or 0))
                ws.cell(row=row, column=15, value=float(record.overtime_hours or 0))
            
            # 調整欄寬
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 生成檔案名稱 - 使用中文名稱
            filename = f"月工作時數報表_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # 確保目錄存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 儲存檔案
            wb.save(file_path)
            
            return file_path
            
        except ImportError:
            logger.error("未安裝 openpyxl，無法生成 Excel 報表")
            return None
        except Exception as e:
            logger.error(f"生成 Excel 報表失敗: {str(e)}")
            return None

class WorkOrderAnalysisService:
    """工單分析服務"""
    
    @staticmethod
    def analyze_completed_workorder(workorder_id, company_code, product_code=None, force=False):
        """
        分析單一已完工工單
        
        Args:
            workorder_id: 工單編號
            company_code: 公司代號
            force: 是否強制重新分析，即使已經分析過
            
        Returns:
            dict: 分析結果
        """
        from workorder.models import CompletedWorkOrder, CompletedWorkOrderProcess
        from .models import CompletedWorkOrderAnalysis
        from datetime import datetime, timedelta
        from decimal import Decimal
        
        try:
            # 檢查是否已經分析過
            if not force:
                existing_analysis = CompletedWorkOrderAnalysis.objects.filter(
                    workorder_id=workorder_id,
                    company_code=company_code
                ).first()
                
                if existing_analysis:
                    return {
                        'success': True,
                        'message': f'工單 {workorder_id} 已經分析過，跳過分析',
                        'analysis_id': existing_analysis.id,
                        'created': False,
                        'skipped': True
                    }
            
            # 取得已完工工單（使用公司名稱+工單號碼+產品編號作為唯一識別）
            filter_kwargs = {
                'order_number': workorder_id,
                'company_code': company_code
            }
            if product_code:
                filter_kwargs['product_code'] = product_code
                
            completed_workorder = CompletedWorkOrder.objects.filter(**filter_kwargs).first()
            
            if not completed_workorder:
                return {
                    'success': False,
                    'error': f'找不到工單 {workorder_id}'
                }
            
            # 從填報記錄取得詳細資料
            from workorder.fill_work.models import FillWork
            
            # 查詢填報記錄（填報記錄的 company_code 可能為 None，所以不加入公司代號條件）
            work_records = FillWork.objects.filter(
                workorder=workorder_id,
                product_id=completed_workorder.product_code
            ).order_by('work_date', 'start_time')
            
            if not work_records.exists():
                return {
                    'success': False,
                    'error': f'工單 {workorder_id} 沒有找到填報記錄'
                }
            
            # 計算時間分析
            first_record = work_records.first()
            last_record = work_records.last()
            first_date = first_record.work_date
            last_date = last_record.work_date
            total_execution_days = (last_date - first_date).days + 1
            
            # 計算總工作時數
            total_work_hours = sum(float(record.work_hours_calculated or 0) for record in work_records)
            total_overtime_hours = sum(float(record.overtime_hours_calculated or 0) for record in work_records)
            average_daily_hours = total_work_hours / total_execution_days if total_execution_days > 0 else 0
            
            # 計算效率比率（假設標準工時為8小時/天）
            standard_hours = total_execution_days * 8
            efficiency_rate = min(999.99, (total_work_hours / standard_hours * 100) if standard_hours > 0 else 0)
            
            # 工序分析
            process_records = {}
            for record in work_records:
                process_name = record.operation or '未知工序'
                if process_name not in process_records:
                    process_records[process_name] = {
                        'total_hours': 0,
                        'records': [],
                        'operators': set()
                    }
                process_records[process_name]['total_hours'] += float(record.work_hours_calculated or 0)
                process_records[process_name]['records'].append({
                    'date': record.work_date.strftime('%Y-%m-%d'),
                    'operator': record.operator,
                    'hours': float(record.work_hours_calculated or 0),
                    'overtime': float(record.overtime_hours_calculated or 0)
                })
                if record.operator:
                    process_records[process_name]['operators'].add(record.operator)
            
            # 對每個工序的記錄按時間排序，並記錄第一筆時間
            for process_name in process_records:
                process_records[process_name]['records'].sort(key=lambda x: x['date'])
                # 記錄第一筆記錄的時間，用於排序
                if process_records[process_name]['records']:
                    process_records[process_name]['first_record_date'] = process_records[process_name]['records'][0]['date']
                else:
                    process_records[process_name]['first_record_date'] = '9999-12-31'  # 沒有記錄的工序排在最後
            
            # 作業員分析
            operator_records = {}
            for record in work_records:
                operator_name = record.operator or '未知作業員'
                if operator_name not in operator_records:
                    operator_records[operator_name] = {
                        'total_hours': 0,
                        'overtime_hours': 0,
                        'processes': set(),
                        'work_days': set()
                    }
                operator_records[operator_name]['total_hours'] += float(record.work_hours_calculated or 0)
                operator_records[operator_name]['overtime_hours'] += float(record.overtime_hours_calculated or 0)
                if record.operation:
                    operator_records[operator_name]['processes'].add(record.operation)
                operator_records[operator_name]['work_days'].add(record.work_date)
            
            # 準備分析資料
            analysis_data = {
                'workorder_id': workorder_id,
                'company_code': company_code,
                'company_name': completed_workorder.company_name,
                'product_code': completed_workorder.product_code,
                'product_name': completed_workorder.product_code,  # CompletedWorkOrder 沒有 product_name 欄位
                'order_quantity': completed_workorder.completed_quantity,
                'first_record_date': first_date,
                'last_record_date': last_date,
                'total_execution_days': total_execution_days,
                'total_work_hours': total_work_hours,
                'total_overtime_hours': total_overtime_hours,
                'average_daily_hours': average_daily_hours,
                'efficiency_rate': efficiency_rate,
                'total_processes': len(process_records),
                'unique_processes': len(set(record.operation for record in work_records if record.operation)),
                'total_operators': len(operator_records),
                'process_details': {
                    name: {
                        'total_hours': data['total_hours'],
                        'records': data['records'][:5],  # 只保留前5筆記錄
                        'operators': list(data['operators']),
                        'first_record_date': data.get('first_record_date', '9999-12-31')
                    }
                    for name, data in process_records.items()
                },
                'operator_details': {
                    name: {
                        'total_hours': data['total_hours'],
                        'overtime_hours': data['overtime_hours'],
                        'processes': list(data['processes']),
                        'work_days': len(data['work_days'])
                    }
                    for name, data in operator_records.items()
                },
                'completion_date': completed_workorder.completed_at.date(),
                'completion_status': 'completed'
            }
            
            # 儲存或更新分析資料
            analysis, created = CompletedWorkOrderAnalysis.objects.update_or_create(
                workorder_id=workorder_id,
                company_code=company_code,
                product_code=completed_workorder.product_code,
                defaults=analysis_data
            )
            
            return {
                'success': True,
                'message': f'工單 {workorder_id} 分析完成',
                'analysis_id': analysis.id,
                'created': created
            }
            
        except CompletedWorkOrder.DoesNotExist:
            return {
                'success': False,
                'error': f'找不到工單 {workorder_id}'
            }
        except Exception as e:
            return {
                'success': False,
                'error': f'分析工單時發生錯誤: {str(e)}'
            }
    
    @staticmethod
    def analyze_completed_workorders_batch(start_date=None, end_date=None, company_code=None):
        """
        批量分析已完工工單
        
        Args:
            start_date: 開始日期 (YYYY-MM-DD 格式)
            end_date: 結束日期 (YYYY-MM-DD 格式)
            company_code: 公司代號
            
        Returns:
            dict: 批量分析結果
        """
        from workorder.models import CompletedWorkOrder
        from datetime import datetime
        import logging
        
        logger = logging.getLogger(__name__)
        
        try:
            logger.info(f"開始批量分析 - 公司代號: {company_code}, 開始日期: {start_date}, 結束日期: {end_date}")
            
            # 查詢已完工工單
            queryset = CompletedWorkOrder.objects.all()
            logger.info(f"初始查詢結果數量: {queryset.count()}")
            
            if company_code:
                queryset = queryset.filter(company_code=company_code)
                logger.info(f"按公司代號過濾後數量: {queryset.count()}")
            
            # 處理日期格式轉換
            if start_date:
                try:
                    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d').date()
                    queryset = queryset.filter(completed_at__date__gte=start_date_obj)
                    logger.info(f"按開始日期過濾後數量: {queryset.count()}")
                except ValueError:
                    logger.error(f"開始日期格式錯誤: {start_date}")
                    return {
                        'success': False,
                        'error': f'開始日期格式錯誤: {start_date}，請使用 YYYY-MM-DD 格式'
                    }
            
            if end_date:
                try:
                    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d').date()
                    queryset = queryset.filter(completed_at__date__lte=end_date_obj)
                    logger.info(f"按結束日期過濾後數量: {queryset.count()}")
                except ValueError:
                    logger.error(f"結束日期格式錯誤: {end_date}")
                    return {
                        'success': False,
                        'error': f'結束日期格式錯誤: {end_date}，請使用 YYYY-MM-DD 格式'
                    }
            
            # 使用公司名稱+工單號碼+產品編號作為唯一識別
            completed_workorders = queryset.values_list('order_number', 'company_code', 'product_code')
            
            if not completed_workorders.exists():
                return {
                    'success': False,
                    'error': '沒有找到符合條件的已完工工單'
                }
            
            # 批量分析
            success_count = 0
            error_count = 0
            errors = []
            
            logger.info(f"開始分析 {len(completed_workorders)} 筆工單")
            
            for order_number, company_code, product_code in completed_workorders:
                logger.info(f"分析工單: {order_number} ({company_code}) - {product_code}")
                result = WorkOrderAnalysisService.analyze_completed_workorder(order_number, company_code, product_code, force=True)
                if result['success']:
                    success_count += 1
                    logger.info(f"工單 {order_number} 分析成功")
                else:
                    error_count += 1
                    error_msg = f'工單 {order_number}-{product_code}: {result["error"]}'
                    errors.append(error_msg)
                    logger.error(error_msg)
            
            return {
                'success': True,
                'message': f'批量分析完成，成功: {success_count}，失敗: {error_count}',
                'success_count': success_count,
                'error_count': error_count,
                'errors': errors
            }
            
        except Exception as e:
            return {
                'success': False,
                'error': f'批量分析時發生錯誤: {str(e)}'
        } 