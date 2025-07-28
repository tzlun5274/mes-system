# 這個檔案是報表管理模組的Celery任務，負責自動同步SMT生產報表與作業員生產報表
from celery import shared_task
from django.core.management import call_command
from django.core.mail import get_connection, send_mail
from django.template.loader import render_to_string
from django.utils import timezone
from datetime import datetime, timedelta
import openpyxl
from io import BytesIO
from .models import (
    ProductionDailyReport,
    OperatorPerformance,
    ReportEmailSchedule,
    ReportEmailLog,
)
from system.models import EmailConfig
import logging

logger = logging.getLogger(__name__)


@shared_task
def sync_smt_report_task():
    """
    Celery任務：自動同步SMT生產報表
    """
    call_command("sync_smt_report")


@shared_task
def sync_operator_performance_task():
    """
    Celery任務：自動同步作業員生產報表
    """
    call_command("sync_operator_performance")


@shared_task
def send_daily_reports_email():
    """
    Celery任務：每天自動發送報表郵件
    """
    logger.info("開始執行每日報表郵件發送任務")

    try:
        # 取得郵件設定
        email_config = EmailConfig.objects.get(id=1)
    except EmailConfig.DoesNotExist:
        logger.error("郵件設定未找到，無法發送報表郵件")
        return

    # 取得所有啟用的郵件發送設定
    schedules = ReportEmailSchedule.objects.filter(is_active=True)

    for schedule in schedules:
        try:
            # 檢查是否應該發送（根據發送頻率）
            if not should_send_report(schedule):
                continue

            # 生成報表數據
            report_data = generate_report_data(schedule.report_type)

            # 生成Excel檔案
            excel_file = generate_excel_report(schedule.report_type, report_data)

            # 生成郵件內容
            subject, message, html_message = generate_email_content(
                schedule, report_data
            )

            # 發送郵件
            success = send_report_email(
                email_config, schedule, subject, message, html_message, excel_file
            )

            # 記錄發送結果
            log_email_send(schedule, subject, success)

        except Exception as e:
            logger.error(f"發送報表郵件失敗 {schedule.report_type}: {str(e)}")
            log_email_send(
                schedule,
                f"MES 系統 - {schedule.get_report_type_display()}",
                False,
                str(e),
            )


def should_send_report(schedule):
    """檢查是否應該發送報表"""
    now = timezone.now()

    if schedule.schedule_type == "daily":
        return True
    elif schedule.schedule_type == "weekly":
        # 每週一發送
        return now.weekday() == 0
    elif schedule.schedule_type == "monthly":
        # 每月第一天發送
        return now.day == 1

    return False


def generate_report_data(report_type):
    """生成報表數據"""
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)

    if report_type == "operator":
        return OperatorPerformance.objects.filter(date=yesterday)
    elif report_type == "production_daily":
        return ProductionDailyReport.objects.filter(date=yesterday)
    elif report_type == "all":
        return {
            "operator": OperatorPerformance.objects.filter(date=yesterday),
            "production_daily": ProductionDailyReport.objects.filter(date=yesterday),
        }

    return []


def generate_excel_report(report_type, data):
    """生成Excel報表檔案"""
    workbook = openpyxl.Workbook()

    if report_type == "operator":
        worksheet = workbook.active
        worksheet.title = "作業員生產報表"

        headers = ["作業員名稱", "設備名稱", "生產數量", "設備使用率 (%)", "日期"]
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)

        for row_num, report in enumerate(data, 2):
            worksheet.cell(row=row_num, column=1, value=report.operator_name)
            worksheet.cell(row=row_num, column=2, value=report.equipment_name)
            worksheet.cell(row=row_num, column=3, value=report.production_quantity)
            worksheet.cell(row=row_num, column=4, value=report.equipment_usage_rate)
            worksheet.cell(
                row=row_num, column=5, value=report.date.strftime("%Y-%m-%d")
            )

    elif report_type == "production_daily":
        worksheet = workbook.active
        worksheet.title = "生產日報表"

        headers = [
            "日期",
            "作業員姓名",
            "設備名稱",
            "生產線",
            "完成數量",
            "工作時數",
            "效率 (%)",
        ]
        for col_num, header in enumerate(headers, 1):
            worksheet.cell(row=1, column=col_num, value=header)

        for row_num, report in enumerate(data, 2):
            worksheet.cell(
                row=row_num, column=1, value=report.date.strftime("%Y-%m-%d")
            )
            worksheet.cell(row=row_num, column=2, value=report.operator_or_line)
            worksheet.cell(row=row_num, column=3, value=report.equipment_name)
            worksheet.cell(row=row_num, column=4, value=report.get_line_display())
            worksheet.cell(row=row_num, column=5, value=report.completed_quantity)
            worksheet.cell(row=row_num, column=6, value=report.work_hours)
            worksheet.cell(row=row_num, column=7, value=report.efficiency_rate)

    # 將Excel檔案轉換為BytesIO物件
    excel_file = BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    return excel_file


def generate_email_content(schedule, data):
    """生成郵件內容"""
    today = timezone.now().date()
    yesterday = today - timedelta(days=1)

    # 生成主旨
    subject = schedule.subject_template.format(
        report_type=schedule.get_report_type_display(),
        date=yesterday.strftime("%Y-%m-%d"),
    )

    # 生成純文字內容
    message = f"""
您好，

這是 {yesterday.strftime('%Y-%m-%d')} 的 {schedule.get_report_type_display()}。

報表數據摘要：
- 報表類型：{schedule.get_report_type_display()}
- 報表日期：{yesterday.strftime('%Y-%m-%d')}
- 數據筆數：{len(data) if not isinstance(data, dict) else sum(len(v) for v in data.values())}

詳細數據請查看附件。

此郵件由 MES 系統自動發送，請勿回覆。

MES 系統團隊
    """.strip()

    # 生成HTML內容
    html_message = render_to_string(
        "reporting/email_report_template.html",
        {
            "report_type": schedule.get_report_type_display(),
            "report_date": yesterday.strftime("%Y-%m-%d"),
            "data_count": (
                len(data)
                if not isinstance(data, dict)
                else sum(len(v) for v in data.values())
            ),
            "data": data,
            "schedule": schedule,
        },
    )

    return subject, message, html_message


def send_report_email(
    email_config, schedule, subject, message, html_message, excel_file
):
    """發送報表郵件"""
    try:
        # 設定SMTP連線
        connection = get_connection(
            backend="django.core.mail.backends.smtp.EmailBackend",
            host=email_config.email_host,
            port=email_config.email_port,
            username=email_config.email_host_user,
            password=email_config.email_host_password,
            use_tls=email_config.email_use_tls,
        )

        # 準備收件人
        recipient_list = schedule.get_recipient_list()
        if not recipient_list:
            logger.warning(f"沒有設定收件人: {schedule.report_type}")
            return False

        # 發送郵件
        from django.core.mail import EmailMessage

        email = EmailMessage(
            subject=subject,
            body=html_message,
            from_email=email_config.default_from_email,
            to=recipient_list,
            cc=schedule.get_cc_list(),
            connection=connection,
        )
        email.content_subtype = "html"

        # 添加Excel附件
        filename = f"{schedule.get_report_type_display()}_{timezone.now().strftime('%Y%m%d')}.xlsx"
        email.attach(
            filename,
            excel_file.getvalue(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        email.send()
        logger.info(f"報表郵件發送成功: {schedule.report_type} -> {recipient_list}")
        return True

    except Exception as e:
        logger.error(f"發送報表郵件失敗: {str(e)}")
        return False


def log_email_send(schedule, subject, success, error_message=""):
    """記錄郵件發送結果"""
    ReportEmailLog.objects.create(
        schedule=schedule,
        report_type=schedule.report_type,
        recipients=schedule.recipients,
        subject=subject,
        status="success" if success else "failed",
        error_message=error_message,
    )
