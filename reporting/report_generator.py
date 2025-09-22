"""
報表生成器
只負責報表生成相關的邏輯
"""

import logging
import os
from datetime import datetime, timedelta, date
from django.utils import timezone
from django.conf import settings

logger = logging.getLogger(__name__)


class ReportGenerator:
    """報表生成器 - 只負責報表生成"""
    
    def __init__(self):
        # 延遲載入避免循環依賴
        self._email_service = None
        self._calendar_service = None
    
    @property
    def email_service(self):
        """延遲載入郵件服務"""
        if self._email_service is None:
            from .email_service import EmailService
            self._email_service = EmailService()
        return self._email_service
    
    @property
    def calendar_service(self):
        """延遲載入工作日曆服務"""
        if self._calendar_service is None:
            from .workday_calendar import WorkdayCalendarService
            self._calendar_service = WorkdayCalendarService()
        return self._calendar_service
    
    def execute_previous_workday_report(self, schedule):
        """執行前一個工作日報表"""
        try:
            # 取得報表日期
            report_date = self._get_previous_workday_date()
            
            # 收集資料
            data = self._collect_previous_workday_data(report_date)
            
            # 檢查是否有資料
            if data['fill_works_count'] == 0 and data['onsite_reports_count'] == 0:
                logger.warning(f"{report_date} 沒有找到任何填報或現場報工資料")
                # 即使沒有資料，也生成空的報表檔案，避免排程執行失敗
                report_file = self._generate_previous_workday_report(data, schedule)
                
                # 處理不同檔案格式的路徑
                excel_path = None
                html_path = None
                
                if schedule and schedule.file_format == 'both':
                    # both 格式：從物件取得兩個檔案路徑
                    excel_path = getattr(self, '_excel_filepath', None)
                    html_path = getattr(self, '_html_filepath', None)
                elif schedule and schedule.file_format == 'excel':
                    # excel 格式
                    excel_path = report_file
                else:
                    # html 格式或預設
                    html_path = report_file
                
                result = {
                    'success': True,
                    'filename': os.path.basename(report_file) if report_file else '',
                    'file_path': report_file or '',
                    'excel_path': excel_path,
                    'html_path': html_path,
                    'message': f'前一個工作日報表生成成功（無資料），報表日期: {report_date}'
                }
            else:
                # 生成報表
                report_file = self._generate_previous_workday_report(data, schedule)
                
                # 記錄執行結果
                logger.info(f"前一個工作日報表生成成功: {report_date}")
                
                # 處理不同檔案格式的路徑
                excel_path = None
                html_path = None
                
                if schedule and schedule.file_format == 'both':
                    # both 格式：從物件取得兩個檔案路徑
                    excel_path = getattr(self, '_excel_filepath', None)
                    html_path = getattr(self, '_html_filepath', None)
                elif schedule and schedule.file_format == 'excel':
                    # excel 格式
                    excel_path = report_file
                else:
                    # html 格式或預設
                    html_path = report_file
                
                result = {
                    'success': True,
                    'filename': os.path.basename(report_file) if report_file else '',
                    'file_path': report_file or '',
                    'excel_path': excel_path,
                    'html_path': html_path,
                    'message': f'前一個工作日報表生成成功，報表日期: {report_date}'
                }
            
            # 如果有設定收件人，直接發送郵件
            if schedule and schedule.email_recipients and schedule.email_recipients.strip():
                self.email_service.send_report_email(schedule, result)
            
            return result
            
        except Exception as e:
            logger.error(f"前一個工作日報表生成失敗: {str(e)}")
            return {
                'success': False,
                'filename': '',
                'file_path': '',
                'message': f'前一個工作日報表生成失敗: {str(e)}'
            }
    
    def execute_weekly_report(self, schedule):
        """執行週報表"""
        try:
            now = timezone.localtime(timezone.now())
            
            if schedule.report_type == 'current_week':
                # 本週報表：從週一到現在
                days_since_monday = now.weekday()
                start_date = now.date() - timedelta(days=days_since_monday)
                end_date = now.date()
                report_title = f"本週報表 ({start_date} ~ {end_date})"
            else:
                # 上週報表：上週一到上週日
                days_since_monday = now.weekday()
                last_monday = now.date() - timedelta(days=days_since_monday + 7)
                last_sunday = last_monday + timedelta(days=6)
                start_date = last_monday
                end_date = last_sunday
                report_title = f"上週報表 ({start_date} ~ {end_date})"
            
            # 使用統一的報表生成方法
            result = self._generate_unified_report(start_date, end_date, report_title, schedule)
            
            # 如果有設定收件人，發送郵件
            if schedule and schedule.email_recipients and schedule.email_recipients.strip():
                self.email_service.send_report_email(schedule, result)
            
            return result
            
        except Exception as e:
            logger.error(f"週報表生成失敗: {str(e)}")
            return {
                'success': False,
                'filename': '',
                'file_path': '',
                'message': f'週報表生成失敗: {str(e)}'
            }
    
    def execute_monthly_report(self, schedule):
        """執行月報表"""
        try:
            now = timezone.localtime(timezone.now())
            
            if schedule.report_type == 'current_month':
                # 本月報表：從本月1號到現在
                start_date = now.date().replace(day=1)
                end_date = now.date()
                report_title = f"本月報表 ({start_date} ~ {end_date})"
            else:
                # 上月報表：上月1號到上月最後一天
                if now.month == 1:
                    # 如果是1月，上月是去年12月
                    start_date = date(now.year - 1, 12, 1)
                    end_date = date(now.year - 1, 12, 31)
                else:
                    # 其他月份
                    start_date = now.date().replace(month=now.month - 1, day=1)
                    # 上月最後一天
                    if now.month - 1 in [1, 3, 5, 7, 8, 10, 12]:
                        end_date = start_date.replace(day=31)
                    elif now.month - 1 in [4, 6, 9, 11]:
                        end_date = start_date.replace(day=30)
                    else:  # 2月
                        # 檢查是否為閏年
                        if now.year % 4 == 0 and (now.year % 100 != 0 or now.year % 400 == 0):
                            end_date = start_date.replace(day=29)
                        else:
                            end_date = start_date.replace(day=28)
                
                report_title = f"上月報表 ({start_date} ~ {end_date})"
            
            # 使用統一的報表生成方法
            result = self._generate_unified_report(start_date, end_date, report_title, schedule)
            
            # 如果有設定收件人，發送郵件
            if schedule and schedule.email_recipients and schedule.email_recipients.strip():
                self.email_service.send_report_email(schedule, result)
            
            return result
            
        except Exception as e:
            logger.error(f"月報表生成失敗: {str(e)}")
            return {
                'success': False,
                'filename': '',
                'file_path': '',
                'message': f'月報表生成失敗: {str(e)}'
            }
    
    def execute_quarterly_report(self, schedule):
        """執行季報表"""
        try:
            now = timezone.localtime(timezone.now())
            
            if schedule.report_type == 'current_quarter':
                # 本季報表：從本季第一天到現在
                quarter = (now.month - 1) // 3 + 1
                start_month = (quarter - 1) * 3 + 1
                start_date = now.date().replace(month=start_month, day=1)
                end_date = now.date()
                report_title = f"本季報表 ({start_date} ~ {end_date})"
            else:
                # 上季報表：上季第一天到上季最後一天
                quarter = (now.month - 1) // 3 + 1
                if quarter == 1:
                    # 上季是去年第4季
                    start_date = date(now.year - 1, 10, 1)
                    end_date = date(now.year - 1, 12, 31)
                else:
                    # 上季是今年
                    start_month = (quarter - 2) * 3 + 1
                    start_date = now.date().replace(month=start_month, day=1)
                    end_month = start_month + 2
                    # 上季最後一天
                    if end_month in [1, 3, 5, 7, 8, 10, 12]:
                        end_date = now.date().replace(month=end_month, day=31)
                    elif end_month in [4, 6, 9, 11]:
                        end_date = now.date().replace(month=end_month, day=30)
                    else:  # 2月
                        if now.year % 4 == 0 and (now.year % 100 != 0 or now.year % 400 == 0):
                            end_date = now.date().replace(month=end_month, day=29)
                        else:
                            end_date = now.date().replace(month=end_month, day=28)
                
                report_title = f"上季報表 ({start_date} ~ {end_date})"
            
            # 使用統一的報表生成方法
            result = self._generate_unified_report(start_date, end_date, report_title, schedule)
            
            # 如果有設定收件人，發送郵件
            if schedule and schedule.email_recipients and schedule.email_recipients.strip():
                self.email_service.send_report_email(schedule, result)
            
            return result
            
        except Exception as e:
            logger.error(f"季報表生成失敗: {str(e)}")
            return {
                'success': False,
                'filename': '',
                'file_path': '',
                'message': f'季報表生成失敗: {str(e)}'
            }
    
    def execute_yearly_report(self, schedule):
        """執行年報表"""
        try:
            now = timezone.localtime(timezone.now())
            
            if schedule.report_type == 'current_year':
                # 本年報表：從今年1月1號到現在
                start_date = now.date().replace(month=1, day=1)
                end_date = now.date()
                report_title = f"本年報表 ({start_date} ~ {end_date})"
            else:
                # 去年報表：去年1月1號到去年12月31號
                start_date = date(now.year - 1, 1, 1)
                end_date = date(now.year - 1, 12, 31)
                report_title = f"去年報表 ({start_date} ~ {end_date})"
            
            # 使用統一的報表生成方法
            result = self._generate_unified_report(start_date, end_date, report_title, schedule)
            
            # 如果有設定收件人，發送郵件
            if schedule and schedule.email_recipients and schedule.email_recipients.strip():
                self.email_service.send_report_email(schedule, result)
            
            return result
            
        except Exception as e:
            logger.error(f"年報表生成失敗: {str(e)}")
            return {
                'success': False,
                'filename': '',
                'file_path': '',
                'message': f'年報表生成失敗: {str(e)}'
            }
    
    def _generate_unified_report(self, start_date, end_date, report_title, schedule):
        """統一的報表生成方法"""
        try:
            # 收集資料
            data = self._collect_workorder_data(start_date, end_date, schedule.company)
            
            # 根據用戶選擇的格式生成檔案
            html_result = None
            excel_result = None
            file_path = None
            filename = None
            
            if schedule.file_format in ['html', 'both']:
                html_result = self._generate_html_report(data, report_title, schedule)
                if not file_path:
                    file_path = html_result
                    filename = os.path.basename(html_result)
            
            if schedule.file_format in ['excel', 'both']:
                excel_result = self._generate_excel_report(data, report_title, schedule)
                if excel_result:
                    file_path = excel_result
                    filename = os.path.basename(excel_result)
            
            # 準備結果
            result = {
                'success': True,
                'filename': filename,
                'file_path': file_path,
                'html_path': html_result,
                'excel_path': excel_result,
                'message': f'{report_title}生成成功'
            }
            
            return result
            
        except Exception as e:
            logger.error(f"報表生成失敗: {str(e)}")
            return {
                'success': False,
                'filename': '',
                'file_path': '',
                'message': f'報表生成失敗: {str(e)}'
            }
    
    # 前一個工作日報表的專用方法
    def _get_previous_workday_date(self):
        """取得前一個工作日的日期"""
        current_datetime = timezone.localtime(timezone.now())
        current_date = current_datetime.date()
        return self.calendar_service.get_previous_workday(current_date)
    
    def _collect_previous_workday_data(self, report_date):
        """收集前一個工作日的資料"""
        from workorder.fill_work.models import FillWork
        from workorder.onsite_reporting.models import OnsiteReport
        
        # 收集填報資料
        fill_works = FillWork.objects.filter(work_date=report_date)
        
        # 收集現場報工資料
        onsite_reports = OnsiteReport.objects.filter(work_date=report_date)
        
        # 統計資料
        data = {
            'report_date': report_date,
            'fill_works_count': fill_works.count(),
            'onsite_reports_count': onsite_reports.count(),
            'fill_works': list(fill_works),
            'onsite_reports': list(onsite_reports),
            'total_work_hours': 0,
            'total_overtime_hours': 0,
            'total_operators': 0,
            'total_equipment': 0,
        }
        
        # 計算總工作時數和加班時數
        for fill_work in fill_works:
            data['total_work_hours'] += float(fill_work.work_hours_calculated or 0)
            data['total_overtime_hours'] += float(fill_work.overtime_hours_calculated or 0)
        
        for onsite_report in onsite_reports:
            data['total_work_hours'] += float(onsite_report.work_hours or 0)
        
        # 統計作業員數量（去重）
        operator_names = set()
        for fill_work in fill_works:
            if fill_work.operator:
                operator_names.add(fill_work.operator)
        
        for onsite_report in onsite_reports:
            if onsite_report.operator_name:
                operator_names.add(onsite_report.operator_name)
        
        data['total_operators'] = len(operator_names)
        
        # 統計設備數量（去重）
        equipment_ids = set()
        for fill_work in fill_works:
            if fill_work.equipment:
                equipment_ids.add(fill_work.equipment)
        
        for onsite_report in onsite_reports:
            if onsite_report.equipment_id:
                equipment_ids.add(onsite_report.equipment_id)
        
        data['total_equipment'] = len(equipment_ids)
        
        logger.info(f"收集 {report_date} 的資料完成：填報 {data['fill_works_count']} 筆，現場報工 {data['onsite_reports_count']} 筆")
        return data
    
    def _generate_previous_workday_report(self, data, schedule=None):
        """生成前一個工作日報表"""
        # 建立報表目錄
        report_dir = os.path.join(settings.MEDIA_ROOT, 'reports', 'previous_workday')
        os.makedirs(report_dir, exist_ok=True)
        
        # 取得檔案格式設定
        file_format = 'html'  # 預設為 HTML
        if schedule and hasattr(schedule, 'file_format'):
            file_format = schedule.file_format
        
        # 生成報表檔案名稱
        report_date = data['report_date']
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        
        # 根據檔案格式決定副檔名和生成方法
        if file_format == 'excel':
            filename = f"前一個工作日報表_{report_date}_{timestamp}.xlsx"
            filepath = os.path.join(report_dir, filename)
            self._generate_previous_workday_excel(data, filepath)
            
        elif file_format == 'both':
            # 同時生成 HTML 和 Excel
            html_filename = f"前一個工作日報表_{report_date}_{timestamp}.html"
            excel_filename = f"前一個工作日報表_{report_date}_{timestamp}.xlsx"
            html_filepath = os.path.join(report_dir, html_filename)
            excel_filepath = os.path.join(report_dir, excel_filename)
            
            # 生成 HTML 報表
            html_content = self._generate_previous_workday_html(data)
            with open(html_filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            # 生成 Excel 報表
            self._generate_previous_workday_excel(data, excel_filepath)
            
            # 返回 Excel 檔案路徑作為主要檔案
            filepath = excel_filepath
            filename = excel_filename
            
            # 儲存兩個檔案路徑供後續使用
            self._html_filepath = html_filepath
            self._excel_filepath = excel_filepath
            
        else:  # 預設為 HTML
            filename = f"前一個工作日報表_{report_date}_{timestamp}.html"
            filepath = os.path.join(report_dir, filename)
            
            # 生成 HTML 報表
            html_content = self._generate_previous_workday_html(data)
            with open(filepath, 'w', encoding='utf-8') as f:
                f.write(html_content)
        
        logger.info(f"前一個工作日報表已生成：{filepath}")
        return filepath
    
    def _generate_previous_workday_html(self, data):
        """生成前一個工作日報表的 HTML 內容"""
        report_date = data['report_date']
        
        html = f"""
<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>前一個工作日報表 - {report_date}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .header {{ text-align: center; margin-bottom: 30px; border-bottom: 2px solid #007bff; padding-bottom: 20px; }}
        .summary {{ margin-bottom: 30px; }}
        .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin: 20px 0; }}
        .summary-card {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px; text-align: center; }}
        .summary-card h4 {{ margin: 0 0 10px 0; font-size: 14px; opacity: 0.9; }}
        .summary-card .value {{ font-size: 24px; font-weight: bold; margin: 0; }}
        .section {{ margin: 30px 0; }}
        .section h3 {{ color: #333; border-left: 4px solid #007bff; padding-left: 15px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; background-color: white; }}
        th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
        th {{ background-color: #f8f9fa; font-weight: bold; color: #495057; }}
        tr:nth-child(even) {{ background-color: #f8f9fa; }}
        tr:hover {{ background-color: #e3f2fd; }}
        .footer {{ margin-top: 40px; text-align: center; color: #666; border-top: 1px solid #ddd; padding-top: 20px; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 前一個工作日報表</h1>
            <h2>報表日期：{report_date}</h2>
            <p>生成時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
        
        <div class="summary">
            <h3>📈 統計摘要</h3>
            <div class="summary-grid">
                <div class="summary-card">
                    <h4>填報記錄數</h4>
                    <div class="value">{data['fill_works_count']}</div>
                </div>
                <div class="summary-card">
                    <h4>現場報工記錄數</h4>
                    <div class="value">{data['onsite_reports_count']}</div>
                </div>
                <div class="summary-card">
                    <h4>正常時數</h4>
                    <div class="value">{data.get('total_work_hours', 0):.2f} 小時</div>
                </div>
                <div class="summary-card">
                    <h4>加班時數</h4>
                    <div class="value">{data.get('total_overtime_hours', 0):.2f} 小時</div>
                </div>
                <div class="summary-card">
                    <h4>總工作時數</h4>
                    <div class="value">{data['total_work_hours']:.2f} 小時</div>
                </div>
                <div class="summary-card">
                    <h4>參與作業員數</h4>
                    <div class="value">{data['total_operators']} 人</div>
                </div>
                <div class="summary-card">
                    <h4>使用設備數</h4>
                    <div class="value">{data['total_equipment']} 台</div>
                </div>
            </div>
        </div>
        
        <!-- 按公司統計 -->
        {self._generate_company_statistics_table(data)}
        
        <!-- 按工序統計 -->
        {self._generate_process_statistics_table(data)}
        
        <!-- 按作業員統計 -->
        {self._generate_operator_statistics_table(data)}
        
        <!-- 詳細資料表格 -->
        {self._generate_detailed_data_table(data)}
        
        <div class="footer">
            <p>此報表由 MES 系統自動生成</p>
        </div>
    </div>
</body>
</html>
        """
        return html
    
    def _generate_company_statistics_table(self, data):
        """生成按公司統計表格"""
        # 統計各公司資料
        company_stats = {}
        for fill_work in data['fill_works']:
            company = fill_work.company_name or '未知公司'
            if company not in company_stats:
                company_stats[company] = {
                    'count': 0,
                    'work_hours': 0,
                    'overtime_hours': 0,
                    'operators': set(),
                    'equipment': set()
                }
            
            company_stats[company]['count'] += 1
            company_stats[company]['work_hours'] += float(fill_work.work_hours_calculated or 0)
            company_stats[company]['overtime_hours'] += float(fill_work.overtime_hours_calculated or 0)
            if fill_work.operator:
                company_stats[company]['operators'].add(fill_work.operator)
            if fill_work.equipment:
                company_stats[company]['equipment'].add(fill_work.equipment)
        
        if not company_stats:
            return ""
        
        table_html = """
        <div class="section">
            <h3>🏢 按公司統計</h3>
            <table>
                <thead>
                    <tr>
                        <th>公司名稱</th>
                        <th>記錄數</th>
                        <th>正常時數</th>
                        <th>加班時數</th>
                        <th>總時數</th>
                        <th>作業員數</th>
                        <th>設備數</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for company, stats in company_stats.items():
            total_hours = stats['work_hours'] + stats['overtime_hours']
            table_html += f"""
                    <tr>
                        <td>{company}</td>
                        <td>{stats['count']}</td>
                        <td>{stats['work_hours']:.2f} 小時</td>
                        <td>{stats['overtime_hours']:.2f} 小時</td>
                        <td>{total_hours:.2f} 小時</td>
                        <td>{len(stats['operators'])}</td>
                        <td>{len(stats['equipment'])}</td>
                    </tr>
            """
        
        table_html += """
                </tbody>
            </table>
        </div>
        """
        
        return table_html
    
    def _generate_process_statistics_table(self, data):
        """生成按工序統計表格"""
        # 統計各工序資料
        process_stats = {}
        for fill_work in data['fill_works']:
            process = fill_work.process_name or '未知工序'
            if process not in process_stats:
                process_stats[process] = {
                    'count': 0,
                    'work_hours': 0,
                    'overtime_hours': 0,
                    'quantity': 0
                }
            
            process_stats[process]['count'] += 1
            process_stats[process]['work_hours'] += float(fill_work.work_hours_calculated or 0)
            process_stats[process]['overtime_hours'] += float(fill_work.overtime_hours_calculated or 0)
            process_stats[process]['quantity'] += float(fill_work.work_quantity or 0)
        
        if not process_stats:
            return ""
        
        table_html = """
        <div class="section">
            <h3>⚙️ 按工序統計</h3>
            <table>
                <thead>
                    <tr>
                        <th>工序名稱</th>
                        <th>記錄數</th>
                        <th>正常時數</th>
                        <th>加班時數</th>
                        <th>總時數</th>
                        <th>完成數量</th>
                        <th>平均效率</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for process, stats in process_stats.items():
            total_hours = stats['work_hours'] + stats['overtime_hours']
            efficiency = stats['quantity'] / total_hours if total_hours > 0 else 0
            table_html += f"""
                    <tr>
                        <td>{process}</td>
                        <td>{stats['count']}</td>
                        <td>{stats['work_hours']:.2f} 小時</td>
                        <td>{stats['overtime_hours']:.2f} 小時</td>
                        <td>{total_hours:.2f} 小時</td>
                        <td>{stats['quantity']:.0f}</td>
                        <td>{efficiency:.2f} 件/小時</td>
                    </tr>
            """
        
        table_html += """
                </tbody>
            </table>
        </div>
        """
        
        return table_html
    
    def _generate_operator_statistics_table(self, data):
        """生成按作業員統計表格"""
        # 統計各作業員資料
        operator_stats = {}
        for fill_work in data['fill_works']:
            operator = fill_work.operator or '未知作業員'
            if operator not in operator_stats:
                operator_stats[operator] = {
                    'count': 0,
                    'work_hours': 0,
                    'overtime_hours': 0,
                    'quantity': 0
                }
            
            operator_stats[operator]['count'] += 1
            operator_stats[operator]['work_hours'] += float(fill_work.work_hours_calculated or 0)
            operator_stats[operator]['overtime_hours'] += float(fill_work.overtime_hours_calculated or 0)
            operator_stats[operator]['quantity'] += float(fill_work.work_quantity or 0)
        
        if not operator_stats:
            return ""
        
        table_html = """
        <div class="section">
            <h3>👥 按作業員統計</h3>
            <table>
                <thead>
                    <tr>
                        <th>作業員</th>
                        <th>記錄數</th>
                        <th>正常時數</th>
                        <th>加班時數</th>
                        <th>總時數</th>
                        <th>完成數量</th>
                        <th>平均效率</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for operator, stats in operator_stats.items():
            total_hours = stats['work_hours'] + stats['overtime_hours']
            efficiency = stats['quantity'] / total_hours if total_hours > 0 else 0
            table_html += f"""
                    <tr>
                        <td>{operator}</td>
                        <td>{stats['count']}</td>
                        <td>{stats['work_hours']:.2f} 小時</td>
                        <td>{stats['overtime_hours']:.2f} 小時</td>
                        <td>{total_hours:.2f} 小時</td>
                        <td>{stats['quantity']:.0f}</td>
                        <td>{efficiency:.2f} 件/小時</td>
                    </tr>
            """
        
        table_html += """
                </tbody>
            </table>
        </div>
        """
        
        return table_html
    
    def _generate_detailed_data_table(self, data):
        """生成詳細資料表格"""
        if data['fill_works_count'] == 0:
            return ""
        
        table_html = """
        <div class="section">
            <h3>📋 詳細資料</h3>
            <table>
                <thead>
                    <tr>
                        <th>公司名稱</th>
                        <th>作業員</th>
                        <th>工單號碼</th>
                        <th>產品編號</th>
                        <th>工序</th>
                        <th>設備</th>
                        <th>開始時間</th>
                        <th>結束時間</th>
                        <th>正常時數</th>
                        <th>加班時數</th>
                        <th>完成數量</th>
                        <th>備註</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for fill_work in data['fill_works']:
            table_html += f"""
                    <tr>
                        <td>{fill_work.company_name or ''}</td>
                        <td>{fill_work.operator or ''}</td>
                        <td>{fill_work.workorder or ''}</td>
                        <td>{fill_work.product_id or ''}</td>
                        <td>{fill_work.process_name or ''}</td>
                        <td>{fill_work.equipment or ''}</td>
                        <td>{fill_work.start_time or ''}</td>
                        <td>{fill_work.end_time or ''}</td>
                        <td>{fill_work.work_hours_calculated or 0}</td>
                        <td>{fill_work.overtime_hours_calculated or 0}</td>
                        <td>{fill_work.work_quantity or 0}</td>
                        <td>{fill_work.remarks or ''}</td>
                    </tr>
            """
        
        table_html += """
                </tbody>
            </table>
        </div>
        """
        
        return table_html
    
    def _generate_previous_workday_excel(self, data, filepath):
        """生成前一個工作日報表的 Excel 檔案"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 建立工作簿
            wb = openpyxl.Workbook()
            
            # 移除預設工作表
            wb.remove(wb.active)
            
            # 1. 統計摘要工作表
            ws_summary = wb.create_sheet("統計摘要")
            self._create_previous_workday_summary_sheet(ws_summary, data)
            
            # 2. 按公司統計工作表
            ws_company = wb.create_sheet("按公司統計")
            self._create_previous_workday_company_sheet(ws_company, data)
            
            # 3. 按工序統計工作表
            ws_process = wb.create_sheet("按工序統計")
            self._create_previous_workday_process_sheet(ws_process, data)
            
            # 4. 按作業員統計工作表
            ws_operator = wb.create_sheet("按作業員統計")
            self._create_previous_workday_operator_sheet(ws_operator, data)
            
            # 5. 詳細資料工作表
            ws_detail = wb.create_sheet("詳細資料")
            self._create_previous_workday_detail_sheet(ws_detail, data)
            
            # 儲存檔案
            wb.save(filepath)
            logger.info(f"前一個工作日 Excel 報表已生成（5個分頁）：{filepath}")
            
        except ImportError:
            logger.error("未安裝 openpyxl，無法生成 Excel 報表")
            raise
        except Exception as e:
            logger.error(f"生成前一個工作日 Excel 報表失敗: {str(e)}")
            raise
    
    def _create_previous_workday_summary_sheet(self, ws, data):
        """建立前一個工作日報表統計摘要工作表"""
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # 標題
        ws['A1'] = f"前一個工作日報表 - {data['report_date']}"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:B1')
        
        # 生成時間
        ws['A2'] = f"生成時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].font = Font(size=10)
        
        # 統計資料
        row = 4
        stats = [
            ("填報記錄數", f"{data['fill_works_count']} 筆"),
            ("現場報工記錄數", f"{data['onsite_reports_count']} 筆"),
            ("正常時數", f"{data.get('total_work_hours', 0):.2f} 小時"),
            ("加班時數", f"{data.get('total_overtime_hours', 0):.2f} 小時"),
            ("總工作時數", f"{data['total_work_hours']:.2f} 小時"),
            ("參與作業員數", f"{data['total_operators']} 人"),
            ("使用設備數", f"{data['total_equipment']} 台"),
        ]
        
        for label, value in stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
        
        # 調整欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_previous_workday_company_sheet(self, ws, data):
        """建立前一個工作日報表按公司統計工作表"""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        # 標題
        ws['A1'] = "按公司統計"
        ws['A1'].font = Font(size=14, bold=True)
        
        # 表頭
        headers = ["公司名稱", "記錄數", "正常時數", "加班時數", "總時數", "作業員數", "設備數"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 按公司統計邏輯
        row = 3
        if data.get('fill_works') or data.get('onsite_reports'):
            # 按公司分組統計
            company_stats = {}
            
            # 統計填報資料
            for fill_work in data.get('fill_works', []):
                company_code = getattr(fill_work, 'company_name', '未知公司')
                if company_code not in company_stats:
                    company_stats[company_code] = {
                        'records': 0,
                        'normal_hours': 0,
                        'overtime_hours': 0,
                        'total_hours': 0,
                        'operators': set(),
                        'equipment': set()
                    }
                
                company_stats[company_code]['records'] += 1
                company_stats[company_code]['normal_hours'] += float(fill_work.work_hours_calculated or 0)
                company_stats[company_code]['overtime_hours'] += float(fill_work.overtime_hours_calculated or 0)
                company_stats[company_code]['total_hours'] += float(fill_work.work_hours_calculated or 0) + float(fill_work.overtime_hours_calculated or 0)
                
                if fill_work.operator:
                    company_stats[company_code]['operators'].add(fill_work.operator)
                if fill_work.equipment:
                    company_stats[company_code]['equipment'].add(fill_work.equipment)
            
            # 統計現場報工資料
            for onsite_report in data.get('onsite_reports', []):
                company_code = getattr(onsite_report, 'company', '未知公司')
                if company_code not in company_stats:
                    company_stats[company_code] = {
                        'records': 0,
                        'normal_hours': 0,
                        'overtime_hours': 0,
                        'total_hours': 0,
                        'operators': set(),
                        'equipment': set()
                    }
                
                company_stats[company_code]['records'] += 1
                company_stats[company_code]['normal_hours'] += float(onsite_report.work_hours or 0)
                company_stats[company_code]['total_hours'] += float(onsite_report.work_hours or 0)
                
                if onsite_report.operator_name:
                    company_stats[company_code]['operators'].add(onsite_report.operator_name)
                if onsite_report.equipment_id:
                    company_stats[company_code]['equipment'].add(onsite_report.equipment_id)
            
            # 寫入統計資料
            for company_code, stats in company_stats.items():
                ws[f'A{row}'] = company_code
                ws[f'B{row}'] = stats['records']
                ws[f'C{row}'] = f"{stats['normal_hours']:.2f}"
                ws[f'D{row}'] = f"{stats['overtime_hours']:.2f}"
                ws[f'E{row}'] = f"{stats['total_hours']:.2f}"
                ws[f'F{row}'] = len(stats['operators'])
                ws[f'G{row}'] = len(stats['equipment'])
                row += 1
        else:
            # 沒有詳細資料時，按公司統計應該顯示無資料
            ws[f'A{row}'] = "無資料"
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
        
        # 調整欄寬
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_previous_workday_detail_sheet(self, ws, data):
        """建立前一個工作日報表詳細資料工作表"""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        # 標題
        ws['A1'] = "詳細資料"
        ws['A1'].font = Font(size=14, bold=True)
        
        # 表頭
        headers = ["類型", "工單號", "作業員", "設備", "工作時數", "加班時數", "公司"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 填報資料
        row = 3
        for fill_work in data['fill_works']:
            ws.cell(row=row, column=1, value="填報")
            ws.cell(row=row, column=2, value=fill_work.workorder)
            ws.cell(row=row, column=3, value=fill_work.operator)
            ws.cell(row=row, column=4, value=fill_work.equipment)
            ws.cell(row=row, column=5, value=fill_work.work_hours_calculated)
            ws.cell(row=row, column=6, value=fill_work.overtime_hours_calculated)
            ws.cell(row=row, column=7, value=fill_work.company_name)
            row += 1
        
        # 現場報工資料
        for onsite_report in data['onsite_reports']:
            ws.cell(row=row, column=1, value="現場報工")
            ws.cell(row=row, column=2, value=onsite_report.workorder_id)
            ws.cell(row=row, column=3, value=onsite_report.operator_name)
            ws.cell(row=row, column=4, value=onsite_report.equipment_id)
            ws.cell(row=row, column=5, value=onsite_report.work_hours)
            ws.cell(row=row, column=6, value="")
            ws.cell(row=row, column=7, value=onsite_report.company)
            row += 1
        
        # 調整欄寬
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_previous_workday_operator_sheet(self, ws, data):
        """建立前一個工作日報表按作業員統計工作表"""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        # 標題
        ws['A1'] = "按作業員統計"
        ws['A1'].font = Font(size=14, bold=True)
        
        # 表頭
        headers = ["作業員", "記錄數", "正常時數", "加班時數", "總時數", "設備數", "公司"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 按作業員統計
        row = 3
        if data.get('fill_works') or data.get('onsite_reports'):
            # 按作業員分組統計
            operator_stats = {}
            
            # 統計填報資料
            for fill_work in data.get('fill_works', []):
                operator = getattr(fill_work, 'operator', '未知作業員')
                if operator not in operator_stats:
                    operator_stats[operator] = {
                        'records': 0,
                        'normal_hours': 0,
                        'overtime_hours': 0,
                        'total_hours': 0,
                        'equipment': set(),
                        'company': getattr(fill_work, 'company_name', '未知公司')
                    }
                
                operator_stats[operator]['records'] += 1
                operator_stats[operator]['normal_hours'] += float(fill_work.work_hours_calculated or 0)
                operator_stats[operator]['overtime_hours'] += float(fill_work.overtime_hours_calculated or 0)
                operator_stats[operator]['total_hours'] += float(fill_work.work_hours_calculated or 0) + float(fill_work.overtime_hours_calculated or 0)
                
                if fill_work.equipment:
                    operator_stats[operator]['equipment'].add(fill_work.equipment)
            
            # 統計現場報工資料
            for onsite_report in data.get('onsite_reports', []):
                operator = getattr(onsite_report, 'operator_name', '未知作業員')
                if operator not in operator_stats:
                    operator_stats[operator] = {
                        'records': 0,
                        'normal_hours': 0,
                        'overtime_hours': 0,
                        'total_hours': 0,
                        'equipment': set(),
                        'company': getattr(onsite_report, 'company', '未知公司')
                    }
                
                operator_stats[operator]['records'] += 1
                operator_stats[operator]['normal_hours'] += float(onsite_report.work_hours or 0)
                operator_stats[operator]['total_hours'] += float(onsite_report.work_hours or 0)
                
                if onsite_report.equipment_id:
                    operator_stats[operator]['equipment'].add(onsite_report.equipment_id)
            
            # 寫入統計資料
            for operator, stats in operator_stats.items():
                ws.cell(row=row, column=1, value=operator)
                ws.cell(row=row, column=2, value=stats['records'])
                ws.cell(row=row, column=3, value=f"{stats['normal_hours']:.2f}")
                ws.cell(row=row, column=4, value=f"{stats['overtime_hours']:.2f}")
                ws.cell(row=row, column=5, value=f"{stats['total_hours']:.2f}")
                ws.cell(row=row, column=6, value=len(stats['equipment']))
                ws.cell(row=row, column=7, value=stats['company'])
                row += 1
        else:
            # 無資料時顯示
            ws.cell(row=row, column=1, value="無資料")
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
        
        # 調整欄寬
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_previous_workday_process_sheet(self, ws, data):
        """建立前一個工作日報表按工序統計工作表"""
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        
        # 標題
        ws['A1'] = "按工序統計"
        ws['A1'].font = Font(size=14, bold=True)
        
        # 表頭
        headers = ["工序", "記錄數", "正常時數", "加班時數", "總時數", "作業員數", "公司"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 按工序統計
        row = 3
        if data.get('fill_works') or data.get('onsite_reports'):
            # 按工序分組統計
            process_stats = {}
            
            # 統計填報資料
            for fill_work in data.get('fill_works', []):
                process = getattr(fill_work, 'process_name', '未知工序')
                if process not in process_stats:
                    process_stats[process] = {
                        'records': 0,
                        'normal_hours': 0,
                        'overtime_hours': 0,
                        'total_hours': 0,
                        'operators': set(),
                        'company': getattr(fill_work, 'company_name', '未知公司')
                    }
                
                process_stats[process]['records'] += 1
                process_stats[process]['normal_hours'] += float(fill_work.work_hours_calculated or 0)
                process_stats[process]['overtime_hours'] += float(fill_work.overtime_hours_calculated or 0)
                process_stats[process]['total_hours'] += float(fill_work.work_hours_calculated or 0) + float(fill_work.overtime_hours_calculated or 0)
                
                if fill_work.operator:
                    process_stats[process]['operators'].add(fill_work.operator)
            
            # 統計現場報工資料
            for onsite_report in data.get('onsite_reports', []):
                process = getattr(onsite_report, 'process_name', '未知工序')
                if process not in process_stats:
                    process_stats[process] = {
                        'records': 0,
                        'normal_hours': 0,
                        'overtime_hours': 0,
                        'total_hours': 0,
                        'operators': set(),
                        'company': getattr(onsite_report, 'company', '未知公司')
                    }
                
                process_stats[process]['records'] += 1
                process_stats[process]['normal_hours'] += float(onsite_report.work_hours or 0)
                process_stats[process]['total_hours'] += float(onsite_report.work_hours or 0)
                
                if onsite_report.operator_name:
                    process_stats[process]['operators'].add(onsite_report.operator_name)
            
            # 寫入統計資料
            for process, stats in process_stats.items():
                ws.cell(row=row, column=1, value=process)
                ws.cell(row=row, column=2, value=stats['records'])
                ws.cell(row=row, column=3, value=f"{stats['normal_hours']:.2f}")
                ws.cell(row=row, column=4, value=f"{stats['overtime_hours']:.2f}")
                ws.cell(row=row, column=5, value=f"{stats['total_hours']:.2f}")
                ws.cell(row=row, column=6, value=len(stats['operators']))
                ws.cell(row=row, column=7, value=stats['company'])
                row += 1
        else:
            # 無資料時顯示
            ws.cell(row=row, column=1, value="無資料")
            ws.merge_cells(f'A{row}:G{row}')
            row += 1
        
        # 調整欄寬
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 通用報表生成方法（需要從原 services.py 移過來）
    def _collect_workorder_data(self, start_date, end_date, company):
        """收集工單資料"""
        try:
            from workorder.models import WorkOrderReportData
            from erp_integration.models import CompanyConfig
            
            # 如果沒有指定公司，則查詢所有公司
            if company and company != 'all':
                # 根據公司代號查找對應的資料庫
                try:
                    company_config = CompanyConfig.objects.get(company_code=company)
                    queryset = WorkOrderReportData.objects.filter(
                        work_date__range=[start_date, end_date]
                    )
                except CompanyConfig.DoesNotExist:
                    logger.warning(f"找不到公司配置: {company}")
                    queryset = WorkOrderReportData.objects.filter(
                        work_date__range=[start_date, end_date]
                    )
            else:
                # 查詢所有公司的資料
                queryset = WorkOrderReportData.objects.filter(
                    work_date__range=[start_date, end_date]
                )
            
            # 轉換為字典格式，方便報表生成使用
            data = []
            for record in queryset:
                data.append({
                    'workorder_number': record.workorder_number,
                    'operator_name': record.operator_name,
                    'operator': record.operator,
                    'work_date': record.work_date,
                    'start_time': record.start_time,
                    'end_time': record.end_time,
                    'work_hours': record.work_hours,
                    'work_hours_calculated': record.work_hours_calculated,
                    'equipment_hours': record.equipment_hours,
                    'status': record.status,
                    'company_code': record.company_code,
                    'created_at': record.created_at,
                    'updated_at': record.updated_at
                })
            
            return data
            
        except Exception as e:
            logger.error(f"收集工單資料失敗: {str(e)}")
            return []
    
    def _generate_html_report(self, data, report_title, schedule):
        """生成 HTML 報表"""
        try:
            # 計算統計資料
            total_records = len(data) if data else 0
            total_work_hours = sum(float(record.get('work_hours', 0) or 0) for record in data) if data else 0
            total_equipment_hours = sum(float(record.get('equipment_hours', 0) or 0) for record in data) if data else 0
            
            # 統計作業員數量（去重）
            operator_names = set()
            for record in data:
                operator_name = record.get('operator_name') or record.get('operator')
                if operator_name:
                    operator_names.add(operator_name)
            total_operators = len(operator_names)
            
            # 生成詳細統計資料
            detailed_stats = self._generate_weekly_detailed_statistics(data)
            
            # 生成HTML內容（使用與前一個工作日報表相同的格式）
            html_content = f"""
<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report_title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .header {{ text-align: center; margin-bottom: 30px; border-bottom: 2px solid #007bff; padding-bottom: 20px; }}
        .summary {{ margin-bottom: 30px; }}
        .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin: 20px 0; }}
        .summary-card {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px; text-align: center; }}
        .summary-card h4 {{ margin: 0 0 10px 0; font-size: 14px; opacity: 0.9; }}
        .summary-card .value {{ font-size: 24px; font-weight: bold; margin: 0; }}
        .section {{ margin: 30px 0; }}
        .section h3 {{ color: #333; border-left: 4px solid #007bff; padding-left: 15px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; background-color: white; }}
        th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
        th {{ background-color: #f8f9fa; font-weight: bold; color: #495057; }}
        tr:nth-child(even) {{ background-color: #f8f9fa; }}
        tr:hover {{ background-color: #e3f2fd; }}
        .footer {{ margin-top: 40px; text-align: center; color: #666; border-top: 1px solid #ddd; padding-top: 20px; }}
        .chart-container {{ margin: 20px 0; text-align: center; }}
        .status-badge {{ display: inline-block; padding: 4px 8px; border-radius: 4px; font-size: 12px; font-weight: bold; }}
        .status-active {{ background-color: #d4edda; color: #155724; }}
        .status-pending {{ background-color: #fff3cd; color: #856404; }}
        .status-completed {{ background-color: #d1ecf1; color: #0c5460; }}
    </style>
</head>
<body>
    <div class="container">
    <div class="header">
            <h1>📊 {report_title}</h1>
        <p>生成時間：{timezone.now().strftime('%Y-%m-%d %H:%M:%S')} | 排程名稱：{schedule.name}</p>
    </div>
    
    <div class="summary">
            <h3>📈 統計摘要</h3>
            <div class="summary-grid">
                <div class="summary-card">
                    <h4>總記錄數</h4>
                    <div class="value">{total_records}</div>
        </div>
                <div class="summary-card">
                    <h4>總工作時數</h4>
                    <div class="value">{total_work_hours:.2f} 小時</div>
        </div>
                <div class="summary-card">
                    <h4>設備使用時數</h4>
                    <div class="value">{total_equipment_hours:.2f} 小時</div>
        </div>
                <div class="summary-card">
                    <h4>參與作業員數</h4>
                    <div class="value">{total_operators} 人</div>
        </div>
                <div class="summary-card">
                    <h4>平均每日工作時數</h4>
                    <div class="value">{(total_work_hours/7):.2f} 小時</div>
                </div>
        </div>
    </div>
        
        {detailed_stats}
    
    <div class="footer">
            <p>此報表由 MES 系統自動生成 | 生成時間：{timezone.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>
        """
        
            # 生成檔案名稱
            filename = f"{report_title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.html"
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # 確保目錄存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 寫入檔案
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            return file_path
            
        except Exception as e:
            logger.error(f"生成 HTML 報表失敗: {str(e)}")
            return None
    
    def _generate_weekly_detailed_statistics(self, data):
        """生成週報表詳細統計資料"""
        if not data:
            return "<div class='section'><h3>📋 詳細資料</h3><p>無資料</p></div>"
        
        # 按日期分組統計
        daily_stats = {}
        operator_stats = {}
        
        for record in data:
            # 取得工作日期
            work_date = None
            if hasattr(record, 'work_date') and record.work_date:
                work_date = record.work_date
            elif hasattr(record, 'created_at') and record.created_at:
                work_date = record.created_at.date()
            
            date_str = work_date.strftime('%Y-%m-%d') if work_date else '未知'
            
            # 日期統計
            if date_str not in daily_stats:
                daily_stats[date_str] = {
                    'records': 0,
                    'work_hours': 0,
                    'equipment_hours': 0,
                    'operators': set()
                }
            
            daily_stats[date_str]['records'] += 1
            
            # 工作時數
            work_hours = 0
            if hasattr(record, 'work_hours') and record.work_hours:
                work_hours = float(record.work_hours)
            elif hasattr(record, 'work_hours_calculated') and record.work_hours_calculated:
                work_hours = float(record.work_hours_calculated)
            
            daily_stats[date_str]['work_hours'] += work_hours
            
            # 設備時數
            equipment_hours = 0
            if hasattr(record, 'equipment_hours') and record.equipment_hours:
                equipment_hours = float(record.equipment_hours)
            
            daily_stats[date_str]['equipment_hours'] += equipment_hours
            
            # 作業員
            operator_name = None
            if hasattr(record, 'operator_name') and record.operator_name:
                operator_name = record.operator_name
            elif hasattr(record, 'operator') and record.operator:
                operator_name = record.operator
            
            if operator_name:
                daily_stats[date_str]['operators'].add(operator_name)
            
            # 作業員統計
            if operator_name:
                if operator_name not in operator_stats:
                    operator_stats[operator_name] = {
                        'records': 0,
                        'work_hours': 0,
                        'equipment_hours': 0
                    }
                operator_stats[operator_name]['records'] += 1
                operator_stats[operator_name]['work_hours'] += work_hours
                operator_stats[operator_name]['equipment_hours'] += equipment_hours
        
        html = ""
        
        # 每日統計
        if daily_stats:
            html += """
    <div class="section">
        <h3>📅 每日統計</h3>
        <table>
            <thead>
                <tr>
                    <th>日期</th>
                    <th>記錄數</th>
                    <th>工作時數</th>
                    <th>設備使用時數</th>
                    <th>作業員數</th>
                </tr>
            </thead>
            <tbody>
            """
            
            for date_str in sorted(daily_stats.keys()):
                stats = daily_stats[date_str]
                html += f"""
                <tr>
                    <td>{date_str}</td>
                    <td>{stats['records']}</td>
                    <td>{stats['work_hours']:.2f}</td>
                    <td>{stats['equipment_hours']:.2f}</td>
                    <td>{len(stats['operators'])}</td>
                </tr>
                """
            
            html += """
            </tbody>
        </table>
    </div>
            """
        
        # 作業員統計
        if operator_stats:
            html += """
    <div class="section">
        <h3>👥 作業員統計</h3>
        <table>
            <thead>
                <tr>
                    <th>作業員</th>
                    <th>記錄數</th>
                    <th>工作時數</th>
                    <th>設備使用時數</th>
                </tr>
            </thead>
            <tbody>
            """
            
            for operator in sorted(operator_stats.keys()):
                stats = operator_stats[operator]
                html += f"""
                <tr>
                    <td>{operator}</td>
                    <td>{stats['records']}</td>
                    <td>{stats['work_hours']:.2f}</td>
                    <td>{stats['equipment_hours']:.2f}</td>
                </tr>
                """
            
            html += """
            </tbody>
        </table>
    </div>
            """
        
        # 詳細記錄
        html += """
    <div class="section">
        <h3>📋 詳細記錄</h3>
        <table>
            <thead>
                <tr>
                    <th>工單編號</th>
                    <th>工作日期</th>
                    <th>作業員</th>
                    <th>產品編號</th>
                    <th>工序名稱</th>
                    <th>開始時間</th>
                    <th>結束時間</th>
                    <th>工作時數</th>
                    <th>設備使用時數</th>
                </tr>
            </thead>
            <tbody>
        """
        
        for record in data:
            # 取得各種欄位值
            workorder_id = getattr(record, 'workorder_id', None) or getattr(record, 'id', '-')
            work_date = getattr(record, 'work_date', None)
            operator_name = getattr(record, 'operator_name', None) or getattr(record, 'operator', '-')
            product_code = getattr(record, 'product_code', None) or getattr(record, 'product_id', '-')
            process_name = getattr(record, 'process_name', None) or getattr(record, 'process', '-')
            start_time = getattr(record, 'start_time', None)
            end_time = getattr(record, 'end_time', None)
            work_hours = getattr(record, 'work_hours', None) or getattr(record, 'work_hours_calculated', '-')
            equipment_hours = getattr(record, 'equipment_hours', None) or '-'
            
            html += f"""
                <tr>
                    <td>{workorder_id}</td>
                    <td>{work_date.strftime('%Y-%m-%d') if work_date else '-'}</td>
                    <td>{operator_name}</td>
                    <td>{product_code}</td>
                    <td>{process_name}</td>
                    <td>{start_time.strftime('%H:%M') if start_time else '-'}</td>
                    <td>{end_time.strftime('%H:%M') if end_time else '-'}</td>
                    <td>{work_hours}</td>
                    <td>{equipment_hours}</td>
                </tr>
            """
        
        html += """
            </tbody>
        </table>
    </div>
        """
        
        return html

    def _generate_excel_report(self, data, report_title, schedule):
        """生成 Excel 報表"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 建立工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "報表資料"
            
            # 設定標題樣式
            title_font = Font(name='微軟正黑體', size=16, bold=True, color='FFFFFF')
            title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            data_font = Font(name='微軟正黑體', size=11)
            
            # 設定邊框
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 寫入標題
            ws.merge_cells('A1:H1')
            ws['A1'] = report_title
            ws['A1'].font = title_font
            ws['A1'].fill = title_fill
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # 寫入生成時間
            ws.merge_cells('A2:H2')
            ws['A2'] = f"生成時間：{timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws['A2'].font = data_font
            ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
            
            # 寫入排程資訊
            ws.merge_cells('A3:H3')
            ws['A3'] = f"排程名稱：{schedule.name}"
            ws['A3'].font = data_font
            ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
            
            # 空行
            current_row = 5
            
            # 計算統計資料
            total_records = len(data) if data else 0
            total_work_hours = sum(float(record.get('work_hours', 0) or 0) for record in data) if data else 0
            total_equipment_hours = sum(float(record.get('equipment_hours', 0) or 0) for record in data) if data else 0
            
            # 統計作業員數量（去重）
            operator_names = set()
            for record in data:
                operator_name = record.get('operator_name') or record.get('operator')
                if operator_name:
                    operator_names.add(operator_name)
            total_operators = len(operator_names)
            
            # 寫入統計摘要
            ws.merge_cells(f'A{current_row}:H{current_row}')
            ws[f'A{current_row}'] = "📈 統計摘要"
            ws[f'A{current_row}'].font = header_font
            ws[f'A{current_row}'].fill = header_fill
            ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
            
            # 統計資料
            stats_data = [
                ['總記錄數', total_records],
                ['總工作時數', f"{total_work_hours:.2f} 小時"],
                ['設備使用時數', f"{total_equipment_hours:.2f} 小時"],
                ['參與作業員數', f"{total_operators} 人"],
                ['平均每日工作時數', f"{(total_work_hours/7):.2f} 小時" if total_work_hours > 0 else "0 小時"]
            ]
            
            for stat in stats_data:
                ws[f'A{current_row}'] = stat[0]
                ws[f'B{current_row}'] = stat[1]
                ws[f'A{current_row}'].font = data_font
                ws[f'B{current_row}'].font = data_font
                ws[f'A{current_row}'].border = thin_border
                ws[f'B{current_row}'].border = thin_border
                current_row += 1
            
            current_row += 1
            
            # 寫入詳細資料標題
            if data:
                ws.merge_cells(f'A{current_row}:H{current_row}')
                ws[f'A{current_row}'] = "📋 詳細資料"
                ws[f'A{current_row}'].font = header_font
                ws[f'A{current_row}'].fill = header_fill
                ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
                
                # 寫入表頭
                headers = ['工單編號', '作業員', '工作日期', '開始時間', '結束時間', '工作時數', '設備時數', '狀態']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
                current_row += 1
                
                # 寫入資料
                for record in data:
                    row_data = [
                        record.get('workorder_number', ''),
                        record.get('operator_name') or record.get('operator', ''),
                        record.get('work_date', '').strftime('%Y-%m-%d') if record.get('work_date') else '',
                        record.get('start_time', '').strftime('%H:%M') if record.get('start_time') else '',
                        record.get('end_time', '').strftime('%H:%M') if record.get('end_time') else '',
                        f"{float(record.get('work_hours', 0) or 0):.2f}",
                        f"{float(record.get('equipment_hours', 0) or 0):.2f}",
                        record.get('status', '')
                    ]
                    
                    for col, value in enumerate(row_data, 1):
                        cell = ws.cell(row=current_row, column=col, value=value)
                        cell.font = data_font
                        cell.border = thin_border
                        if col in [6, 7]:  # 數值欄位右對齊
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                        else:
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    current_row += 1
            else:
                ws.merge_cells(f'A{current_row}:H{current_row}')
                ws[f'A{current_row}'] = "無資料"
                ws[f'A{current_row}'].font = data_font
                ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
                current_row += 1
            
            # 設定欄寬
            column_widths = [15, 12, 12, 10, 10, 12, 12, 10]
            for col, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(col)].width = width
            
            # 生成檔案名稱
            filename = f"{report_title.replace(' ', '_')}_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # 確保目錄存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 儲存檔案
            wb.save(file_path)
            
            return file_path
            
        except Exception as e:
            logger.error(f"生成 Excel 報表失敗: {str(e)}")
            return None


class WorkHourReportService:
    """工作時數報表服務"""
    
    def __init__(self, company_code=None):
        self.company_code = company_code
    
    def get_daily_report(self, company_code, date):
        """日報表"""
        from workorder.models import WorkOrderReportData
        return WorkOrderReportData.objects.filter(
            company=company_code,
            work_date=date
        )
    
    def get_daily_report_by_company_operator(self, company, operator, date):
        """按公司和作業員查詢日報表"""
        from workorder.models import WorkOrderReportData
        queryset = WorkOrderReportData.objects.filter(work_date=date)
        
        if company != 'all':
            queryset = queryset.filter(company=company)
        
        if operator != 'all':
            queryset = queryset.filter(operator_name=operator)
        
        return queryset
    
    def get_daily_summary_by_company_operator(self, company, operator, date):
        """按公司和作業員查詢日報表摘要"""
        from django.db.models import Sum, Avg
        from decimal import Decimal
        
        data = self.get_daily_report_by_company_operator(company, operator, date)
        
        # 計算總工作時數
        total_work_hours = data.aggregate(total=Sum('daily_work_hours'))['total'] or Decimal('0')
        
        # 計算作業員數量（去重）
        if operator == 'all':
            total_operators = data.exclude(operator_name__isnull=True).values('operator_name').distinct().count()
        else:
            total_operators = 1  # 特定作業員
        
        # 計算總設備使用時數
        total_equipment_hours = data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0')
        
        # 計算工單數量
        workorder_count = data.count()
        
        # 計算平均日工作時數
        avg_daily_hours = data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0')
        
        summary = {
            'total_work_hours': total_work_hours,
            'total_operators': total_operators,
            'total_equipment_hours': total_equipment_hours,
            'workorder_count': workorder_count,
            'avg_daily_hours': avg_daily_hours,
        }
        
        return summary
    
    def get_operator_statistics(self, data):
        """生成作業員統計資料"""
        from django.db.models import Sum
        
        if not data:
            return []
        
        # 按作業員分組統計
        operator_stats = data.values('operator_name').annotate(
            total_work_hours=Sum('daily_work_hours'),
            total_overtime_hours=Sum('overtime_hours'),
            total_hours=Sum('daily_work_hours') + Sum('overtime_hours')
        ).order_by('-total_hours')
        
        # 格式化統計資料
        stats_list = []
        for i, stat in enumerate(operator_stats, 1):
            stats_list.append({
                'rank': i,
                'operator_name': stat['operator_name'] or '未指定',
                'work_hours': float(stat['total_work_hours'] or 0),
                'overtime_hours': float(stat['total_overtime_hours'] or 0),
                'total_hours': float(stat['total_hours'] or 0),
            })
        
        return stats_list
    
    def get_custom_report_by_company_operator(self, company, operator, start_date, end_date):
        """按公司和作業員查詢自訂報表"""
        from workorder.models import WorkOrderReportData
        queryset = WorkOrderReportData.objects.filter(
            work_date__range=[start_date, end_date]
        )
        
        if company != 'all':
            queryset = queryset.filter(company=company)
        
        if operator != 'all':
            queryset = queryset.filter(operator_name=operator)
        
        return queryset.order_by('work_date', 'operator_name')
    
    def get_custom_summary_by_company_operator(self, company, operator, start_date, end_date):
        """按公司和作業員查詢自訂報表摘要"""
        from django.db.models import Sum, Avg
        from decimal import Decimal
        
        data = self.get_custom_report_by_company_operator(company, operator, start_date, end_date)
        
        # 計算總工作時數
        total_work_hours = data.aggregate(total=Sum('daily_work_hours'))['total'] or Decimal('0')
        
        # 計算作業員數量（去重）
        if operator == 'all':
            total_operators = data.exclude(operator_name__isnull=True).values('operator_name').distinct().count()
        else:
            total_operators = 1  # 特定作業員
        
        # 計算總設備使用時數
        total_equipment_hours = data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0')
        
        # 計算工單數量
        workorder_count = data.count()
        
        # 計算平均日工作時數
        avg_daily_hours = data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0')
        
        summary = {
            'total_work_hours': total_work_hours,
            'total_operators': total_operators,
            'total_equipment_hours': total_equipment_hours,
            'workorder_count': workorder_count,
            'avg_daily_hours': avg_daily_hours,
        }
        
        return summary
    
    def get_weekly_report(self, company_code, year, week):
        """週報表"""
        from workorder.models import WorkOrderReportData
        return WorkOrderReportData.objects.filter(
            company=company_code,
            work_year=year,
            work_week=week
        )
    
    def get_monthly_report(self, company_code, year, month):
        """月報表"""
        from workorder.models import WorkOrderReportData
        return WorkOrderReportData.objects.filter(
            company=company_code,
            work_year=year,
            work_month=month
        )
    
    def get_quarterly_report(self, company_code, year, quarter):
        """季報表"""
        from workorder.models import WorkOrderReportData
        return WorkOrderReportData.objects.filter(
            company=company_code,
            work_year=year,
            work_quarter=quarter
        )
    
    def get_yearly_report(self, company_code, year):
        """年報表"""
        from workorder.models import WorkOrderReportData
        return WorkOrderReportData.objects.filter(
            company=company_code,
            work_year=year
        )
    
    def get_daily_summary(self, company_code, date):
        """日報表摘要"""
        from django.db.models import Sum, Avg
        from decimal import Decimal
        
        data = self.get_daily_report(company_code, date)
        
        summary = {
            'total_work_hours': data.aggregate(total=Sum('daily_work_hours'))['total'] or Decimal('0'),
            'total_operators': data.aggregate(total=Sum('operator_count'))['total'] or 0,
            'total_equipment_hours': data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0'),
            'workorder_count': data.count(),
            'avg_daily_hours': data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0'),
        }
        
        return summary
    
    def get_weekly_summary(self, company_code, year, week):
        """週報表摘要"""
        from django.db.models import Sum, Avg
        from decimal import Decimal
        
        data = self.get_weekly_report(company_code, year, week)
        
        summary = {
            'total_work_hours': data.aggregate(total=Sum('daily_work_hours'))['total'] or Decimal('0'),
            'total_operators': data.aggregate(total=Sum('operator_count'))['total'] or 0,
            'total_equipment_hours': data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0'),
            'workorder_count': data.count(),
            'avg_daily_hours': data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0'),
        }
        
        return summary
    
    def get_monthly_summary(self, company_code, year, month):
        """月報表摘要"""
        from django.db.models import Sum, Avg
        from decimal import Decimal
        
        data = self.get_monthly_report(company_code, year, month)
        
        summary = {
            'total_work_hours': data.aggregate(total=Sum('monthly_work_hours'))['total'] or Decimal('0'),
            'total_operators': data.aggregate(total=Sum('operator_count'))['total'] or 0,
            'total_equipment_hours': data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0'),
            'workorder_count': data.count(),
            'avg_daily_hours': data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0'),
        }
        
        return summary
    
    def get_quarterly_summary(self, company_code, year, quarter):
        """季報表摘要"""
        from django.db.models import Sum, Avg
        from decimal import Decimal
        
        data = self.get_quarterly_report(company_code, year, quarter)
        
        summary = {
            'total_work_hours': data.aggregate(total=Sum('monthly_work_hours'))['total'] or Decimal('0'),
            'total_operators': data.aggregate(total=Sum('operator_count'))['total'] or 0,
            'total_equipment_hours': data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0'),
            'workorder_count': data.count(),
            'avg_daily_hours': data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0'),
        }
        
        return summary
    
    def get_yearly_summary(self, company_code, year):
        """年報表摘要"""
        from django.db.models import Sum, Avg
        from decimal import Decimal
        
        data = self.get_yearly_report(company_code, year)
        
        summary = {
            'total_work_hours': data.aggregate(total=Sum('monthly_work_hours'))['total'] or Decimal('0'),
            'total_operators': data.aggregate(total=Sum('operator_count'))['total'] or 0,
            'total_equipment_hours': data.aggregate(total=Sum('equipment_hours'))['total'] or Decimal('0'),
            'workorder_count': data.count(),
            'avg_daily_hours': data.aggregate(avg=Avg('daily_work_hours'))['avg'] or Decimal('0'),
        }
        
        return summary


class ReportGeneratorService:
    """報表生成服務"""
    
    def __init__(self):
        self.work_hour_service = WorkHourReportService()
    
    def generate_daily_report_by_operator(self, operator, date, format='preview'):
        """按作業員生成日報表"""
        try:
            data = self.work_hour_service.get_daily_report_by_company_operator('all', operator, date)
            summary = self.work_hour_service.get_daily_summary_by_company_operator('all', operator, date)
            
            report_data = {
                'report_type': '日報表',
                'operator': operator,
                'date': date,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
            }
            
            if format == 'excel':
                return self._export_to_excel(report_data, f'作業員日報表_{date}')
            elif format == 'csv':
                return self._export_to_csv(report_data, f'作業員日報表_{date}')
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成日報表失敗: {str(e)}")
            raise
    
    def generate_daily_report(self, company_code, date, format='excel'):
        """生成日報表"""
        try:
            data = self.work_hour_service.get_daily_report(company_code, date)
            summary = self.work_hour_service.get_daily_summary(company_code, date)
            
            report_data = {
                'report_type': '日報表',
                'company_code': company_code,
                'date': date,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
            }
            
            if format == 'excel':
                return self._export_to_excel(report_data, f'日報表_{company_code}_{date}')
            elif format == 'csv':
                return self._export_to_csv(report_data, f'日報表_{company_code}_{date}')
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成日報表失敗: {str(e)}")
            raise
    
    def generate_weekly_report(self, company_code, year, week, format='excel'):
        """生成週報表"""
        try:
            data = self.work_hour_service.get_weekly_report(company_code, year, week)
            summary = self.work_hour_service.get_weekly_summary(company_code, year, week)
            
            report_data = {
                'report_type': '週報表',
                'company_code': company_code,
                'year': year,
                'week': week,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
            }
            
            if format == 'excel':
                return self._export_to_excel(report_data, f'週報表_{company_code}_{year}_{week}')
            elif format == 'csv':
                return self._export_to_csv(report_data, f'週報表_{company_code}_{year}_{week}')
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成週報表失敗: {str(e)}")
            raise
    
    def generate_monthly_report(self, company_code, year, month, format='excel'):
        """生成月報表"""
        try:
            data = self.work_hour_service.get_monthly_report(company_code, year, month)
            summary = self.work_hour_service.get_monthly_summary(company_code, year, month)
            
            report_data = {
                'report_type': '月報表',
                'company_code': company_code,
                'year': year,
                'month': month,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
            }
            
            if format == 'excel':
                return self._export_to_excel(report_data, f'月報表_{company_code}_{year}_{month}')
            elif format == 'csv':
                return self._export_to_csv(report_data, f'月報表_{company_code}_{year}_{month}')
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成月報表失敗: {str(e)}")
            raise
    
    def generate_quarterly_report(self, company_code, year, quarter, format='excel'):
        """生成季報表"""
        try:
            data = self.work_hour_service.get_quarterly_report(company_code, year, quarter)
            summary = self.work_hour_service.get_quarterly_summary(company_code, year, quarter)
            
            report_data = {
                'report_type': '季報表',
                'company_code': company_code,
                'year': year,
                'quarter': quarter,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
            }
            
            if format == 'excel':
                return self._export_to_excel(report_data, f'季報表_{company_code}_{year}_{quarter}')
            elif format == 'csv':
                return self._export_to_csv(report_data, f'季報表_{company_code}_{year}_{quarter}')
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成季報表失敗: {str(e)}")
            raise
    
    def generate_yearly_report(self, company_code, year, format='excel'):
        """生成年報表"""
        try:
            data = self.work_hour_service.get_yearly_report(company_code, year)
            summary = self.work_hour_service.get_yearly_summary(company_code, year)
            
            report_data = {
                'report_type': '年報表',
                'company_code': company_code,
                'year': year,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
            }
            
            if format == 'excel':
                return self._export_to_excel(report_data, f'年報表_{company_code}_{year}')
            elif format == 'csv':
                return self._export_to_csv(report_data, f'年報表_{company_code}_{year}')
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成年報表失敗: {str(e)}")
            raise
    
    def generate_custom_report_by_company_operator(self, company, operator, start_date, end_date, format='excel'):
        """按公司和作業員生成自訂報表"""
        try:
            data = self.work_hour_service.get_custom_report_by_company_operator(company, operator, start_date, end_date)
            summary = self.work_hour_service.get_custom_summary_by_company_operator(company, operator, start_date, end_date)
            operator_stats = self.work_hour_service.get_operator_statistics(data)
            
            report_data = {
                'report_type': '自訂報表',
                'company': company,
                'operator': operator,
                'start_date': start_date,
                'end_date': end_date,
                'generated_at': timezone.now(),
                'summary': summary,
                'details': list(data.values()),
                'operator_stats': operator_stats,
            }
            
            # 生成檔案名稱
            company_name = company if company != 'all' else '全部公司'
            operator_name = operator if operator != 'all' else '全部作業員'
            filename = f'{company_name}_{operator_name}_{start_date}_{end_date}'
            
            if format == 'excel':
                return self._export_to_excel(report_data, filename)
            elif format == 'csv':
                return self._export_to_csv(report_data, filename)
            else:
                return report_data
                
        except Exception as e:
            logger.error(f"生成自訂報表失敗: {str(e)}")
            raise
    
    def _export_to_excel(self, report_data, filename):
        """匯出為Excel格式（包含三個活頁簿：統計摘要、詳細、統計）"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
            
            wb = openpyxl.Workbook()
            
            # 移除預設的工作表
            wb.remove(wb.active)
            
            # 1. 統計摘要活頁簿
            ws_summary = wb.create_sheet("統計摘要")
            self._create_summary_sheet(ws_summary, report_data)
            
            # 2. 詳細活頁簿
            ws_detail = wb.create_sheet("詳細")
            self._create_detail_sheet(ws_detail, report_data)
            
            # 3. 統計活頁簿
            ws_stats = wb.create_sheet("統計")
            self._create_stats_sheet(ws_stats, report_data)
            
            # 儲存檔案
            file_path = f"/tmp/{filename}.xlsx"
            wb.save(file_path)
            
            return {
                'file_path': file_path,
                'filename': f"{filename}.xlsx",
                'format': 'excel'
            }
            
        except ImportError:
            logger.error("openpyxl未安裝，無法生成Excel檔案")
            raise
        except Exception as e:
            logger.error(f"生成Excel檔案失敗: {str(e)}")
            raise
    
    def _create_summary_sheet(self, ws, report_data):
        """建立統計摘要活頁簿"""
        # 設定樣式
        from openpyxl.styles import Font, PatternFill
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        
        # 報表標題
        ws['A1'] = f"{report_data['report_type']} - 統計摘要"
        ws['A1'].font = title_font
        ws.merge_cells('A1:D1')
        
        # 基本資訊
        if report_data['report_type'] == '自訂報表':
            ws['A3'] = "公司"
            ws['B3'] = report_data.get('company', '全部')
            ws['C3'] = "作業員"
            ws['D3'] = report_data.get('operator', '全部')
            
            ws['A4'] = "起始日期"
            ws['B4'] = report_data.get('start_date', '').strftime('%Y-%m-%d') if hasattr(report_data.get('start_date', ''), 'strftime') else str(report_data.get('start_date', ''))
            ws['C4'] = "結束日期"
            ws['D4'] = report_data.get('end_date', '').strftime('%Y-%m-%d') if hasattr(report_data.get('end_date', ''), 'strftime') else str(report_data.get('end_date', ''))
            
            ws['A5'] = "生成時間"
            ws['B5'] = report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')
        else:
            ws['A3'] = "作業員"
            ws['B3'] = report_data.get('operator', '全部')
            ws['C3'] = "報表日期"
            ws['D3'] = report_data.get('date', '').strftime('%Y-%m-%d') if hasattr(report_data.get('date', ''), 'strftime') else str(report_data.get('date', ''))
            
            ws['A4'] = "生成時間"
            ws['B4'] = report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')
        
        # 摘要資料
        summary = report_data['summary']
        ws['A6'] = "總工作時數"
        ws['B6'] = float(summary['total_work_hours'])
        ws['C6'] = "總作業員數"
        ws['D6'] = summary['total_operators']
        
        ws['A7'] = "總設備使用時數"
        ws['B7'] = float(summary['total_equipment_hours'])
        ws['C7'] = "工單數量"
        ws['D7'] = summary['workorder_count']
        
        ws['A8'] = "平均日工作時數"
        ws['B8'] = float(summary['avg_daily_hours'])
        
        # 設定欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15
    
    def _create_detail_sheet(self, ws, report_data):
        """建立詳細活頁簿"""
        # 設定樣式
        from openpyxl.styles import Font, PatternFill, Border, Side
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 標題
        ws['A1'] = f"{report_data['report_type']} - 詳細資料"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:K1')
        
        # 表頭
        headers = ['作業員', '公司名稱', '工單編號', '產品編號', '工序', '工作日期', '開始時間', '結束時間', '工作時數', '加班時數', '合計時數']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # 資料行
        details = report_data['details']
        for row, detail in enumerate(details, 4):
            ws.cell(row=row, column=1, value=detail.get('operator_name', '-')).border = border
            ws.cell(row=row, column=2, value=detail.get('company', '-')).border = border
            ws.cell(row=row, column=3, value=detail.get('workorder_id', '-')).border = border
            ws.cell(row=row, column=4, value=detail.get('product_code', '-')).border = border
            ws.cell(row=row, column=5, value=detail.get('process_name', '-')).border = border
            ws.cell(row=row, column=6, value=detail.get('work_date', '-')).border = border
            ws.cell(row=row, column=7, value=detail.get('start_time', '-')).border = border
            ws.cell(row=row, column=8, value=detail.get('end_time', '-')).border = border
            ws.cell(row=row, column=9, value=float(detail.get('work_hours', 0))).border = border
            ws.cell(row=row, column=10, value=float(detail.get('overtime_hours', 0))).border = border
            ws.cell(row=row, column=11, value=float(detail.get('total_hours', 0))).border = border
        
        # 設定欄寬
        from openpyxl.utils import get_column_letter
        column_widths = [15, 15, 15, 20, 15, 12, 10, 10, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_stats_sheet(self, ws, report_data):
        """建立統計活頁簿"""
        # 設定樣式
        from openpyxl.styles import Font, PatternFill, Border, Side
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 標題
        ws['A1'] = f"{report_data['report_type']} - 作業員統計"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:E1')
        
        # 表頭
        headers = ['排名', '作業員', '工作時數', '加班時數', '合計時數']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # 計算作業員統計
        details = report_data['details']
        operator_stats = {}
        
        for detail in details:
            operator = detail.get('operator_name', '未指定')
            if operator not in operator_stats:
                operator_stats[operator] = {
                    'work_hours': 0,
                    'overtime_hours': 0,
                    'total_hours': 0
                }
            operator_stats[operator]['work_hours'] += float(detail.get('work_hours', 0))
            operator_stats[operator]['overtime_hours'] += float(detail.get('overtime_hours', 0))
            operator_stats[operator]['total_hours'] += float(detail.get('total_hours', 0))
        
        # 排序並寫入資料
        sorted_operators = sorted(operator_stats.items(), key=lambda x: x[1]['total_hours'], reverse=True)
        
        for row, (operator, stats) in enumerate(sorted_operators, 4):
            ws.cell(row=row, column=1, value=row-3).border = border  # 排名
            ws.cell(row=row, column=2, value=operator).border = border  # 作業員
            ws.cell(row=row, column=3, value=stats['work_hours']).border = border  # 工作時數
            ws.cell(row=row, column=4, value=stats['overtime_hours']).border = border  # 加班時數
            ws.cell(row=row, column=5, value=stats['total_hours']).border = border  # 合計時數
        
        # 設定欄寬
        from openpyxl.utils import get_column_letter
        column_widths = [8, 20, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _export_to_csv(self, report_data, filename):
        """匯出為CSV格式"""
        try:
            import csv
            import io
            
            output = io.StringIO()
            writer = csv.writer(output)
            
            # 寫入標題
            writer.writerow([f"{report_data['report_type']}"])
            writer.writerow([])
            
            # 寫入基本資訊
            writer.writerow(["公司代號", report_data['company_code'], "生成時間", report_data['generated_at'].strftime('%Y-%m-%d %H:%M:%S')])
            writer.writerow([])
            
            # 寫入摘要
            summary = report_data['summary']
            writer.writerow(["總工作時數", float(summary['total_work_hours'])])
            writer.writerow(["總作業員數", summary['total_operators']])
            writer.writerow(["總設備時數", float(summary['total_equipment_hours'])])
            writer.writerow(["工單數量", summary['workorder_count']])
            writer.writerow(["平均日工作時數", float(summary['avg_daily_hours'])])
            writer.writerow([])
            
            # 寫入詳細資料
            if report_data['details']:
                writer.writerow(["工單編號", "工作日期", "日工作時數", "作業員人數", "設備時數"])
                for detail in report_data['details']:
                    writer.writerow([
                        detail.get('workorder_id', ''),
                        detail.get('work_date', ''),
                        float(detail.get('daily_work_hours', 0)),
                        detail.get('operator_count', 0),
                        float(detail.get('equipment_hours', 0))
                    ])
            
            # 儲存檔案
            file_path = f"/tmp/{filename}.csv"
            with open(file_path, 'w', newline='', encoding='utf-8') as f:
                f.write(output.getvalue())
            
            return {
                'file_path': file_path,
                'filename': f"{filename}.csv",
                'format': 'csv'
            }
            
        except Exception as e:
            logger.error(f"生成CSV檔案失敗: {str(e)}")
            raise
