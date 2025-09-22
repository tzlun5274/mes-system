"""
報表模組視圖
提供各種報表的查詢和匯出功能
"""

import logging
from datetime import datetime, timedelta
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Avg, Count, Sum
from django.db.models.functions import ExtractYear, ExtractMonth
from django.utils import timezone
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.conf import settings
from django.views.generic import TemplateView, ListView, DetailView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt

from .models import (
    WorkOrderReportData, ReportSchedule, ReportExecutionLog,
    OperatorProcessCapacityScore, CompletedWorkOrderAnalysis
)
from .forms import (
    GenerateScoringReportForm, 
    SupervisorScoreForm, OperatorScoreFilterForm
)
from .scheduler import ReportScheduler
from .report_generator import ReportGenerator
from .report_schedule_sync_service import ReportScheduleSyncService
from django.utils import timezone
from datetime import timedelta

# 動態導入其他模組的模型
try:
    from workorder.fill_work.models import FillWork
except ImportError:
    FillWork = None

logger = logging.getLogger(__name__)


def is_report_user(user):
    """檢查使用者是否有報表權限"""
    return user.is_authenticated and (user.is_staff or user.is_superuser)


@login_required
def index(request):
    """報表首頁"""
    try:
        # 統計資料 - 只使用存在的模型
        total_workorders = WorkOrderReportData.objects.count()
        
        # 今日報表統計
        today = timezone.now().date()
        today_reports = WorkOrderReportData.objects.filter(
            work_date=today
        ).count()
        
        # 本週報表統計
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        weekly_reports = WorkOrderReportData.objects.filter(
            work_date__range=[start_of_week, end_of_week]
        ).count()
        
        # 本月報表統計
        start_of_month = today.replace(day=1)
        monthly_reports = WorkOrderReportData.objects.filter(
            work_date__gte=start_of_month
        ).count()
        
        # 計算總工作時數
        total_work_hours = WorkOrderReportData.objects.aggregate(
            total=Sum('daily_work_hours')
        )['total'] or 0
        
        # 計算今日工作時數
        today_work_hours = WorkOrderReportData.objects.filter(
            work_date=today
        ).aggregate(
            total=Sum('daily_work_hours')
        )['total'] or 0
        
        # 計算作業員數量（去重）
        operator_count = WorkOrderReportData.objects.exclude(
            operator_name__isnull=True
        ).values('operator_name').distinct().count()
        
        context = {
            'total_reports': total_workorders,
            'today_reports': today_reports,
            'pending_reports': weekly_reports,  # 使用週報表數量作為待處理
            'completed_reports': monthly_reports,  # 使用月報表數量作為已完成
            'total_work_hours': total_work_hours,
            'today_work_hours': today_work_hours,
            'operator_count': operator_count,
        }
    except Exception as e:
        logger.error(f"取得報表首頁統計資料失敗: {str(e)}")
        context = {
            'total_reports': 0,
            'today_reports': 0,
            'pending_reports': 0,
            'completed_reports': 0,
            'total_work_hours': 0,
            'today_work_hours': 0,
            'operator_count': 0,
        }
    
    return render(request, 'reporting/index.html', context)

def get_date_range(date_range):
    """
    根據日期範圍獲取開始和結束日期
    
    Args:
        date_range: 日期範圍字串
        
    Returns:
        tuple: (開始日期, 結束日期)
    """
    today = timezone.now().date()
    
    if date_range == 'today':
        return today, today
    elif date_range == 'yesterday':
        yesterday = today - timedelta(days=1)
        return yesterday, yesterday
    elif date_range == 'this_week':
        # 週一到週日
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        return start_of_week, end_of_week
    elif date_range == 'last_week':
        # 上週一到上週日
        start_of_week = today - timedelta(days=today.weekday())
        start_of_last_week = start_of_week - timedelta(days=7)
        end_of_last_week = start_of_last_week + timedelta(days=6)
        return start_of_last_week, end_of_last_week
    elif date_range == 'this_month':
        start_of_month = today.replace(day=1)
        if today.month == 12:
            end_of_month = today.replace(year=today.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end_of_month = today.replace(month=today.month + 1, day=1) - timedelta(days=1)
        return start_of_month, end_of_month
    elif date_range == 'last_month':
        if today.month == 1:
            start_of_last_month = today.replace(year=today.year - 1, month=12, day=1)
        else:
            start_of_last_month = today.replace(month=today.month - 1, day=1)
        end_of_last_month = today.replace(day=1) - timedelta(days=1)
        return start_of_last_month, end_of_last_month
    else:
        # 預設為今天
        return today, today

def get_report_types():
    """
    獲取報表類型選項
    """
    return [
        ('daily_work', '日工作報表'),
        ('weekly_work', '週工作報表'),
        ('monthly_work', '月工作報表'),
        ('daily_work_hour_operator', '日工時報表 (作業員)'),
        ('weekly_work_hour_operator', '週工時報表 (作業員)'),
        ('monthly_work_hour_operator', '月工時報表 (作業員)'),
        ('daily_work_hour_smt', '日工時報表 (SMT設備)'),
        ('weekly_work_hour_smt', '週工時報表 (SMT設備)'),
        ('monthly_work_hour_smt', '月工時報表 (SMT設備)'),
        ('operator_performance', '作業員績效報表'),
        ('smt_equipment', 'SMT設備效率報表'),
        ('abnormal_report', '異常報工報表'),
        ('efficiency_analysis', '效率分析報表'),
    ]

def get_date_ranges():
    """
    獲取日期範圍選項
    """
    return [
        ('today', '今天'),
        ('yesterday', '昨天'),
        ('this_week', '本週'),
        ('last_week', '上週'),
        ('this_month', '本月'),
        ('last_month', '上月'),
        ('custom', '自訂範圍'),
    ]

def get_daily_work_report_data(start_date, end_date):
    """
    獲取日工作報表資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 從填報記錄中獲取資料
    fill_work_reports = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    )
    
    for report in fill_work_reports:
        data.append({
            'operator': report.operator,
            'work_date': report.work_date.strftime('%Y-%m-%d'),
            'workorder': report.workorder,
            'product_id': report.product_id,
            'start_time': report.start_time,
            'end_time': report.end_time,
            'process': report.operation or report.process_name or '',
            'work_quantity': report.work_quantity,
            'defect_quantity': report.defect_quantity,
        })
    
    return data

def get_weekly_work_report_data(start_date, end_date):
    """
    獲取週工作報表資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 按作業員分組統計
    operator_stats = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    ).values('operator').annotate(
        total_work_quantity=Sum('work_quantity'),
        total_defect_quantity=Sum('defect_quantity')
    )
    
    for stat in operator_stats:
        data.append({
            'operator': stat['operator'],
            'total_work_quantity': stat['total_work_quantity'] or 0,
            'total_defect_quantity': stat['total_defect_quantity'] or 0,
        })
    
    return data

def get_monthly_work_report_data(start_date, end_date):
    """
    獲取月工作報表資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 按作業員分組統計
    operator_stats = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    ).values('operator').annotate(
        total_work_quantity=Sum('work_quantity'),
        total_defect_quantity=Sum('defect_quantity')
    )
    
    for stat in operator_stats:
        data.append({
            'operator': stat['operator'],
            'total_work_quantity': stat['total_work_quantity'] or 0,
            'total_defect_quantity': stat['total_defect_quantity'] or 0,
        })
    
    return data

def get_daily_work_hour_operator_report_data(start_date, end_date):
    """
    獲取日工時報表 (作業員) 資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 從填報記錄中獲取作業員工時資料
    operator_stats = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    ).exclude(
        Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
    ).values('operator', 'work_date').annotate(
        total_work_hours=Sum('work_hours_calculated'),
        total_good_quantity=Sum('work_quantity'),
        total_defect_quantity=Sum('defect_quantity')
    )
    
    for stat in operator_stats:
        data.append({
            'operator': stat['operator'],
            'work_date': stat['work_date'].strftime('%Y-%m-%d'),
            'work_hours': round(stat['total_work_hours'] or 0, 2),
            'good_quantity': stat['total_good_quantity'] or 0,
            'defect_quantity': stat['total_defect_quantity'] or 0,
        })
    
    return data

def get_weekly_work_hour_operator_report_data(start_date, end_date):
    """
    獲取週工時報表 (作業員) 資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 按作業員分組統計週工時
    operator_stats = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    ).exclude(
        Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
    ).values('operator').annotate(
        total_work_hours=Sum('work_hours_calculated'),
        total_good_quantity=Sum('work_quantity'),
        total_defect_quantity=Sum('defect_quantity')
    )
    
    for stat in operator_stats:
        data.append({
            'operator': stat['operator'],
            'work_hours': round(stat['total_work_hours'] or 0, 2),
            'good_quantity': stat['total_good_quantity'] or 0,
            'defect_quantity': stat['total_defect_quantity'] or 0,
        })
    
    return data

def get_monthly_work_hour_operator_report_data(start_date, end_date):
    """
    獲取月工時報表 (作業員) 資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 按作業員分組統計月工時
    operator_stats = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    ).exclude(
        Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
    ).values('operator').annotate(
        total_work_hours=Sum('work_hours_calculated'),
        total_good_quantity=Sum('work_quantity'),
        total_defect_quantity=Sum('defect_quantity')
    )
    
    for stat in operator_stats:
        data.append({
            'operator': stat['operator'],
            'work_hours': round(stat['total_work_hours'] or 0, 2),
            'good_quantity': stat['total_good_quantity'] or 0,
            'defect_quantity': stat['total_defect_quantity'] or 0,
        })
    
    return data

def get_daily_work_hour_smt_report_data(start_date, end_date):
    """
    獲取日工時報表 (SMT設備) 資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 從填報記錄中獲取SMT設備工時資料
    smt_stats = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    ).filter(
        Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
    ).values('equipment__name', 'work_date').annotate(
        total_work_hours=Sum('work_hours_calculated'),
        total_good_quantity=Sum('work_quantity'),
        total_defect_quantity=Sum('defect_quantity')
    )
    
    for stat in smt_stats:
        data.append({
            'equipment': stat['equipment__name'] or 'SMT設備',
            'work_date': stat['work_date'].strftime('%Y-%m-%d'),
            'work_hours': round(stat['total_work_hours'] or 0, 2),
            'good_quantity': stat['total_good_quantity'] or 0,
            'defect_quantity': stat['total_defect_quantity'] or 0,
        })
    
    return data

def get_weekly_work_hour_smt_report_data(start_date, end_date):
    """
    獲取週工時報表 (SMT設備) 資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 按SMT設備分組統計週工時
    smt_stats = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    ).filter(
        Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
    ).values('equipment__name').annotate(
        total_work_hours=Sum('work_hours_calculated'),
        total_good_quantity=Sum('work_quantity'),
        total_defect_quantity=Sum('defect_quantity')
    )
    
    for stat in smt_stats:
        data.append({
            'equipment': stat['equipment__name'] or 'SMT設備',
            'work_hours': round(stat['total_work_hours'] or 0, 2),
            'good_quantity': stat['total_good_quantity'] or 0,
            'defect_quantity': stat['total_defect_quantity'] or 0,
        })
    
    return data

def get_monthly_work_hour_smt_report_data(start_date, end_date):
    """
    獲取月工時報表 (SMT設備) 資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 按SMT設備分組統計月工時
    smt_stats = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    ).filter(
        Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
    ).values('equipment__name').annotate(
        total_work_hours=Sum('work_hours_calculated'),
        total_good_quantity=Sum('work_quantity'),
        total_defect_quantity=Sum('defect_quantity')
    )
    
    for stat in smt_stats:
        data.append({
            'equipment': stat['equipment__name'] or 'SMT設備',
            'work_hours': round(stat['total_work_hours'] or 0, 2),
            'good_quantity': stat['total_good_quantity'] or 0,
            'defect_quantity': stat['total_defect_quantity'] or 0,
        })
    
    return data

def get_operator_performance_report_data(start_date, end_date):
    """
    獲取作業員績效報表資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 從填報記錄中獲取作業員績效資料
    operator_stats = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    ).exclude(
        Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
    ).values('operator').annotate(
        total_good_quantity=Sum('work_quantity'),
        total_defect_quantity=Sum('defect_quantity'),
        total_work_hours=Sum('work_hours_calculated')
    )
    
    for stat in operator_stats:
        total_quantity = (stat['total_good_quantity'] or 0) + (stat['total_defect_quantity'] or 0)
        defect_rate = 0
        if total_quantity > 0:
            defect_rate = round((stat['total_defect_quantity'] or 0) / total_quantity * 100, 2)
        
        efficiency = 0
        if stat['total_work_hours'] and stat['total_work_hours'] > 0:
            efficiency = round((stat['total_good_quantity'] or 0) / stat['total_work_hours'], 2)
        
        data.append({
            'operator': stat['operator'],
            'good_quantity': stat['total_good_quantity'] or 0,
            'defect_quantity': stat['total_defect_quantity'] or 0,
            'defect_rate': defect_rate,
            'efficiency': efficiency,
        })
    
    return data

def get_smt_equipment_report_data(start_date, end_date):
    """
    獲取SMT設備效率報表資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 從填報記錄中獲取SMT設備效率資料
    smt_stats = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    ).filter(
        Q(operator__icontains='SMT') | Q(process__name__icontains='SMT')
    ).values('equipment__name').annotate(
        total_good_quantity=Sum('work_quantity'),
        total_defect_quantity=Sum('defect_quantity'),
        total_work_hours=Sum('work_hours_calculated')
    )
    
    for stat in smt_stats:
        total_quantity = (stat['total_good_quantity'] or 0) + (stat['total_defect_quantity'] or 0)
        defect_rate = 0
        if total_quantity > 0:
            defect_rate = round((stat['total_defect_quantity'] or 0) / total_quantity * 100, 2)
        
        efficiency = 0
        if stat['total_work_hours'] and stat['total_work_hours'] > 0:
            efficiency = round((stat['total_good_quantity'] or 0) / stat['total_work_hours'], 2)
        
        data.append({
            'equipment': stat['equipment__name'] or 'SMT設備',
            'good_quantity': stat['total_good_quantity'] or 0,
            'defect_quantity': stat['total_defect_quantity'] or 0,
            'defect_rate': defect_rate,
            'efficiency': efficiency,
        })
    
    return data

def get_abnormal_report_data(start_date, end_date):
    """
    獲取異常報工報表資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 從填報記錄中獲取異常報工資料
    abnormal_reports = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    ).exclude(
        abnormal_notes__isnull=True
    ).exclude(
        abnormal_notes__exact=''
    )
    
    for report in abnormal_reports:
        data.append({
            'operator': report.operator,
            'work_date': report.work_date.strftime('%Y-%m-%d'),
            'workorder': report.workorder,
            'process': report.operation or report.process_name or '',
            'abnormal_notes': report.abnormal_notes,
            'good_quantity': report.work_quantity,
            'defect_quantity': report.defect_quantity,
        })
    
    return data

def get_efficiency_analysis_report_data(start_date, end_date):
    """
    獲取效率分析報表資料
    
    Args:
        start_date: 開始日期
        end_date: 結束日期
        
    Returns:
        list: 報表資料
    """
    data = []
    
    # 從填報記錄中獲取效率分析資料
    efficiency_stats = FillWork.objects.filter(
        work_date__range=[start_date, end_date],
        approval_status='approved'
    ).values('operator', 'process__name').annotate(
        total_good_quantity=Sum('work_quantity'),
        total_defect_quantity=Sum('defect_quantity'),
        total_work_hours=Sum('work_hours_calculated')
    )
    
    for stat in efficiency_stats:
        total_quantity = (stat['total_good_quantity'] or 0) + (stat['total_defect_quantity'] or 0)
        defect_rate = 0
        if total_quantity > 0:
            defect_rate = round((stat['total_defect_quantity'] or 0) / total_quantity * 100, 2)
        
        efficiency = 0
        if stat['total_work_hours'] and stat['total_work_hours'] > 0:
            efficiency = round((stat['total_good_quantity'] or 0) / stat['total_work_hours'], 2)
        
        data.append({
            'operator': stat['operator'],
            'process': stat['process__name'] or '',
            'good_quantity': stat['total_good_quantity'] or 0,
            'defect_quantity': stat['total_defect_quantity'] or 0,
            'defect_rate': defect_rate,
            'efficiency': efficiency,
        })
    
    return data

def get_report_title(report_type):
    """
    獲取報表標題
    """
    titles = {
        'daily_work': '日工作報表',
        'weekly_work': '週工作報表',
        'monthly_work': '月工作報表',
        'daily_work_hour_operator': '日工時報表 (作業員)',
        'weekly_work_hour_operator': '週工時報表 (作業員)',
        'monthly_work_hour_operator': '月工時報表 (作業員)',
        'daily_work_hour_smt': '日工時報表 (SMT設備)',
        'weekly_work_hour_smt': '週工時報表 (SMT設備)',
        'monthly_work_hour_smt': '月工時報表 (SMT設備)',
        'operator_performance': '作業員績效報表',
        'smt_equipment': 'SMT設備效率報表',
        'abnormal_report': '異常報工報表',
        'efficiency_analysis': '效率分析報表',
    }
    return titles.get(report_type, '未知報表')

def get_report_headers(report_type):
    """
    獲取報表表頭
    """
    headers = {
        'daily_work': ['作業員', '報工日期', '工單號', '產品編號', '開始時間', '結束時間', '工序', '工作數量', '不良品數量'],
        'weekly_work': ['作業員', '週工作數量', '週不良品數量'],
        'monthly_work': ['作業員', '月工作數量', '月不良品數量'],
        'daily_work_hour_operator': ['作業員', '報工日期', '工時(小時)', '良品數量', '不良品數量'],
        'weekly_work_hour_operator': ['作業員', '週工時(小時)', '週良品數量', '週不良品數量'],
        'monthly_work_hour_operator': ['作業員', '月工時(小時)', '月良品數量', '月不良品數量'],
        'daily_work_hour_smt': ['設備', '報工日期', '運行時間(小時)', '良品數量', '不良品數量'],
        'weekly_work_hour_smt': ['設備', '週運行時間(小時)', '週良品數量', '週不良品數量'],
        'monthly_work_hour_smt': ['設備', '月運行時間(小時)', '月良品數量', '月不良品數量'],
        'operator_performance': ['作業員', '良品數量', '不良品數量', '不良率(%)', '平均效率(%)'],
        'smt_equipment': ['設備', '良品數量', '不良品數量', '不良率(%)', '平均效率(%)'],
        'abnormal_report': ['作業員', '報工日期', '工單號', '工序', '異常紀錄', '良品數量', '不良品數量'],
        'efficiency_analysis': ['作業員', '工序', '良品數量', '不良品數量', '不良率(%)', '平均效率(%)'],
    }
    return headers.get(report_type, [])

def get_report_filename(report_type, start_date, end_date):
    """
    獲取報表檔案名稱
    """
    title = get_report_title(report_type)
    start_str = start_date.strftime('%Y%m%d')
    end_str = end_date.strftime('%Y%m%d')
    
    if start_date == end_date:
        return f"{title}_{start_str}.xlsx"
    else:
        return f"{title}_{start_str}_{end_str}.xlsx"

def export_to_excel(data, headers, filename):
    """
    匯出Excel檔案
    
    Args:
        data: 報表資料
        headers: 表頭
        filename: 檔案名稱
        
    Returns:
        HttpResponse: Excel檔案回應
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # 建立工作簿
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "報表"
        
        # 設定表頭樣式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 寫入表頭
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 寫入資料
        for row, record in enumerate(data, 2):
            for col, header in enumerate(headers, 1):
                value = record.get(header, '')
                ws.cell(row=row, column=col, value=value)
        
        # 調整欄寬
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # 建立回應
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # 儲存檔案
        wb.save(response)
        return response
        
    except ImportError:
        # 如果沒有 openpyxl，回退到 CSV
        return export_to_csv(data, headers, filename.replace('.xlsx', '.csv'))

def export_to_csv(data, headers, filename):
    """
    匯出CSV檔案
    
    Args:
        data: 報表資料
        headers: 表頭
        filename: 檔案名稱
        
    Returns:
        HttpResponse: CSV檔案回應
    """
    import csv
    from io import StringIO
    
    # 建立CSV檔案
    output = StringIO()
    writer = csv.writer(output)
    
    # 寫入表頭
    writer.writerow(headers)
    
    # 寫入資料
    for record in data:
        row = [record.get(header, '') for header in headers]
        writer.writerow(row)
    
    # 建立回應
    response = HttpResponse(output.getvalue(), content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    return response

@login_required
def report_export(request):
    """
    報表匯出頁面
    """
    context = {
        'report_types': get_report_types(),
        'date_ranges': get_date_ranges(),
    }
    return render(request, 'reporting/report_export.html', context)

@login_required
def execute_report_export(request):
    """
    執行報表匯出
    """
    if request.method != 'POST':
        return JsonResponse({'error': '只支援 POST 請求'}, status=405)
    
    try:
        # 取得參數
        report_type = request.POST.get('report_type')
        date_range = request.POST.get('date_range')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        export_format = request.POST.get('export_format', 'excel')
        
        # 驗證參數
        if not report_type:
            return JsonResponse({'error': '請選擇報表類型'}, status=400)
        
        # 取得日期範圍
        if date_range == 'custom':
            if not start_date_str or not end_date_str:
                return JsonResponse({'error': '請選擇自訂日期範圍'}, status=400)
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        else:
            start_date, end_date = get_date_range(date_range)
        
        # 驗證日期範圍（最多4個月）
        date_diff = (end_date - start_date).days
        if date_diff > 120:  # 4個月 = 120天
            return JsonResponse({'error': '查詢期間不能超過4個月'}, status=400)
        
        # 根據報表類型獲取資料
        data = []
        if report_type == 'daily_work':
            data = get_daily_work_report_data(start_date, end_date)
        elif report_type == 'weekly_work':
            data = get_weekly_work_report_data(start_date, end_date)
        elif report_type == 'monthly_work':
            data = get_monthly_work_report_data(start_date, end_date)
        elif report_type == 'daily_work_hour_operator':
            data = get_daily_work_hour_operator_report_data(start_date, end_date)
        elif report_type == 'weekly_work_hour_operator':
            data = get_weekly_work_hour_operator_report_data(start_date, end_date)
        elif report_type == 'monthly_work_hour_operator':
            data = get_monthly_work_hour_operator_report_data(start_date, end_date)
        elif report_type == 'daily_work_hour_smt':
            data = get_daily_work_hour_smt_report_data(start_date, end_date)
        elif report_type == 'weekly_work_hour_smt':
            data = get_weekly_work_hour_smt_report_data(start_date, end_date)
        elif report_type == 'monthly_work_hour_smt':
            data = get_monthly_work_hour_smt_report_data(start_date, end_date)
        elif report_type == 'operator_performance':
            data = get_operator_performance_report_data(start_date, end_date)
        elif report_type == 'smt_equipment':
            data = get_smt_equipment_report_data(start_date, end_date)
        elif report_type == 'abnormal_report':
            data = get_abnormal_report_data(start_date, end_date)
        elif report_type == 'efficiency_analysis':
            data = get_efficiency_analysis_report_data(start_date, end_date)
        else:
            return JsonResponse({'error': '不支援的報表類型'}, status=400)
        
        # 獲取表頭和檔案名稱
        headers = get_report_headers(report_type)
        filename = get_report_filename(report_type, start_date, end_date)
        
        # 根據格式匯出
        if export_format == 'csv':
            filename = filename.replace('.xlsx', '.csv')
            return export_to_csv(data, headers, filename)
        else:
            return export_to_excel(data, headers, filename)
        
    except Exception as e:
        logger.error(f"匯出報表時發生錯誤: {str(e)}")
        return JsonResponse({'error': f'匯出失敗: {str(e)}'}, status=500) 

def placeholder_view(request, function_name):
    """
    通用佔位符視圖 - 用於所有未實現的功能
    """
    context = {
        'function_name': function_name,
        'message': f'功能 "{function_name}" 正在開發中，敬請期待！'
    }
    return render(request, 'reporting/placeholder.html', context) 


@login_required
def work_hour_report_index(request):
    """工作時數報表首頁"""
    return render(request, 'reporting/reporting/work_hour_report.html')


@login_required
def daily_report(request):
    """日報表預覽"""
    if request.method == 'POST':
        company = request.POST.get('company', 'all')
        operator = request.POST.get('operator', 'all')
        date_str = request.POST.get('date')
        
        try:
            # 解析日期
            date = datetime.strptime(date_str, '%Y-%m-%d').date()
            
            service = WorkHourReportService()
            
            # 取得報表資料
            data = service.get_daily_report_by_company_operator(company, operator, date)
            summary = service.get_daily_summary_by_company_operator(company, operator, date)
            
            # 取得公司選項
            companies = WorkOrderReportData.objects.exclude(
                company__isnull=True
            ).values_list('company', flat=True).distinct().order_by('company')
            
            # 取得作業員選項
            operators = WorkOrderReportData.objects.exclude(
                operator_name__isnull=True
            ).values_list('operator_name', flat=True).distinct().order_by('operator_name')
            
            # 顯示預覽報表
            context = {
                'report_type': '日報表',
                'company': company,
                'operator': operator,
                'date': date,
                'summary': summary,
                'data': data,
                'generated_at': timezone.now(),
                'companies': companies,
                'operators': operators,
            }
            return render(request, 'reporting/reporting/daily_report.html', context)
                
        except Exception as e:
            messages.error(request, f'生成日報表失敗: {str(e)}')
            return redirect('reporting:work_hour_report_index')
    
    # GET 請求顯示預設頁面
    # 取得最近有資料的日期，如果沒有則使用今天
    latest_date = WorkOrderReportData.objects.order_by('-work_date').first()
    if latest_date:
        current_date = latest_date.work_date
    else:
        current_date = timezone.now().date()
    
    # 取得預設資料（最近有資料的日期全部公司和作業員）
    service = WorkHourReportService()
    data = service.get_daily_report_by_company_operator('all', 'all', current_date)
    summary = service.get_daily_summary_by_company_operator('all', 'all', current_date)
    
    # 取得公司選項
    companies = WorkOrderReportData.objects.exclude(
        company__isnull=True
    ).values_list('company', flat=True).distinct().order_by('company')
    
    # 取得作業員選項
    operators = WorkOrderReportData.objects.exclude(
        operator_name__isnull=True
    ).values_list('operator_name', flat=True).distinct().order_by('operator_name')
    
    context = {
        'report_type': '日報表',
        'company': 'all',
        'operator': 'all',
        'date': current_date,
        'summary': summary,
        'data': data,
        'generated_at': timezone.now(),
        'companies': companies,
        'operators': operators,
    }
    return render(request, 'reporting/reporting/daily_report.html', context)


@login_required
def custom_report(request):
    """自訂報表預覽"""
    if request.method == 'POST':
        company = request.POST.get('company', 'all')
        operator = request.POST.get('operator', 'all')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        
        try:
            # 解析日期
            start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
            end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
            
            service = WorkHourReportService()
            
            # 取得報表資料
            data = service.get_custom_report_by_company_operator(company, operator, start_date, end_date)
            summary = service.get_custom_summary_by_company_operator(company, operator, start_date, end_date)
            
            # 取得公司選項
            companies = WorkOrderReportData.objects.exclude(
                company__isnull=True
            ).values_list('company', flat=True).distinct().order_by('company')
            
            # 取得作業員選項
            operators = WorkOrderReportData.objects.exclude(
                operator_name__isnull=True
            ).values_list('operator_name', flat=True).distinct().order_by('operator_name')
            
            # 顯示預覽報表
            context = {
                'report_type': '自訂報表',
                'company': company,
                'operator': operator,
                'start_date': start_date,
                'end_date': end_date,
                'summary': summary,
                'data': data,
                'generated_at': timezone.now(),
                'companies': companies,
                'operators': operators,
            }
            return render(request, 'reporting/reporting/custom_report.html', context)
                
        except Exception as e:
            messages.error(request, f'生成自訂報表失敗: {str(e)}')
            return redirect('reporting:work_hour_report_index')
    
    # GET 請求顯示預設頁面
    # 取得最近30天的資料
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    # 取得預設資料
    service = WorkHourReportService()
    data = service.get_custom_report_by_company_operator('all', 'all', start_date, end_date)
    summary = service.get_custom_summary_by_company_operator('all', 'all', start_date, end_date)
    
    # 取得公司選項
    companies = WorkOrderReportData.objects.exclude(
        company__isnull=True
    ).values_list('company', flat=True).distinct().order_by('company')
    
    # 取得作業員選項
    operators = WorkOrderReportData.objects.exclude(
        operator_name__isnull=True
    ).values_list('operator_name', flat=True).distinct().order_by('operator_name')
    
    context = {
        'report_type': '自訂報表',
        'company': 'all',
        'operator': 'all',
        'start_date': start_date,
        'end_date': end_date,
        'summary': summary,
        'data': data,
        'generated_at': timezone.now(),
        'companies': companies,
        'operators': operators,
    }
    return render(request, 'reporting/reporting/custom_report.html', context)


@login_required
def unified_report_form(request):
    """統一報表生成器 - 簡化為自訂日期範圍報表"""
    from django.core.paginator import Paginator
    
    # 準備通用上下文資料
    current_year = timezone.now().year
    current_month = timezone.now().month
    current_week = timezone.now().isocalendar()[1]
    
    years = range(current_year - 5, current_year + 2)
    months = range(1, 13)
    weeks = range(1, 54)  # 一年最多53週
    
    # 準備自訂報表的預設日期範圍（最近30天）
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    # 獲取真實的公司資料
    try:
        from erp_integration.models import CompanyConfig
        companies = CompanyConfig.objects.all().order_by('company_name')
    except ImportError:
        companies = []
    
    # 獲取真實的作業員資料
    try:
        from process.models import Operator
        operators = Operator.objects.all().order_by('name')
    except ImportError:
        operators = []
    
    # 如果沒有作業員資料，嘗試從填報記錄中獲取
    if not operators:
        try:
            from workorder.fill_work.models import FillWork
            from django.db.models import Q
            operators_data = FillWork.objects.values_list('operator', flat=True).distinct().exclude(
                Q(operator__isnull=True) | Q(operator='')
            ).order_by('operator')
            operators = [{'name': op} for op in operators_data]
        except ImportError:
            operators = []
    
    # 如果還是沒有作業員資料，嘗試從現場報工記錄中獲取
    if not operators:
        try:
            from workorder.onsite_reporting.models import OnsiteReport
            from django.db.models import Q
            operators_data = OnsiteReport.objects.values_list('operator', flat=True).distinct().exclude(
                Q(operator__isnull=True) | Q(operator='')
            ).order_by('operator')
            operators = [{'name': op} for op in operators_data]
        except ImportError:
            operators = []
    
    # 處理查詢請求
    if request.method == 'POST':
        company = request.POST.get('company', 'all')
        operator = request.POST.get('operator', 'all')
        start_date_str = request.POST.get('start_date')
        end_date_str = request.POST.get('end_date')
        
        try:
            if start_date_str and end_date_str:
                start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
                
                service = WorkHourReportService()
                data = service.get_custom_report_by_company_operator(company, operator, start_date, end_date)
                summary = service.get_custom_summary_by_company_operator(company, operator, start_date, end_date)
                operator_stats = service.get_operator_statistics(data)
                
                # 分頁處理 - 每頁20行
                paginator = Paginator(data, 20)
                page_number = request.GET.get('page', 1)
                report_data = paginator.get_page(page_number)
                
                # 統計資料分頁
                stats_paginator = Paginator(operator_stats, 20)
                stats_page_number = request.GET.get('page', 1)
                operator_stats_paginated = stats_paginator.get_page(stats_page_number)
                
                context = {
                    'current_year': current_year,
                    'current_month': current_month,
                    'current_week': current_week,
                    'years': years,
                    'months': months,
                    'weeks': weeks,
                    'current_date': timezone.now().date(),
                    'start_date': start_date,
                    'end_date': end_date,
                    'report_data': report_data,
                    'summary': summary,
                    'operator_stats': operator_stats_paginated,
                    'company': company,
                    'operator': operator,
                    'companies': companies,
                    'operators': operators,
                    'timezone': timezone,
                }
                return render(request, 'reporting/reporting/unified_report_form.html', context)
            else:
                messages.error(request, '請選擇起始日期和結束日期')
                
        except Exception as e:
            messages.error(request, f'生成報表失敗: {str(e)}')
    
    # GET 請求或查詢失敗時的預設上下文
    context = {
        'current_year': current_year,
        'current_month': current_month,
        'current_week': current_week,
        'years': years,
        'months': months,
        'weeks': weeks,
        'current_date': timezone.now().date(),
        'start_date': start_date,
        'end_date': end_date,
        # 初始化報表相關變數，確保模板能正常顯示
        'report_data': None,
        'summary': {
            'total_work_hours': 0,
            'total_operators': 0,
            'total_equipment_hours': 0,
            'workorder_count': 0,
            'avg_daily_hours': 0,
        },
        'operator_stats': None,
        'report_type': None,
        # 添加真實資料到上下文
        'companies': companies,
        'operators': operators,
        # 添加 timezone 到上下文
        'timezone': timezone,
    }
    
    return render(request, 'reporting/reporting/unified_report_form.html', context)


@login_required
def custom_report_export(request):
    """自訂報表匯出"""
    company = request.GET.get('company', 'all')
    operator = request.GET.get('operator', 'all')
    start_date_str = request.GET.get('start_date')
    end_date_str = request.GET.get('end_date')
    format = request.GET.get('format', 'excel')
    
    try:
        # 驗證必要條件
        if not start_date_str or not end_date_str:
            messages.error(request, '請選擇起始日期和結束日期')
            return redirect('reporting:unified_report')
        
        # 解析日期
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        
        # 驗證日期範圍
        if start_date > end_date:
            messages.error(request, '起始日期不能大於結束日期')
            return redirect('reporting:unified_report')
        
        service = WorkHourReportService()
        generator = ReportGeneratorService()
        
        # 使用與查詢相同的條件生成檔案
        result = generator.generate_custom_report_by_company_operator(company, operator, start_date, end_date, format)
        
        # 下載檔案
        with open(result['file_path'], 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/octet-stream')
            # 設定檔案名稱，支援中文
            filename = result["filename"]
            import urllib.parse
            # 使用 URL 編碼處理中文檔案名稱
            encoded_filename = urllib.parse.quote(filename)
            response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"; filename*=UTF-8\'\'{encoded_filename}'
            # 添加安全標頭
            response['X-Content-Type-Options'] = 'nosniff'
            response['X-Frame-Options'] = 'DENY'
            return response
            
    except ValueError as e:
        messages.error(request, f'日期格式錯誤: {str(e)}')
        return redirect('reporting:unified_report')
    except Exception as e:
        messages.error(request, f'匯出自訂報表失敗: {str(e)}')
        return redirect('reporting:unified_report')


@login_required
def daily_report_export(request):
    """日報表匯出"""
    operator = request.GET.get('operator')
    date_str = request.GET.get('date')
    format = request.GET.get('format', 'excel')
    
    try:
        # 解析日期
        date = datetime.strptime(date_str, '%Y-%m-%d').date()
        
        service = WorkHourReportService()
        generator = ReportGeneratorService()
        
        # 生成檔案
        result = generator.generate_daily_report_by_operator(operator, date, format)
        
        # 下載檔案
        with open(result['file_path'], 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/octet-stream')
            # 設定檔案名稱，支援中文
            filename = result["filename"]
            import urllib.parse
            # 使用 URL 編碼處理中文檔案名稱
            encoded_filename = urllib.parse.quote(filename)
            response['Content-Disposition'] = f'attachment; filename="{encoded_filename}"; filename*=UTF-8\'\'{encoded_filename}'
            # 添加安全標頭
            response['X-Content-Type-Options'] = 'nosniff'
            response['X-Frame-Options'] = 'DENY'
            return response
            
    except Exception as e:
        messages.error(request, f'匯出日報表失敗: {str(e)}')
        return redirect('reporting:daily_report')


@login_required
def weekly_report(request):
    """週報表"""
    if request.method == 'POST':
        company_code = request.POST.get('company_code')
        year = int(request.POST.get('year'))
        week = int(request.POST.get('week'))
        format = request.POST.get('format', 'excel')
        
        try:
            service = WorkHourReportService()
            generator = ReportGeneratorService()
            
            # 取得報表資料
            data = service.get_weekly_report(company_code, year, week)
            summary = service.get_weekly_summary(company_code, year, week)
            
            if format in ['excel', 'csv']:
                # 生成檔案
                result = generator.generate_weekly_report(company_code, year, week, format)
                
                # 下載檔案
                with open(result['file_path'], 'rb') as f:
                    response = HttpResponse(f.read(), content_type='application/octet-stream')
                    response['Content-Disposition'] = f'attachment; filename="{result["filename"]}"'
                    return response
            else:
                # 顯示網頁報表
                context = {
                    'report_type': '週報表',
                    'company_code': company_code,
                    'year': year,
                    'week': week,
                    'summary': summary,
                    'data': data,
                    'generated_at': timezone.now(),
                }
                return render(request, 'reporting/reporting/weekly_report.html', context)
                
        except Exception as e:
            messages.error(request, f'生成週報表失敗: {str(e)}')
            return redirect('reporting:work_hour_report_index')
    
    # GET 請求顯示表單
    current_year = timezone.now().year
    current_week = timezone.now().isocalendar()[1]
    
    context = {
        'current_year': current_year,
        'current_week': current_week,
        'years': range(current_year - 5, current_year + 1),
        'weeks': range(1, 53),
    }
    return render(request, 'reporting/reporting/weekly_report_form.html', context)


@login_required
def monthly_report(request):
    """月報表"""
    if request.method == 'POST':
        company_code = request.POST.get('company_code')
        year = int(request.POST.get('year'))
        month = int(request.POST.get('month'))
        format = request.POST.get('format', 'excel')
        
        try:
            service = WorkHourReportService()
            generator = ReportGeneratorService()
            
            # 取得報表資料
            data = service.get_monthly_report(company_code, year, month)
            summary = service.get_monthly_summary(company_code, year, month)
            
            if format in ['excel', 'csv']:
                # 生成檔案
                result = generator.generate_monthly_report(company_code, year, month, format)
                
                # 下載檔案
                with open(result['file_path'], 'rb') as f:
                    response = HttpResponse(f.read(), content_type='application/octet-stream')
                    response['Content-Disposition'] = f'attachment; filename="{result["filename"]}"'
                    return response
            else:
                # 顯示網頁報表
                context = {
                    'report_type': '月報表',
                    'company_code': company_code,
                    'year': year,
                    'month': month,
                    'summary': summary,
                    'data': data,
                    'generated_at': timezone.now(),
                }
                return render(request, 'reporting/reporting/monthly_report.html', context)
                
        except Exception as e:
            messages.error(request, f'生成月報表失敗: {str(e)}')
            return redirect('reporting:work_hour_report_index')
    
    # GET 請求顯示表單
    current_year = timezone.now().year
    current_month = timezone.now().month
    
    context = {
        'current_year': current_year,
        'current_month': current_month,
        'years': range(current_year - 5, current_year + 1),
        'months': range(1, 13),
    }
    return render(request, 'reporting/reporting/monthly_report_form.html', context)


@login_required
def quarterly_report(request):
    """季報表"""
    if request.method == 'POST':
        company_code = request.POST.get('company_code')
        year = int(request.POST.get('year'))
        quarter = int(request.POST.get('quarter'))
        format = request.POST.get('format', 'excel')
        
        try:
            service = WorkHourReportService()
            generator = ReportGeneratorService()
            
            # 取得報表資料
            data = service.get_quarterly_report(company_code, year, quarter)
            summary = service.get_quarterly_summary(company_code, year, quarter)
            
            if format in ['excel', 'csv']:
                # 生成檔案
                result = generator.generate_quarterly_report(company_code, year, quarter, format)
                
                # 下載檔案
                with open(result['file_path'], 'rb') as f:
                    response = HttpResponse(f.read(), content_type='application/octet-stream')
                    response['Content-Disposition'] = f'attachment; filename="{result["filename"]}"'
                    return response
            else:
                # 顯示網頁報表
                context = {
                    'report_type': '季報表',
                    'company_code': company_code,
                    'year': year,
                    'quarter': quarter,
                    'summary': summary,
                    'data': data,
                    'generated_at': timezone.now(),
                }
                return render(request, 'reporting/reporting/quarterly_report.html', context)
                
        except Exception as e:
            messages.error(request, f'生成季報表失敗: {str(e)}')
            return redirect('reporting:work_hour_report_index')
    
    # GET 請求顯示表單
    current_year = timezone.now().year
    current_quarter = (timezone.now().month - 1) // 3 + 1
    
    context = {
        'current_year': current_year,
        'current_quarter': current_quarter,
        'years': range(current_year - 5, current_year + 1),
        'quarters': range(1, 5),
    }
    return render(request, 'reporting/reporting/quarterly_report_form.html', context)


@login_required
def yearly_report(request):
    """年報表"""
    if request.method == 'POST':
        company_code = request.POST.get('company_code')
        year = int(request.POST.get('year'))
        format = request.POST.get('format', 'excel')
        
        try:
            service = WorkHourReportService()
            generator = ReportGeneratorService()
            
            # 取得報表資料
            data = service.get_yearly_report(company_code, year)
            summary = service.get_yearly_summary(company_code, year)
            
            if format in ['excel', 'csv']:
                # 生成檔案
                result = generator.generate_yearly_report(company_code, year, format)
                
                # 下載檔案
                with open(result['file_path'], 'rb') as f:
                    response = HttpResponse(f.read(), content_type='application/octet-stream')
                    response['Content-Disposition'] = f'attachment; filename="{result["filename"]}"'
                    return response
            else:
                # 顯示網頁報表
                context = {
                    'report_type': '年報表',
                    'company_code': company_code,
                    'year': year,
                    'summary': summary,
                    'data': data,
                    'generated_at': timezone.now(),
                }
                return render(request, 'reporting/reporting/yearly_report.html', context)
                
        except Exception as e:
            messages.error(request, f'生成年報表失敗: {str(e)}')
            return redirect('reporting:work_hour_report_index')
    
    # GET 請求顯示表單
    current_year = timezone.now().year
    
    context = {
        'current_year': current_year,
        'years': range(current_year - 10, current_year + 1),
    }
    return render(request, 'reporting/reporting/yearly_report_form.html', context)


@login_required
def report_schedule_list(request):
    """報表排程列表"""
    schedules = ReportSchedule.objects.all().order_by('-created_at')
    
    # 分頁
    paginator = Paginator(schedules, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 檢查每個排程的同步狀態
    from django_celery_beat.models import PeriodicTask
    schedule_sync_status = {}
    for schedule in schedules:
        task = PeriodicTask.objects.filter(name=f'report_schedule_{schedule.id}').first()
        schedule_sync_status[schedule.id] = {
            'synced': task is not None,
            'enabled': task.enabled if task else False,
            'last_run': task.last_run_at if task else None,
        }
    
    context = {
        'page_obj': page_obj,
        'schedules': page_obj,
        'schedule_sync_status': schedule_sync_status,
    }
    return render(request, 'reporting/reporting/report_schedule_list.html', context)


@login_required
def report_schedule_form(request, schedule_id=None):
    """報表排程表單"""
    if schedule_id:
        schedule = get_object_or_404(ReportSchedule, id=schedule_id)
    else:
        schedule = None
    
    if request.method == 'POST':
        try:
            # 處理公司代號：資料同步類型自動設為 ALL
            report_type = request.POST.get('report_type')
            company = request.POST.get('company')
            if report_type == 'data_sync':
                company = 'ALL'  # 填報與現場記錄資料同步不分公司，自動設為 ALL
            
            # 處理執行時間：填報與現場記錄資料同步類型可以為空
            schedule_time = request.POST.get('schedule_time')
            if report_type == 'data_sync' and not schedule_time:
                schedule_time = '09:00:00'  # 填報與現場記錄資料同步預設時間
            
            if schedule:
                # 更新現有排程
                schedule.name = request.POST.get('name')
                schedule.report_type = report_type
                schedule.company = company
                schedule.schedule_time = schedule_time
                schedule.schedule_day = request.POST.get('schedule_day') or None
                schedule.file_format = request.POST.get('file_format', 'html')
                schedule.sync_execution_type = request.POST.get('sync_execution_type', 'interval')
                schedule.sync_interval_minutes = request.POST.get('sync_interval_minutes', 60)
                schedule.sync_fixed_time = request.POST.get('sync_fixed_time') or None
                schedule.status = request.POST.get('status')
                schedule.email_recipients = request.POST.get('email_recipients', '')
                schedule.save()
                
                # 同步到 Celery Beat
                from .report_schedule_sync_service import ReportScheduleSyncService
                sync_result = ReportScheduleSyncService.sync_report_schedules_to_celery()
                
                messages.success(request, '報表排程更新成功')
            else:
                # 建立新排程
                schedule = ReportSchedule.objects.create(
                    name=request.POST.get('name'),
                    report_type=report_type,
                    company=company,
                    schedule_time=schedule_time,
                    schedule_day=request.POST.get('schedule_day') or None,
                    file_format=request.POST.get('file_format', 'html'),
                    sync_execution_type=request.POST.get('sync_execution_type', 'interval'),
                    sync_interval_minutes=request.POST.get('sync_interval_minutes', 60),
                    sync_fixed_time=request.POST.get('sync_fixed_time') or None,
                    status=request.POST.get('status'),
                    email_recipients=request.POST.get('email_recipients', ''),
                )
                
                # 同步到 Celery Beat
                from .report_schedule_sync_service import ReportScheduleSyncService
                sync_result = ReportScheduleSyncService.sync_report_schedules_to_celery()
                
                messages.success(request, '報表排程建立成功')
            
            return redirect('reporting:report_schedule_list')
            
        except Exception as e:
            messages.error(request, f'儲存報表排程失敗: {str(e)}')
    
    # 獲取公司選項
    try:
        from erp_integration.models import CompanyConfig
        companies = CompanyConfig.objects.all().order_by('company_code')
    except Exception:
        companies = []
    
    context = {
        'schedule': schedule,
        'report_types': ReportSchedule.REPORT_TYPES,
        'file_formats': ReportSchedule.FILE_FORMATS,
        'status_choices': ReportSchedule.STATUS_CHOICES,
        'companies': companies,
    }
    return render(request, 'reporting/reporting/report_schedule_form.html', context)


@login_required
def report_execution_log(request):
    """報表執行日誌"""
    logs = ReportExecutionLog.objects.all().order_by('-execution_time')
    
    # 分頁
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'logs': page_obj,
    }
    return render(request, 'reporting/reporting/report_execution_log.html', context)


@require_http_methods(["GET"])
@csrf_exempt
def api_report_execution_logs(request):
    """API: 獲取報表執行日誌"""
    try:
        schedule_id = request.GET.get('schedule_id')
        
        if schedule_id:
            # 查詢特定排程的執行記錄
            logs = ReportExecutionLog.objects.filter(
                report_schedule_id=schedule_id
            ).order_by('-execution_time')
        else:
            # 查詢所有執行記錄
            logs = ReportExecutionLog.objects.all().order_by('-execution_time')
        
        # 轉換為 JSON 格式
        logs_data = []
        for log in logs:
            logs_data.append({
                'id': log.id,
                'report_schedule_id': log.report_schedule_id,
                'report_schedule_name': log.report_schedule_name,
                'execution_time': log.execution_time.strftime('%Y-%m-%d %H:%M:%S'),
                'status': log.status,
                'message': log.message,
                'file_path': log.file_path,
            })
        
        return JsonResponse({
            'success': True,
            'data': logs_data,
            'count': len(logs_data)
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
def view_report_file(request, file_path):
    """檢視報表檔案"""
    import os
    from django.http import HttpResponse, Http404
    from django.conf import settings
    
    try:
        # 解碼檔案路徑
        import urllib.parse
        decoded_file_path = urllib.parse.unquote(file_path)
        
        # 檢查檔案是否存在
        if not os.path.exists(decoded_file_path):
            raise Http404("報表檔案不存在")
        
        # 檢查檔案是否在允許的目錄內（安全檢查）
        media_root = settings.MEDIA_ROOT
        if not decoded_file_path.startswith(media_root):
            raise Http404("無權限存取此檔案")
        
        # 根據檔案類型決定回應方式
        file_extension = os.path.splitext(decoded_file_path)[1].lower()
        
        if file_extension == '.html':
            # HTML 檔案直接顯示
            with open(decoded_file_path, 'r', encoding='utf-8') as f:
                content = f.read()
            return HttpResponse(content, content_type='text/html; charset=utf-8')
        
        elif file_extension in ['.xlsx', '.xls']:
            # Excel 檔案下載
            with open(decoded_file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                filename = os.path.basename(decoded_file_path)
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
        
        elif file_extension == '.csv':
            # CSV 檔案下載
            with open(decoded_file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='text/csv; charset=utf-8')
                filename = os.path.basename(decoded_file_path)
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
        
        else:
            # 其他檔案類型下載
            with open(decoded_file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/octet-stream')
                filename = os.path.basename(decoded_file_path)
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
                return response
                
    except Exception as e:
        raise Http404(f"無法檢視報表檔案: {str(e)}")


@login_required
def execute_report_schedule(request, schedule_id):
    """手動執行報表排程"""
    try:
        schedule = get_object_or_404(ReportSchedule, id=schedule_id)
        scheduler = ReportScheduler()
        
        # 手動執行排程（跳過時間檢查）
        result = scheduler.execute_single_schedule(schedule)
        
        if result.get('success'):
            # 統一調度器已經處理了郵件發送，這裡只需要顯示結果
            if schedule.email_recipients:
                messages.success(request, f'報表排程 {schedule.name} 執行成功，郵件已發送')
            else:
                messages.success(request, f'報表排程 {schedule.name} 執行成功')
        else:
            messages.error(request, f'執行報表排程失敗: {result.get("message", "未知錯誤")}')
        
    except Exception as e:
        import traceback
        error_msg = f'執行報表排程失敗: {str(e)}'
        print(f"錯誤詳情: {error_msg}")
        print(f"錯誤追蹤: {traceback.format_exc()}")
        messages.error(request, error_msg)
    
    return redirect('reporting:report_schedule_list')


@login_required
def delete_report_schedule(request, schedule_id):
    """刪除報表排程"""
    try:
        schedule = get_object_or_404(ReportSchedule, id=schedule_id)
        schedule_name = schedule.name
        
        # 查詢相關的執行記錄
        execution_count = ReportExecutionLog.objects.filter(report_schedule_id=schedule.id).count()
        
        if request.method == 'POST':
            # 確認刪除
            schedule.delete()
            
            # 同步到 Celery Beat
            from .report_schedule_sync_service import ReportScheduleSyncService
            sync_result = ReportScheduleSyncService.sync_report_schedules_to_celery()
            
            messages.success(request, f'報表排程「{schedule_name}」已成功刪除')
            return redirect('reporting:report_schedule_list')
        else:
            # 顯示確認頁面
            context = {
                'schedule': schedule,
                'execution_count': execution_count,
            }
            return render(request, 'reporting/reporting/report_schedule_delete_confirm.html', context)
            
    except Exception as e:
        messages.error(request, f'刪除報表排程失敗: {str(e)}')
        return redirect('reporting:report_schedule_list')


@login_required
@require_POST
def sync_report_schedules(request):
    """同步報表排程到 Celery Beat"""
    try:
        from .report_schedule_sync_service import ReportScheduleSyncService
        
        # 執行同步
        result = ReportScheduleSyncService.sync_report_schedules_to_celery()
        
        if result['success']:
            messages.success(request, f"✅ {result['message']}")
        else:
            messages.error(request, f"❌ 同步失敗: {result.get('error', '未知錯誤')}")
            
    except Exception as e:
        messages.error(request, f"❌ 同步時發生錯誤: {str(e)}")
    
    return redirect('reporting:report_schedule_list')




@login_required
def sync_report_data(request):
    """同步報表資料"""
    try:
        from .data_sync import DataSyncService
        
        # 使用統一的資料同步核心函數
        # 手動同步：同步所有資料，不強制同步（會跳過已存在的記錄）
        result = DataSyncService.sync_fill_work_and_onsite_data(force_sync=False)
        
        if result['success']:
            # 檢查是否為 AJAX 請求
            if request.headers.get('Content-Type') == 'application/json' or request.headers.get('Accept') == 'application/json':
                return JsonResponse({
                    'success': True,
                    'message': result['message'],
                    'fill_works_synced': result['fill_works_synced'],
                    'onsite_synced': result['onsite_synced'],
                    'total_synced': result['total_synced'],
                    'fill_works_skipped': result.get('fill_works_skipped', 0),
                    'onsite_skipped': result.get('onsite_skipped', 0),
                    'total_skipped': result.get('total_skipped', 0)
                })
            else:
                messages.success(request, result['message'])
                return redirect('reporting:work_hour_report_index')
        else:
            # 檢查是否為 AJAX 請求
            if request.headers.get('Content-Type') == 'application/json' or request.headers.get('Accept') == 'application/json':
                return JsonResponse({
                    'success': False,
                    'error': result['error']
                })
            else:
                messages.error(request, result['error'])
                return redirect('reporting:work_hour_report_index')
        
    except Exception as e:
        error_message = f'同步報表資料失敗: {str(e)}'
        logger.error(error_message)
        
        # 檢查是否為 AJAX 請求
        if request.headers.get('Content-Type') == 'application/json' or request.headers.get('Accept') == 'application/json':
            return JsonResponse({
                'success': False,
                'error': error_message
            }, status=500)
        else:
            messages.error(request, error_message)
            return redirect('reporting:work_hour_report_index') 

@login_required
def chart_data(request):
    """提供圖表資料的 API"""
    from django.http import JsonResponse
    
    chart_type = request.GET.get('chart_type')
    
    if chart_type == 'work_hours_trend':
        # 工作時數趨勢圖資料
        end_date = timezone.now().date()
        start_date = end_date - timedelta(days=90)  # 改為90天
        
        # 取得最近90天的每日工作時數
        daily_data = WorkOrderReportData.objects.filter(
            work_date__range=[start_date, end_date]
        ).values('work_date').annotate(
            total_hours=Sum('daily_work_hours')
        ).order_by('work_date')
        
        labels = []
        values = []
        
        for item in daily_data:
            labels.append(item['work_date'].strftime('%m/%d'))
            values.append(float(item['total_hours'] or 0))
        
        return JsonResponse({
            'labels': labels,
            'values': values
        })
    
    elif chart_type == 'company_distribution':
        # 公司工作時數分布圖資料
        company_data = WorkOrderReportData.objects.values('company').annotate(
            total_hours=Sum('daily_work_hours')
        ).order_by('-total_hours')[:5]
        
        labels = []
        values = []
        
        for item in company_data:
            company_name = item['company'] or '未指定'
            labels.append(company_name)
            values.append(float(item['total_hours'] or 0))
        
        return JsonResponse({
            'labels': labels,
            'values': values
        })
    
    elif chart_type == 'completed_workorder_trend':
        # 已完工工單趨勢圖資料
        trend_data = CompletedWorkOrderReportService.get_completed_workorder_trend()
        return JsonResponse(trend_data)
    
    elif chart_type == 'completed_workorder_distribution':
        # 已完工工單公司分布圖資料
        company_data = CompletedWorkOrderReportService.get_company_completion_distribution()
        return JsonResponse(company_data)
    
    return JsonResponse({'error': '無效的圖表類型'})


@login_required
def report_data_list(request):
    """提供報表資料列表的 API"""
    from django.http import JsonResponse
    from django.core.paginator import Paginator
    
    # 取得分頁參數
    page = request.GET.get('page', 1)
    per_page = 20  # 每頁顯示20筆
    
    # 取得所有資料，按日期排序
    queryset = WorkOrderReportData.objects.all().order_by('-work_date')
    
    # 建立分頁器
    paginator = Paginator(queryset, per_page)
    
    try:
        page_obj = paginator.page(page)
    except:
        page_obj = paginator.page(1)
    
    data = []
    for item in page_obj:
        data.append({
            'workorder_id': item.workorder_id,
            'company': item.company or '未指定',
            'work_date': item.work_date.strftime('%Y-%m-%d'),
            'daily_work_hours': float(item.daily_work_hours or 0),
            'weekly_work_hours': float(item.weekly_work_hours or 0),
            'monthly_work_hours': float(item.monthly_work_hours or 0),
            'operator_count': item.operator_count or 0,
        })
    
    # 返回分頁資訊
    response_data = {
        'data': data,
        'pagination': {
            'current_page': page_obj.number,
            'total_pages': paginator.num_pages,
            'total_count': paginator.count,
            'has_previous': page_obj.has_previous(),
            'has_next': page_obj.has_next(),
            'previous_page': page_obj.previous_page_number() if page_obj.has_previous() else None,
            'next_page': page_obj.next_page_number() if page_obj.has_next() else None,
        }
    }
    
    return JsonResponse(response_data)

@login_required
def dashboard_stats(request):
    """提供儀表板統計資料的 API"""
    from django.http import JsonResponse
    
    try:
        # 統計資料 - 只使用存在的模型
        total_reports = WorkOrderReportData.objects.count()
        
        # 今日報表統計
        today = timezone.now().date()
        today_reports = WorkOrderReportData.objects.filter(
            work_date=today
        ).count()
        
        # 本週報表統計
        start_of_week = today - timedelta(days=today.weekday())
        end_of_week = start_of_week + timedelta(days=6)
        weekly_reports = WorkOrderReportData.objects.filter(
            work_date__range=[start_of_week, end_of_week]
        ).count()
        
        # 本月報表統計
        start_of_month = today.replace(day=1)
        monthly_reports = WorkOrderReportData.objects.filter(
            work_date__gte=start_of_month
        ).count()
        
        return JsonResponse({
            'total_reports': total_reports,
            'today_reports': today_reports,
            'pending_reports': weekly_reports,  # 使用週報表數量作為待處理
            'completed_reports': monthly_reports,  # 使用月報表數量作為已完成
        })
    except Exception as e:
        logger.error(f"取得儀表板統計資料失敗: {str(e)}")
        return JsonResponse({
            'total_reports': 0,
            'today_reports': 0,
            'pending_reports': 0,
            'completed_reports': 0,
        })

@login_required
def completed_workorder_report_index(request):
    """已完工工單報表首頁"""
    try:
        # 取得統計摘要
        summary = CompletedWorkOrderReportService.get_completed_workorder_summary()
        
        # 取得趨勢資料
        trend_data = CompletedWorkOrderReportService.get_completed_workorder_trend()
        
        # 取得公司分布資料
        company_data = CompletedWorkOrderReportService.get_company_completion_distribution()
        
        context = {
            'summary': summary,
            'trend_data': trend_data,
            'company_data': company_data,
        }
    except Exception as e:
        logger.error(f"取得已完工工單報表資料失敗: {str(e)}")
        context = {
            'summary': {},
            'trend_data': {},
            'company_data': {},
        }
    
    return render(request, 'reporting/reporting/completed_workorder_report.html', context)


 

# ==================== 評分報表視圖 ====================

@login_required
def scoring_report_list(request):
    """評分報表清單"""
    # 暫時重定向到作業員評分清單
    return redirect('reporting:operator_score_list')

@login_required
def scoring_report_detail(request, report_id):
    """評分報表詳情"""
    # 暫時重定向到作業員評分清單
    return redirect('reporting:operator_score_list')

@login_required
def generate_scoring_report(request):
    """生成評分報表"""
    if request.method == 'POST':
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        report_period = request.POST.get('report_period', 'monthly')
        
        if start_date and end_date:
            try:
                # 取得指定日期範圍內的評分資料
                scores = OperatorProcessCapacityScore.objects.filter(
                    work_date__range=[start_date, end_date]
                ).order_by('-work_date')
                
                if scores.exists():
                    # 計算統計資料
                    total_records = scores.count()
                    avg_capacity_score = scores.aggregate(Avg('capacity_score'))['capacity_score__avg'] or 0
                    avg_supervisor_score = scores.aggregate(Avg('supervisor_score'))['supervisor_score__avg'] or 80
                    avg_total_score = scores.aggregate(Avg('total_score'))['total_score__avg'] or 0
                    
                    # 等級分布
                    excellent_count = scores.filter(overall_grade='優秀').count()
                    good_count = scores.filter(overall_grade='良好').count()
                    pass_count = scores.filter(overall_grade='及格').count()
                    fail_count = scores.filter(overall_grade='不及格').count()
                    
                    # 建立報表記錄
                    report_data = {
                        'start_date': start_date,
                        'end_date': end_date,
                        'report_period': report_period,
                        'total_records': total_records,
                        'capacity_score': avg_capacity_score,
                        'supervisor_score': avg_supervisor_score,
                        'total_score': avg_total_score,
                        'excellent_count': excellent_count,
                        'good_count': good_count,
                        'pass_count': pass_count,
                        'fail_count': fail_count,
                    }
                    
                    messages.success(request, f'成功生成評分報表！共處理 {total_records} 筆資料')
                    return render(request, 'reporting/scoring_dashboard.html', {
                        'latest_report': report_data,
                        'operator_stats': {
                            'total_records': total_records,
                            'avg_capacity_score': avg_capacity_score,
                            'avg_supervisor_score': avg_supervisor_score,
                            'avg_total_score': avg_total_score,
                        },
                        'grade_distribution': [
                            {'overall_grade': '優秀', 'count': excellent_count},
                            {'overall_grade': '良好', 'count': good_count},
                            {'overall_grade': '及格', 'count': pass_count},
                            {'overall_grade': '不及格', 'count': fail_count},
                        ]
                    })
                else:
                    messages.warning(request, '指定日期範圍內沒有評分資料')
            except Exception as e:
                messages.error(request, f'生成評分報表失敗：{str(e)}')
        else:
            messages.error(request, '請選擇開始和結束日期')
    
    # 提供報表週期選項
    report_periods = [
        ('monthly', '月評分'),
        ('quarterly', '季評分'),
        ('yearly', '年評分'),
    ]
    
    context = {
        'report_periods': report_periods,
    }
    
    return render(request, 'reporting/generate_scoring_report.html', context)

@login_required
def operator_score_list(request):
    """作業員評分清單"""
    scores = OperatorProcessCapacityScore.objects.all().order_by('-work_date')
    
    # 篩選
    form = OperatorScoreFilterForm(request.GET)
    if form.is_valid():
        company_code = form.cleaned_data.get('company_code')
        operator_id = form.cleaned_data.get('operator_id')
        process_name = form.cleaned_data.get('process_name')
        start_date = form.cleaned_data.get('start_date')
        end_date = form.cleaned_data.get('end_date')
        has_supervisor_score = form.cleaned_data.get('has_supervisor_score')
        
        if company_code:
            scores = scores.filter(company_code=company_code)
        if operator_id:
            scores = scores.filter(operator_id=operator_id)
        if process_name:
            scores = scores.filter(process_name__icontains=process_name)
        if start_date:
            scores = scores.filter(work_date__gte=start_date)
        if end_date:
            scores = scores.filter(work_date__lte=end_date)
        if has_supervisor_score == 'yes':
            scores = scores.filter(is_supervisor_scored=True)
        elif has_supervisor_score == 'no':
            scores = scores.filter(is_supervisor_scored=False)
    
    # 分頁
    paginator = Paginator(scores, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'form': form,
    }
    
    return render(request, 'reporting/operator_score_list.html', context)

@login_required
def supervisor_score_form(request, score_id):
    """主管評分表單"""
    score_record = get_object_or_404(OperatorProcessCapacityScore, id=score_id)
    
    if request.method == 'POST':
        form = SupervisorScoreForm(request.POST, instance=score_record)
        if form.is_valid():
            # 更新主管評分資訊
            score_record.supervisor_name = request.user.get_full_name() or request.user.username
            score_record.supervisor_date = timezone.now()
            score_record.is_supervisor_scored = True
            
            # 重新計算總評分
            score_record.total_score = score_record.calculate_total_score()
            score_record.overall_grade = score_record.get_grade(score_record.total_score)
            
            score_record.save()
            
            messages.success(request, '主管評分已成功儲存！')
            return redirect('reporting:operator_score_list')
    else:
        form = SupervisorScoreForm(instance=score_record)
    
    context = {
        'form': form,
        'score_record': score_record,
    }
    
    return render(request, 'reporting/supervisor_score_form.html', context)

@login_required
def operator_score_detail(request, score_id):
    """作業員評分詳情"""
    score_record = get_object_or_404(OperatorProcessCapacityScore, id=score_id)
    
    context = {
        'score_record': score_record,
    }
    
    return render(request, 'reporting/operator_score_detail.html', context)

@login_required
def export_scoring_report(request, report_id):
    """匯出評分報表"""
    # ScoringReport 和 ScoringDetail 模型已移除，此功能不再可用
    messages.warning(request, '評分報表匯出功能已移除，請直接查看詳情。')
    return redirect('reporting:operator_score_list')

@login_required
def scoring_dashboard(request):
    """評分儀表板"""
    # 取得作業員評分統計
    operator_stats = OperatorProcessCapacityScore.objects.aggregate(
        total_records=Count('id'),
        avg_capacity_score=Avg('capacity_score'),
        avg_total_score=Avg('total_score'),
        scored_by_supervisor=Count('id', filter=Q(is_supervisor_scored=True))
    )
    
    # 取得各等級分布
    grade_distribution = OperatorProcessCapacityScore.objects.values('overall_grade').annotate(
        count=Count('id')
    ).order_by('overall_grade')
    
    context = {
        'operator_stats': operator_stats,
        'grade_distribution': grade_distribution,
    }
    
    return render(request, 'reporting/scoring_dashboard.html', context)

@login_required
def score_period_management(request):
    """評分週期管理"""
    from .score_period_service import ScorePeriodService
    
    if request.method == 'POST':
        action = request.POST.get('action')
        company_code = request.POST.get('company_code')
        period_type = request.POST.get('period_type', 'monthly')
        
        if action == 'create_period':
            try:
                records = ScorePeriodService.create_period_score_records(
                    company_code, period_type, force=True
                )
                messages.success(request, f'成功建立 {len(records)} 筆{period_type}評分記錄')
            except Exception as e:
                messages.error(request, f'建立評分記錄失敗：{str(e)}')
        
        elif action == 'close_period':
            try:
                count = ScorePeriodService.close_period(company_code, period_type)
                messages.success(request, f'成功關閉 {count} 筆{period_type}評分記錄')
            except Exception as e:
                messages.error(request, f'關閉評分週期失敗：{str(e)}')
    
    # 取得各週期的摘要
    company_code = request.GET.get('company_code', '')
    period_summaries = {}
    
    if company_code:
        for period_type in ['monthly', 'quarterly', 'yearly']:
            summary = ScorePeriodService.get_period_summary(company_code, period_type)
            if summary:
                period_summaries[period_type] = summary
    
    context = {
        'company_code': company_code,
        'period_summaries': period_summaries,
    }
    
    return render(request, 'reporting/score_period_management.html', context)

@login_required
def score_period_detail(request, period_type):
    """評分週期詳情"""
    from .score_period_service import ScorePeriodService
    
    company_code = request.GET.get('company_code', '')
    if not company_code:
        messages.error(request, '請選擇公司代號')
        return redirect('reporting:score_period_management')
    
    # 取得週期摘要
    summary = ScorePeriodService.get_period_summary(company_code, period_type)
    if not summary:
        messages.error(request, f'找不到{period_type}評分資料')
        return redirect('reporting:score_period_management')
    
    # 取得詳細記錄
    start_date, end_date = ScorePeriodService.get_current_period_dates(period_type)
    records = OperatorProcessCapacityScore.objects.filter(
        company_code=company_code,
        score_period=period_type,
        period_start_date=start_date,
        period_end_date=end_date
    ).order_by('-work_date')
    
    # 分頁
    paginator = Paginator(records, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'summary': summary,
        'page_obj': page_obj,
        'period_type': period_type,
        'company_code': company_code,
    }
    
    return render(request, 'reporting/score_period_detail.html', context)

@login_required
def work_hour_stats(request):
    """提供工作時數統計資料的 API"""
    from django.http import JsonResponse
    
    try:
        # 基本統計資料
        total_records = WorkOrderReportData.objects.count()
        
        # 總工作時數
        total_hours = WorkOrderReportData.objects.aggregate(
            total=Sum('daily_work_hours')
        )['total'] or 0
        
        # 作業員數量（去重）
        total_operators = WorkOrderReportData.objects.exclude(
            operator_name__isnull=True
        ).values('operator_name').distinct().count()
        
        # 平均日工作時數
        avg_daily_hours = WorkOrderReportData.objects.aggregate(
            avg=Avg('daily_work_hours')
        )['avg'] or 0
        
        return JsonResponse({
            'total_records': total_records,
            'total_hours': float(total_hours),
            'total_operators': total_operators,
            'avg_daily_hours': float(avg_daily_hours),
        })
    except Exception as e:
        logger.error(f"取得工作時數統計資料失敗: {str(e)}")
        return JsonResponse({
            'total_records': 0,
            'total_hours': 0,
            'total_operators': 0,
            'avg_daily_hours': 0,
        })


@login_required
def detailed_stats(request):
    """提供詳細統計資料的 API"""
    from django.http import JsonResponse
    
    try:
        # 按月份統計
        monthly_stats = WorkOrderReportData.objects.annotate(
            year_month=ExtractYear('work_date') * 100 + ExtractMonth('work_date')
        ).values('year_month').annotate(
            total_hours=Sum('daily_work_hours'),
            record_count=Count('id')
        ).order_by('year_month')
        
        monthly_data = {
            'labels': [],
            'hours': [],
            'counts': []
        }
        
        for item in monthly_stats:
            year = item['year_month'] // 100
            month = item['year_month'] % 100
            monthly_data['labels'].append(f"{year}-{month:02d}")
            monthly_data['hours'].append(float(item['total_hours'] or 0))
            monthly_data['counts'].append(item['record_count'])
        
        # 按公司統計
        company_stats = WorkOrderReportData.objects.values('company').annotate(
            total_hours=Sum('daily_work_hours'),
            record_count=Count('id')
        ).order_by('-total_hours')
        
        company_data = {
            'labels': [],
            'hours': [],
            'counts': []
        }
        
        for item in company_stats:
            company_name = item['company'] or '未指定'
            company_data['labels'].append(company_name)
            company_data['hours'].append(float(item['total_hours'] or 0))
            company_data['counts'].append(item['record_count'])
        
        return JsonResponse({
            'monthly': monthly_data,
            'company': company_data
        })
    except Exception as e:
        logger.error(f"取得詳細統計資料失敗: {str(e)}")
        return JsonResponse({
            'monthly': {'labels': [], 'hours': [], 'counts': []},
            'company': {'labels': [], 'hours': [], 'counts': []}
        })

@login_required
def test_workday_calendar(request):
    """測試工作日曆服務"""
    from .workday_calendar import WorkdayCalendarService
    from datetime import date, timedelta
    
    calendar_service = WorkdayCalendarService()
    current_date = date.today()
    
    # 繁體中文星期對應
    weekday_names = {
        0: '星期一',
        1: '星期二', 
        2: '星期三',
        3: '星期四',
        4: '星期五',
        5: '星期六',
        6: '星期日'
    }
    
    # 測試資料
    test_results = {
        'current_date': current_date,
        'current_weekday': weekday_names[current_date.weekday()],
        'is_current_workday': calendar_service.is_workday(current_date),
        'previous_workday': calendar_service.get_previous_workday(current_date),
        'previous_weekday': weekday_names[calendar_service.get_previous_workday(current_date).weekday()],
        'next_5_days': [],
        'previous_5_days': [],
    }
    
    # 測試未來5天
    for i in range(1, 6):
        test_date = current_date + timedelta(days=i)
        test_results['next_5_days'].append({
            'date': test_date,
            'is_workday': calendar_service.is_workday(test_date),
            'weekday': weekday_names[test_date.weekday()]
        })
    
    # 測試過去5天
    for i in range(1, 6):
        test_date = current_date - timedelta(days=i)
        test_results['previous_5_days'].append({
            'date': test_date,
            'is_workday': calendar_service.is_workday(test_date),
            'weekday': weekday_names[test_date.weekday()]
        })
    
    context = {
        'test_results': test_results,
    }
    return render(request, 'reporting/reporting/test_workday_calendar.html', context)

@login_required
def holiday_setup_management(request):
    """假期設定管理"""
    from .holiday_auto_setup_service import HolidayAutoSetupService
    from .workday_calendar import WorkdayCalendarService
    from datetime import date, timedelta
    
    setup_service = HolidayAutoSetupService()
    calendar_service = WorkdayCalendarService()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'setup_2025':
            # 設定 2025 年假期
            result = setup_service.setup_all_2025()
            messages.success(request, f'2025年假期設定完成！國定假日：{result["holidays"]}個，補班日：{result["makeup_days"]}個')
            
        elif action == 'clear_2025':
            # 清除 2025 年假期
            deleted_count = setup_service.clear_all_holidays(2025)
            messages.success(request, f'2025年假期設定已清除，共刪除 {deleted_count} 個事件')
            
        elif action == 'add_holiday':
            # 手動新增假期
            holiday_date = request.POST.get('holiday_date')
            holiday_name = request.POST.get('holiday_name')
            description = request.POST.get('description', '')
            
            if holiday_date and holiday_name:
                try:
                    date_obj = datetime.strptime(holiday_date, '%Y-%m-%d').date()
                    calendar_service.add_holiday(
                        date=date_obj,
                        holiday_name=holiday_name,
                        description=description,
                        created_by=request.user.username
                    )
                    messages.success(request, f'成功新增假期：{holiday_name} ({holiday_date})')
                except Exception as e:
                    messages.error(request, f'新增假期失敗：{str(e)}')
            else:
                messages.error(request, '請填寫假期日期和名稱')
                
        elif action == 'add_workday':
            # 手動新增補班日
            workday_date = request.POST.get('workday_date')
            description = request.POST.get('description', '')
            
            if workday_date:
                try:
                    date_obj = datetime.strptime(workday_date, '%Y-%m-%d').date()
                    calendar_service.add_workday(
                        date=date_obj,
                        description=description,
                        created_by=request.user.username
                    )
                    messages.success(request, f'成功新增補班日：{workday_date}')
                except Exception as e:
                    messages.error(request, f'新增補班日失敗：{str(e)}')
            else:
                messages.error(request, '請填寫補班日期')
    
    # 取得當前日期
    current_date = date.today()
    
    context = {
        'current_date': current_date,
    }
    return render(request, 'reporting/reporting/holiday_setup_management.html', context)

@login_required
def execute_previous_workday_report(request):
    """手動執行前一個工作日報表"""
    if request.method == 'POST':
        try:
            scheduler = ReportScheduler()
            
            # 使用 ReportGenerator 生成前一個工作日報表
            generator = ReportGenerator()
            report_date = generator._collect_previous_workday_data(timezone.now().date() - timedelta(days=1))['date']
            
            # 生成報表（手動執行時預設生成 HTML 格式）
            file_paths = generator.generate_previous_workday_report(report_date, 'html')
            
            # 手動執行不發送郵件，因為沒有排程設定
            messages.success(request, f'前一個工作日報表執行成功！報表日期：{report_date}')
            
        except Exception as e:
            messages.error(request, f'執行前一個工作日報表失敗：{str(e)}')
    
    return redirect('reporting:index')


@login_required
def test_previous_workday_report(request):
    """測試前一個工作日報表功能"""
    from datetime import date, timedelta
    
    scheduler = ReportScheduler()
    
    # 測試資料
    test_results = {
        'current_date': date.today(),
        'current_time': timezone.now().time(),
        'should_execute': scheduler.should_execute(),
        'report_date': scheduler.get_report_date(),
        'next_5_days': [],
    }
    
    # 測試未來5天的報表日期
    for i in range(5):
        test_date = date.today() + timedelta(days=i)
        test_results['next_5_days'].append({
            'date': test_date,
            'report_date': scheduler.calendar_service.get_previous_workday(test_date),
            'is_workday': scheduler.calendar_service.is_workday(test_date),
        })
    
    # 測試資料收集
    try:
        data = scheduler.collect_data(test_results['report_date'])
        test_results['data_collection'] = {
            'success': True,
            'fill_works_count': data['fill_works_count'],
            'onsite_reports_count': data['onsite_reports_count'],
            'total_work_hours': data['total_work_hours'],
            'total_operators': data['total_operators'],
            'total_equipment': data['total_equipment'],
        }
    except Exception as e:
        test_results['data_collection'] = {
            'success': False,
            'error': str(e)
        }
    
    context = {
        'test_results': test_results,
    }
    return render(request, 'reporting/reporting/test_previous_workday_report.html', context)

@login_required
def previous_workday_report_management(request):
    """前一個工作日報表管理"""
    from .workday_calendar import WorkdayCalendarService
    from datetime import date, timedelta
    import os
    
    scheduler = ReportScheduler()
    calendar_service = WorkdayCalendarService()
    
    # 取得當前狀態
    current_date = date.today()
    # 修正時區問題，確保使用正確的本地時間
    current_datetime = timezone.localtime(timezone.now())
    current_time = current_datetime.time()
    report_date = scheduler.get_report_date()
    should_execute = scheduler.should_execute()
    
    # 檢查報表檔案
    report_dir = os.path.join(settings.MEDIA_ROOT, 'reports', 'previous_workday')
    report_files = []
    if os.path.exists(report_dir):
        files = os.listdir(report_dir)
        files.sort(reverse=True)  # 最新的檔案在前
        for file in files[:10]:  # 只顯示最近10個檔案
            file_path = os.path.join(report_dir, file)
            file_stat = os.stat(file_path)
            report_files.append({
                'filename': file,
                'filepath': file_path,
                'size': file_stat.st_size,
                'created_time': datetime.fromtimestamp(file_stat.st_ctime),
                'modified_time': datetime.fromtimestamp(file_stat.st_mtime),
            })
    
    # 取得未來7天的報表日期預測
    next_7_days = []
    for i in range(7):
        test_date = current_date + timedelta(days=i)
        next_7_days.append({
            'date': test_date,
            'report_date': calendar_service.get_previous_workday(test_date),
            'is_workday': calendar_service.is_workday(test_date),
        })
    
    # 取得最近7天的報表日期
    past_7_days = []
    for i in range(7):
        test_date = current_date - timedelta(days=i)
        past_7_days.append({
            'date': test_date,
            'report_date': calendar_service.get_previous_workday(test_date),
            'is_workday': calendar_service.is_workday(test_date),
        })
    
    # 測試資料收集
    try:
        data = scheduler.collect_data(report_date)
        data_collection = {
            'success': True,
            'fill_works_count': data['fill_works_count'],
            'onsite_reports_count': data['onsite_reports_count'],
            'total_work_hours': data['total_work_hours'],
            'total_operators': data['total_operators'],
            'total_equipment': data['total_equipment'],
        }
    except Exception as e:
        data_collection = {
            'success': False,
            'error': str(e)
        }
    
    context = {
        'current_date': current_date,
        'current_time': current_time,
        'report_date': report_date,
        'should_execute': should_execute,
        'report_files': report_files,
        'next_7_days': next_7_days,
        'past_7_days': past_7_days,
        'data_collection': data_collection,
    }
    
    return render(request, 'reporting/reporting/previous_workday_report_management.html', context)

@login_required
def government_calendar_sync(request):
    """政府行事曆 API 同步管理"""
    from .taiwan_government_calendar_service import TaiwanGovernmentCalendarService
    from datetime import datetime
    
    gov_service = TaiwanGovernmentCalendarService()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'sync_single_year':
            # 同步單一年份
            year = int(request.POST.get('year', datetime.now().year))
            result = gov_service.sync_government_holidays(year)
            
            if result['success']:
                messages.success(request, result['message'])
            else:
                messages.error(request, result['message'])
                
        elif action == 'sync_multiple_years':
            # 同步多個年份
            start_year = int(request.POST.get('start_year', datetime.now().year))
            end_year = int(request.POST.get('end_year', datetime.now().year + 1))
            result = gov_service.sync_multiple_years(start_year, end_year)
            
            if result['success']:
                messages.success(request, result['message'])
            else:
                messages.error(request, result['message'])
                
        elif action == 'check_status':
            # 檢查狀態（不執行同步，只檢查）
            year = int(request.POST.get('year', datetime.now().year))
            status = gov_service.get_holiday_status(year)
            
            if status['government_data_available']:
                if status['sync_needed']:
                    messages.info(request, f"{year}年政府資料可用，建議同步（系統: {status['system_holidays']}個，政府: {status['government_holidays']}個）")
                else:
                    messages.success(request, f"{year}年假期資料已是最新（系統: {status['system_holidays']}個）")
            else:
                messages.warning(request, f"{year}年政府資料不可用")
    
    # 取得可用年份
    available_years = gov_service.get_available_years()
    current_year = datetime.now().year
    
    # 檢查當前年份狀態
    current_status = gov_service.get_holiday_status(current_year)
    
    # 檢查未來幾年狀態
    future_status = {}
    for year in range(current_year, current_year + 3):
        future_status[year] = gov_service.get_holiday_status(year)
    
    context = {
        'available_years': available_years,
        'current_year': current_year,
        'current_status': current_status,
        'future_status': future_status,
    }
    
    return render(request, 'reporting/reporting/government_calendar_sync.html', context)

@login_required
def csv_holiday_import(request):
    """CSV 國定假日匯入管理"""
    from .csv_holiday_import_service import CSVHolidayImportService
    from django.http import HttpResponse
    
    csv_service = CSVHolidayImportService()
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'import_csv':
            # 處理 CSV 檔案匯入
            if 'csv_file' in request.FILES:
                csv_file = request.FILES['csv_file']
                
                # 檢查檔案類型
                if not csv_file.name.endswith('.csv'):
                    messages.error(request, '請上傳 CSV 格式的檔案')
                    return redirect('reporting:csv_holiday_import')
                
                # 檢查檔案大小（限制 5MB）
                if csv_file.size > 5 * 1024 * 1024:
                    messages.error(request, '檔案大小不能超過 5MB')
                    return redirect('reporting:csv_holiday_import')
                
                # 匯入 CSV 檔案
                result = csv_service.import_holidays_from_csv(csv_file)
                
                if result['success']:
                    messages.success(request, result['message'])
                else:
                    messages.error(request, result['message'])
                
                # 如果有錯誤，顯示詳細錯誤訊息
                if result.get('errors'):
                    for error in result['errors'][:5]:  # 只顯示前5個錯誤
                        messages.warning(request, error)
                    if len(result['errors']) > 5:
                        messages.warning(request, f"... 還有 {len(result['errors']) - 5} 個錯誤")
                
            else:
                messages.error(request, '請選擇要匯入的 CSV 檔案')
                
        elif action == 'download_sample':
            # 下載範例 CSV 檔案
            sample_content = csv_service.generate_sample_csv()
            response = HttpResponse(sample_content, content_type='text/csv; charset=utf-8')
            response['Content-Disposition'] = 'attachment; filename="holiday_sample.csv"'
            return response
    
    # 取得匯入統計
    from scheduling.models import Event
    from datetime import datetime
    
    current_year = datetime.now().year
    csv_imported_holidays = Event.objects.filter(
        type='holiday',
        created_by='csv_import',
        start__year=current_year
    ).count()
    
    total_holidays = Event.objects.filter(
        type='holiday',
        start__year=current_year
    ).count()
    
    context = {
        'current_year': current_year,
        'csv_imported_holidays': csv_imported_holidays,
        'total_holidays': total_holidays,
    }
    
    return render(request, 'reporting/reporting/csv_holiday_import.html', context)

class CompletedWorkOrderAnalysisIndexView(LoginRequiredMixin, TemplateView):
    """已完工工單分析報表首頁"""
    template_name = 'reporting/completed_workorder_analysis.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 取得統計摘要
        from .models import CompletedWorkOrderAnalysis
        from workorder.models import CompletedWorkOrder
        
        # 已分析工單數
        total_analyzed = CompletedWorkOrderAnalysis.objects.count()
        
        # 已完工工單數
        total_completed = CompletedWorkOrder.objects.count()
        
        # 待分析工單數
        pending_analysis = total_completed - total_analyzed
        
        # 最後分析時間
        last_analysis = CompletedWorkOrderAnalysis.objects.order_by('-created_at').first()
        if last_analysis:
            # 轉換為本地時間
            from django.utils import timezone
            local_time = timezone.localtime(last_analysis.created_at)
            last_analysis_date = local_time.strftime('%Y-%m-%d %H:%M')
        else:
            last_analysis_date = '無'
        
        # 平均執行天數
        avg_execution_days = CompletedWorkOrderAnalysis.objects.aggregate(
            avg_days=Avg('total_execution_days')
        )['avg_days'] or 0
        
        # 平均工作時數
        avg_work_hours = CompletedWorkOrderAnalysis.objects.aggregate(
            avg_hours=Avg('total_work_hours')
        )['avg_hours'] or 0
        
        # 平均工序數
        avg_processes = CompletedWorkOrderAnalysis.objects.aggregate(
            avg_processes=Avg('total_processes')
        )['avg_processes'] or 0
        
        context.update({
            'summary': {
                'total_workorders': total_analyzed,
                'avg_execution_days': avg_execution_days,
                'avg_work_hours': avg_work_hours,
                'avg_processes': avg_processes,
            }
        })
        
        return context


class CompletedWorkOrderAnalysisListView(LoginRequiredMixin, ListView):
    """已完工工單分析列表"""
    model = CompletedWorkOrderAnalysis
    template_name = 'reporting/completed_workorder_analysis_list.html'
    context_object_name = 'analyses'
    paginate_by = 20
    ordering = ['-created_at']
    
    def get_template_names(self):
        # 如果是 AJAX 請求，使用部分模板
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return ['reporting/completed_workorder_analysis_list_ajax.html']
        return super().get_template_names()
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        # 公司篩選
        company = self.request.GET.get('company')
        if company:
            queryset = queryset.filter(company_code=company)
        
        # 工單編號篩選
        workorder_id = self.request.GET.get('workorder_id')
        if workorder_id:
            queryset = queryset.filter(workorder_id__icontains=workorder_id)
        
        # 產品編號篩選
        product_code = self.request.GET.get('product_code')
        if product_code:
            queryset = queryset.filter(product_code__icontains=product_code)
        
        # 日期範圍篩選
        start_date = self.request.GET.get('start_date')
        if start_date:
            queryset = queryset.filter(completion_date__gte=start_date)
        
        end_date = self.request.GET.get('end_date')
        if end_date:
            queryset = queryset.filter(completion_date__lte=end_date)
        
        # 執行天數篩選
        min_days = self.request.GET.get('min_days')
        if min_days:
            queryset = queryset.filter(total_execution_days__gte=min_days)
        
        max_days = self.request.GET.get('max_days')
        if max_days:
            queryset = queryset.filter(total_execution_days__lte=max_days)
        
        # 工作時數篩選
        min_hours = self.request.GET.get('min_hours')
        if min_hours:
            queryset = queryset.filter(total_work_hours__gte=min_hours)
        
        max_hours = self.request.GET.get('max_hours')
        if max_hours:
            queryset = queryset.filter(total_work_hours__lte=max_hours)
        
        # 效率比率篩選
        efficiency_min = self.request.GET.get('efficiency_min')
        if efficiency_min:
            queryset = queryset.filter(efficiency_rate__gte=efficiency_min)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 取得公司列表
        from erp_integration.models import CompanyConfig
        companies = CompanyConfig.objects.values_list('company_code', 'company_name')
        context['companies'] = companies
        
        # 篩選條件
        context['selected_company'] = self.request.GET.get('company', '')
        context['selected_workorder_id'] = self.request.GET.get('workorder_id', '')
        context['selected_product_code'] = self.request.GET.get('product_code', '')
        context['selected_start_date'] = self.request.GET.get('start_date', '')
        context['selected_end_date'] = self.request.GET.get('end_date', '')
        context['selected_min_days'] = self.request.GET.get('min_days', '')
        context['selected_max_days'] = self.request.GET.get('max_days', '')
        context['selected_min_hours'] = self.request.GET.get('min_hours', '')
        context['selected_max_hours'] = self.request.GET.get('max_hours', '')
        context['selected_efficiency_min'] = self.request.GET.get('efficiency_min', '')
        
        return context


class CompletedWorkOrderAnalysisDetailView(LoginRequiredMixin, DetailView):
    """已完工工單分析詳情"""
    model = CompletedWorkOrderAnalysis
    template_name = 'reporting/completed_workorder_analysis_detail.html'
    context_object_name = 'analysis'
    
    def get_object(self, queryset=None):
        return get_object_or_404(CompletedWorkOrderAnalysis, id=self.kwargs['analysis_id'])
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 按照工序的實際執行順序排列統計資料，與已完工工單詳情頁面完全一致
        if self.object.process_details:
            # 定義工序優先順序，出貨包裝必須排在最後
            def get_process_priority(process_name):
                if process_name == "出貨包裝":
                    return 9999  # 出貨包裝排在最後，不按時間順序
                # 其他工序按第一次出現的順序排列
                process_data = self.object.process_details.get(process_name, {})
                return process_data.get('first_appearance_order', 999)
            
            # 按照工序執行順序排序統計資料
            sorted_process_details = dict(sorted(
                self.object.process_details.items(),
                key=lambda x: get_process_priority(x[0])
            ))
            
            # 轉換為列表格式供模板使用
            context['sorted_process_details'] = list(sorted_process_details.items())
        else:
            context['sorted_process_details'] = []
        
        return context
    
    def _get_sorted_process_details(self, process_details):
        """取得按工序執行順序排序的工序詳細資料，與已完工工單詳情頁面完全一致"""
        if not process_details:
            return []
        
        # 定義工序優先順序，出貨包裝必須排在最後
        def get_process_priority(process_name):
            if process_name == "出貨包裝":
                return 9999  # 出貨包裝排在最後，不按時間順序
            # 其他工序按第一次出現的順序排列
            process_data = process_details.get(process_name, {})
            return process_data.get('first_appearance_order', 999)
        
        # 按照工序執行順序排序統計資料
        sorted_process_details = dict(sorted(
            process_details.items(),
            key=lambda x: get_process_priority(x[0])
        ))
        
        # 轉換為列表格式
        return list(sorted_process_details.items())
    
    def post(self, request, *args, **kwargs):
        """處理匯出請求"""
        action = request.POST.get('action')
        export_format = request.POST.get('format', 'excel')
        
        if action == 'export':
            if export_format == 'html':
                return self.export_analysis_html()
            else:
                return self.export_analysis_excel()
        
        return super().post(request, *args, **kwargs)
    
    def export_analysis_excel(self):
        """匯出工單分析詳情為 Excel 格式"""
        import pandas as pd
        from io import BytesIO
        from django.http import HttpResponse
        from datetime import datetime
        
        analysis = self.get_object()
        
        # 建立 Excel 檔案
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            
            # 工作表1：基本資訊
            basic_data = {
                '項目': [
                    '工單編號', '公司代號', '公司名稱', '產品編號', '產品名稱',
                    '訂單數量', '完工日期', '分析時間', '總執行天數', '總工作時數',
                    '總工序數', '效率比率'
                ],
                '數值': [
                    analysis.workorder_id,
                    analysis.company_code,
                    analysis.company_name,
                    analysis.product_code,
                    analysis.product_name,
                    analysis.order_quantity,
                    analysis.completion_date.strftime('%Y-%m-%d') if analysis.completion_date else '',
                    analysis.created_at.strftime('%Y-%m-%d %H:%M') if analysis.created_at else '',
                    analysis.total_execution_days,
                    f"{analysis.total_work_hours:.1f}",
                    analysis.total_processes,
                    f"{analysis.efficiency_rate:.1%}" if analysis.efficiency_rate else ''
                ]
            }
            
            df_basic = pd.DataFrame(basic_data)
            df_basic.to_excel(writer, sheet_name='基本資訊', index=False)
            
            # 工作表2：時間分析
            time_data = [
                ['第一筆紀錄日期', analysis.first_record_date.strftime('%Y-%m-%d') if analysis.first_record_date else ''],
                ['最後一筆紀錄日期', analysis.last_record_date.strftime('%Y-%m-%d') if analysis.last_record_date else ''],
                ['總執行天數', analysis.total_execution_days],
                ['總工作時數', f"{analysis.total_work_hours:.1f}"],
                ['總加班時數', f"{analysis.total_overtime_hours:.1f}"],
                ['平均每日工作時數', f"{analysis.average_daily_hours:.1f}"],
                ['效率比率', f"{analysis.efficiency_rate:.1%}" if analysis.efficiency_rate else ''],
                ['完工日期', analysis.completion_date.strftime('%Y-%m-%d') if analysis.completion_date else ''],
                ['完工狀態', analysis.completion_status],
            ]
            
            df_time = pd.DataFrame(time_data, columns=['項目', '數值'])
            df_time.to_excel(writer, sheet_name='時間分析', index=False)
            
            # 工作表3：工序分析
            if analysis.process_details:
                process_data = []
                # 按時間排序工序（與網頁顯示一致）
                sorted_processes = self._get_sorted_process_details(analysis.process_details)
                for process_name, process_info in sorted_processes:
                    process_data.append({
                        '工序名稱': process_name,
                        '總工作時數': process_info.get('total_hours', 0),
                        '總加班時數': process_info.get('total_overtime_hours', 0),
                        '總完成數量': process_info.get('total_quantity', 0),
                        '每小時產能': f"{process_info.get('hourly_capacity', 0):.1f} pcs/hr" if process_info.get('hourly_capacity', 0) > 0 else '-',
                        '參與作業員數': process_info.get('operator_count', 0),
                        '第一筆記錄': process_info.get('first_record_date', ''),
                        '最後一筆記錄': process_info.get('last_record_date', ''),
                        '平均每日工作時數': process_info.get('avg_daily_hours', 0)
                    })
                
                df_process = pd.DataFrame(process_data)
                df_process.to_excel(writer, sheet_name='工序分析', index=False)
            
            # 工作表4：作業員分析
            if analysis.operator_details:
                operator_data = []
                for operator_name, operator_info in analysis.operator_details.items():
                    operator_data.append({
                        '作業員姓名': operator_name,
                        '總工作時數': operator_info.get('total_hours', 0),
                        '加班時數': operator_info.get('overtime_hours', 0),
                        '參與工序': ', '.join(operator_info.get('processes', [])),
                        '工作天數': operator_info.get('work_days', 0),
                        '參與工序數': operator_info.get('process_count', 0)
                    })
                
                df_operator = pd.DataFrame(operator_data)
                df_operator.to_excel(writer, sheet_name='作業員分析', index=False)
            
            # 工作表5：詳細統計
            stats_data = [
                ['不重複工序數', analysis.unique_processes],
                ['參與作業員數', analysis.total_operators],
                ['分析建立時間', analysis.created_at.strftime('%Y-%m-%d %H:%M') if analysis.created_at else ''],
                ['分析更新時間', analysis.updated_at.strftime('%Y-%m-%d %H:%M') if analysis.updated_at else ''],
            ]
            
            df_stats = pd.DataFrame(stats_data, columns=['統計項目', '數值'])
            df_stats.to_excel(writer, sheet_name='詳細統計', index=False)
        
        # 設定檔案名稱
        import re
        from urllib.parse import quote
        # 清理檔案名稱，移除或替換可能導致問題的字符
        company_name = re.sub(r'[<>:"/\\|?*\s]', '_', analysis.company_name)
        workorder_id = re.sub(r'[<>:"/\\|?*\s]', '_', analysis.workorder_id)
        product_code = re.sub(r'[<>:"/\\|?*\s]', '_', analysis.product_code)
        raw_filename = f"{company_name}_{workorder_id}_{product_code}.xlsx"
        filename = raw_filename
        
        # 設定回應標頭
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{quote(filename)}"; filename*=UTF-8\'\'{quote(filename)}'
        
        return response
    
    def export_analysis_html(self):
        """匯出工單分析詳情為 HTML 格式"""
        from django.template.loader import render_to_string
        from django.http import HttpResponse
        from datetime import datetime
        
        analysis = self.get_object()
        
        # 準備匯出用的資料
        export_data = {
            'analysis': analysis,
            'export_time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'company_info': f"{analysis.company_name} ({analysis.company_code})",
            'basic_info': {
                '工單編號': analysis.workorder_id,
                '公司': f"{analysis.company_name} ({analysis.company_code})",
                '產品編號': analysis.product_code,
                '產品名稱': analysis.product_name,
                '訂單數量': analysis.order_quantity,
                '完工日期': analysis.completion_date.strftime('%Y-%m-%d') if analysis.completion_date else '',
                '分析時間': analysis.created_at.strftime('%Y-%m-%d %H:%M') if analysis.created_at else '',
            },
            'metrics': {
                '總執行天數': analysis.total_execution_days,
                '總工作時數': f"{analysis.total_work_hours:.1f}",
                '總工序數': analysis.total_processes,
                '效率比率': f"{analysis.efficiency_rate:.1%}" if analysis.efficiency_rate else '',
            },
            'time_analysis': {
                '第一筆紀錄日期': analysis.first_record_date.strftime('%Y-%m-%d') if analysis.first_record_date else '',
                '最後一筆紀錄日期': analysis.last_record_date.strftime('%Y-%m-%d') if analysis.last_record_date else '',
                '總執行天數': analysis.total_execution_days,
                '總工作時數': f"{analysis.total_work_hours:.1f}",
                '總加班時數': f"{analysis.total_overtime_hours:.1f}",
                '平均每日工作時數': f"{analysis.average_daily_hours:.1f}",
                '效率比率': f"{analysis.efficiency_rate:.1%}" if analysis.efficiency_rate else '',
                '完工日期': analysis.completion_date.strftime('%Y-%m-%d') if analysis.completion_date else '',
                '完工狀態': analysis.completion_status,
            },
            'process_details': self._get_sorted_process_details(analysis.process_details or {}),
            'operator_details': analysis.operator_details or {},
            'statistics': {
                '不重複工序數': analysis.unique_processes,
                '參與作業員數': analysis.total_operators,
                '分析建立時間': analysis.created_at.strftime('%Y-%m-%d %H:%M') if analysis.created_at else '',
                '分析更新時間': analysis.updated_at.strftime('%Y-%m-%d %H:%M') if analysis.updated_at else '',
            },
        }
        
        # 渲染 HTML 模板
        html_content = render_to_string('reporting/export_workorder_analysis_detail.html', export_data)
        
        # 設定檔案名稱
        import re
        from urllib.parse import quote
        # 清理檔案名稱，移除或替換可能導致問題的字符
        company_name = re.sub(r'[<>:"/\\|?*\s]', '_', analysis.company_name)
        workorder_id = re.sub(r'[<>:"/\\|?*\s]', '_', analysis.workorder_id)
        product_code = re.sub(r'[<>:"/\\|?*\s]', '_', analysis.product_code)
        raw_filename = f"{company_name}_{workorder_id}_{product_code}.html"
        filename = raw_filename
        
        # 設定回應標頭
        response = HttpResponse(html_content, content_type='text/html; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{quote(filename)}"; filename*=UTF-8\'\'{quote(filename)}'
        
        return response


class WorkOrderAnalysisManagementView(LoginRequiredMixin, TemplateView):
    """工單分析管理頁面"""
    template_name = 'reporting/workorder_analysis_management.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 取得統計資料
        from .models import CompletedWorkOrderAnalysis
        from workorder.models import CompletedWorkOrder
        
        # 計算已分析工單數（唯一工單）
        total_analyzed = CompletedWorkOrderAnalysis.objects.values('workorder_id', 'company_code', 'product_code').distinct().count()
        total_completed = CompletedWorkOrder.objects.count()
        pending_analysis = total_completed - total_analyzed
        
        last_analysis = CompletedWorkOrderAnalysis.objects.order_by('-created_at').first()
        if last_analysis:
            # 轉換為本地時間
            from django.utils import timezone
            local_time = timezone.localtime(last_analysis.created_at)
            last_analysis_date = local_time.strftime('%Y-%m-%d %H:%M')
        else:
            last_analysis_date = '無'
        
        # 取得公司列表
        from erp_integration.models import CompanyConfig
        companies = CompanyConfig.objects.values_list('company_code', 'company_name')
        
        # 取得最近的分析記錄
        recent_analyses = CompletedWorkOrderAnalysis.objects.order_by('-created_at')[:10]
        
        context.update({
            'stats': {
                'total_analyzed': total_analyzed,
                'total_completed': total_completed,
                'pending_analysis': pending_analysis,
                'last_analysis_date': last_analysis_date,
            },
            'companies': companies,
            'recent_analyses': recent_analyses,
        })
        
        return context


@login_required
@require_POST
def analyze_single_workorder(request):
    """分析單一工單"""
    from django.http import JsonResponse
    
    workorder_id = request.POST.get('workorder_id')
    company_code = request.POST.get('company_code')
    force = request.POST.get('force') == 'true'
    
    if not workorder_id or not company_code:
        return JsonResponse({
            'success': False,
            'error': '請提供工單編號和公司代號'
        })
    
    try:
        from .workorder_analysis_service import WorkOrderAnalysisService
        result = WorkOrderAnalysisService.analyze_completed_workorder(
            workorder_id, company_code, force=force
        )
        
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'分析時發生錯誤: {str(e)}'
        })


@login_required
@require_POST
def analyze_batch_workorders(request):
    """批量分析工單"""
    from django.http import JsonResponse
    
    company_code = request.POST.get('company_code')
    start_date = request.POST.get('start_date')
    end_date = request.POST.get('end_date')
    
    try:
        from .workorder_analysis_service import WorkOrderAnalysisService
        result = WorkOrderAnalysisService.analyze_completed_workorders_batch(
            start_date=start_date,
            end_date=end_date,
            company_code=company_code
        )
        
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': f'批量分析時發生錯誤: {str(e)}'
        })


@login_required
@require_POST
def setup_analysis_schedule(request):
    """設定分析定時任務"""
    from django.http import JsonResponse
    import json
    
    try:
        data = json.loads(request.body)
        action = data.get('action')
        interval = data.get('interval', 60)
        
        if action == 'setup':
            # 設定定時任務
            from django_celery_beat.models import PeriodicTask, IntervalSchedule
            
            interval_schedule, created = IntervalSchedule.objects.get_or_create(
                every=interval,
                period=IntervalSchedule.MINUTES,
            )
            
            task, created = PeriodicTask.objects.get_or_create(
                name='auto_analyze_completed_workorders',
                defaults={
                    'task': 'reporting.tasks.auto_analyze_completed_workorders',
                    'interval': interval_schedule,
                    'enabled': True,
                    'description': f'自動分析已完工工單（每{interval}分鐘執行）'
                }
            )
            
            if not created:
                task.interval = interval_schedule
                task.enabled = True  # 確保任務被啟用
                task.description = f'自動分析已完工工單（每{interval}分鐘執行）'
                task.save()
            
            return JsonResponse({
                'success': True,
                'message': f'定時任務設定成功，每{interval}分鐘執行一次'
            })
            
        elif action == 'remove':
            # 移除定時任務
            from django_celery_beat.models import PeriodicTask
            
            try:
                task = PeriodicTask.objects.get(name='auto_analyze_completed_workorders')
                task.delete()
                return JsonResponse({
                    'success': True,
                    'message': '定時任務已移除'
                })
            except PeriodicTask.DoesNotExist:
                return JsonResponse({
                    'success': True,
                    'message': '定時任務不存在'
                })
        
        else:
            return JsonResponse({
                'success': False,
                'message': '無效的操作'
            })
            
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'設定失敗: {str(e)}'
        })


@login_required
def get_analysis_schedule_status(request):
    """取得分析定時任務狀態"""
    from django.http import JsonResponse
    
    try:
        from django_celery_beat.models import PeriodicTask
        
        task = PeriodicTask.objects.filter(name='auto_analyze_completed_workorders').first()
        
        if task:
            from django.utils import timezone
            
            last_run = None
            
            if task.last_run_at:
                local_last_run = timezone.localtime(task.last_run_at)
                last_run = local_last_run.strftime('%Y-%m-%d %H:%M:%S')
            
            return JsonResponse({
                'has_schedule': True,
                'interval': task.interval.every if task.interval else 0,
                'last_run': last_run,
                'next_run': '自動計算',  # PeriodicTask 沒有 next_run_at 屬性
            })
        else:
            return JsonResponse({
                'has_schedule': False
            })
            
    except Exception as e:
        return JsonResponse({
            'has_schedule': False,
            'error': str(e)
        })