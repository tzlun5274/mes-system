"""
匯出工具類別
提供 Excel、CSV、PDF 格式的報表匯出功能
"""

import logging
import csv
import io
from datetime import datetime
from typing import Dict, List, Any
from django.http import HttpResponse
from django.utils import timezone

logger = logging.getLogger(__name__)


class ExportUtils:
    """匯出工具類別"""
    
    def export_report(self, report_data: Dict[str, Any], report_type: str, 
                     export_format: str, date_range: Dict[str, Any]) -> HttpResponse:
        """
        匯出報表
        
        Args:
            report_data: 報表數據
            report_type: 報表類型
            export_format: 匯出格式
            date_range: 日期範圍
        
        Returns:
            HttpResponse 物件
        """
        try:
            if export_format == 'EXCEL':
                return self._export_to_excel(report_data, report_type, date_range)
            elif export_format == 'CSV':
                return self._export_to_csv(report_data, report_type, date_range)
            elif export_format == 'PDF':
                return self._export_to_pdf(report_data, report_type, date_range)
            else:
                raise ValueError(f"不支援的匯出格式: {export_format}")
                
        except Exception as e:
            logger.error(f"匯出報表失敗: {str(e)}")
            raise
    
    def _export_to_excel(self, report_data: Dict[str, Any], report_type: str, 
                        date_range: Dict[str, Any]) -> HttpResponse:
        """匯出為 Excel 格式"""
        try:
            # 嘗試導入 openpyxl
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
            except ImportError:
                logger.error("openpyxl 未安裝，無法匯出 Excel")
                raise ImportError("請安裝 openpyxl: pip install openpyxl")
            
            # 建立工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = f"{report_type}_{datetime.now().strftime('%Y%m%d')}"
            
            # 設定標題樣式
            title_font = Font(bold=True, size=14)
            header_font = Font(bold=True)
            header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 寫入標題
            title = f"{report_data.get('report_type', report_type)} - {date_range.get('start_date')} 至 {date_range.get('end_date')}"
            ws['A1'] = title
            ws['A1'].font = title_font
            ws.merge_cells('A1:H1')
            
            # 寫入生成時間
            ws['A2'] = f"生成時間: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws.merge_cells('A2:H2')
            
            # 根據報表類型設定欄位
            if report_type == 'WORK_REPORT':
                headers = ['作業員', '報工日期', '工單號', '產品編號', '開始時間', '結束時間', '工序', '工作數量', '不良品數量', '工作時長(小時)', '異常紀錄']
                data = report_data.get('report_data', [])
                
            elif report_type == 'WORKORDER_REPORT':
                headers = ['工單號', '產品編號', '產品名稱', '總數量', '完成數量', '不良品數量', '完成率(%)', '開始日期', '結束日期', '狀態', '效率率(%)']
                data = report_data.get('report_data', [])
                
            elif report_type == 'WORK_HOUR_REPORT':
                headers = ['日期', '人員/設備', '總工時', '正常工時', '加班工時', '休息時間', '效率率(%)', '利用率(%)']
                data = report_data.get('report_data', [])
                
            else:
                headers = ['欄位1', '欄位2', '欄位3']
                data = report_data.get('report_data', [])
            
            # 寫入表頭
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # 寫入數據
            for row, record in enumerate(data, 5):
                for col, header in enumerate(headers, 1):
                    value = record.get(header, '')
                    ws.cell(row=row, column=col, value=value)
            
            # 自動調整欄寬
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 建立回應
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # 儲存到回應
            wb.save(response)
            return response
            
        except Exception as e:
            logger.error(f"Excel 匯出失敗: {str(e)}")
            raise
    
    def _export_to_csv(self, report_data: Dict[str, Any], report_type: str, 
                      date_range: Dict[str, Any]) -> HttpResponse:
        """匯出為 CSV 格式"""
        try:
            # 建立回應
            response = HttpResponse(content_type='text/csv; charset=utf-8')
            filename = f"{report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            # 寫入 BOM 以支援中文
            response.write('\ufeff')
            
            # 建立 CSV writer
            writer = csv.writer(response)
            
            # 寫入標題
            title = f"{report_data.get('report_type', report_type)} - {date_range.get('start_date')} 至 {date_range.get('end_date')}"
            writer.writerow([title])
            writer.writerow([f"生成時間: {timezone.now().strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])  # 空行
            
            # 根據報表類型設定欄位
            if report_type == 'WORK_REPORT':
                headers = ['作業員', '報工日期', '工單號', '產品編號', '開始時間', '結束時間', '工序', '工作數量', '不良品數量', '工作時長(小時)', '異常紀錄']
                data = report_data.get('report_data', [])
                
            elif report_type == 'WORKORDER_REPORT':
                headers = ['工單號', '產品編號', '產品名稱', '總數量', '完成數量', '不良品數量', '完成率(%)', '開始日期', '結束日期', '狀態', '效率率(%)']
                data = report_data.get('report_data', [])
                
            elif report_type == 'WORK_HOUR_REPORT':
                headers = ['日期', '人員/設備', '總工時', '正常工時', '加班工時', '休息時間', '效率率(%)', '利用率(%)']
                data = report_data.get('report_data', [])
                
            else:
                headers = ['欄位1', '欄位2', '欄位3']
                data = report_data.get('report_data', [])
            
            # 寫入表頭
            writer.writerow(headers)
            
            # 寫入數據
            for record in data:
                row = []
                for header in headers:
                    value = record.get(header, '')
                    row.append(str(value))
                writer.writerow(row)
            
            return response
            
        except Exception as e:
            logger.error(f"CSV 匯出失敗: {str(e)}")
            raise
    
    def _export_to_pdf(self, report_data: Dict[str, Any], report_type: str, 
                      date_range: Dict[str, Any]) -> HttpResponse:
        """匯出為 PDF 格式（暫時重定向到 Excel）"""
        try:
            # 暫時重定向到 Excel 匯出
            logger.info("PDF 匯出功能正在開發中，暫時使用 Excel 格式")
            
            # 建立回應
            response = HttpResponse(content_type='text/html; charset=utf-8')
            response.write(f"""
            <html>
            <head>
                <title>PDF 匯出功能開發中</title>
                <meta charset="utf-8">
            </head>
            <body>
                <h2>PDF 匯出功能正在開發中</h2>
                <p>報表類型: {report_type}</p>
                <p>日期範圍: {date_range.get('start_date')} 至 {date_range.get('end_date')}</p>
                <p>請暫時使用 Excel 或 CSV 格式匯出報表。</p>
                <p><a href="javascript:history.back()">返回上一頁</a></p>
            </body>
            </html>
            """)
            
            return response
            
        except Exception as e:
            logger.error(f"PDF 匯出失敗: {str(e)}")
            raise 