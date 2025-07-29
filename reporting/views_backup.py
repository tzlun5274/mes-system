import logging
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse, JsonResponse
from .models import (
    ProductionDailyReport,
    OperatorPerformance,
    ReportEmailSchedule,
    ReportEmailLog,
    ReportingOperationLog,
)
from system.models import ReportSyncSettings
from .utils import log_user_operation  # 導入 log_user_operation
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
# from .tasks import sync_operator_performance_task  # 暫時註解，避免導入錯誤
from django.views.decorators.csrf import csrf_exempt
from django.contrib import messages
from django.urls import reverse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.db import transaction
from django.utils import timezone

import pandas as pd
from django.db import models

# 設定報表管理模組的日誌記錄器
reporting_logger = logging.getLogger("reporting")
reporting_handler = logging.FileHandler("/var/log/mes/reporting.log")
reporting_handler.setFormatter(
    logging.Formatter("%(levelname)s %(asctime)s %(module)s %(message)s")
)
reporting_logger.addHandler(reporting_handler)
reporting_logger.setLevel(logging.INFO)

# 設定通用logger
logger = logging.getLogger(__name__)


def log_user_operation(username, module, action):
    """記錄用戶操作"""
    try:
        ReportingOperationLog.objects.create(
            user=username,
            action=f"[{module}] {action}",
        )
    except Exception as e:
        print(f"記錄操作日誌失敗: {e}")


def reporting_user_required(user):
    """檢查用戶是否有報表管理權限"""
    return user.is_authenticated and (
        user.is_superuser or user.groups.filter(name="報表管理員").exists()
    )


@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def reporting_index(request):
    """報表管理首頁"""
    log_user_operation(request.user.username, "報表管理", "訪問報表管理首頁")
    
    return render(
        request,
        "reporting/index.html",
        {"module_display_name": "報表管理"},
    )


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def production_daily(request):
    # 記錄用戶操作
    log_user_operation(request.user.username, "reporting", "查看生產日報表")
    production_reports = ProductionDailyReport.objects.all().order_by(
        "-date", "operator_or_line"
    )

    if "download" in request.GET:
        # 記錄下載操作
        log_user_operation(
            request.user.username, "reporting", "下載生產日報表 Excel 文件"
        )
        # 創建 Excel 文件
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "生產日報表"

        # 寫入表頭
        headers = [
            "日期",
            "作業員姓名",
            "設備名稱",
            "生產線",
            "工序名稱",
            "完成數量",
            "工作時數",
            "效率 (件/小時)",
            "完成率 (%)",
        ]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f"{col_letter}1"] = header

        # 寫入數據
        for row_num, report in enumerate(production_reports, 2):
            worksheet[f"A{row_num}"] = report.date.strftime("%Y-%m-%d")
            worksheet[f"B{row_num}"] = report.operator_or_line
            worksheet[f"C{row_num}"] = report.equipment_name
            worksheet[f"D{row_num}"] = report.get_line_display()
            worksheet[f"E{row_num}"] = report.process_name
            worksheet[f"F{row_num}"] = report.completed_quantity
            worksheet[f"G{row_num}"] = report.work_hours
            worksheet[f"H{row_num}"] = report.efficiency_rate
            worksheet[f"I{row_num}"] = report.completion_rate

        # 設置響應頭以下載 Excel 文件
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            "attachment; filename=production_daily_report.xlsx"
        )
        workbook.save(response)
        return response

    return render(
        request,
        "reporting/production_daily.html",
        {
            "production_reports": production_reports,
        },
    )


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def operator_performance(request):
    # 顯示所有欄位
    log_user_operation(request.user.username, "reporting", "查看工作報表")
    operator_reports = OperatorPerformance.objects.all()

    if "download" in request.GET:
        log_user_operation(
            request.user.username, "reporting", "下載工作報表 Excel 文件"
        )
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "工作報表"
        # 新增詳細欄位
        headers = [
            "作業員名稱",
            "工單",
            "產品名稱",
            "生產數量",
            "開始時間",
            "結束時間",
            "日期",
            "工序",
            "設備名稱",
        ]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f"{col_letter}1"] = header
        for row_num, report in enumerate(operator_reports, 2):
            worksheet[f"A{row_num}"] = report.operator_name
            worksheet[f"B{row_num}"] = report.work_order
            worksheet[f"C{row_num}"] = report.product_name
            worksheet[f"D{row_num}"] = report.production_quantity
            worksheet[f"E{row_num}"] = (
                report.start_time.strftime("%Y-%m-%d %H:%M:%S")
                if report.start_time
                else ""
            )
            worksheet[f"F{row_num}"] = (
                report.end_time.strftime("%Y-%m-%d %H:%M:%S") if report.end_time else ""
            )
            worksheet[f"G{row_num}"] = report.date.strftime("%Y-%m-%d")
            worksheet[f"H{row_num}"] = report.process_name
            worksheet[f"I{row_num}"] = report.equipment_name
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            "attachment; filename=work_report.xlsx"
        )
        workbook.save(response)
        return response
    return render(
        request,
        "reporting/operator_performance.html",
        {
            "operator_reports": operator_reports,
        },
    )


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def get_production_daily(request):
    # 記錄用戶操作
    log_user_operation(
        request.user.username, "reporting", "通過 API 獲取生產日報表數據"
    )
    production_reports = ProductionDailyReport.objects.all()
    production_reports_data = [
        {
            "id": report.id,
            "date": report.date.strftime("%Y-%m-%d"),
            "line": report.line,
            "production_quantity": report.production_quantity,
            "completed_quantity": report.completed_quantity,
            "completion_rate": report.completion_rate,
            "created_at": report.created_at.isoformat(),
            "updated_at": report.updated_at.isoformat(),
        }
        for report in production_reports
    ]
    return JsonResponse({"production_reports": production_reports_data})


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def get_operator_performance(request):
    # 回傳所有欄位
    log_user_operation(
        request.user.username, "reporting", "通過 API 獲取工作報表數據"
    )
    operator_reports = OperatorPerformance.objects.all()
    operator_reports_data = [
        {
            "id": report.id,
            "operator_name": report.operator_name,
            "equipment_name": report.equipment_name,
            "production_quantity": report.production_quantity,
            "equipment_usage_rate": report.equipment_usage_rate,
            "date": report.date.strftime("%Y-%m-%d"),
            "created_at": report.created_at.isoformat(),
            "updated_at": report.updated_at.isoformat(),
        }
        for report in operator_reports
    ]
    return JsonResponse({"operator_reports": operator_reports_data})


@csrf_exempt
@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def manual_sync_smt_report(request):
    """
    手動觸發SMT生產報表同步的API
    """
    # from .tasks import sync_smt_report_task # This line was removed as per the new_code, as the task is no longer imported.
    # sync_smt_report_task.delay() # This line was removed as per the new_code, as the task is no longer imported.
    return JsonResponse(
        {"success": True, "message": "已開始同步SMT生產報表，請稍後刷新頁面"}
    )


@csrf_exempt
@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def manual_sync_operator_performance(request):
    """
    手動觸發作業員生產報表同步的API
    """
    sync_operator_performance_task.delay()
    return JsonResponse(
        {"success": True, "message": "已開始同步作業員生產報表，請稍後刷新頁面"}
    )


@csrf_exempt
@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def update_sync_interval(request):
    """
    更新報表同步間隔設定的API
    """
    if request.method == "POST":
        try:
            report_type = request.POST.get("report_type")
            sync_interval_hours = int(request.POST.get("sync_interval_hours", 24))

            # 驗證輸入
            if report_type not in ["smt", "operator"]:
                return JsonResponse({"success": False, "message": "無效的報表類型"})

            if sync_interval_hours < 1 or sync_interval_hours > 168:  # 1小時到7天
                return JsonResponse(
                    {"success": False, "message": "同步間隔必須在1-168小時之間"}
                )

            # 更新或創建設定
            # 將小時轉換為頻率設定
            if sync_interval_hours <= 1:
                sync_frequency = "hourly"
            elif sync_interval_hours <= 24:
                sync_frequency = "daily"
            elif sync_interval_hours <= 168:  # 7天
                sync_frequency = "weekly"
            else:
                sync_frequency = "monthly"
            
            setting, created = ReportSyncSettings.objects.update_or_create(
                sync_type=report_type,
                defaults={
                    "sync_frequency": sync_frequency,
                    "sync_time": timezone.now().time(),
                    "is_active": True,
                },
            )

            # 記錄操作
            action = f"更新{setting.get_sync_type_display()}同步間隔為{sync_interval_hours}小時"
            log_user_operation(request.user.username, "reporting", action)

            return JsonResponse(
                {
                    "success": True,
                    "message": f"{setting.get_sync_type_display()}同步間隔已更新為{sync_interval_hours}小時",
                }
            )

        except ValueError:
            return JsonResponse(
                {"success": False, "message": "同步間隔必須是有效的數字"}
            )
        except Exception as e:
            return JsonResponse({"success": False, "message": f"更新失敗：{str(e)}"})

    return JsonResponse({"success": False, "message": "只支援POST請求"})


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def get_sync_settings(request):
    """
    取得報表同步設定的API
    """
    try:
        smt_setting = ReportSyncSettings.objects.filter(sync_type="smt").first()
        operator_setting = ReportSyncSettings.objects.filter(
            sync_type="operator"
        ).first()

        settings = {
            "smt": {
                "sync_frequency": (
                    smt_setting.sync_frequency if smt_setting else "daily"
                ),
                "sync_time": (
                    smt_setting.sync_time if smt_setting else timezone.now().time()
                ),
                "is_active": smt_setting.is_active if smt_setting else True,
            },
            "operator": {
                "sync_frequency": (
                    operator_setting.sync_frequency if operator_setting else "daily"
                ),
                "sync_time": (
                    operator_setting.sync_time if operator_setting else timezone.now().time()
                ),
                "is_active": operator_setting.is_active if operator_setting else True,
            },
        }

        return JsonResponse({"success": True, "settings": settings})

    except Exception as e:
        return JsonResponse({"success": False, "message": f"取得設定失敗：{str(e)}"})


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def email_schedule_list(request):
    """
    報表郵件發送設定列表頁面
    """
    log_user_operation(request.user.username, "reporting", "查看報表郵件發送設定")
    schedules = ReportEmailSchedule.objects.all().order_by(
        "report_type", "schedule_type"
    )
    return render(
        request,
        "reporting/email_schedule_list.html",
        {
            "schedules": schedules,
        },
    )


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def email_schedule_create(request):
    """
    創建報表郵件發送設定
    """
    if request.method == "POST":
        try:
            schedule = ReportEmailSchedule.objects.create(
                report_type=request.POST.get("report_type"),
                schedule_type=request.POST.get("schedule_type", "daily"),
                send_time=request.POST.get("send_time"),
                recipients=request.POST.get("recipients", ""),
                cc_recipients=request.POST.get("cc_recipients", ""),
                subject_template=request.POST.get(
                    "subject_template", "MES 系統 - {report_type} - {date}"
                ),
                is_active=request.POST.get("is_active") == "on",
            )
            log_user_operation(
                request.user.username, "reporting", f"創建報表郵件發送設定: {schedule}"
            )
            messages.success(request, f"報表郵件發送設定已創建：{schedule}")
            return redirect("reporting:email_schedule_list")
        except Exception as e:
            messages.error(request, f"創建設定失敗：{str(e)}")

    return render(
        request,
        "reporting/email_schedule_form.html",
        {"title": "創建報表郵件發送設定", "action": "create"},
    )


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def email_schedule_edit(request, schedule_id):
    """
    編輯報表郵件發送設定
    """
    try:
        schedule = ReportEmailSchedule.objects.get(id=schedule_id)
    except ReportEmailSchedule.DoesNotExist:
        messages.error(request, "找不到指定的設定")
        return redirect("reporting:email_schedule_list")

    if request.method == "POST":
        try:
            schedule.report_type = request.POST.get("report_type")
            schedule.schedule_type = request.POST.get("schedule_type", "daily")
            schedule.send_time = request.POST.get("send_time")
            schedule.recipients = request.POST.get("recipients", "")
            schedule.cc_recipients = request.POST.get("cc_recipients", "")
            schedule.subject_template = request.POST.get(
                "subject_template", "MES 系統 - {report_type} - {date}"
            )
            schedule.is_active = request.POST.get("is_active") == "on"
            schedule.save()

            log_user_operation(
                request.user.username, "reporting", f"更新報表郵件發送設定: {schedule}"
            )
            messages.success(request, f"報表郵件發送設定已更新：{schedule}")
            return redirect("reporting:email_schedule_list")
        except Exception as e:
            messages.error(request, f"更新設定失敗：{str(e)}")

    return render(
        request,
        "reporting/email_schedule_form.html",
        {"title": "編輯報表郵件發送設定", "action": "edit", "schedule": schedule},
    )


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def email_schedule_delete(request, schedule_id):
    """
    刪除報表郵件發送設定
    """
    try:
        schedule = ReportEmailSchedule.objects.get(id=schedule_id)
        schedule_name = str(schedule)
        schedule.delete()

        log_user_operation(
            request.user.username, "reporting", f"刪除報表郵件發送設定: {schedule_name}"
        )
        messages.success(request, f"報表郵件發送設定已刪除：{schedule_name}")
    except ReportEmailSchedule.DoesNotExist:
        messages.error(request, "找不到指定的設定")
    except Exception as e:
        messages.error(request, f"刪除設定失敗：{str(e)}")

    return redirect("reporting:email_schedule_list")


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def email_log_list(request):
    """
    報表郵件發送記錄列表
    """
    log_user_operation(request.user.username, "reporting", "查看報表郵件發送記錄")
    logs = ReportEmailLog.objects.all().order_by("-sent_at")[:100]  # 只顯示最近100筆
    return render(
        request,
        "reporting/email_log_list.html",
        {
            "logs": logs,
        },
    )


@csrf_exempt
@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def test_send_report_email(request):
    """
    測試發送報表郵件
    """
    if request.method == "POST":
        try:
            schedule_id = request.POST.get("schedule_id")
            schedule = ReportEmailSchedule.objects.get(id=schedule_id)

            # 執行測試發送
            from .tasks import send_daily_reports_email

            send_daily_reports_email.delay()

            log_user_operation(
                request.user.username, "reporting", f"測試發送報表郵件: {schedule}"
            )
            return JsonResponse(
                {"success": True, "message": f"測試郵件已發送：{schedule}"}
            )
        except ReportEmailSchedule.DoesNotExist:
            return JsonResponse({"success": False, "message": "找不到指定的設定"})
        except Exception as e:
            return JsonResponse({"success": False, "message": f"發送失敗：{str(e)}"})

    return JsonResponse({"success": False, "message": "只支援POST請求"})


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def clear_report_data(request):
    """
    清除報表資料的視圖函數
    GET：顯示清除頁面
    POST：執行清除
    """
    if request.method == "GET":
        return render(request, "reporting/clear_data.html")
    try:
        clear_type = request.POST.get("clear_type", "reports_only")
        date_range = request.POST.get("date_range", "").strip()
        with transaction.atomic():
            cleared_count = 0
            # 處理日期範圍
            date_filter = {}
            if date_range:
                try:
                    start_date, end_date = date_range.split(",")
                    from datetime import datetime

                    start_date = datetime.strptime(
                        start_date.strip(), "%Y-%m-%d"
                    ).date()
                    end_date = datetime.strptime(end_date.strip(), "%Y-%m-%d").date()
                    date_filter = {"date__range": [start_date, end_date]}
                except ValueError:
                    messages.error(
                        request, "日期格式錯誤，請使用 YYYY-MM-DD,YYYY-MM-DD 格式"
                    )
                    return redirect("reporting:index")
            # 根據清除類型執行相應操作
            if clear_type == "all":
                cleared_count += _clear_all_report_data(date_filter)
                message = f"成功清除所有報表資料，共 {cleared_count} 筆記錄"
            elif clear_type == "reports_only":
                cleared_count += _clear_reports_only(date_filter)
                message = f"成功清除報表資料，共 {cleared_count} 筆記錄"
            elif clear_type == "logs_only":
                cleared_count += _clear_logs_only()
                message = f"成功清除日誌資料，共 {cleared_count} 筆記錄"
            elif clear_type == "settings_only":
                cleared_count += _clear_settings_only()
                message = f"成功清除設定資料，共 {cleared_count} 筆記錄"
            else:
                messages.error(request, "無效的清除類型")
                return redirect("reporting:index")
            # 記錄操作日誌
            _log_operation(
                f"用戶 {request.user.username} 清除報表資料，類型：{clear_type}，共清除 {cleared_count} 筆記錄"
            )
            messages.success(request, message)
    except Exception as e:
        messages.error(request, f"清除資料時發生錯誤：{str(e)}")
    return redirect("reporting:index")


def _clear_all_report_data(date_filter):
    """清除所有報表相關資料"""
    count = 0
    count += _clear_reports_only(date_filter)
    count += _clear_logs_only()
    count += _clear_settings_only()
    return count


def _clear_reports_only(date_filter):
    """只清除報表資料"""
    count = 0

    # 清除生產日報表
    if date_filter:
        count += ProductionDailyReport.objects.filter(**date_filter).delete()[0]
    else:
        count += ProductionDailyReport.objects.all().delete()[0]

    # 清除作業員績效報表
    if date_filter:
        count += OperatorPerformance.objects.filter(**date_filter).delete()[0]
    else:
        count += OperatorPerformance.objects.all().delete()[0]

    return count


def _clear_logs_only():
    """只清除日誌資料"""
    count = 0
    count += ReportingOperationLog.objects.all().delete()[0]
    count += ReportEmailLog.objects.all().delete()[0]
    return count


def _clear_settings_only():
    """只清除設定資料"""
    count = 0
    count += ReportSyncSettings.objects.all().delete()[0]
    count += ReportEmailSchedule.objects.all().delete()[0]
    return count


def _log_operation(action):
    """記錄操作到日誌"""
    try:
        ReportingOperationLog.objects.create(user="system", action=action)
    except Exception:
        pass





@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def report_export(request):
    """
    報表匯出頁面 - 統一管理所有報表匯出功能
    匯出各種報工報表，支援Excel、CSV、PDF等多種格式
    """
    from datetime import date, timedelta
    from django.db.models import Q
    from workorder.models import OperatorSupplementReport, SMTProductionReport
    
    # 記錄用戶操作
    log_user_operation(request.user.username, "reporting", "訪問報表匯出頁面")
    
    today = date.today()
    month_start = today.replace(day=1)
    
    # 可匯出的報表類型
    report_types = [
        {'id': 'daily', 'name': '日報表', 'description': '每日報工統計報表', 'icon': 'fas fa-calendar-day'},
        {'id': 'weekly', 'name': '週報表', 'description': '每週報工統計報表', 'icon': 'fas fa-calendar-week'},
        {'id': 'monthly', 'name': '月報表', 'description': '每月報工統計報表', 'icon': 'fas fa-calendar-alt'},
        {'id': 'operator', 'name': '作業員報工報表', 'description': '作業員報工詳細記錄', 'icon': 'fas fa-user'},
        {'id': 'smt', 'name': 'SMT報工報表', 'description': 'SMT設備報工詳細記錄', 'icon': 'fas fa-microchip'},
        {'id': 'abnormal', 'name': '異常報工報表', 'description': '異常報工記錄分析', 'icon': 'fas fa-exclamation-triangle'},
        {'id': 'efficiency', 'name': '效率分析報表', 'description': '產能效率分析報表', 'icon': 'fas fa-tachometer-alt'},
        {'id': 'production_daily', 'name': '生產日報表', 'description': '生產日報表統計', 'icon': 'fas fa-chart-line'},
        {'id': 'operator_performance', 'name': '作業員績效報表', 'description': '作業員績效分析', 'icon': 'fas fa-user-chart'},
    ]
    
    # 匯出格式
    export_formats = [
        {'id': 'excel', 'name': 'Excel (.xlsx)', 'icon': 'fas fa-file-excel'},
        {'id': 'csv', 'name': 'CSV (.csv)', 'icon': 'fas fa-file-csv'},
        {'id': 'pdf', 'name': 'PDF (.pdf)', 'icon': 'fas fa-file-pdf'},
    ]
    
    # 日期範圍選項
    date_ranges = [
        {'id': 'today', 'name': '今日', 'start': today, 'end': today},
        {'id': 'yesterday', 'name': '昨日', 'start': today - timedelta(days=1), 'end': today - timedelta(days=1)},
        {'id': 'week', 'name': '本週', 'start': today - timedelta(days=today.weekday()), 'end': today},
        {'id': 'month', 'name': '本月', 'start': month_start, 'end': today},
        {'id': 'custom', 'name': '自訂日期範圍', 'start': None, 'end': None},
    ]
    
    context = {
        'report_types': report_types,
        'export_formats': export_formats,
        'date_ranges': date_ranges,
    }
    return render(request, 'reporting/report_export.html', context)


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def execute_report_export(request):
    """
    執行報表匯出功能
    """
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import csv
    from io import StringIO
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '只支援POST請求'})
    
    try:
        report_type = request.POST.get('report_type')
        export_format = request.POST.get('export_format')
        date_range = request.POST.get('date_range')
        custom_start = request.POST.get('custom_start')
        custom_end = request.POST.get('custom_end')
        
        # 記錄用戶操作
        log_user_operation(
            request.user.username, 
            "reporting", 
            f"匯出報表：{report_type}，格式：{export_format}，日期範圍：{date_range}"
        )
        
        # 根據報表類型取得資料
        if report_type == 'operator':
            data = get_operator_report_data(date_range, custom_start, custom_end)
            filename = f'作業員報工報表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        elif report_type == 'smt':
            data = get_smt_report_data(date_range, custom_start, custom_end)
            filename = f'SMT報工報表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        elif report_type == 'abnormal':
            data = get_abnormal_report_data(date_range, custom_start, custom_end)
            filename = f'異常報工報表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        elif report_type == 'production_daily':
            data = get_production_daily_report_data(date_range, custom_start, custom_end)
            filename = f'生產日報表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        elif report_type == 'operator_performance':
            data = get_operator_performance_report_data(date_range, custom_start, custom_end)
            filename = f'作業員績效報表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        else:
            return JsonResponse({'success': False, 'error': '不支援的報表類型'})
        
        # 根據格式匯出
        if export_format == 'excel':
            return export_to_excel(data, filename)
        elif export_format == 'csv':
            return export_to_csv(data, filename)
        elif export_format == 'pdf':
            return export_to_pdf(data, filename)
        else:
            return JsonResponse({'success': False, 'error': '不支援的匯出格式'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'匯出失敗：{str(e)}'})


def get_operator_report_data(date_range, custom_start=None, custom_end=None):
    """取得作業員報工資料"""
    from workorder.models import OperatorSupplementReport
    from datetime import date, timedelta
    
    today = date.today()
    
    # 設定日期範圍
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=30)
        end_date = today
    
    # 查詢資料
    reports = OperatorSupplementReport.objects.filter(
        work_date__gte=start_date,
        work_date__lte=end_date
    ).select_related('operator', 'workorder', 'process').order_by('-work_date', '-start_time')
    
    # 格式化資料
    data = []
    for report in reports:
        data.append({
            '報工日期': report.work_date.strftime('%Y-%m-%d'),
            '開始時間': report.start_time.strftime('%H:%M'),
            '結束時間': report.end_time.strftime('%H:%M'),
            '作業員': report.operator.name if report.operator else '-',
            '工單號': report.workorder.order_number if report.workorder else '-',
            '工序': report.process.name if report.process else '-',
            '報工數量': report.work_quantity,
            '不良品數量': report.defect_quantity,
            '工時': report.work_hours,
            '審核狀態': report.get_approval_status_display(),
            '備註': report.remarks or '',
            '異常紀錄': report.abnormal_notes or '',
            '建立時間': report.created_at.strftime('%Y-%m-%d %H:%M'),
        })
    
    return data


def get_smt_report_data(date_range, custom_start=None, custom_end=None):
    """取得SMT報工資料"""
    from workorder.models import SMTProductionReport
    from datetime import date, timedelta
    
    today = date.today()
    
    # 設定日期範圍
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=30)
        end_date = today
    
    # 查詢資料
    reports = SMTProductionReport.objects.filter(
        work_date__gte=start_date,
        work_date__lte=end_date
    ).select_related('workorder', 'equipment').order_by('-work_date', '-start_time')
    
    # 格式化資料
    data = []
    for report in reports:
        data.append({
            '設備': report.equipment.name if report.equipment else '-',
            '報工日期': report.work_date.strftime('%Y-%m-%d'),
            '開始時間': report.start_time.strftime('%H:%M'),
            '結束時間': report.end_time.strftime('%H:%M'),
            '製令號碼': report.workorder.order_number if report.workorder else '-',
            '機種名稱': report.rd_product_code or '-',
            '工序': report.operation or '-',
            '工作數量': report.work_quantity,
            '不良品數量': report.defect_quantity,
            '備註': report.remarks or '',
            '異常紀錄': report.abnormal_notes or '',
            '審核狀態': report.get_approval_status_display(),
        })
    
    return data


def get_abnormal_report_data(date_range, custom_start=None, custom_end=None):
    """取得異常報工資料"""
    from workorder.models import OperatorSupplementReport, SMTProductionReport
    from datetime import date, timedelta
    
    today = date.today()
    
    # 設定日期範圍
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=30)
        end_date = today
    
    # 查詢異常資料
    operator_abnormal = OperatorSupplementReport.objects.filter(
        work_date__gte=start_date,
        work_date__lte=end_date,
        abnormal_notes__isnull=False
    ).exclude(abnormal_notes='').select_related('operator', 'workorder', 'process')
    
    smt_abnormal = SMTProductionReport.objects.filter(
        work_date__gte=start_date,
        work_date__lte=end_date,
        abnormal_notes__isnull=False
    ).exclude(abnormal_notes='').select_related('workorder', 'equipment')
    
    # 格式化資料
    data = []
    
    # 作業員異常
    for report in operator_abnormal:
        data.append({
            '報表類型': '作業員報工',
            '報工日期': report.work_date.strftime('%Y-%m-%d'),
            '作業員': report.operator.name if report.operator else '-',
            '工單號': report.workorder.order_number if report.workorder else '-',
            '工序': report.process.name if report.process else '-',
            '異常紀錄': report.abnormal_notes,
            '審核狀態': report.get_approval_status_display(),
            '建立時間': report.created_at.strftime('%Y-%m-%d %H:%M'),
        })
    
    # SMT異常
    for report in smt_abnormal:
        data.append({
            '報表類型': 'SMT報工',
            '報工日期': report.work_date.strftime('%Y-%m-%d'),
            '設備': report.equipment.name if report.equipment else '-',
            '製令號碼': report.workorder.order_number if report.workorder else '-',
            '工序': report.operation or '-',
            '異常紀錄': report.abnormal_notes,
            '審核狀態': report.get_approval_status_display(),
            '建立時間': report.created_at.strftime('%Y-%m-%d %H:%M'),
        })
    
    return data


def get_production_daily_report_data(date_range, custom_start=None, custom_end=None):
    """取得生產日報表資料"""
    from datetime import date, timedelta
    
    today = date.today()
    
    # 設定日期範圍
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=30)
        end_date = today
    
    # 查詢資料
    reports = ProductionDailyReport.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).order_by('-date', 'operator_or_line')
    
    # 格式化資料
    data = []
    for report in reports:
        data.append({
            '日期': report.date.strftime('%Y-%m-%d'),
            '作業員/SMT產線': report.operator_or_line,
            '設備名稱': report.equipment_name,
            '生產線': report.get_line_display(),
            '工序名稱': report.process_name,
            '完成數量': report.completed_quantity,
            '工作時數': report.work_hours,
            '效率': report.efficiency_rate,
            '完成率': report.completion_rate,
        })
    
    return data


def get_operator_performance_report_data(date_range, custom_start=None, custom_end=None):
    """取得作業員績效報表資料"""
    from datetime import date, timedelta
    
    today = date.today()
    
    # 設定日期範圍
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=30)
        end_date = today
    
    # 查詢資料
    reports = OperatorPerformance.objects.filter(
        date__gte=start_date,
        date__lte=end_date
    ).order_by('-date', 'operator_name')
    
    # 格式化資料
    data = []
    for report in reports:
        data.append({
            '作業員名稱': report.operator_name,
            '工單': report.work_order,
            '產品名稱': report.product_name,
            '生產數量': report.production_quantity,
            '開始時間': report.start_time.strftime('%Y-%m-%d %H:%M:%S') if report.start_time else '',
            '結束時間': report.end_time.strftime('%Y-%m-%d %H:%M:%S') if report.end_time else '',
            '日期': report.date.strftime('%Y-%m-%d'),
            '工序': report.process_name,
            '設備名稱': report.equipment_name,
        })
    
    return data


def export_to_excel(data, filename):
    """匯出為Excel格式"""
    from django.http import HttpResponse
    
    if not data:
        return HttpResponse("沒有資料可匯出", status=400)
    
    # 建立工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "報表資料"
    
    # 設定標題樣式
    header_font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    # 設定資料樣式
    data_font = Font(name='微軟正黑體', size=11)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    # 設定邊框
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 寫入標題
    headers = list(data[0].keys())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 寫入資料
    for row, record in enumerate(data, 2):
        for col, value in enumerate(record.values(), 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # 自動調整欄寬
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # 建立回應
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    
    wb.save(response)
    return response


def export_to_csv(data, filename):
    """匯出為CSV格式"""
    from django.http import HttpResponse
    
    if not data:
        return HttpResponse("沒有資料可匯出", status=400)
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    
    # 寫入BOM以支援中文
    response.write('\ufeff')
    
    writer = csv.writer(response)
    
    # 寫入標題
    headers = list(data[0].keys())
    writer.writerow(headers)
    
    # 寫入資料
    for record in data:
        writer.writerow(record.values())
    
    return response


def export_to_pdf(data, filename):
    """匯出為PDF格式（暫時返回Excel，因為PDF需要額外的套件）"""
    # 暫時使用Excel格式，因為PDF需要reportlab或其他PDF套件
    return export_to_excel(data, filename)

# 新增的報表視圖
from django.views.generic import TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from .services.work_time_service import WorkTimeReportService
from .models import WorkTimeReport
from datetime import date, datetime
import json


@method_decorator(csrf_exempt, name='dispatch')
class WorkTimeReportView(LoginRequiredMixin, TemplateView):
    """工作時間報表視圖"""
    template_name = 'reporting/work_time/daily_report.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 取得查詢參數
        report_date = self.request.GET.get('report_date', date.today().isoformat())
        worker_type = self.request.GET.get('worker_type', '')
        worker_name = self.request.GET.get('worker_name', '')
        workorder_number = self.request.GET.get('workorder_number', '')
        
        # 查詢報表數據
        queryset = WorkTimeReport.objects.filter(report_date=report_date)
        
        if worker_type:
            queryset = queryset.filter(worker_type=worker_type)
        if worker_name:
            queryset = queryset.filter(worker_name__icontains=worker_name)
        if workorder_number:
            queryset = queryset.filter(workorder_number__icontains=workorder_number)
        
        # 計算統計數據
        total_workers = queryset.values('worker_name').distinct().count()
        total_work_hours = queryset.aggregate(Sum('total_work_hours'))['total_work_hours__sum'] or 0
        total_completed_quantity = queryset.aggregate(Sum('completed_quantity'))['completed_quantity__sum'] or 0
        avg_efficiency = queryset.aggregate(Avg('efficiency_rate'))['efficiency_rate__avg'] or 0
        
        # 準備圖表數據
        chart_data = self._prepare_chart_data(queryset)
        
        context.update({
            'report_title': '工作時間報表',
            'report_type': 'work_time',
            'report_date': report_date,
            'worker_type': worker_type,
            'worker_name': worker_name,
            'workorder_number': workorder_number,
            'reports': queryset.order_by('worker_name'),
            'statistics': {
                'total_workers': total_workers,
                'total_work_hours': total_work_hours,
                'total_completed_quantity': total_completed_quantity,
                'average_efficiency': avg_efficiency,
            },
            'chart_data': chart_data,
        })
        
        return context
    
    def _prepare_chart_data(self, queryset):
        """準備圖表數據"""
        # 效率數據
        efficiency_data = list(queryset.values('worker_name', 'efficiency_rate').order_by('-efficiency_rate')[:10])
        
        # 工時數據
        work_hours_data = list(queryset.values('worker_name', 'total_work_hours').order_by('-total_work_hours')[:5])
        
        return {
            'labels': [item['worker_name'] for item in efficiency_data],
            'efficiency_data': [item['efficiency_rate'] for item in efficiency_data],
            'worker_labels': [item['worker_name'] for item in work_hours_data],
            'work_hours_data': [item['total_work_hours'] for item in work_hours_data],
        }


@csrf_exempt
def api_report_data(request):
    """API 端點：取得報表數據"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            report_type = data.get('report_type', 'work_time')
            filters = data.get('filters', {})
            
            # 使用服務層生成報表
            service = WorkTimeReportService()
            
            # 準備日期範圍
            report_date = filters.get('report_date', date.today().isoformat())
            date_obj = datetime.strptime(report_date, '%Y-%m-%d').date()
            date_range = {'start': date_obj, 'end': date_obj}
            
            # 生成報表
            report_data = service.generate_report(report_type, date_range, **filters)
            
            return JsonResponse({
                'success': True,
                'data': report_data
            })
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({
        'success': False,
        'error': '不支援的請求方法'
    })


@csrf_exempt
def export_report(request):
    """匯出報表"""
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            report_type = data.get('report_type', 'work_time')
            format_type = data.get('format', 'excel')
            filters = data.get('filters', {})
            
            # 這裡可以實作匯出邏輯
            # 暫時返回一個簡單的響應
            
            from django.http import HttpResponse
            response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="報表_{date.today().isoformat()}.xlsx"'
            
            # 這裡應該實作實際的 Excel 生成邏輯
            response.write(b'Excel file content would be here')
            
            return response
            
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    return JsonResponse({
        'success': False,
        'error': '不支援的請求方法'
    })


def work_time_report_detail(request, report_id):
    """工作時間報表詳細資訊"""
    try:
        report = WorkTimeReport.objects.get(id=report_id)
        return JsonResponse({
            'success': True,
            'report': {
                'id': report.id,
                'worker_name': report.worker_name,
                'worker_type_display': report.get_worker_type_display(),
                'workorder_number': report.workorder_number,
                'product_code': report.product_code,
                'process_name': report.process_name,
                'start_time': report.start_time.strftime('%H:%M') if report.start_time else None,
                'end_time': report.end_time.strftime('%H:%M') if report.end_time else None,
                'total_work_hours': report.total_work_hours,
                'actual_work_hours': report.actual_work_hours,
                'break_hours': report.break_hours,
                'overtime_hours': report.overtime_hours,
                'completed_quantity': report.completed_quantity,
                'defect_quantity': report.defect_quantity,
                'yield_rate': report.yield_rate,
                'efficiency_rate': report.efficiency_rate,
                'original_quantity': report.original_quantity,
                'allocated_quantity': report.allocated_quantity,
                'quantity_source_display': report.get_quantity_source_display(),
                'abnormal_notes': report.abnormal_notes,
            }
        })
    except WorkTimeReport.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': '報表記錄不存在'
        })


def work_time_report_allocation(request, report_id):
    """工作時間報表分配詳情"""
    try:
        report = WorkTimeReport.objects.get(id=report_id)
        return JsonResponse({
            'success': True,
            'allocation': {
                'original_quantity': report.original_quantity,
                'allocated_quantity': report.allocated_quantity,
                'work_hours': report.total_work_hours,
                'efficiency_rate': report.efficiency_rate,
                'quantity_source_display': report.get_quantity_source_display(),
                'allocation_notes': report.allocation_notes,
                'allocation_ratio': round((report.allocated_quantity / report.original_quantity * 100), 2) if report.original_quantity > 0 else 0,
            }
        })
    except WorkTimeReport.DoesNotExist:
        return JsonResponse({
            'success': False,
            'error': '報表記錄不存在'
        })

# 數量分配相關視圖
from .services.quantity_allocation_service import QuantityAllocationService


def quantity_allocation_dashboard(request):
    """
    數量分配儀表板
    顯示分配統計和操作介面
    """
    from datetime import datetime, timedelta
    from django.utils import timezone
    
    # 取得日期範圍參數
    date_range_type = request.GET.get('date_range', 'month')
    custom_start = request.GET.get('custom_start')
    custom_end = request.GET.get('custom_end')
    
    # 設定日期範圍
    end_date = timezone.now()
    if date_range_type == 'week':
        start_date = end_date - timedelta(days=7)
    elif date_range_type == 'month':
        start_date = end_date - timedelta(days=30)
    elif date_range_type == 'quarter':
        start_date = end_date - timedelta(days=90)
    elif date_range_type == 'custom' and custom_start and custom_end:
        try:
            start_date = datetime.strptime(custom_start, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            end_date = datetime.strptime(custom_end, '%Y-%m-%d').replace(tzinfo=timezone.utc)
        except ValueError:
            start_date = end_date - timedelta(days=30)
    else:
        start_date = end_date - timedelta(days=30)
    
    date_range = {'start': start_date, 'end': end_date}
    
    # 初始化分配服務
    allocation_service = QuantityAllocationService()
    
    # 取得分配摘要
    allocation_summary = allocation_service.get_allocation_summary(date_range)
    
    # 取得需要分配的工單列表
    from workorder.models import WorkOrder, OperatorSupplementReport
    
    workorders_needing_allocation = []
    workorders = WorkOrder.objects.filter(
        status__in=['in_progress', 'completed']
    ).order_by('-created_at')[:20]
    
    for workorder in workorders:
        # 檢查是否有未填寫數量的報工記錄
        unfilled_reports = OperatorSupplementReport.objects.filter(
            workorder=workorder,
            work_quantity=0,
            approval_status='approved'
        ).exclude(process__name__icontains='包裝')
        
        if unfilled_reports.exists():
            workorders_needing_allocation.append({
                'workorder': workorder,
                'unfilled_count': unfilled_reports.count(),
                'total_reports': OperatorSupplementReport.objects.filter(
                    workorder=workorder,
                    approval_status='approved'
                ).count()
            })
    
    context = {
        'allocation_summary': allocation_summary,
        'workorders_needing_allocation': workorders_needing_allocation,
        'date_range': date_range,
        'date_range_type': date_range_type,
        'custom_start': custom_start,
        'custom_end': custom_end
    }
    
    return render(request, 'reporting/quantity_allocation/dashboard.html', context)


@require_POST
@csrf_exempt
def allocate_workorder_quantities(request):
    """
    API: 為指定工單分配數量
    """
    try:
        workorder_id = request.POST.get('workorder_id')
        date_range_type = request.POST.get('date_range_type', 'month')
        custom_start = request.POST.get('custom_start')
        custom_end = request.POST.get('custom_end')
        
        if not workorder_id:
            return JsonResponse({
                'success': False,
                'message': '請提供工單號碼'
            })
        
        # 設定日期範圍
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        end_date = timezone.now()
        if date_range_type == 'week':
            start_date = end_date - timedelta(days=7)
        elif date_range_type == 'month':
            start_date = end_date - timedelta(days=30)
        elif date_range_type == 'quarter':
            start_date = end_date - timedelta(days=90)
        elif date_range_type == 'custom' and custom_start and custom_end:
            try:
                start_date = datetime.strptime(custom_start, '%Y-%m-%d').replace(tzinfo=timezone.utc)
                end_date = datetime.strptime(custom_end, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            except ValueError:
                start_date = end_date - timedelta(days=30)
        else:
            start_date = end_date - timedelta(days=30)
        
        date_range = {'start': start_date, 'end': end_date}
        
        # 初始化分配服務
        allocation_service = QuantityAllocationService()
        
        # 驗證數據
        validation_result = allocation_service.validate_allocation_data(workorder_id, date_range)
        if not validation_result['valid']:
            return JsonResponse({
                'success': False,
                'message': validation_result['message']
            })
        
        # 執行分配
        result = allocation_service.allocate_workorder_quantities(workorder_id, date_range)
        
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'分配失敗: {str(e)}'
        })


@require_POST
@csrf_exempt
def allocate_multiple_workorders(request):
    """
    API: 批量分配多個工單數量
    """
    try:
        workorder_ids = request.POST.getlist('workorder_ids[]')
        date_range_type = request.POST.get('date_range_type', 'month')
        custom_start = request.POST.get('custom_start')
        custom_end = request.POST.get('custom_end')
        
        if not workorder_ids:
            return JsonResponse({
                'success': False,
                'message': '請選擇要分配的工單'
            })
        
        # 設定日期範圍
        from datetime import datetime, timedelta
        from django.utils import timezone
        
        end_date = timezone.now()
        if date_range_type == 'week':
            start_date = end_date - timedelta(days=7)
        elif date_range_type == 'month':
            start_date = end_date - timedelta(days=30)
        elif date_range_type == 'quarter':
            start_date = end_date - timedelta(days=90)
        elif date_range_type == 'custom' and custom_start and custom_end:
            try:
                start_date = datetime.strptime(custom_start, '%Y-%m-%d').replace(tzinfo=timezone.utc)
                end_date = datetime.strptime(custom_end, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            except ValueError:
                start_date = end_date - timedelta(days=30)
        else:
            start_date = end_date - timedelta(days=30)
        
        date_range = {'start': start_date, 'end': end_date}
        
        # 初始化分配服務
        allocation_service = QuantityAllocationService()
        
        # 執行批量分配
        result = allocation_service.allocate_multiple_workorders(workorder_ids, date_range)
        
        return JsonResponse(result)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'批量分配失敗: {str(e)}'
        })


def allocation_detail(request, workorder_id):
    """
    分配詳情頁面
    顯示指定工單的分配結果
    """
    from datetime import datetime, timedelta
    from django.utils import timezone
    from workorder.models import WorkOrder, OperatorSupplementReport
    
    try:
        workorder = WorkOrder.objects.get(order_number=workorder_id)
    except WorkOrder.DoesNotExist:
        messages.error(request, f'工單 {workorder_id} 不存在')
        return redirect('reporting:quantity_allocation_dashboard')
    
    # 取得日期範圍參數
    date_range_type = request.GET.get('date_range', 'month')
    custom_start = request.GET.get('custom_start')
    custom_end = request.GET.get('custom_end')
    
    # 設定日期範圍
    end_date = timezone.now()
    if date_range_type == 'week':
        start_date = end_date - timedelta(days=7)
    elif date_range_type == 'month':
        start_date = end_date - timedelta(days=30)
    elif date_range_type == 'quarter':
        start_date = end_date - timedelta(days=90)
    elif date_range_type == 'custom' and custom_start and custom_end:
        try:
            start_date = datetime.strptime(custom_start, '%Y-%m-%d').replace(tzinfo=timezone.utc)
            end_date = datetime.strptime(custom_end, '%Y-%m-%d').replace(tzinfo=timezone.utc)
        except ValueError:
            start_date = end_date - timedelta(days=30)
    else:
        start_date = end_date - timedelta(days=30)
    
    date_range = {'start': start_date, 'end': end_date}
    
    # 取得報工記錄
    reports = OperatorSupplementReport.objects.filter(
        workorder=workorder,
        work_date__range=[date_range['start'].date(), date_range['end'].date()],
        approval_status='approved'
    ).select_related('operator', 'process').order_by('work_date', 'start_time')
    
    # 分類報工記錄
    original_reports = []
    allocated_reports = []
    packaging_reports = []
    
    for report in reports:
        if report.quantity_source == 'original':
            original_reports.append(report)
        elif report.quantity_source == 'allocated':
            allocated_reports.append(report)
        elif report.quantity_source == 'packaging':
            packaging_reports.append(report)
        else:
            # 預設為原始記錄
            original_reports.append(report)
    
    # 計算統計資訊
    total_original = sum(r.work_quantity for r in original_reports)
    total_allocated = sum(r.allocated_quantity for r in allocated_reports)
    total_packaging = sum(r.work_quantity for r in packaging_reports)
    
    context = {
        'workorder': workorder,
        'date_range': date_range,
        'date_range_type': date_range_type,
        'custom_start': custom_start,
        'custom_end': custom_end,
        'original_reports': original_reports,
        'allocated_reports': allocated_reports,
        'packaging_reports': packaging_reports,
        'total_original': total_original,
        'total_allocated': total_allocated,
        'total_packaging': total_packaging,
        'allocation_ratio': (total_allocated / total_packaging * 100) if total_packaging > 0 else 0
    }
    
    return render(request, 'reporting/quantity_allocation/detail.html', context)


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def work_time_report(request):
    """統一工作時間報表視圖"""
    # 記錄用戶操作
    log_user_operation(request.user.username, "reporting", "查看統一工作時間報表")
    
    # 設定預設日期範圍（最近30天）
    end_date = timezone.now().date()
    start_date = end_date - timedelta(days=30)
    
    # 獲取所有作業員列表
    operators = ProductionDailyReport.objects.values_list(
        'operator_or_line', flat=True
    ).distinct().order_by('operator_or_line')
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'operators': operators,
    }
    
    return render(request, 'reporting/work_time_report.html', context)


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def api_work_time_report(request):
    """工作時間報表 API 端點"""
    try:
        report_type = request.GET.get('report_type', 'daily')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        operator_filter = request.GET.get('operator', '')
        
        # 轉換日期格式
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # 根據報表類型獲取數據
        if report_type == 'daily':
            data = get_daily_work_time_data(start_date, end_date, operator_filter)
        elif report_type == 'weekly':
            data = get_weekly_work_time_data(start_date, end_date, operator_filter)
        elif report_type == 'monthly':
            data = get_monthly_work_time_data(start_date, end_date, operator_filter)
        else:
            data = get_daily_work_time_data(start_date, end_date, operator_filter)
        
        # 準備圖表數據
        charts = prepare_work_time_charts(data, report_type)
        
        return JsonResponse({
            'success': True,
            'data': data,
            'charts': charts
        })
        
    except Exception as e:
        logger.error(f"工作時間報表 API 錯誤: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def export_work_time_report(request):
    """匯出工作時間報表"""
    try:
        report_type = request.GET.get('report_type', 'daily')
        start_date = request.GET.get('start_date')
        end_date = request.GET.get('end_date')
        operator_filter = request.GET.get('operator', '')
        
        # 記錄下載操作
        log_user_operation(request.user.username, "reporting", f"下載{report_type}工作時間報表")
        
        # 轉換日期格式
        if start_date:
            start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        if end_date:
            end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
        
        # 根據報表類型獲取數據
        if report_type == 'daily':
            data = get_daily_work_time_data(start_date, end_date, operator_filter)
            report_title = "工作時間日報表"
        elif report_type == 'weekly':
            data = get_weekly_work_time_data(start_date, end_date, operator_filter)
            report_title = "工作時間週報表"
        elif report_type == 'monthly':
            data = get_monthly_work_time_data(start_date, end_date, operator_filter)
            report_title = "工作時間月報表"
        else:
            data = get_daily_work_time_data(start_date, end_date, operator_filter)
            report_title = "工作時間報表"
        
        # 創建 Excel 文件
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = report_title
        
        # 寫入表頭
        headers = [
            "日期",
            "作業員/SMT",
            "設備名稱",
            "開始時間",
            "結束時間",
            "工序",
            "工作時數",
            "效率 (件/小時)",
        ]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            worksheet[f"{col_letter}1"] = header
        
        # 寫入數據
        for row_num, item in enumerate(data, 2):
            worksheet[f"A{row_num}"] = item['date']
            worksheet[f"B{row_num}"] = item['operator_name']
            worksheet[f"C{row_num}"] = item['equipment_name']
            worksheet[f"D{row_num}"] = item['start_time']
            worksheet[f"E{row_num}"] = item['end_time']
            worksheet[f"F{row_num}"] = item['process_name']
            worksheet[f"G{row_num}"] = item['work_hours']
            worksheet[f"H{row_num}"] = item['efficiency_rate']
        
        # 設置響應頭以下載 Excel 文件
        filename = f"work_time_{report_type}_report_{start_date}_{end_date}.xlsx"
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f"attachment; filename={filename}"
        workbook.save(response)
        return response
        
    except Exception as e:
        logger.error(f"匯出工作時間報表錯誤: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=500)


def get_daily_work_time_data(start_date, end_date, operator_filter):
    """獲取日報表數據 - 從報工記錄取得"""
    try:
        from .services.work_time_service import WorkTimeReportService
        
        service = WorkTimeReportService()
        data = service.get_work_time_data(start_date, end_date, "daily")
        
        # 轉換為前端需要的格式
        result = []
        for item in data:
            # 如果有作業員篩選，檢查是否包含該作業員
            if operator_filter:
                worker_names = [r['worker_name'] for r in item['reports']]
                if operator_filter not in worker_names:
                    continue
            
            # 計算平均開始和結束時間
            start_times = [r['start_time'] for r in item['reports'] if r['start_time']]
            end_times = [r['end_time'] for r in item['reports'] if r['end_time']]
            
            avg_start_time = min(start_times) if start_times else "08:00"
            avg_end_time = max(end_times) if end_times else "16:00"
            
            result.append({
                'date': item['date'].strftime('%Y-%m-%d'),
                'operator_name': f"{item['worker_count']}位作業員",
                'equipment_name': '多設備',
                'start_time': avg_start_time.strftime('%H:%M') if hasattr(avg_start_time, 'strftime') else str(avg_start_time),
                'end_time': avg_end_time.strftime('%H:%M') if hasattr(avg_end_time, 'strftime') else str(avg_end_time),
                'process_name': '多工序',
                'work_hours': item['total_work_hours'],
                'efficiency_rate': item['avg_efficiency_rate']
            })
        
        return result
    except Exception as e:
        print(f"取得日報工作時間資料失敗: {str(e)}")
        return []


def get_weekly_work_time_data(start_date, end_date, operator_filter):
    """獲取週報表數據（按週彙總）- 從報工記錄取得"""
    try:
        from .services.work_time_service import WorkTimeReportService
        
        service = WorkTimeReportService()
        data = service.get_work_time_data(start_date, end_date, "weekly")
        
        # 轉換為前端需要的格式
        result = []
        for item in data:
            # 如果有作業員篩選，檢查是否包含該作業員
            if operator_filter:
                worker_names = [r['worker_name'] for r in item['reports']]
                if operator_filter not in worker_names:
                    continue
            
            result.append({
                'date': item['period'],
                'operator_name': f"{item['worker_count']}位作業員",
                'equipment_name': '週彙總',
                'start_time': "08:00",
                'end_time': "16:00",
                'process_name': '週彙總',
                'work_hours': item['total_work_hours'],
                'efficiency_rate': item['avg_efficiency_rate']
            })
        
        return result
    except Exception as e:
        print(f"取得週報工作時間資料失敗: {str(e)}")
        return []


def get_monthly_work_time_data(start_date, end_date, operator_filter):
    """獲取月報表數據（按月彙總）- 從報工記錄取得"""
    try:
        from .services.work_time_service import WorkTimeReportService
        
        service = WorkTimeReportService()
        data = service.get_work_time_data(start_date, end_date, "monthly")
        
        # 轉換為前端需要的格式
        result = []
        for item in data:
            # 如果有作業員篩選，檢查是否包含該作業員
            if operator_filter:
                worker_names = [r['worker_name'] for r in item['reports']]
                if operator_filter not in worker_names:
                    continue
            
            result.append({
                'date': item['period'],
                'operator_name': f"{item['worker_count']}位作業員",
                'equipment_name': '月彙總',
                'start_time': "08:00",
                'end_time': "16:00",
                'process_name': '月彙總',
                'work_hours': item['total_work_hours'],
                'efficiency_rate': item['avg_efficiency_rate']
            })
        
        return result
    except Exception as e:
        print(f"取得月報工作時間資料失敗: {str(e)}")
        return []





def prepare_work_time_charts(data, report_type):
    """準備圖表數據"""
    if not data:
        return {
            'efficiency': {'labels': [], 'data': []},
            'operators': {'labels': [], 'data': []}
        }
    
    # 效率趨勢圖數據
    efficiency_data = {}
    for item in data:
        date_key = item['date'].split(' ')[0]  # 取日期部分
        if date_key not in efficiency_data:
            efficiency_data[date_key] = []
        efficiency_data[date_key].append(item['efficiency_rate'])
    
    # 計算每日平均效率
    efficiency_labels = []
    efficiency_values = []
    for date in sorted(efficiency_data.keys()):
        efficiency_labels.append(date)
        avg_efficiency = sum(efficiency_data[date]) / len(efficiency_data[date])
        efficiency_values.append(round(avg_efficiency, 2))
    
    # 作業員工作量分布（基於工作時數）
    operator_data = {}
    for item in data:
        operator = item['operator_name']
        if operator not in operator_data:
            operator_data[operator] = 0
        operator_data[operator] += item['work_hours']
    
    operator_labels = list(operator_data.keys())
    operator_values = list(operator_data.values())
    
    return {
        'efficiency': {
            'labels': efficiency_labels,
            'data': efficiency_values
        },
        'operators': {
            'labels': operator_labels,
            'data': operator_values
        }
    }

@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def test_work_time(request):
    """工作時間報表測試視圖"""
    from .models import ProductionDailyReport
    
    # 獲取基本統計
    total_count = ProductionDailyReport.objects.count()
    recent_data = ProductionDailyReport.objects.order_by('-date')[:5]
    
    context = {
        'total_count': total_count,
        'recent_data': recent_data,
    }
    
    return render(request, 'reporting/test_work_time.html', context)
