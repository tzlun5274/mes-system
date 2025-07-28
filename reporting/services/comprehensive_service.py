# 綜合分析報表服務
# 本檔案定義了綜合分析報表的業務邏輯服務
# 負責整合多個維度的綜合分析報表

from typing import Dict, List, Any, Optional
from datetime import date, datetime, timedelta
from django.db import models
from django.utils import timezone
import logging

from .base_service import BaseReportService
from ..models import ComprehensiveAnalysisReport
from ..queries.base_query import BaseQuery
from ..validators.base_validator import BaseValidator
from ..transformers.base_transformer import BaseTransformer

# 設定日誌記錄器
logger = logging.getLogger(__name__)


class ComprehensiveQuery(BaseQuery):
    """綜合查詢器"""
    
    def get_data(self, **kwargs) -> List[Dict[str, Any]]:
        """取得綜合分析數據"""
        try:
            start_date = kwargs.get('start_date')
            end_date = kwargs.get('end_date')
            
            # 從各個模組查詢數據
            from workorder.models import WorkOrder, OperatorSupplementReport, SMTProductionReport
            from equip.models import Equipment, EquipmentMaintenance
            
            # 查詢工單數據
            workorders = WorkOrder.objects.all()
            if start_date and end_date:
                workorders = self.filter_by_date_range(
                    workorders, start_date, end_date, 'created_at'
                )
            
            # 查詢報工數據
            operator_reports = OperatorSupplementReport.objects.all()
            smt_reports = SMTProductionReport.objects.all()
            if start_date and end_date:
                operator_reports = self.filter_by_date_range(
                    operator_reports, start_date, end_date, 'report_date'
                )
                smt_reports = self.filter_by_date_range(
                    smt_reports, start_date, end_date, 'report_date'
                )
            
            # 查詢設備數據
            equipment_list = Equipment.objects.all()
            maintenance_records = EquipmentMaintenance.objects.all()
            if start_date and end_date:
                maintenance_records = self.filter_by_date_range(
                    maintenance_records, start_date, end_date, 'maintenance_date'
                )
            
            # 按日期分組統計
            daily_stats = {}
            
            # 統計工單數據
            for workorder in workorders:
                date_key = workorder.created_at.date()
                if date_key not in daily_stats:
                    daily_stats[date_key] = self._init_daily_stats()
                
                stats = daily_stats[date_key]
                stats['total_workorders'] += 1
                stats['planned_quantity'] += workorder.planned_quantity or 0
                
                # 計算交期
                if workorder.planned_end_date and workorder.actual_end_date:
                    delay_days = (workorder.actual_end_date - workorder.planned_end_date).days
                    if delay_days <= 0:
                        stats['on_time_deliveries'] += 1
                    stats['total_deliveries'] += 1
                    stats['total_delay_days'] += max(0, delay_days)
            
            # 統計報工數據
            for report in operator_reports:
                date_key = report.report_date
                if date_key not in daily_stats:
                    daily_stats[date_key] = self._init_daily_stats()
                
                stats = daily_stats[date_key]
                stats['total_production_quantity'] += report.quantity or 0
                stats['total_defect_quantity'] += report.defect_quantity or 0
                stats['total_work_hours'] += report.work_hours or 0
                stats['total_overtime_hours'] += report.overtime_hours or 0
                
                if report.abnormal_notes:
                    stats['abnormal_count'] += 1
            
            for report in smt_reports:
                date_key = report.report_date
                if date_key not in daily_stats:
                    daily_stats[date_key] = self._init_daily_stats()
                
                stats = daily_stats[date_key]
                stats['total_production_quantity'] += report.production_quantity or 0
                stats['total_defect_quantity'] += report.defect_quantity or 0
                stats['total_work_hours'] += report.operation_hours or 0
                
                if report.abnormal_notes:
                    stats['abnormal_count'] += 1
            
            # 統計設備數據
            for record in maintenance_records:
                date_key = record.maintenance_date
                if date_key not in daily_stats:
                    daily_stats[date_key] = self._init_daily_stats()
                
                stats = daily_stats[date_key]
                stats['maintenance_count'] += 1
                stats['maintenance_hours'] += record.duration or 0
                stats['maintenance_cost'] += record.cost or 0
                
                if record.maintenance_type == 'breakdown':
                    stats['breakdown_count'] += 1
            
            # 轉換為列表
            data = []
            for date_key, stats in daily_stats.items():
                # 計算效率指標
                efficiency = (stats['total_production_quantity'] / stats['total_work_hours'] * 100) if stats['total_work_hours'] > 0 else 0
                yield_rate = ((stats['total_production_quantity'] - stats['total_defect_quantity']) / stats['total_production_quantity'] * 100) if stats['total_production_quantity'] > 0 else 0
                on_time_delivery_rate = (stats['on_time_deliveries'] / stats['total_deliveries'] * 100) if stats['total_deliveries'] > 0 else 0
                average_delay = (stats['total_delay_days'] / stats['total_deliveries']) if stats['total_deliveries'] > 0 else 0
                
                # 計算成本（假設值）
                material_cost = stats['total_production_quantity'] * 10  # 假設每件材料成本10元
                labor_cost = stats['total_work_hours'] * 200  # 假設每小時人工成本200元
                overhead_cost = stats['total_work_hours'] * 50  # 假設每小時間接成本50元
                total_cost = material_cost + labor_cost + overhead_cost + stats['maintenance_cost']
                unit_cost = (total_cost / stats['total_production_quantity']) if stats['total_production_quantity'] > 0 else 0
                
                daily_data = {
                    'date': date_key,
                    'total_production_quantity': stats['total_production_quantity'],
                    'total_work_hours': stats['total_work_hours'],
                    'overall_efficiency': efficiency,
                    'capacity_utilization': min(efficiency / 100 * 100, 100),  # 假設最大產能為100%
                    'material_cost': material_cost,
                    'labor_cost': labor_cost,
                    'overhead_cost': overhead_cost,
                    'maintenance_cost': stats['maintenance_cost'],
                    'total_cost': total_cost,
                    'unit_cost': unit_cost,
                    'overall_yield_rate': yield_rate,
                    'quality_cost_rate': (stats['total_defect_quantity'] * 50 / total_cost * 100) if total_cost > 0 else 0,  # 假設每件不良品成本50元
                    'on_time_delivery_rate': on_time_delivery_rate,
                    'average_lead_time': average_delay,
                    'production_score': min(efficiency, 100),
                    'quality_score': yield_rate,
                    'cost_score': max(100 - (unit_cost / 100 * 100), 0),  # 假設標準單位成本100元
                    'delivery_score': on_time_delivery_rate,
                    'overall_score': (efficiency + yield_rate + max(100 - (unit_cost / 100 * 100), 0) + on_time_delivery_rate) / 4,
                    'abnormal_count': stats['abnormal_count'],
                    'maintenance_count': stats['maintenance_count'],
                    'breakdown_count': stats['breakdown_count'],
                    'start_date': start_date,
                    'end_date': end_date
                }
                data.append(daily_data)
            
            return data
            
        except Exception as e:
            self.logger.error(f"綜合分析數據查詢失敗: {e}")
            return []
    
    def _init_daily_stats(self) -> Dict[str, Any]:
        """初始化每日統計數據"""
        return {
            'total_workorders': 0,
            'planned_quantity': 0,
            'total_production_quantity': 0,
            'total_defect_quantity': 0,
            'total_work_hours': 0,
            'total_overtime_hours': 0,
            'on_time_deliveries': 0,
            'total_deliveries': 0,
            'total_delay_days': 0,
            'abnormal_count': 0,
            'maintenance_count': 0,
            'maintenance_hours': 0,
            'maintenance_cost': 0,
            'breakdown_count': 0
        }


class ComprehensiveValidator(BaseValidator):
    """綜合分析數據驗證器"""
    
    def validate(self, data: Any) -> bool:
        """驗證綜合分析數據"""
        try:
            if not isinstance(data, list):
                self.add_error("數據必須是列表格式")
                return False
            
            for i, item in enumerate(data):
                if not isinstance(item, dict):
                    self.add_error(f"第 {i+1} 個項目必須是字典格式", f"item[{i}]")
                    continue
                
                # 驗證必填欄位
                self.validate_required_field(item, 'date')
                
                # 驗證數值欄位
                if 'total_production_quantity' in item:
                    self.validate_integer_field(item, 'total_production_quantity', min_value=0)
                
                if 'total_work_hours' in item:
                    self.validate_float_field(item, 'total_work_hours', min_value=0.0)
                
                if 'overall_efficiency' in item:
                    self.validate_percentage_field(item, 'overall_efficiency')
                
                if 'overall_yield_rate' in item:
                    self.validate_percentage_field(item, 'overall_yield_rate')
                
                if 'on_time_delivery_rate' in item:
                    self.validate_percentage_field(item, 'on_time_delivery_rate')
                
                if 'overall_score' in item:
                    self.validate_percentage_field(item, 'overall_score')
                
                # 驗證成本欄位
                cost_fields = ['material_cost', 'labor_cost', 'overhead_cost', 'maintenance_cost', 'total_cost', 'unit_cost']
                for field in cost_fields:
                    if field in item:
                        self.validate_float_field(item, field, min_value=0.0)
                
                # 跨欄位驗證
                if 'total_production_quantity' in item and 'total_defect_quantity' in item:
                    if item['total_defect_quantity'] > item['total_production_quantity']:
                        self.add_error(
                            f"第 {i+1} 個項目的不良品數量不能大於總產量",
                            f"item[{i}].quantity_validation"
                        )
            
            return not self.has_errors()
            
        except Exception as e:
            self.logger.error(f"綜合分析數據驗證失敗: {e}")
            return False


class ComprehensiveTransformer(BaseTransformer):
    """綜合分析數據轉換器"""
    
    def transform(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """轉換綜合分析數據"""
        try:
            transformed_data = []
            
            for item in data:
                # 計算綜合評分
                production_score = item.get('production_score', 0)
                quality_score = item.get('quality_score', 0)
                cost_score = item.get('cost_score', 0)
                delivery_score = item.get('delivery_score', 0)
                overall_score = item.get('overall_score', 0)
                
                # 格式化數據
                transformed_item = {
                    'date': self.format_date(item.get('date')),
                    'total_production_quantity': self.format_number(item.get('total_production_quantity', 0), 0),
                    'total_work_hours': self.format_number(item.get('total_work_hours', 0), 2),
                    'overall_efficiency': self.format_percentage(item.get('overall_efficiency', 0)),
                    'capacity_utilization': self.format_percentage(item.get('capacity_utilization', 0)),
                    'material_cost': self.format_currency(item.get('material_cost', 0)),
                    'labor_cost': self.format_currency(item.get('labor_cost', 0)),
                    'overhead_cost': self.format_currency(item.get('overhead_cost', 0)),
                    'maintenance_cost': self.format_currency(item.get('maintenance_cost', 0)),
                    'total_cost': self.format_currency(item.get('total_cost', 0)),
                    'unit_cost': self.format_currency(item.get('unit_cost', 0)),
                    'overall_yield_rate': self.format_percentage(item.get('overall_yield_rate', 0)),
                    'quality_cost_rate': self.format_percentage(item.get('quality_cost_rate', 0)),
                    'on_time_delivery_rate': self.format_percentage(item.get('on_time_delivery_rate', 0)),
                    'average_lead_time': self.format_number(item.get('average_lead_time', 0), 1),
                    'production_score': self.format_percentage(production_score),
                    'quality_score': self.format_percentage(quality_score),
                    'cost_score': self.format_percentage(cost_score),
                    'delivery_score': self.format_percentage(delivery_score),
                    'overall_score': self.format_percentage(overall_score),
                    'performance_level': self._get_performance_level(overall_score),
                    'key_improvements': self._get_key_improvements(item),
                    'abnormal_count': self.format_number(item.get('abnormal_count', 0), 0),
                    'maintenance_count': self.format_number(item.get('maintenance_count', 0), 0),
                    'breakdown_count': self.format_number(item.get('breakdown_count', 0), 0),
                    'start_date': self.format_date(item.get('start_date')),
                    'end_date': self.format_date(item.get('end_date'))
                }
                
                transformed_data.append(transformed_item)
            
            return transformed_data
            
        except Exception as e:
            self.logger.error(f"綜合分析數據轉換失敗: {e}")
            return data
    
    def _get_performance_level(self, overall_score: float) -> str:
        """取得績效等級"""
        if overall_score >= 90:
            return "優秀"
        elif overall_score >= 80:
            return "良好"
        elif overall_score >= 70:
            return "一般"
        elif overall_score >= 60:
            return "待改進"
        else:
            return "不合格"
    
    def _get_key_improvements(self, item: Dict[str, Any]) -> List[str]:
        """取得關鍵改善項目"""
        improvements = []
        
        # 基於各項指標提出改善建議
        if item.get('overall_efficiency', 0) < 80:
            improvements.append("提升生產效率")
        
        if item.get('overall_yield_rate', 0) < 95:
            improvements.append("改善產品品質")
        
        if item.get('unit_cost', 0) > 100:
            improvements.append("降低生產成本")
        
        if item.get('on_time_delivery_rate', 0) < 90:
            improvements.append("改善交期管理")
        
        if item.get('abnormal_count', 0) > 5:
            improvements.append("減少異常發生")
        
        if item.get('breakdown_count', 0) > 2:
            improvements.append("加強設備維護")
        
        return improvements if improvements else ["整體表現良好"]


class ComprehensiveReportService(BaseReportService):
    """綜合分析報表服務"""
    
    def __init__(self):
        """初始化服務"""
        super().__init__()
        self.query = ComprehensiveQuery()
        self.validator = ComprehensiveValidator()
        self.transformer = ComprehensiveTransformer()
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def generate_report(self, report_type: str, date_range: Dict[str, date], **kwargs) -> Dict[str, Any]:
        """
        生成綜合分析報表
        
        Args:
            report_type: 報表類型（daily/weekly/monthly）
            date_range: 日期範圍
            **kwargs: 其他參數
            
        Returns:
            Dict[str, Any]: 報表數據
        """
        try:
            start_time = timezone.now()
            
            # 檢查快取
            cache_key = self.get_cache_key(report_type, date_range['start'], **kwargs)
            cached_data = self.get_cached_report(report_type, date_range['start'], **kwargs)
            if cached_data:
                self.logger.info(f"使用快取的綜合分析報表: {cache_key}")
                return cached_data
            
            # 查詢數據
            query_params = {
                'start_date': date_range['start'],
                'end_date': date_range['end'],
                **kwargs
            }
            
            raw_data = self.query.get_data(**query_params)
            
            # 驗證數據
            if not self.validator.validate(raw_data):
                validation_summary = self.validator.get_validation_summary()
                self.logger.error(f"綜合分析數據驗證失敗: {validation_summary}")
                return {
                    'success': False,
                    'error': '數據驗證失敗',
                    'validation_errors': validation_summary
                }
            
            # 轉換數據
            transformed_data = self.transformer.transform(raw_data)
            
            # 計算統計摘要
            summary = self._calculate_summary(transformed_data)
            
            # 按報表類型聚合數據
            aggregated_data = self._aggregate_by_report_type(transformed_data, report_type)
            
            # 準備報表數據
            report_data = {
                'success': True,
                'report_type': report_type,
                'date_range': date_range,
                'data': transformed_data,
                'aggregated_data': aggregated_data,
                'summary': summary,
                'generated_at': timezone.now().isoformat(),
                'total_records': len(transformed_data)
            }
            
            # 快取報表數據
            self.set_cached_report(report_type, date_range['start'], report_data, **kwargs)
            
            # 記錄執行時間
            execution_time = (timezone.now() - start_time).total_seconds()
            self.logger.info(f"綜合分析報表生成完成，執行時間: {execution_time:.2f}秒")
            
            return report_data
            
        except Exception as e:
            self.logger.error(f"綜合分析報表生成失敗: {e}")
            return {
                'success': False,
                'error': str(e),
                'report_type': report_type,
                'date_range': date_range
            }
    
    def _calculate_summary(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """計算統計摘要"""
        try:
            if not data:
                return {}
            
            # 計算總計
            total_production_quantity = sum(
                float(item.get('total_production_quantity', '0').replace(',', '')) 
                for item in data
            )
            total_work_hours = sum(
                float(item.get('total_work_hours', '0').replace(',', '')) 
                for item in data
            )
            
            # 計算平均評分
            production_scores = [
                float(item.get('production_score', '0%').replace('%', '')) 
                for item in data
            ]
            quality_scores = [
                float(item.get('quality_score', '0%').replace('%', '')) 
                for item in data
            ]
            cost_scores = [
                float(item.get('cost_score', '0%').replace('%', '')) 
                for item in data
            ]
            delivery_scores = [
                float(item.get('delivery_score', '0%').replace('%', '')) 
                for item in data
            ]
            overall_scores = [
                float(item.get('overall_score', '0%').replace('%', '')) 
                for item in data
            ]
            
            avg_production = sum(production_scores) / len(production_scores) if production_scores else 0
            avg_quality = sum(quality_scores) / len(quality_scores) if quality_scores else 0
            avg_cost = sum(cost_scores) / len(cost_scores) if cost_scores else 0
            avg_delivery = sum(delivery_scores) / len(delivery_scores) if delivery_scores else 0
            avg_overall = sum(overall_scores) / len(overall_scores) if overall_scores else 0
            
            # 計算總成本
            total_material_cost = sum(
                float(item.get('material_cost', 'NT$0').replace('NT$', '').replace(',', '')) 
                for item in data
            )
            total_labor_cost = sum(
                float(item.get('labor_cost', 'NT$0').replace('NT$', '').replace(',', '')) 
                for item in data
            )
            total_overhead_cost = sum(
                float(item.get('overhead_cost', 'NT$0').replace('NT$', '').replace(',', '')) 
                for item in data
            )
            total_maintenance_cost = sum(
                float(item.get('maintenance_cost', 'NT$0').replace('NT$', '').replace(',', '')) 
                for item in data
            )
            total_cost = total_material_cost + total_labor_cost + total_overhead_cost + total_maintenance_cost
            
            return {
                'total_production_quantity': self.transformer.format_number(total_production_quantity, 0),
                'total_work_hours': self.transformer.format_number(total_work_hours, 2),
                'average_production_score': self.transformer.format_percentage(avg_production),
                'average_quality_score': self.transformer.format_percentage(avg_quality),
                'average_cost_score': self.transformer.format_percentage(avg_cost),
                'average_delivery_score': self.transformer.format_percentage(avg_delivery),
                'average_overall_score': self.transformer.format_percentage(avg_overall),
                'total_material_cost': self.transformer.format_currency(total_material_cost),
                'total_labor_cost': self.transformer.format_currency(total_labor_cost),
                'total_overhead_cost': self.transformer.format_currency(total_overhead_cost),
                'total_maintenance_cost': self.transformer.format_currency(total_maintenance_cost),
                'total_cost': self.transformer.format_currency(total_cost),
                'overall_efficiency': self.transformer.format_percentage(
                    (total_production_quantity / total_work_hours * 100) if total_work_hours > 0 else 0
                )
            }
            
        except Exception as e:
            self.logger.error(f"統計摘要計算失敗: {e}")
            return {}
    
    def _aggregate_by_report_type(self, data: List[Dict[str, Any]], report_type: str) -> Dict[str, Any]:
        """按報表類型聚合數據"""
        try:
            if report_type == 'daily':
                # 按績效等級聚合
                performance_groups = {}
                for item in data:
                    level = item.get('performance_level', '未知')
                    if level not in performance_groups:
                        performance_groups[level] = []
                    performance_groups[level].append(item)
                
                return performance_groups
            
            elif report_type == 'weekly':
                # 按週聚合
                weekly_data = {}
                for item in data:
                    date_obj = datetime.strptime(item.get('date', ''), '%Y-%m-%d').date()
                    week_start = date_obj - timedelta(days=date_obj.weekday())
                    week_key = week_start.strftime('%Y-W%U')
                    
                    if week_key not in weekly_data:
                        weekly_data[week_key] = []
                    weekly_data[week_key].append(item)
                
                return weekly_data
            
            elif report_type == 'monthly':
                # 按月聚合
                monthly_data = {}
                for item in data:
                    date_obj = datetime.strptime(item.get('date', ''), '%Y-%m-%d').date()
                    month_key = date_obj.strftime('%Y-%m')
                    
                    if month_key not in monthly_data:
                        monthly_data[month_key] = []
                    monthly_data[month_key].append(item)
                
                return monthly_data
            
            else:
                return {}
                
        except Exception as e:
            self.logger.error(f"數據聚合失敗: {e}")
            return {}
    
    def export_to_excel(self, report_data: Dict[str, Any], filename: str = None) -> bytes:
        """匯出為Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 建立工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "綜合分析報表"
            
            # 設定標題
            headers = [
                '日期', '總生產數量', '總工作時數', '整體效率', '產能利用率',
                '材料成本', '人工成本', '間接成本', '維護成本', '總成本',
                '單位成本', '整體良率', '品質成本率', '準時交貨率', '平均交期',
                '生產評分', '品質評分', '成本評分', '交期評分', '綜合評分',
                '績效等級', '關鍵改善項目', '異常次數', '維護次數', '故障次數'
            ]
            
            # 寫入標題
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 寫入數據
            data = report_data.get('data', [])
            for row, item in enumerate(data, 2):
                ws.cell(row=row, column=1, value=item.get('date', ''))
                ws.cell(row=row, column=2, value=item.get('total_production_quantity', ''))
                ws.cell(row=row, column=3, value=item.get('total_work_hours', ''))
                ws.cell(row=row, column=4, value=item.get('overall_efficiency', ''))
                ws.cell(row=row, column=5, value=item.get('capacity_utilization', ''))
                ws.cell(row=row, column=6, value=item.get('material_cost', ''))
                ws.cell(row=row, column=7, value=item.get('labor_cost', ''))
                ws.cell(row=row, column=8, value=item.get('overhead_cost', ''))
                ws.cell(row=row, column=9, value=item.get('maintenance_cost', ''))
                ws.cell(row=row, column=10, value=item.get('total_cost', ''))
                ws.cell(row=row, column=11, value=item.get('unit_cost', ''))
                ws.cell(row=row, column=12, value=item.get('overall_yield_rate', ''))
                ws.cell(row=row, column=13, value=item.get('quality_cost_rate', ''))
                ws.cell(row=row, column=14, value=item.get('on_time_delivery_rate', ''))
                ws.cell(row=row, column=15, value=item.get('average_lead_time', ''))
                ws.cell(row=row, column=16, value=item.get('production_score', ''))
                ws.cell(row=row, column=17, value=item.get('quality_score', ''))
                ws.cell(row=row, column=18, value=item.get('cost_score', ''))
                ws.cell(row=row, column=19, value=item.get('delivery_score', ''))
                ws.cell(row=row, column=20, value=item.get('overall_score', ''))
                ws.cell(row=row, column=21, value=item.get('performance_level', ''))
                ws.cell(row=row, column=22, value=', '.join(item.get('key_improvements', [])))
                ws.cell(row=row, column=23, value=item.get('abnormal_count', ''))
                ws.cell(row=row, column=24, value=item.get('maintenance_count', ''))
                ws.cell(row=row, column=25, value=item.get('breakdown_count', ''))
            
            # 調整欄寬
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 儲存到記憶體
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Excel匯出失敗: {e}")
            return b""
    
    def get_report_templates(self) -> List[Dict[str, Any]]:
        """取得報表模板列表"""
        return [
            {
                'name': '綜合分析日報表',
                'type': 'daily',
                'description': '顯示每日綜合分析狀況的詳細報表'
            },
            {
                'name': '綜合分析週報表',
                'type': 'weekly',
                'description': '顯示每週綜合分析狀況的匯總報表'
            },
            {
                'name': '綜合分析月報表',
                'type': 'monthly',
                'description': '顯示每月綜合分析狀況的統計報表'
            }
        ] 