# 工單機種報表服務
# 本檔案定義了工單機種報表的業務邏輯服務
# 負責生成工單和產品的完成狀況統計報表

from typing import Dict, List, Any, Optional
from datetime import date, datetime, timedelta
from django.db import models
from django.utils import timezone
import logging

from .base_service import BaseReportService
from ..models import WorkOrderProductReport
from ..queries.base_query import BaseQuery
from ..validators.base_validator import BaseValidator
from ..transformers.base_transformer import BaseTransformer

# 設定日誌記錄器
logger = logging.getLogger(__name__)


class WorkOrderQuery(BaseQuery):
    """工單查詢器"""
    
    def get_data(self, **kwargs) -> List[Dict[str, Any]]:
        """取得工單數據"""
        try:
            start_date = kwargs.get('start_date')
            end_date = kwargs.get('end_date')
            workorder_number = kwargs.get('workorder_number')
            product_code = kwargs.get('product_code')
            
            # 從工單管理模組查詢數據
            from workorder.models import WorkOrder, OperatorSupplementReport
            
            queryset = WorkOrder.objects.all()
            
            # 按日期範圍篩選
            if start_date and end_date:
                queryset = self.filter_by_date_range(
                    queryset, start_date, end_date, 'created_at'
                )
            
            # 按工單號碼篩選
            if workorder_number:
                queryset = queryset.filter(workorder_number__icontains=workorder_number)
            
            # 按產品編號篩選
            if product_code:
                queryset = queryset.filter(product_code__icontains=product_code)
            
            # 轉換為字典列表
            data = []
            for workorder in queryset:
                # 取得相關的報工記錄
                reports = OperatorSupplementReport.objects.filter(
                    workorder_number=workorder.workorder_number
                )
                
                # 計算統計數據
                total_quantity = sum(report.quantity for report in reports if report.quantity)
                defect_quantity = sum(report.defect_quantity for report in reports if report.defect_quantity)
                
                workorder_data = {
                    'workorder_number': workorder.workorder_number,
                    'product_code': workorder.product_code,
                    'product_name': workorder.product_name,
                    'planned_quantity': workorder.planned_quantity,
                    'planned_start_date': workorder.planned_start_date,
                    'planned_end_date': workorder.planned_end_date,
                    'completed_quantity': total_quantity,
                    'defect_quantity': defect_quantity,
                    'actual_start_date': workorder.actual_start_date,
                    'actual_end_date': workorder.actual_end_date,
                    'status': workorder.status,
                    'created_at': workorder.created_at,
                    'updated_at': workorder.updated_at
                }
                data.append(workorder_data)
            
            return data
            
        except Exception as e:
            self.logger.error(f"工單數據查詢失敗: {e}")
            return []


class WorkOrderValidator(BaseValidator):
    """工單數據驗證器"""
    
    def validate(self, data: Any) -> bool:
        """驗證工單數據"""
        try:
            if not isinstance(data, list):
                self.add_error("數據必須是列表格式")
                return False
            
            for i, item in enumerate(data):
                if not isinstance(item, dict):
                    self.add_error(f"第 {i+1} 個項目必須是字典格式", f"item[{i}]")
                    continue
                
                # 驗證必填欄位
                self.validate_required_field(item, 'workorder_number')
                self.validate_required_field(item, 'product_code')
                self.validate_required_field(item, 'planned_quantity')
                
                # 驗證數值欄位
                if 'planned_quantity' in item:
                    self.validate_integer_field(item, 'planned_quantity', min_value=0)
                
                if 'completed_quantity' in item:
                    self.validate_integer_field(item, 'completed_quantity', min_value=0)
                
                if 'defect_quantity' in item:
                    self.validate_integer_field(item, 'defect_quantity', min_value=0)
                
                # 驗證日期欄位
                if 'planned_start_date' in item and 'planned_end_date' in item:
                    self.validate_date_field(item, 'planned_start_date')
                    self.validate_date_field(item, 'planned_end_date')
                    
                    # 跨欄位驗證：開始日期不能晚於結束日期
                    if item.get('planned_start_date') and item.get('planned_end_date'):
                        if item['planned_start_date'] > item['planned_end_date']:
                            self.add_error(
                                f"工單 {item.get('workorder_number')} 的計劃開始日期不能晚於結束日期",
                                f"item[{i}].planned_dates"
                            )
            
            return not self.has_errors()
            
        except Exception as e:
            self.logger.error(f"工單數據驗證失敗: {e}")
            return False


class WorkOrderTransformer(BaseTransformer):
    """工單數據轉換器"""
    
    def transform(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """轉換工單數據"""
        try:
            transformed_data = []
            
            for item in data:
                # 計算完成率
                planned_qty = item.get('planned_quantity', 0)
                completed_qty = item.get('completed_quantity', 0)
                completion_rate = (completed_qty / planned_qty * 100) if planned_qty > 0 else 0
                
                # 計算良率
                defect_qty = item.get('defect_quantity', 0)
                yield_rate = ((completed_qty - defect_qty) / completed_qty * 100) if completed_qty > 0 else 0
                
                # 計算交期差異
                planned_end = item.get('planned_end_date')
                actual_end = item.get('actual_end_date')
                delivery_delay = 0
                if planned_end and actual_end:
                    delivery_delay = (actual_end - planned_end).days
                
                # 格式化數據
                transformed_item = {
                    'workorder_number': item.get('workorder_number', ''),
                    'product_code': item.get('product_code', ''),
                    'product_name': item.get('product_name', ''),
                    'planned_quantity': self.format_number(planned_qty, 0),
                    'completed_quantity': self.format_number(completed_qty, 0),
                    'defect_quantity': self.format_number(defect_qty, 0),
                    'completion_rate': self.format_percentage(completion_rate),
                    'yield_rate': self.format_percentage(yield_rate),
                    'planned_start_date': self.format_date(planned_end),
                    'planned_end_date': self.format_date(planned_end),
                    'actual_start_date': self.format_date(actual_end),
                    'actual_end_date': self.format_date(actual_end),
                    'delivery_delay': f"{delivery_delay} 天" if delivery_delay != 0 else "準時",
                    'status': item.get('status', ''),
                    'created_at': self.format_date(item.get('created_at')),
                    'updated_at': self.format_date(item.get('updated_at'))
                }
                
                transformed_data.append(transformed_item)
            
            return transformed_data
            
        except Exception as e:
            self.logger.error(f"工單數據轉換失敗: {e}")
            return data


class WorkOrderReportService(BaseReportService):
    """工單機種報表服務"""
    
    def __init__(self):
        """初始化服務"""
        super().__init__()
        self.query = WorkOrderQuery()
        self.validator = WorkOrderValidator()
        self.transformer = WorkOrderTransformer()
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def generate_report(self, report_type: str, date_range: Dict[str, date], **kwargs) -> Dict[str, Any]:
        """
        生成工單機種報表
        
        Args:
            report_type: 報表類型（daily/weekly/monthly）
            date_range: 日期範圍
            **kwargs: 其他參數
            
        Returns:
            Dict[str, Any]: 報表數據
        """
        try:
            start_time = timezone.now()
            
            # 檢查快取
            cache_key = self.get_cache_key(report_type, date_range['start'], **kwargs)
            cached_data = self.get_cached_report(report_type, date_range['start'], **kwargs)
            if cached_data:
                self.logger.info(f"使用快取的工單報表: {cache_key}")
                return cached_data
            
            # 查詢數據
            query_params = {
                'start_date': date_range['start'],
                'end_date': date_range['end'],
                **kwargs
            }
            
            raw_data = self.query.get_data(**query_params)
            
            # 驗證數據
            if not self.validator.validate(raw_data):
                validation_summary = self.validator.get_validation_summary()
                self.logger.error(f"工單數據驗證失敗: {validation_summary}")
                return {
                    'success': False,
                    'error': '數據驗證失敗',
                    'validation_errors': validation_summary
                }
            
            # 轉換數據
            transformed_data = self.transformer.transform(raw_data)
            
            # 計算統計摘要
            summary = self._calculate_summary(transformed_data)
            
            # 按報表類型聚合數據
            aggregated_data = self._aggregate_by_report_type(transformed_data, report_type)
            
            # 準備報表數據
            report_data = {
                'success': True,
                'report_type': report_type,
                'date_range': date_range,
                'data': transformed_data,
                'aggregated_data': aggregated_data,
                'summary': summary,
                'generated_at': timezone.now().isoformat(),
                'total_records': len(transformed_data)
            }
            
            # 快取報表數據
            self.set_cached_report(report_type, date_range['start'], report_data, **kwargs)
            
            # 記錄執行時間
            execution_time = (timezone.now() - start_time).total_seconds()
            self.logger.info(f"工單報表生成完成，執行時間: {execution_time:.2f}秒")
            
            return report_data
            
        except Exception as e:
            self.logger.error(f"工單報表生成失敗: {e}")
            return {
                'success': False,
                'error': str(e),
                'report_type': report_type,
                'date_range': date_range
            }
    
    def _calculate_summary(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """計算統計摘要"""
        try:
            if not data:
                return {}
            
            # 計算總計
            total_planned = sum(
                float(item.get('planned_quantity', '0').replace(',', '')) 
                for item in data
            )
            total_completed = sum(
                float(item.get('completed_quantity', '0').replace(',', '')) 
                for item in data
            )
            total_defect = sum(
                float(item.get('defect_quantity', '0').replace(',', '')) 
                for item in data
            )
            
            # 計算平均完成率和良率
            completion_rates = [
                float(item.get('completion_rate', '0%').replace('%', '')) 
                for item in data
            ]
            yield_rates = [
                float(item.get('yield_rate', '0%').replace('%', '')) 
                for item in data
            ]
            
            avg_completion_rate = sum(completion_rates) / len(completion_rates) if completion_rates else 0
            avg_yield_rate = sum(yield_rates) / len(yield_rates) if yield_rates else 0
            
            return {
                'total_workorders': len(data),
                'total_planned_quantity': self.transformer.format_number(total_planned, 0),
                'total_completed_quantity': self.transformer.format_number(total_completed, 0),
                'total_defect_quantity': self.transformer.format_number(total_defect, 0),
                'average_completion_rate': self.transformer.format_percentage(avg_completion_rate),
                'average_yield_rate': self.transformer.format_percentage(avg_yield_rate),
                'overall_efficiency': self.transformer.format_percentage(
                    (total_completed / total_planned * 100) if total_planned > 0 else 0
                )
            }
            
        except Exception as e:
            self.logger.error(f"統計摘要計算失敗: {e}")
            return {}
    
    def _aggregate_by_report_type(self, data: List[Dict[str, Any]], report_type: str) -> Dict[str, Any]:
        """按報表類型聚合數據"""
        try:
            if report_type == 'daily':
                # 按日期聚合
                return self.transformer.group_by_field(data, 'created_at')
            
            elif report_type == 'weekly':
                # 按週聚合
                weekly_data = {}
                for item in data:
                    created_date = datetime.strptime(item.get('created_at', ''), '%Y-%m-%d').date()
                    week_start = created_date - timedelta(days=created_date.weekday())
                    week_key = week_start.strftime('%Y-W%U')
                    
                    if week_key not in weekly_data:
                        weekly_data[week_key] = []
                    weekly_data[week_key].append(item)
                
                return weekly_data
            
            elif report_type == 'monthly':
                # 按月聚合
                monthly_data = {}
                for item in data:
                    created_date = datetime.strptime(item.get('created_at', ''), '%Y-%m-%d').date()
                    month_key = created_date.strftime('%Y-%m')
                    
                    if month_key not in monthly_data:
                        monthly_data[month_key] = []
                    monthly_data[month_key].append(item)
                
                return monthly_data
            
            else:
                return {}
                
        except Exception as e:
            self.logger.error(f"數據聚合失敗: {e}")
            return {}
    
    def export_to_excel(self, report_data: Dict[str, Any], filename: str = None) -> bytes:
        """匯出為Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 建立工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "工單機種報表"
            
            # 設定標題
            headers = [
                '工單號碼', '產品編號', '產品名稱', '計劃數量', '完成數量',
                '不良品數量', '完成率', '良率', '計劃開始日期', '計劃結束日期',
                '實際開始日期', '實際結束日期', '交期延遲', '狀態'
            ]
            
            # 寫入標題
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 寫入數據
            data = report_data.get('data', [])
            for row, item in enumerate(data, 2):
                ws.cell(row=row, column=1, value=item.get('workorder_number', ''))
                ws.cell(row=row, column=2, value=item.get('product_code', ''))
                ws.cell(row=row, column=3, value=item.get('product_name', ''))
                ws.cell(row=row, column=4, value=item.get('planned_quantity', ''))
                ws.cell(row=row, column=5, value=item.get('completed_quantity', ''))
                ws.cell(row=row, column=6, value=item.get('defect_quantity', ''))
                ws.cell(row=row, column=7, value=item.get('completion_rate', ''))
                ws.cell(row=row, column=8, value=item.get('yield_rate', ''))
                ws.cell(row=row, column=9, value=item.get('planned_start_date', ''))
                ws.cell(row=row, column=10, value=item.get('planned_end_date', ''))
                ws.cell(row=row, column=11, value=item.get('actual_start_date', ''))
                ws.cell(row=row, column=12, value=item.get('actual_end_date', ''))
                ws.cell(row=row, column=13, value=item.get('delivery_delay', ''))
                ws.cell(row=row, column=14, value=item.get('status', ''))
            
            # 調整欄寬
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 儲存到記憶體
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Excel匯出失敗: {e}")
            return b""
    
    def get_report_templates(self) -> List[Dict[str, Any]]:
        """取得報表模板列表"""
        return [
            {
                'name': '工單完成狀況日報表',
                'type': 'daily',
                'description': '顯示每日工單完成狀況的詳細報表'
            },
            {
                'name': '工單完成狀況週報表',
                'type': 'weekly',
                'description': '顯示每週工單完成狀況的匯總報表'
            },
            {
                'name': '工單完成狀況月報表',
                'type': 'monthly',
                'description': '顯示每月工單完成狀況的統計報表'
            }
        ] 