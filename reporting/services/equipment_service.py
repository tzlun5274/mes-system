# 設備效率報表服務
# 本檔案定義了設備效率報表的業務邏輯服務
# 負責生成設備的使用效率和維護狀況統計報表

from typing import Dict, List, Any, Optional
from datetime import date, datetime, timedelta
from django.db import models
from django.utils import timezone
import logging

from .base_service import BaseReportService
from ..models import EquipmentEfficiencyReport
from ..queries.base_query import BaseQuery
from ..validators.base_validator import BaseValidator
from ..transformers.base_transformer import BaseTransformer

# 設定日誌記錄器
logger = logging.getLogger(__name__)


class EquipmentQuery(BaseQuery):
    """設備查詢器"""
    
    def get_data(self, **kwargs) -> List[Dict[str, Any]]:
        """取得設備效率數據"""
        try:
            start_date = kwargs.get('start_date')
            end_date = kwargs.get('end_date')
            equipment_name = kwargs.get('equipment_name')
            equipment_type = kwargs.get('equipment_type')
            production_line = kwargs.get('production_line')
            
            # 從設備管理模組查詢數據
            from equip.models import Equipment, EquipmentMaintenance
            
            queryset = Equipment.objects.all()
            
            # 按設備名稱篩選
            if equipment_name:
                queryset = queryset.filter(name__icontains=equipment_name)
            
            # 按設備類型篩選
            if equipment_type:
                queryset = queryset.filter(equipment_type=equipment_type)
            
            # 按生產線篩選
            if production_line:
                queryset = queryset.filter(production_line__icontains=production_line)
            
            # 從工單管理模組查詢報工數據
            from workorder.models import OperatorSupplementReport, SMTProductionReport
            
            # 轉換為字典列表
            data = []
            for equipment in queryset:
                # 查詢相關的報工記錄
                operator_reports = OperatorSupplementReport.objects.filter(
                    equipment_name=equipment.name
                )
                
                smt_reports = SMTProductionReport.objects.filter(
                    equipment_name=equipment.name
                )
                
                # 按日期範圍篩選
                if start_date and end_date:
                    operator_reports = self.filter_by_date_range(
                        operator_reports, start_date, end_date, 'report_date'
                    )
                    smt_reports = self.filter_by_date_range(
                        smt_reports, start_date, end_date, 'report_date'
                    )
                
                # 計算統計數據
                total_operation_hours = sum(
                    report.work_hours for report in operator_reports if report.work_hours
                )
                total_operation_hours += sum(
                    report.operation_hours for report in smt_reports if report.operation_hours
                )
                
                # 計算實際運作時數（假設為總時數的80%）
                actual_operation_hours = total_operation_hours * 0.8
                idle_hours = total_operation_hours * 0.2
                
                # 查詢維護記錄
                maintenance_records = EquipmentMaintenance.objects.filter(
                    equipment=equipment
                )
                
                if start_date and end_date:
                    maintenance_records = self.filter_by_date_range(
                        maintenance_records, start_date, end_date, 'maintenance_date'
                    )
                
                maintenance_hours = sum(
                    record.duration for record in maintenance_records if record.duration
                )
                
                # 計算產量統計
                total_output = sum(
                    report.quantity for report in operator_reports if report.quantity
                )
                total_output += sum(
                    report.production_quantity for report in smt_reports if report.production_quantity
                )
                
                defect_output = sum(
                    report.defect_quantity for report in operator_reports if report.defect_quantity
                )
                defect_output += sum(
                    report.defect_quantity for report in smt_reports if report.defect_quantity
                )
                
                # 計算效率指標
                availability_rate = ((total_operation_hours - maintenance_hours) / total_operation_hours * 100) if total_operation_hours > 0 else 0
                performance_rate = (actual_operation_hours / total_operation_hours * 100) if total_operation_hours > 0 else 0
                quality_rate = ((total_output - defect_output) / total_output * 100) if total_output > 0 else 0
                oee_rate = (availability_rate * performance_rate * quality_rate) / 10000
                
                # 計算維護統計
                maintenance_count = maintenance_records.count()
                breakdown_count = maintenance_records.filter(maintenance_type='breakdown').count()
                
                # 計算平均修復時間和故障間隔（假設值）
                mttr_hours = 2.0 if breakdown_count > 0 else 0.0
                mtbf_hours = (total_operation_hours / breakdown_count) if breakdown_count > 0 else total_operation_hours
                
                equipment_data = {
                    'equipment_name': equipment.name,
                    'equipment_type': equipment.equipment_type,
                    'equipment_model': equipment.model,
                    'production_line': equipment.production_line,
                    'total_operation_hours': total_operation_hours,
                    'actual_operation_hours': actual_operation_hours,
                    'idle_hours': idle_hours,
                    'maintenance_hours': maintenance_hours,
                    'availability_rate': availability_rate,
                    'performance_rate': performance_rate,
                    'quality_rate': quality_rate,
                    'oee_rate': oee_rate,
                    'total_output': total_output,
                    'defect_output': defect_output,
                    'yield_rate': ((total_output - defect_output) / total_output * 100) if total_output > 0 else 0,
                    'throughput_rate': (total_output / total_operation_hours) if total_operation_hours > 0 else 0,
                    'maintenance_count': maintenance_count,
                    'breakdown_count': breakdown_count,
                    'mttr_hours': mttr_hours,
                    'mtbf_hours': mtbf_hours,
                    'energy_cost': equipment.energy_cost if hasattr(equipment, 'energy_cost') else 0,
                    'maintenance_cost': sum(record.cost for record in maintenance_records if record.cost),
                    'start_date': start_date,
                    'end_date': end_date
                }
                data.append(equipment_data)
            
            return data
            
        except Exception as e:
            self.logger.error(f"設備效率數據查詢失敗: {e}")
            return []


class EquipmentValidator(BaseValidator):
    """設備效率數據驗證器"""
    
    def validate(self, data: Any) -> bool:
        """驗證設備效率數據"""
        try:
            if not isinstance(data, list):
                self.add_error("數據必須是列表格式")
                return False
            
            for i, item in enumerate(data):
                if not isinstance(item, dict):
                    self.add_error(f"第 {i+1} 個項目必須是字典格式", f"item[{i}]")
                    continue
                
                # 驗證必填欄位
                self.validate_required_field(item, 'equipment_name')
                self.validate_required_field(item, 'equipment_type')
                
                # 驗證數值欄位
                if 'total_operation_hours' in item:
                    self.validate_float_field(item, 'total_operation_hours', min_value=0.0)
                
                if 'actual_operation_hours' in item:
                    self.validate_float_field(item, 'actual_operation_hours', min_value=0.0)
                
                if 'idle_hours' in item:
                    self.validate_float_field(item, 'idle_hours', min_value=0.0)
                
                if 'maintenance_hours' in item:
                    self.validate_float_field(item, 'maintenance_hours', min_value=0.0)
                
                # 驗證百分比欄位
                if 'availability_rate' in item:
                    self.validate_percentage_field(item, 'availability_rate')
                
                if 'performance_rate' in item:
                    self.validate_percentage_field(item, 'performance_rate')
                
                if 'quality_rate' in item:
                    self.validate_percentage_field(item, 'quality_rate')
                
                if 'oee_rate' in item:
                    self.validate_percentage_field(item, 'oee_rate')
                
                if 'yield_rate' in item:
                    self.validate_percentage_field(item, 'yield_rate')
                
                if 'throughput_rate' in item:
                    self.validate_float_field(item, 'throughput_rate', min_value=0.0)
                
                # 驗證整數欄位
                if 'total_output' in item:
                    self.validate_integer_field(item, 'total_output', min_value=0)
                
                if 'defect_output' in item:
                    self.validate_integer_field(item, 'defect_output', min_value=0)
                
                if 'maintenance_count' in item:
                    self.validate_integer_field(item, 'maintenance_count', min_value=0)
                
                if 'breakdown_count' in item:
                    self.validate_integer_field(item, 'breakdown_count', min_value=0)
                
                # 跨欄位驗證
                if 'total_operation_hours' in item and 'actual_operation_hours' in item:
                    if item['actual_operation_hours'] > item['total_operation_hours']:
                        self.add_error(
                            f"設備 {item.get('equipment_name')} 的實際運作時數不能大於總運作時數",
                            f"item[{i}].operation_hours_validation"
                        )
                
                if 'total_output' in item and 'defect_output' in item:
                    if item['defect_output'] > item['total_output']:
                        self.add_error(
                            f"設備 {item.get('equipment_name')} 的不良品產出不能大於總產出",
                            f"item[{i}].output_validation"
                        )
            
            return not self.has_errors()
            
        except Exception as e:
            self.logger.error(f"設備效率數據驗證失敗: {e}")
            return False


class EquipmentTransformer(BaseTransformer):
    """設備效率數據轉換器"""
    
    def transform(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """轉換設備效率數據"""
        try:
            transformed_data = []
            
            for item in data:
                # 計算總成本
                energy_cost = item.get('energy_cost', 0)
                maintenance_cost = item.get('maintenance_cost', 0)
                total_cost = energy_cost + maintenance_cost
                
                # 格式化數據
                transformed_item = {
                    'equipment_name': item.get('equipment_name', ''),
                    'equipment_type': item.get('equipment_type', ''),
                    'equipment_model': item.get('equipment_model', ''),
                    'production_line': item.get('production_line', ''),
                    'total_operation_hours': self.format_number(item.get('total_operation_hours', 0), 2),
                    'actual_operation_hours': self.format_number(item.get('actual_operation_hours', 0), 2),
                    'idle_hours': self.format_number(item.get('idle_hours', 0), 2),
                    'maintenance_hours': self.format_number(item.get('maintenance_hours', 0), 2),
                    'availability_rate': self.format_percentage(item.get('availability_rate', 0)),
                    'performance_rate': self.format_percentage(item.get('performance_rate', 0)),
                    'quality_rate': self.format_percentage(item.get('quality_rate', 0)),
                    'oee_rate': self.format_percentage(item.get('oee_rate', 0)),
                    'total_output': self.format_number(item.get('total_output', 0), 0),
                    'defect_output': self.format_number(item.get('defect_output', 0), 0),
                    'yield_rate': self.format_percentage(item.get('yield_rate', 0)),
                    'throughput_rate': self.format_number(item.get('throughput_rate', 0), 2),
                    'maintenance_count': self.format_number(item.get('maintenance_count', 0), 0),
                    'breakdown_count': self.format_number(item.get('breakdown_count', 0), 0),
                    'mttr_hours': self.format_number(item.get('mttr_hours', 0), 2),
                    'mtbf_hours': self.format_number(item.get('mtbf_hours', 0), 2),
                    'energy_cost': self.format_currency(energy_cost),
                    'maintenance_cost': self.format_currency(maintenance_cost),
                    'total_cost': self.format_currency(total_cost),
                    'efficiency_level': self._get_efficiency_level(item.get('oee_rate', 0)),
                    'maintenance_status': self._get_maintenance_status(item.get('breakdown_count', 0)),
                    'start_date': self.format_date(item.get('start_date')),
                    'end_date': self.format_date(item.get('end_date'))
                }
                
                transformed_data.append(transformed_item)
            
            return transformed_data
            
        except Exception as e:
            self.logger.error(f"設備效率數據轉換失敗: {e}")
            return data
    
    def _get_efficiency_level(self, oee_rate: float) -> str:
        """取得效率等級"""
        if oee_rate >= 85:
            return "優秀"
        elif oee_rate >= 70:
            return "良好"
        elif oee_rate >= 50:
            return "一般"
        elif oee_rate >= 30:
            return "待改進"
        else:
            return "不合格"
    
    def _get_maintenance_status(self, breakdown_count: int) -> str:
        """取得維護狀態"""
        if breakdown_count == 0:
            return "正常"
        elif breakdown_count <= 2:
            return "輕微"
        elif breakdown_count <= 5:
            return "中等"
        else:
            return "嚴重"


class EquipmentReportService(BaseReportService):
    """設備效率報表服務"""
    
    def __init__(self):
        """初始化服務"""
        super().__init__()
        self.query = EquipmentQuery()
        self.validator = EquipmentValidator()
        self.transformer = EquipmentTransformer()
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def generate_report(self, report_type: str, date_range: Dict[str, date], **kwargs) -> Dict[str, Any]:
        """
        生成設備效率報表
        
        Args:
            report_type: 報表類型（daily/weekly/monthly）
            date_range: 日期範圍
            **kwargs: 其他參數
            
        Returns:
            Dict[str, Any]: 報表數據
        """
        try:
            start_time = timezone.now()
            
            # 檢查快取
            cache_key = self.get_cache_key(report_type, date_range['start'], **kwargs)
            cached_data = self.get_cached_report(report_type, date_range['start'], **kwargs)
            if cached_data:
                self.logger.info(f"使用快取的設備效率報表: {cache_key}")
                return cached_data
            
            # 查詢數據
            query_params = {
                'start_date': date_range['start'],
                'end_date': date_range['end'],
                **kwargs
            }
            
            raw_data = self.query.get_data(**query_params)
            
            # 驗證數據
            if not self.validator.validate(raw_data):
                validation_summary = self.validator.get_validation_summary()
                self.logger.error(f"設備效率數據驗證失敗: {validation_summary}")
                return {
                    'success': False,
                    'error': '數據驗證失敗',
                    'validation_errors': validation_summary
                }
            
            # 轉換數據
            transformed_data = self.transformer.transform(raw_data)
            
            # 計算統計摘要
            summary = self._calculate_summary(transformed_data)
            
            # 按報表類型聚合數據
            aggregated_data = self._aggregate_by_report_type(transformed_data, report_type)
            
            # 準備報表數據
            report_data = {
                'success': True,
                'report_type': report_type,
                'date_range': date_range,
                'data': transformed_data,
                'aggregated_data': aggregated_data,
                'summary': summary,
                'generated_at': timezone.now().isoformat(),
                'total_records': len(transformed_data)
            }
            
            # 快取報表數據
            self.set_cached_report(report_type, date_range['start'], report_data, **kwargs)
            
            # 記錄執行時間
            execution_time = (timezone.now() - start_time).total_seconds()
            self.logger.info(f"設備效率報表生成完成，執行時間: {execution_time:.2f}秒")
            
            return report_data
            
        except Exception as e:
            self.logger.error(f"設備效率報表生成失敗: {e}")
            return {
                'success': False,
                'error': str(e),
                'report_type': report_type,
                'date_range': date_range
            }
    
    def _calculate_summary(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """計算統計摘要"""
        try:
            if not data:
                return {}
            
            # 計算總計
            total_equipment = len(data)
            total_operation_hours = sum(
                float(item.get('total_operation_hours', '0').replace(',', '')) 
                for item in data
            )
            total_maintenance_hours = sum(
                float(item.get('maintenance_hours', '0').replace(',', '')) 
                for item in data
            )
            total_output = sum(
                float(item.get('total_output', '0').replace(',', '')) 
                for item in data
            )
            total_defect_output = sum(
                float(item.get('defect_output', '0').replace(',', '')) 
                for item in data
            )
            
            # 計算平均效率指標
            oee_rates = [
                float(item.get('oee_rate', '0%').replace('%', '')) 
                for item in data
            ]
            availability_rates = [
                float(item.get('availability_rate', '0%').replace('%', '')) 
                for item in data
            ]
            performance_rates = [
                float(item.get('performance_rate', '0%').replace('%', '')) 
                for item in data
            ]
            quality_rates = [
                float(item.get('quality_rate', '0%').replace('%', '')) 
                for item in data
            ]
            
            avg_oee = sum(oee_rates) / len(oee_rates) if oee_rates else 0
            avg_availability = sum(availability_rates) / len(availability_rates) if availability_rates else 0
            avg_performance = sum(performance_rates) / len(performance_rates) if performance_rates else 0
            avg_quality = sum(quality_rates) / len(quality_rates) if quality_rates else 0
            
            # 計算總成本
            total_energy_cost = sum(
                float(item.get('energy_cost', 'NT$0').replace('NT$', '').replace(',', '')) 
                for item in data
            )
            total_maintenance_cost = sum(
                float(item.get('maintenance_cost', 'NT$0').replace('NT$', '').replace(',', '')) 
                for item in data
            )
            total_cost = total_energy_cost + total_maintenance_cost
            
            return {
                'total_equipment': total_equipment,
                'total_operation_hours': self.transformer.format_number(total_operation_hours, 2),
                'total_maintenance_hours': self.transformer.format_number(total_maintenance_hours, 2),
                'total_output': self.transformer.format_number(total_output, 0),
                'total_defect_output': self.transformer.format_number(total_defect_output, 0),
                'average_oee_rate': self.transformer.format_percentage(avg_oee),
                'average_availability_rate': self.transformer.format_percentage(avg_availability),
                'average_performance_rate': self.transformer.format_percentage(avg_performance),
                'average_quality_rate': self.transformer.format_percentage(avg_quality),
                'total_energy_cost': self.transformer.format_currency(total_energy_cost),
                'total_maintenance_cost': self.transformer.format_currency(total_maintenance_cost),
                'total_cost': self.transformer.format_currency(total_cost),
                'overall_efficiency': self.transformer.format_percentage(
                    (total_output / total_operation_hours * 100) if total_operation_hours > 0 else 0
                )
            }
            
        except Exception as e:
            self.logger.error(f"統計摘要計算失敗: {e}")
            return {}
    
    def _aggregate_by_report_type(self, data: List[Dict[str, Any]], report_type: str) -> Dict[str, Any]:
        """按報表類型聚合數據"""
        try:
            if report_type == 'daily':
                # 按生產線聚合
                return self.transformer.group_by_field(data, 'production_line')
            
            elif report_type == 'weekly':
                # 按設備類型聚合
                return self.transformer.group_by_field(data, 'equipment_type')
            
            elif report_type == 'monthly':
                # 按效率等級聚合
                efficiency_groups = {}
                for item in data:
                    level = item.get('efficiency_level', '未知')
                    if level not in efficiency_groups:
                        efficiency_groups[level] = []
                    efficiency_groups[level].append(item)
                
                return efficiency_groups
            
            else:
                return {}
                
        except Exception as e:
            self.logger.error(f"數據聚合失敗: {e}")
            return {}
    
    def export_to_excel(self, report_data: Dict[str, Any], filename: str = None) -> bytes:
        """匯出為Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 建立工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "設備效率報表"
            
            # 設定標題
            headers = [
                '設備名稱', '設備類型', '設備型號', '生產線', '總運作時數',
                '實際運作時數', '閒置時數', '維護時數', '可用率', '性能率',
                '品質率', 'OEE', '總產出', '不良品產出', '良率', '產能利用率',
                '維護次數', '故障次數', '平均修復時間', '平均故障間隔',
                '能源成本', '維護成本', '總成本', '效率等級', '維護狀態'
            ]
            
            # 寫入標題
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 寫入數據
            data = report_data.get('data', [])
            for row, item in enumerate(data, 2):
                ws.cell(row=row, column=1, value=item.get('equipment_name', ''))
                ws.cell(row=row, column=2, value=item.get('equipment_type', ''))
                ws.cell(row=row, column=3, value=item.get('equipment_model', ''))
                ws.cell(row=row, column=4, value=item.get('production_line', ''))
                ws.cell(row=row, column=5, value=item.get('total_operation_hours', ''))
                ws.cell(row=row, column=6, value=item.get('actual_operation_hours', ''))
                ws.cell(row=row, column=7, value=item.get('idle_hours', ''))
                ws.cell(row=row, column=8, value=item.get('maintenance_hours', ''))
                ws.cell(row=row, column=9, value=item.get('availability_rate', ''))
                ws.cell(row=row, column=10, value=item.get('performance_rate', ''))
                ws.cell(row=row, column=11, value=item.get('quality_rate', ''))
                ws.cell(row=row, column=12, value=item.get('oee_rate', ''))
                ws.cell(row=row, column=13, value=item.get('total_output', ''))
                ws.cell(row=row, column=14, value=item.get('defect_output', ''))
                ws.cell(row=row, column=15, value=item.get('yield_rate', ''))
                ws.cell(row=row, column=16, value=item.get('throughput_rate', ''))
                ws.cell(row=row, column=17, value=item.get('maintenance_count', ''))
                ws.cell(row=row, column=18, value=item.get('breakdown_count', ''))
                ws.cell(row=row, column=19, value=item.get('mttr_hours', ''))
                ws.cell(row=row, column=20, value=item.get('mtbf_hours', ''))
                ws.cell(row=row, column=21, value=item.get('energy_cost', ''))
                ws.cell(row=row, column=22, value=item.get('maintenance_cost', ''))
                ws.cell(row=row, column=23, value=item.get('total_cost', ''))
                ws.cell(row=row, column=24, value=item.get('efficiency_level', ''))
                ws.cell(row=row, column=25, value=item.get('maintenance_status', ''))
            
            # 調整欄寬
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 儲存到記憶體
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Excel匯出失敗: {e}")
            return b""
    
    def get_report_templates(self) -> List[Dict[str, Any]]:
        """取得報表模板列表"""
        return [
            {
                'name': '設備效率日報表',
                'type': 'daily',
                'description': '顯示每日設備效率狀況的詳細報表'
            },
            {
                'name': '設備效率週報表',
                'type': 'weekly',
                'description': '顯示每週設備效率狀況的匯總報表'
            },
            {
                'name': '設備效率月報表',
                'type': 'monthly',
                'description': '顯示每月設備效率狀況的統計報表'
            }
        ] 