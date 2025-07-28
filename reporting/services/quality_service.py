# 品質分析報表服務
# 本檔案定義了品質分析報表的業務邏輯服務
# 負責生成品質相關的統計分析報表

from typing import Dict, List, Any, Optional
from datetime import date, datetime, timedelta
from django.db import models
from django.utils import timezone
import logging
from collections import Counter

from .base_service import BaseReportService
from ..models import QualityAnalysisReport
from ..queries.base_query import BaseQuery
from ..validators.base_validator import BaseValidator
from ..transformers.base_transformer import BaseTransformer

# 設定日誌記錄器
logger = logging.getLogger(__name__)


class QualityQuery(BaseQuery):
    """品質查詢器"""
    
    def get_data(self, **kwargs) -> List[Dict[str, Any]]:
        """取得品質分析數據"""
        try:
            start_date = kwargs.get('start_date')
            end_date = kwargs.get('end_date')
            product_code = kwargs.get('product_code')
            defect_type = kwargs.get('defect_type')
            
            # 從工單管理模組查詢報工數據
            from workorder.models import OperatorSupplementReport, SMTProductionReport
            
            # 查詢作業員報工記錄
            operator_reports = OperatorSupplementReport.objects.all()
            smt_reports = SMTProductionReport.objects.all()
            
            # 按日期範圍篩選
            if start_date and end_date:
                operator_reports = self.filter_by_date_range(
                    operator_reports, start_date, end_date, 'report_date'
                )
                smt_reports = self.filter_by_date_range(
                    smt_reports, start_date, end_date, 'report_date'
                )
            
            # 按產品編號篩選
            if product_code:
                operator_reports = operator_reports.filter(product_code__icontains=product_code)
                smt_reports = smt_reports.filter(product_code__icontains=product_code)
            
            # 按不良品類型篩選
            if defect_type:
                operator_reports = operator_reports.filter(defect_type__icontains=defect_type)
                smt_reports = smt_reports.filter(defect_type__icontains=defect_type)
            
            # 合併數據
            all_reports = []
            
            # 處理作業員報工記錄
            for report in operator_reports:
                if report.quantity or report.defect_quantity:
                    all_reports.append({
                        'date': report.report_date,
                        'product_code': report.product_code,
                        'product_name': report.product_name,
                        'operator_name': report.operator_name,
                        'equipment_name': report.equipment_name,
                        'quantity': report.quantity or 0,
                        'defect_quantity': report.defect_quantity or 0,
                        'defect_type': report.defect_type,
                        'defect_reason': report.defect_reason,
                        'abnormal_notes': report.abnormal_notes,
                        'report_type': 'operator'
                    })
            
            # 處理SMT報工記錄
            for report in smt_reports:
                if report.production_quantity or report.defect_quantity:
                    all_reports.append({
                        'date': report.report_date,
                        'product_code': report.product_code,
                        'product_name': report.product_name,
                        'operator_name': report.operator_name,
                        'equipment_name': report.equipment_name,
                        'quantity': report.production_quantity or 0,
                        'defect_quantity': report.defect_quantity or 0,
                        'defect_type': report.defect_type,
                        'defect_reason': report.defect_reason,
                        'abnormal_notes': report.abnormal_notes,
                        'report_type': 'smt'
                    })
            
            return all_reports
            
        except Exception as e:
            self.logger.error(f"品質分析數據查詢失敗: {e}")
            return []


class QualityValidator(BaseValidator):
    """品質分析數據驗證器"""
    
    def validate(self, data: Any) -> bool:
        """驗證品質分析數據"""
        try:
            if not isinstance(data, list):
                self.add_error("數據必須是列表格式")
                return False
            
            for i, item in enumerate(data):
                if not isinstance(item, dict):
                    self.add_error(f"第 {i+1} 個項目必須是字典格式", f"item[{i}]")
                    continue
                
                # 驗證必填欄位
                self.validate_required_field(item, 'date')
                self.validate_required_field(item, 'product_code')
                
                # 驗證數值欄位
                if 'quantity' in item:
                    self.validate_integer_field(item, 'quantity', min_value=0)
                
                if 'defect_quantity' in item:
                    self.validate_integer_field(item, 'defect_quantity', min_value=0)
                
                # 驗證日期欄位
                self.validate_date_field(item, 'date')
                
                # 跨欄位驗證
                if 'quantity' in item and 'defect_quantity' in item:
                    if item['defect_quantity'] > item['quantity']:
                        self.add_error(
                            f"第 {i+1} 個項目的不良品數量不能大於總數量",
                            f"item[{i}].quantity_validation"
                        )
            
            return not self.has_errors()
            
        except Exception as e:
            self.logger.error(f"品質分析數據驗證失敗: {e}")
            return False


class QualityTransformer(BaseTransformer):
    """品質分析數據轉換器"""
    
    def transform(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """轉換品質分析數據"""
        try:
            if not data:
                return []
            
            # 按日期分組統計
            daily_stats = {}
            for item in data:
                date_key = item.get('date')
                if date_key not in daily_stats:
                    daily_stats[date_key] = {
                        'date': date_key,
                        'total_inspected': 0,
                        'passed_quantity': 0,
                        'failed_quantity': 0,
                        'defect_categories': Counter(),
                        'defect_types': Counter(),
                        'products': set(),
                        'operators': set(),
                        'equipment': set(),
                        'abnormal_count': 0
                    }
                
                stats = daily_stats[date_key]
                
                # 累計檢驗數量
                quantity = item.get('quantity', 0)
                defect_quantity = item.get('defect_quantity', 0)
                passed_quantity = quantity - defect_quantity
                
                stats['total_inspected'] += quantity
                stats['passed_quantity'] += passed_quantity
                stats['failed_quantity'] += defect_quantity
                
                # 統計不良品分類
                defect_type = item.get('defect_type', '未知')
                stats['defect_types'][defect_type] += defect_quantity
                
                # 統計產品、作業員、設備
                stats['products'].add(item.get('product_code', ''))
                stats['operators'].add(item.get('operator_name', ''))
                stats['equipment'].add(item.get('equipment_name', ''))
                
                # 統計異常記錄
                if item.get('abnormal_notes'):
                    stats['abnormal_count'] += 1
            
            # 轉換為列表並計算統計數據
            transformed_data = []
            for date_key, stats in daily_stats.items():
                # 計算良率
                yield_rate = (stats['passed_quantity'] / stats['total_inspected'] * 100) if stats['total_inspected'] > 0 else 0
                
                # 取得主要不良品類型
                top_defect_types = [item[0] for item in stats['defect_types'].most_common(5)]
                
                # 格式化數據
                transformed_item = {
                    'date': self.format_date(stats['date']),
                    'total_inspected': self.format_number(stats['total_inspected'], 0),
                    'passed_quantity': self.format_number(stats['passed_quantity'], 0),
                    'failed_quantity': self.format_number(stats['failed_quantity'], 0),
                    'yield_rate': self.format_percentage(yield_rate),
                    'defect_categories': dict(stats['defect_types']),
                    'top_defect_types': top_defect_types,
                    'product_count': len(stats['products']),
                    'operator_count': len(stats['operators']),
                    'equipment_count': len(stats['equipment']),
                    'abnormal_count': self.format_number(stats['abnormal_count'], 0),
                    'quality_level': self._get_quality_level(yield_rate),
                    'defect_trend': self._get_defect_trend(stats['failed_quantity']),
                    'products': list(stats['products']),
                    'operators': list(stats['operators']),
                    'equipment': list(stats['equipment'])
                }
                
                transformed_data.append(transformed_item)
            
            return transformed_data
            
        except Exception as e:
            self.logger.error(f"品質分析數據轉換失敗: {e}")
            return data
    
    def _get_quality_level(self, yield_rate: float) -> str:
        """取得品質等級"""
        if yield_rate >= 99:
            return "優秀"
        elif yield_rate >= 95:
            return "良好"
        elif yield_rate >= 90:
            return "一般"
        elif yield_rate >= 80:
            return "待改進"
        else:
            return "不合格"
    
    def _get_defect_trend(self, defect_quantity: int) -> str:
        """取得不良品趨勢"""
        if defect_quantity == 0:
            return "無不良品"
        elif defect_quantity <= 5:
            return "輕微"
        elif defect_quantity <= 20:
            return "中等"
        else:
            return "嚴重"


class QualityReportService(BaseReportService):
    """品質分析報表服務"""
    
    def __init__(self):
        """初始化服務"""
        super().__init__()
        self.query = QualityQuery()
        self.validator = QualityValidator()
        self.transformer = QualityTransformer()
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def generate_report(self, report_type: str, date_range: Dict[str, date], **kwargs) -> Dict[str, Any]:
        """
        生成品質分析報表
        
        Args:
            report_type: 報表類型（daily/weekly/monthly）
            date_range: 日期範圍
            **kwargs: 其他參數
            
        Returns:
            Dict[str, Any]: 報表數據
        """
        try:
            start_time = timezone.now()
            
            # 檢查快取
            cache_key = self.get_cache_key(report_type, date_range['start'], **kwargs)
            cached_data = self.get_cached_report(report_type, date_range['start'], **kwargs)
            if cached_data:
                self.logger.info(f"使用快取的品質分析報表: {cache_key}")
                return cached_data
            
            # 查詢數據
            query_params = {
                'start_date': date_range['start'],
                'end_date': date_range['end'],
                **kwargs
            }
            
            raw_data = self.query.get_data(**query_params)
            
            # 驗證數據
            if not self.validator.validate(raw_data):
                validation_summary = self.validator.get_validation_summary()
                self.logger.error(f"品質分析數據驗證失敗: {validation_summary}")
                return {
                    'success': False,
                    'error': '數據驗證失敗',
                    'validation_errors': validation_summary
                }
            
            # 轉換數據
            transformed_data = self.transformer.transform(raw_data)
            
            # 計算統計摘要
            summary = self._calculate_summary(transformed_data)
            
            # 按報表類型聚合數據
            aggregated_data = self._aggregate_by_report_type(transformed_data, report_type)
            
            # 準備報表數據
            report_data = {
                'success': True,
                'report_type': report_type,
                'date_range': date_range,
                'data': transformed_data,
                'aggregated_data': aggregated_data,
                'summary': summary,
                'generated_at': timezone.now().isoformat(),
                'total_records': len(transformed_data)
            }
            
            # 快取報表數據
            self.set_cached_report(report_type, date_range['start'], report_data, **kwargs)
            
            # 記錄執行時間
            execution_time = (timezone.now() - start_time).total_seconds()
            self.logger.info(f"品質分析報表生成完成，執行時間: {execution_time:.2f}秒")
            
            return report_data
            
        except Exception as e:
            self.logger.error(f"品質分析報表生成失敗: {e}")
            return {
                'success': False,
                'error': str(e),
                'report_type': report_type,
                'date_range': date_range
            }
    
    def _calculate_summary(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """計算統計摘要"""
        try:
            if not data:
                return {}
            
            # 計算總計
            total_inspected = sum(
                float(item.get('total_inspected', '0').replace(',', '')) 
                for item in data
            )
            total_passed = sum(
                float(item.get('passed_quantity', '0').replace(',', '')) 
                for item in data
            )
            total_failed = sum(
                float(item.get('failed_quantity', '0').replace(',', '')) 
                for item in data
            )
            
            # 計算整體良率
            overall_yield_rate = (total_passed / total_inspected * 100) if total_inspected > 0 else 0
            
            # 統計不良品類型
            all_defect_categories = {}
            for item in data:
                defect_categories = item.get('defect_categories', {})
                for defect_type, count in defect_categories.items():
                    if defect_type not in all_defect_categories:
                        all_defect_categories[defect_type] = 0
                    all_defect_categories[defect_type] += count
            
            # 取得主要不良品類型
            top_defect_types = sorted(all_defect_categories.items(), key=lambda x: x[1], reverse=True)[:5]
            
            # 計算異常統計
            total_abnormal = sum(
                float(item.get('abnormal_count', '0').replace(',', '')) 
                for item in data
            )
            
            # 計算品質等級分布
            quality_levels = Counter(item.get('quality_level', '') for item in data)
            
            return {
                'total_inspected': self.transformer.format_number(total_inspected, 0),
                'total_passed': self.transformer.format_number(total_passed, 0),
                'total_failed': self.transformer.format_number(total_failed, 0),
                'overall_yield_rate': self.transformer.format_percentage(overall_yield_rate),
                'total_abnormal_count': self.transformer.format_number(total_abnormal, 0),
                'top_defect_types': top_defect_types,
                'quality_level_distribution': dict(quality_levels),
                'average_daily_yield': self.transformer.format_percentage(
                    sum(float(item.get('yield_rate', '0%').replace('%', '')) for item in data) / len(data) if data else 0
                ),
                'defect_rate': self.transformer.format_percentage(
                    (total_failed / total_inspected * 100) if total_inspected > 0 else 0
                )
            }
            
        except Exception as e:
            self.logger.error(f"統計摘要計算失敗: {e}")
            return {}
    
    def _aggregate_by_report_type(self, data: List[Dict[str, Any]], report_type: str) -> Dict[str, Any]:
        """按報表類型聚合數據"""
        try:
            if report_type == 'daily':
                # 按品質等級聚合
                quality_groups = {}
                for item in data:
                    level = item.get('quality_level', '未知')
                    if level not in quality_groups:
                        quality_groups[level] = []
                    quality_groups[level].append(item)
                
                return quality_groups
            
            elif report_type == 'weekly':
                # 按週聚合
                weekly_data = {}
                for item in data:
                    date_obj = datetime.strptime(item.get('date', ''), '%Y-%m-%d').date()
                    week_start = date_obj - timedelta(days=date_obj.weekday())
                    week_key = week_start.strftime('%Y-W%U')
                    
                    if week_key not in weekly_data:
                        weekly_data[week_key] = []
                    weekly_data[week_key].append(item)
                
                return weekly_data
            
            elif report_type == 'monthly':
                # 按月聚合
                monthly_data = {}
                for item in data:
                    date_obj = datetime.strptime(item.get('date', ''), '%Y-%m-%d').date()
                    month_key = date_obj.strftime('%Y-%m')
                    
                    if month_key not in monthly_data:
                        monthly_data[month_key] = []
                    monthly_data[month_key].append(item)
                
                return monthly_data
            
            else:
                return {}
                
        except Exception as e:
            self.logger.error(f"數據聚合失敗: {e}")
            return {}
    
    def export_to_excel(self, report_data: Dict[str, Any], filename: str = None) -> bytes:
        """匯出為Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 建立工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "品質分析報表"
            
            # 設定標題
            headers = [
                '日期', '總檢驗數量', '合格數量', '不合格數量', '良率',
                '不良品類型', '主要不良品類型', '產品數量', '作業員數量',
                '設備數量', '異常次數', '品質等級', '不良品趨勢'
            ]
            
            # 寫入標題
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 寫入數據
            data = report_data.get('data', [])
            for row, item in enumerate(data, 2):
                ws.cell(row=row, column=1, value=item.get('date', ''))
                ws.cell(row=row, column=2, value=item.get('total_inspected', ''))
                ws.cell(row=row, column=3, value=item.get('passed_quantity', ''))
                ws.cell(row=row, column=4, value=item.get('failed_quantity', ''))
                ws.cell(row=row, column=5, value=item.get('yield_rate', ''))
                ws.cell(row=row, column=6, value=str(item.get('defect_categories', {})))
                ws.cell(row=row, column=7, value=', '.join(item.get('top_defect_types', [])))
                ws.cell(row=row, column=8, value=item.get('product_count', ''))
                ws.cell(row=row, column=9, value=item.get('operator_count', ''))
                ws.cell(row=row, column=10, value=item.get('equipment_count', ''))
                ws.cell(row=row, column=11, value=item.get('abnormal_count', ''))
                ws.cell(row=row, column=12, value=item.get('quality_level', ''))
                ws.cell(row=row, column=13, value=item.get('defect_trend', ''))
            
            # 調整欄寬
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 儲存到記憶體
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Excel匯出失敗: {e}")
            return b""
    
    def get_report_templates(self) -> List[Dict[str, Any]]:
        """取得報表模板列表"""
        return [
            {
                'name': '品質分析日報表',
                'type': 'daily',
                'description': '顯示每日品質狀況的詳細報表'
            },
            {
                'name': '品質分析週報表',
                'type': 'weekly',
                'description': '顯示每週品質狀況的匯總報表'
            },
            {
                'name': '品質分析月報表',
                'type': 'monthly',
                'description': '顯示每月品質狀況的統計報表'
            }
        ] 