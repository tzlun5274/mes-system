# 人員績效報表服務
# 本檔案定義了人員績效報表的業務邏輯服務
# 負責生成作業員和主管的績效統計報表

from typing import Dict, List, Any, Optional
from datetime import date, datetime, timedelta
from django.db import models
from django.utils import timezone
import logging

from .base_service import BaseReportService
from ..models import PersonnelPerformanceReport
from ..queries.base_query import BaseQuery
from ..validators.base_validator import BaseValidator
from ..transformers.base_transformer import BaseTransformer

# 設定日誌記錄器
logger = logging.getLogger(__name__)


class PersonnelQuery(BaseQuery):
    """人員查詢器"""
    
    def get_data(self, **kwargs) -> List[Dict[str, Any]]:
        """取得人員績效數據"""
        try:
            start_date = kwargs.get('start_date')
            end_date = kwargs.get('end_date')
            personnel_name = kwargs.get('personnel_name')
            personnel_type = kwargs.get('personnel_type')
            department = kwargs.get('department')
            
            # 從工單管理模組查詢報工數據
            from workorder.models import OperatorSupplementReport
            
            queryset = OperatorSupplementReport.objects.all()
            
            # 按日期範圍篩選
            if start_date and end_date:
                queryset = self.filter_by_date_range(
                    queryset, start_date, end_date, 'report_date'
                )
            
            # 按人員姓名篩選
            if personnel_name:
                queryset = queryset.filter(operator_name__icontains=personnel_name)
            
            # 按人員類型篩選
            if personnel_type:
                queryset = queryset.filter(operator_type=personnel_type)
            
            # 按部門篩選
            if department:
                queryset = queryset.filter(department__icontains=department)
            
            # 按人員分組並計算統計數據
            personnel_stats = {}
            for report in queryset:
                operator_name = report.operator_name
                
                if operator_name not in personnel_stats:
                    personnel_stats[operator_name] = {
                        'personnel_name': operator_name,
                        'personnel_type': report.operator_type,
                        'department': report.department,
                        'total_work_hours': 0.0,
                        'overtime_hours': 0.0,
                        'work_days': set(),
                        'total_completed_quantity': 0,
                        'total_defect_quantity': 0,
                        'completed_workorders': set(),
                        'abnormal_count': 0,
                        'abnormal_hours': 0.0,
                        'reports': []
                    }
                
                stats = personnel_stats[operator_name]
                
                # 累計工作時數
                if report.work_hours:
                    stats['total_work_hours'] += report.work_hours
                
                # 累計加班時數
                if report.overtime_hours:
                    stats['overtime_hours'] += report.overtime_hours
                
                # 記錄工作天數
                if report.report_date:
                    stats['work_days'].add(report.report_date)
                
                # 累計完成數量
                if report.quantity:
                    stats['total_completed_quantity'] += report.quantity
                
                # 累計不良品數量
                if report.defect_quantity:
                    stats['total_defect_quantity'] += report.defect_quantity
                
                # 記錄完成的工單
                if report.workorder_number:
                    stats['completed_workorders'].add(report.workorder_number)
                
                # 累計異常記錄
                if report.abnormal_notes:
                    stats['abnormal_count'] += 1
                    # 假設異常時數為1小時
                    stats['abnormal_hours'] += 1.0
                
                stats['reports'].append(report)
            
            # 轉換為列表格式
            data = []
            for operator_name, stats in personnel_stats.items():
                # 計算出勤率
                total_days = (end_date - start_date).days + 1 if start_date and end_date else 30
                work_days_count = len(stats['work_days'])
                attendance_rate = (work_days_count / total_days * 100) if total_days > 0 else 0
                
                # 計算平均效率
                avg_efficiency = (stats['total_completed_quantity'] / stats['total_work_hours']) if stats['total_work_hours'] > 0 else 0
                
                # 計算良率
                yield_rate = ((stats['total_completed_quantity'] - stats['total_defect_quantity']) / stats['total_completed_quantity'] * 100) if stats['total_completed_quantity'] > 0 else 0
                
                personnel_data = {
                    'personnel_name': stats['personnel_name'],
                    'personnel_type': stats['personnel_type'],
                    'department': stats['department'],
                    'total_work_hours': stats['total_work_hours'],
                    'overtime_hours': stats['overtime_hours'],
                    'work_days': work_days_count,
                    'attendance_rate': attendance_rate,
                    'total_completed_quantity': stats['total_completed_quantity'],
                    'total_defect_quantity': stats['total_defect_quantity'],
                    'average_efficiency': avg_efficiency,
                    'yield_rate': yield_rate,
                    'completed_workorders': len(stats['completed_workorders']),
                    'abnormal_count': stats['abnormal_count'],
                    'abnormal_hours': stats['abnormal_hours'],
                    'start_date': start_date,
                    'end_date': end_date
                }
                data.append(personnel_data)
            
            return data
            
        except Exception as e:
            self.logger.error(f"人員績效數據查詢失敗: {e}")
            return []


class PersonnelValidator(BaseValidator):
    """人員績效數據驗證器"""
    
    def validate(self, data: Any) -> bool:
        """驗證人員績效數據"""
        try:
            if not isinstance(data, list):
                self.add_error("數據必須是列表格式")
                return False
            
            for i, item in enumerate(data):
                if not isinstance(item, dict):
                    self.add_error(f"第 {i+1} 個項目必須是字典格式", f"item[{i}]")
                    continue
                
                # 驗證必填欄位
                self.validate_required_field(item, 'personnel_name')
                self.validate_required_field(item, 'personnel_type')
                
                # 驗證數值欄位
                if 'total_work_hours' in item:
                    self.validate_float_field(item, 'total_work_hours', min_value=0.0)
                
                if 'overtime_hours' in item:
                    self.validate_float_field(item, 'overtime_hours', min_value=0.0)
                
                if 'work_days' in item:
                    self.validate_integer_field(item, 'work_days', min_value=0)
                
                if 'attendance_rate' in item:
                    self.validate_percentage_field(item, 'attendance_rate')
                
                if 'total_completed_quantity' in item:
                    self.validate_integer_field(item, 'total_completed_quantity', min_value=0)
                
                if 'total_defect_quantity' in item:
                    self.validate_integer_field(item, 'total_defect_quantity', min_value=0)
                
                if 'average_efficiency' in item:
                    self.validate_float_field(item, 'average_efficiency', min_value=0.0)
                
                if 'yield_rate' in item:
                    self.validate_percentage_field(item, 'yield_rate')
                
                # 跨欄位驗證
                if 'total_completed_quantity' in item and 'total_defect_quantity' in item:
                    if item['total_defect_quantity'] > item['total_completed_quantity']:
                        self.add_error(
                            f"人員 {item.get('personnel_name')} 的不良品數量不能大於完成數量",
                            f"item[{i}].quantity_validation"
                        )
            
            return not self.has_errors()
            
        except Exception as e:
            self.logger.error(f"人員績效數據驗證失敗: {e}")
            return False


class PersonnelTransformer(BaseTransformer):
    """人員績效數據轉換器"""
    
    def transform(self, data: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        """轉換人員績效數據"""
        try:
            transformed_data = []
            
            for item in data:
                # 計算績效評分
                productivity_score = self._calculate_productivity_score(item)
                quality_score = self._calculate_quality_score(item)
                attendance_score = self._calculate_attendance_score(item)
                overall_score = (productivity_score + quality_score + attendance_score) / 3
                
                # 格式化數據
                transformed_item = {
                    'personnel_name': item.get('personnel_name', ''),
                    'personnel_type': item.get('personnel_type', ''),
                    'department': item.get('department', ''),
                    'total_work_hours': self.format_number(item.get('total_work_hours', 0), 2),
                    'overtime_hours': self.format_number(item.get('overtime_hours', 0), 2),
                    'work_days': self.format_number(item.get('work_days', 0), 0),
                    'attendance_rate': self.format_percentage(item.get('attendance_rate', 0)),
                    'total_completed_quantity': self.format_number(item.get('total_completed_quantity', 0), 0),
                    'total_defect_quantity': self.format_number(item.get('total_defect_quantity', 0), 0),
                    'average_efficiency': self.format_number(item.get('average_efficiency', 0), 2),
                    'yield_rate': self.format_percentage(item.get('yield_rate', 0)),
                    'completed_workorders': self.format_number(item.get('completed_workorders', 0), 0),
                    'abnormal_count': self.format_number(item.get('abnormal_count', 0), 0),
                    'abnormal_hours': self.format_number(item.get('abnormal_hours', 0), 2),
                    'productivity_score': self.format_percentage(productivity_score),
                    'quality_score': self.format_percentage(quality_score),
                    'attendance_score': self.format_percentage(attendance_score),
                    'overall_score': self.format_percentage(overall_score),
                    'performance_level': self._get_performance_level(overall_score),
                    'start_date': self.format_date(item.get('start_date')),
                    'end_date': self.format_date(item.get('end_date'))
                }
                
                transformed_data.append(transformed_item)
            
            return transformed_data
            
        except Exception as e:
            self.logger.error(f"人員績效數據轉換失敗: {e}")
            return data
    
    def _calculate_productivity_score(self, item: Dict[str, Any]) -> float:
        """計算生產力評分"""
        try:
            efficiency = item.get('average_efficiency', 0)
            completed_qty = item.get('total_completed_quantity', 0)
            work_hours = item.get('total_work_hours', 0)
            
            # 基於效率和產量的綜合評分
            if work_hours > 0:
                # 效率評分（假設標準效率為10件/小時）
                efficiency_score = min(efficiency / 10 * 100, 100)
                # 產量評分（假設標準產量為100件/天）
                quantity_score = min(completed_qty / (work_hours / 8 * 100) * 100, 100)
                return (efficiency_score + quantity_score) / 2
            else:
                return 0.0
        except:
            return 0.0
    
    def _calculate_quality_score(self, item: Dict[str, Any]) -> float:
        """計算品質評分"""
        try:
            yield_rate = item.get('yield_rate', 0)
            abnormal_count = item.get('abnormal_count', 0)
            
            # 基於良率和異常次數的評分
            yield_score = yield_rate  # 良率直接作為評分
            abnormal_score = max(100 - abnormal_count * 10, 0)  # 每次異常扣10分
            
            return (yield_score + abnormal_score) / 2
        except:
            return 0.0
    
    def _calculate_attendance_score(self, item: Dict[str, Any]) -> float:
        """計算出勤評分"""
        try:
            attendance_rate = item.get('attendance_rate', 0)
            return attendance_rate  # 出勤率直接作為評分
        except:
            return 0.0
    
    def _get_performance_level(self, score: float) -> str:
        """取得績效等級"""
        if score >= 90:
            return "優秀"
        elif score >= 80:
            return "良好"
        elif score >= 70:
            return "一般"
        elif score >= 60:
            return "待改進"
        else:
            return "不合格"


class PersonnelReportService(BaseReportService):
    """人員績效報表服務"""
    
    def __init__(self):
        """初始化服務"""
        super().__init__()
        self.query = PersonnelQuery()
        self.validator = PersonnelValidator()
        self.transformer = PersonnelTransformer()
        self.logger = logging.getLogger(self.__class__.__name__)
    
    def generate_report(self, report_type: str, date_range: Dict[str, date], **kwargs) -> Dict[str, Any]:
        """
        生成人員績效報表
        
        Args:
            report_type: 報表類型（daily/weekly/monthly）
            date_range: 日期範圍
            **kwargs: 其他參數
            
        Returns:
            Dict[str, Any]: 報表數據
        """
        try:
            start_time = timezone.now()
            
            # 檢查快取
            cache_key = self.get_cache_key(report_type, date_range['start'], **kwargs)
            cached_data = self.get_cached_report(report_type, date_range['start'], **kwargs)
            if cached_data:
                self.logger.info(f"使用快取的人員績效報表: {cache_key}")
                return cached_data
            
            # 查詢數據
            query_params = {
                'start_date': date_range['start'],
                'end_date': date_range['end'],
                **kwargs
            }
            
            raw_data = self.query.get_data(**query_params)
            
            # 驗證數據
            if not self.validator.validate(raw_data):
                validation_summary = self.validator.get_validation_summary()
                self.logger.error(f"人員績效數據驗證失敗: {validation_summary}")
                return {
                    'success': False,
                    'error': '數據驗證失敗',
                    'validation_errors': validation_summary
                }
            
            # 轉換數據
            transformed_data = self.transformer.transform(raw_data)
            
            # 計算統計摘要
            summary = self._calculate_summary(transformed_data)
            
            # 按報表類型聚合數據
            aggregated_data = self._aggregate_by_report_type(transformed_data, report_type)
            
            # 準備報表數據
            report_data = {
                'success': True,
                'report_type': report_type,
                'date_range': date_range,
                'data': transformed_data,
                'aggregated_data': aggregated_data,
                'summary': summary,
                'generated_at': timezone.now().isoformat(),
                'total_records': len(transformed_data)
            }
            
            # 快取報表數據
            self.set_cached_report(report_type, date_range['start'], report_data, **kwargs)
            
            # 記錄執行時間
            execution_time = (timezone.now() - start_time).total_seconds()
            self.logger.info(f"人員績效報表生成完成，執行時間: {execution_time:.2f}秒")
            
            return report_data
            
        except Exception as e:
            self.logger.error(f"人員績效報表生成失敗: {e}")
            return {
                'success': False,
                'error': str(e),
                'report_type': report_type,
                'date_range': date_range
            }
    
    def _calculate_summary(self, data: List[Dict[str, Any]]) -> Dict[str, Any]:
        """計算統計摘要"""
        try:
            if not data:
                return {}
            
            # 計算總計
            total_personnel = len(data)
            total_work_hours = sum(
                float(item.get('total_work_hours', '0').replace(',', '')) 
                for item in data
            )
            total_overtime_hours = sum(
                float(item.get('overtime_hours', '0').replace(',', '')) 
                for item in data
            )
            total_completed_quantity = sum(
                float(item.get('total_completed_quantity', '0').replace(',', '')) 
                for item in data
            )
            total_defect_quantity = sum(
                float(item.get('total_defect_quantity', '0').replace(',', '')) 
                for item in data
            )
            
            # 計算平均評分
            productivity_scores = [
                float(item.get('productivity_score', '0%').replace('%', '')) 
                for item in data
            ]
            quality_scores = [
                float(item.get('quality_score', '0%').replace('%', '')) 
                for item in data
            ]
            attendance_scores = [
                float(item.get('attendance_score', '0%').replace('%', '')) 
                for item in data
            ]
            overall_scores = [
                float(item.get('overall_score', '0%').replace('%', '')) 
                for item in data
            ]
            
            avg_productivity = sum(productivity_scores) / len(productivity_scores) if productivity_scores else 0
            avg_quality = sum(quality_scores) / len(quality_scores) if quality_scores else 0
            avg_attendance = sum(attendance_scores) / len(attendance_scores) if attendance_scores else 0
            avg_overall = sum(overall_scores) / len(overall_scores) if overall_scores else 0
            
            return {
                'total_personnel': total_personnel,
                'total_work_hours': self.transformer.format_number(total_work_hours, 2),
                'total_overtime_hours': self.transformer.format_number(total_overtime_hours, 2),
                'total_completed_quantity': self.transformer.format_number(total_completed_quantity, 0),
                'total_defect_quantity': self.transformer.format_number(total_defect_quantity, 0),
                'average_productivity_score': self.transformer.format_percentage(avg_productivity),
                'average_quality_score': self.transformer.format_percentage(avg_quality),
                'average_attendance_score': self.transformer.format_percentage(avg_attendance),
                'average_overall_score': self.transformer.format_percentage(avg_overall),
                'overall_efficiency': self.transformer.format_percentage(
                    (total_completed_quantity / total_work_hours * 100) if total_work_hours > 0 else 0
                )
            }
            
        except Exception as e:
            self.logger.error(f"統計摘要計算失敗: {e}")
            return {}
    
    def _aggregate_by_report_type(self, data: List[Dict[str, Any]], report_type: str) -> Dict[str, Any]:
        """按報表類型聚合數據"""
        try:
            if report_type == 'daily':
                # 按部門聚合
                return self.transformer.group_by_field(data, 'department')
            
            elif report_type == 'weekly':
                # 按人員類型聚合
                return self.transformer.group_by_field(data, 'personnel_type')
            
            elif report_type == 'monthly':
                # 按績效等級聚合
                performance_groups = {}
                for item in data:
                    level = item.get('performance_level', '未知')
                    if level not in performance_groups:
                        performance_groups[level] = []
                    performance_groups[level].append(item)
                
                return performance_groups
            
            else:
                return {}
                
        except Exception as e:
            self.logger.error(f"數據聚合失敗: {e}")
            return {}
    
    def export_to_excel(self, report_data: Dict[str, Any], filename: str = None) -> bytes:
        """匯出為Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 建立工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "人員績效報表"
            
            # 設定標題
            headers = [
                '人員姓名', '人員類型', '部門', '總工作時數', '加班時數',
                '工作天數', '出勤率', '完成數量', '不良品數量', '平均效率',
                '良率', '完成工單數', '異常次數', '異常時數', '生產力評分',
                '品質評分', '出勤評分', '綜合評分', '績效等級'
            ]
            
            # 寫入標題
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            # 寫入數據
            data = report_data.get('data', [])
            for row, item in enumerate(data, 2):
                ws.cell(row=row, column=1, value=item.get('personnel_name', ''))
                ws.cell(row=row, column=2, value=item.get('personnel_type', ''))
                ws.cell(row=row, column=3, value=item.get('department', ''))
                ws.cell(row=row, column=4, value=item.get('total_work_hours', ''))
                ws.cell(row=row, column=5, value=item.get('overtime_hours', ''))
                ws.cell(row=row, column=6, value=item.get('work_days', ''))
                ws.cell(row=row, column=7, value=item.get('attendance_rate', ''))
                ws.cell(row=row, column=8, value=item.get('total_completed_quantity', ''))
                ws.cell(row=row, column=9, value=item.get('total_defect_quantity', ''))
                ws.cell(row=row, column=10, value=item.get('average_efficiency', ''))
                ws.cell(row=row, column=11, value=item.get('yield_rate', ''))
                ws.cell(row=row, column=12, value=item.get('completed_workorders', ''))
                ws.cell(row=row, column=13, value=item.get('abnormal_count', ''))
                ws.cell(row=row, column=14, value=item.get('abnormal_hours', ''))
                ws.cell(row=row, column=15, value=item.get('productivity_score', ''))
                ws.cell(row=row, column=16, value=item.get('quality_score', ''))
                ws.cell(row=row, column=17, value=item.get('attendance_score', ''))
                ws.cell(row=row, column=18, value=item.get('overall_score', ''))
                ws.cell(row=row, column=19, value=item.get('performance_level', ''))
            
            # 調整欄寬
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 儲存到記憶體
            from io import BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            self.logger.error(f"Excel匯出失敗: {e}")
            return b""
    
    def get_report_templates(self) -> List[Dict[str, Any]]:
        """取得報表模板列表"""
        return [
            {
                'name': '人員績效日報表',
                'type': 'daily',
                'description': '顯示每日人員績效狀況的詳細報表'
            },
            {
                'name': '人員績效週報表',
                'type': 'weekly',
                'description': '顯示每週人員績效狀況的匯總報表'
            },
            {
                'name': '人員績效月報表',
                'type': 'monthly',
                'description': '顯示每月人員績效狀況的統計報表'
            }
        ] 