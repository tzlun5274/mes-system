"""
批次匯出服務
提供已完工工單分析的批次匯出功能
"""

import os
import json
import zipfile
import tempfile
from datetime import datetime
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import csv
import io

from .models import CompletedWorkOrderAnalysis


class BatchExportService:
    """批次匯出服務"""
    
    @staticmethod
    def export_workorder_analysis(analysis_ids, export_type='single', export_format='excel', include_details=True):
        """
        匯出工單分析資料
        
        Args:
            analysis_ids: 分析ID列表
            export_type: 匯出類型 ('single' 或 'multiple')
            export_format: 檔案格式 ('excel' 或 'csv')
            include_details: 是否包含詳細資料
            
        Returns:
            HttpResponse 或檔案路徑
        """
        try:
            # 取得分析資料
            analyses = CompletedWorkOrderAnalysis.objects.filter(id__in=analysis_ids)
            
            if not analyses.exists():
                return None
            
            if export_type == 'single':
                # 全部匯出到一個檔案
                return BatchExportService._export_single_file(analyses, export_format, include_details)
            else:
                # 個別匯出
                return BatchExportService._export_multiple_files(analyses, export_format, include_details)
                
        except Exception as e:
            print(f"批次匯出錯誤: {str(e)}")
            return None
    
    @staticmethod
    def _export_single_file(analyses, export_format, include_details):
        """匯出到單一檔案"""
        if export_format == 'excel':
            return BatchExportService._create_excel_file(analyses, include_details)
        else:
            return BatchExportService._create_csv_file(analyses, include_details)
    
    @staticmethod
    def _export_multiple_files(analyses, export_format, include_details):
        """匯出多個檔案並打包成 ZIP"""
        temp_dir = tempfile.mkdtemp()
        zip_path = os.path.join(temp_dir, f'工單分析個別匯出_{datetime.now().strftime("%Y%m%d_%H%M%S")}.zip')
        
        with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for analysis in analyses:
                if export_format == 'excel':
                    file_content = BatchExportService._create_excel_file([analysis], include_details)
                    file_extension = 'xlsx'
                else:
                    file_content = BatchExportService._create_csv_file([analysis], include_details)
                    file_extension = 'csv'
                
                # 將檔案內容寫入 ZIP
                filename = f"{analysis.workorder_id}_{analysis.company_name}_分析報告.{file_extension}"
                zip_file.writestr(filename, file_content.content)
        
        # 讀取 ZIP 檔案內容
        with open(zip_path, 'rb') as f:
            zip_content = f.read()
        
        # 清理暫存檔案
        os.remove(zip_path)
        os.rmdir(temp_dir)
        
        # 建立回應
        response = HttpResponse(zip_content, content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="工單分析個別匯出_{datetime.now().strftime("%Y%m%d")}.zip"'
        return response
    
    @staticmethod
    def _create_excel_file(analyses, include_details):
        """建立 Excel 檔案"""
        wb = Workbook()
        ws = wb.active
        ws.title = "工單分析報告"
        
        # 設定標題樣式
        title_font = Font(name='微軟正黑體', size=14, bold=True)
        header_font = Font(name='微軟正黑體', size=12, bold=True)
        header_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # 標題
        ws['A1'] = '已完工工單分析報告'
        ws['A1'].font = title_font
        ws.merge_cells('A1:J1')
        ws['A1'].alignment = Alignment(horizontal='center')
        
        # 生成時間
        ws['A2'] = f'生成時間: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'
        ws['A2'].font = Font(name='微軟正黑體', size=10)
        
        # 表頭
        headers = [
            '公司名稱', '公司代號', '工單編號', '產品編號', '產品名稱',
            '完工日期', '執行天數', '工作時數', '工單預定數量', '工序數', '效率比率'
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # 資料行
        for row, analysis in enumerate(analyses, 5):
            ws.cell(row=row, column=1, value=analysis.company_name)  # 公司名稱
            ws.cell(row=row, column=2, value=analysis.company_code)  # 公司代號
            ws.cell(row=row, column=3, value=analysis.workorder_id)  # 工單編號
            ws.cell(row=row, column=4, value=analysis.product_code)  # 產品編號
            ws.cell(row=row, column=5, value=analysis.product_name)  # 產品名稱
            ws.cell(row=row, column=6, value=analysis.completion_date.strftime('%Y-%m-%d'))  # 完工日期
            ws.cell(row=row, column=7, value=analysis.total_execution_days)  # 執行天數
            ws.cell(row=row, column=8, value=round(analysis.total_work_hours, 2))  # 工作時數
            ws.cell(row=row, column=9, value=analysis.order_quantity)  # 工單預定數量
            ws.cell(row=row, column=10, value=analysis.total_processes)  # 工序數
            ws.cell(row=row, column=11, value=f"{analysis.efficiency_rate:.1f}%")  # 效率比率
        
        # 調整欄寬
        column_widths = [15, 10, 15, 20, 30, 12, 10, 10, 12, 8, 10]
        for col, width in enumerate(column_widths, 1):
            try:
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
            except AttributeError:
                # 跳過合併儲存格
                continue
        
        # 如果包含詳細資料，增加工序詳細資料工作表
        if include_details:
            BatchExportService._add_process_details_sheet(wb, analyses)
        
        # 儲存到記憶體
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # 建立回應
        response = HttpResponse(output.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="工單分析批次匯出_{datetime.now().strftime("%Y%m%d")}.xlsx"'
        return response
    
    @staticmethod
    def _create_csv_file(analyses, include_details):
        """建立 CSV 檔案"""
        output = io.StringIO()
        writer = csv.writer(output)
        
        # 寫入標題
        writer.writerow(['已完工工單分析報告'])
        writer.writerow([f'生成時間: {timezone.now().strftime("%Y-%m-%d %H:%M:%S")}'])
        writer.writerow([])  # 空行
        
        # 寫入表頭
        headers = [
            '工單編號', '公司代號', '公司名稱', '產品編號', '產品名稱',
            '完工日期', '執行天數', '工作時數', '工序數', '效率比率'
        ]
        writer.writerow(headers)
        
        # 寫入資料
        for analysis in analyses:
            row = [
                analysis.workorder_id,
                analysis.company_code,
                analysis.company_name,
                analysis.product_code,
                analysis.product_name,
                analysis.completion_date.strftime('%Y-%m-%d'),
                analysis.total_execution_days,
                round(analysis.total_work_hours, 2),
                analysis.total_processes,
                f"{analysis.efficiency_rate:.1f}%"
            ]
            writer.writerow(row)
        
        # 轉換為 BytesIO
        csv_content = output.getvalue()
        output.close()
        
        # 建立回應
        response = HttpResponse(csv_content.encode('utf-8-sig'), content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = f'attachment; filename="工單分析批次匯出_{datetime.now().strftime("%Y%m%d")}.csv"'
        return response
    
    @staticmethod
    def _add_process_details_sheet(wb, analyses):
        """增加工序詳細資料工作表"""
        ws = wb.create_sheet("工序詳細資料")
        
        # 表頭
        headers = ['公司名稱', '工單編號', '產品編號', '工序名稱', '工作日期', '工作開始時間', '工作結束時間', '工作時數', '工作數量', '不良品數量', '作業員', '使用的設備']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(name='微軟正黑體', size=12, bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        row = 2
        for analysis in analyses:
            # 查詢該工單的填報記錄（工序詳細資料）
            from workorder.fill_work.models import FillWork
            
            fill_works = FillWork.objects.filter(
                workorder=analysis.workorder_id,
                product_id=analysis.product_code
            ).order_by('work_date', 'start_time')
            
            for fill_work in fill_works:
                # 寫入工序詳細資料
                ws.cell(row=row, column=1, value=analysis.company_name)  # 公司名稱
                ws.cell(row=row, column=2, value=analysis.workorder_id)  # 工單編號
                ws.cell(row=row, column=3, value=analysis.product_code)   # 產品編號
                ws.cell(row=row, column=4, value=fill_work.operation or fill_work.process_name or '未指定')  # 工序名稱
                ws.cell(row=row, column=5, value=fill_work.work_date.strftime('%Y-%m-%d') if fill_work.work_date else '')  # 工作日期
                ws.cell(row=row, column=6, value=fill_work.start_time.strftime('%H:%M') if fill_work.start_time else '')  # 工作開始時間
                ws.cell(row=row, column=7, value=fill_work.end_time.strftime('%H:%M') if fill_work.end_time else '')  # 工作結束時間
                ws.cell(row=row, column=8, value=round(fill_work.work_hours_calculated or 0, 2))  # 工作時數
                ws.cell(row=row, column=9, value=fill_work.work_quantity or 0)  # 工作數量
                ws.cell(row=row, column=10, value=fill_work.defect_quantity or 0)  # 不良品數量
                ws.cell(row=row, column=11, value=fill_work.operator or '未指定')  # 作業員
                ws.cell(row=row, column=12, value=fill_work.equipment or '未指定')  # 使用的設備
                row += 1
        
        # 調整欄寬
        column_widths = [15, 15, 20, 15, 12, 12, 12, 10, 10, 10, 15, 15]
        for col, width in enumerate(column_widths, 1):
            try:
                ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
            except AttributeError:
                # 跳過合併儲存格
                continue
