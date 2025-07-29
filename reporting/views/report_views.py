# -*- coding: utf-8 -*-
"""
報表查看視圖
提供報表列表、詳情、匯出等功能
"""

from django.views.generic import ListView, DetailView, View, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib import messages
from django.shortcuts import redirect, render
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.db import models
from django.urls import reverse
from django.contrib.auth.decorators import login_required, user_passes_test
from datetime import datetime, timedelta, date
from reporting.models import WorkTimeReport, WorkOrderProductReport
import json
import logging
from reporting.calculators.time_calculator import TimeCalculator

logger = logging.getLogger(__name__)


# 權限檢查函數
def reporting_user_required(user):
    """檢查用戶是否為超級用戶或屬於「報表使用者」群組"""
    return user.is_superuser or user.groups.filter(name="報表使用者").exists()


# 用戶操作日誌函數
def log_user_operation(username, module, action):
    """記錄用戶操作"""
    logger.info(f"用戶操作 - 用戶: {username}, 模組: {module}, 動作: {action}")


class ReportDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """報表儀表板視圖"""
    template_name = 'reporting/dashboard.html'
    
    def test_func(self):
        """檢查用戶權限"""
        return self.request.user.is_authenticated
    
    def get_context_data(self, **kwargs):
        """取得上下文資料"""
        context = super().get_context_data(**kwargs)
        dashboard_data = self.get_dashboard_data()
        context.update(dashboard_data)
        return context
    
    def get_dashboard_data(self):
        """取得儀表板資料"""
        # 今日報表統計
        today = timezone.now().date()
        today_work_time = WorkTimeReport.objects.filter(report_date=today)
        today_work_order = WorkOrderProductReport.objects.filter(report_date=today)
        
        # 本週報表統計
        week_start = today - timedelta(days=today.weekday())
        week_work_time = WorkTimeReport.objects.filter(report_date__gte=week_start)
        week_work_order = WorkOrderProductReport.objects.filter(report_date__gte=week_start)
        
        # 本月報表統計
        month_start = today.replace(day=1)
        month_work_time = WorkTimeReport.objects.filter(report_date__gte=month_start)
        month_work_order = WorkOrderProductReport.objects.filter(report_date__gte=month_start)
        
        # 計算總計數據
        total_production = WorkTimeReport.objects.count()
        today_production = today_work_time.count()
        total_operator = WorkTimeReport.objects.values('worker_name').distinct().count()
        active_operators = WorkTimeReport.objects.filter(report_date__gte=today - timedelta(days=7)).values('worker_name').distinct().count()
        
        # 安全地查詢 WorkOrderProductReport，避免在沒有資料時出錯
        try:
            total_smt = WorkOrderProductReport.objects.count()
            running_equipment = WorkOrderProductReport.objects.filter(report_date=today).count()
        except Exception:
            total_smt = 0
            running_equipment = 0
        
        # 計算平均效率（這裡使用一個簡單的計算方式）
        total_completed = WorkTimeReport.objects.aggregate(total=models.Sum('completed_quantity'))['total'] or 0
        total_defects = WorkTimeReport.objects.aggregate(total=models.Sum('defect_quantity'))['total'] or 0
        total_quantity = total_completed + total_defects
        avg_efficiency = round((total_completed / total_quantity * 100) if total_quantity > 0 else 0, 1)
        
        return {
            'total_production': total_production,
            'today_production': today_production,
            'total_operator': total_operator,
            'active_operators': active_operators,
            'total_smt': total_smt,
            'running_equipment': running_equipment,
            'avg_efficiency': avg_efficiency,
            'now': timezone.now(),
            'today_stats': {
                'work_time_count': today_work_time.count(),
                'work_order_count': today_work_order.count(),
                'total_completed': today_work_time.aggregate(total=models.Sum('completed_quantity'))['total'] or 0,
                'total_defects': today_work_time.aggregate(total=models.Sum('defect_quantity'))['total'] or 0,
            },
            'week_stats': {
                'work_time_count': week_work_time.count(),
                'work_order_count': week_work_order.count(),
                'total_completed': week_work_time.aggregate(total=models.Sum('completed_quantity'))['total'] or 0,
                'total_defects': week_work_time.aggregate(total=models.Sum('defect_quantity'))['total'] or 0,
            },
            'month_stats': {
                'work_time_count': month_work_time.count(),
                'work_order_count': month_work_order.count(),
                'total_completed': month_work_time.aggregate(total=models.Sum('completed_quantity'))['total'] or 0,
                'total_defects': month_work_time.aggregate(total=models.Sum('defect_quantity'))['total'] or 0,
            },
            'recent_work_time': WorkTimeReport.objects.order_by('-report_date', '-created_at')[:10],
            'recent_work_order': WorkOrderProductReport.objects.order_by('-report_date', '-created_at')[:10] if WorkOrderProductReport.objects.exists() else [],
        }


class WorkTimeReportListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """工作報表列表視圖"""
    model = WorkTimeReport
    template_name = 'reporting/work_time_list.html'
    context_object_name = 'work_time_reports'
    paginate_by = 20
    
    def test_func(self):
        """檢查用戶權限"""
        return self.request.user.is_authenticated
    
    def get_queryset(self):
        """取得查詢集"""
        queryset = super().get_queryset()
        
        # 取得報表類型
        report_type = self.request.GET.get('type', 'daily')
        
        # 搜尋功能
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(worker_name__icontains=search) |
                models.Q(workorder_number__icontains=search) |
                models.Q(process_name__icontains=search) |
                models.Q(product_code__icontains=search)
            )
        
        # 日期篩選
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            queryset = queryset.filter(report_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(report_date__lte=date_to)
        
        # 作業員篩選
        worker = self.request.GET.get('worker')
        if worker:
            queryset = queryset.filter(worker_name=worker)
        
        # 工序篩選（週報表和月報表不支援工序篩選）
        if report_type not in ['weekly', 'monthly']:
            process = self.request.GET.get('process')
            if process:
                queryset = queryset.filter(process_name=process)
        
        # 週報表專用：簡單的週期篩選（週一到週日）
        if report_type == 'weekly':
            week_period = self.request.GET.get('week_period', 'current')
            from datetime import date, timedelta
            
            today = date.today()
            
            if week_period == 'current':
                # 本週（週一到週日）
                week_start = today - timedelta(days=today.weekday())
                week_end = week_start + timedelta(days=6)
            elif week_period == 'last':
                # 上週（週一到週日）
                week_start = today - timedelta(days=today.weekday() + 7)
                week_end = week_start + timedelta(days=6)
            elif week_period == 'custom':
                # 自訂日期範圍
                if date_from and date_to:
                    try:
                        week_start = datetime.strptime(date_from, '%Y-%m-%d').date()
                        week_end = datetime.strptime(date_to, '%Y-%m-%d').date()
                    except ValueError:
                        # 如果日期格式錯誤，使用本週
                        week_start = today - timedelta(days=today.weekday())
                        week_end = week_start + timedelta(days=6)
                else:
                    # 如果沒有提供日期，使用本週
                    week_start = today - timedelta(days=today.weekday())
                    week_end = week_start + timedelta(days=6)
            else:
                # 預設使用本週
                week_start = today - timedelta(days=today.weekday())
                week_end = week_start + timedelta(days=6)
            
            queryset = queryset.filter(report_date__range=[week_start, week_end])
        
        # 月報表專用：簡單的週期篩選（月初到月末）
        elif report_type == 'monthly':
            month_period = self.request.GET.get('month_period', 'current')
            from datetime import date, timedelta
            from calendar import monthrange
            
            today = date.today()
            
            if month_period == 'current':
                # 本月（月初到月末）
                month_start = date(today.year, today.month, 1)
                _, last_day = monthrange(today.year, today.month)
                month_end = date(today.year, today.month, last_day)
            elif month_period == 'last':
                # 上月（月初到月末）
                if today.month == 1:
                    last_month_year = today.year - 1
                    last_month = 12
                else:
                    last_month_year = today.year
                    last_month = today.month - 1
                month_start = date(last_month_year, last_month, 1)
                _, last_day = monthrange(last_month_year, last_month)
                month_end = date(last_month_year, last_month, last_day)
            elif month_period == 'custom':
                # 自訂日期範圍
                if date_from and date_to:
                    try:
                        month_start = datetime.strptime(date_from, '%Y-%m-%d').date()
                        month_end = datetime.strptime(date_to, '%Y-%m-%d').date()
                    except ValueError:
                        # 如果日期格式錯誤，使用本月
                        month_start = date(today.year, today.month, 1)
                        _, last_day = monthrange(today.year, today.month)
                        month_end = date(today.year, today.month, last_day)
                else:
                    # 如果沒有提供日期，使用本月
                    month_start = date(today.year, today.month, 1)
                    _, last_day = monthrange(today.year, today.month)
                    month_end = date(today.year, today.month, last_day)
            else:
                # 預設使用本月
                month_start = date(today.year, today.month, 1)
                _, last_day = monthrange(today.year, today.month)
                month_end = date(today.year, today.month, last_day)
            
            queryset = queryset.filter(report_date__range=[month_start, month_end])
        
        # 根據報表類型進行不同的處理
        if report_type == 'weekly':
            # 週報表：按作業員分組，顯示每個作業員的週報工統計
            from django.db.models import Sum, Count, Avg
            
            queryset = queryset.values('worker_name').annotate(
                total_work_hours=Sum('total_work_hours'),
                completed_quantity=Sum('completed_quantity'),
                defect_quantity=Sum('defect_quantity'),
                report_count=Count('id')
            ).order_by('worker_name')
            
        elif report_type == 'monthly':
            # 月報表：按作業員分組，顯示每個作業員的月報工統計
            from django.db.models import Sum, Count, Avg
            
            queryset = queryset.values('worker_name').annotate(
                total_work_hours=Sum('total_work_hours'),
                completed_quantity=Sum('completed_quantity'),
                defect_quantity=Sum('defect_quantity'),
                report_count=Count('id')
            ).order_by('worker_name')
            
        else:
            # 日報表：保持原有邏輯
            order_by = self.request.GET.get('order_by', '-report_date')
            queryset = queryset.order_by(order_by)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """取得上下文資料"""
        context = super().get_context_data(**kwargs)
        
        # 報表類型參數
        context['report_type'] = self.request.GET.get('type', 'daily')
        
        # 搜尋和篩選參數
        context['search'] = self.request.GET.get('search', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['order_by'] = self.request.GET.get('order_by', '-report_date')
        
        # 統計資料 - 根據報表類型分別計算
        report_type = self.request.GET.get('type', 'daily')
        today = date.today()
        
        # 取得基礎查詢集（不包含分組）
        base_queryset = super().get_queryset()
        
        # 搜尋功能
        search = self.request.GET.get('search')
        if search:
            base_queryset = base_queryset.filter(
                models.Q(worker_name__icontains=search) |
                models.Q(workorder_number__icontains=search) |
                models.Q(process_name__icontains=search)
            )
        
        # 日期篩選
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            base_queryset = base_queryset.filter(report_date__gte=date_from)
        if date_to:
            base_queryset = base_queryset.filter(report_date__lte=date_to)
        
        if report_type == 'weekly':
            # 週報表統計 - 使用與查詢集相同的篩選條件
            from django.db.models import Sum
            from django.db.models.functions import TruncWeek
            
            # 重新應用周報表的篩選條件
            weekly_queryset = base_queryset
            week_period = self.request.GET.get('week_period', 'current')
            
            if week_period == 'current':
                week_start = today - timedelta(days=today.weekday())
                week_end = week_start + timedelta(days=6)
            elif week_period == 'last':
                week_start = today - timedelta(days=today.weekday() + 7)
                week_end = week_start + timedelta(days=6)
            elif week_period == 'last2':
                week_start = today - timedelta(days=today.weekday() + 14)
                week_end = week_start + timedelta(days=6)
            elif week_period == 'last4':
                week_start = today - timedelta(days=today.weekday() + 28)
                week_end = week_start + timedelta(days=6)
            elif week_period == 'last8':
                week_start = today - timedelta(days=today.weekday() + 56)
                week_end = week_start + timedelta(days=6)
            elif week_period == 'last12':
                week_start = today - timedelta(days=today.weekday() + 84)
                week_end = week_start + timedelta(days=6)
            elif week_period == 'this_year':
                week_start = date(today.year, 1, 1)
                week_end = today
            elif week_period == 'last_year':
                week_start = date(today.year - 1, 1, 1)
                week_end = date(today.year - 1, 12, 31)
            elif week_period.startswith('week_'):
                try:
                    week_num = int(week_period.split('_')[1])
                    year_start = date(today.year, 1, 1)
                    while year_start.weekday() != 0:
                        year_start += timedelta(days=1)
                    week_start = year_start + timedelta(weeks=week_num - 1)
                    week_end = week_start + timedelta(days=6)
                except (ValueError, IndexError):
                    week_start = today - timedelta(days=today.weekday())
                    week_end = week_start + timedelta(days=6)
            elif week_period == 'custom':
                if date_from and date_to:
                    try:
                        week_start = datetime.strptime(date_from, '%Y-%m-%d').date()
                        week_end = datetime.strptime(date_to, '%Y-%m-%d').date()
                    except ValueError:
                        week_start = today - timedelta(days=today.weekday())
                        week_end = week_start + timedelta(days=6)
                else:
                    week_start = today - timedelta(days=today.weekday())
                    week_end = week_start + timedelta(days=6)
            else:
                week_start = today - timedelta(days=today.weekday())
                week_end = week_start + timedelta(days=6)
            
            weekly_queryset = weekly_queryset.filter(report_date__range=[week_start, week_end])
            
            weekly_stats = weekly_queryset.aggregate(
                total_completed=Sum('completed_quantity'),
                total_defects=Sum('defect_quantity')
            )
            context['total_completed'] = weekly_stats['total_completed'] or 0
            context['total_defects'] = weekly_stats['total_defects'] or 0
            
        elif report_type == 'monthly':
            # 月報表統計 - 使用與查詢集相同的篩選條件
            from django.db.models import Sum
            from django.db.models.functions import TruncMonth
            
            monthly_stats = base_queryset.aggregate(
                total_completed=Sum('completed_quantity'),
                total_defects=Sum('defect_quantity')
            )
            context['total_completed'] = monthly_stats['total_completed'] or 0
            context['total_defects'] = monthly_stats['total_defects'] or 0
            
        else:
            # 日報表統計
            context['total_completed'] = base_queryset.aggregate(
                total=models.Sum('completed_quantity')
            )['total'] or 0
            context['total_defects'] = base_queryset.aggregate(
                total=models.Sum('defect_quantity')
            )['total'] or 0
        
        # 計算平均效率
        total_quantity = context['total_completed'] + context['total_defects']
        if total_quantity > 0:
            context['avg_efficiency'] = round(
                (context['total_completed'] / total_quantity * 100), 1
            )
        else:
            context['avg_efficiency'] = 0
        
        # 作業員選項
        context['workers'] = WorkTimeReport.objects.values_list(
            'worker_name', flat=True
        ).distinct().order_by('worker_name')
        
        # 工序選項（週報表和月報表不顯示工序篩選）
        if report_type not in ['weekly', 'monthly']:
            context['processes'] = WorkTimeReport.objects.values_list(
                'process_name', flat=True
            ).distinct().order_by('process_name')
        else:
            context['processes'] = []
        
        # 周報表專用：周期選項
        if report_type == 'weekly':
            context['week_period'] = self.request.GET.get('week_period', 'current')
            
            # 週報表週期選項 - 簡單明瞭
            week_periods = [
                ('current', '本週'),
                ('last', '上週'),
                ('custom', '自訂日期'),
            ]
            
            context['week_periods'] = week_periods
            
            # 計算當前選擇的週期日期範圍
            week_period = self.request.GET.get('week_period', 'current')
            today = date.today()
            
            if week_period == 'current':
                week_start = today - timedelta(days=today.weekday())
                week_end = week_start + timedelta(days=6)
            elif week_period == 'last':
                week_start = today - timedelta(days=today.weekday() + 7)
                week_end = week_start + timedelta(days=6)
            elif week_period == 'custom':
                if date_from and date_to:
                    try:
                        week_start = datetime.strptime(date_from, '%Y-%m-%d').date()
                        week_end = datetime.strptime(date_to, '%Y-%m-%d').date()
                    except ValueError:
                        week_start = today - timedelta(days=today.weekday())
                        week_end = week_start + timedelta(days=6)
                else:
                    week_start = today - timedelta(days=today.weekday())
                    week_end = week_start + timedelta(days=6)
            else:
                week_start = today - timedelta(days=today.weekday())
                week_end = week_start + timedelta(days=6)
            
            context['week_start'] = week_start
            context['week_end'] = week_end
        
        # 月報表專用：週期選項
        elif report_type == 'monthly':
            context['month_period'] = self.request.GET.get('month_period', 'current')
            
            # 月報表週期選項 - 簡單明瞭
            month_periods = [
                ('current', '本月'),
                ('last', '上月'),
                ('custom', '自訂日期'),
            ]
            
            context['month_periods'] = month_periods
            
            # 計算當前選擇的週期日期範圍
            month_period = self.request.GET.get('month_period', 'current')
            today = date.today()
            from calendar import monthrange
            
            if month_period == 'current':
                month_start = date(today.year, today.month, 1)
                _, last_day = monthrange(today.year, today.month)
                month_end = date(today.year, today.month, last_day)
            elif month_period == 'last':
                if today.month == 1:
                    last_month_year = today.year - 1
                    last_month = 12
                else:
                    last_month_year = today.year
                    last_month = today.month - 1
                month_start = date(last_month_year, last_month, 1)
                _, last_day = monthrange(last_month_year, last_month)
                month_end = date(last_month_year, last_month, last_day)
            elif month_period == 'custom':
                if date_from and date_to:
                    try:
                        month_start = datetime.strptime(date_from, '%Y-%m-%d').date()
                        month_end = datetime.strptime(date_to, '%Y-%m-%d').date()
                    except ValueError:
                        month_start = date(today.year, today.month, 1)
                        _, last_day = monthrange(today.year, today.month)
                        month_end = date(today.year, today.month, last_day)
                else:
                    month_start = date(today.year, today.month, 1)
                    _, last_day = monthrange(today.year, today.month)
                    month_end = date(today.year, today.month, last_day)
            else:
                month_start = date(today.year, today.month, 1)
                _, last_day = monthrange(today.year, today.month)
                month_end = date(today.year, today.month, last_day)
            
            context['month_start'] = month_start
            context['month_end'] = month_end
        
        return context
    
    def paginate_queryset(self, queryset, page_size):
        """自定義分頁邏輯以支援週報表和月報表"""
        # 週報表和月報表現在與日報表邏輯相同，使用預設分頁
        return super().paginate_queryset(queryset, page_size)


class WorkTimeReportDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """工作報表詳情視圖"""
    model = WorkTimeReport
    template_name = 'reporting/work_time_detail.html'
    context_object_name = 'work_time_report'
    
    def test_func(self):
        """檢查用戶權限"""
        return self.request.user.is_authenticated


class WorkTimeReportExportView(LoginRequiredMixin, UserPassesTestMixin, View):
    """工作報表匯出視圖 - 支援日、週、月報表格式"""
    
    def test_func(self):
        """檢查用戶權限"""
        return self.request.user.is_authenticated
    
    def get(self, request, *args, **kwargs):
        """匯出工作報表"""
        try:
            from datetime import datetime, date
            import os
            
            format_type = request.GET.get('format', 'excel')
            report_type = request.GET.get('report_type', 'daily')  # daily, weekly, monthly
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')
            
            # 記錄用戶操作
            log_user_operation(
                request.user.username, 
                "reporting", 
                f"匯出工作{report_type}報表，格式：{format_type}"
            )
            
            # 查詢資料
            queryset = WorkTimeReport.objects.all()
            
            if date_from:
                queryset = queryset.filter(report_date__gte=date_from)
            if date_to:
                queryset = queryset.filter(report_date__lte=date_to)
            
            # 根據報表類型進行資料分組和統計
            if report_type == 'daily':
                data = self._prepare_daily_data(queryset)
                report_title = "工作日報表"
            elif report_type == 'weekly':
                data = self._prepare_weekly_data(queryset)
                report_title = "工作週報表"
            elif report_type == 'monthly':
                data = self._prepare_monthly_data(queryset)
                report_title = "工作月報表"
            else:
                data = self._prepare_daily_data(queryset)
                report_title = "工作報表"
            
            # 生成檔案名稱（包含時間戳記）
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            if format_type == 'csv':
                filename = f'工作{report_type}報表_{timestamp}.csv'
                return self.export_to_csv(data, filename, report_title)
            elif format_type == 'pdf':
                filename = f'工作{report_type}報表_{timestamp}.pdf'
                return self.export_to_pdf(data, filename, report_title)
            else:
                filename = f'工作{report_type}報表_{timestamp}.xlsx'
                return self.export_to_excel(data, filename, report_title)
                
        except Exception as e:
            logger.error(f"匯出工作報表失敗：{str(e)}")
            messages.error(request, f"匯出失敗：{str(e)}")
            return redirect('reporting:work_time_list')
    
    def _prepare_daily_data(self, queryset):
        """準備日報表資料"""
        from reporting.calculators.time_calculator import TimeCalculator
        work_time_calc = TimeCalculator()
        
        data = []
        for report in queryset:
            # 使用新的工作時數計算器重新計算工作時數
            work_time_data = work_time_calc.calculate_work_time_for_report(report)
            
            data.append({
                '報表日期': report.report_date.strftime('%Y-%m-%d'),
                '作業員': report.worker_name,
                '工單號': report.workorder_number,
                '產品編號': getattr(report, 'product_code', ''),
                '工序': report.process_name,
                '開始時間': report.start_time.strftime('%H:%M') if report.start_time else '',
                '結束時間': report.end_time.strftime('%H:%M') if report.end_time else '',
                '總工作時數': round(work_time_data['total_work_hours'], 2),
                '實際工作時數': round(work_time_data['actual_work_hours'], 2),
                '休息時數': round(work_time_data['break_hours'], 2),
                '加班時數': round(work_time_data['overtime_hours'], 2),
                '正常工時': round(work_time_data['regular_hours'], 2),
                '完成數量': report.completed_quantity or 0,
                '不良品': report.defect_quantity or 0,
                '效率(件/小時)': round((report.completed_quantity or 0) / (work_time_data['actual_work_hours'] or 1), 2) if work_time_data['actual_work_hours'] else 0,
            })
        return data
    
    def _prepare_weekly_data(self, queryset):
        """準備週報表資料 - 按週分組統計"""
        from django.db.models import Sum, Avg
        from datetime import timedelta
        
        # 按週分組統計
        weekly_data = {}
        for report in queryset:
            # 計算週開始日期
            week_start = report.report_date - timedelta(days=report.report_date.weekday())
            week_key = week_start.strftime('%Y-%m-%d')
            
            if week_key not in weekly_data:
                weekly_data[week_key] = {
                    '週開始日期': week_start.strftime('%Y-%m-%d'),
                    '週結束日期': (week_start + timedelta(days=6)).strftime('%Y-%m-%d'),
                    '作業員': report.worker_name,
                    '總工作時數': 0,
                    '總完成數量': 0,
                    '總不良品': 0,
                    '平均效率': 0,
                    '工單數量': 0,
                }
            
            weekly_data[week_key]['總工作時數'] += report.total_work_hours or 0
            weekly_data[week_key]['總完成數量'] += report.completed_quantity or 0
            weekly_data[week_key]['總不良品'] += report.defect_quantity or 0
            weekly_data[week_key]['工單數量'] += 1
        
        # 計算平均效率
        for week_data in weekly_data.values():
            if week_data['總工作時數'] > 0:
                week_data['平均效率'] = round(week_data['總完成數量'] / week_data['總工作時數'], 2)
            week_data['總工作時數'] = round(week_data['總工作時數'], 2)
        
        return list(weekly_data.values())
    
    def _prepare_monthly_data(self, queryset):
        """準備月報表資料 - 按月分組統計"""
        from datetime import timedelta
        
        # 按月分組統計
        monthly_data = {}
        for report in queryset:
            # 計算月開始日期
            month_start = report.report_date.replace(day=1)
            month_key = month_start.strftime('%Y-%m')
            
            if month_key not in monthly_data:
                # 計算月末日期
                if month_start.month == 12:
                    next_month = month_start.replace(year=month_start.year + 1, month=1)
                else:
                    next_month = month_start.replace(month=month_start.month + 1)
                month_end = next_month - timedelta(days=1)
                
                monthly_data[month_key] = {
                    '月份': month_start.strftime('%Y年%m月'),
                    '月開始日期': month_start.strftime('%Y-%m-%d'),
                    '月結束日期': month_end.strftime('%Y-%m-%d'),
                    '作業員': report.worker_name,
                    '總工作時數': 0,
                    '總完成數量': 0,
                    '總不良品': 0,
                    '平均效率': 0,
                    '工單數量': 0,
                }
            
            monthly_data[month_key]['總工作時數'] += report.total_work_hours or 0
            monthly_data[month_key]['總完成數量'] += report.completed_quantity or 0
            monthly_data[month_key]['總不良品'] += report.defect_quantity or 0
            monthly_data[month_key]['工單數量'] += 1
        
        # 計算平均效率
        for month_data in monthly_data.values():
            if month_data['總工作時數'] > 0:
                month_data['平均效率'] = round(month_data['總完成數量'] / month_data['總工作時數'], 2)
            month_data['總工作時數'] = round(month_data['總工作時數'], 2)
        
        return list(monthly_data.values())
    
    def export_to_excel(self, data, filename, report_title):
        """匯出為Excel檔案 - 完整樣式設定"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = report_title
        
        # 設定樣式
        header_font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        data_font = Font(name='微軟正黑體', size=11)
        data_alignment = Alignment(horizontal='center', vertical='center')
        
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 寫入表頭
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                cell = worksheet.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # 寫入資料
            for row, record in enumerate(data, 2):
                for col, value in enumerate(record.values(), 1):
                    cell = worksheet.cell(row=row, column=col, value=value)
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border
            
            # 調整欄寬
            for col in range(1, len(headers) + 1):
                max_length = 0
                column_letter = get_column_letter(col)
                for row in range(1, len(data) + 2):
                    cell_value = worksheet.cell(row=row, column=col).value
                    if cell_value:
                        max_length = max(max_length, len(str(cell_value)))
                adjusted_width = min(max_length + 2, 30)
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # 建立回應
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
    
    def export_to_csv(self, data, filename, report_title):
        """匯出為CSV檔案 - UTF-8編碼支援中文"""
        import csv
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # 寫入BOM以支援中文
        response.write('\ufeff')
        
        if data:
            writer = csv.DictWriter(response, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return response
    
    def export_to_pdf(self, data, filename, report_title):
        """匯出為PDF檔案 - 暫時重定向到Excel"""
        # 暫時使用Excel格式，未來實作真正的PDF匯出
        return self.export_to_excel(data, filename, report_title)


class WorkOrderProductReportListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """工單機種報表列表視圖"""
    model = WorkOrderProductReport
    template_name = 'reporting/work_order_list.html'
    context_object_name = 'work_order_reports'
    paginate_by = 20
    
    def test_func(self):
        """檢查用戶權限"""
        return self.request.user.is_authenticated
    
    def get_queryset(self):
        """取得查詢集"""
        queryset = super().get_queryset()
        
        # 搜尋功能
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                models.Q(work_order_no__icontains=search) |
                models.Q(product_name__icontains=search) |
                models.Q(process_name__icontains=search)
            )
        
        # 日期篩選
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            queryset = queryset.filter(report_date__gte=date_from)
        if date_to:
            queryset = queryset.filter(report_date__lte=date_to)
        
        # 排序
        order_by = self.request.GET.get('order_by', '-report_date')
        queryset = queryset.order_by(order_by)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """取得上下文資料"""
        context = super().get_context_data(**kwargs)
        
        # 搜尋和篩選參數
        context['search'] = self.request.GET.get('search', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['order_by'] = self.request.GET.get('order_by', '-report_date')
        
        # 統計資料
        queryset = self.get_queryset()
        context['total_production'] = queryset.aggregate(
            total=models.Sum('actual_quantity')
        )['total'] or 0
        
        # 計算完成率
        total_planned = queryset.aggregate(
            total=models.Sum('planned_quantity')
        )['total'] or 0
        if total_planned > 0:
            context['completion_rate'] = round(
                (context['total_production'] / total_planned * 100), 1
            )
        else:
            context['completion_rate'] = 0
        
        # 計算平均效率
        efficiency_sum = queryset.aggregate(
            total=models.Sum('efficiency')
        )['total'] or 0
        efficiency_count = queryset.count()
        if efficiency_count > 0:
            context['avg_efficiency'] = round(efficiency_sum / efficiency_count, 1)
        else:
            context['avg_efficiency'] = 0
        
        # 機種選項
        context['products'] = WorkOrderProductReport.objects.values_list(
            'product_name', flat=True
        ).distinct().order_by('product_name')
        
        return context


class WorkOrderProductReportDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """工單機種報表詳情視圖"""
    model = WorkOrderProductReport
    template_name = 'reporting/work_order_detail.html'
    context_object_name = 'work_order_report'
    
    def test_func(self):
        """檢查用戶權限"""
        return self.request.user.is_authenticated


class WorkOrderProductReportExportView(LoginRequiredMixin, UserPassesTestMixin, View):
    """工單機種報表匯出視圖"""
    
    def test_func(self):
        """檢查用戶權限"""
        return self.request.user.is_authenticated
    
    def get(self, request, *args, **kwargs):
        """匯出工單機種報表"""
        try:
            format_type = request.GET.get('format', 'excel')
            date_from = request.GET.get('date_from')
            date_to = request.GET.get('date_to')
            
            # 查詢資料
            queryset = WorkOrderProductReport.objects.all()
            
            if date_from:
                queryset = queryset.filter(report_date__gte=date_from)
            if date_to:
                queryset = queryset.filter(report_date__lte=date_to)
            
            # 準備資料
            data = []
            for report in queryset:
                data.append({
                    '報表日期': report.report_date.strftime('%Y-%m-%d'),
                    '作業員': getattr(report, 'worker_name', ''),
                    '工單號': report.work_order_no,
                    '產品編號': report.product_name,
                    '工序': getattr(report, 'process_name', ''),
                    '開始時間': report.start_date.strftime('%Y-%m-%d') if report.start_date else '',
                    '結束時間': report.completion_date.strftime('%Y-%m-%d') if report.completion_date else '',
                    '工作時數': getattr(report, 'work_hours', 0) or 0,
                    '完成數量': report.actual_quantity or 0,
                    '不良品': report.defect_quantity or 0,
                })
            
            # 匯出檔案
            if format_type == 'csv':
                return self.export_to_csv(data, 'work_order_report.csv')
            else:
                return self.export_to_excel(data, 'work_order_report.xlsx')
                
        except Exception as e:
            logger.error(f"匯出工單機種報表失敗：{str(e)}")
            messages.error(request, f"匯出失敗：{str(e)}")
            return redirect('reporting:work_order_list')
    
    def export_to_excel(self, data, filename):
        """匯出為Excel檔案"""
        import openpyxl
        from openpyxl.utils import get_column_letter
        
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "工單機種報表"
        
        # 寫入表頭
        if data:
            headers = list(data[0].keys())
            for col, header in enumerate(headers, 1):
                worksheet.cell(row=1, column=col, value=header)
            
            # 寫入資料
            for row, record in enumerate(data, 2):
                for col, value in enumerate(record.values(), 1):
                    worksheet.cell(row=row, column=col, value=value)
            
            # 調整欄寬
            for col in range(1, len(headers) + 1):
                worksheet.column_dimensions[get_column_letter(col)].width = 15
        
        # 建立回應
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        workbook.save(response)
        return response
    
    def export_to_csv(self, data, filename):
        """匯出為CSV檔案"""
        import csv
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # 寫入BOM以支援中文
        response.write('\ufeff')
        
        if data:
            writer = csv.DictWriter(response, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        return response


class ReportExportView(LoginRequiredMixin, UserPassesTestMixin, View):
    """報表匯出視圖"""
    
    def test_func(self):
        """檢查用戶權限"""
        return self.request.user.is_authenticated
    
    def get(self, request, *args, **kwargs):
        """顯示報表匯出頁面"""
        return render(request, 'reporting/report_export.html')
    
    def post(self, request, *args, **kwargs):
        """執行報表匯出"""
        try:
            report_type = request.POST.get('report_type')
            date_from = request.POST.get('date_from')
            date_to = request.POST.get('date_to')
            format_type = request.POST.get('format', 'excel')
            
            if report_type == 'work_time':
                return redirect(f"{reverse('reporting:work_time_export')}?date_from={date_from}&date_to={date_to}&format={format_type}")
            elif report_type == 'work_order':
                return redirect(f"{reverse('reporting:work_order_export')}?date_from={date_from}&date_to={date_to}&format={format_type}")
            else:
                messages.error(request, "不支援的報表類型")
                return redirect('reporting:export')
                
        except Exception as e:
            logger.error(f"報表匯出失敗：{str(e)}")
            messages.error(request, f"匯出失敗：{str(e)}")
            return redirect('reporting:export')


# 報表匯出函數
@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def report_export(request):
    """
    報表匯出頁面 - 統一管理所有報表匯出功能
    匯出各種報工報表，支援Excel、CSV、PDF等多種格式
    """
    from django.db.models import Q
    from workorder.models import OperatorSupplementReport, SMTProductionReport
    
    # 記錄用戶操作
    log_user_operation(request.user.username, "reporting", "訪問報表匯出頁面")
    
    today = date.today()
    month_start = today.replace(day=1)
    
    # 可匯出的報表類型
    report_types = [
        {'id': 'daily', 'name': '工作日報表', 'description': '每日工作報表統計', 'icon': 'fas fa-calendar-day'},
        {'id': 'weekly', 'name': '工作週報表', 'description': '每週工作報表統計', 'icon': 'fas fa-calendar-week'},
        {'id': 'monthly', 'name': '工作月報表', 'description': '每月工作報表統計', 'icon': 'fas fa-calendar-alt'},
        {'id': 'operator', 'name': '作業員報工報表', 'description': '作業員報工詳細記錄', 'icon': 'fas fa-user'},
        {'id': 'smt', 'name': 'SMT報工報表', 'description': 'SMT設備報工詳細記錄', 'icon': 'fas fa-microchip'},
        {'id': 'abnormal', 'name': '異常報工報表', 'description': '異常報工記錄分析', 'icon': 'fas fa-exclamation-triangle'},
        {'id': 'efficiency', 'name': '效率分析報表', 'description': '產能效率分析報表', 'icon': 'fas fa-tachometer-alt'},
        {'id': 'production_daily', 'name': '生產日報表', 'description': '生產日報表統計', 'icon': 'fas fa-chart-line'},
        {'id': 'operator_performance', 'name': '作業員績效報表', 'description': '作業員績效分析', 'icon': 'fas fa-user-chart'},
    ]
    
    # 匯出格式
    export_formats = [
        {'id': 'excel', 'name': 'Excel (.xlsx)', 'icon': 'fas fa-file-excel'},
        {'id': 'csv', 'name': 'CSV (.csv)', 'icon': 'fas fa-file-csv'},
        {'id': 'pdf', 'name': 'PDF (.pdf)', 'icon': 'fas fa-file-pdf'},
    ]
    
    # 日期範圍選項
    date_ranges = [
        {'id': 'today', 'name': '今日', 'start': today, 'end': today},
        {'id': 'yesterday', 'name': '昨日', 'start': today - timedelta(days=1), 'end': today - timedelta(days=1)},
        {'id': 'week', 'name': '本週', 'start': today - timedelta(days=today.weekday()), 'end': today},
        {'id': 'month', 'name': '本月', 'start': month_start, 'end': today},
        {'id': 'custom', 'name': '自訂日期範圍', 'start': None, 'end': None},
    ]
    
    context = {
        'report_types': report_types,
        'export_formats': export_formats,
        'date_ranges': date_ranges,
    }
    return render(request, 'reporting/report_export.html', context)


@login_required
@user_passes_test(reporting_user_required, login_url="/accounts/login/")
def execute_report_export(request):
    """
    執行報表匯出功能
    """
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from datetime import datetime
    import csv
    from io import StringIO
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '只支援POST請求'})
    
    try:
        report_type = request.POST.get('report_type')
        export_format = request.POST.get('export_format')
        date_range = request.POST.get('date_range')
        custom_start = request.POST.get('custom_start')
        custom_end = request.POST.get('custom_end')
        
        # 記錄用戶操作
        log_user_operation(
            request.user.username, 
            "reporting", 
            f"匯出報表：{report_type}，格式：{export_format}，日期範圍：{date_range}"
        )
        
        # 根據報表類型取得資料
        if report_type in ['daily', 'weekly', 'monthly']:
            data = get_work_report_data(report_type, date_range, custom_start, custom_end)
            filename = f'工作{report_type}報表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        elif report_type == 'operator':
            data = get_operator_report_data(date_range, custom_start, custom_end)
            filename = f'作業員報工報表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        elif report_type == 'smt':
            data = get_smt_report_data(date_range, custom_start, custom_end)
            filename = f'SMT報工報表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        elif report_type == 'abnormal':
            data = get_abnormal_report_data(date_range, custom_start, custom_end)
            filename = f'異常報工報表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        elif report_type == 'production_daily':
            data = get_production_daily_report_data(date_range, custom_start, custom_end)
            filename = f'生產日報表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        elif report_type == 'operator_performance':
            data = get_operator_performance_report_data(date_range, custom_start, custom_end)
            filename = f'作業員績效報表_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
        else:
            return JsonResponse({'success': False, 'error': '不支援的報表類型'})
        
        # 根據格式匯出
        if export_format == 'excel':
            return export_to_excel(data, filename)
        elif export_format == 'csv':
            return export_to_csv(data, filename)
        elif export_format == 'pdf':
            return export_to_pdf(data, filename)
        else:
            return JsonResponse({'success': False, 'error': '不支援的匯出格式'})
            
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'匯出失敗：{str(e)}'})


# 資料獲取函數
def get_operator_report_data(date_range, custom_start=None, custom_end=None):
    """取得作業員報工資料"""
    from workorder.models import OperatorSupplementReport
    from reporting.calculators.time_calculator import TimeCalculator
    
    today = date.today()
    
    # 設定日期範圍
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=30)
        end_date = today
    
    # 查詢資料
    reports = OperatorSupplementReport.objects.filter(
        work_date__gte=start_date,
        work_date__lte=end_date
    ).order_by('work_date', 'worker_name')
    
    # 初始化工作時數計算器
    work_time_calc = TimeCalculator()
    
    # 準備資料
    data = []
    for report in reports:
        # 使用工作時數計算器計算正確的工作時數
        work_time_data = work_time_calc.calculate_work_time_for_report(report)
        
        data.append({
            '報工日期': report.work_date.strftime('%Y-%m-%d'),
            '作業員': report.worker_name,
            '工單號': report.work_order_no,
            '工序': report.process_name,
            '完成數量': report.completed_quantity or 0,
            '不良數量': report.defect_quantity or 0,
            '開始時間': report.start_time.strftime('%H:%M') if report.start_time else '',
            '結束時間': report.end_time.strftime('%H:%M') if report.end_time else '',
            '總工作時數': round(work_time_data['total_work_hours'], 2),
            '實際工作時數': round(work_time_data['actual_work_hours'], 2),
            '休息時數': round(work_time_data['break_hours'], 2),
            '加班時數': round(work_time_data['overtime_hours'], 2),
            '正常工時': round(work_time_data['regular_hours'], 2),
            '異常紀錄': report.abnormal_notes or '',
        })
    
    return data


def get_smt_report_data(date_range, custom_start=None, custom_end=None):
    """取得SMT報工資料"""
    from workorder.models import SMTProductionReport
    from reporting.calculators.time_calculator import TimeCalculator
    
    today = date.today()
    
    # 設定日期範圍
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=30)
        end_date = today
    
    # 查詢資料
    reports = SMTProductionReport.objects.filter(
        work_date__gte=start_date,
        work_date__lte=end_date
    ).order_by('work_date', 'equipment_name')
    
    # 初始化工作時數計算器
    work_time_calc = TimeCalculator()
    
    # 準備資料
    data = []
    for report in reports:
        # 使用工作時數計算器計算正確的工作時數
        work_time_data = work_time_calc.calculate_work_time_for_report(report)
        
        data.append({
            '報工日期': report.work_date.strftime('%Y-%m-%d'),
            '設備名稱': report.equipment_name,
            '製令號碼': report.work_order_no,
            '機種名稱': report.product_name,
            '完成數量': report.completed_quantity or 0,
            '不良數量': report.defect_quantity or 0,
            '總工作時數': round(work_time_data['total_work_hours'], 2),
            '實際工作時數': round(work_time_data['actual_work_hours'], 2),
            '休息時數': round(work_time_data['break_hours'], 2),
            '加班時數': round(work_time_data['overtime_hours'], 2),
            '正常工時': round(work_time_data['regular_hours'], 2),
            '異常紀錄': report.abnormal_notes or '',
        })
    
    return data


def get_abnormal_report_data(date_range, custom_start=None, custom_end=None):
    """取得異常報工資料"""
    from workorder.models import OperatorSupplementReport, SMTProductionReport
    from reporting.calculators.time_calculator import TimeCalculator
    
    today = date.today()
    
    # 設定日期範圍
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=30)
        end_date = today
    
    # 查詢作業員異常報工
    operator_reports = OperatorSupplementReport.objects.filter(
        work_date__gte=start_date,
        work_date__lte=end_date,
        abnormal_notes__isnull=False
    ).exclude(abnormal_notes='')
    
    # 查詢SMT異常報工
    smt_reports = SMTProductionReport.objects.filter(
        work_date__gte=start_date,
        work_date__lte=end_date,
        abnormal_notes__isnull=False
    ).exclude(abnormal_notes='')
    
    # 初始化工作時數計算器
    work_time_calc = TimeCalculator()
    
    # 準備資料
    data = []
    
    # 作業員異常報工
    for report in operator_reports:
        # 使用工作時數計算器計算正確的工作時數
        work_time_data = work_time_calc.calculate_work_time_for_report(report)
        
        data.append({
            '報工日期': report.work_date.strftime('%Y-%m-%d'),
            '報工類型': '作業員報工',
            '作業員/設備': report.worker_name,
            '工單號': report.work_order_no,
            '工序/機種': report.process_name,
            '總工作時數': round(work_time_data['total_work_hours'], 2),
            '實際工作時數': round(work_time_data['actual_work_hours'], 2),
            '休息時數': round(work_time_data['break_hours'], 2),
            '加班時數': round(work_time_data['overtime_hours'], 2),
            '正常工時': round(work_time_data['regular_hours'], 2),
            '異常紀錄': report.abnormal_notes,
        })
    
    # SMT異常報工
    for report in smt_reports:
        # 使用工作時數計算器計算正確的工作時數
        work_time_data = work_time_calc.calculate_work_time_for_report(report)
        
        data.append({
            '報工日期': report.work_date.strftime('%Y-%m-%d'),
            '報工類型': 'SMT報工',
            '作業員/設備': report.equipment_name,
            '工單號': report.work_order_no,
            '工序/機種': report.product_name,
            '總工作時數': round(work_time_data['total_work_hours'], 2),
            '實際工作時數': round(work_time_data['actual_work_hours'], 2),
            '休息時數': round(work_time_data['break_hours'], 2),
            '加班時數': round(work_time_data['overtime_hours'], 2),
            '正常工時': round(work_time_data['regular_hours'], 2),
            '異常紀錄': report.abnormal_notes,
        })
    
    # 按日期排序
    data.sort(key=lambda x: x['報工日期'])
    
    return data


def get_production_daily_report_data(date_range, custom_start=None, custom_end=None):
    """取得生產日報表資料"""
    from workorder.models import OperatorSupplementReport, SMTProductionReport
    from reporting.calculators.time_calculator import TimeCalculator
    
    today = date.today()
    
    # 設定日期範圍
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=30)
        end_date = today
    
    # 查詢資料
    operator_reports = OperatorSupplementReport.objects.filter(
        work_date__gte=start_date,
        work_date__lte=end_date
    )
    
    smt_reports = SMTProductionReport.objects.filter(
        work_date__gte=start_date,
        work_date__lte=end_date
    )
    
    # 初始化工作時數計算器
    work_time_calc = TimeCalculator()
    
    # 按日期分組統計
    daily_stats = {}
    
    # 統計作業員報工
    for report in operator_reports:
        date_str = report.work_date.strftime('%Y-%m-%d')
        if date_str not in daily_stats:
            daily_stats[date_str] = {
                '日期': date_str,
                '作業員報工數': 0,
                '作業員完成數量': 0,
                '作業員不良數量': 0,
                '作業員總工作時數': 0,
                '作業員實際工作時數': 0,
                '作業員休息時數': 0,
                '作業員加班時數': 0,
                'SMT報工數': 0,
                'SMT完成數量': 0,
                'SMT不良數量': 0,
                'SMT總工作時數': 0,
                'SMT實際工作時數': 0,
                'SMT休息時數': 0,
                'SMT加班時數': 0,
            }
        
        # 使用工作時數計算器計算正確的工作時數
        work_time_data = work_time_calc.calculate_work_time_for_report(report)
        
        daily_stats[date_str]['作業員報工數'] += 1
        daily_stats[date_str]['作業員完成數量'] += report.completed_quantity or 0
        daily_stats[date_str]['作業員不良數量'] += report.defect_quantity or 0
        daily_stats[date_str]['作業員總工作時數'] += work_time_data['total_work_hours']
        daily_stats[date_str]['作業員實際工作時數'] += work_time_data['actual_work_hours']
        daily_stats[date_str]['作業員休息時數'] += work_time_data['break_hours']
        daily_stats[date_str]['作業員加班時數'] += work_time_data['overtime_hours']
    
    # 統計SMT報工
    for report in smt_reports:
        date_str = report.work_date.strftime('%Y-%m-%d')
        if date_str not in daily_stats:
            daily_stats[date_str] = {
                '日期': date_str,
                '作業員報工數': 0,
                '作業員完成數量': 0,
                '作業員不良數量': 0,
                '作業員總工作時數': 0,
                '作業員實際工作時數': 0,
                '作業員休息時數': 0,
                '作業員加班時數': 0,
                'SMT報工數': 0,
                'SMT完成數量': 0,
                'SMT不良數量': 0,
                'SMT總工作時數': 0,
                'SMT實際工作時數': 0,
                'SMT休息時數': 0,
                'SMT加班時數': 0,
            }
        
        # 使用工作時數計算器計算正確的工作時數
        work_time_data = work_time_calc.calculate_work_time_for_report(report)
        
        daily_stats[date_str]['SMT報工數'] += 1
        daily_stats[date_str]['SMT完成數量'] += report.completed_quantity or 0
        daily_stats[date_str]['SMT不良數量'] += report.defect_quantity or 0
        daily_stats[date_str]['SMT總工作時數'] += work_time_data['total_work_hours']
        daily_stats[date_str]['SMT實際工作時數'] += work_time_data['actual_work_hours']
        daily_stats[date_str]['SMT休息時數'] += work_time_data['break_hours']
        daily_stats[date_str]['SMT加班時數'] += work_time_data['overtime_hours']
    
    # 轉換為列表並排序，同時四捨五入時數
    data = []
    for stats in daily_stats.values():
        stats['作業員總工作時數'] = round(stats['作業員總工作時數'], 2)
        stats['作業員實際工作時數'] = round(stats['作業員實際工作時數'], 2)
        stats['作業員休息時數'] = round(stats['作業員休息時數'], 2)
        stats['作業員加班時數'] = round(stats['作業員加班時數'], 2)
        stats['SMT總工作時數'] = round(stats['SMT總工作時數'], 2)
        stats['SMT實際工作時數'] = round(stats['SMT實際工作時數'], 2)
        stats['SMT休息時數'] = round(stats['SMT休息時數'], 2)
        stats['SMT加班時數'] = round(stats['SMT加班時數'], 2)
        data.append(stats)
    
    data.sort(key=lambda x: x['日期'])
    
    return data


def get_operator_performance_report_data(date_range, custom_start=None, custom_end=None):
    """取得作業員績效報表資料"""
    from workorder.models import OperatorSupplementReport
    from reporting.calculators.time_calculator import TimeCalculator
    
    today = date.today()
    
    # 設定日期範圍
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=30)
        end_date = today
    
    # 查詢資料
    reports = OperatorSupplementReport.objects.filter(
        work_date__gte=start_date,
        work_date__lte=end_date
    ).order_by('worker_name', 'work_date')
    
    # 初始化工作時數計算器
    work_time_calc = TimeCalculator()
    
    # 按作業員分組統計
    operator_stats = {}
    
    for report in reports:
        worker_name = report.worker_name
        if worker_name not in operator_stats:
            operator_stats[worker_name] = {
                '作業員': worker_name,
                '報工次數': 0,
                '總完成數量': 0,
                '總不良數量': 0,
                '總工作時數': 0,
                '實際工作時數': 0,
                '休息時數': 0,
                '加班時數': 0,
                '正常工時': 0,
                '平均效率': 0,
                '平均良率': 0,
            }
        
        # 使用工作時數計算器計算正確的工作時數
        work_time_data = work_time_calc.calculate_work_time_for_report(report)
        
        operator_stats[worker_name]['報工次數'] += 1
        operator_stats[worker_name]['總完成數量'] += report.completed_quantity or 0
        operator_stats[worker_name]['總不良數量'] += report.defect_quantity or 0
        operator_stats[worker_name]['總工作時數'] += work_time_data['total_work_hours']
        operator_stats[worker_name]['實際工作時數'] += work_time_data['actual_work_hours']
        operator_stats[worker_name]['休息時數'] += work_time_data['break_hours']
        operator_stats[worker_name]['加班時數'] += work_time_data['overtime_hours']
        operator_stats[worker_name]['正常工時'] += work_time_data['regular_hours']
    
    # 計算平均效率和良率
    for stats in operator_stats.values():
        if stats['實際工作時數'] > 0:
            stats['平均效率'] = round(stats['總完成數量'] / stats['實際工作時數'], 2)
        
        total_quantity = stats['總完成數量'] + stats['總不良數量']
        if total_quantity > 0:
            stats['平均良率'] = round((stats['總完成數量'] / total_quantity) * 100, 2)
        
        # 四捨五入時數
        stats['總工作時數'] = round(stats['總工作時數'], 2)
        stats['實際工作時數'] = round(stats['實際工作時數'], 2)
        stats['休息時數'] = round(stats['休息時數'], 2)
        stats['加班時數'] = round(stats['加班時數'], 2)
        stats['正常工時'] = round(stats['正常工時'], 2)
    
    # 轉換為列表並排序
    data = list(operator_stats.values())
    data.sort(key=lambda x: x['作業員'])
    
    return data


# 匯出函數
def export_to_excel(data, filename):
    """匯出為Excel檔案"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "報表資料"
    
    # 寫入表頭
    if data:
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            cell = worksheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # 寫入資料
        for row, record in enumerate(data, 2):
            for col, value in enumerate(record.values(), 1):
                worksheet.cell(row=row, column=col, value=value)
        
        # 調整欄寬
        for col in range(1, len(headers) + 1):
            worksheet.column_dimensions[get_column_letter(col)].width = 15
    
    # 建立回應
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    workbook.save(response)
    return response


def export_to_csv(data, filename):
    """匯出為CSV檔案"""
    import csv
    
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    
    # 寫入BOM以支援中文
    response.write('\ufeff')
    
    if data:
        writer = csv.DictWriter(response, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)
    
    return response


def export_to_pdf(data, filename):
    """匯出為PDF檔案（暫時重定向到Excel）"""
    # 暫時使用Excel格式，未來實作真正的PDF匯出
    return export_to_excel(data, filename) 


# 資料獲取函數
def get_work_report_data(report_type, date_range, custom_start=None, custom_end=None):
    """取得工作報表資料 - 支援日、週、月報表"""
    from reporting.models import WorkTimeReport
    from datetime import timedelta
    
    today = date.today()
    
    # 設定日期範圍
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = today - timedelta(days=1)
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'custom':
        if custom_start and custom_end:
            start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
            end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
        else:
            start_date = today
            end_date = today
    else:
        start_date = today
        end_date = today
    
    # 查詢資料
    queryset = WorkTimeReport.objects.filter(
        report_date__gte=start_date,
        report_date__lte=end_date
    ).order_by('report_date', 'worker_name')
    
    # 根據報表類型準備資料
    if report_type == 'daily':
        return _prepare_work_daily_data(queryset)
    elif report_type == 'weekly':
        return _prepare_work_weekly_data(queryset)
    elif report_type == 'monthly':
        return _prepare_work_monthly_data(queryset)
    else:
        return _prepare_work_daily_data(queryset)

def _prepare_work_daily_data(queryset):
    """準備工作日報表資料"""
    data = []
    for report in queryset:
        data.append({
            '報表日期': report.report_date.strftime('%Y-%m-%d'),
            '作業員': report.worker_name,
            '工單號': report.workorder_number,
            '產品編號': getattr(report, 'product_code', ''),
            '工序': report.process_name,
            '開始時間': report.start_time.strftime('%H:%M') if report.start_time else '',
            '結束時間': report.end_time.strftime('%H:%M') if report.end_time else '',
            '工作時數': round(report.total_work_hours or 0, 2),
            '完成數量': report.completed_quantity or 0,
            '不良品': report.defect_quantity or 0,
            '效率(件/小時)': round((report.completed_quantity or 0) / (report.total_work_hours or 1), 2) if report.total_work_hours else 0,
        })
    return data

def _prepare_work_weekly_data(queryset):
    """準備工作週報表資料 - 按週分組統計"""
    from datetime import timedelta
    
    # 按週分組統計
    weekly_data = {}
    for report in queryset:
        # 計算週開始日期
        week_start = report.report_date - timedelta(days=report.report_date.weekday())
        week_key = week_start.strftime('%Y-%m-%d')
        
        if week_key not in weekly_data:
            weekly_data[week_key] = {
                '週開始日期': week_start.strftime('%Y-%m-%d'),
                '週結束日期': (week_start + timedelta(days=6)).strftime('%Y-%m-%d'),
                '作業員': report.worker_name,
                '總工作時數': 0,
                '總完成數量': 0,
                '總不良品': 0,
                '平均效率': 0,
                '工單數量': 0,
            }
        
        weekly_data[week_key]['總工作時數'] += report.total_work_hours or 0
        weekly_data[week_key]['總完成數量'] += report.completed_quantity or 0
        weekly_data[week_key]['總不良品'] += report.defect_quantity or 0
        weekly_data[week_key]['工單數量'] += 1
    
    # 計算平均效率
    for week_data in weekly_data.values():
        if week_data['總工作時數'] > 0:
            week_data['平均效率'] = round(week_data['總完成數量'] / week_data['總工作時數'], 2)
        week_data['總工作時數'] = round(week_data['總工作時數'], 2)
    
    return list(weekly_data.values())

def _prepare_work_monthly_data(queryset):
    """準備工作月報表資料 - 按月分組統計"""
    from datetime import timedelta
    
    # 按月分組統計
    monthly_data = {}
    for report in queryset:
        # 計算月開始日期
        month_start = report.report_date.replace(day=1)
        month_key = month_start.strftime('%Y-%m')
        
        if month_key not in monthly_data:
            # 計算月末日期
            if month_start.month == 12:
                next_month = month_start.replace(year=month_start.year + 1, month=1)
            else:
                next_month = month_start.replace(month=month_start.month + 1)
            month_end = next_month - timedelta(days=1)
            
            monthly_data[month_key] = {
                '月份': month_start.strftime('%Y年%m月'),
                '月開始日期': month_start.strftime('%Y-%m-%d'),
                '月結束日期': month_end.strftime('%Y-%m-%d'),
                '作業員': report.worker_name,
                '總工作時數': 0,
                '總完成數量': 0,
                '總不良品': 0,
                '平均效率': 0,
                '工單數量': 0,
            }
        
        monthly_data[month_key]['總工作時數'] += report.total_work_hours or 0
        monthly_data[month_key]['總完成數量'] += report.completed_quantity or 0
        monthly_data[month_key]['總不良品'] += report.defect_quantity or 0
        monthly_data[month_key]['工單數量'] += 1
    
    # 計算平均效率
    for month_data in monthly_data.values():
        if month_data['總工作時數'] > 0:
            month_data['平均效率'] = round(month_data['總完成數量'] / month_data['總工作時數'], 2)
        month_data['總工作時數'] = round(month_data['總工作時數'], 2)
    
    return list(monthly_data.values())

def get_work_hour_report_data(date_range, custom_start=None, custom_end=None):
    """取得工時報表資料 - 即時計算，不回寫原始資料"""
    from workorder.models import OperatorSupplementReport
    from reporting.calculators.time_calculator import TimeCalculator
    
    today = date.today()
    
    # 設定日期範圍
    if date_range == 'today':
        start_date = today
        end_date = today
    elif date_range == 'yesterday':
        start_date = today - timedelta(days=1)
        end_date = start_date
    elif date_range == 'week':
        start_date = today - timedelta(days=today.weekday())
        end_date = today
    elif date_range == 'month':
        start_date = today.replace(day=1)
        end_date = today
    elif date_range == 'custom' and custom_start and custom_end:
        start_date = datetime.strptime(custom_start, '%Y-%m-%d').date()
        end_date = datetime.strptime(custom_end, '%Y-%m-%d').date()
    else:
        start_date = today - timedelta(days=30)
        end_date = today
    
    # 查詢原始報工資料
    reports = OperatorSupplementReport.objects.filter(
        work_date__gte=start_date,
        work_date__lte=end_date
    ).order_by('worker_name', 'work_date')
    
    # 初始化工時計算器
    time_calc = TimeCalculator()
    
    # 按作業員和日期分組統計
    work_hour_stats = {}
    
    for report in reports:
        worker_name = report.worker_name
        report_date = report.work_date
        
        # 建立唯一鍵
        key = f"{worker_name}_{report_date}"
        
        if key not in work_hour_stats:
            work_hour_stats[key] = {
                '作業員': worker_name,
                '報表日期': report_date.strftime('%Y-%m-%d'),
                '報工次數': 0,
                '總工作時數': 0.0,
                '實際工作時數': 0.0,
                '休息時數': 0.0,
                '正常工時': 0.0,
                '加班時數': 0.0,
                '完成數量': 0,
                '不良品數量': 0,
                '效率': 0.0,
                '良率': 0.0,
            }
        
        # 使用工時計算器計算扣除午休時間的工時
        if report.start_time and report.end_time:
            work_time_data = time_calc.calculate_work_time_with_breaks(
                report.start_time, 
                report.end_time,
                report.work_date
            )
            
            work_hour_stats[key]['報工次數'] += 1
            work_hour_stats[key]['總工作時數'] += work_time_data['total_work_hours']
            work_hour_stats[key]['實際工作時數'] += work_time_data['actual_work_hours']
            work_hour_stats[key]['休息時數'] += work_time_data['break_hours']
            work_hour_stats[key]['正常工時'] += work_time_data['regular_hours']
            work_hour_stats[key]['加班時數'] += work_time_data['overtime_hours']
        
        # 統計產量
        work_hour_stats[key]['完成數量'] += report.completed_quantity or 0
        work_hour_stats[key]['不良品數量'] += report.defect_quantity or 0
    
    # 計算效率和良率
    for stats in work_hour_stats.values():
        # 計算效率（件/小時）
        if stats['實際工作時數'] > 0:
            stats['效率'] = round(stats['完成數量'] / stats['實際工作時數'], 2)
        
        # 計算良率
        total_quantity = stats['完成數量'] + stats['不良品數量']
        if total_quantity > 0:
            stats['良率'] = round((stats['完成數量'] / total_quantity) * 100, 2)
        
        # 四捨五入時數
        stats['總工作時數'] = round(stats['總工作時數'], 2)
        stats['實際工作時數'] = round(stats['實際工作時數'], 2)
        stats['休息時數'] = round(stats['休息時數'], 2)
        stats['正常工時'] = round(stats['正常工時'], 2)
        stats['加班時數'] = round(stats['加班時數'], 2)
    
    # 轉換為列表並排序
    data = list(work_hour_stats.values())
    data.sort(key=lambda x: (x['作業員'], x['報表日期']))
    
    return data