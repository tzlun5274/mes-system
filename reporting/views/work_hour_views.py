# -*- coding: utf-8 -*-
"""
工時報表視圖
專門處理工時報表的顯示、查詢和匯出功能
"""

from django.views.generic import ListView, DetailView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.shortcuts import render, get_object_or_404
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum, Avg
from django.utils import timezone
from datetime import datetime, timedelta, date
import json
import logging

from reporting.models import WorkTimeReport
from reporting.calculators.time_calculator import TimeCalculator
from workorder.models import OperatorSupplementReport

logger = logging.getLogger(__name__)


def reporting_user_required(user):
    """檢查用戶是否有報表權限"""
    return user.is_superuser or user.groups.filter(name="報表使用者").exists()


class WorkHourReportListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """工時報表列表視圖"""
    model = WorkTimeReport
    template_name = 'reporting/work_hour_list.html'
    context_object_name = 'work_hour_reports'
    paginate_by = 20
    
    def test_func(self):
        """權限檢查"""
        return reporting_user_required(self.request.user)
    
    def get_queryset(self):
        """獲取查詢集"""
        queryset = WorkTimeReport.objects.filter(report_category='work_hour_report').order_by('-report_date', 'worker_name')
        
        # 搜尋功能
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(worker_name__icontains=search) |
                Q(workorder_number__icontains=search) |
                Q(product_code__icontains=search) |
                Q(process_name__icontains=search)
            )
        
        # 日期範圍篩選
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(report_date__gte=date_from)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(report_date__lte=date_to)
            except ValueError:
                pass
        
        # 作業員篩選
        worker = self.request.GET.get('worker')
        if worker:
            queryset = queryset.filter(worker_name__icontains=worker)
        
        # 工序篩選
        process = self.request.GET.get('process')
        if process:
            queryset = queryset.filter(process_name__icontains=process)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """添加上下文數據"""
        context = super().get_context_data(**kwargs)
        
        # 統計資訊
        queryset = self.get_queryset()
        context['total_workers'] = queryset.values('worker_name').distinct().count()
        context['total_work_hours'] = queryset.aggregate(Sum('actual_work_hours'))['actual_work_hours__sum'] or 0
        context['total_overtime_hours'] = queryset.aggregate(Sum('overtime_hours'))['overtime_hours__sum'] or 0
        context['total_completed_quantity'] = queryset.aggregate(Sum('completed_quantity'))['completed_quantity__sum'] or 0
        context['avg_efficiency'] = queryset.aggregate(Avg('efficiency_rate'))['efficiency_rate__avg'] or 0
        
        # 篩選選項
        context['workers'] = WorkTimeReport.objects.filter(report_category='work_hour_report').values_list('worker_name', flat=True).distinct().order_by('worker_name')
        context['processes'] = WorkTimeReport.objects.filter(report_category='work_hour_report').values_list('process_name', flat=True).distinct().order_by('process_name')
        
        # 搜尋參數
        context['search'] = self.request.GET.get('search', '')
        context['date_from'] = self.request.GET.get('date_from', '')
        context['date_to'] = self.request.GET.get('date_to', '')
        context['worker'] = self.request.GET.get('worker', '')
        context['process'] = self.request.GET.get('process', '')
        
        return context


class WorkHourReportDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """工時報表詳情視圖"""
    model = WorkTimeReport
    template_name = 'reporting/work_hour_detail.html'
    context_object_name = 'work_hour_report'
    
    def test_func(self):
        """權限檢查"""
        return reporting_user_required(self.request.user)


class WorkHourReportExportView(LoginRequiredMixin, UserPassesTestMixin, View):
    """工時報表匯出視圖"""
    
    def test_func(self):
        """權限檢查"""
        return reporting_user_required(self.request.user)
    
    def get(self, request, *args, **kwargs):
        """處理匯出請求"""
        report_id = request.GET.get('id')
        export_format = request.GET.get('format', 'excel')
        
        if report_id:
            # 匯出單一報表
            report = get_object_or_404(WorkTimeReport, id=report_id)
            data = [self._prepare_report_data(report)]
            filename = f"工時報表_{report.worker_name}_{report.report_date}"
        else:
            # 匯出查詢結果
            queryset = self._get_filtered_queryset(request)
            data = [self._prepare_report_data(report) for report in queryset]
            filename = f"工時報表_{date.today()}"
        
        if export_format == 'csv':
            return self._export_to_csv(data, filename)
        elif export_format == 'pdf':
            return self._export_to_pdf(data, filename)
        else:
            return self._export_to_excel(data, filename)
    
    def _get_filtered_queryset(self, request):
        """獲取篩選後的查詢集"""
        queryset = WorkTimeReport.objects.filter(report_category='work_hour_report').order_by('-report_date', 'worker_name')
        
        # 應用篩選條件
        search = request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(worker_name__icontains=search) |
                Q(workorder_number__icontains=search) |
                Q(product_code__icontains=search) |
                Q(process_name__icontains=search)
            )
        
        date_from = request.GET.get('date_from')
        date_to = request.GET.get('date_to')
        
        if date_from:
            try:
                date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
                queryset = queryset.filter(report_date__gte=date_from)
            except ValueError:
                pass
        
        if date_to:
            try:
                date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
                queryset = queryset.filter(report_date__lte=date_to)
            except ValueError:
                pass
        
        worker = request.GET.get('worker')
        if worker:
            queryset = queryset.filter(worker_name__icontains=worker)
        
        process = request.GET.get('process')
        if process:
            queryset = queryset.filter(process_name__icontains=process)
        
        return queryset
    
    def _prepare_report_data(self, report):
        """準備報表數據"""
        return {
            '報表日期': report.report_date.strftime('%Y-%m-%d'),
            '作業員': report.worker_name,
            '作業員編號': report.worker_id or '',
            '工單號': report.workorder_number or '',
            '產品編號': report.product_code or '',
            '工序': report.process_name or '',
            '開始時間': report.start_time.strftime('%H:%M') if report.start_time else '',
            '結束時間': report.end_time.strftime('%H:%M') if report.end_time else '',
            '總工作時數': round(report.total_work_hours, 2),
            '實際工作時數': round(report.actual_work_hours, 2),
            '休息時數': round(report.break_hours, 2),
            '正常工時': round(report.regular_hours, 2),
            '加班時數': round(report.overtime_hours, 2),
            '完成數量': report.completed_quantity,
            '不良品': report.defect_quantity,
            '效率(件/小時)': round(report.efficiency_rate, 2),
            '良率(%)': round(report.yield_rate, 2),
        }
    
    def _export_to_excel(self, data, filename):
        """匯出為Excel格式"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 建立工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "工時報表"
            
            # 設定標題
            title = f"工時報表 - {filename}"
            ws['A1'] = title
            ws.merge_cells('A1:P1')
            
            # 設定標題樣式
            title_font = Font(bold=True, size=14)
            title_alignment = Alignment(horizontal='center')
            title_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            
            ws['A1'].font = title_font
            ws['A1'].alignment = title_alignment
            ws['A1'].fill = title_fill
            
            # 設定欄位標題
            headers = [
                '報表日期', '作業員', '作業員編號', '工單號', '產品編號', '工序',
                '開始時間', '結束時間', '總工作時數', '實際工作時數', '休息時數',
                '正常工時', '加班時數', '完成數量', '不良品', '效率(件/小時)', '良率(%)'
            ]
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
                cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
            
            # 填入數據
            for row, record in enumerate(data, 4):
                for col, value in enumerate(record.values(), 1):
                    ws.cell(row=row, column=col, value=value)
            
            # 調整欄寬
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # 建立回應
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            
            wb.save(response)
            return response
            
        except ImportError:
            return HttpResponse("需要安裝 openpyxl 套件才能匯出 Excel 檔案", status=500)
    
    def _export_to_csv(self, data, filename):
        """匯出為CSV格式"""
        import csv
        
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        
        # 寫入 BOM 以支援中文
        response.write('\ufeff')
        
        writer = csv.writer(response)
        
        # 寫入標題
        if data:
            writer.writerow(data[0].keys())
        
        # 寫入數據
        for record in data:
            writer.writerow(record.values())
        
        return response
    
    def _export_to_pdf(self, data, filename):
        """匯出為PDF格式（暫時重定向到Excel）"""
        return self._export_to_excel(data, filename)


class WorkHourSyncView(LoginRequiredMixin, UserPassesTestMixin, View):
    """工時報表同步視圖"""
    
    def test_func(self):
        """權限檢查"""
        return reporting_user_required(self.request.user)
    
    def post(self, request, *args, **kwargs):
        """同步工時數據"""
        try:
            date_from = request.POST.get('date_from')
            date_to = request.POST.get('date_to')
            
            if not date_from or not date_to:
                return JsonResponse({'success': False, 'message': '請指定日期範圍'})
            
            date_from = datetime.strptime(date_from, '%Y-%m-%d').date()
            date_to = datetime.strptime(date_to, '%Y-%m-%d').date()
            
            # 同步工時數據
            sync_count = self._sync_work_hour_data(date_from, date_to)
            
            return JsonResponse({
                'success': True, 
                'message': f'成功同步 {sync_count} 筆工時數據',
                'sync_count': sync_count
            })
            
        except Exception as e:
            logger.error(f"同步工時數據失敗: {str(e)}")
            return JsonResponse({'success': False, 'message': f'同步失敗: {str(e)}'})
    
    def _sync_work_hour_data(self, date_from, date_to):
        """同步工時數據"""
        work_time_calc = TimeCalculator()
        sync_count = 0
        
        # 獲取作業員報工記錄
        operator_reports = OperatorSupplementReport.objects.filter(
            work_date__gte=date_from,
            work_date__lte=date_to,
            approval_status='approved'
        )
        
        for report in operator_reports:
            try:
                # 計算工時
                work_time_data = work_time_calc.calculate_work_time_for_report(report)
                
                # 建立或更新工時報表記錄
                work_hour_report, created = WorkTimeReport.objects.update_or_create(
                    report_type='daily',
                    report_date=report.work_date,
                    report_category='work_hour_report',
                    worker_name=report.operator.name if report.operator else '未知作業員',
                    workorder_number=report.workorder.order_number if report.workorder else '',
                    process_name=report.process_name or '',
                    defaults={
                        'report_period_start': report.work_date,
                        'report_period_end': report.work_date,
                        'data_source': 'workorder',
                        'calculation_method': 'time_calculator',
                        'created_by': 'system',
                        'worker_type': 'operator',
                        'worker_id': getattr(report.operator, 'employee_id', '') if report.operator else '',
                        'product_code': report.workorder.product_code if report.workorder else '',
                        'start_time': report.start_time,
                        'end_time': report.end_time,
                        'total_work_hours': work_time_data['total_work_hours'],
                        'actual_work_hours': work_time_data['actual_work_hours'],
                        'break_hours': work_time_data['break_hours'],
                        'regular_hours': work_time_data['regular_hours'],
                        'overtime_hours': work_time_data['overtime_hours'],
                        'completed_quantity': report.completed_quantity or 0,
                        'defect_quantity': report.defect_quantity or 0,
                    }
                )
                
                sync_count += 1
                
            except Exception as e:
                logger.error(f"同步工時數據失敗 - 記錄 {report.id}: {str(e)}")
                continue
        
        return sync_count 