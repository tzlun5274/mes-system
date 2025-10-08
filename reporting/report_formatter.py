"""
報表格式生成器
負責生成不同格式的報表檔案（Excel、HTML）
"""

import logging
import os
import pytz
from datetime import datetime
from django.utils import timezone
from django.conf import settings

logger = logging.getLogger(__name__)


class ReportFormatter:
    """報表格式生成器 - 負責生成 Excel 和 HTML 格式的報表"""
    
    def __init__(self):
        pass
    
    def generate_excel_report(self, data, report_title, schedule):
        """生成 Excel 報表"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # 確保資料是字典格式
            if not isinstance(data, dict):
                logger.error(f"Excel 報表資料格式錯誤，期望字典，實際: {type(data)}")
                return None
            
            # 建立工作簿
            wb = openpyxl.Workbook()
            
            # 移除預設工作表
            wb.remove(wb.active)
            
            # 1. 統計摘要工作表
            ws_summary = wb.create_sheet("統計摘要")
            self._create_summary_sheet(ws_summary, data, report_title, schedule)
            
            # 2. 按公司統計工作表
            if data.get('company_stats'):
                ws_company = wb.create_sheet("按公司統計")
                self._create_company_sheet(ws_company, data)
            
            # 3. 按工序統計工作表
            if data.get('process_stats'):
                ws_process = wb.create_sheet("按工序統計")
                self._create_process_sheet(ws_process, data)
            
            # 4. 按作業員統計工作表
            if data.get('operator_stats'):
                ws_operator = wb.create_sheet("按作業員統計")
                self._create_operator_sheet(ws_operator, data)
            
            # 5. 詳細資料工作表
            if data.get('detailed_data'):
                ws_detail = wb.create_sheet("詳細資料")
                self._create_detail_sheet(ws_detail, data)
            
            # 生成檔案名稱 - 使用報表實際日期而非當前時間
            filename = self._generate_filename(report_title, data, 'xlsx')
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # 確保目錄存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 儲存檔案
            wb.save(file_path)
            
            return file_path
            
        except Exception as e:
            logger.error(f"生成 Excel 報表失敗: {str(e)}")
            return None
    
    def generate_html_report(self, data, report_title, schedule):
        """生成 HTML 報表"""
        try:
            # 判斷資料格式並標準化
            normalized_data = self._normalize_data_for_html(data)
            
            # 生成統計資料
            company_stats_html = self._generate_company_statistics_table(normalized_data.get('company_stats', []))
            process_stats_html = self._generate_process_statistics_table(normalized_data.get('process_stats', []))
            operator_stats_html = self._generate_operator_statistics_table(normalized_data.get('operator_stats', []))
            detailed_data_html = self._generate_detailed_data_table(normalized_data.get('detailed_data', []))
            
            # 生成日期資訊
            date_info = self._get_report_date_info(normalized_data, report_title)
            
            # 生成統計摘要
            summary = normalized_data.get('summary', {})
            total_records = summary.get('total_records', 0)
            total_work_hours = summary.get('total_work_hours', 0)
            overtime_hours = summary.get('overtime_hours', 0)
            total_operators = summary.get('operator_count', 0)
            equipment_count = summary.get('equipment_count', 0)
            
            html_content = f"""
<!DOCTYPE html>
<html lang="zh-TW">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>{report_title}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
        .container {{ max-width: 1200px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 10px rgba(0,0,0,0.1); }}
        .header {{ text-align: center; margin-bottom: 30px; border-bottom: 2px solid #007bff; padding-bottom: 20px; }}
        .summary {{ margin-bottom: 30px; }}
        .summary-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 15px; margin: 20px 0; }}
        .summary-card {{ background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); color: white; padding: 20px; border-radius: 8px; text-align: center; }}
        .summary-card h4 {{ margin: 0 0 10px 0; font-size: 14px; opacity: 0.9; }}
        .summary-card .value {{ font-size: 24px; font-weight: bold; margin: 0; }}
        .section {{ margin: 30px 0; }}
        .section h3 {{ color: #333; border-left: 4px solid #007bff; padding-left: 15px; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; background-color: white; }}
        th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
        th {{ background-color: #f8f9fa; font-weight: bold; color: #495057; }}
        tr:nth-child(even) {{ background-color: #f8f9fa; }}
        tr:hover {{ background-color: #e3f2fd; }}
        .footer {{ margin-top: 40px; text-align: center; color: #666; border-top: 1px solid #ddd; padding-top: 20px; }}
    </style>
</head>
<body>
    <div class="container">
    <div class="header">
            <h1>📊 {report_title}</h1>
        <p>生成時間：{timezone.now().astimezone(pytz.timezone('Asia/Taipei')).strftime('%Y-%m-%d %H:%M:%S')} | 排程名稱：{schedule.name}</p>
        {date_info}
    </div>
    
    <div class="summary">
            <h3>📈 統計摘要</h3>
            <div class="summary-grid">
                <div class="summary-card">
                    <h4>總記錄數</h4>
                    <div class="value">{total_records}</div>
        </div>
                <div class="summary-card">
                    <h4>正常時數</h4>
                    <div class="value">{total_work_hours:.2f} 小時</div>
        </div>
                <div class="summary-card">
                    <h4>加班時數</h4>
                    <div class="value">{overtime_hours:.2f} 小時</div>
                </div>
                <div class="summary-card">
                    <h4>總工作時數</h4>
                    <div class="value">{(total_work_hours + overtime_hours):.2f} 小時</div>
        </div>
                <div class="summary-card">
                    <h4>參與作業員數</h4>
                    <div class="value">{total_operators} 人</div>
        </div>
                <div class="summary-card">
                    <h4>使用設備數</h4>
                    <div class="value">{equipment_count} 台</div>
                </div>
                <div class="summary-card">
                    <h4>平均每日工作時數</h4>
                    <div class="value">{(total_work_hours/7):.2f} 小時</div>
                </div>
        </div>
    </div>
        
        <!-- 按公司統計 -->
        {company_stats_html}
        
        <!-- 按工序統計 -->
        {process_stats_html}
        
        <!-- 按作業員統計 -->
        {operator_stats_html}
        
        <!-- 詳細資料 -->
        {detailed_data_html}
    
    <div class="footer">
            <p>此報表由 MES 系統自動生成 | 生成時間：{timezone.now().astimezone(pytz.timezone('Asia/Taipei')).strftime('%Y-%m-%d %H:%M:%S')}</p>
        </div>
    </div>
</body>
</html>
        """
        
            # 生成檔案名稱 - 使用報表實際日期而非當前時間
            filename = self._generate_filename(report_title, data, 'html')
            file_path = os.path.join(settings.MEDIA_ROOT, 'reports', filename)
            
            # 確保目錄存在
            os.makedirs(os.path.dirname(file_path), exist_ok=True)
            
            # 寫入檔案
            with open(file_path, 'w', encoding='utf-8') as f:
                f.write(html_content)
            
            return file_path
            
        except Exception as e:
            logger.error(f"生成 HTML 報表失敗: {str(e)}")
            return None
    
    def _normalize_data_for_html(self, data):
        """標準化資料格式用於 HTML 報表"""
        if isinstance(data, dict) and 'summary' in data:
            # 完整的資料結構 - 直接返回
            return data
        elif isinstance(data, dict) and 'fill_works_count' in data:
            # 前一個工作日報表格式
            return {
                'summary': {
                    'total_records': data.get('fill_works_count', 0),
                    'total_work_hours': data.get('total_work_hours', 0),
                    'overtime_hours': data.get('total_overtime_hours', 0),
                    'equipment_count': data.get('total_equipment', 0),
                    'operator_count': data.get('total_operators', 0)
                },
                'detailed_data': data.get('report_data', []),
                'company_stats': data.get('company_stats', []),
                'process_stats': data.get('process_stats', []),
                'operator_stats': data.get('operator_stats', [])
            }
        elif isinstance(data, dict) and 'total_records' in data:
            # 週報表格式
            return {
                'summary': {
                    'total_records': data.get('total_records', 0),
                    'total_work_hours': data.get('total_work_hours', 0),
                    'overtime_hours': data.get('total_overtime_hours', 0),
                    'equipment_count': data.get('total_equipment', 0),
                    'operator_count': data.get('total_operators', 0)
                },
                'detailed_data': data.get('report_data', []),
                'company_stats': data.get('company_stats', []),
                'process_stats': data.get('process_stats', []),
                'operator_stats': data.get('operator_stats', [])
            }
        else:
            # 預設格式
            return data
    
    def _get_report_date_info(self, data, report_title):
        """取得報表日期資訊"""
        if isinstance(data, dict) and 'start_date' in data and 'end_date' in data:
            # 週報表格式
            start_date = data.get('start_date')
            end_date = data.get('end_date')
            if start_date and end_date:
                return f"<p>報表期間：{start_date} 到 {end_date}</p>"
        elif isinstance(data, dict) and 'report_date' in data:
            # 前一個工作日報表格式
            report_date = data.get('report_date')
            if report_date:
                return f"<p>報表日期：{report_date}</p>"
        
        return ""
    
    def _generate_company_statistics_table(self, company_stats):
        """生成按公司統計表格"""
        if not company_stats:
            return """
        <div class="section">
            <h3>🏢 按公司統計</h3>
            <p>無資料</p>
        </div>
        """
        
        table_html = """
        <div class="section">
            <h3>🏢 按公司統計</h3>
            <table>
                <thead>
                    <tr>
                        <th>公司名稱</th>
                        <th>記錄數</th>
                        <th>正常時數</th>
                        <th>加班時數</th>
                        <th>總時數</th>
                        <th>作業員數</th>
                        <th>設備數</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for stat in company_stats:
            total_hours = stat['normal_hours'] + stat['overtime_hours']
            table_html += f"""
                    <tr>
                        <td>{stat['company_name']}</td>
                        <td>{stat['record_count']}</td>
                        <td>{stat['normal_hours']:.2f} 小時</td>
                        <td>{stat['overtime_hours']:.2f} 小時</td>
                        <td>{total_hours:.2f} 小時</td>
                        <td>{stat['operator_count']}</td>
                        <td>{stat['equipment_count']}</td>
                    </tr>
            """
        
        table_html += """
                </tbody>
            </table>
        </div>
        """
        
        return table_html
    
    def _generate_process_statistics_table(self, process_stats):
        """生成按工序統計表格"""
        if not process_stats:
            return """
        <div class="section">
            <h3>⚙️ 按工序統計</h3>
            <p>無資料</p>
        </div>
        """
        
        table_html = """
        <div class="section">
            <h3>⚙️ 按工序統計</h3>
            <table>
                <thead>
                    <tr>
                        <th>工序名稱</th>
                        <th>記錄數</th>
                        <th>正常時數</th>
                        <th>加班時數</th>
                        <th>總時數</th>
                        <th>工作數量</th>
                        <th>不良品數量</th>
                        <th>平均效率</th>
                        <th>作業員數</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for stat in process_stats:
            total_hours = stat['normal_hours'] + stat['overtime_hours']
            efficiency = stat.get('efficiency', 0)
            table_html += f"""
                    <tr>
                        <td>{stat['process_name']}</td>
                        <td>{stat['record_count']}</td>
                        <td>{stat['normal_hours']:.2f} 小時</td>
                        <td>{stat['overtime_hours']:.2f} 小時</td>
                        <td>{total_hours:.2f} 小時</td>
                        <td>{stat.get('work_quantity', 0)}</td>
                        <td>{stat.get('defect_quantity', 0)}</td>
                        <td>{efficiency:.2f} 件/小時</td>
                        <td>{stat.get('operator_count', 0)}</td>
                    </tr>
            """
        
        table_html += """
                </tbody>
            </table>
        </div>
        """
        
        return table_html
    
    def _generate_operator_statistics_table(self, operator_stats):
        """生成按作業員統計表格"""
        if not operator_stats:
            return """
        <div class="section">
            <h3>👥 按作業員統計</h3>
            <p>無資料</p>
        </div>
        """
        
        table_html = """
        <div class="section">
            <h3>👥 按作業員統計</h3>
            <table>
                <thead>
                    <tr>
                        <th>作業員</th>
                        <th>記錄數</th>
                        <th>正常時數</th>
                        <th>加班時數</th>
                        <th>總時數</th>
                        <th>工作數量</th>
                        <th>不良品數量</th>
                        <th>平均效率</th>
                        <th>設備數</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for stat in operator_stats:
            total_hours = stat['normal_hours'] + stat['overtime_hours']
            efficiency = stat.get('efficiency', 0)
            table_html += f"""
                    <tr>
                        <td>{stat['operator_name']}</td>
                        <td>{stat['record_count']}</td>
                        <td>{stat['normal_hours']:.2f} 小時</td>
                        <td>{stat['overtime_hours']:.2f} 小時</td>
                        <td>{total_hours:.2f} 小時</td>
                        <td>{stat.get('work_quantity', 0)}</td>
                        <td>{stat.get('defect_quantity', 0)}</td>
                        <td>{efficiency:.2f} 件/小時</td>
                        <td>{stat.get('equipment_count', 0)}</td>
                    </tr>
            """
        
        table_html += """
                </tbody>
            </table>
        </div>
        """
        
        return table_html
    
    def _generate_detailed_data_table(self, detailed_data):
        """生成詳細資料表格"""
        if not detailed_data:
            return "<div class='section'><h3>📋 詳細資料</h3><p>無資料</p></div>"
        
        table_html = """
        <div class="section">
            <h3>📋 詳細資料</h3>
            <table>
                <thead>
                    <tr>
                        <th>公司名稱</th>
                        <th>作業員</th>
                        <th>工單編號</th>
                        <th>產品編號</th>
                        <th>工序名稱</th>
                        <th>工作日期</th>
                        <th>開始時間</th>
                        <th>結束時間</th>
                        <th>工作時數</th>
                        <th>加班時數</th>
                        <th>工作數量</th>
                        <th>不良品數量</th>
                    </tr>
                </thead>
                <tbody>
        """
        
        for record in detailed_data:
            table_html += f"""
                    <tr>
                        <td>{record.get('company_name', '')}</td>
                        <td>{record.get('operator_name', '')}</td>
                        <td>{record.get('workorder_id', '')}</td>
                        <td>{record.get('product_code', '')}</td>
                        <td>{record.get('process_name', '')}</td>
                        <td>{record.get('work_date', '')}</td>
                        <td>{record.get('start_time', '')}</td>
                        <td>{record.get('end_time', '')}</td>
                        <td>{record.get('work_hours', 0):.2f}</td>
                        <td>{record.get('overtime_hours', 0):.2f}</td>
                        <td>{record.get('work_quantity', 0):.0f}</td>
                        <td>{record.get('defect_quantity', 0):.0f}</td>
                    </tr>
            """
        
        table_html += """
                </tbody>
            </table>
        </div>
        """
        
        return table_html
    
    def _generate_filename(self, report_title, data, file_extension):
        """生成檔案名稱，使用報表實際日期而非當前時間"""
        try:
            # 嘗試從資料中取得報表日期
            report_date = None
            
            # 檢查是否有報表日期資訊
            if isinstance(data, dict):
                if 'report_date' in data:
                    report_date = data['report_date']
                elif 'start_date' in data:
                    report_date = data['start_date']
                elif 'end_date' in data:
                    report_date = data['end_date']
            
            # 如果沒有找到報表日期，使用當前時間
            if not report_date:
                taiwan_tz = pytz.timezone('Asia/Taipei')
                local_time = timezone.now().astimezone(taiwan_tz)
                date_str = local_time.strftime('%Y%m%d_%H%M%S')
            else:
                # 使用報表實際日期
                if hasattr(report_date, 'strftime'):
                    # 如果是日期物件
                    date_str = report_date.strftime('%Y%m%d')
                else:
                    # 如果是字串，嘗試解析
                    try:
                        from datetime import datetime
                        parsed_date = datetime.strptime(str(report_date), '%Y-%m-%d')
                        date_str = parsed_date.strftime('%Y%m%d')
                    except:
                        # 解析失敗，使用當前時間
                        taiwan_tz = pytz.timezone('Asia/Taipei')
                        local_time = timezone.now().astimezone(taiwan_tz)
                        date_str = local_time.strftime('%Y%m%d_%H%M%S')
            
            # 生成檔案名稱
            safe_title = report_title.replace(' ', '_').replace('(', '').replace(')', '')
            filename = f"{safe_title}_{date_str}.{file_extension}"
            
            return filename
            
        except Exception as e:
            logger.error(f"生成檔案名稱失敗: {str(e)}")
            # 發生錯誤時使用當前時間作為備案
            taiwan_tz = pytz.timezone('Asia/Taipei')
            local_time = timezone.now().astimezone(taiwan_tz)
            safe_title = report_title.replace(' ', '_').replace('(', '').replace(')', '')
            return f"{safe_title}_{local_time.strftime('%Y%m%d_%H%M%S')}.{file_extension}"
    
    # Excel 報表的工作表創建方法
    def _create_summary_sheet(self, ws, data, report_title, schedule):
        """創建統計摘要工作表"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        
        # 設定樣式
        title_font = Font(name='微軟正黑體', size=16, bold=True, color='FFFFFF')
        title_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        data_font = Font(name='微軟正黑體', size=11)
        
        # 設定邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        current_row = 1
        
        # 標題
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = report_title
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].fill = title_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # 生成時間
        ws.merge_cells(f'A{current_row}:H{current_row}')
        taiwan_tz = pytz.timezone('Asia/Taipei')
        local_time = timezone.now().astimezone(taiwan_tz)
        ws[f'A{current_row}'] = f"生成時間：{local_time.strftime('%Y-%m-%d %H:%M:%S')}"
        ws[f'A{current_row}'].font = data_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # 排程名稱
        ws.merge_cells(f'A{current_row}:H{current_row}')
        schedule_name = getattr(schedule, 'name', '手動執行報表')
        ws[f'A{current_row}'] = f"排程名稱：{schedule_name}"
        ws[f'A{current_row}'].font = data_font
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        current_row += 2
        
        # 統計摘要標題
        ws.merge_cells(f'A{current_row}:H{current_row}')
        ws[f'A{current_row}'] = "📈 統計摘要"
        ws[f'A{current_row}'].font = header_font
        ws[f'A{current_row}'].fill = header_fill
        ws[f'A{current_row}'].alignment = Alignment(horizontal='center', vertical='center')
        current_row += 1
        
        # 提取統計資料
        summary = data.get('summary', {})
        total_records = summary.get('total_records', 0)
        total_work_hours = summary.get('total_work_hours', 0)
        total_overtime_hours = summary.get('overtime_hours', 0)
        total_equipment_hours = summary.get('equipment_count', 0)
        total_operators = summary.get('operator_count', 0)
        
        # 寫入統計資料
        stats_data = [
            ['總記錄數', total_records],
            ['正常時數', f"{total_work_hours:.2f} 小時"],
            ['加班時數', f"{total_overtime_hours:.2f} 小時"],
            ['總工作時數', f"{total_work_hours:.2f} 小時"],
            ['參與作業員數', f"{total_operators} 人"],
            ['使用設備數', f"{total_equipment_hours} 台"],
            ['平均每日工作時數', f"{(total_work_hours/7):.2f} 小時" if total_work_hours > 0 else "0 小時"]
        ]
        
        for stat in stats_data:
            ws[f'A{current_row}'] = stat[0]
            ws[f'B{current_row}'] = stat[1]
            ws[f'A{current_row}'].font = data_font
            ws[f'B{current_row}'].font = data_font
            ws[f'A{current_row}'].border = thin_border
            ws[f'B{current_row}'].border = thin_border
            current_row += 1
        
        # 設定欄寬
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
    
    def _create_company_sheet(self, ws, data):
        """創建按公司統計工作表"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 設定樣式
        header_font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        data_font = Font(name='微軟正黑體', size=11)
        
        # 設定邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 標題
        ws['A1'] = "🏢 按公司統計"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:H1')
        
        # 表頭
        headers = ['公司名稱', '記錄數', '正常時數', '加班時數', '總時數', '作業員數', '設備數']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 資料
        current_row = 3
        company_stats = data.get('company_stats', [])
        for company_stat in company_stats:
            if isinstance(company_stat, dict):
                row_data = [
                    company_stat.get('company_name', ''),
                    company_stat.get('record_count', 0),
                    f"{company_stat.get('normal_hours', 0):.2f}",
                    f"{company_stat.get('overtime_hours', 0):.2f}",
                    f"{company_stat.get('total_hours', 0):.2f}",
                    company_stat.get('operator_count', 0),
                    company_stat.get('equipment_count', 0)
                ]
            else:
                logger.warning(f"跳過非字典類型的公司統計記錄: {type(company_stat)} - {company_stat}")
                continue
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
        
        # 設定欄寬
        column_widths = [15, 10, 12, 12, 12, 10, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_process_sheet(self, ws, data):
        """創建按工序統計工作表"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 設定樣式
        header_font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        data_font = Font(name='微軟正黑體', size=11)
        
        # 設定邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 標題
        ws['A1'] = "⚙️ 按工序統計"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:H1')
        
        # 表頭
        headers = ['工序名稱', '記錄數', '正常時數', '加班時數', '總時數', '工作數量', '不良品數量', '平均效率', '作業員數']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 資料
        current_row = 3
        process_stats = data.get('process_stats', [])
        for process_stat in process_stats:
            if isinstance(process_stat, dict):
                row_data = [
                    process_stat.get('process_name', ''),
                    process_stat.get('record_count', 0),
                    f"{process_stat.get('normal_hours', 0):.2f}",
                    f"{process_stat.get('overtime_hours', 0):.2f}",
                    f"{process_stat.get('total_hours', 0):.2f}",
                    process_stat.get('work_quantity', 0),
                    process_stat.get('defect_quantity', 0),
                    f"{process_stat.get('efficiency', 0):.2f}",
                    process_stat.get('operator_count', 0)
                ]
            else:
                logger.warning(f"跳過非字典類型的工序統計記錄: {type(process_stat)} - {process_stat}")
                continue
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
        
        # 設定欄寬
        column_widths = [15, 10, 12, 12, 12, 12, 12, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_operator_sheet(self, ws, data):
        """創建按作業員統計工作表"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 設定樣式
        header_font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        data_font = Font(name='微軟正黑體', size=11)
        
        # 設定邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 標題
        ws['A1'] = "👥 按作業員統計"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:H1')
        
        # 表頭
        headers = ['作業員', '記錄數', '正常時數', '加班時數', '總時數', '工作數量', '不良品數量', '平均效率', '設備數']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 資料
        current_row = 3
        operator_stats = data.get('operator_stats', [])
        for operator_stat in operator_stats:
            if isinstance(operator_stat, dict):
                row_data = [
                    operator_stat.get('operator_name', ''),
                    operator_stat.get('record_count', 0),
                    f"{operator_stat.get('normal_hours', 0):.2f}",
                    f"{operator_stat.get('overtime_hours', 0):.2f}",
                    f"{operator_stat.get('total_hours', 0):.2f}",
                    operator_stat.get('work_quantity', 0),
                    operator_stat.get('defect_quantity', 0),
                    f"{operator_stat.get('efficiency', 0):.2f}",
                    operator_stat.get('equipment_count', 0)
                ]
            else:
                logger.warning(f"跳過非字典類型的作業員統計記錄: {type(operator_stat)} - {operator_stat}")
                continue
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
        
        # 設定欄寬
        column_widths = [15, 10, 12, 12, 12, 12, 12, 10]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
    
    def _create_detail_sheet(self, ws, data):
        """創建詳細資料工作表"""
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # 設定樣式
        header_font = Font(name='微軟正黑體', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        data_font = Font(name='微軟正黑體', size=11)
        
        # 設定邊框
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 標題
        ws['A1'] = "📋 詳細資料"
        ws['A1'].font = header_font
        ws['A1'].fill = header_fill
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:L1')
        
        # 表頭
        headers = ['公司名稱', '作業員', '工單編號', '產品編號', '工序名稱', '工作日期', '開始時間', '結束時間', '工作時數', '加班時數', '工作數量', '不良品數量']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # 資料
        current_row = 3
        detailed_data = data.get('detailed_data', [])
        for record in detailed_data:
            if isinstance(record, dict):
                # 安全處理時間欄位
                start_time = record.get('start_time', '')
                end_time = record.get('end_time', '')
                
                # 如果是字串，直接使用；如果是時間物件，格式化
                if hasattr(start_time, 'strftime'):
                    start_time_str = start_time.strftime('%H:%M')
                else:
                    start_time_str = str(start_time) if start_time else ''
                    
                if hasattr(end_time, 'strftime'):
                    end_time_str = end_time.strftime('%H:%M')
                else:
                    end_time_str = str(end_time) if end_time else ''
                
                row_data = [
                    record.get('company_name', ''),
                    record.get('operator_name', ''),
                    record.get('workorder_id', ''),
                    record.get('product_code', ''),
                    record.get('process_name', ''),
                    record.get('work_date', ''),
                    start_time_str,
                    end_time_str,
                    f"{record.get('work_hours', 0):.2f}",
                    f"{record.get('overtime_hours', 0):.2f}",
                    record.get('work_quantity', 0),
                    record.get('defect_quantity', 0)
                ]
            else:
                logger.warning(f"跳過非字典類型的詳細記錄: {type(record)} - {record}")
                continue
            
            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.font = data_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            current_row += 1
        
        # 設定欄寬
        column_widths = [15, 12, 12, 12, 12, 10, 10, 12, 12, 12]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width
