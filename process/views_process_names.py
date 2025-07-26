from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse
from .models import ProcessName, ProcessEquipment
from .utils import log_user_operation
import requests
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
import json
import logging
from equip.models import Equipment

logger = logging.getLogger(__name__)

EQUIPMENT_API = "http://localhost:8000/equip/api/equipments/"


def get_equipment_options(request):
    """取得設備選項"""
    equipments = []

    try:
        # 取得一般設備
        equipments = list(Equipment.objects.values("id", "name"))

    except Exception as e:
        print(f"取得設備選項時發生錯誤: {e}")

    return equipments


def process_user_required(user):
    return user.is_superuser or user.groups.filter(name="工序使用者").exists()


def superuser_required(user):
    return user.is_superuser


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def process_names(request):
    log_user_operation(request.user.username, "process", "查看工序工藝名稱設定")
    process_names = ProcessName.objects.all().prefetch_related("equipments")
    equipments = get_equipment_options(request)
    equip_id_to_name = {
        str(e["id"]): e["name"] for e in equipments if "id" in e and "name" in e
    }
    enhanced_process_names = []
    for process in process_names:
        equip_ids = [str(e.equipment_id) for e in process.equipments.all()]
        equip_names = [
            equip_id_to_name.get(eid, f"未知設備({eid})") for eid in equip_ids
        ]
        enhanced_process_names.append(
            {
                "id": process.id,
                "name": process.name,
                "description": process.description,
                "usable_equipment_names": (
                    ", ".join(equip_names) if equip_names else "-"
                ),
            }
        )
    # 統計數字
    total_count = len(enhanced_process_names)
    with_equipment_count = sum(
        1 for p in enhanced_process_names if p["usable_equipment_names"] != "-"
    )
    with_description_count = sum(1 for p in enhanced_process_names if p["description"])
    return render(
        request,
        "process/process_names.html",
        {
            "process_names": enhanced_process_names,
            "total_count": total_count,
            "with_equipment_count": with_equipment_count,
            "with_description_count": with_description_count,
        },
    )


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def add_process_name(request):
    log_user_operation(request.user.username, "process", "嘗試添加工序工藝名稱")
    equipments = get_equipment_options(request)
    if request.method == "POST":
        name = request.POST.get("name")
        description = request.POST.get("description", "")
        if ProcessName.objects.filter(name=name).exists():
            messages.error(request, f"工序名稱 '{name}' 已存在")
            return render(
                request,
                "process/add_process_name.html",
                {"equipments": equipments, "name": name, "description": description},
            )
        process_name = ProcessName.objects.create(name=name, description=description)
        equip_ids = request.POST.getlist("usable_equipment_ids[]")
        for equip_id in equip_ids:
            if equip_id.strip().isdigit():
                ProcessEquipment.objects.get_or_create(
                    process_name=process_name, equipment_id=int(equip_id)
                )
        log_user_operation(
            request.user.username, "process", f"成功添加工序工藝名稱: {process_name}"
        )
        messages.success(request, "工序工藝名稱添加成功！")
        return redirect("process:process_names")
    return render(request, "process/add_process_name.html", {"equipments": equipments})


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def edit_process_name(request, name_id):
    log_user_operation(
        request.user.username, "process", f"嘗試編輯工序工藝名稱: {name_id}"
    )
    process_name = get_object_or_404(ProcessName, id=name_id)
    equipments = get_equipment_options(request)
    # 預處理設備 ID 列表
    normal_equip_ids = list(
        process_name.equipments.values_list("equipment_id", flat=True)
    )
    if request.method == "POST":
        new_name = request.POST.get("name")
        description = request.POST.get("description", "")
        if (
            new_name != process_name.name
            and ProcessName.objects.filter(name=new_name).exists()
        ):
            messages.error(request, f"工序名稱 '{new_name}' 已存在")
            return render(
                request,
                "process/edit_process_name.html",
                {
                    "process_name": process_name,
                    "equipments": equipments,
                    "normal_equip_ids": normal_equip_ids,
                },
            )
        process_name.name = new_name
        process_name.description = description
        process_name.save()
        process_name.equipments.all().delete()
        equip_ids = request.POST.getlist("usable_equipment_ids[]")
        for equip_id in equip_ids:
            if equip_id.strip().isdigit():
                ProcessEquipment.objects.get_or_create(
                    process_name=process_name, equipment_id=int(equip_id)
                )
        log_user_operation(
            request.user.username, "process", f"成功編輯工序工藝名稱: {process_name}"
        )
        messages.success(request, "工序工藝名稱更新成功！")
        return redirect("process:process_names")
    return render(
        request,
        "process/edit_process_name.html",
        {
            "process_name": process_name,
            "equipments": equipments,
            "normal_equip_ids": normal_equip_ids,
        },
    )


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def delete_process_name(request, name_id):
    log_user_operation(
        request.user.username, "process", f"嘗試刪除工序工藝名稱: {name_id}"
    )
    process_name = get_object_or_404(ProcessName, id=name_id)
    process_name.delete()
    log_user_operation(
        request.user.username, "process", f"成功刪除工序工藝名稱: {name_id}"
    )
    messages.success(request, "工序工藝名稱刪除成功！")
    return redirect("process:process_names")


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def export_process_names(request):
    log_user_operation(request.user.username, "process", "匯出工序工藝名稱數據")
    equipments = get_equipment_options(request)
    equip_id_to_name = {
        str(e["id"]): e["name"] for e in equipments if "id" in e and "name" in e
    }
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Process Names"
    headers = ["工序名稱", "描述", "能使用的設備"]
    for col_num, header in enumerate(headers, 1):
        worksheet[f"{get_column_letter(col_num)}1"] = header
    process_names = ProcessName.objects.all().prefetch_related("equipments")
    row_num = 2
    for process in process_names:
        equip_ids = [str(e.equipment_id) for e in process.equipments.all()]
        equip_names = [
            equip_id_to_name.get(eid, f"未知設備({eid})") for eid in equip_ids
        ]
        worksheet[f"A{row_num}"] = process.name
        worksheet[f"B{row_num}"] = process.description
        worksheet[f"C{row_num}"] = ", ".join(equip_names) if equip_names else ""
        row_num += 1
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="工序名稱.xlsx"'
    return response


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def import_process_names(request):
    log_user_operation(request.user.username, "process", "嘗試匯入工序工藝名稱數據")
    equipments = get_equipment_options(request)
    equip_name_to_id = {
        e["name"]: str(e["id"]) for e in equipments if "id" in e and "name" in e
    }
    if request.method == "POST":
        if "action" in request.POST:
            action = request.POST.get("action")
            dataset_json = request.POST.get("dataset")
            dataset = json.loads(dataset_json)
            if action == "overwrite":
                errors = []
                for row in dataset:
                    name = row["工序名稱"]
                    description = row.get("描述", "")
                    equip_names = (
                        row.get("能使用的設備", "").split(",")
                        if row.get("能使用的設備")
                        else []
                    )
                    process_name, created = ProcessName.objects.get_or_create(name=name)
                    process_name.description = description
                    process_name.save()
                    process_name.equipments.all().delete()
                    for equip_name in equip_names:
                        equip_id = equip_name_to_id.get(equip_name.strip())
                        if equip_id and equip_id.isdigit():
                            ProcessEquipment.objects.get_or_create(
                                process_name=process_name, equipment_id=int(equip_id)
                            )
                log_user_operation(
                    request.user.username, "process", "匯入工序工藝名稱數據（覆蓋模式）"
                )
                if errors:
                    for error in errors[:5]:
                        messages.warning(request, error)
                    if len(errors) > 5:
                        messages.warning(
                            request, f"還有 {len(errors) - 5} 個警告未顯示"
                        )
                messages.success(request, "工序工藝名稱數據匯入成功！")
                return redirect("process:process_names")
            else:
                messages.info(request, "匯入操作已取消")
                return redirect("process:process_names")
        if "file" in request.FILES:
            file = request.FILES["file"]
            if not file.name.endswith(".xlsx"):
                messages.error(request, "請上傳 Excel 文件 (.xlsx 格式)")
                return redirect("process:process_names")
            try:
                workbook = openpyxl.load_workbook(file)
                worksheet = workbook.active
                required_headers = ["工序名稱", "描述", "能使用的設備"]
                actual_headers = [cell.value for cell in worksheet[1]]
                if not all(header in actual_headers for header in required_headers):
                    messages.error(
                        request, f"Excel 文件缺少必要欄位: {required_headers}"
                    )
                    return redirect("process:process_names")
                dataset = []
                for row in worksheet.iter_rows(min_row=2, values_only=True):
                    if any(row):
                        dataset.append(
                            {
                                "工序名稱": row[actual_headers.index("工序名稱")],
                                "描述": (
                                    row[actual_headers.index("描述")]
                                    if row[actual_headers.index("描述")]
                                    else ""
                                ),
                                "能使用的設備": (
                                    row[actual_headers.index("能使用的設備")]
                                    if row[actual_headers.index("能使用的設備")]
                                    else ""
                                ),
                            }
                        )
                existing_names = set(ProcessName.objects.values_list("name", flat=True))
                import_names = set(row["工序名稱"] for row in dataset)
                duplicate_names = list(import_names & existing_names)
                if duplicate_names:
                    return render(
                        request,
                        "process/import_process_names.html",
                        {
                            "duplicate_names": duplicate_names,
                            "dataset": json.dumps(dataset),
                        },
                    )
                for row in dataset:
                    process_name = ProcessName.objects.create(
                        name=row["工序名稱"], description=row["描述"]
                    )
                    for equip_name in row["能使用的設備"].split(","):
                        equip_id = equip_name_to_id.get(equip_name.strip())
                        if equip_id and equip_id.isdigit():
                            ProcessEquipment.objects.get_or_create(
                                process_name=process_name, equipment_id=int(equip_id)
                            )
                log_user_operation(
                    request.user.username, "process", "匯入工序工藝名稱數據"
                )
                messages.success(request, "工序工藝名稱數據匯入成功！")
                return redirect("process:process_names")
            except Exception as e:
                messages.error(request, f"匯入失敗: {str(e)}")
                return redirect("process:process_names")
    return render(request, "process/import_process_names.html", {})
