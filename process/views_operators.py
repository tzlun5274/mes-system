from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse
from .models import Operator, OperatorSkill, ProcessName
from .utils import log_user_operation
from .services import OperatorService, OperatorStatisticsService, OperatorImportExportService
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
import json
from production.models import ProductionLine


# 檢查用戶是否屬於「工序使用者」群組，或者是超級用戶
def process_user_required(user):
    return user.is_superuser or user.groups.filter(name="工序使用者").exists()


# 檢查是否為超級管理員
def superuser_required(user):
    return user.is_superuser


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def operators(request):
    """作業員與技能管理頁面"""
    log_user_operation(request.user.username, "process", "查看作業員與技能設定")
    
    operators = Operator.objects.all()
    statistics = OperatorStatisticsService.get_operator_statistics()
    
    return render(
        request,
        "process/operators.html",
        {
            "operators": operators,
            **statistics,
        },
    )


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def add_operator(request):
    """新增作業員頁面"""
    log_user_operation(request.user.username, "process", "嘗試添加作業員")
    
    if request.method == "POST":
        name = request.POST.get("name")
        production_line_id = request.POST.get("production_line")
        process_name_ids = request.POST.getlist("process_name[]")
        priorities = request.POST.getlist("priority[]")
        
        # 使用服務層建立作業員
        operator, error_message = OperatorService.create_operator_with_skills(
            name, production_line_id, process_name_ids, priorities
        )
        
        if error_message:
            messages.error(request, error_message)
            process_names = ProcessName.objects.all()
            production_lines = ProductionLine.objects.filter(is_active=True)
            return render(
                request,
                "process/add_operator.html",
                {
                    "name": name,
                    "process_names": process_names,
                    "production_lines": production_lines,
                },
            )
        
        log_user_operation(
            request.user.username, "process", f"成功添加作業員: {operator}"
        )
        messages.success(request, "作業員添加成功！")
        return redirect("process:operators")
    
    process_names = ProcessName.objects.all()
    production_lines = ProductionLine.objects.filter(is_active=True)
    return render(
        request,
        "process/add_operator.html",
        {
            "process_names": process_names,
            "production_lines": production_lines,
        },
    )


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def edit_operator(request, operator_id):
    log_user_operation(
        request.user.username, "process", f"嘗試編輯作業員: {operator_id}"
    )
    operator = get_object_or_404(Operator, id=operator_id)
    process_names = ProcessName.objects.all()
    production_lines = ProductionLine.objects.filter(is_active=True)
    if request.method == "POST":
        new_name = request.POST.get("name")
        production_line_id = request.POST.get("production_line")
        process_name_ids = request.POST.getlist("process_name[]")
        priorities = request.POST.getlist("priority[]")
        skill_ids = request.POST.getlist("skill_id[]")
        if (
            new_name != operator.name
            and Operator.objects.filter(name=new_name).exists()
        ):
            messages.error(request, f"作業員名稱 '{new_name}' 已存在，請選擇其他名稱！")
            return render(
                request,
                "process/edit_operator.html",
                {
                    "operator": operator,
                    "process_names": process_names,
                    "production_lines": production_lines,
                },
            )
        operator.name = new_name
        
        # 透過API查詢產線資訊
        production_line_name = ""
        if production_line_id:
            try:
                production_line = ProductionLine.objects.filter(id=production_line_id).first()
                if production_line:
                    production_line_name = production_line.line_name
            except:
                pass
        
        operator.production_line_id = production_line_id
        operator.production_line_name = production_line_name
        operator.save()
        if len(process_name_ids) != len(priorities):
            messages.error(request, "工序名稱和優先順序數量不匹配！")
            return render(
                request,
                "process/edit_operator.html",
                {
                    "operator": operator,
                    "process_names": process_names,
                    "production_lines": production_lines,
                },
            )
        submitted_skill_ids = [int(sid) for sid in skill_ids if sid]
        existing_skills = OperatorSkill.objects.filter(operator_id=str(operator.id)).exclude(
            id__in=submitted_skill_ids
        )
        for skill in existing_skills:
            skill.delete()
        for i, (process_name_id, priority) in enumerate(
            zip(process_name_ids, priorities)
        ):
            if process_name_id and priority:
                try:
                    priority = int(priority)
                    if priority < 1:
                        messages.error(request, "技能優先順序必須是正整數！")
                        return render(
                            request,
                            "process/edit_operator.html",
                            {
                                "operator": operator,
                                "process_names": process_names,
                                "production_lines": production_lines,
                            },
                        )
                    process_name = ProcessName.objects.get(id=process_name_id)
                    if i < len(skill_ids) and skill_ids[i]:
                        skill = OperatorSkill.objects.get(id=skill_ids[i])
                        skill.process_name_id = str(process_name.id)
                        skill.process_name = process_name.name
                        skill.priority = priority
                        skill.save()
                    else:
                        OperatorSkill.objects.create(
                            operator_id=str(operator.id),
                            operator_name=operator.name,
                            process_name_id=str(process_name.id),
                            process_name=process_name.name,
                            priority=priority,
                        )
                except (ValueError, ProcessName.DoesNotExist):
                    messages.error(request, "無效的工序或優先順序數據！")
                    return render(
                        request,
                        "process/edit_operator.html",
                        {
                            "operator": operator,
                            "process_names": process_names,
                            "production_lines": production_lines,
                        },
                    )
        log_user_operation(
            request.user.username, "process", f"成功編輯作業員: {operator}"
        )
        messages.success(request, "作業員更新成功！")
        return redirect("process:operators")
    return render(
        request,
        "process/edit_operator.html",
        {
            "operator": operator,
            "process_names": process_names,
            "production_lines": production_lines,
        },
    )


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def delete_operator(request, operator_id):
    log_user_operation(
        request.user.username, "process", f"嘗試刪除作業員: {operator_id}"
    )
    operator = get_object_or_404(Operator, id=operator_id)
    
    # 先刪除相關的技能記錄
    from .models import OperatorSkill
    OperatorSkill.objects.filter(operator_id=str(operator_id)).delete()
    
    # 然後刪除作業員
    operator.delete()
    log_user_operation(
        request.user.username, "process", f"成功刪除作業員: {operator_id}"
    )
    messages.success(request, "作業員刪除成功！")
    return redirect("process:operators")


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def export_operators(request):
    log_user_operation(request.user.username, "process", "匯出作業員與技能數據")
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Operators and Skills"
    # 修正欄位順序：作業員名稱、所屬單位、工序名稱、技能優先順序
    headers = ["作業員名稱", "所屬單位", "工序名稱", "技能優先順序"]
    for col_num, header in enumerate(headers, 1):
        worksheet[f"{get_column_letter(col_num)}1"] = header
    operator_skills = OperatorSkill.objects.all()
    row_num = 2
    for skill in operator_skills:
        worksheet[f"A{row_num}"] = skill.operator_name
        # 透過API查詢產線名稱
        production_line_name = ""
        if skill.operator_id:
            try:
                operator = Operator.objects.get(id=skill.operator_id)
                production_line_name = operator.production_line_name or ""
            except:
                pass
        worksheet[f"B{row_num}"] = production_line_name
        worksheet[f"C{row_num}"] = skill.process_name
        worksheet[f"D{row_num}"] = skill.priority
        row_num += 1
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="operators_and_skills.xlsx"'
    return response


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def import_operators(request):
    log_user_operation(request.user.username, "process", "嘗試匯入作業員與技能數據")
    if request.method == "POST":
        if "action" in request.POST:
            action = request.POST.get("action")
            dataset_json = request.POST.get("dataset")
            dataset = json.loads(dataset_json)
            if action == "overwrite":
                for row in dataset:
                    operator_name = row["作業員名稱"]
                    production_line_name = row.get("所屬單位", "")
                    process_name = row["工序名稱"]
                    priority = row["技能優先順序"]
                    if not operator_name or not process_name:
                        continue
                    try:
                        process = ProcessName.objects.get(name=process_name)
                    except ProcessName.DoesNotExist:
                        continue
                    try:
                        priority = int(priority) if priority else 1
                    except (ValueError, TypeError):
                        priority = 1

                    # 根據產線名稱找到對應的 ProductionLine
                    production_line = None
                    if production_line_name:
                        try:
                            production_line = ProductionLine.objects.get(
                                line_name=production_line_name
                            )
                        except ProductionLine.DoesNotExist:
                            # 如果找不到產線，記錄警告但繼續處理
                            messages.warning(
                                request,
                                f"找不到產線：{production_line_name}，作業員 {operator_name} 將不設定產線",
                            )

                    operator, created = Operator.objects.get_or_create(
                        name=operator_name
                    )

                    # 更新作業員的產線資訊
                    if production_line:
                        operator.production_line_id = str(production_line.id)
                        operator.production_line_name = production_line.line_name
                        operator.save()

                    existing_skill = OperatorSkill.objects.filter(
                        operator_id=str(operator.id), process_name_id=str(process.id)
                    ).first()
                    if existing_skill:
                        existing_skill.priority = priority
                        existing_skill.save()
                    else:
                        OperatorSkill.objects.create(
                            operator_id=str(operator.id),
                            operator_name=operator.name,
                            process_name_id=str(process.id),
                            process_name=process.name,
                            priority=priority
                        )
                log_user_operation(
                    request.user.username, "process", "匯入作業員與技能數據（覆蓋模式）"
                )
                messages.success(request, "作業員與技能數據匯入成功！")
                return redirect("process:operators")
            else:
                messages.info(request, "匯入操作已取消。")
                return redirect("process:operators")
        if "file" not in request.FILES:
            messages.error(request, "請上傳一個文件！")
            return redirect("process:operators")
        file = request.FILES["file"]
        if not file.name.endswith(".xlsx"):
            messages.error(request, "請上傳 Excel 文件（.xlsx 格式）！")
            return redirect("process:operators")
        try:
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook.active
        except Exception as e:
            messages.error(request, f"無法讀取文件：{str(e)}")
            return redirect("process:operators")
        # 修正必要欄位檢查，支援「所屬單位」欄位
        required_headers = ["作業員名稱", "工序名稱", "技能優先順序"]
        optional_headers = ["所屬單位"]
        actual_headers = [cell.value for cell in worksheet[1]]
        if not all(header in actual_headers for header in required_headers):
            messages.error(
                request, f"Excel 文件缺少必要欄位，必須包含：{required_headers}"
            )
            return redirect("process:operators")

        # 檢查是否有「所屬單位」欄位
        has_production_line = "所屬單位" in actual_headers

        dataset = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            operator_name = row[actual_headers.index("作業員名稱")]
            process_name = row[actual_headers.index("工序名稱")]
            priority = row[actual_headers.index("技能優先順序")]

            # 取得所屬單位（產線名稱）
            production_line_name = ""
            if has_production_line:
                production_line_name = row[actual_headers.index("所屬單位")] or ""

            dataset.append(
                {
                    "作業員名稱": operator_name,
                    "所屬單位": production_line_name,
                    "工序名稱": process_name,
                    "技能優先順序": priority,
                }
            )
        existing_skills = set(
            f"{skill.operator_name}_{skill.process_name}"
            for skill in OperatorSkill.objects.all()
        )
        import_skills = set(f"{row['作業員名稱']}_{row['工序名稱']}" for row in dataset)
        duplicate_skills = [
            skill for skill in import_skills if skill in existing_skills
        ]
        if duplicate_skills:
            dataset_json = json.dumps(dataset)
            return render(
                request,
                "process/import_operators.html",
                {
                    "duplicate_skills": duplicate_skills,
                    "dataset": dataset_json,
                },
            )
        for row in dataset:
            operator_name = row["作業員名稱"]
            production_line_name = row.get("所屬單位", "")
            process_name = row["工序名稱"]
            priority = row["技能優先順序"]
            if not operator_name or not process_name:
                continue
            try:
                process = ProcessName.objects.get(name=process_name)
            except ProcessName.DoesNotExist:
                continue
            try:
                priority = int(priority) if priority else 1
            except (ValueError, TypeError):
                priority = 1

            # 根據產線名稱找到對應的 ProductionLine
            production_line = None
            if production_line_name:
                try:
                    production_line = ProductionLine.objects.get(
                        line_name=production_line_name
                    )
                except ProductionLine.DoesNotExist:
                    # 如果找不到產線，記錄警告但繼續處理
                    messages.warning(
                        request,
                        f"找不到產線：{production_line_name}，作業員 {operator_name} 將不設定產線",
                    )

            operator, created = Operator.objects.get_or_create(name=operator_name)

            # 更新作業員的產線資訊
            if production_line:
                operator.production_line_id = str(production_line.id)
                operator.production_line_name = production_line.line_name
                operator.save()

            existing_skill = OperatorSkill.objects.filter(
                operator_id=str(operator.id), process_name_id=str(process.id)
            ).first()
            if existing_skill:
                existing_skill.priority = priority
                existing_skill.save()
            else:
                OperatorSkill.objects.create(
                    operator_id=str(operator.id),
                    operator_name=operator.name,
                    process_name_id=str(process.id),
                    process_name=process.name,
                    priority=priority
                )
        log_user_operation(request.user.username, "process", "匯入作業員與技能數據")
        messages.success(request, "作業員與技能數據匯入成功！")
        return redirect("process:operators")
    return render(request, "process/import_operators.html", {})
