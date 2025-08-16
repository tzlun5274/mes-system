from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.db import transaction
from .models import ProcessName, ProductProcessRoute
from equip.models import Equipment
from .utils import log_user_operation
from .views_process_names import get_equipment_options
import openpyxl
from openpyxl.utils import get_column_letter
from io import BytesIO
import logging
import re
import json
import unicodedata
from django.views.decorators.http import require_POST
import math
from urllib.parse import unquote
from process.models import ProductProcessStandardCapacity

logger = logging.getLogger(__name__)


def standardize_product_id(product_id):
    """
    標準化產品編號：
    1. 去除前後空白
    2. 全形轉半形
    3. 去除不可見字元
    4. 統一為大寫
    5. 去除多餘空格
    """
    if not product_id:
        return ""

    # 轉為字串
    product_id = str(product_id)

    # 去除前後空白
    product_id = product_id.strip()

    # 全形轉半形
    product_id = unicodedata.normalize("NFKC", product_id)

    # 去除不可見字元（保留空格、數字、字母、連字號、底線）
    product_id = re.sub(r"[^\w\s\-]", "", product_id)

    # 統一為大寫
    product_id = product_id.upper()

    # 去除多餘空格（將多個空格替換為單個空格）
    product_id = re.sub(r"\s+", " ", product_id)

    # 再次去除前後空白
    product_id = product_id.strip()

    return product_id


def check_product_id_variations(dataset):
    """
    檢查產品編號變體，找出可能的重複
    回傳：{標準化編號: [原始編號列表]}
    """
    variations = {}
    for row in dataset:
        original_id = row["產品編號"]
        standardized_id = standardize_product_id(original_id)

        if standardized_id not in variations:
            variations[standardized_id] = []
        variations[standardized_id].append(original_id)

    # 找出有變體的產品編號
    problematic_variations = {}
    for std_id, original_ids in variations.items():
        if len(set(original_ids)) > 1:  # 有多個不同的原始編號
            problematic_variations[std_id] = list(set(original_ids))

    return problematic_variations


def process_user_required(user):
    return user.is_superuser or user.groups.filter(name="工序使用者").exists()


def superuser_required(user):
    return user.is_superuser


def clean_equipment_ids(equipment_ids):
    if not equipment_ids or equipment_ids == [""] or equipment_ids == "":
        logger.debug(f"設備 ID 為空: {equipment_ids}")
        return ""
    try:
        if isinstance(equipment_ids, str):
            ids = [
                id.strip()
                for id in re.split(r"[,;]", equipment_ids)
                if id.strip() and id.strip().isdigit()
            ]
        elif isinstance(equipment_ids, list):
            ids = [
                str(id).strip()
                for id in equipment_ids
                if str(id).strip() and str(id).strip().isdigit()
            ]
        else:
            ids = []
        valid_ids = sorted(set(ids))
        logger.debug(f"清理設備 ID: 原始={equipment_ids}, 清潔後={valid_ids}")
        return ",".join(valid_ids) if valid_ids else ""
    except Exception as e:
        logger.error(f"清理設備 ID 失敗: {equipment_ids}, 錯誤: {str(e)}")
        return ""


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def product_routes(request):
    log_user_operation(request.user.username, "process", "查看產品工藝路線設定")

    # 取得搜尋參數
    search = request.GET.get("search", "").strip()
    sort = request.GET.get("sort", "product_id")  # 預設按產品編號排序

    # 查詢產品編號
    product_ids = ProductProcessRoute.objects.values("product_id").distinct()

    # 如果有搜尋條件，過濾產品編號
    if search:
        product_ids = product_ids.filter(product_id__icontains=search)
        log_user_operation(
            request.user.username, "process", f"搜尋產品工藝路線: {search}"
        )

    # 排序
    if sort == "product_id":
        product_ids = product_ids.order_by("product_id")
    elif sort == "product_id_desc":
        product_ids = product_ids.order_by("-product_id")

    return render(
        request,
        "process/product_routes.html",
        {
            "product_ids": product_ids,
            "search": search,
            "sort": sort,
        },
    )


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def view_product_route(request, product_id):
    product_id = unquote(product_id)  # 解碼 URL 編碼
    log_user_operation(
        request.user.username, "process", f"檢視產品工藝路線: {product_id}"
    )
    routes = ProductProcessRoute.objects.filter(product_id=product_id).order_by(
        "step_order"
    )
    if not routes.exists():
        messages.error(request, f"產品編號 {product_id} 不存在！")
        return redirect("process:product_routes")

    equipments = get_equipment_options(request)
    equip_id_to_name = {
        str(equip["id"]): equip["name"]
        for equip in equipments
        if isinstance(equip, dict) and "id" in equip and "name" in equip
    }
    logger.debug(f"設備 ID 到名稱映射: {equip_id_to_name}")

    enhanced_routes = []
    for route in routes:
        equipment_ids = (
            [
                eid.strip()
                for eid in route.usable_equipment_ids.split(",")
                if eid.strip() and eid.strip().isdigit()
            ]
            if route.usable_equipment_ids
            else []
        )
        equipment_names = [
            equip_id_to_name.get(eid, f"未知設備({eid})") for eid in equipment_ids
        ]
        logger.debug(
            f"檢視產品工藝路線: product_id={product_id}, step_order={route.step_order}, process_name={route.process_name.name}, equipment_ids={equipment_ids}, equipment_names={equipment_names}"
        )
        enhanced_routes.append(
            {
                "id": route.id,
                "process_name": route.process_name,
                "step_order": route.step_order,
                "usable_equipment_names": equipment_names,
                "dependent_semi_product": route.dependent_semi_product or "",
            }
        )

    return render(
        request,
        "process/view_product_route.html",
        {
            "product_id": product_id,
            "routes": enhanced_routes,
        },
    )


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def add_product_route(request):
    log_user_operation(request.user.username, "process", "嘗試添加產品工藝路線")
    if not request.user.has_perm("process.can_add_product_route"):
        messages.error(request, "您沒有添加產品工藝路線的權限！")
        return redirect("process:product_routes")
    equipments = get_equipment_options(request)
    psp_product_ids = (
        ProductProcessRoute.objects.filter(product_id__startswith="PSP-")
        .values("product_id")
        .distinct()
    )
    process_names = ProcessName.objects.all().prefetch_related("equipments")
    if request.method == "POST":
        product_id = request.POST.get("product_id").strip()
        logger.debug(f"添加產品工藝路線: product_id={product_id}")
        # 新增防呆：產品編號唯一性檢查
        if ProductProcessRoute.objects.filter(product_id=product_id).exists():
            messages.error(request, f"產品編號 '{product_id}' 已存在，請勿重複建立！")
            return render(
                request,
                "process/add_product_route.html",
                {
                    "process_names": process_names,
                    "product_id": product_id,
                    "equipments": equipments,
                    "psp_product_ids": psp_product_ids,
                },
            )
        if not product_id:
            messages.error(request, "產品編號不能為空！")
            return render(
                request,
                "process/add_product_route.html",
                {
                    "process_names": process_names,
                    "product_id": product_id,
                    "equipments": equipments,
                    "psp_product_ids": psp_product_ids,
                },
            )
        if ProductProcessRoute.objects.filter(product_id=product_id).exists():
            messages.error(request, f"產品編號 '{product_id}' 已存在，請選擇其他編號！")
            return render(
                request,
                "process/add_product_route.html",
                {
                    "process_names": process_names,
                    "product_id": product_id,
                    "equipments": equipments,
                    "psp_product_ids": psp_product_ids,
                },
            )
        process_names_list = request.POST.getlist("process_name[]")
        step_orders = request.POST.getlist("step_order[]")
        # capacity_per_hour 相關程式碼已移除，不再處理此欄位
        dependent_semi_products = request.POST.getlist("dependent_semi_product[]")
        logger.debug(
            f"POST 數據: product_id={product_id}, process_names={process_names_list}, step_orders={step_orders}, semi_products={dependent_semi_products}"
        )
        if len(process_names_list) != len(step_orders):
            messages.error(request, "工序名稱和工序順序數量不匹配！")
            return render(
                request,
                "process/add_product_route.html",
                {
                    "process_names": process_names,
                    "product_id": product_id,
                    "equipments": equipments,
                    "psp_product_ids": psp_product_ids,
                },
            )
        step_orders_set = set()
        valid_psp_ids = {p["product_id"] for p in psp_product_ids}
        for idx, step_order in enumerate(step_orders):
            try:
                step_order = int(step_order)
                if step_order < 1:
                    messages.error(request, f"工序 {idx+1} 的工序順序必須是正整數！")
                    return render(
                        request,
                        "process/add_product_route.html",
                        {
                            "process_names": process_names,
                            "product_id": product_id,
                            "equipments": equipments,
                            "psp_product_ids": psp_product_ids,
                        },
                    )
                if step_order in step_orders_set:
                    messages.error(
                        request, f"工序順序 {step_order} 重複，請確保每個工序順序唯一！"
                    )
                    return render(
                        request,
                        "process/add_product_route.html",
                        {
                            "process_names": process_names,
                            "product_id": product_id,
                            "equipments": equipments,
                            "psp_product_ids": psp_product_ids,
                        },
                    )
                step_orders_set.add(step_order)
            except ValueError:
                messages.error(request, f"工序 {idx+1} 的工序順序必須是整數！")
                return render(
                    request,
                    "process/add_product_route.html",
                    {
                        "process_names": process_names,
                        "product_id": product_id,
                        "equipments": equipments,
                        "psp_product_ids": psp_product_ids,
                    },
                )
        # capacity_per_hour 相關程式碼已移除，不再處理此欄位
        try:
            with transaction.atomic():
                for idx, (process_name_id, step_order) in enumerate(
                    zip(process_names_list, step_orders)
                ):
                    try:
                        process_name = ProcessName.objects.get(id=process_name_id)
                    except ProcessName.DoesNotExist:
                        messages.error(request, f"工序 {idx+1} 的工序名稱無效！")
                        return render(
                            request,
                            "process/add_product_route.html",
                            {
                                "process_names": process_names,
                                "product_id": product_id,
                                "equipments": equipments,
                                "psp_product_ids": psp_product_ids,
                            },
                        )
                    equipment_ids_key = f"usable_equipment_ids_{idx}[]"
                    usable_equipment_ids = request.POST.getlist(equipment_ids_key)
                    dependent_semi_product = (
                        dependent_semi_products[idx]
                        if idx < len(dependent_semi_products)
                        else ""
                    )
                    if (
                        dependent_semi_product
                        and dependent_semi_product not in valid_psp_ids
                    ):
                        messages.error(
                            request,
                            f"工序 {idx+1} 的依賴半成品 '{dependent_semi_product}' 無效！",
                        )
                        return render(
                            request,
                            "process/add_product_route.html",
                            {
                                "process_names": process_names,
                                "product_id": product_id,
                                "equipments": equipments,
                                "psp_product_ids": psp_product_ids,
                            },
                        )
                    cleaned_equipment_ids = clean_equipment_ids(usable_equipment_ids)
                    logger.debug(
                        f"工序 {idx+1} 設備 ID: 鍵={equipment_ids_key}, 原始={usable_equipment_ids}, 清潔後={cleaned_equipment_ids}"
                    )
                    ProductProcessRoute.objects.create(
                        product_id=product_id,
                        process_name=process_name,
                        step_order=int(step_order),
                        usable_equipment_ids=cleaned_equipment_ids,
                        dependent_semi_product=dependent_semi_product or None,
                    )
            log_user_operation(
                request.user.username, "process", f"成功添加產品工藝路線: {product_id}"
            )
            messages.success(request, "產品工藝路線添加成功！")
            return redirect("process:product_routes")
        except Exception as e:
            logger.error(
                f"添加產品工藝路線失敗: product_id={product_id}, 錯誤: {str(e)}"
            )
            messages.error(request, f"無法添加產品工藝路線：{str(e)}")
            return render(
                request,
                "process/add_product_route.html",
                {
                    "process_names": process_names,
                    "product_id": product_id,
                    "equipments": equipments,
                    "psp_product_ids": psp_product_ids,
                },
            )
    return render(
        request,
        "process/add_product_route.html",
        {
            "process_names": process_names,
            "equipments": equipments,
            "psp_product_ids": psp_product_ids,
        },
    )


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def edit_product_route(request, product_id):
    log_user_operation(
        request.user.username, "process", f"嘗試編輯產品工藝路線: {product_id}"
    )
    if not request.user.has_perm("process.can_edit_product_route"):
        messages.error(request, "您沒有編輯產品工藝路線的權限！")
        return redirect("process:product_routes")
    routes = ProductProcessRoute.objects.filter(product_id=product_id).order_by(
        "step_order"
    )
    equipments = get_equipment_options(request)
    psp_product_ids = (
        ProductProcessRoute.objects.filter(product_id__startswith="PSP-")
        .values("product_id")
        .distinct()
    )
    if not routes.exists():
        messages.error(request, f"產品編號 {product_id} 不存在！")
        return redirect("process:product_routes")

    enhanced_routes = []
    for route in routes:
        equipment_ids = (
            [
                eid.strip()
                for eid in route.usable_equipment_ids.split(",")
                if eid.strip() and eid.strip().isdigit()
            ]
            if route.usable_equipment_ids
            else []
        )
        logger.debug(
            f"編輯頁面加載: product_id={product_id}, step_order={route.step_order}, process_name={route.process_name.name}, equipment_ids={equipment_ids}"
        )
        enhanced_routes.append(
            {
                "id": route.id,
                "process_name": route.process_name,
                "step_order": route.step_order,
                "usable_equipment_ids": equipment_ids,
                "dependent_semi_product": route.dependent_semi_product or "",
            }
        )

    if request.method == "POST":
        product_id_new = request.POST.get("product_id", "").strip()
        process_names = request.POST.getlist("process_name[]")
        step_orders = request.POST.getlist("step_order[]")
        # capacity_per_hour 相關程式碼已移除，不再處理此欄位
        dependent_semi_products = request.POST.getlist("dependent_semi_product[]")
        route_ids = request.POST.getlist("route_id[]")
        logger.debug(
            f"POST 數據: product_id_new={product_id_new}, process_names={process_names}, step_orders={step_orders}, semi_products={dependent_semi_products}, route_ids={route_ids}"
        )
        if not product_id_new:
            messages.error(request, "產品編號不能為空！")
            return render(
                request,
                "process/edit_product_route.html",
                {
                    "product_id": product_id,
                    "routes": enhanced_routes,
                    "process_names": ProcessName.objects.all(),
                    "equipments": equipments,
                    "psp_product_ids": psp_product_ids,
                },
            )
        if (
            product_id_new != product_id
            and ProductProcessRoute.objects.filter(product_id=product_id_new).exists()
        ):
            messages.error(
                request, f"產品編號 '{product_id_new}' 已存在，請選擇其他編號！"
            )
            return render(
                request,
                "process/edit_product_route.html",
                {
                    "product_id": product_id,
                    "routes": enhanced_routes,
                    "process_names": ProcessName.objects.all(),
                    "equipments": equipments,
                    "psp_product_ids": psp_product_ids,
                },
            )
        if len(process_names) != len(step_orders):
            messages.error(request, "工序名稱和工序順序數量不匹配！")
            return render(
                request,
                "process/edit_product_route.html",
                {
                    "product_id": product_id,
                    "routes": enhanced_routes,
                    "process_names": ProcessName.objects.all(),
                    "equipments": equipments,
                    "psp_product_ids": psp_product_ids,
                },
            )
        step_orders_set = set()
        valid_psp_ids = {p["product_id"] for p in psp_product_ids}
        for idx, step_order in enumerate(step_orders):
            try:
                step_order = int(step_order)
                if step_order < 1:
                    messages.error(request, f"工序 {idx+1} 的步驟順序必須是正整數！")
                    return render(
                        request,
                        "process/edit_product_route.html",
                        {
                            "product_id": product_id,
                            "routes": enhanced_routes,
                            "process_names": ProcessName.objects.all(),
                            "equipments": equipments,
                            "psp_product_ids": psp_product_ids,
                        },
                    )
                if step_order in step_orders_set:
                    messages.error(
                        request, f"步驟順序 {step_order} 重複，請確保每個步驟順序唯一！"
                    )
                    return render(
                        request,
                        "process/edit_product_route.html",
                        {
                            "product_id": product_id,
                            "routes": enhanced_routes,
                            "process_names": ProcessName.objects.all(),
                            "equipments": equipments,
                            "psp_product_ids": psp_product_ids,
                        },
                    )
                step_orders_set.add(step_order)
            except ValueError:
                messages.error(request, f"工序 {idx+1} 的步驟順序必須是整數！")
                return render(
                    request,
                    "process/edit_product_route.html",
                    {
                        "product_id": product_id,
                        "routes": enhanced_routes,
                        "process_names": ProcessName.objects.all(),
                        "equipments": equipments,
                        "psp_product_ids": psp_product_ids,
                    },
                )
        # capacity_per_hour 相關程式碼已移除，不再處理此欄位
        try:
            with transaction.atomic():
                # 處理工序移除：刪除未提交的工序（route_id 不在 route_ids 中）
                submitted_route_ids = []
                for rid in route_ids:
                    if rid.strip():  # 只處理非空的 route_id
                        try:
                            submitted_route_ids.append(int(rid))
                        except ValueError:
                            continue

                # 找出需要刪除的工序（存在於資料庫但不在提交的列表中）
                existing_routes_to_delete = routes.exclude(id__in=submitted_route_ids)
                for route in existing_routes_to_delete:
                    logger.debug(
                        f"刪除被移除的工序: product_id={product_id}, route_id={route.id}, step_order={route.step_order}"
                    )
                    route.delete()

                # 處理工序更新和新增
                for idx, (process_name_id, step_order) in enumerate(
                    zip(process_names, step_orders)
                ):
                    try:
                        process_name = ProcessName.objects.get(id=process_name_id)
                    except ProcessName.DoesNotExist:
                        messages.error(request, f"工序 {idx+1} 的工序名稱無效！")
                        return render(
                            request,
                            "process/edit_product_route.html",
                            {
                                "product_id": product_id,
                                "routes": enhanced_routes,
                                "process_names": ProcessName.objects.all(),
                                "equipments": equipments,
                                "psp_product_ids": psp_product_ids,
                            },
                        )
                    equipment_ids_key = f"usable_equipment_ids_{idx}[]"
                    usable_equipment_ids = request.POST.getlist(equipment_ids_key)
                    dependent_semi_product = (
                        dependent_semi_products[idx]
                        if idx < len(dependent_semi_products)
                        else ""
                    )
                    if (
                        dependent_semi_product
                        and dependent_semi_product not in valid_psp_ids
                    ):
                        messages.error(
                            request,
                            f"工序 {idx+1} 的依賴半成品 '{dependent_semi_product}' 無效！",
                        )
                        return render(
                            request,
                            "process/edit_product_route.html",
                            {
                                "product_id": product_id,
                                "routes": enhanced_routes,
                                "process_names": ProcessName.objects.all(),
                                "equipments": equipments,
                                "psp_product_ids": psp_product_ids,
                            },
                        )
                    cleaned_equipment_ids = clean_equipment_ids(usable_equipment_ids)
                    logger.debug(
                        f"工序 {idx+1} 設備 ID: 鍵={equipment_ids_key}, 原始={usable_equipment_ids}, 清潔後={cleaned_equipment_ids}"
                    )
                    step_order = int(step_order)

                    # 檢查是否有對應的現有工序（更新）或需要新增
                    if idx < len(route_ids) and route_ids[idx].strip():
                        try:
                            route = ProductProcessRoute.objects.get(id=route_ids[idx])
                            
                            # 檢查是否需要更新產品編號或步驟順序
                            if route.product_id != product_id_new or route.step_order != step_order:
                                # 檢查新產品編號和步驟順序是否已存在（排除當前記錄）
                                existing_duplicate = ProductProcessRoute.objects.filter(
                                    product_id=product_id_new,
                                    step_order=step_order
                                ).exclude(id=route.id).first()
                                
                                if existing_duplicate:
                                    # 如果存在重複，先刪除舊記錄，再更新重複記錄
                                    logger.debug(f"刪除舊記錄: route_id={route.id}")
                                    route.delete()
                                    # 更新重複記錄
                                    existing_duplicate.process_name = process_name
                                    existing_duplicate.usable_equipment_ids = cleaned_equipment_ids
                                    existing_duplicate.dependent_semi_product = dependent_semi_product or None
                                    existing_duplicate.save()
                                    logger.debug(f"更新重複記錄: route_id={existing_duplicate.id}, step_order={step_order}")
                                else:
                                    # 更新現有工序
                                    route.product_id = product_id_new
                                    route.process_name = process_name
                                    route.step_order = step_order
                                    route.usable_equipment_ids = cleaned_equipment_ids
                                    route.dependent_semi_product = (
                                        dependent_semi_product or None
                                    )
                                    route.save()
                                    logger.debug(
                                        f"更新工序: route_id={route.id}, step_order={step_order}"
                                    )
                            else:
                                # 只更新其他欄位，不改變產品編號和步驟順序
                                route.process_name = process_name
                                route.usable_equipment_ids = cleaned_equipment_ids
                                route.dependent_semi_product = (
                                    dependent_semi_product or None
                                )
                                route.save()
                                logger.debug(
                                    f"更新工序: route_id={route.id}, step_order={step_order}"
                                )
                        except ProductProcessRoute.DoesNotExist:
                            # 如果找不到對應的工序，檢查是否已存在相同的產品和步驟順序
                            existing_route = ProductProcessRoute.objects.filter(
                                product_id=product_id_new,
                                step_order=step_order
                            ).first()
                            
                            if existing_route:
                                # 如果已存在，則更新現有記錄
                                existing_route.process_name = process_name
                                existing_route.usable_equipment_ids = cleaned_equipment_ids
                                existing_route.dependent_semi_product = dependent_semi_product or None
                                existing_route.save()
                                logger.debug(f"更新已存在的工序: route_id={existing_route.id}, step_order={step_order}")
                            else:
                                # 如果不存在，則創建新記錄
                                ProductProcessRoute.objects.create(
                                    product_id=product_id_new,
                                    process_name=process_name,
                                    step_order=step_order,
                                    usable_equipment_ids=cleaned_equipment_ids,
                                    dependent_semi_product=dependent_semi_product or None,
                                )
                                logger.debug(f"新增工序: step_order={step_order}")
                    else:
                        # 新增工序 - 檢查是否已存在相同的產品和步驟順序
                        existing_route = ProductProcessRoute.objects.filter(
                            product_id=product_id_new,
                            step_order=step_order
                        ).first()
                        
                        if existing_route:
                            # 如果已存在，則更新現有記錄
                            existing_route.process_name = process_name
                            existing_route.usable_equipment_ids = cleaned_equipment_ids
                            existing_route.dependent_semi_product = dependent_semi_product or None
                            existing_route.save()
                            logger.debug(f"更新已存在的工序: route_id={existing_route.id}, step_order={step_order}")
                        else:
                            # 如果不存在，則創建新記錄
                            ProductProcessRoute.objects.create(
                                product_id=product_id_new,
                                process_name=process_name,
                                step_order=step_order,
                                usable_equipment_ids=cleaned_equipment_ids,
                                dependent_semi_product=dependent_semi_product or None,
                            )
                            logger.debug(f"新增工序: step_order={step_order}")
            log_user_operation(
                request.user.username,
                "process",
                f"成功編輯產品工藝路線: {product_id_new}",
            )
            messages.success(request, "產品工藝路線更新成功！")
            return redirect("process:product_routes")
        except Exception as e:
            logger.error(
                f"編輯產品工藝路線失敗: product_id={product_id}, 錯誤: {str(e)}"
            )
            messages.error(request, f"無法編輯產品工藝路線：{str(e)}")
            return render(
                request,
                "process/edit_product_route.html",
                {
                    "product_id": product_id,
                    "routes": enhanced_routes,
                    "process_names": ProcessName.objects.all(),
                    "equipments": equipments,
                    "psp_product_ids": psp_product_ids,
                },
            )
    return render(
        request,
        "process/edit_product_route.html",
        {
            "product_id": product_id,
            "routes": enhanced_routes,
            "process_names": ProcessName.objects.all(),
            "equipments": equipments,
            "psp_product_ids": psp_product_ids,
        },
    )


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
def delete_product_route(request, product_id):
    log_user_operation(
        request.user.username, "process", f"嘗試刪除產品工藝路線: {product_id}"
    )
    if not request.user.has_perm("process.can_delete_product_route"):
        messages.error(request, "您沒有刪除產品工藝路線的權限！")
        return redirect("process:product_routes")
    ProductProcessRoute.objects.filter(product_id=product_id).delete()
    log_user_operation(
        request.user.username, "process", f"成功刪除產品工藝路線: {product_id}"
    )
    messages.success(request, "產品工藝路線刪除成功！")
    return redirect("process:product_routes")


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def export_product_routes(request):
    """
    匯出產品工藝路線（Excel），欄位順序：產品編號、工序順序、工序名稱、指定可用設備、依賴半成品
    """
    log_user_operation(request.user.username, "process", "匯出產品工藝路線數據")
    equipments = get_equipment_options(request)
    equip_id_to_name = {
        str(equip["id"]): equip["name"]
        for equip in equipments
        if isinstance(equip, dict) and "id" in equip and "name" in equip
    }
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Product Process Routes"
    # 新格式：產品編號、工序順序、工序名稱、指定可用設備、依賴半成品
    headers = ["產品編號", "工序順序", "工序名稱", "指定可用設備", "依賴半成品"]
    for col_num, header in enumerate(headers, 1):
        worksheet[f"{get_column_letter(col_num)}1"] = header
    routes = ProductProcessRoute.objects.all().order_by("product_id", "step_order")
    row_num = 2
    for route in routes:
        equip_ids = (
            [
                eid.strip()
                for eid in route.usable_equipment_ids.split(",")
                if eid.strip()
            ]
            if route.usable_equipment_ids
            else []
        )
        equip_names = []
        for eid in equip_ids:
            if eid in equip_id_to_name:
                equip_names.append(equip_id_to_name[eid])
            else:
                equip_names.append(f"未知設備({eid})")
        worksheet[f"A{row_num}"] = route.product_id
        worksheet[f"B{row_num}"] = route.step_order
        worksheet[f"C{row_num}"] = route.process_name.name
        worksheet[f"D{row_num}"] = ", ".join(equip_names) if equip_names else ""
        worksheet[f"E{row_num}"] = route.dependent_semi_product or ""
        row_num += 1
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="產品工藝路線.xlsx"'
    return response


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def import_product_routes(request):
    """
    匯入產品工藝路線（Excel），欄位順序：產品編號、工序順序、工序名稱、指定可用設備、依賴半成品
    """
    log_user_operation(request.user.username, "process", "嘗試匯入產品工藝路線數據")
    equipments = get_equipment_options(request)
    equip_name_to_id = {
        equip["name"]: str(equip["id"])
        for equip in equipments
        if isinstance(equip, dict) and "id" in equip and "name" in equip
    }
    psp_product_ids = (
        ProductProcessRoute.objects.filter(product_id__startswith="PSP-")
        .values("product_id")
        .distinct()
    )
    valid_psp_ids = {p["product_id"] for p in psp_product_ids}
    if request.method == "POST":
        if "action" in request.POST:
            action = request.POST.get("action")
            dataset_json = request.POST.get("dataset")
            try:
                dataset = json.loads(dataset_json)
            except json.JSONDecodeError:
                messages.error(request, "無效的數據格式！")
                return redirect("process:product_routes")
            if action == "overwrite":
                errors = []
                success_count = 0
                product_step_map = {}
                for row in dataset:
                    original_pid = row["產品編號"]
                    # 標準化產品編號
                    pid = standardize_product_id(original_pid)
                    try:
                        step = int(row["工序順序"])
                    except Exception:
                        continue
                    if pid not in product_step_map:
                        product_step_map[pid] = set()
                    product_step_map[pid].add(step)
                with transaction.atomic():
                    for pid, steps in product_step_map.items():
                        ProductProcessRoute.objects.filter(product_id=pid).exclude(
                            step_order__in=steps
                        ).delete()
                    for row in dataset:
                        original_product_id = row["產品編號"]
                        # 標準化產品編號
                        product_id = standardize_product_id(original_product_id)
                        step_order = row["工序順序"]
                        process_name = row["工序名稱"]
                        equip_names = (
                            row.get("指定可用設備", "").split(",")
                            if row.get("指定可用設備")
                            else []
                        )
                        dependent_semi_product = row.get("依賴半成品", "")
                        if (
                            not product_id
                            or str(product_id).strip() == ""
                            or product_id.lower() == "null"
                        ):
                            errors.append(
                                f"跳過記錄：產品編號為空或無效（工序：{process_name}，順序：{step_order}）"
                            )
                            continue
                        try:
                            step_order = int(step_order)
                        except (ValueError, TypeError):
                            errors.append(
                                f"跳過記錄：無效的工序順序 {step_order}（產品編號：{product_id}）"
                            )
                            continue
                        try:
                            process = ProcessName.objects.get(name=process_name)
                        except ProcessName.DoesNotExist:
                            errors.append(
                                f"跳過記錄：工序 {process_name} 不存在（產品編號：{product_id}）"
                            )
                            continue
                        if (
                            dependent_semi_product
                            and dependent_semi_product not in valid_psp_ids
                        ):
                            errors.append(
                                f"跳過記錄：依賴半成品 {dependent_semi_product} 無效（產品編號：{product_id}）"
                            )
                            continue
                        equipment_ids = []
                        for equip_name in equip_names:
                            equip_name = equip_name.strip()
                            if equip_name:
                                if equip_name in equip_name_to_id:
                                    equipment_ids.append(equip_name_to_id[equip_name])
                                else:
                                    errors.append(
                                        f"無效設備名稱: {equip_name} (產品編號: {product_id})"
                                    )
                                    continue
                        ProductProcessRoute.objects.filter(
                            product_id=product_id, step_order=step_order
                        ).delete()
                        try:
                            ProductProcessRoute.objects.create(
                                product_id=product_id,
                                process_name=process,
                                step_order=step_order,
                                capacity_per_hour=0,  # 已移除產能，預設為 0
                                usable_equipment_ids=(
                                    ",".join(equipment_ids) if equipment_ids else ""
                                ),
                                dependent_semi_product=dependent_semi_product or None,
                            )
                            success_count += 1
                            # 記錄標準化過程
                            if original_product_id != product_id:
                                logger.info(
                                    f"產品編號標準化: '{original_product_id}' -> '{product_id}'"
                                )
                        except Exception as e:
                            errors.append(
                                f"儲存記錄失敗（產品編號：{product_id}，工序：{process_name}，順序：{step_order}）：{str(e)}"
                            )
                            logger.error(
                                f"匯入記錄失敗: product_id={product_id}, process_name={process_name}, step_order={step_order}, error={str(e)}"
                            )
                            continue
                log_user_operation(
                    request.user.username, "process", "匯入產品工藝路線數據（覆蓋模式）"
                )
                if success_count > 0:
                    messages.success(
                        request, f"成功匯入 {success_count} 條產品工藝路線記錄"
                    )
                if errors:
                    for error in errors[:5]:
                        messages.warning(request, error)
                    if len(errors) > 5:
                        messages.warning(
                            request, f"還有 {len(errors) - 5} 個警告未顯示，請檢查日誌"
                        )
                return redirect("process:product_routes")
            else:
                messages.info(request, "匯入操作已取消。")
                return redirect("process:product_routes")
        if "file" not in request.FILES:
            messages.error(request, "請上傳一個文件！")
            return redirect("process:product_routes")
        file = request.FILES["file"]
        if not file.name.endswith(".xlsx"):
            messages.error(request, "請上傳 Excel 文件（.xlsx 格式）！")
            return redirect("process:product_routes")
        try:
            workbook = openpyxl.load_workbook(file)
            worksheet = workbook.active
        except Exception as e:
            messages.error(request, f"無法讀取文件：{str(e)}")
            return redirect("process:product_routes")
        required_headers = ["產品編號", "工序順序", "工序名稱"]
        actual_headers = [cell.value for cell in worksheet[1]]
        if not all(header in actual_headers for header in required_headers):
            messages.error(
                request, f"Excel 文件缺少必要欄位，必須包含：{required_headers}"
            )
            return redirect("process:product_routes")
        dataset = []
        errors = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            product_id = row[actual_headers.index("產品編號")]
            step_order = row[actual_headers.index("工序順序")]
            process_name = row[actual_headers.index("工序名稱")]
            equip_names = (
                row[actual_headers.index("指定可用設備")]
                if "指定可用設備" in actual_headers
                else ""
            )
            dependent_semi_product = (
                row[actual_headers.index("依賴半成品")]
                if "依賴半成品" in actual_headers
                else ""
            )
            if (
                not product_id
                or str(product_id).strip() == ""
                or product_id.lower() == "null"
            ):
                errors.append(
                    f"跳過記錄：產品編號為空或無效（工序：{process_name}，順序：{step_order}）"
                )
                continue
            dataset.append(
                {
                    "產品編號": product_id,
                    "工序順序": step_order,
                    "工序名稱": process_name,
                    "指定可用設備": equip_names,
                    "依賴半成品": dependent_semi_product,
                }
            )

        # 檢查產品編號變體
        problematic_variations = check_product_id_variations(dataset)
        if problematic_variations:
            # 顯示產品編號變體警告
            variation_messages = []
            for std_id, original_ids in problematic_variations.items():
                variation_messages.append(
                    f"產品編號 '{std_id}' 發現多種寫法：{', '.join(original_ids)}"
                )

            # 將原始資料轉為 JSON 供後續處理
            dataset_json = json.dumps(dataset)
            return render(
                request,
                "process/import_product_routes.html",
                {
                    "product_id_variations": problematic_variations,
                    "variation_messages": variation_messages,
                    "dataset": dataset_json,
                    "show_variation_warning": True,
                },
            )

        duplicate_routes = []
        for row in dataset:
            product_id = row["產品編號"]
            step_order = row["工序順序"]
            try:
                step_order = int(step_order)
                if ProductProcessRoute.objects.filter(
                    product_id=product_id, step_order=step_order
                ).exists():
                    duplicate_routes.append((product_id, step_order))
            except (ValueError, TypeError):
                errors.append(
                    f"跳過記錄：無效的工序順序 {step_order}（產品編號：{product_id}）"
                )
                continue
        if duplicate_routes:
            dataset_json = json.dumps(dataset)
            return render(
                request,
                "process/import_product_routes.html",
                {
                    "duplicate_routes": duplicate_routes,
                    "dataset": dataset_json,
                },
            )
        success_count = 0
        with transaction.atomic():
            for row in dataset:
                original_product_id = row["產品編號"]
                # 標準化產品編號
                product_id = standardize_product_id(original_product_id)
                step_order = row["工序順序"]
                process_name = row["工序名稱"]
                equip_names = (
                    row.get("指定可用設備", "").split(",")
                    if row.get("指定可用設備")
                    else []
                )
                dependent_semi_product = row.get("依賴半成品", "")
                if (
                    not product_id
                    or str(product_id).strip() == ""
                    or product_id.lower() == "null"
                ):
                    errors.append(
                        f"跳過記錄：產品編號為空或無效（工序：{process_name}，順序：{step_order}）"
                    )
                    continue
                try:
                    step_order = int(step_order)
                except (ValueError, TypeError):
                    errors.append(
                        f"跳過記錄：無效的工序順序 {step_order}（產品編號：{product_id}）"
                    )
                    continue
                try:
                    process = ProcessName.objects.get(name=process_name)
                except ProcessName.DoesNotExist:
                    errors.append(
                        f"跳過記錄：工序 {process_name} 不存在（產品編號：{product_id}）"
                    )
                    continue
                if (
                    dependent_semi_product
                    and dependent_semi_product not in valid_psp_ids
                ):
                    errors.append(
                        f"跳過記錄：依賴半成品 {dependent_semi_product} 無效（產品編號：{product_id}）"
                    )
                    continue
                equipment_ids = []
                for equip_name in equip_names:
                    equip_name = equip_name.strip()
                    if equip_name:
                        if equip_name in equip_name_to_id:
                            equipment_ids.append(equip_name_to_id[equip_name])
                        else:
                            errors.append(
                                f"無效設備名稱: {equip_name} (產品編號: {product_id})"
                            )
                            continue
                try:
                    ProductProcessRoute.objects.filter(
                        product_id=product_id, step_order=step_order
                    ).delete()
                    ProductProcessRoute.objects.create(
                        product_id=product_id,
                        process_name=process,
                        step_order=step_order,
                        usable_equipment_ids=(
                            ",".join(equipment_ids) if equipment_ids else ""
                        ),
                        dependent_semi_product=dependent_semi_product or None,
                    )
                    success_count += 1
                    # 記錄標準化過程
                    if original_product_id != product_id:
                        logger.info(
                            f"產品編號標準化: '{original_product_id}' -> '{product_id}'"
                        )
                except Exception as e:
                    errors.append(
                        f"儲存記錄失敗（產品編號：{product_id}，工序：{process_name}，順序：{step_order}）：{str(e)}"
                    )
                    logger.error(
                        f"匯入記錄失敗: product_id={product_id}, process_name={process_name}, step_order={step_order}, error={str(e)}"
                    )
                    continue
        log_user_operation(request.user.username, "process", "匯入產品工藝路線數據")
        if success_count > 0:
            messages.success(request, f"成功匯入 {success_count} 條產品工藝路線記錄")
        if errors:
            for error in errors[:5]:
                messages.warning(request, error)
            if len(errors) > 5:
                messages.warning(
                    request, f"還有 {len(errors) - 5} 個警告未顯示，請檢查日誌"
                )
        return redirect("process:product_routes")
    return render(request, "process/import_product_routes.html", {})


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
@require_POST
def api_calculate_capacity(request):
    """
    產能計算 API：
    接收產品編號與每個工序的併行工序數，計算每道工序的工時（小時）、天數（每日8+3小時），並加總總工時與總天數。
    回傳 JSON 結果。
    """
    product_id = request.POST.get("product_id")
    parallel_steps = request.POST.get("parallel_steps")  # JSON 格式: {工序順序: 併行數}
    if not product_id:
        return JsonResponse({"success": False, "msg": "缺少產品編號"}, status=400)
    try:
        parallel_steps = json.loads(parallel_steps) if parallel_steps else {}
    except Exception:
        parallel_steps = {}
    routes = ProductProcessRoute.objects.filter(product_id=product_id).order_by(
        "step_order"
    )
    if not routes.exists():
        return JsonResponse({"success": False, "msg": "查無產品工藝路線"})
    result = []
    total_hours = 0.0
    total_days = 0.0
    for route in routes:
        step_order = route.step_order
        parallel = int(parallel_steps.get(str(step_order), 1))

        # 查詢標準產能資料
        capacity_data = (
            ProductProcessStandardCapacity.objects.filter(
                product_code=product_id,
                process_name=route.process_name.name,
                is_active=True,
            )
            .order_by("-version")
            .first()
        )

        # 使用標準產能或預設值
        capacity_per_hour = (
            capacity_data.standard_capacity_per_hour if capacity_data else 1000
        )

        # 假設每道工序需生產 1 批（可依需求調整）
        batch_qty = 1
        # 工時 = 批量 / (併行數 * 每小時產能)
        hours = (
            batch_qty / (parallel * capacity_per_hour)
            if parallel > 0 and capacity_per_hour > 0
            else 0
        )
        # 天數 = 工時 / 11（每日8+3小時）
        days = hours / 11 if hours > 0 else 0
        total_hours += hours
        total_days += days
        result.append(
            {
                "step_order": step_order,
                "process_name": route.process_name.name,
                "parallel": parallel,
                "capacity_per_hour": capacity_per_hour,
                "hours": round(hours, 2),
                "days": round(days, 2),
            }
        )
    return JsonResponse(
        {
            "success": True,
            "routes": result,
            "total_hours": round(total_hours, 2),
            "total_days": round(total_days, 2),
        }
    )


@login_required
@user_passes_test(process_user_required, login_url="/accounts/login/")
@require_POST
def force_delete_product_route(request, product_id):
    """
    徹底刪除所有包含指定產品編號的產品工藝路線（支援模糊比對、空白、特殊字元）
    """
    from process.models import ProductProcessRoute
    from urllib.parse import unquote
    import logging

    logger = logging.getLogger("process")
    keyword = unquote(product_id)
    qs = ProductProcessRoute.objects.filter(product_id__icontains=keyword)
    count = qs.count()
    if count == 0:
        messages.error(request, f"找不到任何包含「{keyword}」的產品工藝路線。")
    else:
        for obj in qs:
            logger.info(
                f"徹底刪除產品工藝路線：產品編號={obj.product_id}，工序={obj.process_name}，順序={obj.step_order}"
            )
        qs.delete()
        messages.success(
            request, f"已徹底刪除 {count} 筆產品工藝路線，關鍵字：「{keyword}」"
        )
    return redirect("process:product_routes")
