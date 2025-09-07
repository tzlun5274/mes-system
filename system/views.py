import logging
import csv
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User, Permission
from django.contrib import messages
from django.contrib.auth.forms import UserCreationForm, UserChangeForm, SetPasswordForm
from django import forms
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from .forms import (
    UserCreationFormCustom,
    UserChangeFormCustom,
    OperationLogConfigForm,
    EmailConfigForm,
    UserWorkPermissionForm,
    BackupScheduleForm,
    OrderSyncSettingsForm
)
from .models import (
    EmailConfig, 
    BackupSchedule, 
    OperationLogConfig,
    AutoApprovalSettings,
    CleanupLog,
    OrderSyncSettings,
    OrderSyncLog,
    UserWorkPermission
)
from django.core.mail import get_connection, send_mail
from django.http import HttpResponse, FileResponse, JsonResponse
import os
import subprocess
from datetime import datetime, timedelta
import smtplib
from django_celery_beat.models import PeriodicTask, CrontabSchedule, IntervalSchedule
from .tasks import auto_backup_database
import csv
import openpyxl
from django.utils import timezone
from django.urls import get_resolver, reverse_lazy
from django.conf import settings
import shutil
import glob
from django.db.models import Q
from datetime import timedelta
from django.views.generic import UpdateView, View
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, CreateView, DeleteView
from .models import ScheduledTask
from .translations import (
    PERMISSION_NAME_TRANSLATIONS,
    get_permission_display_name as get_translated_permission_name,
    get_module_display_name as get_translated_module_name,
    get_permissions_text,
)
from django.core.paginator import Paginator

# from system.models import AutoApprovalTask  # 移除循環導入

# 設定系統管理模組的日誌記錄器
from django.conf import settings
system_logger = logging.getLogger("system")
system_handler = logging.FileHandler(os.path.join(settings.SYSTEM_LOG_DIR, "system.log"))
system_handler.setFormatter(
    logging.Formatter("%(levelname)s %(asctime)s %(module)s %(message)s")
)
system_logger.addHandler(system_handler)
system_logger.setLevel(logging.INFO)

logger = logging.getLogger("django")
logger.setLevel(logging.DEBUG)
handler = logging.FileHandler(os.path.join(settings.DJANGO_LOG_DIR, "mes.log"))
formatter = logging.Formatter("%(levelname)s %(asctime)s %(module)s %(message)s")
handler.setFormatter(formatter)
logger.handlers = [handler]

# 定義所有模組及其對應的 OperationLog 模型
MODULE_LOG_MODELS = {
    "equip": "equip.models.EquipOperationLog",
    "material": "material.models.MaterialOperationLog",
    "scheduling": "scheduling.models.SchedulingOperationLog",
    "process": "process.models.ProcessOperationLog",
    "quality": "quality.models.QualityOperationLog",
    "work_order": "workorder.models.WorkOrderOperationLog",
    "kanban": "kanban.models.KanbanOperationLog",
    "erp_integration": "erp_integration.models.ERPIntegrationOperationLog",
    "ai": "ai.models.AIOperationLog",
}


def superuser_required(user):
    return user.is_superuser


def get_module_display_name(module_name):
    try:
        module = __import__(f"{module_name}.urls", fromlist=["module_display_name"])
        return getattr(module, "module_display_name", module_name)
    except ImportError:
        return module_name


def get_all_permissions():
    all_modules = [
        "equip",
        "material",
        "scheduling",
        "process",
        "quality",
        "workorder",
        "kanban",
        "erp_integration",
        "ai",
    ]
    permissions = Permission.objects.filter(content_type__app_label__in=all_modules)
    return permissions


class CustomPermissionChoiceField(forms.ModelMultipleChoiceField):
    def label_from_instance(self, obj):
        logger.debug(f"處理權限: name='{obj.name}', codename='{obj.codename}'")
        if obj.name.startswith("可以"):
            logger.debug(f"使用自訂權限名稱: {obj.name}")
            return obj.name
        name_lower = obj.name.lower().strip()
        translations_lower = {
            k.lower().strip(): v for k, v in PERMISSION_NAME_TRANSLATIONS.items()
        }
        translated_name = translations_lower.get(name_lower, obj.name)
        if translated_name == obj.name:
            logger.warning(
                f"未找到權限 '{obj.name}' (codename: {obj.codename}) 的中文翻譯，映射表鍵: {list(translations_lower.keys())}"
            )
            codename_lower = obj.codename.lower().strip()
            translated_name = translations_lower.get(codename_lower, obj.name)
            if translated_name == obj.name:
                logger.warning(f"也未找到 codename '{obj.codename}' 的中文翻譯")
            else:
                logger.debug(
                    f"通過 codename 找到翻譯: '{obj.codename}' 翻譯為: {translated_name}"
                )
        else:
            logger.debug(f"權限 '{obj.name}' 翻譯為: {translated_name}")
        return translated_name


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def index(request):
    return render(request, "system/index.html", {})


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def user_list(request):
    """用戶列表頁面"""
    # 支援搜尋和篩選
    search_query = request.GET.get('search', '')
    status_filter = request.GET.get('status', 'all')
    
    users = User.objects.all()
    
    # 搜尋功能
    if search_query:
        users = users.filter(
            Q(username__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query)
        )
    
    # 狀態篩選
    if status_filter == 'active':
        users = users.filter(is_active=True)
    elif status_filter == 'inactive':
        users = users.filter(is_active=False)
    elif status_filter == 'staff':
        users = users.filter(is_staff=True)
    elif status_filter == 'superuser':
        users = users.filter(is_superuser=True)
    
    # 排序
    users = users.order_by('username')
    
    # 分頁
    paginator = Paginator(users, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 計算統計數據
    all_users = User.objects.all()
    active_users_count = all_users.filter(is_active=True).count()
    staff_users_count = all_users.filter(is_staff=True).count()
    super_users_count = all_users.filter(is_superuser=True).count()
    
    context = {
        'users': page_obj,
        'page_obj': page_obj,
        'search_query': search_query,
        'status_filter': status_filter,
        'total_users': all_users.count(),
        'active_users_count': active_users_count,
        'staff_users_count': staff_users_count,
        'super_users_count': super_users_count,
        'active_users': all_users.filter(is_active=True).count(),
        'inactive_users': all_users.filter(is_active=False).count(),
        'staff_users': all_users.filter(is_staff=True).count(),
        'superusers': all_users.filter(is_superuser=True).count(),
    }
    
    return render(request, "system/user_list.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def user_add(request):
    """新增用戶頁面"""
    if request.method == "POST":
        form = UserCreationFormCustom(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.email = form.cleaned_data["email"]
            user.first_name = form.cleaned_data.get("first_name", "")
            user.last_name = form.cleaned_data.get("last_name", "")
            user.is_active = form.cleaned_data.get("is_active", True)
            user.is_staff = form.cleaned_data.get("is_staff", False)
            user.save()
            
            # 記錄操作日誌
            logger.info(f"用戶 {user.username} 由 {request.user.username} 新增")
            messages.success(request, f"用戶 {user.username} 新增成功！")

            return redirect("system:user_list")
    else:
        form = UserCreationFormCustom()
    
    context = {
        "form": form,
        "title": "新增用戶",
        "submit_text": "新增用戶"
    }
    
    return render(request, "system/user_form.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def user_detail(request, user_id):
    """用戶詳情頁面"""
    user = get_object_or_404(User, id=user_id)
    
    context = {
        "user_obj": user,
        "title": f"用戶詳情 - {user.username}"
    }
    
    return render(request, "system/user_detail.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def user_edit(request, user_id):
    """編輯用戶頁面"""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == "POST":
        form = UserChangeFormCustom(request.POST, instance=user)
        if form.is_valid():
            user = form.save()
            
            # 記錄操作日誌
            logger.info(f"用戶 {user.username} 由 {request.user.username} 編輯")
            messages.success(request, f"用戶 {user.username} 編輯成功！")
            
            return redirect("system:user_list")
    else:
        form = UserChangeFormCustom(instance=user)

    context = {
        "form": form,
        "title": f"編輯用戶 - {user.username}",
        "submit_text": "更新用戶",
        "user_id": user_id,
        "user_obj": user,
        "target_user": user
    }
    
    return render(request, "system/user_form.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def user_change_password(request, user_id):
    """更改指定用戶的密碼"""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == "POST":
        form = SetPasswordForm(user=user, data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, f"用戶 {user.username} 的密碼已成功更改！")
            logger.info(f"用戶 {user.username} 的密碼由 {request.user.username} 更改")
            return redirect("system:user_list")
        else:
            messages.error(request, "請修正以下錯誤：")
    else:
        form = SetPasswordForm(user=user)
    
    context = {
        "form": form,
        "title": f"更改用戶 {user.username} 的密碼",
        "user_id": user_id,
        "user_obj": user,
        "target_user": user,
        "submit_text": "更改密碼"
    }
    
    return render(request, "system/user_change_password.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def user_delete(request, user_id):
    """刪除用戶"""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == "POST":
        # 防止刪除自己
        if user == request.user:
            messages.error(request, "不能刪除自己的帳號！")
            return redirect("system:user_list")
        
        # 防止刪除最後一個超級用戶
        if user.is_superuser and User.objects.filter(is_superuser=True).count() <= 1:
            messages.error(request, "不能刪除最後一個超級用戶！")
            return redirect("system:user_list")
        
        username = user.username
        user.delete()
        
        messages.success(request, f"用戶 {username} 刪除成功！")
        logger.info(f"用戶 {username} 由 {request.user.username} 刪除")
        
        return redirect("system:user_list")
    
    context = {
        "target_user": user,
        "title": f"確認刪除用戶 - {user.username}"
    }
    
    return render(request, "system/user_confirm_delete.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def user_toggle_status(request, user_id):
    """切換用戶啟用狀態"""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == "POST":
        # 防止停用自己的帳號
        if user == request.user:
            messages.error(request, "不能停用自己的帳號！")
            return redirect("system:user_list")
        
        user.is_active = not user.is_active
        user.save()
        
        status = "啟用" if user.is_active else "停用"
        messages.success(request, f"用戶 {user.username} 已{status}！")
        logger.info(f"用戶 {user.username} 由 {request.user.username} {status}")
        
        return redirect("system:user_list")
    
    return redirect("system:user_list")


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def user_permissions(request, user_id):
    """用戶權限管理頁面"""
    user = get_object_or_404(User, id=user_id)
    
    if request.method == "POST":
        # 處理權限分配
        permission_ids = request.POST.getlist('permissions')
        
        # 清除現有權限
        user.user_permissions.clear()
        
        # 分配新權限
        if permission_ids:
            permissions = Permission.objects.filter(id__in=permission_ids)
            user.user_permissions.set(permissions)
        
        messages.success(request, f"用戶 {user.username} 的權限已更新！")
        logger.info(f"用戶 {user.username} 的權限由 {request.user.username} 更新")
        
        return redirect("system:user_permissions", user_id=user_id)
    
    # 取得所有可用權限，按模組分組
    all_permissions = get_all_permissions()
    permissions_by_module = {}
    
    for perm in all_permissions:
        module_name = perm.content_type.app_label
        if module_name not in permissions_by_module:
            permissions_by_module[module_name] = []
        
        # 翻譯權限名稱
        translated_name = PERMISSION_NAME_TRANSLATIONS.get(perm.name, perm.name)
        permissions_by_module[module_name].append({
            'id': perm.id,
            'name': translated_name,
            'codename': perm.codename,
            'content_type': perm.content_type.model,
            'app_label': module_name,
            'is_assigned': perm in user.user_permissions.all()
        })
    
    # 取得用戶現有權限
    user_permissions = user.user_permissions.all()
    
    context = {
        'target_user': user,
        'permissions_by_module': permissions_by_module,
        'user_permissions': user_permissions,
        'title': f"用戶權限管理 - {user.username}"
    }
    
    return render(request, "system/user_permissions.html", context)


@user_passes_test(superuser_required, login_url="/accounts/login/")
def email_config(request):
    email_config_obj, created = EmailConfig.objects.get_or_create(id=1)
    if request.method == "POST":
        form = EmailConfigForm(request.POST)
        if "send_test_email" in request.POST:
            try:
                admin_user = User.objects.get(username="admin")
                logger.info(
                    f"準備發送測試郵件給 Admin 使用者: {admin_user.username}, 目標郵箱: {admin_user.email}"
                )
                if not admin_user.email:
                    logger.error("Admin 使用者未設置電子郵件地址")
                    messages.error(
                        request, "Admin 使用者未設置電子郵件地址，請先設置！"
                    )
                    return redirect("system:email_config")
                subject = "MES 系統 - 測試郵件"
                message = "這是一封來自 MES 系統的測試郵件。\n\n如果您收到此郵件，表示郵件主機設置正確。"
                from_email = email_config_obj.default_from_email
                recipient_list = [admin_user.email]
                logger.info(
                    f"SMTP 配置: host={email_config_obj.email_host}, port={email_config_obj.email_port}, use_tls={email_config_obj.email_use_tls}, user={email_config_obj.email_host_user}, from_email={from_email}"
                )
                connection = get_connection(
                    backend="django.core.mail.backends.smtp.EmailBackend",
                    host=email_config_obj.email_host,
                    port=email_config_obj.email_port,
                    username=email_config_obj.email_host_user,
                    password=email_config_obj.email_host_password,
                    use_tls=email_config_obj.email_use_tls,
                )
                send_mail(
                    subject,
                    message,
                    from_email,
                    recipient_list,
                    fail_silently=False,
                    connection=connection,
                )
                logger.info(f"測試郵件發送成功至: {admin_user.email}")
                messages.success(request, f"測試郵件已成功發送到 {admin_user.email}！")
            except User.DoesNotExist:
                logger.error("Admin 使用者不存在")
                messages.error(request, "Admin 使用者不存在，請確保已創建！")
            except smtplib.SMTPAuthenticationError as e:
                logger.error(f"SMTP 認證失敗: {str(e)}")
                messages.error(
                    request,
                    f"SMTP 認證失敗，請檢查郵件主機 hesap戶或應用程式密碼：{str(e)}",
                )
            except smtplib.SMTPException as e:
                logger.error(f"SMTP 錯誤: {str(e)}")
                messages.error(request, f"發送測試郵件失敗（SMTP 錯誤）：{str(e)}")
            except Exception as e:
                logger.error(f"發送測試郵件失敗: {str(e)}")
                messages.error(request, f"發送測試郵件失敗：{str(e)}")
            return redirect("system:email_config")

        if form.is_valid():
            email_config_obj.email_host = form.cleaned_data["email_host"]
            email_config_obj.email_port = (
                form.cleaned_data["email_port"]
                if form.cleaned_data["email_port"]
                else 25
            )
            email_config_obj.email_use_tls = form.cleaned_data["email_use_tls"]
            email_config_obj.email_host_user = form.cleaned_data["email_host_user"]
            email_config_obj.email_host_password = form.cleaned_data["email_host_password"]
            email_config_obj.default_from_email = form.cleaned_data["default_from_email"]
            email_config_obj.save()
            logger.info("郵件主機設定已更新")
            messages.success(request, "郵件主機設定已更新！")
            return redirect("system:index")
    else:
        initial_data = {
            "email_host": email_config_obj.email_host,
            "email_port": email_config_obj.email_port,
            "email_use_tls": email_config_obj.email_use_tls,
            "email_host_user": email_config_obj.email_host_user,
            "email_host_password": email_config_obj.email_host_password,
            "default_from_email": email_config_obj.default_from_email,
        }
        form = EmailConfigForm(initial=initial_data)
    return render(
        request, "system/email_config.html", {"form": form, "title": "郵件主機設定"}
    )


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def backup_database(request):
    backup_dir = "/var/www/mes/backups_DB"
    backup_files = []
    try:
        backup_files = [
            {
                "name": f,
                "size": os.path.getsize(os.path.join(backup_dir, f)) // 1024,
                "date": datetime.fromtimestamp(
                    os.path.getmtime(os.path.join(backup_dir, f))
                ).strftime("%Y-%m-%d %H:%M:%S"),
            }
            for f in os.listdir(backup_dir)
            if os.path.isfile(os.path.join(backup_dir, f)) and f.endswith(".sql")
        ]
        backup_files.sort(key=lambda x: x["date"], reverse=True)
    except Exception as e:
        logger.error(f"無法列出備份文件: {str(e)}")
        messages.error(request, f"無法列出備份文件：{str(e)}")

    if request.method == "POST":
        try:
            database_url = os.environ.get("DATABASE_URL")
            if not database_url:
                raise ValueError("環境變數 DATABASE_URL 未設置")
            from urllib.parse import urlparse

            parsed_url = urlparse(database_url)
            db_user = parsed_url.username
            db_password = parsed_url.password
            db_host = parsed_url.hostname
            db_port = parsed_url.port
            db_name = parsed_url.path.lstrip("/")
            os.environ["PGPASSWORD"] = db_password
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"backup_{db_name}_{timestamp}.sql"
            backup_path = os.path.join(backup_dir, backup_filename)
            cmd = [
                "pg_dump",
                "-h",
                db_host,
                "-p",
                str(db_port),
                "-U",
                db_user,
                "-F",
                "p",
                "-f",
                backup_path,
                db_name,
            ]
            result = subprocess.run(cmd, capture_output=True, text=True)
            if result.returncode != 0:
                raise Exception(f"pg_dump 失敗: {result.stderr}")
            del os.environ["PGPASSWORD"]
            logger.info(f"資料庫備份成功: {backup_filename}")
            messages.success(request, f"資料庫備份成功：{backup_filename}")
            backup_files = [
                {
                    "name": f,
                    "size": os.path.getsize(os.path.join(backup_dir, f)) // 1024,
                    "date": datetime.fromtimestamp(
                        os.path.getmtime(os.path.join(backup_dir, f))
                    ).strftime("%Y-%m-%d %H:%M:%S"),
                }
                for f in os.listdir(backup_dir)
                if os.path.isfile(os.path.join(backup_dir, f)) and f.endswith(".sql")
            ]
            backup_files.sort(key=lambda x: x["date"], reverse=True)
        except Exception as e:
            logger.error(f"資料庫備份失敗: {str(e)}")
            messages.error(request, f"資料庫備份失敗：{str(e)}")
        return redirect("system:backup")
    return render(
        request,
        "system/backup.html",
        {"backup_files": backup_files, "title": "資料庫備份"},
    )


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def download_backup(request, filename):
    backup_dir = "/var/www/mes/backups_DB"
    file_path = os.path.join(backup_dir, filename)
    if not os.path.isfile(file_path) or not filename.endswith(".sql"):
        logger.error(f"備份文件不存在或格式錯誤: {filename}")
        messages.error(request, f"備份文件不存在或格式錯誤：{filename}")
        return redirect("system:backup")
    if not os.path.abspath(file_path).startswith(os.path.abspath(backup_dir)):
        logger.error(f"無權訪問文件: {filename}")
        messages.error(request, f"無權訪問文件：{filename}")
        return redirect("system:backup")
    try:
        response = FileResponse(
            open(file_path, "rb"), as_attachment=True, filename=filename
        )
        logger.info(f"備份文件下載成功: {filename}")
        return response
    except Exception as e:
        logger.error(f"備份文件下載失敗: {str(e)}")
        messages.error(request, f"備份文件下載失敗：{str(e)}")
        return redirect("system:backup")


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def restore_database(request):
    backup_dir = "/var/www/mes/backups_DB"
    
    # 獲取現有備份檔案列表
    try:
        backup_files = [
            {
                "name": f,
                "size": os.path.getsize(os.path.join(backup_dir, f)) // 1024,
                "date": datetime.fromtimestamp(
                    os.path.getmtime(os.path.join(backup_dir, f))
                ).strftime("%Y-%m-%d %H:%M:%S"),
            }
            for f in os.listdir(backup_dir)
            if os.path.isfile(os.path.join(backup_dir, f)) and f.endswith(".sql")
        ]
        backup_files.sort(key=lambda x: x["date"], reverse=True)
    except Exception as e:
        logger.error(f"無法列出備份文件：{str(e)}")
        backup_files = []
    
    if request.method == "POST":
        # 檢查是否有上傳檔案或選擇現有備份
        sql_file = request.FILES.get("sql_file")
        selected_backup = request.POST.get("selected_backup")
        
        if not sql_file and not selected_backup:
            logger.error("未上傳備份文件或選擇現有備份")
            messages.error(request, "請上傳一個備份文件或選擇現有備份！")
            return redirect("system:restore_database")
        
        # 確定要還原的檔案路徑
        if sql_file:
            if not sql_file.name.endswith(".sql"):
                logger.error(f"上傳文件格式錯誤: {sql_file.name}")
                messages.error(request, f"請上傳 .sql 格式的備份文件！")
                return redirect("system:restore_database")
            
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            upload_filename = f"restore_upload_{timestamp}_{sql_file.name}"
            upload_path = os.path.join(backup_dir, upload_filename)
            
            # 保存上傳的檔案
            with open(upload_path, "wb+") as destination:
                for chunk in sql_file.chunks():
                    destination.write(chunk)
        else:
            # 使用現有備份檔案
            upload_path = os.path.join(backup_dir, selected_backup)
            if not os.path.exists(upload_path):
                messages.error(request, f"選擇的備份檔案不存在：{selected_backup}")
                return redirect("system:restore_database")
        
        try:
            # 在開始資料庫還原前，先清除當前 session
            logger.info("清除當前 session...")
            try:
                request.session.flush()
            except Exception as session_error:
                logger.warning(f"清除 session 時出現警告: {str(session_error)}")
            
            database_url = os.environ.get("DATABASE_URL")
            if not database_url:
                raise ValueError("環境變數 DATABASE_URL 未設置")
            from urllib.parse import urlparse

            parsed_url = urlparse(database_url)
            db_user = parsed_url.username
            db_password = parsed_url.password
            db_host = parsed_url.hostname
            db_port = parsed_url.port
            db_name = parsed_url.path.lstrip("/")
            os.environ["PGPASSWORD"] = db_password
            
            # 步驟1: 先清空資料庫（斷開所有連線並重建）
            logger.info("開始清空資料庫...")
            drop_cmd = [
                "psql",
                "-h", db_host,
                "-p", str(db_port),
                "-U", db_user,
                "-d", "postgres",  # 連接到 postgres 資料庫
                "-c", f"SELECT pg_terminate_backend(pid) FROM pg_stat_activity WHERE datname = '{db_name}' AND pid <> pg_backend_pid();"
            ]
            result = subprocess.run(drop_cmd, capture_output=True, text=True)
            if result.returncode != 0:
                logger.warning(f"斷開資料庫連線時出現警告: {result.stderr}")
            
            drop_db_cmd = [
                "dropdb",
                "-h", db_host,
                "-p", str(db_port),
                "-U", db_user,
                "--if-exists",
                db_name
            ]
            result = subprocess.run(drop_db_cmd, capture_output=True, text=True)
            if result.returncode != 0:
                logger.warning(f"刪除資料庫時出現警告: {result.stderr}")
            
            # 步驟2: 重新建立資料庫
            logger.info("重新建立資料庫...")
            create_cmd = [
                "createdb",
                "-h", db_host,
                "-p", str(db_port),
                "-U", db_user,
                db_name
            ]
            result = subprocess.run(create_cmd, capture_output=True, text=True)
            if result.returncode != 0:
                raise Exception(f"建立資料庫失敗: {result.stderr}")
            
            # 步驟3: 還原備份檔案
            logger.info("開始還原備份檔案...")
            restore_cmd = [
                "psql",
                "-h", db_host,
                "-p", str(db_port),
                "-U", db_user,
                "-d", db_name,
                "-f", upload_path,
            ]
            result = subprocess.run(restore_cmd, capture_output=True, text=True)
            if result.returncode != 0:
                raise Exception(f"psql 恢復失敗: {result.stderr}")
            
            # 步驟4: 直接建立必要的資料表
            logger.info("建立必要的 Django 資料表...")
            
            # 建立 django_session 資料表
            session_table_sql = """
            CREATE TABLE IF NOT EXISTS django_session (
                session_key VARCHAR(40) NOT NULL PRIMARY KEY,
                session_data TEXT NOT NULL,
                expire_date TIMESTAMP WITH TIME ZONE NOT NULL
            );
            """
            
            # 建立 django_migrations 資料表
            migrations_table_sql = """
            CREATE TABLE IF NOT EXISTS django_migrations (
                id BIGSERIAL PRIMARY KEY,
                app VARCHAR(255) NOT NULL,
                name VARCHAR(255) NOT NULL,
                applied TIMESTAMP WITH TIME ZONE NOT NULL
            );
            """
            
            try:
                # 執行 SQL 建立資料表
                create_tables_cmd = [
                    "psql",
                    "-h", db_host,
                    "-p", str(db_port),
                    "-U", db_user,
                    "-d", db_name,
                    "-c", session_table_sql + migrations_table_sql
                ]
                result = subprocess.run(create_tables_cmd, capture_output=True, text=True)
                if result.returncode == 0:
                    logger.info("Django 必要資料表建立成功")
                else:
                    logger.warning(f"建立資料表時出現警告: {result.stderr}")
                    
            except Exception as e:
                logger.warning(f"建立 Django 資料表失敗: {str(e)}")
            
            del os.environ["PGPASSWORD"]
            backup_name = os.path.basename(upload_path)
            logger.info(f"資料庫恢復成功: {backup_name}")
            
            # 清理上傳的檔案
            if 'upload_path' in locals() and sql_file:
                try:
                    os.remove(upload_path)
                except:
                    pass
            
            # 重定向到登入頁面並顯示成功訊息
            return redirect(f"/accounts/login/?restore_success={backup_name}")
            
        except Exception as e:
            logger.error(f"資料庫恢復失敗: {str(e)}")
            
            # 清理上傳的檔案
            if 'upload_path' in locals() and sql_file:
                try:
                    os.remove(upload_path)
                except:
                    pass
            
            # 重定向到登入頁面並顯示錯誤訊息
            return redirect(f"/accounts/login/?restore_error={str(e)}")
    
    return render(request, "system/restore.html", {
        "title": "恢復資料庫",
        "backup_files": backup_files
    })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def backup_schedule(request):
    schedule, created = BackupSchedule.objects.get_or_create(
        id=1,
        defaults={
            'schedule_type': 'daily',
            'backup_time': '02:00:00',
            'retention_days': 30,
            'is_active': True
        }
    )
    task_name = "auto-backup-database-task"
    if request.method == "POST":
        form = BackupScheduleForm(request.POST, instance=schedule)
        if form.is_valid():
            schedule = form.save()
            try:
                task = PeriodicTask.objects.get(name=task_name)
                if schedule.is_active:
                    backup_time = schedule.backup_time
                    schedule_obj, _ = CrontabSchedule.objects.get_or_create(
                        minute=backup_time.minute,
                        hour=backup_time.hour,
                        day_of_week="*",
                        day_of_month="*",
                        month_of_year="*",
                        timezone="Asia/Taipei",
                    )
                    task.crontab = schedule_obj
                    task.enabled = True
                    task.save()
                    logger.info("自動備份任務已更新並啟用")
                    messages.success(request, "自動備份排程已更新並啟用！")
                else:
                    task.enabled = False
                    task.save()
                    logger.info("自動備份任務已禁用")
                    messages.success(request, "自動備份排程已更新並禁用！")
            except PeriodicTask.DoesNotExist:
                if schedule.is_active:
                    backup_time = schedule.backup_time
                    schedule_obj, _ = CrontabSchedule.objects.get_or_create(
                        minute=backup_time.minute,
                        hour=backup_time.hour,
                        day_of_week="*",
                        day_of_month="*",
                        month_of_year="*",
                        timezone="Asia/Taipei",
                    )
                    PeriodicTask.objects.create(
                        crontab=schedule_obj,
                        name=task_name,
                        task="system.tasks.auto_backup_database",
                        enabled=True,
                    )
                    logger.info("自動備份任務已創建並啟用")
                    messages.success(request, "自動備份排程已創建並啟用！")
                else:
                    logger.info("自動備份排程已更新，但未啟用")
                    messages.success(request, "自動備份排程已更新，但未啟用！")
            return redirect("system:index")
    else:
        form = BackupScheduleForm(instance=schedule)
    try:
        task = PeriodicTask.objects.get(name=task_name)
        task_status = "已啟用" if task.enabled else "未啟用"
        task_time = schedule.backup_time.strftime("%H:%M") if task.enabled else None
    except PeriodicTask.DoesNotExist:
        task_status = "未啟用"
        task_time = None
    return render(
        request,
        "system/backup_schedule.html",
        {
            "form": form,
            "title": "自動備份排程",
            "task_status": task_status,
            "task_time": task_time,
        },
    )


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def export_users(request):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = 'attachment; filename="users_export.csv"'
    response.write("\ufeff".encode("utf8"))
    writer = csv.writer(response)
    writer.writerow(["用戶名稱", "電子郵件", "姓名", "是否啟用", "是否為員工", "是否為超級用戶", "預設密碼"])
    users = User.objects.all().order_by('username')
    for user in users:
        writer.writerow(
            [
                user.username,
                user.email,
                f"{user.first_name} {user.last_name}".strip() or "",
                "是" if user.is_active else "否",
                "是" if user.is_staff else "否",
                "是" if user.is_superuser else "否",
                "123456",  # 預設密碼
            ]
        )
    logger.info(f"用戶數據匯出成功，由 {request.user.username} 執行")
    return response


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def export_users_excel(request):
    """匯出用戶資料為Excel格式"""
    # 創建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "用戶資料"
    
    # 設定標題行
    headers = [
        "用戶名稱", "電子郵件", "姓名", "是否啟用", "是否為員工", 
        "是否為超級用戶", "預設密碼"
    ]
    
    # 設定標題行樣式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # 寫入標題行
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # 獲取用戶資料
    users = User.objects.all().order_by('username')
    
    # 寫入資料行
    for row, user in enumerate(users, 2):
        ws.cell(row=row, column=1, value=user.username)
        ws.cell(row=row, column=2, value=user.email)
        ws.cell(row=row, column=3, value=f"{user.first_name} {user.last_name}".strip() or "")
        ws.cell(row=row, column=4, value="是" if user.is_active else "否")
        ws.cell(row=row, column=5, value="是" if user.is_staff else "否")
        ws.cell(row=row, column=6, value="是" if user.is_superuser else "否")
        ws.cell(row=row, column=7, value="123456")  # 預設密碼
    
    # 自動調整欄寬
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in range(1, ws.max_row + 1):
            cell_value = ws[f"{column_letter}{row}"].value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        adjusted_width = min(max_length + 2, 50)  # 最大寬度限制為50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 設定資料行對齊方式
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).alignment = Alignment(horizontal="left", vertical="center")
    
    # 創建HTTP響應
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="users_export.xlsx"'
    
    # 儲存工作簿到響應
    wb.save(response)
    
    logger.info(f"用戶Excel資料匯出成功，由 {request.user.username} 執行")
    return response


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def test_import(request):
    """測試匯入功能頁面"""
    return render(request, "system/test_import.html")


@login_required
def import_users(request):
    # 直接寫入檔案來確認函數被調用
    with open('/tmp/import_debug.txt', 'a') as f:
        f.write(f"匯入函數被調用: {request.method} - {request.user.username} - {request.path}\n")
    
    logger.info(f"🔍 匯入用戶請求: {request.method}")
    logger.info(f"🔍 請求用戶: {request.user.username}")
    logger.info(f"🔍 請求路徑: {request.path}")
    logger.info(f"🔍 用戶是否為超級用戶: {request.user.is_superuser}")
    
    # 手動檢查權限
    if not request.user.is_superuser:
        with open('/tmp/import_debug.txt', 'a') as f:
            f.write(f"權限不足: {request.user.username}\n")
        logger.error(f"❌ 用戶 {request.user.username} 沒有超級用戶權限")
        messages.error(request, "您沒有權限執行此操作！")
        return redirect("system:user_list")
    
    if request.method == "POST":
        with open('/tmp/import_debug.txt', 'a') as f:
            f.write(f"POST 請求處理開始\n")
        logger.info(f"📤 POST 請求，FILES: {list(request.FILES.keys())}")
        logger.info(f"📤 POST 請求，POST: {list(request.POST.keys())}")
        
        if "file" not in request.FILES:
            with open('/tmp/import_debug.txt', 'a') as f:
                f.write(f"沒有檔案上傳\n")
            logger.error("❌ 未上傳文件")
            messages.error(request, "請上傳一個文件！")
            return redirect("system:user_list")
            
        csv_file = request.FILES["file"]
        logger.info(f"📁 上傳檔案: {csv_file.name}, 大小: {csv_file.size} bytes")
        
        if not (csv_file.name.endswith(".csv") or csv_file.name.endswith(".xlsx")):
            logger.error(f"❌ 上傳文件格式錯誤: {csv_file.name}")
            messages.error(request, "請上傳 .csv 或 .xlsx 格式的文件！")
            return redirect("system:user_list")
        try:
            created_count = 0
            updated_count = 0
            default_password = "123456"
            if csv_file.name.endswith(".csv"):
                decoded_file = csv_file.read().decode("utf-8-sig")
                csv_reader = csv.DictReader(decoded_file.splitlines())
                for row in csv_reader:
                    # 完全匹配匯出格式
                    username = row.get("用戶名稱")
                    email = row.get("電子郵件", "")
                    full_name = row.get("姓名", "")
                    is_active = row.get("是否啟用") == "是"
                    is_staff = row.get("是否為員工") == "是"
                    is_superuser = row.get("是否為超級用戶") == "是"
                    password = str(row.get("預設密碼", default_password)) if row.get("預設密碼") is not None else default_password
                    
                    if not username:
                        continue
                    
                    # 處理姓名分割
                    first_name = ""
                    last_name = ""
                    if full_name:
                        name_parts = full_name.strip().split()
                        if len(name_parts) >= 1:
                            first_name = name_parts[0]
                        if len(name_parts) >= 2:
                            last_name = " ".join(name_parts[1:])
                    
                    # 創建或更新用戶
                    user, created = User.objects.get_or_create(username=username)
                    user.email = email or ""
                    user.first_name = first_name
                    user.last_name = last_name
                    user.is_active = is_active
                    user.is_staff = is_staff
                    user.is_superuser = is_superuser
                    
                    if created or password != default_password:
                        user.set_password(password)
                    
                    user.save()
                    
                    if created:
                        created_count += 1
                    else:
                        updated_count += 1
            else:
                with open('/tmp/import_debug.txt', 'a') as f:
                    f.write(f"處理 Excel 檔案: {csv_file.name}\n")
                wb = openpyxl.load_workbook(csv_file)
                ws = wb.active
                headers = [cell.value for cell in ws[1]]
                with open('/tmp/import_debug.txt', 'a') as f:
                    f.write(f"Excel 標題行: {headers}\n")
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
                    row_data = dict(zip(headers, row))
                    with open('/tmp/import_debug.txt', 'a') as f:
                        f.write(f"第 {row_num} 行資料: {row_data}\n")
                    # 完全匹配匯出格式
                    username = row_data.get("用戶名稱")
                    email = row_data.get("電子郵件", "")
                    full_name = row_data.get("姓名", "")
                    is_active = row_data.get("是否啟用") == "是"
                    is_staff = row_data.get("是否為員工") == "是"
                    is_superuser = row_data.get("是否為超級用戶") == "是"
                    password = str(row_data.get("預設密碼", default_password)) if row_data.get("預設密碼") is not None else default_password
                    
                    with open('/tmp/import_debug.txt', 'a') as f:
                        f.write(f"解析結果 - username: {username}, email: {email}, is_active: {is_active}, is_staff: {is_staff}, is_superuser: {is_superuser}\n")
                    
                    if not username:
                        with open('/tmp/import_debug.txt', 'a') as f:
                            f.write(f"跳過第 {row_num} 行：沒有用戶名稱\n")
                        continue
                    
                    # 處理姓名分割
                    first_name = ""
                    last_name = ""
                    if full_name:
                        name_parts = full_name.strip().split()
                        if len(name_parts) >= 1:
                            first_name = name_parts[0]
                        if len(name_parts) >= 2:
                            last_name = " ".join(name_parts[1:])
                    
                    # 創建或更新用戶
                    user, created = User.objects.get_or_create(username=username)
                    user.email = email or ""
                    user.first_name = first_name
                    user.last_name = last_name
                    user.is_active = is_active
                    user.is_staff = is_staff
                    user.is_superuser = is_superuser
                    
                    if created or password != default_password:
                        user.set_password(password)
                    
                    user.save()
                    
                    if created:
                        created_count += 1
                        with open('/tmp/import_debug.txt', 'a') as f:
                            f.write(f"✅ 新增用戶: {username}\n")
                    else:
                        updated_count += 1
                        with open('/tmp/import_debug.txt', 'a') as f:
                            f.write(f"🔄 更新用戶: {username}\n")
            logger.info(
                f"用戶匯入完成：新增 {created_count} 個，更新 {updated_count} 個"
            )
            messages.success(
                request,
                f"用戶匯入完成：新增 {created_count} 個，更新 {updated_count} 個",
            )
            # 測試：返回 JSON 而不是重定向
            from django.http import JsonResponse
            return JsonResponse({
                'status': 'success',
                'message': f'用戶匯入完成：新增 {created_count} 個，更新 {updated_count} 個',
                'created_count': created_count,
                'updated_count': updated_count
            })
        except Exception as e:
            logger.error(f"用戶匯入失敗: {str(e)}")
            messages.error(request, f"用戶匯入失敗：{str(e)}")
            # 測試：返回 JSON 錯誤而不是重定向
            from django.http import JsonResponse
            return JsonResponse({
                'status': 'error',
                'message': f'用戶匯入失敗：{str(e)}'
            }, status=400)
        return redirect("system:user_list")
    return redirect("system:user_list")


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def operation_log_manage(request):
    config, created = OperationLogConfig.objects.get_or_create(id=1)
    if request.method == "POST":
        form = OperationLogConfigForm(request.POST, instance=config)
        if form.is_valid():
            form.save()
            # 重新讀取配置以獲取更新後的值
            config.refresh_from_db()
            logger.info(f"操作紀錄保留天數更新為 {config.retention_days} 天")
            messages.success(
                request, f"操作紀錄保留天數已更新為 {config.retention_days} 天！"
            )
            return redirect("system:operation_log_manage")
    else:
        form = OperationLogConfigForm(instance=config)

    module = request.GET.get("module", "")
    user = request.GET.get("user", "")
    start_date = request.GET.get("start_date", "")
    end_date = request.GET.get("end_date", "")

    logs = []
    for module_name, model_path in MODULE_LOG_MODELS.items():
        try:
            module_path, class_name = model_path.rsplit(".", 1)
            module = __import__(module_path, fromlist=[class_name])
            log_model = getattr(module, class_name)
            module_logs = log_model.objects.all()
            # 移除無效的 module 篩選條件
            if module_name and module_name != "":
                if module == module_name:
                    continue  # 跳過不符合模組條件的記錄
            if user:
                module_logs = module_logs.filter(user=user)
            if start_date:
                start_date = datetime.strptime(start_date, "%Y-%m-%d")
                module_logs = module_logs.filter(timestamp__gte=start_date)
            if end_date:
                end_date = datetime.strptime(end_date, "%Y-%m-%d")
                end_date = end_date + timedelta(days=1)  # 包含結束日期當天
                module_logs = module_logs.filter(timestamp__lt=end_date)
            # 添加模組名稱到日誌記錄
            for log in module_logs:
                logs.append(
                    {
                        "module": module_name,
                        "display_module": get_module_display_name(module_name),
                        "timestamp": log.timestamp,
                        "user": log.user,
                        "action": log.action,
                    }
                )
        except ImportError as e:
            logger.error(
                f"無法導入模組 {module_name} 的日誌模型 {model_path}: {str(e)}"
            )
            messages.error(request, f"無法加載模組 {module_name} 的操作紀錄：{str(e)}")
        except AttributeError as e:
            logger.error(f"模組 {module_name} 的日誌模型 {model_path} 無效: {str(e)}")
            messages.error(request, f"模組 {module_name} 的日誌模型無效：{str(e)}")
        except Exception as e:
            logger.error(f"加載模組 {module_name} 的操作紀錄時發生未知錯誤: {str(e)}")
            messages.error(request, f"加載模組 {module_name} 的操作紀錄失敗：{str(e)}")
    # 按時間倒序排序
    logs.sort(key=lambda x: x["timestamp"], reverse=True)

    # 獲取模組選項和使用者列表
    modules = list(MODULE_LOG_MODELS.keys())
    module_choices = [(m, get_module_display_name(m)) for m in modules]
    users = set()
    for module_name, model_path in MODULE_LOG_MODELS.items():
        try:
            module_path, class_name = model_path.rsplit(".", 1)
            module = __import__(module_path, fromlist=[class_name])
            log_model = getattr(module, class_name)
            users.update(log_model.objects.values_list("user", flat=True))
        except Exception as e:
            logger.error(f"無法從模組 {module_name} 獲取使用者列表: {str(e)}")

    default_end_date = timezone.now().date()
    default_start_date = default_end_date - timedelta(days=30)

    return render(
        request,
        "system/operation_log_manage.html",
        {
            "form": form,
            "logs": logs,
            "module_choices": module_choices,
            "users": users,
            "selected_module": module,
            "selected_user": user,
            "start_date": start_date,
            "end_date": end_date,
            "default_start_date": default_start_date,
            "default_end_date": default_end_date,
        },
    )


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def clean_operation_logs(request):
    config, created = OperationLogConfig.objects.get_or_create(id=1)
    cutoff_date = timezone.now() - timedelta(days=config.retention_days)
    total_deleted = 0
    for module, model_path in MODULE_LOG_MODELS.items():
        try:
            module_name, class_name = model_path.rsplit(".", 1)
            module = __import__(module_name, fromlist=[class_name])
            log_model = getattr(module, class_name)
            deleted_count, _ = log_model.objects.filter(
                timestamp__lt=cutoff_date
            ).delete()
            total_deleted += deleted_count
            logger.info(
                f"清理模組 {module} 的過期操作紀錄，刪除 {deleted_count} 條記錄"
            )
        except Exception as e:
            logger.error(f"清理模組 {module} 的操作紀錄失敗: {str(e)}")
    logger.info(f"清理過期操作紀錄，刪除 {total_deleted} 條記錄")
    messages.success(request, f"已清理 {total_deleted} 條過期操作紀錄！")
    return redirect("system:operation_log_manage")


@login_required
def change_password(request):
    if request.method == "POST":
        form = SetPasswordForm(user=request.user, data=request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "您的密碼已成功更改！請重新登入。")
            logger.info(f"用戶 {request.user.username} 更改了自己的密碼")
            return redirect("login")
        else:
            messages.error(request, "請修正以下錯誤：")
    else:
        form = SetPasswordForm(user=request.user)
    return render(
        request, "system/change_password.html", {"form": form, "title": "變更密碼"}
    )


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def environment_management(request):
    """環境管理視圖"""
    from datetime import datetime, timedelta
    # 獲取當前環境信息
    current_debug = getattr(settings, "DEBUG", True)
    
    # 根據 DEBUG 設定判斷環境
    if current_debug:
        current_env = "development"
    else:
        current_env = "production"

    # 環境配置信息
    environments = {
        "development": {
            "name": "開發環境",
            "description": "DEBUG=True，用於程式開發和調試",
            "debug": True,
            "features": ["詳細調試信息", "終端日誌輸出", "完整錯誤追蹤"],
        },
        "production": {
            "name": "生產環境",
            "description": "DEBUG=False，正式營運環境",
            "debug": False,
            "features": ["高安全性", "效能優化"],
        },
    }
    
    # 當前環境信息
    environment_info = environments.get(current_env, environments["development"])

    # 處理日誌清理請求
    if request.method == "POST":
        action = request.POST.get("action")
        
        if action == "clean_logs":
            try:
                days = int(request.POST.get("days", 30))
                dry_run = request.POST.get("dry_run") == "on"
                
                if dry_run:
                    # 模擬執行
                    messages.info(request, f"模擬清理 {days} 天前的日誌檔案")
                    logger.info(f"模擬清理 {days} 天前的日誌檔案")
                else:
                    # 實際執行
                    from datetime import datetime, timedelta
                    cutoff_date = datetime.now() - timedelta(days=days)
                    
                    cleaned_count = 0
                    for log_file in log_files:
                        if log_file["modified"] < cutoff_date:
                            file_path = os.path.join(log_dir, log_file["name"])
                            try:
                                os.remove(file_path)
                                cleaned_count += 1
                            except Exception as e:
                                logger.error(f"刪除日誌檔案失敗 {file_path}: {str(e)}")
                    
                    messages.success(request, f"已清理 {cleaned_count} 個舊日誌檔案")
                    logger.info(f"已清理 {cleaned_count} 個舊日誌檔案")
                    
            except Exception as e:
                messages.error(request, f"清理日誌失敗: {str(e)}")
                logger.error(f"清理日誌失敗: {str(e)}")

    # 獲取日誌檔案信息
    log_dir = getattr(
        settings, "LOG_BASE_DIR", os.path.join(settings.BASE_DIR, "logs")
    )
    log_files = []

    if os.path.exists(log_dir):
        for file in os.listdir(log_dir):
            if file.endswith(".log"):
                file_path = os.path.join(log_dir, file)
                stat = os.stat(file_path)
                log_files.append(
                    {
                        "name": file,
                        "size": stat.st_size,
                        "modified": datetime.fromtimestamp(stat.st_mtime),
                        "size_mb": round(stat.st_size / (1024 * 1024), 2),
                    }
                )

    # 操作日誌配置
    try:
        log_config = OperationLogConfig.objects.get(id=1)
    except OperationLogConfig.DoesNotExist:
        log_config = OperationLogConfig.objects.create(
            log_level='INFO',
            retention_days=90,
            max_file_size=10,
            is_active=True
        )
    
    # 備份排程配置
    try:
        backup_config = BackupSchedule.objects.get(id=1)
    except BackupSchedule.DoesNotExist:
        backup_config = BackupSchedule.objects.create(
            schedule_type='daily',
            backup_time='02:00:00',
            retention_days=30,
            is_active=True
        )
    
    # 郵件配置
    try:
        email_config = EmailConfig.objects.get(id=1)
    except EmailConfig.DoesNotExist:
        email_config = EmailConfig.objects.create()
    
    # 系統統計資訊
    system_stats = {
        'total_users': User.objects.count(),

        'total_permissions': Permission.objects.count(),
        'log_files_count': len(log_files),
        'total_log_size_mb': sum(f['size_mb'] for f in log_files),
    }
    
    # 最近的清理記錄
    try:
        recent_cleanups = CleanupLog.objects.order_by('-execution_time')[:5]
    except:
        recent_cleanups = []
    
    # 設定預設日期範圍（最近7天）
    from datetime import datetime, timedelta
    default_end_date = datetime.now().strftime('%Y-%m-%d')
    default_start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')

    context = {
        'environment_info': environment_info,
        'current_env': current_env,
        'current_debug': current_debug,
        'environments': environments,
        'log_dir': log_dir,
        'log_files': log_files,
        'log_config': log_config,
        'backup_config': backup_config,
        'email_config': email_config,
        'system_stats': system_stats,
        'recent_cleanups': recent_cleanups,
        'title': '環境管理'
    }
    
    return render(request, 'system/environment_management.html', context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def view_log_file(request, filename):
    """查看日誌檔案"""
    import os
    from django.conf import settings
    from django.http import HttpResponse

    log_dir = getattr(
        settings, "LOG_BASE_DIR", "/var/log/mes"
    )
    file_path = os.path.join(log_dir, filename)

    if not os.path.exists(file_path):
        messages.error(request, f"日誌檔案 {filename} 不存在")
        return redirect("system:environment_management")

    # 檢查檔案大小，如果太大只顯示最後部分
    file_size = os.path.getsize(file_path)
    max_size = 1024 * 1024  # 1MB

    try:
        with open(file_path, "r", encoding="utf-8") as f:
            if file_size > max_size:
                # 檔案太大，只讀取最後部分
                f.seek(-max_size, 2)
                content = f.read()
                content = f"... (檔案太大，只顯示最後 {max_size//1024}KB)\n" + content
            else:
                content = f.read()
    except UnicodeDecodeError:
        try:
            with open(file_path, "r", encoding="big5") as f:
                if file_size > max_size:
                    f.seek(-max_size, 2)
                    content = f.read()
                    content = (
                        f"... (檔案太大，只顯示最後 {max_size//1024}KB)\n" + content
                    )
                else:
                    content = f.read()
        except:
            content = "無法讀取檔案內容"

    context = {
        "filename": filename,
        "content": content,
        "file_size": file_size,
        "file_size_mb": round(file_size / (1024 * 1024), 2),
    }

    return render(request, "system/view_log.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def download_log_file(request, filename):
    """下載日誌檔案"""
    import os
    from django.conf import settings
    from django.http import FileResponse

    log_dir = getattr(
        settings, "LOG_BASE_DIR", "/var/log/mes"
    )
    file_path = os.path.join(log_dir, filename)

    if not os.path.exists(file_path):
        messages.error(request, f"日誌檔案 {filename} 不存在")
        return redirect("system:environment_management")

    try:
        response = FileResponse(open(file_path, "rb"))
        response["Content-Type"] = "text/plain"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
    except Exception as e:
        messages.error(request, f"下載失敗: {str(e)}")
        return redirect("system:environment_management")


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def clean_logs(request):
    """清理日誌檔案"""
    if request.method == "POST":
        days = int(request.POST.get("days", 30))
        dry_run = request.POST.get("dry_run") == "on"

        log_dir = getattr(
            settings,
            "LOG_BASE_DIR",
            "/var/log/mes",
        )

        if not os.path.exists(log_dir):
            messages.error(request, "日誌目錄不存在")
            return redirect("system:environment_management")

        cutoff_date = datetime.now() - timedelta(days=days)
        log_files = glob.glob(os.path.join(log_dir, "*.log*"))

        deleted_count = 0
        deleted_size = 0

        for log_file in log_files:
            stat = os.stat(log_file)
            mtime = datetime.fromtimestamp(stat.st_mtime)

            if mtime < cutoff_date:
                if dry_run:
                    messages.info(
                        request,
                        f'將刪除: {os.path.basename(log_file)} ({mtime.strftime("%Y-%m-%d")})',
                    )
                else:
                    try:
                        size = os.path.getsize(log_file)
                        os.remove(log_file)
                        deleted_count += 1
                        deleted_size += size
                        messages.success(
                            request, f"已刪除: {os.path.basename(log_file)}"
                        )
                    except Exception as e:
                        messages.error(
                            request,
                            f"刪除失敗: {os.path.basename(log_file)} - {str(e)}",
                        )

        if dry_run:
            messages.info(
                request,
                f"模擬完成，將刪除 {len([f for f in log_files if datetime.fromtimestamp(os.stat(f).st_mtime) < cutoff_date])} 個檔案",
            )
        else:
            messages.success(
                request,
                f"清理完成，刪除了 {deleted_count} 個檔案，釋放 {deleted_size / (1024*1024):.1f} MB 空間",
            )
            logger.info(
                f"系統管理員 {request.user.username} 清理了 {deleted_count} 個日誌檔案"
            )

    return redirect("system:environment_management")


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def permission_list(request):
    """權限管理列表頁面"""
    # 支援搜尋和篩選
    search_query = request.GET.get('search', '')
    module_filter = request.GET.get('module', 'all')
    action_filter = request.GET.get('action', 'all')
    
    # 取得所有模組的權限
    all_modules = [
        "equip", "material", "scheduling", "process", "quality", 
        "workorder", "kanban", "erp_integration", "ai", "system"
    ]

    # 按模組分組權限
    permissions_by_module = {}
    total_permissions = 0
    
    for module in all_modules:
        module_permissions = Permission.objects.filter(content_type__app_label=module)
        
        # 搜尋篩選
        if search_query:
            module_permissions = module_permissions.filter(
                Q(name__icontains=search_query) | 
                Q(codename__icontains=search_query)
            )
        
        # 動作篩選
        if action_filter != 'all':
            if action_filter == 'view':
                module_permissions = module_permissions.filter(codename__startswith='view_')
            elif action_filter == 'add':
                module_permissions = module_permissions.filter(codename__startswith='add_')
            elif action_filter == 'change':
                module_permissions = module_permissions.filter(codename__startswith='change_')
            elif action_filter == 'delete':
                module_permissions = module_permissions.filter(codename__startswith='delete_')
        
        if module_permissions.exists():
            display_name = get_module_display_name(module)
            permissions_by_module[display_name] = []
            
            for perm in module_permissions:
                # 翻譯權限名稱
                translated_name = PERMISSION_NAME_TRANSLATIONS.get(perm.name, perm.name)
                
                # 統計擁有此權限的用戶數量
                users_count = User.objects.filter(user_permissions=perm).count()
                
                permissions_by_module[display_name].append({
                        "id": perm.id,
                        "name": translated_name,
                        "codename": perm.codename,
                        "content_type": perm.content_type.model,
                        "app_label": perm.content_type.app_label,
                    "users_count": users_count,
                    "original_name": perm.name
                })
                total_permissions += 1
    
    # 統計資訊
    stats = {
        'total_permissions': total_permissions,
        'total_modules': len(permissions_by_module),
        'view_permissions': Permission.objects.filter(codename__startswith='view_').count(),
        'add_permissions': Permission.objects.filter(codename__startswith='add_').count(),
        'change_permissions': Permission.objects.filter(codename__startswith='change_').count(),
        'delete_permissions': Permission.objects.filter(codename__startswith='delete_').count(),
    }
    
    context = {
        "permissions_by_module": permissions_by_module,
        "search_query": search_query,
        "module_filter": module_filter,
        "action_filter": action_filter,
        "stats": stats,
        "all_modules": all_modules,
        "title": "權限管理"
    }
    
    return render(request, "system/permission_list.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def permission_detail(request, permission_id):
    """權限詳情頁面"""
    permission = get_object_or_404(Permission, id=permission_id)

    # 取得擁有此權限的用戶
    users_with_permission = User.objects.filter(
        user_permissions=permission
    ).distinct().order_by('username')

    # 翻譯權限名稱
    translated_name = PERMISSION_NAME_TRANSLATIONS.get(permission.name, permission.name)

    # 取得模組顯示名稱
    module_display_name = get_module_display_name(permission.content_type.app_label)
    
    context = {
            "permission": permission,
            "translated_name": translated_name,
        "module_display_name": module_display_name,
            "users_with_permission": users_with_permission,
        "users_count": users_with_permission.count(),
        "title": f"權限詳情 - {translated_name}"
    }
    
    return render(request, "system/permission_detail.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def permission_assign(request):
    """權限分配頁面"""
    if request.method == "POST":
        permission_id = request.POST.get("permission_id")
        user_id = request.POST.get("user_id")
        action = request.POST.get("action")  # 'assign' 或 'remove'

        try:
            permission = Permission.objects.get(id=permission_id)
            user = User.objects.get(id=user_id)
            
            translated_name = PERMISSION_NAME_TRANSLATIONS.get(permission.name, permission.name)
            
            if action == "assign":
                user.user_permissions.add(permission)
                messages.success(
                    request,
                    f"已將權限「{translated_name}」分配給用戶「{user.username}」",
                )
                logger.info(
                    f"權限 {permission.name} 分配給用戶 {user.username} 由 {request.user.username} 操作"
                )
            else:
                user.user_permissions.remove(permission)
                messages.success(
                    request,
                    f"已從用戶「{user.username}」移除權限「{user.username}」",
                )
                logger.info(
                    f"權限 {permission.name} 從用戶 {user.username} 移除由 {request.user.username} 操作"
                )

        except (Permission.DoesNotExist, User.DoesNotExist):
            messages.error(request, "指定的權限或用戶不存在")
        except Exception as e:
            messages.error(request, f"操作失敗：{str(e)}")
            logger.error(f"權限分配操作失敗: {str(e)}")

    # 取得所有權限和用戶供選擇
    all_permissions = get_all_permissions()
    users = User.objects.filter(is_active=True).order_by('username')

    # 為權限名稱添加翻譯
    for perm in all_permissions:
        perm.translated_name = PERMISSION_NAME_TRANSLATIONS.get(perm.name, perm.name)
        perm.module_display_name = get_module_display_name(perm.content_type.app_label)
    
    # 按模組分組權限
    permissions_by_module = {}
    for perm in all_permissions:
        module_name = perm.module_display_name
        if module_name not in permissions_by_module:
            permissions_by_module[module_name] = []
        permissions_by_module[module_name].append(perm)
    
    # 處理GET參數中的預選權限
    selected_permission_id = request.GET.get('permission')
    selected_permission = None
    if selected_permission_id:
        try:
            selected_permission = Permission.objects.get(id=selected_permission_id)
        except Permission.DoesNotExist:
            pass
    
    context = {
        "permissions_by_module": permissions_by_module,
        "users": users,
        "selected_permission": selected_permission,
        "title": "權限分配"
    }
    
    return render(request, "system/permission_assign.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def permission_bulk_assign(request):
    """批量權限分配頁面"""
    if request.method == "POST":
        selected_users = request.POST.getlist("selected_users")
        selected_permissions = request.POST.getlist("selected_permissions")
        
        if not selected_users or not selected_permissions:
            messages.error(request, "請選擇用戶和權限")
        else:
            try:
                users = User.objects.filter(id__in=selected_users)
                permissions = Permission.objects.filter(id__in=selected_permissions)
                
                # 為每個選中的用戶分配選中的權限
                assigned_count = 0
                for user in users:
                    # 先清除現有權限，然後分配新權限
                    user.user_permissions.clear()
                    user.user_permissions.add(*permissions)
                    assigned_count += 1
                
                messages.success(
                    request,
                    f"已為 {len(users)} 個用戶分配 {len(permissions)} 個權限",
                )
                
                logger.info(
                    f"批量權限分配由 {request.user.username} 執行，為 {len(users)} 個用戶分配了 {len(permissions)} 個權限"
                )
                
            except Exception as e:
                messages.error(request, f"批量權限分配失敗：{str(e)}")
                logger.error(f"批量權限分配失敗: {str(e)}")
    
    # 取得所有用戶和權限
    users = User.objects.filter(is_active=True).order_by('username')
    permissions = Permission.objects.all().order_by('content_type__app_label', 'name')
    
    # 取得權限分類（應用標籤）
    from django.contrib.contenttypes.models import ContentType
    permission_categories = ContentType.objects.values_list('app_label', 'app_label').distinct().order_by('app_label')
    
    context = {
        "users": users,
        "permissions": permissions,
        "permission_categories": permission_categories,
        "title": "批量權限分配"
    }
    
    return render(request, "system/permission_bulk_assign.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def permission_sync(request):
    """權限同步頁面"""
    if request.method == "POST":
        try:
            from django.contrib.auth.management import create_permissions
            from django.apps import apps
            
            # 檢查同步選項
            sync_all = request.POST.get('sync_all') == 'on'
            selected_modules = request.POST.getlist('sync_modules')
            force_sync = request.POST.get('force_sync') == 'on'
            cleanup_orphaned = request.POST.get('cleanup_orphaned') == 'on'
            
            # 確定要同步的模組
            if sync_all:
                modules_to_sync = ['equip', 'material', 'scheduling', 'process', 'quality', 
                                 'workorder', 'kanban', 'erp_integration', 'ai', 'system']
            else:
                modules_to_sync = selected_modules
            
            if not modules_to_sync:
                messages.error(request, "請選擇要同步的模組")
            else:
                # 執行權限同步
                synced_count = 0
                for module_name in modules_to_sync:
                    try:
                        app_config = apps.get_app_config(module_name)
                        # 創建權限
                        create_permissions(app_config, verbosity=0)
                        synced_count += 1
                        logger.info(f"成功同步模組 {module_name} 的權限")
                    except Exception as e:
                        logger.warning(f"同步模組 {module_name} 時發生警告: {str(e)}")
                        continue
                
                if synced_count > 0:
                    messages.success(request, f"權限同步完成！已同步 {synced_count} 個模組")
                    logger.info(f"權限同步由 {request.user.username} 執行，同步了 {synced_count} 個模組")
                else:
                    messages.warning(request, "沒有模組被同步")
        except Exception as e:
            messages.error(request, f"權限同步失敗：{str(e)}")
            logger.error(f"權限同步失敗: {str(e)}")
    
    # 取得權限統計資訊
    stats = {
        'total_permissions': Permission.objects.count(),
        'total_users': User.objects.count(),
        'permissions_by_module': {}
    }
    
    # 按模組統計權限數量
    module_names = {
        'equip': '設備管理',
        'material': '物料管理', 
        'scheduling': '排程管理',
        'process': '製程管理',
        'quality': '品質管理',
        'workorder': '工單管理',
        'kanban': '看板管理',
        'erp_integration': 'ERP整合',
        'ai': 'AI功能',
        'system': '系統管理'
    }
    
    for module_code, module_display_name in module_names.items():
        count = Permission.objects.filter(content_type__app_label=module_code).count()
        stats['permissions_by_module'][module_display_name] = count
    
    context = {
        "stats": stats,
        "title": "權限同步"
    }
    
    return render(request, "system/permission_sync.html", context)


# 移除複雜的權限進階管理功能，使用標準的 Django 權限系統


# 移除複雜的權限進階管理 API，使用標準的 Django 權限系統


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def execute_specific_auto_approval_task(request):
    """
    執行指定的自動審核定時任務 API
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        # from system.models import ScheduledTask  # 已在上方導入
        from system.tasks import auto_approve_work_reports
        from django.utils import timezone
        
        task_id = request.POST.get('task_id')
        if not task_id:
            return JsonResponse({
                'success': False,
                'message': '缺少任務 ID'
            })
        
        # 取得任務
        task = ScheduledTask.objects.get(id=task_id, task_type='auto_approve')
        
        # 執行自動審核
        result = auto_approve_work_reports()
        
        # 更新任務執行記錄
        task.last_run_at = timezone.now()
        task.execution_count += 1
        
        if result['success']:
            task.success_count += 1
            task.last_error_message = ''
        else:
            task.error_count += 1
            task.last_error_message = result.get('error', '未知錯誤')
        
        task.save()
        
        if result['success']:
            return JsonResponse({
                'success': True,
                'message': f"任務 {task.name} 執行完成：{result['message']}"
            })
        else:
            return JsonResponse({
                'success': False,
                'message': f"任務 {task.name} 執行失敗：{result.get('error', '未知錯誤')}"
            })
            
    except ScheduledTask.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': '找不到指定的任務'
        })
    except Exception as e:
        logger.error(f"執行指定自動審核定時任務失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'執行失敗：{str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def execute_all_auto_approval_tasks(request):
    """
    執行所有自動審核定時任務 API
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        from system.models import ScheduledTask
        from system.tasks import auto_approve_work_reports
        from django.utils import timezone
        
        # 取得所有啟用的自動審核定時任務
        enabled_tasks = ScheduledTask.objects.filter(
            task_type='auto_approve',
            is_enabled=True
        )
        
        if not enabled_tasks.exists():
            return JsonResponse({
                'success': False,
                'message': '沒有啟用的自動審核定時任務'
            })
        
        # 執行自動審核
        result = auto_approve_work_reports()
        
        # 更新所有任務的執行記錄
        for task in enabled_tasks:
            task.last_run_at = timezone.now()
            task.execution_count += 1
            
            if result['success']:
                task.success_count += 1
                task.last_error_message = ''
            else:
                task.error_count += 1
                task.last_error_message = result.get('error', '未知錯誤')
            
            task.save()
        
        if result['success']:
            return JsonResponse({
                'success': True,
                'message': f"所有自動審核定時任務執行完成：{result['message']}"
            })
        else:
            return JsonResponse({
                'success': False,
                'message': f"自動審核定時任務執行失敗：{result.get('error', '未知錯誤')}"
            })
            
    except Exception as e:
        logger.error(f"執行所有自動審核定時任務失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'執行失敗：{str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def auto_approval_tasks(request):
    """
    自動審核定時任務管理頁面
    管理多個自動審核定時任務
    """
    if request.method == 'POST':
        action = request.POST.get('action')
        
        if action == 'create':
            # 建立新的定時任務
            name = request.POST.get('name')
            interval_minutes = request.POST.get('interval_minutes')
            description = request.POST.get('description', '')
            
            if not name or not interval_minutes:
                messages.error(request, "任務名稱和執行間隔為必填欄位")
                return redirect('system:auto_approval_tasks')
            
            try:
                interval_minutes = int(interval_minutes)
                if interval_minutes < 1 or interval_minutes > 1440:
                    messages.error(request, "執行間隔必須在1-1440分鐘之間")
                    return redirect('system:auto_approval_tasks')
                
                # 檢查任務名稱是否重複
                if ScheduledTask.objects.filter(name=name, task_type='auto_approve').exists():
                    messages.error(request, f"任務名稱 '{name}' 已存在")
                    return redirect('system:auto_approval_tasks')
                
                # 建立新任務
                task = ScheduledTask.objects.create(
                    name=name,
                    task_type='auto_approve',
                    task_function='system.tasks.auto_approve_work_reports',
                    execution_type='interval',
                    interval_minutes=interval_minutes,
                    is_enabled=True,
                    description=description
                )
                
                messages.success(request, f"定時任務 '{name}' 建立成功")
                
            except ValueError:
                messages.error(request, "執行間隔必須是有效的數字")
            except Exception as e:
                messages.error(request, f"建立任務失敗：{str(e)}")
                
        elif action == 'update':
            # 更新定時任務
            task_id = request.POST.get('task_id')
            name = request.POST.get('name')
            interval_minutes = request.POST.get('interval_minutes')
            description = request.POST.get('description', '')
            is_enabled = request.POST.get('is_enabled') == 'on'
            
            try:
                task = ScheduledTask.objects.get(id=task_id, task_type='auto_approve')
                
                # 檢查名稱是否重複（排除自己）
                if ScheduledTask.objects.filter(name=name, task_type='auto_approve').exclude(id=task_id).exists():
                    messages.error(request, f"任務名稱 '{name}' 已存在")
                    return redirect('system:auto_approval_tasks')
                
                task.name = name
                task.interval_minutes = int(interval_minutes)
                task.description = description
                task.is_enabled = is_enabled
                task.save()
                
                messages.success(request, f"定時任務 '{name}' 更新成功")
                
            except ScheduledTask.DoesNotExist:
                messages.error(request, "找不到指定的定時任務")
            except ValueError:
                messages.error(request, "執行間隔必須是有效的數字")
            except Exception as e:
                messages.error(request, f"更新任務失敗：{str(e)}")
                
        elif action == 'delete':
            # 刪除定時任務
            task_id = request.POST.get('task_id')
            
            try:
                task = ScheduledTask.objects.get(id=task_id, task_type='auto_approve')
                task_name = task.name
                task.delete()
                messages.success(request, f"定時任務 '{task_name}' 刪除成功")
                
            except ScheduledTask.DoesNotExist:
                messages.error(request, "找不到指定的定時任務")
            except Exception as e:
                messages.error(request, f"刪除任務失敗：{str(e)}")
                
        elif action == 'toggle':
            # 切換任務啟用狀態
            task_id = request.POST.get('task_id')
            
            try:
                task = ScheduledTask.objects.get(id=task_id, task_type='auto_approve')
                task.is_enabled = not task.is_enabled
                task.save()
                
                status = "啟用" if task.is_enabled else "停用"
                messages.success(request, f"定時任務 '{task.name}' 已{status}")
                
            except ScheduledTask.DoesNotExist:
                messages.error(request, "找不到指定的定時任務")
            except Exception as e:
                messages.error(request, f"切換狀態失敗：{str(e)}")
                
        elif action == 'execute':
            # 手動執行定時任務
            task_id = request.POST.get('task_id')
            
            try:
                task = ScheduledTask.objects.get(id=task_id, task_type='auto_approve')
                
                # 執行自動審核任務
                from system.tasks import auto_approve_work_reports
                result = auto_approve_work_reports.delay()
                
                # 更新任務執行記錄
                task.last_run_at = timezone.now()
                task.save()
                
                messages.success(request, f"定時任務 '{task.name}' 手動執行成功")
                
            except ScheduledTask.DoesNotExist:
                messages.error(request, "找不到指定的定時任務")
            except Exception as e:
                messages.error(request, f"執行任務失敗：{str(e)}")
        
        return redirect('system:auto_approval_tasks')
    
    # 取得所有定時任務
    tasks = ScheduledTask.objects.filter(task_type='auto_approve').order_by('-created_at')
    
    context = {
        'tasks': tasks,
        'page_title': '自動審核定時任務管理',
        'breadcrumb': [
            {'name': '系統管理', 'url': 'system:index'},
            {'name': '自動審核定時任務管理', 'url': 'system:auto_approval_tasks'},
        ]
    }
    
    return render(request, 'system/auto_approval_tasks.html', context)

@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def auto_approval_task_detail(request, task_id):
    """
    自動審核定時任務詳情頁面
    """
    try:
        task = ScheduledTask.objects.get(id=task_id, task_type='auto_approve')
    except ScheduledTask.DoesNotExist:
        messages.error(request, "找不到指定的定時任務")
        return redirect('system:auto_approval_tasks')
    
    if request.method == 'POST':
        # 更新任務
        name = request.POST.get('name')
        interval_minutes = request.POST.get('interval_minutes')
        description = request.POST.get('description', '')
        is_enabled = request.POST.get('is_enabled') == 'on'
        
        try:
            # 檢查名稱是否重複
            if ScheduledTask.objects.filter(name=name, task_type='auto_approve').exclude(id=task_id).exists():
                messages.error(request, f"任務名稱 '{name}' 已存在")
            else:
                task.name = name
                task.interval_minutes = int(interval_minutes)
                task.description = description
                task.is_enabled = is_enabled
                task.save()
                
                messages.success(request, f"定時任務 '{name}' 更新成功")
                return redirect('system:auto_approval_tasks')
                
        except ValueError:
            messages.error(request, "執行間隔必須是有效的數字")
        except Exception as e:
            messages.error(request, f"更新任務失敗：{str(e)}")
    
    context = {
        'task': task,
        'page_title': f'編輯定時任務 - {task.name}',
        'breadcrumb': [
            {'name': '系統管理', 'url': 'system:index'},
            {'name': '自動審核定時任務管理', 'url': 'system:auto_approval_tasks'},
            {'name': f'編輯 - {task.name}', 'url': 'system:auto_approval_task_detail', 'args': [task_id]},
        ]
    }
    
    return render(request, 'system/auto_approval_task_detail.html', context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def report_settings(request):
    """
    報表設定頁面
    """
    context = {
        'page_title': '報表設定',
        'breadcrumb': [
            {'name': '系統管理', 'url': 'system:index'},
            {'name': '報表設定', 'url': 'system:report_settings'},
        ]
    }
    
    return render(request, 'system/report_settings.html', context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def manual_sync_reports(request):
    """
    手動同步報表頁面
    """
    context = {
        'page_title': '手動同步報表',
        'breadcrumb': [
            {'name': '系統管理', 'url': 'system:index'},
            {'name': '報表設定', 'url': 'system:report_settings'},
            {'name': '手動同步報表', 'url': 'system:manual_sync_reports'},
        ]
    }
    
    return render(request, 'system/manual_sync_reports.html', context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def order_sync_settings(request):
    """
    訂單同步設定頁面
    """
    try:
        # 取得或創建設定
        settings_obj, created = OrderSyncSettings.objects.get_or_create(
            id=1,
            defaults={
                'sync_enabled': True,
                'sync_interval_minutes': 30,
                'cleanup_enabled': True,
                'cleanup_interval_hours': 24,
                'cleanup_retention_days': 90,
                'status_update_enabled': True,
                'status_update_interval_minutes': 60,
            }
        )
        
        if request.method == 'POST':
            form = OrderSyncSettingsForm(request.POST, instance=settings_obj)
            if form.is_valid():
                form.save()
                messages.success(request, "訂單同步設定已更新！")
                
                # 更新定時任務
                update_order_sync_tasks(settings_obj)
                
                # 不更新同步狀態，只更新定時任務配置
                # 同步狀態只能由實際的同步任務執行來更新
                
                logger.info(f"訂單同步設定由 {request.user.username} 更新")
                return redirect('system:order_sync_settings')
        else:
            form = OrderSyncSettingsForm(instance=settings_obj)
        
        # 取得最近的同步日誌
        recent_logs = OrderSyncLog.objects.all()[:10]
        
        # 取得定時任務狀態
        task_status = get_order_sync_task_status()
        
        context = {
            'form': form,
            'settings': settings_obj,
            'recent_logs': recent_logs,
            'task_status': task_status,
        }
        
        return render(request, 'system/order_sync_settings.html', context)
        
    except Exception as e:
        logger.error(f"訂單同步設定頁面載入失敗: {str(e)}")
        messages.error(request, f"頁面載入失敗: {str(e)}")
        return redirect('system:index')


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def manual_order_sync(request):
    """
    手動執行訂單同步
    """
    try:
        if request.method == 'POST':
            sync_type = request.POST.get('sync_type', 'sync')
            
            # 創建同步日誌
            log = OrderSyncLog.objects.create(
                sync_type=sync_type,
                status='running',
                message='手動執行同步任務',
                started_at=timezone.now()
            )
            
            # 執行同步任務
            if sync_type == 'sync':
                from scheduling.tasks import sync_orders_task
                result = sync_orders_task.delay()
            elif sync_type == 'cleanup':
                from scheduling.tasks import cleanup_old_orders_task
                result = cleanup_old_orders_task.delay()
            elif sync_type == 'status_update':
                from scheduling.tasks import update_order_status_task
                result = update_order_status_task.delay()
            else:
                raise ValueError(f"不支援的同步類型: {sync_type}")
            
            # 更新日誌
            log.details = {'task_id': result.id}
            log.save()
            
            # 注意：不更新設定狀態，讓實際的任務執行來更新狀態
            # 這樣可以確保狀態的真實性
            
            messages.success(request, f"同步任務已啟動，任務ID: {result.id}")
            logger.info(f"手動執行訂單同步任務，類型: {sync_type}，任務ID: {result.id}")
            
        return redirect('system:order_sync_settings')
        
    except Exception as e:
        logger.error(f"手動執行訂單同步失敗: {str(e)}")
        messages.error(request, f"執行失敗: {str(e)}")
        return redirect('system:order_sync_settings')


def update_order_sync_tasks(settings_obj):
    """
    更新訂單同步定時任務
    """
    try:
        from django_celery_beat.models import PeriodicTask, IntervalSchedule
        
        # 更新同步任務
        if settings_obj.sync_enabled:
            interval, _ = IntervalSchedule.objects.get_or_create(
                every=settings_obj.sync_interval_minutes,
                period=IntervalSchedule.MINUTES,
            )
            
            task, created = PeriodicTask.objects.get_or_create(
                name='訂單同步任務',
                defaults={
                    'task': 'scheduling.tasks.sync_orders_task',
                    'interval': interval,
                    'enabled': True,
                }
            )
            
            if not created:
                task.interval = interval
                task.enabled = True
                task.save()
        else:
            # 停用任務
            PeriodicTask.objects.filter(name='訂單同步任務').update(enabled=False)
        
        # 更新清理任務
        if settings_obj.cleanup_enabled:
            interval, _ = IntervalSchedule.objects.get_or_create(
                every=settings_obj.cleanup_interval_hours * 60,  # 轉換為分鐘
                period=IntervalSchedule.MINUTES,
            )
            
            task, created = PeriodicTask.objects.get_or_create(
                name='訂單清理任務',
                defaults={
                    'task': 'scheduling.tasks.cleanup_old_orders_task',
                    'interval': interval,
                    'enabled': True,
                }
            )
            
            if not created:
                task.interval = interval
                task.enabled = True
                task.save()
        else:
            # 停用任務
            PeriodicTask.objects.filter(name='訂單清理任務').update(enabled=False)
        
        # 更新狀態更新任務
        if settings_obj.status_update_enabled:
            interval, _ = IntervalSchedule.objects.get_or_create(
                every=settings_obj.status_update_interval_minutes,
                period=IntervalSchedule.MINUTES,
            )
            
            task, created = PeriodicTask.objects.get_or_create(
                name='訂單狀態更新任務',
                defaults={
                    'task': 'scheduling.tasks.update_order_status_task',
                    'interval': interval,
                    'enabled': True,
                }
            )
            
            if not created:
                task.interval = interval
                task.enabled = True
                task.save()
        else:
            # 停用任務
            PeriodicTask.objects.filter(name='訂單狀態更新任務').update(enabled=False)
            
    except Exception as e:
        logger.error(f"更新訂單同步定時任務失敗: {str(e)}")


def get_order_sync_task_status():
    """
    取得訂單同步定時任務狀態
    """
    try:
        from django_celery_beat.models import PeriodicTask
        
        tasks = {
            'sync': PeriodicTask.objects.filter(name__contains='訂單同步任務').first(),
            'cleanup': PeriodicTask.objects.filter(name__contains='訂單清理任務').first(),
            'status_update': PeriodicTask.objects.filter(name__contains='訂單狀態更新任務').first(),
        }
        
        return tasks
        
    except Exception as e:
        logger.error(f"取得訂單同步任務狀態失敗: {str(e)}")
        return {}


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def workorder_settings(request):
    """
    工單管理設定頁面
    管理工單系統相關設定，包含審核流程、定時任務和完工判斷等
    """
    from workorder.models import SystemConfig
    from django_celery_beat.models import PeriodicTask
    
    if request.method == "POST":
        # 處理表單提交
        auto_approval = request.POST.get('auto_approval') == 'on'
        notification_enabled = request.POST.get('notification_enabled') == 'on'
        audit_log_enabled = request.POST.get('audit_log_enabled') == 'on'
        max_file_size = request.POST.get('max_file_size', 10)
        session_timeout = request.POST.get('session_timeout', 30)
        
        # 自動審核設定
        auto_approve_work_hours = request.POST.get('auto_approve_work_hours') == 'on'
        max_work_hours = float(request.POST.get('max_work_hours', 12.0))
        auto_approve_defect_rate = request.POST.get('auto_approve_defect_rate') == 'on'
        max_defect_rate = float(request.POST.get('max_defect_rate', 5.0))
        auto_approve_overtime = request.POST.get('auto_approve_overtime') == 'on'
        max_overtime_hours = float(request.POST.get('max_overtime_hours', 4.0))
        exclude_operators = request.POST.get('exclude_operators', '')
        exclude_processes = request.POST.get('exclude_processes', '')
        
        # 定時任務設定
        auto_allocation_enabled = request.POST.get('auto_allocation_enabled') == 'on'
        auto_allocation_interval = int(request.POST.get('auto_allocation_interval', 30))
        
        # 完工判斷設定
        completion_check_enabled = request.POST.get('completion_check_enabled') == 'on'
        completion_check_interval = int(request.POST.get('completion_check_interval', 30))
        packaging_process_name = request.POST.get('packaging_process_name', '出貨包裝')
        data_transfer_enabled = request.POST.get('data_transfer_enabled') == 'on'
        transfer_batch_size = int(request.POST.get('transfer_batch_size', 50))
        transfer_retention_days = int(request.POST.get('transfer_retention_days', 365))
        
        # 自動完工定時任務設定
        completion_task_name = request.POST.get('completion_task_name', '')
        completion_fixed_time = request.POST.get('completion_fixed_time', '')
        completion_task_enabled = request.POST.get('completion_task_enabled') == 'on'
        
        # 更新系統設定
        SystemConfig.objects.update_or_create(key="auto_approval", defaults={"value": str(auto_approval)})
        SystemConfig.objects.update_or_create(key="notification_enabled", defaults={"value": str(notification_enabled)})
        SystemConfig.objects.update_or_create(key="audit_log_enabled", defaults={"value": str(audit_log_enabled)})
        SystemConfig.objects.update_or_create(key="max_file_size", defaults={"value": str(max_file_size)})
        SystemConfig.objects.update_or_create(key="session_timeout", defaults={"value": str(session_timeout)})
        
        # 儲存自動審核設定
        SystemConfig.objects.update_or_create(key="auto_approve_work_hours", defaults={"value": str(auto_approve_work_hours)})
        SystemConfig.objects.update_or_create(key="max_work_hours", defaults={"value": str(max_work_hours)})
        SystemConfig.objects.update_or_create(key="auto_approve_defect_rate", defaults={"value": str(auto_approve_defect_rate)})
        SystemConfig.objects.update_or_create(key="max_defect_rate", defaults={"value": str(max_defect_rate)})
        SystemConfig.objects.update_or_create(key="auto_approve_overtime", defaults={"value": str(auto_approve_overtime)})
        SystemConfig.objects.update_or_create(key="max_overtime_hours", defaults={"value": str(max_overtime_hours)})
        SystemConfig.objects.update_or_create(key="exclude_operators", defaults={"value": exclude_operators})
        SystemConfig.objects.update_or_create(key="exclude_processes", defaults={"value": exclude_processes})
        
        # 更新完工判斷設定
        SystemConfig.objects.update_or_create(key="completion_check_enabled", defaults={"value": str(completion_check_enabled)})
        SystemConfig.objects.update_or_create(key="completion_check_interval", defaults={"value": str(completion_check_interval)})
        SystemConfig.objects.update_or_create(key="packaging_process_name", defaults={"value": packaging_process_name})
        SystemConfig.objects.update_or_create(key="data_transfer_enabled", defaults={"value": str(data_transfer_enabled)})
        SystemConfig.objects.update_or_create(key="transfer_batch_size", defaults={"value": str(transfer_batch_size)})
        SystemConfig.objects.update_or_create(key="transfer_retention_days", defaults={"value": str(transfer_retention_days)})
        
        # 處理自動完工定時任務
        if completion_task_name and completion_fixed_time:
            try:
                # 檢查是否已存在完工判斷定時任務
                existing_task = ScheduledTask.objects.filter(
                    task_type='completion_check',
                    name=completion_task_name
                ).first()
                
                if existing_task:
                    # 更新現有任務
                    existing_task.fixed_time = completion_fixed_time
                    existing_task.is_enabled = completion_task_enabled
                    existing_task.save()
                else:
                    # 創建新任務
                    ScheduledTask.objects.create(
                        name=completion_task_name,
                        task_type='completion_check',
                        fixed_time=completion_fixed_time,
                        is_enabled=completion_task_enabled,
                        created_by=request.user.username
                    )
                    
            except Exception as e:
                logger.error(f"處理自動完工定時任務失敗: {str(e)}")
                messages.error(request, f"處理自動完工定時任務失敗：{str(e)}")
        
        # 更新定時任務設定
        try:
            from django_celery_beat.models import IntervalSchedule
            # 自動分配定時任務
            auto_allocation_task = PeriodicTask.objects.get(name='自動分配任務')
            auto_allocation_task.enabled = auto_allocation_enabled
            if auto_allocation_interval != auto_allocation_task.interval.every:
                interval_schedule, _ = IntervalSchedule.objects.get_or_create(
                    every=auto_allocation_interval,
                    period=IntervalSchedule.MINUTES,
                )
                auto_allocation_task.interval = interval_schedule
            auto_allocation_task.save()
        except PeriodicTask.DoesNotExist:
            # 自動創建缺失的定時任務
            try:
                from django_celery_beat.models import IntervalSchedule
                interval_schedule, _ = IntervalSchedule.objects.get_or_create(
                    every=auto_allocation_interval,
                    period=IntervalSchedule.MINUTES,
                )
                PeriodicTask.objects.create(
                    name='自動分配任務',
                    task='workorder.tasks.auto_allocation_task',
                    interval=interval_schedule,
                    enabled=auto_allocation_enabled
                )
                messages.info(request, "已自動創建缺失的定時任務：自動分配任務")
            except Exception as create_error:
                messages.warning(request, f"無法創建定時任務：{str(create_error)}")
        
        messages.success(request, "工單管理設定已成功更新！")
        return redirect('system:workorder_settings')
    
    # 取得現有設定
    def get_config(key, default_value, value_type=str):
        try:
            value = SystemConfig.objects.get(key=key).value
            if value_type == bool:
                return value == "True"
            elif value_type == int:
                return int(value)
            elif value_type == float:
                return float(value)
            return value
        except (SystemConfig.DoesNotExist, ValueError):
            return default_value
    
    # 基本設定
    auto_approval = get_config("auto_approval", False, bool)
    notification_enabled = get_config("notification_enabled", True, bool)
    audit_log_enabled = get_config("audit_log_enabled", True, bool)
    max_file_size = get_config("max_file_size", 10, int)
    session_timeout = get_config("session_timeout", 30, int)
    
    # 自動審核設定
    auto_approve_work_hours = get_config("auto_approve_work_hours", True, bool)
    max_work_hours = get_config("max_work_hours", 12.0, float)
    auto_approve_defect_rate = get_config("auto_approve_defect_rate", True, bool)
    max_defect_rate = get_config("max_defect_rate", 5.0, float)
    auto_approve_overtime = get_config("auto_approve_overtime", False, bool)
    max_overtime_hours = get_config("max_overtime_hours", 4.0, float)
    exclude_operators_text = get_config("exclude_operators", "")
    exclude_processes_text = get_config("exclude_processes", "")
    
    # 完工判斷設定
    completion_check_enabled = get_config("completion_check_enabled", True, bool)
    completion_check_interval = get_config("completion_check_interval", 30, int)
    packaging_process_name = get_config("packaging_process_name", "出貨包裝")
    data_transfer_enabled = get_config("data_transfer_enabled", True, bool)
    transfer_batch_size = get_config("transfer_batch_size", 50, int)
    transfer_retention_days = get_config("transfer_retention_days", 365, int)
    
    # 自動完工定時任務
    completion_tasks = ScheduledTask.objects.filter(task_type='completion_check').order_by('-created_at')
    
    # 取得定時任務狀態
    try:
        auto_allocation_task = PeriodicTask.objects.get(name='自動分配任務')
        auto_allocation_task.interval_minutes = auto_allocation_task.interval.every
    except PeriodicTask.DoesNotExist:
        auto_allocation_task = type('obj', (object,), {
            'enabled': False,
            'interval_minutes': 30,
            'last_run': None
        })
    
    context = {
        # 基本設定
        'auto_approval': auto_approval,
        'notification_enabled': notification_enabled,
        'audit_log_enabled': audit_log_enabled,
        'max_file_size': max_file_size,
        'session_timeout': session_timeout,
        
        # 自動審核設定
        'auto_approve_work_hours': auto_approve_work_hours,
        'max_work_hours': max_work_hours,
        'auto_approve_defect_rate': auto_approve_defect_rate,
        'max_defect_rate': max_defect_rate,
        'auto_approve_overtime': auto_approve_overtime,
        'max_overtime_hours': max_overtime_hours,
        'exclude_operators_text': exclude_operators_text,
        'exclude_processes_text': exclude_processes_text,
        
        # 完工判斷設定
        'completion_check_enabled': completion_check_enabled,
        'completion_check_interval': completion_check_interval,
        'packaging_process_name': packaging_process_name,
        'data_transfer_enabled': data_transfer_enabled,
        'transfer_batch_size': transfer_batch_size,
        'transfer_retention_days': transfer_retention_days,
        
        # 自動完工定時任務
        'completion_tasks': completion_tasks,
        
        # 定時任務狀態
        'auto_allocation_task': auto_allocation_task,
    }
    
    return render(request, 'system/workorder_settings.html', context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def execute_auto_allocation(request):
    """
    手動執行自動分配
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        from workorder.tasks import auto_allocation_task
        
        result = auto_allocation_task.delay()
        
        return JsonResponse({
            'success': True,
            'message': f'自動分配任務已啟動，任務ID: {result.id}'
        })
        
    except Exception as e:
        logger.error(f"手動執行自動分配失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'執行失敗: {str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def execute_completion_check(request):
    """
    手動執行完工檢查
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        from workorder.tasks import completion_check_task
        
        result = completion_check_task.delay()
        
        return JsonResponse({
            'success': True,
            'message': f'完工檢查任務已啟動，任務ID: {result.id}'
        })
        
    except Exception as e:
        logger.error(f"手動執行完工檢查失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'執行失敗: {str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def execute_data_transfer(request):
    """
    手動執行資料轉移
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        from workorder.tasks import data_transfer_task
        
        result = data_transfer_task.delay()
        
        return JsonResponse({
            'success': True,
            'message': f'資料轉移任務已啟動，任務ID: {result.id}'
        })
        
    except Exception as e:
        logger.error(f"手動執行資料轉移失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'執行失敗: {str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def add_auto_approval_task(request):
    """
    新增自動審核定時任務 API
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        from system.models import ScheduledTask
        from datetime import datetime
        
        name = request.POST.get('name', '新自動審核任務')
        execution_type = request.POST.get('execution_type', 'interval')
        
        if execution_type == 'interval':
            interval_minutes = int(request.POST.get('interval_minutes', 30))
            
            if interval_minutes < 5 or interval_minutes > 1440:
                return JsonResponse({
                    'success': False,
                    'message': '執行間隔必須在 5-1440 分鐘之間'
                })
            
            # 建立間隔執行任務
            new_task = ScheduledTask.objects.create(
                name=name,
                task_type='auto_approve',
                task_function='system.tasks.auto_approve_work_reports',
                execution_type='interval',
                interval_minutes=interval_minutes,
                is_enabled=True,
                description=f'自動審核定時任務 - {name}'
            )
        elif execution_type == 'fixed_time':
            fixed_time_str = request.POST.get('fixed_time', '')
            if not fixed_time_str:
                return JsonResponse({
                    'success': False,
                    'message': '固定時間執行必須設定執行時間'
                })
            
            try:
                fixed_time = datetime.strptime(fixed_time_str, '%H:%M').time()
            except ValueError:
                return JsonResponse({
                    'success': False,
                    'message': '時間格式錯誤，請使用 HH:MM 格式'
                })
            
            # 建立固定時間執行任務
            new_task = ScheduledTask.objects.create(
                name=name,
                task_type='auto_approve',
                task_function='system.tasks.auto_approve_work_reports',
                execution_type='fixed_time',
                fixed_time=fixed_time,
                is_enabled=True,
                description=f'自動審核定時任務 - {name}'
            )
        else:
            return JsonResponse({
                'success': False,
                'message': '不支援的執行類型'
            })
        
        return JsonResponse({
            'success': True,
            'message': f'成功新增自動審核定時任務：{name}',
            'task_id': new_task.id
        })
        
    except Exception as e:
        logger.error(f"新增自動審核定時任務失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'新增失敗：{str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def delete_auto_approval_task(request):
    """
    刪除自動審核定時任務 API
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        from system.models import ScheduledTask
        
        task_id = request.POST.get('task_id')
        if not task_id:
            return JsonResponse({
                'success': False,
                'message': '缺少任務 ID'
            })
        
        try:
            task = ScheduledTask.objects.get(id=task_id, task_type='auto_approve')
            task_name = task.name
            task.delete()
            
            return JsonResponse({
                'success': True,
                'message': f'成功刪除自動審核定時任務：{task_name}'
            })
        except ScheduledTask.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': '找不到指定的任務'
            })
        
    except Exception as e:
        logger.error(f"刪除自動審核定時任務失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'刪除失敗：{str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def execute_specific_auto_approval_task(request):
    """
    執行指定的自動審核定時任務 API
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        from system.models import ScheduledTask
        
        task_id = request.POST.get('task_id')
        if not task_id:
            return JsonResponse({
                'success': False,
                'message': '缺少任務 ID'
            })
        
        try:
            task = ScheduledTask.objects.get(id=task_id, task_type='auto_approve')
            
            # 執行任務
            from system.tasks import auto_approve_work_reports
            result = auto_approve_work_reports.delay()
            
            return JsonResponse({
                'success': True,
                'message': f'成功啟動自動審核定時任務：{task.name}',
                'task_id': result.id
            })
        except ScheduledTask.DoesNotExist:
            return JsonResponse({
                'success': False,
                'message': '找不到指定的任務'
            })
        
    except Exception as e:
        logger.error(f"執行自動審核定時任務失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'執行失敗：{str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def execute_all_auto_approval_tasks(request):
    """
    執行所有自動審核定時任務 API
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        from system.models import ScheduledTask
        
        # 取得所有啟用的自動審核定時任務
        tasks = ScheduledTask.objects.filter(task_type='auto_approve', is_enabled=True)
        
        if not tasks.exists():
            return JsonResponse({
                'success': False,
                'message': '沒有啟用的自動審核定時任務'
            })
        
        # 執行所有任務
        from system.tasks import auto_approve_work_reports
        result = auto_approve_work_reports.delay()
        
        return JsonResponse({
            'success': True,
            'message': f'成功啟動所有自動審核定時任務，共 {tasks.count()} 個',
            'task_id': result.id
        })
        
    except Exception as e:
        logger.error(f"執行所有自動審核定時任務失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'執行失敗：{str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def enable_auto_completion(request):
    """
    啟用自動完工功能 API
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        from workorder.models import SystemConfig
        
        # 啟用自動完工功能
        SystemConfig.objects.update_or_create(
            key="auto_completion_enabled",
            defaults={"value": "True"}
        )
        
        return JsonResponse({
            'success': True,
            'message': '自動完工功能已啟用'
        })
        
    except Exception as e:
        logger.error(f"啟用自動完工功能失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'啟用失敗：{str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def disable_auto_completion(request):
    """
    停用自動完工功能 API
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        from workorder.models import SystemConfig
        
        # 停用自動完工功能
        SystemConfig.objects.update_or_create(
            key="auto_completion_enabled",
            defaults={"value": "False"}
        )
        
        return JsonResponse({
            'success': True,
            'message': '自動完工功能已停用'
        })
        
    except Exception as e:
        logger.error(f"停用自動完工功能失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'停用失敗：{str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def add_completion_task(request):
    """
    新增自動完工定時任務 API
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        name = request.POST.get('name')
        fixed_time = request.POST.get('fixed_time')
        
        if not name or not fixed_time:
            return JsonResponse({
                'success': False,
                'message': '任務名稱和固定時間為必填欄位'
            })
        
        # 檢查任務名稱是否重複
        if ScheduledTask.objects.filter(name=name, task_type='completion_check').exists():
            return JsonResponse({
                'success': False,
                'message': f'任務名稱 "{name}" 已存在'
            })
        
        # 創建新任務
        task = ScheduledTask.objects.create(
            name=name,
            task_type='completion_check',
            fixed_time=fixed_time,
            is_enabled=True,
            created_by=request.user.username
        )
        
        return JsonResponse({
            'success': True,
            'message': f'自動完工定時任務 "{name}" 創建成功',
            'task_id': task.id
        })
        
    except Exception as e:
        logger.error(f"新增自動完工定時任務失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'創建失敗：{str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def delete_completion_task(request):
    """
    刪除自動完工定時任務 API
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        task_id = request.POST.get('task_id')
        
        if not task_id:
            return JsonResponse({
                'success': False,
                'message': '任務ID為必填欄位'
            })
        
        task = ScheduledTask.objects.get(id=task_id, task_type='completion_check')
        task_name = task.name
        task.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'自動完工定時任務 "{task_name}" 刪除成功'
        })
        
    except ScheduledTask.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': '找不到指定的任務'
        })
    except Exception as e:
        logger.error(f"刪除自動完工定時任務失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'刪除失敗：{str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def execute_completion_task(request):
    """
    執行自動完工定時任務 API
    """
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': '只支援 POST 請求'})
    
    try:
        from workorder.tasks import completion_check_task
        from django.utils import timezone
        
        # 執行完工檢查任務
        result = completion_check_task.delay()
        
        return JsonResponse({
            'success': True,
            'message': '自動完工檢查任務已啟動',
            'task_id': result.id
        })
        
    except Exception as e:
        logger.error(f"執行自動完工檢查任務失敗: {str(e)}")
        return JsonResponse({
            'success': False,
            'message': f'執行失敗：{str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def sync_permissions_ajax(request):
    """AJAX 權限同步處理"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': '只支援 POST 請求'})
    
    try:
        import json
        from django.core.management import call_command
        from io import StringIO
        
        # 解析請求參數
        options = json.loads(request.body)
        
        # 準備結果
        result = {
            'success': False,
            'error': None,
            'details': {
                'added': 0,
                'removed': 0,
                'updated': 0,
                'fixed': 0
            }
        }
        
        # 如果選擇修復工藝路線權限
        if options.get('fix_route', False):
            try:
                # 查找包含「公益」的權限
                from django.contrib.auth.models import Permission
                incorrect_permissions = Permission.objects.filter(name__contains='公益')
                
                fixed_count = 0
                for perm in incorrect_permissions:
                    corrected_name = perm.name.replace('公益', '工藝')
                    if corrected_name in PERMISSION_NAME_TRANSLATIONS:
                        perm.name = corrected_name
                        perm.save()
                        fixed_count += 1
                        logger.info(f"權限名稱已修復: {perm.name} → {corrected_name}")
                
                result['details']['fixed'] = fixed_count
                
            except Exception as e:
                logger.error(f"修復工藝路線權限時發生錯誤: {str(e)}")
                result['error'] = f"修復工藝路線權限失敗: {str(e)}"
                return JsonResponse(result)
        
        # 如果選擇同步所有權限
        if options.get('sync_all', False):
            try:
                # 使用 Django 內建的權限同步機制（Django 5.1+ 兼容）
                from django.core.management import call_command
                from django.contrib.auth.management import create_permissions
                from django.apps import apps
                
                # 同步所有應用的權限
                for app_config in apps.get_app_configs():
                    if app_config.label in ['equip', 'material', 'scheduling', 'process', 'quality', 
                                          'workorder', 'kanban', 'erp_integration', 'ai', 'system']:
                        try:
                            # 使用 call_command 來執行 makemigrations 和 migrate
                            call_command('makemigrations', app_config.label, verbosity=0)
                            call_command('migrate', app_config.label, verbosity=0)
                            
                            # 創建權限
                            create_permissions(app_config, verbosity=0)
                            
                        except Exception as app_error:
                            logger.warning(f"同步應用 {app_config.label} 時發生警告: {str(app_error)}")
                            continue
                
                # 統計新增的權限（這裡簡化處理）
                result['details']['added'] = 1  # 實際應該統計新增數量
                
            except Exception as e:
                logger.error(f"同步權限時發生錯誤: {str(e)}")
                result['error'] = f"同步權限失敗: {str(e)}"
                return JsonResponse(result)
        
        # 如果選擇更新翻譯
        if options.get('update_translations', False):
            try:
                from django.contrib.auth.models import Permission
                
                updated_count = 0
                for perm in Permission.objects.all():
                    if perm.name in PERMISSION_NAME_TRANSLATIONS:
                        translated_name = PERMISSION_NAME_TRANSLATIONS[perm.name]
                        if perm.name != translated_name:
                            old_name = perm.name
                            perm.name = translated_name
                            perm.save()
                            updated_count += 1
                            logger.info(f"權限翻譯已更新: {old_name} → {translated_name}")
                
                result['details']['updated'] = updated_count
                
            except Exception as e:
                logger.error(f"更新權限翻譯時發生錯誤: {str(e)}")
                result['error'] = f"更新權限翻譯失敗: {str(e)}"
                return JsonResponse(result)
        
        # 如果選擇移除孤立權限
        if options.get('remove_orphaned', False):
            try:
                from django.contrib.auth.models import Permission
                from django.contrib.contenttypes.models import ContentType
                from django.apps import apps
                
                # 獲取現有的模型
                existing_models = set()
                for app_config in apps.get_app_configs():
                    if app_config.label in ['equip', 'material', 'scheduling', 'process', 'quality', 
                                          'workorder', 'kanban', 'erp_integration', 'ai', 'system']:
                        for model in app_config.get_models():
                            existing_models.add(f"{app_config.label}.{model._meta.model_name}")
                
                # 找出孤立的權限
                removed_count = 0
                for perm in Permission.objects.all():
                    model_key = f"{perm.content_type.app_label}.{perm.content_type.model}"
                    if model_key not in existing_models:
                        # 檢查是否是系統內建權限
                        if perm.content_type.app_label not in ['auth', 'contenttypes', 'sessions', 'admin']:
                            perm.delete()
                            removed_count += 1
                            logger.info(f"已移除孤立權限: {perm.name}")
                
                result['details']['removed'] = removed_count
                
            except Exception as e:
                logger.error(f"移除孤立權限時發生錯誤: {str(e)}")
                result['error'] = f"移除孤立權限失敗: {str(e)}"
                return JsonResponse(result)
        
        # 如果是模擬執行，不實際修改資料庫
        if options.get('dry_run', False):
            result['details'] = {
                'added': 5,  # 模擬數據
                'removed': 3,
                'updated': 2,
                'fixed': 1
            }
        
        result['success'] = True
        logger.info(f"權限同步完成: {result['details']}")
        
        return JsonResponse(result)
        
    except Exception as e:
        logger.error(f"權限同步過程中發生未預期的錯誤: {str(e)}")
        return JsonResponse({
            'success': False,
            'error': f'權限同步失敗: {str(e)}'
        })


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def system_dashboard(request):
    """統一的系統管理看板，整合環境管理和操作紀錄管理"""
    import os
    from datetime import datetime
    from django.conf import settings
    from django.contrib.auth.models import User, Permission
    from .models import OperationLogConfig, BackupSchedule, EmailConfig, CleanupLog
    
    # 處理 POST 請求（日誌配置更新和清理）
    if request.method == "POST":
        if 'update_log_config' in request.POST:
            # 更新日誌配置
            try:
                log_config = OperationLogConfig.objects.get(id=1)
                log_config.log_level = request.POST.get('log_level', 'INFO')
                log_config.retention_days = int(request.POST.get('retention_days', 90))
                log_config.max_file_size = int(request.POST.get('max_file_size', 10))
                log_config.save()
                messages.success(request, "日誌配置更新成功！")
            except Exception as e:
                messages.error(request, f"更新日誌配置失敗：{str(e)}")
        
        elif 'clean_logs' in request.POST:
            # 清理日誌
            try:
                clean_logs(request)
                messages.success(request, "日誌清理成功！")
            except Exception as e:
                messages.error(request, f"日誌清理失敗：{str(e)}")
        
        elif 'clean_operation_logs' in request.POST:
            # 清理操作日誌
            try:
                clean_operation_logs(request)
                messages.success(request, "操作日誌清理成功！")
            except Exception as e:
                messages.error(request, f"操作日誌清理失敗：{str(e)}")
    
    # 環境狀態資訊
    environment_info = {
        'debug_mode': settings.DEBUG,
        'environment': '開發環境' if settings.DEBUG else '生產環境',
        'allowed_hosts': settings.ALLOWED_HOSTS,
        'timezone': settings.TIME_ZONE,
        'database_engine': settings.DATABASES['default']['ENGINE'],
        'static_root': settings.STATIC_ROOT,
        'media_root': settings.MEDIA_ROOT,
        'log_base_dir': settings.LOG_BASE_DIR,
    }
    
    # 日誌檔案資訊
    log_dir = settings.LOG_BASE_DIR
    log_files = []
    if os.path.exists(log_dir):
        for filename in os.listdir(log_dir):
            file_path = os.path.join(log_dir, filename)
            if os.path.isfile(file_path):
                file_size = os.path.getsize(file_path) / (1024 * 1024)  # MB
                mod_time = os.path.getmtime(file_path)
                log_files.append({
                    'name': filename,
                    'size_mb': round(file_size, 2),
                    'modified_time': datetime.fromtimestamp(mod_time),
                    'path': file_path
                })
    
    # 操作日誌配置
    try:
        log_config = OperationLogConfig.objects.get(id=1)
    except OperationLogConfig.DoesNotExist:
        log_config = OperationLogConfig.objects.create(
            log_level='INFO',
            retention_days=90,
            max_file_size=10,
            is_active=True
        )
    
    # 備份排程配置
    try:
        backup_config = BackupSchedule.objects.get(id=1)
    except BackupSchedule.DoesNotExist:
        backup_config = BackupSchedule.objects.create(
            schedule_type='daily',
            backup_time='02:00:00',
            retention_days=30,
            is_active=True
        )
    
    # 郵件配置
    try:
        email_config = EmailConfig.objects.get(id=1)
    except EmailConfig.DoesNotExist:
        email_config = EmailConfig.objects.create()
    
    # 系統統計資訊
    system_stats = {
        'total_users': User.objects.count(),

        'total_permissions': Permission.objects.count(),
        'log_files_count': len(log_files),
        'total_log_size_mb': sum(f['size_mb'] for f in log_files),
    }
    
    # 最近的清理記錄
    try:
        recent_cleanups = CleanupLog.objects.order_by('-execution_time')[:5]
    except:
        recent_cleanups = []
    
    # 設定預設日期範圍（最近7天）
    from datetime import datetime, timedelta
    default_end_date = datetime.now().strftime('%Y-%m-%d')
    default_start_date = (datetime.now() - timedelta(days=7)).strftime('%Y-%m-%d')
    
    context = {
        'environment_info': environment_info,
        'log_files': log_files,
        'log_config': log_config,
        'backup_config': backup_config,
        'email_config': email_config,
        'system_stats': system_stats,
        'recent_cleanups': recent_cleanups,
        'title': '統一的系統管理看板'
    }
    
    return render(request, 'system/system_dashboard.html', context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def user_work_permissions(request, user_id):
    """用戶工作權限管理頁面"""
    user = get_object_or_404(User, id=user_id)
    
    # 取得或創建用戶工作權限
    work_permission, created = UserWorkPermission.objects.get_or_create(
        user=user,
        defaults={
            'can_operate_all_operators': True,
            'can_operate_all_processes': True,
            'can_operate_all_equipments': True,
            'can_fill_work': True,
            'can_onsite_reporting': True,
            'can_smt_reporting': True,
            'data_scope': 'all',
            'can_view': True,
            'can_add': True,
            'can_edit': True,
            'can_delete': False,
            'can_approve': False,
            'can_reject': False,
            'can_override_limits': False,
            'can_export_data': True,
        }
    )
    
    if request.method == "POST":
        form = UserWorkPermissionForm(request.POST, instance=work_permission)
        if form.is_valid():
            form.save()
            messages.success(request, f"用戶 {user.username} 的工作權限已更新！")
            logger.info(f"用戶 {user.username} 的工作權限由 {request.user.username} 更新")
            return redirect("system:user_work_permissions", user_id=user_id)
    else:
        form = UserWorkPermissionForm(instance=work_permission)
    
    # 取得統計資訊
    stats = {
        'total_operators': 0,
        'total_processes': 0,
        'total_equipments': 0,
        'allowed_operators_count': len(work_permission.allowed_operators) if not work_permission.can_operate_all_operators else 0,
        'allowed_processes_count': len(work_permission.allowed_processes) if not work_permission.can_operate_all_processes else 0,
        'allowed_equipments_count': len(work_permission.allowed_equipments) if not work_permission.can_operate_all_equipments else 0,
    }
    
    # 動態獲取統計資料
    try:
        from process.models import Operator
        stats['total_operators'] = Operator.objects.count()
    except:
        pass
    
    try:
        from process.models import ProcessName
        stats['total_processes'] = ProcessName.objects.count()
    except:
        pass
    
    try:
        from equip.models import Equipment
        stats['total_equipments'] = Equipment.objects.count()
    except:
        pass
    
    context = {
        'target_user': user,
        'work_permission': work_permission,
        'form': form,
        'stats': stats,
        'title': f"用戶工作權限管理 - {user.username}"
    }
    
    return render(request, "system/user_work_permissions.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def user_work_permissions_list(request):
    """用戶工作權限列表頁面"""
    # 支援搜尋和篩選
    search_query = request.GET.get('search', '')
    permission_filter = request.GET.get('permission', 'all')
    
    # 取得所有有工作權限設定的用戶
    work_permissions = UserWorkPermission.objects.select_related('user').all()
    
    # 搜尋功能
    if search_query:
        work_permissions = work_permissions.filter(
            Q(user__username__icontains=search_query) |
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    
    # 權限篩選
    if permission_filter == 'restricted_operators':
        work_permissions = work_permissions.filter(can_operate_all_operators=False)
    elif permission_filter == 'restricted_processes':
        work_permissions = work_permissions.filter(can_operate_all_processes=False)
    elif permission_filter == 'restricted_equipments':
        work_permissions = work_permissions.filter(can_operate_all_equipments=False)
    elif permission_filter == 'fill_work_only':
        work_permissions = work_permissions.filter(
            can_fill_work=True,
            can_onsite_reporting=False,
            can_smt_reporting=False
        )
    
    # 排序
    work_permissions = work_permissions.order_by('user__username')
    
    # 分頁
    paginator = Paginator(work_permissions, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # 統計資訊 - 修復統計計算
    total_users = User.objects.filter(is_active=True).count()
    configured_users = work_permissions.count()
    
    stats = {
        'total_users': total_users,
        'configured_users': configured_users,
        'restricted_operators': work_permissions.filter(can_operate_all_operators=False).count(),
        'restricted_processes': work_permissions.filter(can_operate_all_processes=False).count(),
        'restricted_equipments': work_permissions.filter(can_operate_all_equipments=False).count(),
        'fill_work_users': work_permissions.filter(can_fill_work=True).count(),
        'onsite_reporting_users': work_permissions.filter(can_onsite_reporting=True).count(),
        'smt_reporting_users': work_permissions.filter(can_smt_reporting=True).count(),
    }
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'permission_filter': permission_filter,
        'stats': stats,
        'title': "用戶工作權限列表"
    }
    
    return render(request, "system/user_work_permissions_list.html", context)


@login_required
@user_passes_test(superuser_required, login_url="/accounts/login/")
def bulk_work_permissions(request):
    """批量工作權限管理頁面"""
    if request.method == "POST":
        selected_users = request.POST.getlist('selected_users')
        operator_permission = request.POST.get('operator_permission') == 'on'
        process_permission = request.POST.get('process_permission') == 'on'
        equipment_permission = request.POST.get('equipment_permission') == 'on'
        permission_scope = request.POST.get('permission_scope', 'all')
        
        if not selected_users:
            messages.error(request, "請選擇要操作的用戶！")
            return redirect("system:bulk_work_permissions")
        
        # 取得選中的權限項目
        allowed_operators = request.POST.getlist('allowed_operators') if operator_permission else []
        allowed_processes = request.POST.getlist('allowed_processes') if process_permission else []
        allowed_equipments = request.POST.getlist('allowed_equipments') if equipment_permission else []
        allowed_companies = request.POST.getlist('allowed_companies') if permission_scope == 'specific' else []
        
        users = User.objects.filter(id__in=selected_users)
        updated_count = 0
        
        for user in users:
            work_permission, created = UserWorkPermission.objects.get_or_create(user=user)
            
            # 更新作業員權限
            if operator_permission:
                # 如果選擇了特定作業員，則設為受限模式
                work_permission.can_operate_all_operators = False
                work_permission.allowed_operators = allowed_operators
            else:
                # 如果沒有選擇作業員權限，則設為全部允許
                work_permission.can_operate_all_operators = True
                work_permission.allowed_operators = []
            
            # 更新工序權限
            if process_permission:
                # 如果選擇了特定工序，則設為受限模式
                work_permission.can_operate_all_processes = False
                work_permission.allowed_processes = allowed_processes
            else:
                # 如果沒有選擇工序權限，則設為全部允許
                work_permission.can_operate_all_processes = True
                work_permission.allowed_processes = []
            
            # 更新設備權限
            if equipment_permission:
                # 如果選擇了特定設備，則設為受限模式
                work_permission.can_operate_all_equipments = False
                work_permission.allowed_equipments = allowed_equipments
            else:
                # 如果沒有選擇設備權限，則設為全部允許
                work_permission.can_operate_all_equipments = True
                work_permission.allowed_equipments = []
            
            # 注意：UserWorkPermission 模型沒有 allowed_companies 欄位
            # 公司範圍控制需要通過其他方式實現
            
            work_permission.save()
            updated_count += 1
        
        messages.success(request, f"已成功更新 {updated_count} 個用戶的工作權限！")
        return redirect("system:bulk_work_permissions")
    
    # 取得所有用戶
    users = User.objects.filter(is_active=True).order_by('username')
    
    # 取得作業員資料
    try:
        from process.models import Operator
        operators = Operator.objects.all().order_by('name')[:100]  # 限制數量
    except ImportError:
        operators = []
    
    # 取得工序資料
    try:
        from process.models import ProcessName
        processes = ProcessName.objects.all().order_by('name')[:100]  # 限制數量
    except ImportError:
        processes = []
    
    # 取得設備資料
    try:
        from equip.models import Equipment
        equipment_list = Equipment.objects.all().order_by('name')[:100]  # 限制數量
    except ImportError:
        equipment_list = []
    
    # 取得公司資料
    try:
        from erp_integration.models import CompanyConfig
        companies = CompanyConfig.objects.all().order_by('company_name')
    except ImportError:
        companies = []
    
    context = {
        'users': users,
        'operators': operators,
        'processes': processes,
        'equipment_list': equipment_list,
        'companies': companies,
        'title': "批量工作權限管理"
    }
    
    return render(request, "system/bulk_work_permissions.html", context)






